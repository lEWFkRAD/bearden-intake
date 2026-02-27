#!/usr/bin/env python3
# ============================================================
# PASSION — Extraction Engine
# ============================================================
"""
Document Intake Extractor v6
=================================
OCR-first architecture: Every page gets free Tesseract OCR first. If the OCR
text is readable (200+ chars, good quality), Claude extracts from text (cheap).
If OCR is poor or Claude flags ambiguous fields, falls back to vision (expensive).
Classification always uses vision (layout matters for form identification).

New in v6:
  - OCR-first extraction (~60-90% cost reduction on readable pages)
  - Cost tracking (token usage + estimated $ per run, in JSON log)
  - Checkpointing (crash recovery — resume from last completed phase)
  - Smart dedup (keeps highest-confidence copy when duplicates exist)
  - --resume flag for crash recovery
  - --no-ocr-first flag to force vision-only mode

Pipeline:
  0a. PyMuPDF text-layer extraction (instant, digital PDFs)
  0b. PDF → images (250 DPI, always needed for classification + vision)
  0c. Tesseract OCR every page (skipped if text layer usable)
  1. Claude vision classifies each page
  1.5. Group pages by EIN/entity
  2. Extract fields:
       Text layer good → Claude text call (cheapest, best quality)
       OCR text good → Claude text call (cheap)
       OCR partial → text call + flag ambiguous fields for verification
       No text / poor → Claude vision call (expensive)
  3. Verify critical fields (vision cross-check)
  4. Normalize (split brokerage composites, cross-ref K-1 continuations)
  5. Validate (arithmetic checks, cross-document reconciliation)
  6. Excel output + JSON audit log (with cost data)

Usage:
    python3 extract.py <pdf_file> [--year 2024] [--output output.xlsx]
    python3 extract.py <pdf_file> --resume   # resume after crash
    python3 extract.py <pdf_file> --no-ocr-first  # force vision-only

Requirements:
    pip install anthropic pdf2image openpyxl Pillow pytesseract
    Also: brew install poppler tesseract (Mac) / sudo apt install poppler-utils tesseract-ocr (Linux)

Setup:
    export ANTHROPIC_API_KEY=your_key_here
"""

import argparse
import base64
import hashlib
import json
import math
import os
import sys
import re
import time
from pathlib import Path
from io import BytesIO
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

# ── Load .env file for persistent secrets ────────────────────────────
def _load_env_file(base: Path = None):
    """Read .env from project root and inject into os.environ (skip if missing)."""
    env_path = (base or Path(__file__).resolve().parent) / ".env"
    if not env_path.exists():
        return
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, val = line.partition("=")
            key = key.strip()
            val = val.strip().strip("\"'")
            if key and val and not os.environ.get(key):
                os.environ[key] = val

_load_env_file()

try:
    import anthropic
except ImportError:
    sys.exit("ERROR: pip install anthropic")
try:
    from pdf2image import convert_from_path
except ImportError:
    sys.exit("ERROR: pip install pdf2image  (also install poppler)")
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.comments import Comment
except ImportError:
    sys.exit("ERROR: pip install openpyxl")
try:
    import pytesseract
    from PIL import Image, ImageDraw
    HAS_TESSERACT = True
except ImportError:
    HAS_TESSERACT = False
    print("WARNING: pytesseract not installed. Will use vision-only mode (slower, more expensive).")
    print("  Install: pip install pytesseract && brew install tesseract")
try:
    import fitz  # PyMuPDF — text-layer extraction from digitally-generated PDFs
    HAS_PYMUPDF = True
except ImportError:
    HAS_PYMUPDF = False

# Fix Windows cp1252 console encoding — allow UTF-8 box-drawing / symbols in print()
import io as _io
if hasattr(sys.stdout, 'buffer'):
    sys.stdout = _io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'buffer'):
    sys.stderr = _io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# ─── CONFIGURATION ───────────────────────────────────────────────────────────

MODEL = "claude-sonnet-4-20250514"
MAX_TOKENS = 8000
DPI = 250
MAX_CONCURRENT = 4  # Parallel API calls (Anthropic rate limit friendly)

# Auto-detect poppler on Windows (pdf2image needs it for PDF → image conversion)
POPPLER_PATH = None
if sys.platform == "win32":
    import glob as _glob
    _candidates = _glob.glob(r"C:\tools\poppler*\Library\bin") + \
                  _glob.glob(r"C:\tools\poppler*\bin") + \
                  _glob.glob(r"C:\Program Files\poppler*\Library\bin") + \
                  _glob.glob(r"C:\Program Files\poppler*\bin")
    for _p in _candidates:
        if os.path.isfile(os.path.join(_p, "pdftoppm.exe")):
            POPPLER_PATH = _p
            break
    if not POPPLER_PATH:
        # Check if pdftoppm is on PATH already
        import shutil
        if not shutil.which("pdftoppm"):
            print("WARNING: poppler not found. PDF conversion will fail.")
            print("  Install: download from https://github.com/oschwartz10612/poppler-windows/releases")
            print("  Extract to C:\\tools\\poppler-XX.XX.X\\")

# Auto-detect Tesseract on Windows
if sys.platform == "win32" and HAS_TESSERACT:
    import shutil as _shutil
    if not _shutil.which("tesseract"):
        _tess_path = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
        if os.path.isfile(_tess_path):
            pytesseract.pytesseract.tesseract_cmd = _tess_path

# Minimum OCR character count to consider OCR usable
OCR_MIN_CHARS = 100

# Text-layer thresholds (PyMuPDF) — used by has_meaningful_text()
TEXT_MIN_CHARS_PER_PAGE = 200  # Minimum chars for a page to count as "meaningful"
TEXT_MIN_TOTAL_CHARS = 800     # Minimum total chars across all meaningful pages

# ─── PAGE PREPROCESSING (T1.1) ───────────────────────────────────────────────
# Deskew: max angle correction (degrees). Pages skewed more than this are left alone.
DESKEW_MAX_ANGLE = 5.0
# Deskew: minimum OSD confidence to trust angle detection
DESKEW_MIN_CONFIDENCE = 2.0
# Blank page detection: page is blank if fewer than this many non-white pixels (as %)
BLANK_PAGE_THRESHOLD = 0.5     # 0.5% non-white pixels → blank
# Blank page: minimum OCR chars to NOT be blank (overrides pixel check)
BLANK_MIN_OCR_CHARS = 20
# Contrast boost: target std-dev for page luminance (skip if already good)
CONTRAST_TARGET_STD = 60.0
# Contrast boost: minimum std-dev below which we enhance
# Synthetic pages: ~20-26 std_dev. Real scans: faded photocopies ~5-12, clean ~25-40.
# Set low to avoid false positives on clean pages with sparse text areas.
CONTRAST_MIN_STD = 15.0
# Quality score thresholds
QUALITY_GOOD = 0.5
QUALITY_POOR = 0.15

# Fields where errors matter most — always verify these against the image
CRITICAL_FIELDS = {
    "wages", "federal_wh", "state_wh",
    "ordinary_dividends", "qualified_dividends",
    "interest_income", "us_savings_bonds_and_treasury",
    "div_ordinary_dividends", "div_qualified_dividends",
    "int_interest_income", "int_us_savings_bonds_and_treasury",
    "b_total_gain_loss", "b_short_term_gain_loss", "b_long_term_gain_loss",
    "short_term_gain_loss", "long_term_gain_loss", "total_gain_loss",
    "box1_ordinary_income", "box2_rental_real_estate",
    "gross_distribution", "taxable_amount", "net_benefits",
    "nonemployee_compensation", "mortgage_interest",
    # Bookkeeping critical fields
    "beginning_balance", "ending_balance", "total_deposits", "total_withdrawals",
    "gross_pay", "net_pay", "total_amount",
}

# ─── CONSENSUS LAYER (T1.2.5) ──────────────────────────────────────────────
# Fields verified by multi-candidate consensus: brokerage totals + K-1 boxes.
# These get independent readings from text-layer, OCR, and Claude extraction.
CONSENSUS_FIELDS = {
    # Brokerage totals — 1099-DIV
    "ordinary_dividends", "qualified_dividends",
    "div_ordinary_dividends", "div_qualified_dividends",
    "div_federal_wh", "div_foreign_tax_paid",
    # Brokerage totals — 1099-INT
    "interest_income", "int_interest_income",
    "us_savings_bonds_and_treasury", "int_us_savings_bonds_and_treasury",
    "int_federal_wh", "int_foreign_tax_paid",
    # Brokerage totals — 1099-B / Schedule D
    "b_total_gain_loss", "b_short_term_gain_loss", "b_long_term_gain_loss",
    "total_gain_loss", "short_term_gain_loss", "long_term_gain_loss",
    # Withholding (generic)
    "federal_wh", "foreign_tax_paid",
    # K-1 boxes 1–9
    "box1_ordinary_income", "box2_rental_real_estate", "box3_other_rental",
    "box4a_guaranteed_services", "box5_interest",
    "box6a_ordinary_dividends", "box6b_qualified_dividends",
    "box7_royalties", "box8_short_term_capital_gain",
    "box9a_long_term_capital_gain",
    # K-1 coded box amounts
    "box10_net_1231_gain", "box11_other_income", "box12_section_179",
    "box13_other_deductions",
}

# Consensus scoring thresholds
CONSENSUS_ACCEPT_THRESHOLD = 5.0   # Minimum score to auto-accept top candidate
CONSENSUS_MARGIN = 2.0              # Minimum gap between top and runner-up

# Doc types eligible for consensus verification
CONSENSUS_DOC_TYPES = {"1099-DIV", "1099-INT", "1099-OID", "1099-B", "K-1", "brokerage"}

# ─── PER-PAGE ROUTING (T1.2) ───────────────────────────────────────────────
# Routing thresholds for text-layer and OCR quality
ROUTE_TEXT_MIN_CHARS = 200       # text_chars >= this → text_layer
ROUTE_TEXT_MIN_WORDS = 40        # text_words >= this AND text_chars >= 120 → text_layer
ROUTE_TEXT_MIN_CHARS_ALT = 120   # used with word count threshold
ROUTE_OCR_MIN_CHARS = 200        # ocr_chars >= this → ocr
ROUTE_OCR_MIN_CONF = 70.0        # ocr_conf_avg >= this → ocr

# ─── SECTION / FORM DETECTION (T1.3) ─────────────────────────────────────────
# Keyword scoring heuristic: each section has a list of (phrase, weight) tuples.
# Higher weight = stronger/rarer signal. Score = sum of weights for all phrase hits.
# A page can match multiple sections (multi-label).
SECTION_SCORE_THRESHOLD = 3.0    # Minimum score to assign a section label
SECTION_KEYWORDS = {
    # ── Brokerage 1099 sections ──
    "summary": [
        # Strong signals — unique to brokerage summary pages
        ("summary of income", 5),
        ("tax information statement", 4),
        ("combined summary", 4),
        ("year-end summary", 4),
        ("year end summary", 4),
        ("annual summary", 4),
        ("composite statement", 3),
        ("total ordinary dividends", 3),
        ("total qualified dividends", 3),
        ("total interest income", 3),
        ("federal income tax withheld", 3),
        ("total capital gains", 2),
        ("total dividends", 2),
        ("total interest", 2),
        ("aggregate amounts", 2),
        ("reportable amounts", 2),
    ],
    "div": [
        # Strong signals — IRS form identifiers
        ("1099-div", 6),
        ("form 1099-div", 6),
        ("dividends and distributions", 5),
        # Box labels unique to 1099-DIV
        ("1a ordinary dividends", 4),
        ("1a  ordinary dividends", 4),
        ("1b qualified dividends", 4),
        ("1b  qualified dividends", 4),
        ("nondividend distributions", 3),
        ("section 199a dividends", 3),
        ("collectibles (28%) gain", 3),
        ("section 897 dividends", 3),
        ("foreign tax paid", 2),
        ("exempt-interest dividends", 3),
        ("specified private activity bond interest dividends", 3),
        # Medium signals — common but not unique
        ("ordinary dividends", 2),
        ("qualified dividends", 2),
    ],
    "int": [
        # Strong signals
        ("1099-int", 6),
        ("form 1099-int", 6),
        ("interest income", 4),
        # Box labels unique to 1099-INT
        ("early withdrawal penalty", 4),
        ("us savings bonds and treasury", 3),
        ("u.s. savings bonds", 3),
        ("tax-exempt interest", 3),
        ("tax exempt interest", 3),
        ("specified private activity bond interest", 3),
        ("investment expenses", 2),
        ("bond premium", 2),
        ("market discount", 2),
    ],
    "oid": [
        # Strong signals
        ("1099-oid", 6),
        ("form 1099-oid", 6),
        ("original issue discount", 5),
        # Box labels
        ("other periodic interest", 3),
        ("early withdrawal penalty", 2),  # shared with INT, lower weight
        ("acquisition premium", 3),
        ("oid on us treasury", 3),
    ],
    "b_summary": [
        # Strong signals — 1099-B summary/totals
        ("1099-b", 5),
        ("form 1099-b", 5),
        ("proceeds from broker", 5),
        ("short-term transactions", 4),
        ("long-term transactions", 4),
        ("short term capital gains", 3),
        ("long term capital gains", 3),
        ("aggregate profit or loss", 3),
        ("total proceeds", 3),
        ("total cost basis", 3),
        ("total gain/loss", 3),
        ("total gain or loss", 3),
        ("wash sale loss disallowed", 3),
        ("reported to irs", 2),
        ("basis reported", 2),
        ("basis not reported", 2),
        ("noncovered securities", 2),
    ],
    "b_transactions": [
        # Transaction detail pages — high volume, lower priority
        ("date acquired", 3),
        ("date sold", 3),
        ("date of sale", 3),
        ("quantity sold", 3),
        ("proceeds", 2),
        ("cost basis", 2),
        ("gain/loss", 2),
        ("gain or loss", 2),
        ("cusip", 2),
        ("symbol", 1),
    ],
    # ── K-1 sections ──
    "k1_1065": [
        # Very strong — form-specific
        ("schedule k-1 (form 1065)", 8),
        ("schedule k-1(form 1065)", 8),
        ("form 1065", 5),
        ("partner's share of income", 5),
        ("partner's share", 4),
        ("partnership", 3),
        ("partner's instructions", 3),
        ("general partner", 2),
        ("limited partner", 2),
    ],
    "k1_1120s": [
        # Very strong — form-specific
        ("schedule k-1 (form 1120-s)", 8),
        ("schedule k-1 (form 1120s)", 8),
        ("schedule k-1(form 1120-s)", 8),
        ("schedule k-1(form 1120s)", 8),
        ("form 1120-s", 5),
        ("form 1120s", 5),
        ("shareholder's share of income", 5),
        ("shareholder's share", 4),
        ("s corporation", 3),
        ("shareholder's instructions", 3),
    ],
    "k1_1041": [
        # Very strong — form-specific
        ("schedule k-1 (form 1041)", 8),
        ("schedule k-1(form 1041)", 8),
        ("form 1041", 5),
        ("beneficiary's share of income", 5),
        ("beneficiary's share", 4),
        ("estate or trust", 4),
        ("fiduciary", 3),
        ("beneficiary's instructions", 3),
        ("trust", 2),
        ("estate", 2),
    ],
    # ── Generic / fallback ──
    "cover": [
        ("important tax information", 4),
        ("important tax return information", 4),
        ("tax reporting statement", 4),
        ("this is not a bill", 3),
        ("do not file", 3),
        ("for your records", 3),
        ("enclosed you will find", 3),
        ("enclosed is your", 3),
        ("dear client", 2),
        ("dear investor", 2),
        ("dear shareholder", 2),
        ("dear account holder", 2),
    ],
    "continuation": [
        ("continued from", 4),
        ("continued on next page", 4),
        ("continuation", 3),
        ("see attached", 2),
        ("see statement", 2),
        ("additional detail", 2),
        ("supplemental information", 2),
        ("stmt", 1),
    ],
}

# T1.5: Section-based extraction priority for progressive results.
# Lower number = higher priority = extracted first.
SECTION_PRIORITY = {
    "summary": 1,       # Brokerage summary — totals, dividends, interest
    "k1_1065": 2,       # K-1 partnership
    "k1_1120s": 2,      # K-1 S-corp
    "k1_1041": 2,       # K-1 estate/trust
    "div": 3,           # 1099-DIV section
    "int": 3,           # 1099-INT section
    "oid": 4,           # 1099-OID section
    "b_summary": 4,     # 1099-B totals
    "unknown": 6,       # Unknown section type
    "b_transactions": 7, # Individual trades (high volume, lower priority)
    "cover": 8,         # Cover letters (low value)
    "continuation": 9,  # Continuation pages (handled by multi-page batching)
}


REQUIRES_HUMAN_REVIEW = [
    "K-1 carryover from prior year (basis/at-risk limitations)",
    "K-1 allowed loss vs. suspended loss allocation",
    "K-1 carryover to next year",
    "Capital loss carryover from prior year",
    "Net operating loss carryover",
    "Section 179 carryover",
    "Passive activity loss carryover",
    "Depreciation schedules (continuing assets)",
    "Estimated tax payments (not on source documents)",
    "Prior year state refund (taxability depends on PY itemizing)",
    "At-risk basis calculations",
    "Qualified business income (QBI) carryover",
    "Charitable contribution carryover (if exceeded AGI limit PY)",
    "Installment sale deferred gain tracking",
]


# ─── COST TRACKER ────────────────────────────────────────────────────────────

class CostTracker:
    """Track API token usage and estimate costs per run.

    Pricing (Sonnet, as of 2025):
      Input:  $3 / MTok
      Output: $15 / MTok
      Image:  ~1,600 tokens per image at 250 DPI
    """
    INPUT_COST_PER_MTOK = 3.0
    OUTPUT_COST_PER_MTOK = 15.0

    def __init__(self):
        self.calls = []  # list of {phase, page, input_tokens, output_tokens, call_type}
        self.text_calls = 0
        self.vision_calls = 0

    def record(self, phase, page, usage, call_type="vision"):
        """Record a single API call's usage."""
        inp = getattr(usage, 'input_tokens', 0) if usage else 0
        out = getattr(usage, 'output_tokens', 0) if usage else 0
        self.calls.append({
            "phase": phase, "page": page,
            "input_tokens": inp, "output_tokens": out,
            "call_type": call_type,
        })
        if call_type == "text":
            self.text_calls += 1
        else:
            self.vision_calls += 1  # vision, vision_multi both count as vision

    def total_input(self):
        return sum(c["input_tokens"] for c in self.calls)

    def total_output(self):
        return sum(c["output_tokens"] for c in self.calls)

    def total_cost(self):
        inp = self.total_input() / 1_000_000
        out = self.total_output() / 1_000_000
        return inp * self.INPUT_COST_PER_MTOK + out * self.OUTPUT_COST_PER_MTOK

    def summary(self):
        if not self.calls:
            return "  No API calls made"
        lines = [
            f"  API calls:    {len(self.calls)} ({self.vision_calls} vision, {self.text_calls} text)",
            f"  Input tokens: {self.total_input():,}",
            f"  Output tokens:{self.total_output():,}",
            f"  Est. cost:    ${self.total_cost():.4f}",
        ]
        # Per-phase breakdown
        phases = {}
        for c in self.calls:
            p = c["phase"]
            if p not in phases:
                phases[p] = {"count": 0, "input": 0, "output": 0}
            phases[p]["count"] += 1
            phases[p]["input"] += c["input_tokens"]
            phases[p]["output"] += c["output_tokens"]
        for p, d in sorted(phases.items()):
            cost = (d["input"] / 1e6 * self.INPUT_COST_PER_MTOK +
                    d["output"] / 1e6 * self.OUTPUT_COST_PER_MTOK)
            lines.append(f"    {p}: {d['count']} calls, ${cost:.4f}")
        return "\n".join(lines)

    def to_dict(self):
        return {
            "total_calls": len(self.calls),
            "vision_calls": self.vision_calls,
            "text_calls": self.text_calls,
            "total_input_tokens": self.total_input(),
            "total_output_tokens": self.total_output(),
            "estimated_cost_usd": round(self.total_cost(), 4),
            "per_phase": {p: sum(1 for c in self.calls if c["phase"] == p)
                          for p in set(c["phase"] for c in self.calls)},
        }

# Global tracker — set per run in main()
_cost_tracker = None


class PipelineTimer:
    """Track wall-clock timing per pipeline phase.

    Usage:
        timer = PipelineTimer()
        timer.start("ocr")       # starts timing "ocr" phase
        ...work...
        timer.start("classify")  # auto-stops "ocr", starts "classify"
        ...work...
        timer.stop()             # stops current phase
        print(timer.summary())   # formatted timing table
    """

    def __init__(self):
        self._phases = []       # [{"name": str, "start": float, "end": float|None}]
        self._active = None
        self._run_start = time.monotonic()

    def start(self, name):
        """Start a named phase. Auto-stops the previous phase if still running."""
        now = time.monotonic()
        if self._active and self._phases:
            self._phases[-1]["end"] = now
        self._phases.append({"name": name, "start": now, "end": None})
        self._active = name

    def stop(self):
        """Stop the currently running phase."""
        if self._active and self._phases:
            self._phases[-1]["end"] = time.monotonic()
            self._active = None

    def elapsed(self, name):
        """Get elapsed seconds for a completed phase."""
        for p in self._phases:
            if p["name"] == name and p["end"] is not None:
                return p["end"] - p["start"]
        return 0.0

    def total_elapsed(self):
        """Total elapsed time since timer creation."""
        return time.monotonic() - self._run_start

    def summary(self):
        """Formatted timing table string."""
        lines = []
        for p in self._phases:
            dur = (p["end"] or time.monotonic()) - p["start"]
            lines.append(f"    {p['name']:<30s} {dur:6.2f}s")
        lines.append(f"    {'TOTAL':<30s} {self.total_elapsed():6.2f}s")
        return "\n".join(lines)

    def to_dict(self):
        """Serializable dict for JSON log: {phase_name_s: float, total_s: float}."""
        d = {}
        for p in self._phases:
            d[f"{p['name']}_s"] = round((p["end"] or time.monotonic()) - p["start"], 3)
        d["total_s"] = round(self.total_elapsed(), 3)
        return d


_pipeline_timer = None


# ─── PAGE-LEVEL CACHE (T1.4) ────────────────────────────────────────────────
CACHE_VERSION = 1
CACHE_DIR = os.path.join("data", "cache")
CACHE_MAX_AGE_DAYS = 30


def _cache_key(pdf_path):
    """Compute SHA-256 hash of PDF file content for cache keying."""
    h = hashlib.sha256()
    with open(pdf_path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _save_cache(pdf_path, dpi, page_texts, text_layer_stats,
                b64_images, page_preprocessing, ocr_texts, ocr_confidences, routing_plan):
    """Save Phase 0 outputs to disk cache for fast rerun."""
    try:
        h = _cache_key(pdf_path)
        cache_dir = os.path.join(CACHE_DIR, h[:12])
        img_dir = os.path.join(cache_dir, "images")
        os.makedirs(img_dir, exist_ok=True)

        manifest = {
            "hash": h,
            "dpi": dpi,
            "page_count": len(b64_images) if b64_images else 0,
            "created": datetime.now().isoformat(),
            "cache_version": CACHE_VERSION,
            "pdf_path": str(pdf_path),
        }

        with open(os.path.join(cache_dir, "manifest.json"), "w") as f:
            json.dump(manifest, f, indent=2)
        with open(os.path.join(cache_dir, "page_texts.json"), "w") as f:
            json.dump(page_texts, f)
        with open(os.path.join(cache_dir, "text_layer_stats.json"), "w") as f:
            json.dump(text_layer_stats, f)
        with open(os.path.join(cache_dir, "preprocessing.json"), "w") as f:
            json.dump(page_preprocessing, f, default=str)
        with open(os.path.join(cache_dir, "ocr_texts.json"), "w") as f:
            json.dump(ocr_texts, f)
        with open(os.path.join(cache_dir, "ocr_confidences.json"), "w") as f:
            json.dump(ocr_confidences, f)
        with open(os.path.join(cache_dir, "routing_plan.json"), "w") as f:
            json.dump(routing_plan, f)

        # Write base64 images as individual files (avoids huge single JSON)
        if b64_images:
            for i, img_b64 in enumerate(b64_images):
                img_path = os.path.join(img_dir, f"page_{i+1:03d}.b64")
                with open(img_path, "w") as f:
                    f.write(img_b64 if img_b64 else "")

        print(f"  Cache saved: {cache_dir}")
    except (IOError, OSError) as e:
        print(f"  Cache save failed (non-fatal): {e}")


def _load_cache(pdf_path, dpi):
    """Load Phase 0 outputs from disk cache. Returns dict or None on miss."""
    try:
        h = _cache_key(pdf_path)
        cache_dir = os.path.join(CACHE_DIR, h[:12])
        manifest_path = os.path.join(cache_dir, "manifest.json")

        if not os.path.exists(manifest_path):
            return None

        with open(manifest_path) as f:
            manifest = json.load(f)

        # Validate: hash, DPI, version, freshness
        if manifest.get("hash") != h:
            return None
        if manifest.get("dpi") != dpi:
            return None
        if manifest.get("cache_version", 0) != CACHE_VERSION:
            return None

        created = manifest.get("created")
        if created:
            age_days = (datetime.now() - datetime.fromisoformat(created)).days
            if age_days > CACHE_MAX_AGE_DAYS:
                return None

        # Load all cached data
        with open(os.path.join(cache_dir, "page_texts.json")) as f:
            page_texts = json.load(f)
        with open(os.path.join(cache_dir, "text_layer_stats.json")) as f:
            text_layer_stats = json.load(f)
        with open(os.path.join(cache_dir, "preprocessing.json")) as f:
            page_preprocessing = json.load(f)
        with open(os.path.join(cache_dir, "ocr_texts.json")) as f:
            ocr_texts = json.load(f)
        with open(os.path.join(cache_dir, "ocr_confidences.json")) as f:
            ocr_confidences = json.load(f)
        with open(os.path.join(cache_dir, "routing_plan.json")) as f:
            routing_plan = json.load(f)

        # Load page images
        page_count = manifest.get("page_count", 0)
        b64_images = []
        img_dir = os.path.join(cache_dir, "images")
        for i in range(page_count):
            img_path = os.path.join(img_dir, f"page_{i+1:03d}.b64")
            if os.path.exists(img_path):
                with open(img_path) as f:
                    data = f.read()
                b64_images.append(data if data else None)
            else:
                return None  # incomplete cache

        return {
            "page_texts": page_texts,
            "text_layer_stats": text_layer_stats,
            "b64_images": b64_images,
            "page_preprocessing": page_preprocessing,
            "ocr_texts": ocr_texts,
            "ocr_confidences": ocr_confidences,
            "routing_plan": routing_plan,
        }
    except (IOError, OSError, json.JSONDecodeError, KeyError) as e:
        print(f"  Cache load failed (non-fatal): {e}")
        return None


def _print_routing_summary(routing_plan):
    """Print routing summary from cached data (same format as route_pages for app.py)."""
    counts = {"text_layer": 0, "ocr": 0, "vision": 0, "skip_blank": 0}
    print("\n── [PASSION] Per-Page Routing (from cache) ──")
    for r in routing_plan:
        m = r.get("method", "vision")
        counts[m] = counts.get(m, 0) + 1
        pnum = r.get("page_num", "?")
        reason = r.get("reason", "")
        tc = r.get("text_chars", 0)
        oc = r.get("ocr_chars", 0)
        conf = r.get("ocr_conf_avg")
        if m == "skip_blank":
            print(f"  Page {pnum}: skip_blank ({reason})")
        elif m == "text_layer":
            print(f"  Page {pnum}: text_layer ({reason}, {tc} chars)")
        elif m == "ocr":
            conf_str = f", conf {conf:.0f}%" if conf is not None else ""
            print(f"  Page {pnum}: ocr ({reason}, {oc} chars{conf_str})")
        else:
            print(f"  Page {pnum}: vision ({reason}, text={tc}, ocr={oc})")
    parts = []
    for m in ("text_layer", "ocr", "vision", "skip_blank"):
        if counts[m] > 0:
            parts.append(f"{counts[m]} {m}")
    print(f"  Routing summary: {', '.join(parts)}")


# ─── DOC TYPE → CLASSIFICATION NARROWING ──────────────────────────────────────

DOC_TYPE_CLASSIFICATIONS = {
    "tax_returns": [
        "W-2", "W-2G",
        "1099-INT", "1099-DIV", "1099-B", "1099-R", "1099-NEC", "1099-MISC",
        "1099-SA", "1099-G", "1099-K", "1099-S", "1099-C", "1099-Q", "1099-LTC",
        "SSA-1099", "K-1", "1098", "1098-T", "1098-E",
        "5498", "5498-SA", "1095-A", "1095-B", "1095-C",
        "brokerage_composite", "continuation_statement",
        "property_tax_bill", "charitable_receipt", "estimated_tax_record",
        "farm_income_document", "rental_income_document",
        "schedule_c_summary", "other",
    ],
    "bank_statements": [
        "bank_statement", "bank_statement_deposit_slip", "check", "check_stub",
        "chart_of_accounts", "trial_balance_document", "credit_card_statement", "other",
    ],
    "trust_documents": [
        "K-1", "1099-INT", "1099-DIV", "1099-R", "1099-B",
        "brokerage_composite", "continuation_statement",
        "bank_statement", "trust_accounting_statement", "other",
    ],
    "bookkeeping": [
        "bank_statement", "bank_statement_deposit_slip", "check", "check_stub",
        "credit_card_statement", "invoice", "receipt",
        "profit_loss_statement", "balance_sheet", "aged_receivables", "aged_payables",
        "chart_of_accounts", "trial_balance_document",
        "loan_statement", "mortgage_statement", "other",
    ],
    "payroll": [
        "W-2", "W-3", "check_stub",
        "940", "941", "943", "944", "945",
        "payroll_register", "payroll_summary", "other",
    ],
    "other": [
        "W-2", "1099-INT", "1099-DIV", "1099-B", "1099-R", "1099-NEC", "1099-MISC",
        "1099-K", "1099-S", "1099-C", "1099-G", "1099-SA",
        "K-1", "1098", "bank_statement", "invoice", "receipt", "check",
        "loan_statement", "mortgage_statement", "profit_loss_statement", "balance_sheet",
        "other",
    ],
}

# ─── PII TOKENIZER ───────────────────────────────────────────────────────────

class PIITokenizer:
    """
    Replaces SSNs and personal names with tokens before API calls,
    reverses them after. EINs are NOT tokenized — they're business
    identifiers that Claude needs for classification and grouping.

    Text tokenization: regex detect → replace with [SSN_1], [NAME_1], etc.
    Image tokenization: pytesseract word-level detection → black-box SSN regions.

    The token map lives in memory only and is never written to disk.
    """

    # SSN patterns: 123-45-6789, 123 45 6789, 123456789 (9 digits no dashes)
    # Also partial-masked: XXX-XX-1234, ***-**-1234, xxx-xx-1234
    SSN_PATTERN = re.compile(
        r'\b(\d{3}[-\s]?\d{2}[-\s]?\d{4})\b'
        r'|'
        r'(?<!\w)([Xx*]{3}[-\s]?[Xx*]{2}[-\s]?\d{4})(?!\w)'
    )

    # Last-4 SSN patterns (common on W-2 copies): "SSN: ***-**-1234" already caught above
    # Also catch standalone "last 4: 1234" type references
    SSN_LAST4_PATTERN = re.compile(
        r'(?:SSN|social\s*security)[\s:]*(?:last\s*4[\s:]*)(\d{4})',
        re.IGNORECASE
    )

    def __init__(self):
        self.ssn_map = {}       # token → real value
        self.ssn_reverse = {}   # real value → token
        self._ssn_counter = 0

    def _next_ssn_token(self):
        self._ssn_counter += 1
        return f"[SSN_{self._ssn_counter}]"

    def _register_ssn(self, raw_ssn):
        """Register an SSN and return its token. Reuses token if already seen."""
        # Normalize: strip spaces, keep dashes
        normalized = raw_ssn.strip()
        if normalized in self.ssn_reverse:
            return self.ssn_reverse[normalized]
        token = self._next_ssn_token()
        self.ssn_map[token] = normalized
        self.ssn_reverse[normalized] = token
        return token

    def tokenize_text(self, text):
        """Replace SSNs in OCR text with tokens. Returns (tokenized_text, found_count)."""
        if not text:
            return text, 0

        found = 0

        def replace_ssn(match):
            nonlocal found
            raw = match.group(0)
            found += 1
            return self._register_ssn(raw)

        result = self.SSN_PATTERN.sub(replace_ssn, text)
        return result, found

    def detokenize_text(self, text):
        """Reverse all tokens back to real values in a string."""
        if not text:
            return text
        for token, real in self.ssn_map.items():
            text = text.replace(token, real)
        return text

    def detokenize_json(self, obj):
        """Recursively reverse tokens in a parsed JSON object."""
        if obj is None:
            return None
        if isinstance(obj, str):
            return self.detokenize_text(obj)
        if isinstance(obj, dict):
            return {k: self.detokenize_json(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [self.detokenize_json(item) for item in obj]
        return obj

    def redact_image(self, pil_image):
        """
        Find SSN-pattern text regions in the image and black them out.
        Returns a new PIL Image with SSN regions redacted.

        Uses pytesseract word-level bounding boxes to locate digits,
        then blacks out any sequence matching SSN patterns.
        """
        if not HAS_TESSERACT:
            return pil_image  # Can't redact without tesseract

        img = pil_image.copy()
        draw = ImageDraw.Draw(img)

        try:
            # Get word-level bounding boxes from tesseract
            data = pytesseract.image_to_data(img, output_type=pytesseract.Output.DICT)
        except Exception:
            return img  # If tesseract fails, return unmodified

        n = len(data['text'])
        # Build text with positions for SSN detection
        words = []
        for i in range(n):
            txt = data['text'][i].strip()
            if txt:
                words.append({
                    'text': txt,
                    'left': data['left'][i],
                    'top': data['top'][i],
                    'width': data['width'][i],
                    'height': data['height'][i],
                    'index': i
                })

        # Look for SSN patterns across consecutive words
        # SSNs can appear as: "123-45-6789" (1 word) or "123" "-" "45" "-" "6789" (5 words)
        # or "XXX-XX-2224" (masked, 1 word)
        full_text_positions = []  # (start_char, end_char, word_indices)
        running_text = ""
        for wi, w in enumerate(words):
            start = len(running_text)
            running_text += w['text'] + " "
            full_text_positions.append((start, len(running_text) - 1, wi))

        # Find SSN matches in the concatenated text
        for match in self.SSN_PATTERN.finditer(running_text):
            mstart, mend = match.start(), match.end()
            # Find which words overlap this match
            indices_to_redact = []
            for start, end, wi in full_text_positions:
                if start < mend and end > mstart:
                    indices_to_redact.append(wi)

            # Black out those word regions with padding
            for wi in indices_to_redact:
                w = words[wi]
                pad = 4
                draw.rectangle(
                    [w['left'] - pad, w['top'] - pad,
                     w['left'] + w['width'] + pad, w['top'] + w['height'] + pad],
                    fill='black'
                )

            # Register the matched SSN
            self._register_ssn(match.group(0))

        return img

    def redacted_image_to_b64(self, pil_image):
        """Redact SSNs from image and return base64 JPEG string."""
        redacted = self.redact_image(pil_image)
        buf = BytesIO()
        redacted.save(buf, format="JPEG", quality=85)
        return base64.b64encode(buf.getvalue()).decode("utf-8")

    def get_stats(self):
        """Return summary of tokenization for logging."""
        return {
            "ssns_tokenized": len(self.ssn_map),
            "tokens": list(self.ssn_map.keys()),
        }


# ─── PROMPTS ─────────────────────────────────────────────────────────────────

# Phase 1: Classification (always uses vision — need to see page layout)
CLASSIFICATION_PROMPT = """You are a tax document classification specialist. Examine this scanned page and classify it.

IMPORTANT: A single page from a consolidated brokerage/investment statement may contain data for MULTIPLE tax forms (1099-DIV, 1099-INT, 1099-MISC, 1099-B). If you see multiple form sections on one page, list ALL of them in the sub_types array.

Return JSON only:
{
    "page_number": <int>,
    "document_type": "<primary type - see list>",
    "sub_types": ["<list ALL form types found on this page>"],
    "is_consolidated_brokerage": <bool>,
    "is_duplicate_copy": <bool>,
    "is_supplemental_detail": <bool - true if transaction-level detail, NOT summary>,
    "is_continuation_statement": <bool>,
    "supports_form": "<if supplemental/continuation, which form>",
    "payer_or_entity": "<name>",
    "payer_ein": "<EIN/TIN if visible>",
    "recipient": "<person name>",
    "recipient_ssn_last4": "<last 4 if visible>",
    "tax_year": "<year>",
    "brief_description": "<one line>"
}

Document types:
W-2, W-2G, W-3,
1099-INT, 1099-DIV, 1099-B, 1099-R, 1099-NEC, 1099-MISC, 1099-SA, 1099-G,
1099-K, 1099-S, 1099-C, 1099-Q, 1099-LTC,
SSA-1099, K-1, 1098, 1098-T, 1098-E,
5498, 5498-SA, 1095-A, 1095-B, 1095-C,
brokerage_composite, continuation_statement,
property_tax_bill, charitable_receipt, farm_income_document, rental_income_document,
estimated_tax_record, schedule_c_summary,
bank_statement, bank_statement_deposit_slip, check, check_stub,
credit_card_statement, invoice, receipt,
profit_loss_statement, balance_sheet, aged_receivables, aged_payables,
chart_of_accounts, trial_balance_document,
loan_statement, mortgage_statement,
payroll_register, payroll_summary,
940, 941, 943, 944, 945,
trust_accounting_statement,
other"""

# Phase 2a: OCR-based extraction (cheap, text-only API call)
OCR_EXTRACTION_PROMPT = """You are a tax document extraction specialist. Below is OCR text extracted from a scanned tax document.
The document has been classified as: {doc_type}

Your job:
1. Extract ALL tax-relevant fields from the OCR text
2. Rate the OCR quality for each field
3. Decide if you can confidently extract from this text alone, or if the original image is needed

FIELD NAMING RULES:
For consolidated brokerage statements (multiple forms on one page), use prefixed names:
  1099-DIV: div_ordinary_dividends, div_qualified_dividends, div_capital_gain_distributions, div_nondividend_distributions, div_federal_wh, div_section_199a, div_foreign_tax_paid, div_exempt_interest, div_private_activity_bond, div_state_wh
  1099-INT: int_interest_income, int_early_withdrawal_penalty, int_us_savings_bonds_and_treasury, int_federal_wh, int_investment_expenses, int_foreign_tax_paid, int_tax_exempt_interest, int_market_discount, int_bond_premium, int_bond_premium_treasury, int_state_wh
  1099-B: b_short_term_proceeds, b_short_term_basis, b_short_term_gain_loss, b_long_term_proceeds, b_long_term_basis, b_long_term_gain_loss, b_total_proceeds, b_total_basis, b_total_gain_loss, b_federal_wh, b_market_discount, b_wash_sale_loss
  1099-MISC: misc_royalties, misc_other_income, misc_rents, misc_federal_wh, misc_state_wh

For standalone forms, use unprefixed names:
  W-2: employer_name, employer_ein, wages, federal_wh, ss_wages, ss_wh, medicare_wages, medicare_wh, state_wages, state_wh, local_wages, local_wh, nonqualified_plans_12a, state_id
  1099-R: gross_distribution, taxable_amount, capital_gain, federal_wh, distribution_code, state_wh
  1099-NEC: nonemployee_compensation, federal_wh, state_wh
  SSA-1099: net_benefits, federal_wh
  K-1: partnership_name, partnership_ein, partner_name, entity_type, partner_type,
       profit_share_begin, profit_share_end, loss_share_begin, loss_share_end,
       capital_share_begin, capital_share_end,
       box1_ordinary_income, box2_rental_real_estate, box3_other_rental,
       box4a_guaranteed_services, box5_interest, box6a_ordinary_dividends,
       box6b_qualified_dividends, box7_royalties, box8_short_term_capital_gain,
       box9a_long_term_capital_gain, box10_net_1231_gain, box11_other_income,
       box12_section_179, box13_other_deductions, box14_self_employment,
       box15_credits, box17_alt_min_tax, box18_tax_exempt_income,
       box19_distributions, box20_other_info,
       beginning_capital_account, ending_capital_account,
       current_year_net_income, withdrawals_distributions, capital_contributed
  1098: mortgage_interest, property_tax, mortgage_insurance_premiums
  1098-T: payments_received, scholarships_grants
  1098-E: student_loan_interest
  W-2G: gross_winnings, federal_wh, type_of_wager, state_wh, state_winnings
  1099-K: gross_amount, card_not_present_txns, federal_wh, jan through dec (monthly amounts),
       payment_processor, number_of_transactions
  1099-S: gross_proceeds, buyers_part_of_real_estate_tax, address_of_property, date_of_closing
  1099-C: debt_cancelled, date_cancelled, debt_description, fair_market_value, identifiable_event_code
  1099-Q: gross_distribution, earnings, basis, distribution_type (education/rollover)
  1099-LTC: gross_ltc_benefits, accelerated_death_benefits, per_diem, reimbursed
  5498: ira_contributions, rollover_contributions, roth_conversion, recharacterized,
       fair_market_value, rmd_amount, rmd_date
  5498-SA: hsa_contributions, employer_contributions, total_contributions
  1095-A: monthly_premium, monthly_slcsp, monthly_advance_ptc (for each month)
  Schedule C summary: business_name, gross_income, total_expenses, net_profit

For bookkeeping documents:
  Bank Statement: bank_name, account_number, account_number_last4,
       statement_period_start, statement_period_end,
       beginning_balance, total_deposits, total_withdrawals, ending_balance,
       num_deposits, num_withdrawals, fees_charged, interest_earned
       ALSO extract every individual transaction as numbered fields inside "fields":
       txn_1_date, txn_1_desc, txn_1_amount, txn_1_type (deposit/withdrawal/fee/check/transfer)
       txn_2_date, txn_2_desc, txn_2_amount, txn_2_type
       ...continue numbering for ALL visible transactions (up to 50).
       Include check numbers in desc (e.g. "Check #250 - Final Pmt sewer")
  Credit Card Statement: card_issuer, account_number_last4,
       statement_period_start, statement_period_end,
       previous_balance, payments, credits, purchases, fees_charged, interest_charged,
       new_balance, minimum_payment, payment_due_date
       Extract transactions as txn_N_date, txn_N_desc, txn_N_amount, txn_N_category
  Check Stub (Pay Stub): employer_name, employee_name, pay_period_start, pay_period_end, pay_date,
       gross_pay, federal_wh, state_wh, social_security, medicare, net_pay,
       ytd_gross, ytd_federal_wh, ytd_state_wh, ytd_social_security, ytd_medicare, ytd_net_pay,
       hours_regular, hours_overtime, rate_regular, rate_overtime
  Invoice: vendor_name, invoice_number, invoice_date, due_date, subtotal, tax_amount, total_amount,
       description, payment_terms
       For line items: Return as an array in "line_items" with fields: description, quantity, unit_price, amount
  Receipt: vendor_name, receipt_date, subtotal, tax_amount, total_amount, payment_method,
       category (meals, supplies, travel, equipment, utilities, rent, insurance, professional_services, other)
       For line items: Return as an array in "line_items" with fields: description, amount
  Profit & Loss (P&L): period_start, period_end, total_revenue, total_cogs, gross_profit,
       total_operating_expenses, operating_income, net_income
       Extract line items as: rev_N_desc, rev_N_amount; exp_N_desc, exp_N_amount (numbered)
  Balance Sheet: as_of_date, total_assets, total_liabilities, total_equity
       Extract line items as: asset_N_desc, asset_N_amount; liab_N_desc, liab_N_amount (numbered)
  Loan Statement: lender, account_number, original_amount, current_balance, interest_rate,
       payment_amount, payment_date, principal_paid, interest_paid, maturity_date
  Mortgage Statement: lender, property_address, account_number, original_amount,
       current_balance, interest_rate, escrow_balance,
       payment_amount, principal_paid, interest_paid, escrow_paid, next_due_date

For payroll documents:
  Payroll Register: pay_date, total_gross, total_federal_wh, total_state_wh,
       total_social_security, total_medicare, total_net_pay, num_employees
       Extract per-employee: emp_N_name, emp_N_gross, emp_N_federal_wh, emp_N_net (numbered)
  941/940/943/944/945: quarter, tax_period, total_wages, total_federal_tax,
       total_social_security_tax, total_medicare_tax, total_deposits, balance_due

K-1 CRITICAL NOTES:
  Box 2 = Net rental real estate income (loss). NOT credits.
  Box 15 = Credits. These are DIFFERENT boxes. Do not confuse them.
  Small values like -9, -3 are real. Do not skip them.
  If a box says "STMT" or "* STMT", the value is on a continuation statement.

Return JSON only:
{
    "document_type": "<type>",
    "payer_or_entity": "<name>",
    "payer_ein": "<EIN>",
    "recipient": "<person>",
    "tax_year": "<year>",
    "fields": {
        "<field_name>": {
            "value": <number or string>,
            "label_on_form": "<text context around this value in the OCR>",
            "confidence": "<high|medium|low>",
            "ocr_clear": <true if the digits in the OCR text are unambiguous, false if garbled/unclear>
        }
    },
    "continuation_items": [
        {"line_reference": "<ref>", "description": "<desc>", "amount": <number>, "confidence": "<c>"}
    ],
    "ocr_quality": "<good|partial|poor>",
    "needs_image_review": <true if OCR text was too garbled for confident extraction>,
    "fields_needing_image": ["<field names where OCR text was ambiguous>"],
    "flags": [],
    "notes": []
}"""

# Phase 2b: Vision extraction (expensive, used as fallback or for image-needed fields)
VISION_EXTRACTION_PROMPT = """You are a document extraction specialist for a CPA firm.
The OCR text for this page was insufficient, so you are reading directly from the scanned image.

{context}

CRITICAL ACCURACY RULES:
- Read each number DIGIT BY DIGIT. Do not estimate.
- Pay attention to: commas vs periods, 6 vs 8, 1 vs 7, 0 vs 6, 3 vs 8, 4 vs 9.
- If a value has cents (13,664.79), extract EXACT cents.
- DO NOT add or calculate. Only extract what is printed.
- For dotted-line values, follow the dots carefully to the number.

FIELD NAMING RULES — you MUST use these exact field names:
For consolidated brokerage statements (multiple forms on one page), use prefixed names:
  1099-DIV: div_ordinary_dividends, div_qualified_dividends, div_capital_gain_distributions, div_nondividend_distributions, div_federal_wh, div_section_199a, div_foreign_tax_paid, div_exempt_interest, div_private_activity_bond, div_state_wh
  1099-INT: int_interest_income, int_early_withdrawal_penalty, int_us_savings_bonds_and_treasury, int_federal_wh, int_investment_expenses, int_foreign_tax_paid, int_tax_exempt_interest, int_market_discount, int_bond_premium, int_bond_premium_treasury, int_state_wh
  1099-B: b_short_term_proceeds, b_short_term_basis, b_short_term_gain_loss, b_long_term_proceeds, b_long_term_basis, b_long_term_gain_loss, b_total_proceeds, b_total_basis, b_total_gain_loss, b_federal_wh, b_market_discount, b_wash_sale_loss
  1099-MISC: misc_royalties, misc_other_income, misc_rents, misc_federal_wh, misc_state_wh

For standalone forms, use unprefixed names:
  W-2: employer_name, employer_ein, wages, federal_wh, ss_wages, ss_wh, medicare_wages, medicare_wh, state_wages, state_wh, local_wages, local_wh, nonqualified_plans_12a, state_id
  1099-DIV: ordinary_dividends, qualified_dividends, capital_gain_distributions, nondividend_distributions, federal_wh, section_199a, foreign_tax_paid, exempt_interest, private_activity_bond, state_wh
  1099-INT: interest_income, early_withdrawal_penalty, us_savings_bonds_and_treasury, federal_wh, investment_expenses, foreign_tax_paid, tax_exempt_interest, market_discount, bond_premium, state_wh
  1099-R: gross_distribution, taxable_amount, capital_gain, federal_wh, distribution_code, state_wh
  1099-NEC: nonemployee_compensation, federal_wh, state_wh
  SSA-1099: net_benefits, federal_wh
  K-1: partnership_name, partnership_ein, partner_name, entity_type, partner_type,
       profit_share_begin, profit_share_end, loss_share_begin, loss_share_end,
       capital_share_begin, capital_share_end,
       box1_ordinary_income, box2_rental_real_estate, box3_other_rental,
       box4a_guaranteed_services, box5_interest, box6a_ordinary_dividends,
       box6b_qualified_dividends, box7_royalties, box8_short_term_capital_gain,
       box9a_long_term_capital_gain, box10_net_1231_gain, box11_other_income,
       box12_section_179, box13_other_deductions, box14_self_employment,
       box15_credits, box17_alt_min_tax, box18_tax_exempt_income,
       box19_distributions, box20_other_info,
       beginning_capital_account, ending_capital_account,
       current_year_net_income, withdrawals_distributions, capital_contributed

K-1 CRITICAL: Box 2 = rental real estate income. Box 15 = credits. They are DIFFERENT.
Small values like -9 ARE real values.

BOOKKEEPING DOCUMENTS:
For bank statements: Extract summary fields (beginning_balance, ending_balance, total_deposits,
total_withdrawals, fees_charged, interest_earned, statement_period_start, statement_period_end).
ALSO extract every individual transaction as numbered fields in the fields dict:
  txn_1_date, txn_1_desc, txn_1_amount, txn_1_type (deposit/withdrawal/fee/check/transfer)
  txn_2_date, txn_2_desc, txn_2_amount, txn_2_type
  ...continue numbering for ALL visible transactions (up to 50).
For check stubs: extract gross_pay, federal_wh, state_wh, social_security, medicare, net_pay plus YTD fields.
For invoices and receipts: extract vendor_name, date, subtotal, tax_amount, total_amount, and categorize receipts.

Return JSON:
{
    "document_type": "<type>",
    "payer_or_entity": "<name>",
    "payer_ein": "<EIN>",
    "recipient": "<person>",
    "tax_year": "<year>",
    "fields": {
        "<field_name>": {
            "value": <number or string>,
            "label_on_form": "<exact label text as printed>",
            "confidence": "<high|medium|low>"
        }
    },
    "continuation_items": [
        {"line_reference": "<ref>", "description": "<desc>", "amount": <number>, "confidence": "<c>"}
    ],
    "flags": [],
    "notes": []
}"""

# Phase 3: Verification — cross-check OCR text against image for critical fields
VERIFY_CROSSCHECK_PROMPT = """You are verifying a tax data extraction. I have extracted values using OCR text, and now I need you to check the ORIGINAL IMAGE to verify the critical numeric fields.

Here are the extracted values to verify:
{extracted_json}

INSTRUCTIONS:
1. For EACH field listed, find the corresponding value on the scanned image.
2. Read the number digit by digit from the image.
3. Compare to the extracted value.
4. If they match → status "confirmed"
5. If they differ → status "corrected", provide the value you read from the image
6. If you can't find the field on the image → status "unverifiable"
7. Also look for any fields VISIBLE on the image that are NOT in the extraction.

Return JSON:
{
    "verification_results": {
        "<field_name>": {
            "extracted_value": <from OCR>,
            "image_value": <what you see on image>,
            "final_value": <correct value>,
            "status": "<confirmed|corrected|unverifiable>",
            "notes": "<explanation if corrected>"
        }
    },
    "missing_fields": [
        {"field_name": "<name>", "value": <v>, "label_on_form": "<label>", "notes": "On image but not extracted"}
    ],
    "overall_confidence": "<high|medium|low>"
}"""


# ─── TEMPLATE SECTIONS ───────────────────────────────────────────────────────

TEMPLATE_SECTIONS = [
    {
        "id": "w2",
        "header": "W-2:",
        "match_types": ["W-2"],
        "columns": {"A": "employer_name", "B": "wages", "C": "federal_wh", "D": "state_wh"},
        "col_headers": {"B": "Gross", "C": "Federal WH", "D": "State WH"},
        "sum_cols": ["B", "C", "D"],
    },
    {
        "id": "interest",
        "header": "Interest Income:",
        "match_types": ["1099-INT", "_interest_rollup"],
        "columns": {"A": "_source_name", "B": "interest_income", "C": "us_savings_bonds_and_treasury"},
        "col_headers": {"B": "Interest", "C": "US Bonds"},
        "sum_cols": ["B", "C"],
        "total_formula_col": {"D": "B+C"},
    },
    {
        "id": "dividends",
        "header": "Dividends:",
        "match_types": ["1099-DIV", "_dividend_rollup"],
        "columns": {"A": "_source_name", "B": "ordinary_dividends", "C": "qualified_dividends",
                     "D": "capital_gain_distributions", "E": "section_199a"},
        "col_headers": {"B": "Total Ord", "C": "Qualified", "D": "Cap Gain Dist", "E": "Sec 199A"},
        "sum_cols": ["B", "C", "D", "E"],
    },
    {
        "id": "1099r",
        "header": "1099-R:",
        "match_types": ["1099-R"],
        "columns": {"A": "payer_or_entity", "B": "gross_distribution", "C": "taxable_amount",
                     "D": "federal_wh", "E": "state_wh", "F": "distribution_code"},
        "col_headers": {"B": "Gross", "C": "Taxable", "D": "FWH", "E": "SWH", "F": "Code"},
        "sum_cols": ["B", "C", "D", "E"],
    },
    {
        "id": "ssa",
        "header": "SSA-1099:",
        "match_types": ["SSA-1099"],
        "columns": {"A": "payer_or_entity", "B": "net_benefits", "C": "federal_wh"},
        "col_headers": {"B": "Net Benefits", "C": "Federal WH"},
        "sum_cols": ["B", "C"],
    },
    {
        "id": "1099nec",
        "header": "1099-NEC (Self-Employment):",
        "match_types": ["1099-NEC"],
        "columns": {"A": "payer_or_entity", "B": "nonemployee_compensation", "C": "federal_wh"},
        "col_headers": {"B": "NEC Income", "C": "FWH"},
        "sum_cols": ["B", "C"],
    },
    {
        "id": "1099misc",
        "header": "1099-MISC:",
        "match_types": ["1099-MISC"],
        "columns": {"A": "payer_or_entity", "B": "rents", "C": "royalties", "D": "other_income", "E": "federal_wh"},
        "col_headers": {"B": "Rents", "C": "Royalties", "D": "Other Income", "E": "FWH"},
        "sum_cols": ["B", "C", "D", "E"],
    },
    {
        "id": "1099g",
        "header": "1099-G:",
        "match_types": ["1099-G"],
        "columns": {"A": "payer_or_entity", "B": "unemployment", "C": "state_local_refund", "D": "federal_wh"},
        "col_headers": {"B": "Unemployment", "C": "State Refund", "D": "FWH"},
        "sum_cols": ["B", "C", "D"],
    },
    {
        "id": "schedule_d",
        "header": "Schedule D:",
        "match_types": ["1099-B", "_brokerage_gains"],
        "columns": {"A": "_source_name", "B": "total_proceeds", "C": "total_basis",
                     "D": "wash_sale_loss", "E": "total_gain_loss"},
        "col_headers": {"B": "Proceeds", "C": "Cost Basis", "D": "Wash Sale", "E": "Net Gain/Loss"},
        "sum_cols": ["B", "C", "D", "E"],
        "flags": ["⚠ Check for capital loss carryover from prior year"],
    },
    {
        "id": "k1",
        "header": "K-1s:",
        "match_types": ["K-1"],
        "columns": {"A": "_display_name", "B": "_carryover_prior", "C": "box1_ordinary_income",
                     "D": "box2_rental_real_estate", "E": "_allowed", "F": "_carryover_next"},
        "col_headers": {"B": "C/O from PY", "C": "Box 1", "D": "Box 2", "E": "Allowed", "F": "C/O to NY"},
        "sum_cols": ["B", "C", "D", "E", "F"],
        "flags": [
            "⚠ Column B (PY carryover): REQUIRES prior year data — enter manually",
            "⚠ Column E (Allowed): REQUIRES basis/at-risk/passive analysis — enter manually",
            "⚠ Column F (NY carryover): REQUIRES basis/at-risk/passive analysis — enter manually",
        ],
    },
    {
        "id": "k1_detail",
        "header": "K-1 Additional Detail:",
        "special": "k1_extras",
    },
    {
        "id": "rental",
        "header": "Rental Income (Schedule E):",
        "match_types": ["rental_income_document"],
        "columns": {"A": "property_address", "B": "gross_rents", "C": "total_expenses", "D": "net_rental_income"},
        "col_headers": {"B": "Gross Rents", "C": "Expenses", "D": "Net Income"},
        "sum_cols": ["B", "C", "D"],
    },
    {
        "id": "farm",
        "header": "Farm Income (Schedule F):",
        "match_types": ["farm_income_document"],
        "columns": {"A": "description", "B": "gross_farm_income", "C": "farm_expenses", "D": "net_farm_income"},
        "col_headers": {"B": "Gross", "C": "Expenses", "D": "Net"},
        "sum_cols": ["B", "C", "D"],
    },
    {
        "id": "1098",
        "header": "Form 1098 (Mortgage Interest):",
        "match_types": ["1098"],
        "columns": {"A": "payer_or_entity", "B": "mortgage_interest", "C": "property_tax", "D": "mortgage_insurance_premiums"},
        "col_headers": {"B": "Mortgage Int", "C": "Property Tax", "D": "PMI"},
        "sum_cols": ["B", "C", "D"],
        "field_aliases": {"mortgage_interest": ["mortgage_interest_received"],
                          "property_tax": ["real_estate_tax", "property_taxes_paid"]},
    },
    {
        "id": "1098t",
        "header": "1098-T (Tuition):",
        "match_types": ["1098-T"],
        "columns": {"A": "institution_name", "B": "payments_received", "C": "scholarships_grants"},
        "col_headers": {"B": "Box 1", "C": "Box 5"},
        "sum_cols": ["B", "C"],
    },
    {
        "id": "property_tax",
        "header": "Property Tax Bills:",
        "match_types": ["property_tax_bill"],
        "columns": {"A": "property_address", "B": "tax_amount"},
        "field_aliases": {"tax_amount": ["total_due", "total_tax", "total_amount_due", "total_estimated_tax"]},
        "col_headers": {"B": "Tax Amount"},
        "sum_cols": ["B"],
    },
    {
        "id": "estimated",
        "header": "Estimated Tax Payments:",
        "match_types": ["estimated_tax_record"],
        "columns": {"A": "payment_date", "B": "federal_amount", "C": "state_amount"},
        "col_headers": {"B": "Federal", "C": "State"},
        "sum_cols": ["B", "C"],
        "flags": ["⚠ Estimated payments often NOT on scanned docs — verify with client"],
    },
    {
        "id": "charitable",
        "header": "Charitable Contributions:",
        "match_types": ["charitable_receipt"],
        "columns": {"A": "organization_name", "B": "donation_amount", "C": "donation_type"},
        "col_headers": {"B": "Amount", "C": "Type"},
        "sum_cols": ["B"],
    },
    # ─── Additional Tax Forms ───
    {
        "id": "w2g",
        "header": "W-2G (Gambling Winnings):",
        "match_types": ["W-2G"],
        "columns": {"A": "payer_or_entity", "B": "gross_winnings", "C": "federal_wh",
                     "D": "type_of_wager", "E": "state_wh"},
        "col_headers": {"B": "Winnings", "C": "FWH", "D": "Type", "E": "SWH"},
        "sum_cols": ["B", "C", "E"],
    },
    {
        "id": "1099k",
        "header": "1099-K (Payment Card / Third Party):",
        "match_types": ["1099-K"],
        "columns": {"A": "payer_or_entity", "B": "gross_amount", "C": "number_of_transactions",
                     "D": "federal_wh"},
        "col_headers": {"B": "Gross Amount", "C": "# Txns", "D": "FWH"},
        "sum_cols": ["B", "D"],
    },
    {
        "id": "1099s",
        "header": "1099-S (Real Estate Proceeds):",
        "match_types": ["1099-S"],
        "columns": {"A": "address_of_property", "B": "gross_proceeds", "C": "date_of_closing",
                     "D": "buyers_part_of_real_estate_tax"},
        "col_headers": {"B": "Gross Proceeds", "C": "Closing Date", "D": "RE Tax"},
        "sum_cols": ["B", "D"],
    },
    {
        "id": "1099c",
        "header": "1099-C (Cancellation of Debt):",
        "match_types": ["1099-C"],
        "columns": {"A": "payer_or_entity", "B": "debt_cancelled", "C": "date_cancelled",
                     "D": "fair_market_value", "E": "debt_description"},
        "col_headers": {"B": "Debt Cancelled", "C": "Date", "D": "FMV", "E": "Description"},
        "sum_cols": ["B", "D"],
        "flags": ["⚠ Check insolvency exclusion — may reduce taxable amount"],
    },
    {
        "id": "1098e",
        "header": "1098-E (Student Loan Interest):",
        "match_types": ["1098-E"],
        "columns": {"A": "payer_or_entity", "B": "student_loan_interest"},
        "col_headers": {"B": "Interest Paid"},
        "sum_cols": ["B"],
    },
    {
        "id": "5498",
        "header": "5498 (IRA Contributions):",
        "match_types": ["5498"],
        "columns": {"A": "payer_or_entity", "B": "ira_contributions", "C": "rollover_contributions",
                     "D": "roth_conversion", "E": "fair_market_value", "F": "rmd_amount"},
        "col_headers": {"B": "Contributions", "C": "Rollovers", "D": "Roth Conv", "E": "FMV", "F": "RMD"},
        "sum_cols": ["B", "C", "D"],
    },
    {
        "id": "schedule_c",
        "header": "Schedule C (Business Income):",
        "match_types": ["schedule_c_summary"],
        "columns": {"A": "business_name", "B": "gross_income", "C": "total_expenses", "D": "net_profit"},
        "col_headers": {"B": "Gross Income", "C": "Expenses", "D": "Net Profit"},
        "sum_cols": ["B", "C", "D"],
    },
    # ─── Bookkeeping Sections ───
    {
        "id": "bank_statement",
        "header": "Bank Statements:",
        "match_types": ["bank_statement"],
        "columns": {"A": "bank_name", "B": "beginning_balance", "C": "total_deposits",
                     "D": "total_withdrawals", "E": "fees_charged", "F": "interest_earned", "G": "ending_balance"},
        "col_headers": {"B": "Begin Bal", "C": "Deposits", "D": "Withdrawals", "E": "Fees", "F": "Interest", "G": "End Bal"},
        "sum_cols": ["C", "D", "E", "F"],
    },
    {
        "id": "credit_card",
        "header": "Credit Card Statements:",
        "match_types": ["credit_card_statement"],
        "columns": {"A": "card_issuer", "B": "previous_balance", "C": "purchases",
                     "D": "payments", "E": "interest_charged", "F": "fees_charged", "G": "new_balance"},
        "col_headers": {"B": "Prev Bal", "C": "Purchases", "D": "Payments", "E": "Interest", "F": "Fees", "G": "New Bal"},
        "sum_cols": ["C", "D", "E", "F"],
    },
    {
        "id": "check_stub",
        "header": "Pay Stubs / Check Stubs:",
        "match_types": ["check_stub"],
        "columns": {"A": "employer_name", "B": "gross_pay", "C": "federal_wh",
                     "D": "state_wh", "E": "social_security", "F": "medicare", "G": "net_pay"},
        "col_headers": {"B": "Gross Pay", "C": "FWH", "D": "SWH", "E": "SS", "F": "Medicare", "G": "Net Pay"},
        "sum_cols": ["B", "C", "D", "E", "F", "G"],
    },
    {
        "id": "invoice",
        "header": "Invoices:",
        "match_types": ["invoice"],
        "columns": {"A": "vendor_name", "B": "invoice_number", "C": "invoice_date",
                     "D": "subtotal", "E": "tax_amount", "F": "total_amount"},
        "col_headers": {"B": "Invoice #", "C": "Date", "D": "Subtotal", "E": "Tax", "F": "Total"},
        "sum_cols": ["D", "E", "F"],
    },
    {
        "id": "receipt",
        "header": "Receipts:",
        "match_types": ["receipt"],
        "columns": {"A": "vendor_name", "B": "receipt_date", "C": "category",
                     "D": "subtotal", "E": "tax_amount", "F": "total_amount"},
        "col_headers": {"B": "Date", "C": "Category", "D": "Subtotal", "E": "Tax", "F": "Total"},
        "sum_cols": ["D", "E", "F"],
    },
    # ─── Financial Statements ───
    {
        "id": "profit_loss",
        "header": "Profit & Loss Statements:",
        "match_types": ["profit_loss_statement"],
        "columns": {"A": "payer_or_entity", "B": "total_revenue", "C": "total_cogs",
                     "D": "gross_profit", "E": "total_operating_expenses", "F": "net_income"},
        "col_headers": {"B": "Revenue", "C": "COGS", "D": "Gross Profit", "E": "Op Expenses", "F": "Net Income"},
        "sum_cols": ["B", "C", "D", "E", "F"],
    },
    {
        "id": "balance_sheet",
        "header": "Balance Sheets:",
        "match_types": ["balance_sheet"],
        "columns": {"A": "payer_or_entity", "B": "total_assets", "C": "total_liabilities",
                     "D": "total_equity"},
        "col_headers": {"B": "Total Assets", "C": "Total Liabilities", "D": "Total Equity"},
        "sum_cols": ["B", "C", "D"],
    },
    {
        "id": "loan_statement",
        "header": "Loan Statements:",
        "match_types": ["loan_statement", "mortgage_statement"],
        "columns": {"A": "lender", "B": "current_balance", "C": "interest_rate",
                     "D": "payment_amount", "E": "principal_paid", "F": "interest_paid"},
        "col_headers": {"B": "Balance", "C": "Rate", "D": "Payment", "E": "Principal", "F": "Interest"},
        "sum_cols": ["D", "E", "F"],
    },
    # ─── Payroll Sections ───
    {
        "id": "payroll_register",
        "header": "Payroll Registers:",
        "match_types": ["payroll_register", "payroll_summary"],
        "columns": {"A": "payer_or_entity", "B": "total_gross", "C": "total_federal_wh",
                     "D": "total_state_wh", "E": "total_social_security", "F": "total_medicare", "G": "total_net_pay"},
        "col_headers": {"B": "Gross", "C": "FWH", "D": "SWH", "E": "SS", "F": "Medicare", "G": "Net Pay"},
        "sum_cols": ["B", "C", "D", "E", "F", "G"],
    },
    {
        "id": "payroll_tax",
        "header": "Payroll Tax Forms (940/941/943/944/945):",
        "match_types": ["940", "941", "943", "944", "945"],
        "columns": {"A": "payer_or_entity", "B": "total_wages", "C": "total_federal_tax",
                     "D": "total_social_security_tax", "E": "total_medicare_tax", "F": "balance_due"},
        "col_headers": {"B": "Wages", "C": "Federal Tax", "D": "SS Tax", "E": "Medicare Tax", "F": "Bal Due"},
        "sum_cols": ["B", "C", "D", "E", "F"],
    },
]

ALWAYS_SHOW = ["w2", "interest", "dividends", "schedule_d", "k1"]

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def get_val(fields, key):
    fdata = fields.get(key)
    if fdata is None: return None
    v = fdata.get("value") if isinstance(fdata, dict) else fdata
    try: return float(v) if v is not None else None
    except (ValueError, TypeError): return None

def get_str(fields, key):
    fdata = fields.get(key)
    if fdata is None: return None
    v = fdata.get("value") if isinstance(fdata, dict) else fdata
    return str(v) if v is not None else None

def auto_rotate(img):
    """Detect and fix rotated pages. Returns corrected PIL image.

    Strategy:
    - Landscape pages (w > h * 1.15): always use OSD (any confidence)
    - Portrait pages with 90/270 OSD: apply rotation (content is sideways)
    - Portrait pages with 180 OSD: only rotate if conf >= 10 (180 on portrait
      pages is usually a Tesseract false positive — upright text misread)

    Performance: OSD runs on a downscaled thumbnail (~800px) for speed,
    then rotation is applied to the full-resolution image."""
    w, h = img.size
    is_landscape = w > h * 1.15
    if HAS_TESSERACT:
        try:
            # Use downscaled image for fast OSD detection
            max_dim = max(w, h)
            if max_dim > 1000:
                scale = 800 / max_dim
                osd_img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
            else:
                osd_img = img
            osd = pytesseract.image_to_osd(osd_img, output_type=pytesseract.Output.DICT)
            if osd_img is not img:
                osd_img.close()
            angle = osd.get("rotate", 0)
            conf = float(osd.get("orientation_conf", 0))
            if angle != 0:
                if is_landscape or angle in (90, 270) or conf >= 10:
                    img = img.rotate(-angle, expand=True)
                    return img
                # Portrait + 180deg + low confidence: skip (likely false positive)
            return img
        except Exception as e:
            if is_landscape:
                print(f"  Auto-rotate: Tesseract OSD failed ({e}), rotating 90 CW as fallback")
                img = img.rotate(-90, expand=True)
    else:
        if is_landscape:
            img = img.rotate(-90, expand=True)
    return img


# ─── PAGE PREPROCESSING (T1.1) ──────────────────────────────────────────────

def _is_blank_page(img, page_text=None):
    """Detect blank/near-blank pages by checking non-white pixel percentage.

    Text override: If page_text is provided and contains meaningful content
    (≥40 chars, or money patterns like $X,XXX.XX, or IRS keywords), the page
    is NOT blank even if pixel density is low. This handles sparse-but-meaningful
    pages like K-1 cover sheets or brokerage summaries with a few dollar amounts.

    Returns (is_blank: bool, pct_non_white: float, blank_reason: str).
    blank_reason: "blank_true" | "blank_low_value" | "not_blank" | "overridden_by_text"
    """
    try:
        gray = img.convert("L")
        # Downsample for speed — 400px wide is plenty for blank detection
        w, h = gray.size
        if w > 400:
            scale = 400 / w
            gray = gray.resize((400, int(h * scale)), Image.LANCZOS)
        pixels = list(gray.getdata())
        total = len(pixels)
        if total == 0:
            return True, 0.0, "blank_true"
        # Count pixels darker than 240 (not white/near-white)
        non_white = sum(1 for p in pixels if p < 240)
        pct = (non_white / total) * 100
        gray.close()

        if pct >= BLANK_PAGE_THRESHOLD:
            return False, pct, "not_blank"

        # Below pixel threshold — check text override
        if page_text and len(page_text.strip()) >= BLANK_MIN_OCR_CHARS:
            return False, pct, "overridden_by_text"

        # Check for money patterns or IRS keywords in page_text
        if page_text:
            text = page_text.strip()
            money_pattern = re.compile(r'\$[\d,]+\.?\d*')
            irs_keywords = {"1099", "W-2", "K-1", "Schedule", "Form", "Box",
                            "income", "dividends", "interest", "withholding",
                            "wages", "EIN", "SSN", "tax"}
            has_money = bool(money_pattern.search(text))
            has_irs = any(kw.lower() in text.lower() for kw in irs_keywords)
            if has_money or has_irs:
                return False, pct, "overridden_by_text"

        # Truly blank or low-value page
        if pct == 0:
            return True, pct, "blank_true"
        else:
            return True, pct, "blank_low_value"

    except Exception:
        return False, 100.0, "not_blank"


def _compute_quality_score(img):
    """Compute a 0–1 quality score for a page image based on contrast and sharpness.

    Uses pure PIL (no numpy required). Contrast is measured over the non-white
    bounding box crop (masked std_dev) so that sparse text pages aren't penalized
    by large white margins.

    Factors:
      - Masked std deviation (contrast within content area)
      - Edge density (sharpness) — blurry pages have few edges via ImageFilter
      - Text-area coverage estimate — mostly-white pages score lower

    Returns (score: float, details: dict)."""
    try:
        gray = img.convert("L")
        w, h = gray.size
        # Downsample for speed
        if w > 600:
            scale = 600 / w
            gray = gray.resize((600, int(h * scale)), Image.LANCZOS)

        dw, dh = gray.size
        pixels = list(gray.getdata())
        total = len(pixels)
        if total == 0:
            gray.close()
            return 0.0, {"error": "empty image"}

        # Find non-white bounding box for masked std_dev
        # Scan rows/cols to find extent of non-white content (< 240)
        min_x, min_y, max_x, max_y = dw, dh, 0, 0
        for y_idx in range(dh):
            row_start = y_idx * dw
            for x_idx in range(dw):
                if pixels[row_start + x_idx] < 240:
                    min_x = min(min_x, x_idx)
                    min_y = min(min_y, y_idx)
                    max_x = max(max_x, x_idx)
                    max_y = max(max_y, y_idx)

        has_content_box = max_x > min_x and max_y > min_y
        if has_content_box:
            # Compute std_dev over the content bounding box crop
            crop_pixels = []
            for y_idx in range(min_y, max_y + 1):
                row_start = y_idx * dw
                for x_idx in range(min_x, max_x + 1):
                    crop_pixels.append(pixels[row_start + x_idx])
            crop_total = len(crop_pixels)
            crop_mean = sum(crop_pixels) / crop_total
            crop_var = sum((p - crop_mean) ** 2 for p in crop_pixels) / crop_total
            masked_std = crop_var ** 0.5
        else:
            # No content found — use whole-page
            masked_std = 0.0

        # Also compute whole-page std_dev (for reference/logging)
        mean_val = sum(pixels) / total
        variance = sum((p - mean_val) ** 2 for p in pixels) / total
        whole_std = variance ** 0.5

        # Use masked std_dev for scoring (more accurate for sparse pages)
        std_dev = masked_std if has_content_box else whole_std
        # Normalize: masked std_dev of 70+ is good (content area has real contrast),
        # below 20 is poor (faded within the content region itself)
        contrast_score = min(1.0, max(0.0, (std_dev - 15) / 60))

        # Sharpness: edge density via PIL FIND_EDGES filter
        from PIL import ImageFilter
        edges = gray.filter(ImageFilter.FIND_EDGES)
        edge_pixels = list(edges.getdata())
        edges.close()
        edge_mean = sum(edge_pixels) / len(edge_pixels)
        # Normalize: edge mean of 15+ is sharp, below 3 is blurry
        sharpness_score = min(1.0, max(0.0, (edge_mean - 2) / 15))

        # Content coverage: % of pixels in text range (30-200)
        text_pixels = sum(1 for p in pixels if 30 < p < 200)
        coverage = text_pixels / total
        coverage_score = min(1.0, coverage * 5)  # 20% coverage = full score

        gray.close()

        # Weighted combination
        score = 0.4 * contrast_score + 0.4 * sharpness_score + 0.2 * coverage_score

        details = {
            "std_dev_masked": round(std_dev, 1),
            "std_dev_whole": round(whole_std, 1),
            "edge_mean": round(edge_mean, 1),
            "contrast_score": round(contrast_score, 3),
            "sharpness_score": round(sharpness_score, 3),
            "coverage_score": round(coverage_score, 3),
            "content_coverage_pct": round(coverage * 100, 1),
            "content_bbox": [min_x, min_y, max_x, max_y] if has_content_box else None,
        }
        return round(score, 3), details

    except Exception as e:
        return 0.5, {"error": str(e)}


def _deskew_page(img):
    """Detect and correct small skew angles (< DESKEW_MAX_ANGLE degrees).
    Uses Tesseract OSD for angle detection on a downscaled thumbnail.
    Returns (corrected_img, angle_applied, confidence)."""
    if not HAS_TESSERACT:
        return img, 0.0, 0.0

    try:
        w, h = img.size
        # Downscale for OSD speed
        max_dim = max(w, h)
        if max_dim > 1000:
            scale = 800 / max_dim
            osd_img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
        else:
            osd_img = img

        osd = pytesseract.image_to_osd(osd_img, output_type=pytesseract.Output.DICT)
        if osd_img is not img:
            osd_img.close()

        # OSD reports rotation in 90-degree increments (handled by auto_rotate).
        # For deskew, we look at the script orientation angle if available,
        # or use pytesseract's detailed angle detection.
        # Since standard OSD doesn't give sub-degree skew, we use image_to_osd
        # "rotate" for coarse and rely on the confidence + angle for fine tuning.
        angle = float(osd.get("rotate", 0))
        conf = float(osd.get("orientation_conf", 0))

        # Only handle small residual skew (not 90/180/270 — that's auto_rotate's job)
        if angle in (0, 90, 180, 270):
            # No sub-degree skew detected by OSD — this is expected
            # For fine deskew, try a projection-profile approach
            return img, 0.0, conf

        # Small angle detected
        if abs(angle) <= DESKEW_MAX_ANGLE and conf >= DESKEW_MIN_CONFIDENCE:
            img = img.rotate(-angle, expand=True, fillcolor=(255, 255, 255))
            return img, angle, conf

        return img, 0.0, conf
    except Exception:
        return img, 0.0, 0.0


def _enhance_contrast(img):
    """Boost contrast on low-contrast pages. Returns (enhanced_img, was_enhanced).
    Uses masked std_dev (non-white bounding box) so sparse-text pages with good
    local contrast aren't falsely enhanced. Guardrail: pages with < 0.2% non-white
    pixels skip enhancement (nearly blank — nothing meaningful to boost).
    Uses pure PIL (no numpy required)."""
    try:
        gray = img.convert("L")
        w, h = gray.size
        # Quick check on downsampled version
        if w > 600:
            scale = 600 / w
            check_img = gray.resize((600, int(h * scale)), Image.LANCZOS)
        else:
            check_img = gray

        dw, dh = check_img.size
        pixels = list(check_img.getdata())
        if check_img is not gray:
            check_img.close()
        gray.close()

        total = len(pixels)
        if total == 0:
            return img, False

        # Guardrail: skip near-blank pages (< 0.2% non-white) — nothing to enhance
        non_white_count = sum(1 for p in pixels if p < 240)
        non_white_pct = (non_white_count / total) * 100
        if non_white_pct < 0.2:
            return img, False

        # Masked std_dev: compute over non-white bounding box crop
        min_x, min_y, max_x, max_y = dw, dh, 0, 0
        for y_idx in range(dh):
            row_start = y_idx * dw
            for x_idx in range(dw):
                if pixels[row_start + x_idx] < 240:
                    min_x = min(min_x, x_idx)
                    min_y = min(min_y, y_idx)
                    max_x = max(max_x, x_idx)
                    max_y = max(max_y, y_idx)

        if max_x > min_x and max_y > min_y:
            crop_pixels = []
            for y_idx in range(min_y, max_y + 1):
                row_start = y_idx * dw
                for x_idx in range(min_x, max_x + 1):
                    crop_pixels.append(pixels[row_start + x_idx])
            crop_total = len(crop_pixels)
            crop_mean = sum(crop_pixels) / crop_total
            crop_var = sum((p - crop_mean) ** 2 for p in crop_pixels) / crop_total
            std_dev = crop_var ** 0.5
        else:
            # No content bounding box — use whole-page
            mean_val = sum(pixels) / total
            variance = sum((p - mean_val) ** 2 for p in pixels) / total
            std_dev = variance ** 0.5

        if std_dev >= CONTRAST_MIN_STD:
            return img, False  # Content area has good contrast

        # Apply autocontrast enhancement using PIL
        from PIL import ImageOps
        enhanced = ImageOps.autocontrast(img, cutoff=1)
        return enhanced, True

    except Exception:
        return img, False
    except Exception:
        return img, False


def preprocess_page(img, page_num=None, page_text=None):
    """Run full preprocessing on a single page image.

    Steps (in order):
      1. Blank page detection → skip if blank (with text override)
      2. Deskew (small angle correction)
      3. Contrast enhancement (if needed)
      4. Quality score computation

    Args:
        img: PIL Image (already auto-rotated by auto_rotate())
        page_num: 1-based page number (for logging)
        page_text: Optional text from text-layer or OCR. Used to override
                   blank detection on sparse-but-meaningful pages.

    Returns:
        (processed_img, metadata_dict)
        metadata_dict keys:
            is_blank, blank_reason, pct_non_white, deskew_angle, deskew_conf,
            contrast_enhanced, quality_score, quality_details, dpi,
            original_size, processed_size
    """
    tag = f"Page {page_num}" if page_num else "Page"
    meta = {
        "page_num": page_num,
        "is_blank": False,
        "blank_reason": "not_blank",
        "pct_non_white": 0.0,
        "deskew_angle": 0.0,
        "deskew_conf": 0.0,
        "contrast_enhanced": False,
        "quality_score": 0.5,
        "quality_details": {},
        "original_size": list(img.size),
    }

    # 1. Blank detection (with text override for sparse-but-meaningful pages)
    is_blank, pct_non_white, blank_reason = _is_blank_page(img, page_text=page_text)
    meta["is_blank"] = is_blank
    meta["blank_reason"] = blank_reason
    meta["pct_non_white"] = round(pct_non_white, 2)
    if is_blank:
        meta["quality_score"] = 0.0
        meta["processed_size"] = list(img.size)
        return img, meta

    # 2. Deskew
    img, deskew_angle, deskew_conf = _deskew_page(img)
    meta["deskew_angle"] = round(deskew_angle, 2)
    meta["deskew_conf"] = round(deskew_conf, 1)

    # 3. Contrast enhancement
    img, was_enhanced = _enhance_contrast(img)
    meta["contrast_enhanced"] = was_enhanced

    # 4. Quality score
    score, details = _compute_quality_score(img)
    meta["quality_score"] = score
    meta["quality_details"] = details
    meta["processed_size"] = list(img.size)

    return img, meta


def extract_text_per_page(pdf_path):
    """Extract embedded text layer from a PDF using PyMuPDF (fitz).
    Returns a list of strings, one per page. Pages without text return ''.
    This is instant (no OCR) and only works on digitally-generated PDFs."""
    if not HAS_PYMUPDF:
        return None
    try:
        doc = fitz.open(pdf_path)
        page_texts = []
        for page in doc:
            page_texts.append(page.get_text("text") or "")
        doc.close()
        return page_texts
    except Exception as e:
        print(f"  PyMuPDF text extraction failed: {e}")
        return None


def has_meaningful_text(page_texts):
    """Determine if the PDF has a usable embedded text layer.
    Thresholds:
      - Each page needs TEXT_MIN_CHARS_PER_PAGE (200) chars to count as "meaningful"
      - At least min(2, ceil(n_pages * 0.25)) pages must be meaningful
      - Total chars across meaningful pages >= TEXT_MIN_TOTAL_CHARS (800)
    Returns (bool, stats_dict) — stats dict has counts for logging."""
    if not page_texts:
        return False, {"meaningful_pages": 0, "total_pages": 0, "total_chars": 0, "reason": "no_text_data"}

    n_pages = len(page_texts)
    meaningful_pages = 0
    total_chars = 0
    per_page_chars = []

    for text in page_texts:
        char_count = len(text.strip())
        per_page_chars.append(char_count)
        if char_count >= TEXT_MIN_CHARS_PER_PAGE:
            meaningful_pages += 1
            total_chars += char_count

    min_required = min(2, math.ceil(n_pages * 0.25))
    stats = {
        "meaningful_pages": meaningful_pages,
        "total_pages": n_pages,
        "total_chars": total_chars,
        "min_pages_required": min_required,
        "per_page_chars": per_page_chars,
    }

    if meaningful_pages < min_required:
        stats["reason"] = f"too_few_meaningful_pages ({meaningful_pages}/{min_required})"
        return False, stats
    if total_chars < TEXT_MIN_TOTAL_CHARS:
        stats["reason"] = f"too_few_total_chars ({total_chars}/{TEXT_MIN_TOTAL_CHARS})"
        return False, stats

    stats["reason"] = "pass"
    return True, stats


def pdf_to_images(pdf_path, dpi=DPI, page_texts=None):
    """Convert PDF to base64 JPEG strings (for API). Auto-rotates sideways pages,
    then preprocesses each page (deskew, contrast, blank detection, quality scoring).
    PIL images are freed after encoding to minimize memory usage.

    Args:
        pdf_path: Path to PDF file
        dpi: Render DPI (default 250)
        page_texts: Optional list[str] from extract_text_per_page(). Used to override
                    blank detection on sparse-but-meaningful pages.

    Returns:
        (b64_images, page_preprocessing)
        b64_images: list[str] — base64 JPEG per page (blank pages included but flagged)
        page_preprocessing: list[dict] — per-page preprocessing metadata
    """
    print(f"Converting PDF at {dpi} DPI...")
    convert_kwargs = {"dpi": dpi}
    if POPPLER_PATH:
        convert_kwargs["poppler_path"] = POPPLER_PATH
    raw_images = convert_from_path(pdf_path, **convert_kwargs)
    b64_images = []
    page_preprocessing = []
    rotated_count = 0
    blank_count = 0
    enhanced_count = 0
    deskewed_count = 0
    quality_scores = []

    print(f"\n── [PASSION] Preprocessing ({len(raw_images)} pages) ──")
    for i, img in enumerate(raw_images):
        page_num = i + 1
        orig_size = img.size

        # Step 1: Auto-rotate (existing — handles 90/180/270)
        img = auto_rotate(img)
        was_rotated = img.size != orig_size
        if was_rotated:
            rotated_count += 1

        # Step 2: Preprocess (deskew, contrast, blank detection, quality score)
        pt = page_texts[i] if page_texts and i < len(page_texts) else None
        img, preprocess_meta = preprocess_page(img, page_num=page_num, page_text=pt)
        preprocess_meta["dpi"] = dpi
        preprocess_meta["was_rotated"] = was_rotated
        if was_rotated:
            preprocess_meta["rotation_from"] = list(orig_size)
            preprocess_meta["rotation_to"] = list(img.size)

        # Track stats
        if preprocess_meta["is_blank"]:
            blank_count += 1
        if preprocess_meta["contrast_enhanced"]:
            enhanced_count += 1
        if preprocess_meta["deskew_angle"] != 0:
            deskewed_count += 1
        quality_scores.append(preprocess_meta["quality_score"])

        # Encode to base64 JPEG (even blank pages — downstream uses page index)
        buf = BytesIO()
        img.save(buf, format="JPEG", quality=85)
        b64_images.append(base64.standard_b64encode(buf.getvalue()).decode("utf-8"))
        page_preprocessing.append(preprocess_meta)
        img.close()  # Free PIL image memory

        # Per-page status line
        flags = []
        if preprocess_meta["is_blank"]:
            flags.append("BLANK")
        if was_rotated:
            flags.append("rotated")
        if preprocess_meta["deskew_angle"] != 0:
            flags.append(f"deskew {preprocess_meta['deskew_angle']}°")
        if preprocess_meta["contrast_enhanced"]:
            flags.append("contrast+")
        q = preprocess_meta["quality_score"]
        if q < QUALITY_POOR:
            flags.append(f"quality={q:.2f} LOW")
        elif q < QUALITY_GOOD:
            flags.append(f"quality={q:.2f}")

        if flags:
            print(f"  Page {page_num}: {', '.join(flags)}")

    del raw_images  # Release reference to all raw images

    # Summary
    parts = [f"{len(b64_images)} pages converted"]
    if rotated_count:
        parts.append(f"{rotated_count} rotated")
    if blank_count:
        parts.append(f"{blank_count} blank")
    if deskewed_count:
        parts.append(f"{deskewed_count} deskewed")
    if enhanced_count:
        parts.append(f"{enhanced_count} contrast-enhanced")
    if quality_scores:
        avg_q = sum(quality_scores) / len(quality_scores)
        parts.append(f"avg quality={avg_q:.2f}")
    print(f"  {' | '.join(parts)}")

    return b64_images, page_preprocessing

def ocr_page(pil_image, page_num=None):
    """Run Tesseract OCR on a PIL image. Returns (text, conf_avg) tuple.
    text: OCR text string or None. conf_avg: average word confidence (0-100) or None.
    Pages are already auto-rotated in pdf_to_images(), so we only need one pass."""
    if not HAS_TESSERACT:
        return None, None
    tag = f"[Page {page_num}] " if page_num else ""
    try:
        text = pytesseract.image_to_string(pil_image, config='--oem 3 --psm 6')
        if text and len(text.strip()) >= OCR_MIN_CHARS:
            conf_avg = _ocr_confidence(pil_image)
            return text, conf_avg
        # Fallback: try auto-orientation mode for pages that may have mixed orientations
        text2 = pytesseract.image_to_string(pil_image, config='--oem 3 --psm 1')
        if text2 and len(text2.strip()) > len((text or "").strip()):
            text = text2
        if text and len(text.strip()) > 30:
            conf_avg = _ocr_confidence(pil_image)
            return text, conf_avg
        print(f"  {tag}OCR: too little text ({len(text.strip()) if text else 0} chars)")
        return None, None
    except Exception as e:
        print(f"  {tag}OCR error: {e}")
        return None, None


def _ocr_confidence(pil_image):
    """Extract average per-word OCR confidence from Tesseract image_to_data.
    Returns float (0-100) or None if unavailable."""
    try:
        data = pytesseract.image_to_data(pil_image, output_type=pytesseract.Output.DICT,
                                          config='--oem 3 --psm 6')
        confs = [c for c in data.get("conf", []) if isinstance(c, (int, float)) and c >= 0]
        if confs:
            return round(sum(confs) / len(confs), 1)
        return None
    except Exception:
        return None

def ocr_all_pages(pil_images):
    """OCR every page upfront using parallel threads.
    Returns (ocr_texts, ocr_confidences) — two parallel lists.
    ocr_texts[i]: text string or None. ocr_confidences[i]: avg confidence (0-100) or None.
    Entries in pil_images may be None (blank pages) — these are skipped."""
    print("\n── [PASSION] OCR Pass (Tesseract) ──")
    n = len(pil_images)
    if not HAS_TESSERACT:
        print("  Tesseract not available — all pages will use vision")
        return [None] * n, [None] * n

    ocr_texts = [None] * n
    ocr_confs = [None] * n
    good = 0
    skipped = 0

    def _ocr_one(i, img):
        text, conf = ocr_page(img, i + 1)
        return i, text, conf

    # Only submit non-None images to the thread pool
    ocr_tasks = [(i, img) for i, img in enumerate(pil_images) if img is not None]
    if not ocr_tasks:
        print("  No pages to OCR (all blank)")
        return ocr_texts, ocr_confs

    with ThreadPoolExecutor(max_workers=min(8, len(ocr_tasks))) as pool:
        futures = {pool.submit(_ocr_one, i, img): i for i, img in ocr_tasks}
        for future in as_completed(futures):
            i, text, conf = future.result()
            ocr_texts[i] = text
            ocr_confs[i] = conf

    for i, text in enumerate(ocr_texts):
        if pil_images[i] is None:
            skipped += 1
            print(f"  Page {i+1}: ○ skipped (blank)")
        elif text:
            good += 1
            chars = len(text.strip())
            nums = len(re.findall(r'\d+[,.]?\d*', text))
            conf_str = f", conf {ocr_confs[i]:.0f}%" if ocr_confs[i] is not None else ""
            print(f"  Page {i+1}: ✓ {chars} chars, {nums} numbers found{conf_str}")
        else:
            print(f"  Page {i+1}: ✗ OCR failed — will use vision")
    total_non_blank = n - skipped
    print(f"  OCR success: {good}/{total_non_blank} pages" + (f" ({skipped} blank skipped)" if skipped else ""))
    return ocr_texts, ocr_confs


# ─── PER-PAGE ROUTING (T1.2) ───────────────────────────────────────────────

def route_pages(page_texts, ocr_texts, ocr_confidences, preproc_meta):
    """Determine the optimal extraction method for each page independently.

    Returns list[dict] — one routing entry per page:
      {page_num, blank, text_chars, text_words, digit_ratio,
       ocr_chars, ocr_conf_avg, method, reason}

    Methods: "text_layer", "ocr", "vision", "skip_blank"
    """
    n = len(preproc_meta)
    routing_plan = []
    counts = {"text_layer": 0, "ocr": 0, "vision": 0, "skip_blank": 0}

    print("\n── [PASSION] Per-Page Routing ──")

    for i in range(n):
        page_num = i + 1
        meta = preproc_meta[i] if i < len(preproc_meta) else {}

        # ─── Compute stats ───
        pt = page_texts[i] if page_texts and i < len(page_texts) else None
        ot = ocr_texts[i] if ocr_texts and i < len(ocr_texts) else None

        text_chars = len(pt.strip()) if pt else 0
        text_words = len(pt.split()) if pt else 0
        ocr_chars = len(ot.strip()) if ot else 0
        ocr_conf = ocr_confidences[i] if ocr_confidences and i < len(ocr_confidences) else None

        # Digit ratio: fraction of digits in best available text
        best_text = pt or ot or ""
        total_len = len(best_text.strip())
        digit_count = sum(1 for c in best_text if c.isdigit())
        digit_ratio = round(digit_count / total_len, 3) if total_len > 0 else 0.0

        entry = {
            "page_num": page_num,
            "blank": False,
            "text_chars": text_chars,
            "text_words": text_words,
            "digit_ratio": digit_ratio,
            "ocr_chars": ocr_chars,
            "ocr_conf_avg": ocr_conf,
            "method": "vision",      # default
            "reason": "insufficient_text_and_ocr",
        }

        # ─── Blank check (from preprocessing) ───
        if meta.get("is_blank"):
            entry["blank"] = True
            entry["method"] = "skip_blank"
            reason = meta.get("blank_reason", "blank_true")
            entry["reason"] = f"blank_{reason}" if not reason.startswith("blank_") else reason
            counts["skip_blank"] += 1
            print(f"  Page {page_num}: skip_blank ({entry['reason']})")
            routing_plan.append(entry)
            continue

        # ─── Text-layer checks ───
        if text_chars >= ROUTE_TEXT_MIN_CHARS:
            entry["method"] = "text_layer"
            entry["reason"] = f"text_chars>={ROUTE_TEXT_MIN_CHARS}"
            counts["text_layer"] += 1
            print(f"  Page {page_num}: text_layer ({entry['reason']}, {text_chars} chars)")
            routing_plan.append(entry)
            continue

        if text_words >= ROUTE_TEXT_MIN_WORDS and text_chars >= ROUTE_TEXT_MIN_CHARS_ALT:
            entry["method"] = "text_layer"
            entry["reason"] = f"text_words>={ROUTE_TEXT_MIN_WORDS}+chars>={ROUTE_TEXT_MIN_CHARS_ALT}"
            counts["text_layer"] += 1
            print(f"  Page {page_num}: text_layer ({entry['reason']}, {text_words} words)")
            routing_plan.append(entry)
            continue

        # ─── OCR checks ───
        if ocr_chars >= ROUTE_OCR_MIN_CHARS:
            entry["method"] = "ocr"
            conf_str = f", conf {ocr_conf:.0f}%" if ocr_conf is not None else ""
            entry["reason"] = f"ocr_chars>={ROUTE_OCR_MIN_CHARS}"
            counts["ocr"] += 1
            print(f"  Page {page_num}: ocr ({entry['reason']}, {ocr_chars} chars{conf_str})")
            routing_plan.append(entry)
            continue

        if ocr_conf is not None and ocr_conf >= ROUTE_OCR_MIN_CONF:
            entry["method"] = "ocr"
            entry["reason"] = f"ocr_conf_avg>={ROUTE_OCR_MIN_CONF}"
            counts["ocr"] += 1
            print(f"  Page {page_num}: ocr ({entry['reason']}, conf {ocr_conf:.0f}%)")
            routing_plan.append(entry)
            continue

        # ─── Fallback: vision ───
        counts["vision"] += 1
        print(f"  Page {page_num}: vision ({entry['reason']}, text={text_chars}, ocr={ocr_chars})")
        routing_plan.append(entry)

    parts = []
    for m in ("text_layer", "ocr", "vision", "skip_blank"):
        if counts[m] > 0:
            parts.append(f"{counts[m]} {m}")
    print(f"  Routing summary: {', '.join(parts)}")

    return routing_plan


# ─── PHASE 1.3: SECTION / FORM DETECTION ────────────────────────────────────

def detect_sections(page_texts, routing_plan, ocr_texts=None):
    """Classify each page by document section using keyword scoring.

    Uses the best available text per page (text_layer or OCR, per routing_plan).
    No external API calls — pure keyword matching.

    Args:
        page_texts: list[str|None] — PyMuPDF text-layer text per page
        routing_plan: list[dict] — from route_pages(), one entry per page
        ocr_texts: list[str|None] — Tesseract OCR text per page (optional)

    Returns:
        sections_by_page: dict[str, list[str]] — page number (str) → list of section labels
            Example: {"1": ["cover"], "2": ["summary", "div"], "3": ["div"]}
    """
    n = len(routing_plan) if routing_plan else (len(page_texts) if page_texts else 0)
    if n == 0:
        return {}

    sections_by_page = {}
    label_counts = {}

    print("\n── [PASSION] Section Detection ──")

    for i in range(n):
        page_num = i + 1
        route = routing_plan[i] if routing_plan and i < len(routing_plan) else {}

        # Skip blank pages
        if route.get("blank") or route.get("method") == "skip_blank":
            print(f"  Page {page_num}: (blank)")
            continue

        # Select best text source per routing decision
        method = route.get("method", "vision")
        if method == "text_layer":
            text = page_texts[i] if page_texts and i < len(page_texts) else None
        elif method == "ocr":
            text = ocr_texts[i] if ocr_texts and i < len(ocr_texts) else None
        else:
            # Vision pages: try text_layer first, then OCR
            text = None
            if page_texts and i < len(page_texts) and page_texts[i]:
                text = page_texts[i]
            elif ocr_texts and i < len(ocr_texts) and ocr_texts[i]:
                text = ocr_texts[i]

        if not text or len(text.strip()) < 10:
            sections_by_page[str(page_num)] = ["unknown"]
            label_counts["unknown"] = label_counts.get("unknown", 0) + 1
            print(f"  Page {page_num}: unknown (no text)")
            continue

        # Score against all section keyword lists
        text_lower = text.lower()
        scores = {}
        hits = {}  # track which phrases matched for debugging

        for section, phrases in SECTION_KEYWORDS.items():
            section_score = 0.0
            section_hits = []
            for phrase, weight in phrases:
                if phrase in text_lower:
                    section_score += weight
                    section_hits.append(phrase)
            if section_score > 0:
                scores[section] = section_score
                hits[section] = section_hits

        # Assign labels for sections above threshold
        labels = []
        for section, score in sorted(scores.items(), key=lambda x: -x[1]):
            if score >= SECTION_SCORE_THRESHOLD:
                labels.append(section)

        if not labels:
            labels = ["unknown"]

        sections_by_page[str(page_num)] = labels

        for lbl in labels:
            label_counts[lbl] = label_counts.get(lbl, 0) + 1

        # Print details
        if labels == ["unknown"]:
            top_scores = sorted(scores.items(), key=lambda x: -x[1])[:2]
            score_str = ", ".join(f"{s}={v:.0f}" for s, v in top_scores) if top_scores else "no hits"
            print(f"  Page {page_num}: unknown ({score_str})")
        else:
            detail_parts = []
            for lbl in labels:
                s = scores.get(lbl, 0)
                h = hits.get(lbl, [])
                top_hit = h[0] if h else "?"
                detail_parts.append(f"{lbl}={s:.0f}")
            print(f"  Page {page_num}: {', '.join(labels)} ({', '.join(detail_parts)})")

    # Summary line
    parts = []
    for lbl in ("summary", "div", "int", "oid", "b_summary", "b_transactions",
                "k1_1065", "k1_1120s", "k1_1041", "cover", "continuation", "unknown"):
        cnt = label_counts.get(lbl, 0)
        if cnt > 0:
            parts.append(f"{cnt} {lbl}")
    print(f"  Section summary: {', '.join(parts) if parts else 'none'}")

    return sections_by_page


import time as _time

# Retriable HTTP error codes
_RETRIABLE_ERRORS = (429, 500, 502, 503, 529)
_MAX_RETRIES = 3
_RETRY_BASE_DELAY = 2  # seconds, doubles each attempt

def _is_retriable(exc):
    """Check if an API error is transient and worth retrying."""
    exc_str = str(exc).lower()
    # anthropic SDK raises specific error types with status_code attribute
    status = getattr(exc, 'status_code', None)
    if status and status in _RETRIABLE_ERRORS:
        return True
    # Fallback: check error message text
    if any(s in exc_str for s in ['rate limit', 'overloaded', '529', '502', '503', '500', 'timeout', 'connection', 'timed out']):
        return True
    return False

def call_claude_vision(client, image_b64, prompt, page_num=None, tokenizer=None, max_tokens=None, phase="extract"):
    """Send image to Claude with retry logic. Redacts SSNs if tokenizer provided."""
    global _cost_tracker
    tag = f"[Page {page_num}] " if page_num else ""
    tokens = max_tokens or MAX_TOKENS

    # Redact PII from image before sending (do this once, not per-retry)
    send_b64 = image_b64
    if tokenizer and HAS_TESSERACT:
        try:
            img_bytes = base64.b64decode(image_b64)
            pil_img = Image.open(BytesIO(img_bytes))
            send_b64 = tokenizer.redacted_image_to_b64(pil_img)
        except Exception as e:
            print(f"  {tag}PII redaction warning: {e} (sending unredacted)")

    for attempt in range(_MAX_RETRIES):
        try:
            msg = client.messages.create(
                model=MODEL, max_tokens=tokens,
                messages=[{"role": "user", "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": send_b64}},
                    {"type": "text", "text": prompt}
                ]}]
            )

            # Track cost
            if _cost_tracker and hasattr(msg, 'usage'):
                _cost_tracker.record(phase, page_num, msg.usage, "vision")

            # Check for truncation (max_tokens hit)
            stop_reason = getattr(msg, 'stop_reason', None)
            if stop_reason == 'max_tokens' and tokens < 16000:
                print(f"  {tag}Response truncated (hit {tokens} tokens) — retrying with higher limit")
                tokens = min(tokens * 2, 16000)
                continue

            result = _parse_json_response(msg.content[0].text, tag)

            # If JSON parse failed, retry once with a nudge
            if result is None and attempt < _MAX_RETRIES - 1:
                print(f"  {tag}JSON parse failed — retrying (attempt {attempt + 2}/{_MAX_RETRIES})")
                _time.sleep(1)
                continue

            # Detokenize any tokens that leaked into the response
            if tokenizer and result:
                result = tokenizer.detokenize_json(result)

            return result

        except Exception as e:
            if _is_retriable(e) and attempt < _MAX_RETRIES - 1:
                delay = _RETRY_BASE_DELAY * (2 ** attempt)
                print(f"  {tag}API error: {e} — retrying in {delay}s (attempt {attempt + 2}/{_MAX_RETRIES})")
                _time.sleep(delay)
                continue
            else:
                print(f"  {tag}VISION ERROR: {e}" + (" (no more retries)" if attempt > 0 else ""))
                return None

    print(f"  {tag}VISION ERROR: all {_MAX_RETRIES} attempts failed")
    return None

def call_claude_text(client, text_content, prompt, page_num=None, tokenizer=None, phase="extract"):
    """Send text to Claude with retry logic. Tokenizes SSNs if tokenizer provided."""
    global _cost_tracker
    tag = f"[Page {page_num}] " if page_num else ""

    # Tokenize PII once before retries
    send_text = text_content
    pii_count = 0
    if tokenizer:
        send_text, pii_count = tokenizer.tokenize_text(text_content)
        if pii_count > 0:
            print(f"  {tag}PII: tokenized {pii_count} SSN(s) before API call")

    for attempt in range(_MAX_RETRIES):
        try:
            msg = client.messages.create(
                model=MODEL, max_tokens=MAX_TOKENS,
                messages=[{"role": "user", "content": f"{prompt}\n\n--- OCR TEXT ---\n{send_text}\n--- END OCR TEXT ---"}]
            )

            # Track cost
            if _cost_tracker and hasattr(msg, 'usage'):
                _cost_tracker.record(phase, page_num, msg.usage, "text")
            result = _parse_json_response(msg.content[0].text, tag)

            if result is None and attempt < _MAX_RETRIES - 1:
                print(f"  {tag}JSON parse failed — retrying (attempt {attempt + 2}/{_MAX_RETRIES})")
                _time.sleep(1)
                continue

            if tokenizer and result:
                result = tokenizer.detokenize_json(result)

            return result

        except Exception as e:
            if _is_retriable(e) and attempt < _MAX_RETRIES - 1:
                delay = _RETRY_BASE_DELAY * (2 ** attempt)
                print(f"  {tag}API error: {e} — retrying in {delay}s (attempt {attempt + 2}/{_MAX_RETRIES})")
                _time.sleep(delay)
                continue
            else:
                print(f"  {tag}TEXT ERROR: {e}" + (" (no more retries)" if attempt > 0 else ""))
                return None

    print(f"  {tag}TEXT ERROR: all {_MAX_RETRIES} attempts failed")
    return None


def call_claude_vision_multipage(client, images_b64, prompt, page_nums=None, tokenizer=None, phase="extract"):
    """Send multiple images in one API call so Claude can cross-reference pages.

    Used for K-1s with continuation statements and multi-page brokerage composites
    where values on one page reference data on another (e.g., "STMT" → continuation).
    """
    global _cost_tracker
    tag = f"[Pages {page_nums}] " if page_nums else ""

    # Build content array with all images + prompt
    content = []
    for i, img_b64 in enumerate(images_b64):
        send_b64 = img_b64
        if tokenizer and HAS_TESSERACT:
            try:
                img_bytes = base64.b64decode(img_b64)
                pil_img = Image.open(BytesIO(img_bytes))
                send_b64 = tokenizer.redacted_image_to_b64(pil_img)
            except Exception:
                pass
        pnum = page_nums[i] if page_nums and i < len(page_nums) else i + 1
        content.append({"type": "text", "text": f"--- PAGE {pnum} ---"})
        content.append({"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": send_b64}})
    content.append({"type": "text", "text": prompt})

    tokens = MAX_TOKENS * min(len(images_b64), 3)  # Scale tokens for multi-page

    for attempt in range(_MAX_RETRIES):
        try:
            msg = client.messages.create(
                model=MODEL, max_tokens=tokens,
                messages=[{"role": "user", "content": content}]
            )

            if _cost_tracker and hasattr(msg, 'usage'):
                _cost_tracker.record(phase, page_nums[0] if page_nums else None, msg.usage, "vision_multi")

            stop_reason = getattr(msg, 'stop_reason', None)
            if stop_reason == 'max_tokens' and tokens < 16000:
                tokens = min(tokens * 2, 16000)
                continue

            result = _parse_json_response(msg.content[0].text, tag)

            if result is None and attempt < _MAX_RETRIES - 1:
                _time.sleep(1)
                continue

            if tokenizer and result:
                result = tokenizer.detokenize_json(result)

            return result

        except Exception as e:
            if _is_retriable(e) and attempt < _MAX_RETRIES - 1:
                delay = _RETRY_BASE_DELAY * (2 ** attempt)
                print(f"  {tag}API error: {e} — retrying in {delay}s")
                _time.sleep(delay)
                continue
            else:
                print(f"  {tag}VISION MULTI ERROR: {e}")
                return None

    return None

def _parse_json_response(text, tag=""):
    """Parse JSON from Claude's response, handling markdown fences and truncation."""
    text = text.strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[1] if "\n" in text else text[3:]
        if text.endswith("```"): text = text[:-3]
        text = text.strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        # Try to extract JSON object from surrounding text
        start = text.find("{")
        end = text.rfind("}") + 1
        if start >= 0 and end > start:
            try: return json.loads(text[start:end])
            except (json.JSONDecodeError, ValueError): pass

        # Attempt to repair truncated JSON (missing closing braces)
        if start >= 0:
            fragment = text[start:]
            # Count unmatched braces/brackets and close them
            opens_b = fragment.count("{") - fragment.count("}")
            opens_a = fragment.count("[") - fragment.count("]")
            if opens_b > 0 or opens_a > 0:
                # Strip trailing comma or partial key
                repaired = fragment.rstrip().rstrip(",").rstrip(":")
                # Remove any incomplete key-value at the end
                last_comma = repaired.rfind(",")
                last_brace = max(repaired.rfind("{"), repaired.rfind("["))
                if last_comma > last_brace:
                    repaired = repaired[:last_comma]
                repaired += "]" * max(0, opens_a) + "}" * max(0, opens_b)
                try:
                    result = json.loads(repaired)
                    print(f"  {tag}WARNING: repaired truncated JSON ({opens_b} braces, {opens_a} brackets)")
                    return result
                except (json.JSONDecodeError, ValueError):
                    pass

        print(f"  {tag}WARNING: JSON parse failed ({len(text)} chars)")
        return None


# ─── PHASE 1: CLASSIFY (always vision — need to see page layout) ─────────

def classify_pages(client, b64_images, tokenizer=None, doc_type=None, user_notes="", ai_instructions=""):
    print(f"\n── [PASSION] Phase 1: Classification ({len(b64_images)} pages, {MAX_CONCURRENT} concurrent) ──")
    results = [None] * len(b64_images)

    # Narrow classification prompt to relevant doc types
    prompt = CLASSIFICATION_PROMPT
    if doc_type and doc_type in DOC_TYPE_CLASSIFICATIONS:
        narrowed = ", ".join(DOC_TYPE_CLASSIFICATIONS[doc_type])
        prompt = prompt.rsplit("Document types:", 1)[0] + "Document types:\n" + narrowed
        print(f"  (narrowed to {doc_type} types: {len(DOC_TYPE_CLASSIFICATIONS[doc_type])} categories)")
    if ai_instructions:
        prompt += f"\n\nSPECIAL INSTRUCTIONS FROM OPERATOR:\n{ai_instructions}"
        print(f"  AI instructions: {ai_instructions[:100]}{'...' if len(ai_instructions) > 100 else ''}")
    if user_notes:
        prompt += f"\n\nADDITIONAL CONTEXT: {user_notes}"

    def _classify_one(i, img):
        r = call_claude_vision(client, img, prompt, i+1, tokenizer=tokenizer, phase="classify")
        if r:
            r["page_number"] = i + 1
        else:
            r = {"page_number": i + 1, "document_type": "unknown"}
        return i, r

    with ThreadPoolExecutor(max_workers=MAX_CONCURRENT) as pool:
        futures = {pool.submit(_classify_one, i, img): i for i, img in enumerate(b64_images)}
        for future in as_completed(futures):
            i, r = future.result()
            results[i] = r
            dtype = r.get("document_type", "?")
            subs = r.get("sub_types", [])
            entity = r.get("payer_or_entity", "?")
            ein = r.get("payer_ein", "")
            flags = []
            if r.get("is_consolidated_brokerage"): flags.append("BROKERAGE")
            if r.get("is_supplemental_detail"): flags.append("detail")
            if r.get("is_continuation_statement"): flags.append("continuation")
            flag_str = f" [{', '.join(flags)}]" if flags else ""
            sub_str = f" → {subs}" if subs and len(subs) > 1 else ""
            print(f"  Page {i+1}: {dtype}{sub_str} — {entity} (EIN: {ein}){flag_str}")

    return results


# ─── PHASE 1.5: GROUP PAGES ──────────────────────────────────────────────────

def group_pages(classifications):
    """Group pages by EIN (primary) or entity name (fallback)."""
    groups = []
    current = None

    for cls in classifications:
        pnum = cls["page_number"]
        dtype = cls.get("document_type", "unknown")
        entity = cls.get("payer_or_entity") or ""
        ein = cls.get("payer_ein") or ""
        is_supp = cls.get("is_supplemental_detail", False)
        is_cont = cls.get("is_continuation_statement", False)
        is_brokerage = cls.get("is_consolidated_brokerage", False)

        if is_supp and not is_cont and current:
            current.setdefault("supplemental_pages", []).append(pnum)
            continue

        if is_cont and current:
            current.setdefault("continuation_pages", []).append(pnum)
            continue

        if current:
            same_ein = (ein and ein == current.get("payer_ein", ""))
            same_name = (entity.upper().strip() == current.get("payer_or_entity", "").upper().strip())
            same_dtype = (dtype == current.get("document_type"))
            same_recip = ((cls.get("recipient") or "").upper() == (current.get("recipient") or "").upper())
            if same_dtype and (same_ein or same_name) and same_recip:
                current["pages"].append(pnum)
                if is_brokerage:
                    current["is_consolidated_brokerage"] = True
                    for st in cls.get("sub_types", []):
                        if st not in current.get("sub_types", []):
                            current.setdefault("sub_types", []).append(st)
                continue

        if current:
            groups.append(current)
        current = {
            "document_type": dtype,
            "payer_or_entity": entity,
            "payer_ein": ein,
            "recipient": cls.get("recipient", ""),
            "tax_year": cls.get("tax_year", ""),
            "pages": [pnum],
            "supplemental_pages": [],
            "continuation_pages": [],
            "is_consolidated_brokerage": is_brokerage,
            "sub_types": cls.get("sub_types", []),
        }

    if current:
        groups.append(current)
    return groups


# ─── T1.5: PRIORITY SCHEDULING ──────────────────────────────────────────────

def build_priority_queue(groups, sections_by_page):
    """Reorder groups by section importance for progressive extraction.

    Uses SECTION_PRIORITY to sort groups so that summary/K-1/income pages
    extract first, transaction/cover/continuation pages extract last.
    Stable sort: same-priority groups maintain original document order.
    """
    if not sections_by_page:
        return groups

    def _group_priority(group):
        all_pages = group.get("pages", []) + group.get("continuation_pages", [])
        labels = []
        for pnum in all_pages:
            labels.extend(sections_by_page.get(str(pnum), []))
        if not labels:
            return 5  # default priority for pages with no section labels
        return min(SECTION_PRIORITY.get(lbl, 5) for lbl in labels)

    ordered = sorted(groups, key=_group_priority)

    # Print priority summary
    parts = []
    for g in ordered:
        dtype = g.get("document_type", "?")
        pages = g.get("pages", [])
        pri = _group_priority(g)
        if len(pages) == 1:
            parts.append(f"{dtype}(p.{pages[0]},pri={pri})")
        else:
            parts.append(f"{dtype}(pp.{pages[0]}-{pages[-1]},pri={pri})")
    print(f"  Priority order: {' → '.join(parts)}")

    return ordered


def build_review_queue(extractions, sections_by_page=None):
    """Sort fields by review importance for the operator.

    Returns list of {page, field, priority, confidence} dicts sorted by
    priority (most important to review first).

    Priority order:
      1 = needs_review / low confidence
      2 = consensus disagreement (_consensus_top2 present)
      3 = medium confidence
      4 = high / verified_confirmed
      5 = auto_verified (already verified — lowest review priority)
    """
    queue = []
    for ext in extractions:
        page = ext.get("_page")
        if not page:
            continue
        fields = ext.get("fields", {})
        for fname, fdata in fields.items():
            if not isinstance(fdata, dict):
                continue
            val = fdata.get("value")
            if val is None:
                continue
            conf = fdata.get("confidence", "medium")
            has_disagreement = "_consensus_top2" in fdata

            if conf in ("needs_review", "low"):
                pri = 1
            elif has_disagreement:
                pri = 2
            elif conf == "medium":
                pri = 3
            elif conf in ("high", "verified_confirmed", "verified_corrected",
                          "dual_confirmed", "multipage_verified", "consensus_accepted"):
                pri = 4
            elif conf == "auto_verified":
                pri = 5
            else:
                pri = 3  # default

            queue.append({
                "page": page,
                "field": fname,
                "priority": pri,
                "confidence": conf,
            })

    queue.sort(key=lambda q: (q["priority"], q["page"]))
    return queue


# ─── PHASE 2: EXTRACT (Vision) ───────────────────────────────────────────────

def extract_data(client, b64_images, groups, tokenizer=None, doc_type=None, user_notes="", ai_instructions="", ocr_texts=None, page_texts=None, routing_plan=None, sections_by_page=None, output_path=None):
    """
    For each page, use routing_plan to select extraction method (text_layer/ocr/vision).
    Falls back to original priority cascade if routing_plan is None.
    Priority: PyMuPDF text layer (best) → Tesseract OCR (good) → Claude vision (expensive).
    PII redaction (if tokenizer provided) blacks out sensitive data in images.
    Uses concurrent API calls for speed.
    """
    print(f"\n── [PASSION] Phase 2: Extraction ({MAX_CONCURRENT} concurrent) ──")

    # Build context hints based on doc_type
    doc_type_hints = {
        "bookkeeping": "\nThis is a BOOKKEEPING document. Focus on: transaction dates, amounts, accounts, vendor names, categories. Extract individual line items and transactions when present.",
        "bank_statements": "\nThis is a BANK STATEMENT. Focus on: account balances AND every individual transaction. Use numbered fields: txn_1_date, txn_1_desc, txn_1_amount, txn_1_type, txn_2_date, etc. Extract ALL visible transactions — deposits, withdrawals, checks, fees, transfers. Include check numbers in desc field.",
    }
    extra_context = doc_type_hints.get(doc_type, "")
    if ai_instructions:
        extra_context += f"\n\nSPECIAL INSTRUCTIONS FROM OPERATOR — follow these carefully:\n{ai_instructions}"
        print(f"  AI instructions: {ai_instructions[:100]}{'...' if len(ai_instructions) > 100 else ''}")
    if user_notes:
        extra_context += f"\n\nADDITIONAL CONTEXT: {user_notes}"

    # T1.5: Priority reorder groups so high-value pages extract first
    if sections_by_page:
        groups = build_priority_queue(groups, sections_by_page)

    # Separate groups into single-page work items vs multi-page batches
    work_items = []       # (group, page_number) for single-page extraction
    multipage_groups = [] # groups with continuations → batched extraction

    for group in groups:
        dtype = group["document_type"]
        entity = group["payer_or_entity"]
        ein = group.get("payer_ein", "")
        is_brokerage = group.get("is_consolidated_brokerage", False)
        cont_pages = group.get("continuation_pages", [])
        all_pages = group["pages"] + cont_pages
        print(f"\n  [{dtype}] {entity} (EIN: {ein}){' [BROKERAGE]' if is_brokerage else ''}")

        # Multi-page batch: K-1 with continuations or multi-page brokerage
        # Send all pages in one call so Claude can cross-reference STMT values
        if cont_pages and ("K-1" in dtype or is_brokerage):
            multipage_groups.append(group)
            print(f"    → multi-page batch ({len(all_pages)} pages: {all_pages})")
        else:
            for pnum in all_pages:
                work_items.append((group, pnum))

    # ─── T1.5: Batch extraction with partial writes ───
    extractions = []
    extraction_start = time.monotonic()
    first_values_time = None
    batch_num = 0
    # Count total batches: 1 per multipage group + ceil(single-pages / 3)
    total_batches = len(multipage_groups) + ((len(work_items) + 2) // 3 if work_items else 0)

    # ─── Process multi-page groups first (K-1 + continuations in one call) ───
    for group in multipage_groups:
        dtype = group["document_type"]
        ein = group.get("payer_ein", "")
        entity = group["payer_or_entity"]
        is_brokerage = group.get("is_consolidated_brokerage", False)
        all_pages = group["pages"] + group.get("continuation_pages", [])

        # Build multi-page prompt
        context = f"The document is classified as: {dtype}" + extra_context
        multi_prompt = VISION_EXTRACTION_PROMPT.replace("{context}", context)
        multi_prompt += (
            "\n\nMULTI-PAGE DOCUMENT: You are seeing ALL pages of this document together. "
            "If any box shows 'STMT' or '* STMT', look at the continuation statement pages "
            "to find the actual values. Resolve ALL continuation references and return the "
            "final values (not 'STMT'). Return a SINGLE JSON with all fields from all pages combined."
        )

        imgs = [b64_images[p - 1] for p in all_pages]
        r = call_claude_vision_multipage(client, imgs, multi_prompt,
                                          page_nums=all_pages, tokenizer=tokenizer,
                                          phase="extract_multi")
        if r:
            r["_group"] = group
            r["_page"] = all_pages[0]  # Primary page
            r["_pages_batched"] = all_pages
            r["_is_brokerage"] = is_brokerage
            r["_extraction_method"] = "vision_multipage"
            r["payer_ein"] = ein or r.get("payer_ein", "")
            extractions.append(r)
            print(f"    Multi-page extracted: {entity} ({len(all_pages)} pages → 1 call)")
            for fname, fdata in r.get("fields", {}).items():
                val = fdata.get("value") if isinstance(fdata, dict) else fdata
                if val is not None and val != 0 and val != "0" and val != 0.0:
                    conf = fdata.get("confidence", "?") if isinstance(fdata, dict) else "?"
                    print(f"      {fname}: {val} ({conf})")
        else:
            # Fallback: extract pages individually
            print(f"    Multi-page failed — falling back to individual pages")
            for pnum in all_pages:
                work_items.append((group, pnum))
            # Recompute total_batches since work_items grew
            total_batches = batch_num + len(multipage_groups) - (batch_num) + ((len(work_items) + 2) // 3 if work_items else 0)

        # T1.5: Emit batch event after each multi-page group
        batch_num += 1
        if extractions and first_values_time is None:
            first_values_time = round(time.monotonic() - extraction_start, 2)
            print(f"FIRST_VALUES_READY:{first_values_time}")
        if output_path:
            write_partial_results(output_path, extractions, batch_num, total_batches,
                                  time_to_first_values_s=first_values_time,
                                  sections_by_page=sections_by_page)
        # Count fields for progress event
        total_fields = sum(
            1 for e in extractions for f, fd in (e.get("fields") or {}).items()
            if (fd.get("value") if isinstance(fd, dict) else fd) is not None
        )
        print(f"BATCH_COMPLETE:{batch_num}:{total_batches}:{total_fields}")

    # ─── Process single-page work items ───

    def _extract_one(group, pnum):
        dtype = group["document_type"]
        ein = group.get("payer_ein", "")
        is_brokerage = group.get("is_consolidated_brokerage", False)
        context = f"The document is classified as: {dtype}" + extra_context

        # ─── Routing-plan driven text source selection ───
        text_source = None  # will be "text_layer" or "ocr"
        best_text = None
        route = routing_plan[pnum - 1] if routing_plan and pnum <= len(routing_plan) else None

        if route and route.get("method") == "skip_blank":
            return pnum, None  # blank page — skip extraction

        planned = route.get("method") if route else None

        if planned == "text_layer":
            # Routing says text layer is best for this page
            pt = page_texts[pnum - 1] if page_texts and pnum <= len(page_texts) else None
            if pt and len(pt.strip()) > 0:
                best_text = pt
                text_source = "text_layer"
        elif planned == "ocr":
            # Routing says OCR is best for this page
            ot = ocr_texts[pnum - 1] if ocr_texts and pnum <= len(ocr_texts) else None
            if ot and len(ot.strip()) > 0:
                best_text = ot
                text_source = "ocr"
        elif planned == "vision":
            # Routing says go straight to vision
            pass
        else:
            # No routing plan — original cascade logic (backward compat)
            if page_texts and pnum <= len(page_texts):
                pt = page_texts[pnum - 1]
                if pt and len(pt.strip()) >= TEXT_MIN_CHARS_PER_PAGE:
                    best_text = pt
                    text_source = "text_layer"
            if not best_text:
                ocr_text = ocr_texts[pnum - 1] if ocr_texts and pnum <= len(ocr_texts) else None
                if ocr_text and len(ocr_text.strip()) >= 200:
                    best_text = ocr_text
                    text_source = "ocr"

        # ─── Text-first path: try cheap text call when text is good ───
        method = "vision"  # default

        if best_text:
            text_prompt = OCR_EXTRACTION_PROMPT.replace("{doc_type}", dtype)
            if extra_context:
                text_prompt += extra_context
            phase_tag = "extract_text_layer" if text_source == "text_layer" else "extract_text"
            r = call_claude_text(client, best_text, text_prompt, pnum, tokenizer=tokenizer, phase=phase_tag)
            if r:
                ocr_quality = r.get("ocr_quality", "partial")
                needs_image = r.get("needs_image_review", False)
                fields_needing_image = r.get("fields_needing_image", [])

                if ocr_quality == "good" and not needs_image:
                    method = "text_layer" if text_source == "text_layer" else "ocr_text"
                    src_label = "text layer" if text_source == "text_layer" else "OCR"
                    print(f"    Page {pnum}: {src_label} sufficient → text extraction (saved vision call)")
                elif ocr_quality == "partial" and not needs_image and fields_needing_image:
                    method = "text_layer_partial" if text_source == "text_layer" else "ocr_partial"
                    r["_ambiguous_fields"] = fields_needing_image
                    src_label = "text layer" if text_source == "text_layer" else "OCR"
                    print(f"    Page {pnum}: {src_label} partial — {len(fields_needing_image)} fields need image check")
                else:
                    r = None
                    src_label = "text layer" if text_source == "text_layer" else "OCR"
                    print(f"    Page {pnum}: {src_label} quality '{ocr_quality}' — falling back to vision")

                if r:
                    r["_group"] = group
                    r["_page"] = pnum
                    r["_is_brokerage"] = is_brokerage
                    r["_extraction_method"] = method
                    r["_text_source"] = text_source
                    r["payer_ein"] = ein or r.get("payer_ein", "")
                    return pnum, r

        # ─── Vision path (default or fallback) ───
        prompt_v = VISION_EXTRACTION_PROMPT.replace("{context}", context)
        r = call_claude_vision(client, b64_images[pnum-1], prompt_v, pnum, tokenizer=tokenizer, phase="extract_vision")
        if r:
            r["_group"] = group
            r["_page"] = pnum
            r["_is_brokerage"] = is_brokerage
            r["_extraction_method"] = "vision"
            r["_text_source"] = "none"
            r["payer_ein"] = ein or r.get("payer_ein", "")
            print(f"    Page {pnum}: vision extracted")
        return pnum, r

    total = len(work_items)
    text_layer_saved = 0
    ocr_saved = 0

    # T1.5: Split single-page work into batches of 3 for progressive results
    BATCH_SIZE = 3
    for batch_start in range(0, len(work_items), BATCH_SIZE):
        batch = work_items[batch_start:batch_start + BATCH_SIZE]
        batch_num += 1

        with ThreadPoolExecutor(max_workers=MAX_CONCURRENT) as pool:
            futures = {pool.submit(_extract_one, grp, pn): pn for grp, pn in batch}
            for future in as_completed(futures):
                pnum, r = future.result()
                if r:
                    r["_batch"] = batch_num
                    extractions.append(r)
                    m = r.get("_extraction_method", "vision")
                    if m in ("text_layer", "text_layer_partial"):
                        text_layer_saved += 1
                    elif m in ("ocr_text", "ocr_partial"):
                        ocr_saved += 1
                    for fname, fdata in r.get("fields", {}).items():
                        val = fdata.get("value") if isinstance(fdata, dict) else fdata
                        if val is not None and val != 0 and val != "0" and val != 0.0:
                            conf = fdata.get("confidence", "?") if isinstance(fdata, dict) else "?"
                            print(f"      {fname}: {val} ({conf})")
                else:
                    print(f"    Page {pnum}: extraction failed")

        # T1.5: Emit batch event after each single-page batch
        if extractions and first_values_time is None:
            first_values_time = round(time.monotonic() - extraction_start, 2)
            print(f"FIRST_VALUES_READY:{first_values_time}")
        if output_path:
            write_partial_results(output_path, extractions, batch_num, total_batches,
                                  time_to_first_values_s=first_values_time,
                                  sections_by_page=sections_by_page)
        total_fields = sum(
            1 for e in extractions for f, fd in (e.get("fields") or {}).items()
            if (fd.get("value") if isinstance(fd, dict) else fd) is not None
        )
        print(f"BATCH_COMPLETE:{batch_num}:{total_batches}:{total_fields}")

    print(f"FINALIZE_COMPLETE")

    # Sort by page number to maintain document order
    extractions.sort(key=lambda e: e.get("_page", 0))

    # T1.5: Store streaming metadata for save_log
    _streaming_meta = {
        "time_to_first_values_s": first_values_time,
        "batches_processed": batch_num,
        "fields_streamed": sum(
            1 for e in extractions for f, fd in (e.get("fields") or {}).items()
            if (fd.get("value") if isinstance(fd, dict) else fd) is not None
        ),
    }

    print(f"\n  Extraction stats: {total} pages processed, {len(extractions)} successful")
    if text_layer_saved:
        print(f"  Text-layer: {text_layer_saved}/{total} pages used text layer (OCR + vision skipped)")
    if ocr_saved:
        print(f"  OCR-first: {ocr_saved}/{total} pages used OCR text (vision calls saved)")
    vision_count = total - text_layer_saved - ocr_saved
    if vision_count > 0 and (text_layer_saved > 0 or ocr_saved > 0):
        print(f"  Vision: {vision_count}/{total} pages required vision fallback")
    return extractions, _streaming_meta


# ─── PHASE 3: VERIFY (cross-check OCR vs image for critical fields) ──────────

def verify_extractions(client, b64_images, extractions, tokenizer=None):
    """
    Verification strategy:
      - Vision-extracted pages: send image + extracted values for re-read
      - Focus on CRITICAL_FIELDS to minimize API calls
      - Uses concurrent API calls for speed
    """
    print(f"\n── [PASSION] Phase 3: Verification ({MAX_CONCURRENT} concurrent) ──")
    corrections = 0
    confirmations = 0
    total_fields = 0

    # Separate extractions into those needing API verification and those that can skip
    skip_indices = set()
    verify_work = []  # (index_in_extractions, prompt, page)

    for idx, ext in enumerate(extractions):
        page = ext.get("_page")
        fields = ext.get("fields", {})
        method = ext.get("_extraction_method", "unknown")
        if not fields or page is None:
            skip_indices.add(idx)
            continue

        dtype = ext.get("document_type", "")
        entity = ext.get("payer_or_entity", "")

        # Decide which fields need verification
        fields_to_verify = {}
        ambiguous_fields = ext.get("_ambiguous_fields", [])
        has_critical = False

        for fname, fdata in fields.items():
            val = fdata.get("value") if isinstance(fdata, dict) else fdata
            if val is None: continue
            # Skip fields already auto_verified by consensus layer (Phase 2.5)
            if isinstance(fdata, dict) and fdata.get("confidence") == "auto_verified":
                continue
            is_critical = fname in CRITICAL_FIELDS or fname in ambiguous_fields
            is_unclear = isinstance(fdata, dict) and not fdata.get("ocr_clear", True)
            is_low_conf = isinstance(fdata, dict) and fdata.get("confidence") == "low"
            if is_critical or is_unclear or is_low_conf:
                label = fdata.get("label_on_form", "") if isinstance(fdata, dict) else ""
                fields_to_verify[fname] = {"value": val, "label": label}
                has_critical = True

        if not has_critical and method == "ocr_only":
            print(f"  Page {page}: {dtype} — no critical fields, skipping verification")
            for fname in fields:
                if isinstance(fields[fname], dict):
                    fields[fname]["confidence"] = "ocr_accepted"
            skip_indices.add(idx)
            continue

        # OCR-text extractions with all-high/good confidence can skip verification
        # (the OCR was clear enough that Claude rated it "good" quality)
        if method == "ocr_text" and not has_critical:
            print(f"  Page {page}: {dtype} — OCR-text, all high confidence, skipping verification")
            for fname in fields:
                if isinstance(fields[fname], dict):
                    fields[fname]["confidence"] = "ocr_accepted"
            skip_indices.add(idx)
            continue

        # Multi-page extractions already cross-referenced — mark as higher confidence
        if method == "vision_multipage" and not has_critical:
            print(f"  Page {page}: {dtype} — multi-page extraction, no critical issues, skipping verification")
            for fname in fields:
                if isinstance(fields[fname], dict) and fields[fname].get("confidence") in ("high", "medium"):
                    fields[fname]["confidence"] = "multipage_verified"
            skip_indices.add(idx)
            continue

        # T1.4: Catch-all — if no fields need verification (all critical auto_verified
        # by consensus or no critical fields at all), skip the vision call entirely.
        if not has_critical and not fields_to_verify:
            print(f"  Page {page}: {dtype} — all fields verified by consensus, skipping verification")
            for fname in fields:
                if isinstance(fields[fname], dict) and fields[fname].get("confidence") not in ("auto_verified",):
                    fields[fname]["confidence"] = "consensus_accepted"
            skip_indices.add(idx)
            continue

        if fields_to_verify:
            print(f"  Page {page}: {dtype} — {entity} — verifying {len(fields_to_verify)} fields [{method}]")
            prompt = VERIFY_CROSSCHECK_PROMPT.replace("{extracted_json}", json.dumps(fields_to_verify, indent=2))
        else:
            all_fields = {}
            for fname, fdata in fields.items():
                val = fdata.get("value") if isinstance(fdata, dict) else fdata
                if val is not None:
                    label = fdata.get("label_on_form", "") if isinstance(fdata, dict) else ""
                    all_fields[fname] = {"value": val, "label": label}
            print(f"  Page {page}: {dtype} — {entity} — verifying all {len(all_fields)} fields [{method}]")
            prompt = VERIFY_CROSSCHECK_PROMPT.replace("{extracted_json}", json.dumps(all_fields, indent=2))

        verify_work.append((idx, prompt, page))

    # Run all verification API calls concurrently
    def _verify_one(idx, prompt, page):
        vresult = call_claude_vision(client, b64_images[page - 1], prompt, page, tokenizer=tokenizer, phase="verify")
        return idx, vresult

    results_map = {}
    if verify_work:
        with ThreadPoolExecutor(max_workers=MAX_CONCURRENT) as pool:
            futures = {pool.submit(_verify_one, idx, p, pg): idx for idx, p, pg in verify_work}
            for future in as_completed(futures):
                idx, vresult = future.result()
                results_map[idx] = vresult

    # Apply verification results back to extractions (sequential — mutates fields)
    for idx, ext in enumerate(extractions):
        if idx in skip_indices:
            continue
        fields = ext.get("fields", {})
        method = ext.get("_extraction_method", "unknown")
        page = ext.get("_page")
        vresult = results_map.get(idx)

        if vresult and "verification_results" in vresult:
            vr = vresult["verification_results"]
            for fname, vdata in vr.items():
                total_fields += 1
                status = vdata.get("status", "")
                if status == "corrected":
                    final = vdata.get("final_value") or vdata.get("image_value") or vdata.get("verified_value")
                    extracted = vdata.get("extracted_value")
                    if final is not None:
                        corrections += 1
                        if fname in fields and isinstance(fields[fname], dict):
                            fields[fname]["value"] = final
                            fields[fname]["confidence"] = "verified_corrected"
                            fields[fname]["original_value"] = extracted
                            fields[fname]["correction_note"] = vdata.get("notes", "")
                        print(f"    CORRECTED: {fname}: {extracted} → {final} ({vdata.get('notes','')})")
                elif status == "confirmed":
                    confirmations += 1
                    if fname in fields and isinstance(fields[fname], dict):
                        if method.startswith("ocr"):
                            fields[fname]["confidence"] = "dual_confirmed"
                        else:
                            fields[fname]["confidence"] = "verified_confirmed"

            for mf in vresult.get("missing_fields", []):
                fname = mf.get("field_name", "")
                val = mf.get("value")
                if fname and val is not None:
                    fields[fname] = {
                        "value": val,
                        "label_on_form": mf.get("label_on_form", ""),
                        "confidence": "found_in_verification",
                    }
                    print(f"    NEW: {fname}: {val}")

            ext["_verification"] = vresult
            ext["_overall_confidence"] = vresult.get("overall_confidence", "medium")
        else:
            ext["_overall_confidence"] = "unverified"
            print(f"    Page {page}: Verification failed — keeping extraction as-is")

    pages_verified = len(verify_work)
    pages_skipped = len(skip_indices)
    print(f"\n  Verification: {confirmations} confirmed, {corrections} corrected, {total_fields} fields checked")
    print(f"  Verification: {pages_verified} pages verified, {pages_skipped} pages skipped")
    return extractions, {"pages_verified": pages_verified, "pages_skipped": pages_skipped}


# ─── PHASE 2.5: CONSENSUS VERIFICATION (T1.2.5) ────────────────────────────

# Label patterns for finding amounts in raw text near known field labels.
# Maps field_name → list of (regex_pattern, is_case_sensitive) pairs.
_LABEL_PATTERNS = {
    # 1099-DIV fields
    "ordinary_dividends":       [r"(?:1a|ordinary\s+dividends)", r"total\s+ordinary\s+dividends"],
    "div_ordinary_dividends":   [r"(?:1a|ordinary\s+dividends)", r"total\s+ordinary\s+dividends"],
    "qualified_dividends":      [r"(?:1b|qualified\s+dividends)"],
    "div_qualified_dividends":  [r"(?:1b|qualified\s+dividends)"],
    "federal_wh":               [r"(?:4|federal\s+(?:income\s+)?tax\s+w(?:ith)?h(?:eld)?)"],
    "div_federal_wh":           [r"(?:4|federal\s+(?:income\s+)?tax\s+w(?:ith)?h(?:eld)?)"],
    "foreign_tax_paid":         [r"(?:7|foreign\s+tax\s+paid)"],
    "div_foreign_tax_paid":     [r"(?:7|foreign\s+tax\s+paid)"],
    # 1099-INT fields
    "interest_income":          [r"(?:1|interest\s+income)", r"box\s*1"],
    "int_interest_income":      [r"(?:1|interest\s+income)", r"box\s*1"],
    "us_savings_bonds_and_treasury": [r"(?:3|u\.?s\.?\s+savings)", r"treasury"],
    "int_us_savings_bonds_and_treasury": [r"(?:3|u\.?s\.?\s+savings)", r"treasury"],
    "int_federal_wh":           [r"(?:4|federal\s+(?:income\s+)?tax\s+w(?:ith)?h(?:eld)?)"],
    "int_foreign_tax_paid":     [r"(?:6|foreign\s+tax\s+paid)"],
    # 1099-B / Schedule D
    "b_total_gain_loss":        [r"(?:total\s+gain|net\s+gain)"],
    "b_short_term_gain_loss":   [r"short.?term"],
    "b_long_term_gain_loss":    [r"long.?term"],
    "total_gain_loss":          [r"(?:total\s+gain|net\s+gain)"],
    "short_term_gain_loss":     [r"short.?term"],
    "long_term_gain_loss":      [r"long.?term"],
    # K-1 boxes 1–9
    "box1_ordinary_income":     [r"(?:box\s*1\b|ordinary\s+(?:business\s+)?income)", r"line\s*1\b"],
    "box2_rental_real_estate":  [r"(?:box\s*2\b|net\s+rental\s+real\s+estate)", r"line\s*2\b"],
    "box3_other_rental":        [r"(?:box\s*3\b|other\s+net\s+rental)", r"line\s*3\b"],
    "box4a_guaranteed_services": [r"(?:box\s*4a|guaranteed\s+payments?\s+for\s+services)"],
    "box5_interest":            [r"(?:box\s*5\b|interest\s+income)", r"line\s*5\b"],
    "box6a_ordinary_dividends": [r"(?:box\s*6a|ordinary\s+dividends)"],
    "box6b_qualified_dividends": [r"(?:box\s*6b|qualified\s+dividends)"],
    "box7_royalties":           [r"(?:box\s*7\b|royalties)", r"line\s*7\b"],
    "box8_short_term_capital_gain": [r"(?:box\s*8\b|net\s+short.?term\s+capital)", r"line\s*8\b"],
    "box9a_long_term_capital_gain": [r"(?:box\s*9a|net\s+long.?term\s+capital)"],
    # K-1 coded box amounts
    "box10_net_1231_gain":      [r"(?:box\s*10|net\s+section\s+1231)", r"line\s*10\b"],
    "box11_other_income":       [r"(?:box\s*11|other\s+income)", r"line\s*11\b"],
    "box12_section_179":        [r"(?:box\s*12|section\s+179)", r"line\s*12\b"],
    "box13_other_deductions":   [r"(?:box\s*13|other\s+deductions)", r"line\s*13\b"],
}

# Amount regex: matches dollar amounts like $1,234.56 or 1234.56 or (1,234.56) for negatives
_AMOUNT_RE = re.compile(
    r'[\$]?\s*'                     # optional $
    r'[\(\-]?'                      # optional ( or - for negative
    r'(\d{1,3}(?:,\d{3})*'         # digits with optional comma grouping
    r'(?:\.\d{1,2})?)'             # optional decimal
    r'\)?'                          # optional closing )
)


def _parse_amount_from_text(text, field_name, label_hint=None, method_tag="text_anchor"):
    """Search raw text for a dollar amount near a known label.

    Returns candidate dict or None. Pure regex — no API call.
    {value_num, method, label_anchor_found, parse_ok}
    """
    if not text or not field_name:
        return None

    # Get label patterns for this field
    patterns = _LABEL_PATTERNS.get(field_name)
    if not patterns and label_hint:
        # Build a simple pattern from the label hint
        escaped = re.escape(label_hint.strip())
        if len(escaped) >= 3:
            patterns = [escaped]
    if not patterns:
        return None

    text_lower = text.lower()
    best_match = None
    label_found = False

    for pat in patterns:
        for m in re.finditer(pat, text_lower):
            label_found = True
            # Search for the nearest amount AFTER the label (within 200 chars)
            search_start = m.end()
            search_region = text[search_start:search_start + 200]
            amount_match = _AMOUNT_RE.search(search_region)
            if amount_match:
                raw = amount_match.group(1).replace(",", "")
                try:
                    value = float(raw)
                    # Check for negative indicator
                    prefix = search_region[:amount_match.start()].strip()
                    full_match = search_region[max(0, amount_match.start()-2):amount_match.end()+1]
                    if "(" in full_match or "-" in prefix[-2:]:
                        value = -value
                    # Prefer the match closest to the label
                    dist = amount_match.start()
                    if best_match is None or dist < best_match["_dist"]:
                        best_match = {
                            "value_num": value,
                            "method": method_tag,
                            "label_anchor_found": True,
                            "parse_ok": True,
                            "_dist": dist,
                        }
                except (ValueError, TypeError):
                    continue

    if best_match:
        del best_match["_dist"]
        return best_match

    # Label found but no amount nearby
    if label_found:
        return {"value_num": None, "method": method_tag, "label_anchor_found": True, "parse_ok": False}

    return None


def _values_match(a, b, tolerance=0.01):
    """Check if two numeric values match within tolerance."""
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tolerance
    except (ValueError, TypeError):
        return False


def _validate_candidate(field_name, value, extraction_fields):
    """Check brokerage/K-1 validation rules for a candidate value.
    Returns (passes: bool, violation: str|None)."""
    if value is None:
        return False, "no_value"

    try:
        v = float(value)
    except (ValueError, TypeError):
        return False, "not_numeric"

    # Brokerage: qualified ≤ ordinary
    if field_name in ("qualified_dividends", "div_qualified_dividends"):
        ord_val = get_val(extraction_fields, "ordinary_dividends") or get_val(extraction_fields, "div_ordinary_dividends")
        if ord_val is not None and v > ord_val + 0.01:
            return False, "qualified > ordinary"

    # Withholding ≥ 0
    if "wh" in field_name or "withholding" in field_name:
        if v < 0:
            return False, "negative withholding"

    # Foreign tax ≤ income
    if "foreign_tax" in field_name:
        income = get_val(extraction_fields, "interest_income") or \
                 get_val(extraction_fields, "int_interest_income") or \
                 get_val(extraction_fields, "ordinary_dividends") or \
                 get_val(extraction_fields, "div_ordinary_dividends")
        if income is not None and v > income + 0.01:
            return False, "foreign_tax > income"

    return True, None


def _score_candidate(candidate, others, extraction_fields, field_name):
    """Score a field candidate for consensus selection. Returns float score."""
    score = 0.0
    val = candidate.get("value_num")

    if val is None:
        return -10.0  # unusable candidate

    # Strong positive: label anchor found in text
    if candidate.get("label_anchor_found"):
        score += 3.0

    # Strong positive: value parsed successfully
    if candidate.get("parse_ok"):
        score += 2.0

    # Strong positive: bbox coordinates exist
    if candidate.get("bbox_px"):
        score += 1.0

    # Strong positive: agrees with at least one other candidate
    for other in others:
        if other is candidate:
            continue
        other_val = other.get("value_num")
        if _values_match(val, other_val):
            score += 4.0
            break

    # Validation rules
    passes, violation = _validate_candidate(field_name, val, extraction_fields)
    if passes:
        score += 2.0
        candidate["validation_pass"] = True
    else:
        score -= 5.0
        candidate["rule_violation"] = violation
        candidate["validation_pass"] = False

    # OCR confidence bonus
    conf = candidate.get("ocr_conf_avg")
    if conf is not None:
        if conf >= 80:
            score += 1.5
        elif conf >= 60:
            score += 0.5

    # Negative: conflicts with ALL other valid candidates
    valid_others = [o for o in others if o is not candidate and o.get("value_num") is not None]
    if valid_others and all(not _values_match(val, o["value_num"]) for o in valid_others):
        score -= 2.0

    # Improbable value check (extremely large amounts)
    try:
        if abs(float(val)) > 100_000_000:  # $100M seems improbable for a single field
            score -= 3.0
            candidate["improbable"] = True
    except (ValueError, TypeError):
        pass

    return round(score, 2)


def build_consensus(extractions, page_texts, ocr_texts, ocr_confidences=None):
    """Multi-pass verification for CONSENSUS_FIELDS (brokerage + K-1).

    For each in-scope extraction, generates up to 3 candidates per critical field
    (Claude extraction + text-layer anchor + OCR anchor), scores them, and picks
    the best with confidence scoring.

    Zero additional API calls — uses existing text/OCR data + regex parsing.
    Returns (modified extractions, consensus_data dict for logging).
    """
    print("\n── [PASSION] Phase 2.5: Consensus Verification ──")

    consensus_log = {
        "fields_checked": 0,
        "auto_verified": 0,
        "needs_review": 0,
        "per_extraction": [],
    }

    in_scope = 0
    for ext in extractions:
        dtype = str(ext.get("document_type", ""))
        # Check if this doc type is in scope for consensus
        is_in_scope = any(t in dtype for t in CONSENSUS_DOC_TYPES)
        if not is_in_scope:
            continue

        in_scope += 1
        page = ext.get("_page")
        entity = ext.get("payer_or_entity", "")
        fields = ext.get("fields", {})
        method = ext.get("_extraction_method", "unknown")

        ext_log = {
            "page": page,
            "doc_type": dtype,
            "entity": entity,
            "fields": {},
        }

        # Get raw text sources for this page
        pt = page_texts[page - 1] if page_texts and page and page <= len(page_texts) else None
        ot = ocr_texts[page - 1] if ocr_texts and page and page <= len(ocr_texts) else None
        page_conf = ocr_confidences[page - 1] if ocr_confidences and page and page <= len(ocr_confidences) else None

        fields_checked = 0
        fields_verified = 0
        fields_review = 0

        for fname in CONSENSUS_FIELDS:
            if fname not in fields:
                continue

            fdata = fields[fname]

            # T1.5: Never downgrade — skip fields already verified in a previous batch
            if isinstance(fdata, dict) and fdata.get("confidence") == "auto_verified":
                fields_checked += 1
                fields_verified += 1
                continue

            claude_val = fdata.get("value") if isinstance(fdata, dict) else fdata
            label_hint = fdata.get("label_on_form", "") if isinstance(fdata, dict) else ""

            if claude_val is None:
                continue

            fields_checked += 1
            candidates = []

            # Candidate 1: Claude extraction (already paid — Phase 2 result)
            try:
                claude_num = float(claude_val)
                c1 = {
                    "value_num": claude_num,
                    "method": "claude_extraction",
                    "label_anchor_found": True,  # Claude is instructed to find labels
                    "parse_ok": True,
                    "page": page,
                }
                if page_conf is not None:
                    c1["ocr_conf_avg"] = page_conf
                candidates.append(c1)
            except (ValueError, TypeError):
                # Non-numeric field (string) — skip consensus for this field
                continue

            # Candidate 2: Text-layer anchor (free regex parse)
            if pt:
                c2 = _parse_amount_from_text(pt, fname, label_hint, method_tag="text_anchor")
                if c2 and c2.get("value_num") is not None:
                    c2["page"] = page
                    candidates.append(c2)

            # Candidate 3: OCR anchor (free regex parse)
            if ot and ot != pt:  # skip if OCR text is same as text layer
                c3 = _parse_amount_from_text(ot, fname, label_hint, method_tag="ocr_anchor")
                if c3 and c3.get("value_num") is not None:
                    c3["page"] = page
                    if page_conf is not None:
                        c3["ocr_conf_avg"] = page_conf
                    candidates.append(c3)

            # Score all candidates
            for c in candidates:
                c["score"] = _score_candidate(c, candidates, fields, fname)

            # Sort by score descending
            candidates.sort(key=lambda c: c["score"], reverse=True)
            top = candidates[0]
            runner_up = candidates[1] if len(candidates) > 1 else None

            # Decision policy
            # Margin check only applies when top and runner-up DISAGREE on value.
            # If they agree (same value), auto-accept regardless of score gap.
            runner_agrees = runner_up is not None and _values_match(top["value_num"], runner_up["value_num"])
            if top["score"] >= CONSENSUS_ACCEPT_THRESHOLD:
                if runner_up is None or runner_agrees or (top["score"] - runner_up["score"]) >= CONSENSUS_MARGIN:
                    status = "auto_verified"
                    fields_verified += 1
                    if isinstance(fdata, dict):
                        fdata["confidence"] = "auto_verified"
                        fdata["value"] = top["value_num"]
                else:
                    status = "needs_review"
                    fields_review += 1
                    if isinstance(fdata, dict):
                        fdata["_consensus_top2"] = [
                            {"value": top["value_num"], "method": top["method"], "score": top["score"]},
                            {"value": runner_up["value_num"], "method": runner_up["method"], "score": runner_up["score"]},
                        ]
            else:
                status = "needs_review"
                fields_review += 1
                if isinstance(fdata, dict):
                    fdata["_consensus_top2"] = [
                        {"value": c["value_num"], "method": c["method"], "score": c["score"]}
                        for c in candidates[:2]
                    ]

            # Store consensus metadata on the field
            if isinstance(fdata, dict):
                fdata["_consensus"] = {
                    "status": status,
                    "chosen_method": top["method"],
                    "score": top["score"],
                    "num_candidates": len(candidates),
                }

            # Log entry for this field
            ext_log["fields"][fname] = {
                "status": status,
                "chosen_value": top["value_num"],
                "chosen_method": top["method"],
                "score": top["score"],
                "candidates": [
                    {"value": c["value_num"], "method": c["method"], "score": c["score"],
                     "label_found": c.get("label_anchor_found", False),
                     "validation_pass": c.get("validation_pass", True)}
                    for c in candidates
                ],
            }

        consensus_log["fields_checked"] += fields_checked
        consensus_log["auto_verified"] += fields_verified
        consensus_log["needs_review"] += fields_review
        if fields_checked > 0:
            consensus_log["per_extraction"].append(ext_log)
            print(f"  {dtype} — {entity} (p.{page}): {fields_checked} fields, "
                  f"{fields_verified} auto_verified, {fields_review} needs_review")

    if in_scope == 0:
        print("  No brokerage/K-1 documents — consensus skipped")
    else:
        print(f"\n  Consensus: {consensus_log['auto_verified']} auto_verified, "
              f"{consensus_log['needs_review']} needs_review "
              f"({consensus_log['fields_checked']} fields across {in_scope} documents)")

    return extractions, consensus_log


# ─── PHASE 4: NORMALIZE ──────────────────────────────────────────────────────

def normalize_brokerage_data(extractions):
    """Split brokerage composites, cross-ref K-1 continuations, roll up K-1 interest."""
    print("\n── [PASSION] Phase 4: Normalize ──")
    normalized = []
    rollups = []
    continuation_data = {}

    # Collect continuation statement data by EIN
    for ext in extractions:
        if "continuation" in str(ext.get("document_type", "")).lower():
            ein = ext.get("payer_ein", "")
            entity = ext.get("payer_or_entity", "")
            key = ein if ein else entity.upper().strip()
            if key not in continuation_data:
                continuation_data[key] = []
            for ci in ext.get("continuation_items", []):
                continuation_data[key].append(ci)
            for fname, fdata in ext.get("fields", {}).items():
                val = fdata.get("value") if isinstance(fdata, dict) else fdata
                if val is not None:
                    continuation_data[key].append({
                        "line_reference": fname,
                        "description": fdata.get("label_on_form", "") if isinstance(fdata, dict) else "",
                        "amount": val,
                    })

    for ext in extractions:
        is_brokerage = ext.get("_is_brokerage", False)
        fields = ext.get("fields", {})
        entity = ext.get("payer_or_entity", "")
        ein = ext.get("payer_ein", "")

        if is_brokerage:
            div_f, int_f, b_f, misc_f = {}, {}, {}, {}

            for fname, fdata in fields.items():
                if fname.startswith("div_"):
                    div_f[fname[4:]] = fdata
                elif fname.startswith("int_"):
                    int_f[fname[4:]] = fdata
                elif fname.startswith("b_"):
                    b_f[fname[2:]] = fdata
                elif fname.startswith("misc_"):
                    misc_f[fname[5:]] = fdata
                else:
                    fl = fname.lower()
                    if any(k in fl for k in ["dividend", "qualified"]): div_f[fname] = fdata
                    elif any(k in fl for k in ["interest_income", "us_savings", "treasury", "bond_premium"]): int_f[fname] = fdata
                    elif any(k in fl for k in ["proceeds", "cost_basis", "gain_loss", "wash_sale"]): b_f[fname] = fdata
                    elif any(k in fl for k in ["royalt", "other_income", "rent"]): misc_f[fname] = fdata

            base = {"payer_ein": ein, "recipient": ext.get("recipient", ""),
                    "_page": ext.get("_page"), "_source_name": entity,
                    "_overall_confidence": ext.get("_overall_confidence", "medium")}

            if div_f:
                normalized.append({**base, "document_type": "1099-DIV", "payer_or_entity": entity, "fields": div_f})
                print(f"  Brokerage → 1099-DIV: {entity}")
            if int_f:
                normalized.append({**base, "document_type": "1099-INT", "payer_or_entity": entity, "fields": int_f})
                print(f"  Brokerage → 1099-INT: {entity}")
            if b_f:
                sched_d = {}
                for k, v in b_f.items():
                    kl = k.lower()
                    if "short_term" in kl and "gain" in kl: sched_d["short_term_gain_loss"] = v
                    elif "long_term" in kl and "gain" in kl: sched_d["long_term_gain_loss"] = v
                    elif "total_gain" in kl or kl == "gain_loss": sched_d["total_gain_loss"] = v
                    elif "total_proceeds" in kl or kl == "proceeds": sched_d["total_proceeds"] = v
                    elif "total_basis" in kl or kl == "basis" or "cost_basis" in kl: sched_d["total_basis"] = v
                    elif "wash" in kl: sched_d["wash_sale_loss"] = v
                    else: sched_d[k] = v
                normalized.append({**base, "document_type": "1099-B", "payer_or_entity": entity, "fields": sched_d})
                print(f"  Brokerage → 1099-B: {entity}")
            if misc_f:
                normalized.append({**base, "document_type": "1099-MISC", "payer_or_entity": entity, "fields": misc_f})

            ext["_brokerage_split"] = True
        elif "continuation" in str(ext.get("document_type", "")).lower():
            continue  # Already processed
        else:
            # Strip prefixes from standalone 1099-DIV, 1099-INT, 1099-B, 1099-MISC fields
            # The extraction prompt always uses prefixed names (div_, int_, b_, misc_)
            # but the Excel template expects unprefixed names
            dtype_str = str(ext.get("document_type", "")).upper()
            if any(t in dtype_str for t in ["1099-DIV", "1099-INT", "1099-B", "1099-MISC"]):
                cleaned_fields = {}
                for fname, fdata in fields.items():
                    if fname.startswith("div_"):
                        cleaned_fields[fname[4:]] = fdata
                    elif fname.startswith("int_"):
                        cleaned_fields[fname[4:]] = fdata
                    elif fname.startswith("b_"):
                        cleaned_fields[fname[2:]] = fdata
                    elif fname.startswith("misc_"):
                        cleaned_fields[fname[5:]] = fdata
                    else:
                        cleaned_fields[fname] = fdata
                ext["fields"] = cleaned_fields
            normalized.append(ext)

        # K-1: cross-ref continuations for missing Box 2
        if "K-1" in str(ext.get("document_type", "")):
            k1_name = get_str(fields, "partnership_name") or entity
            k1_ein = ext.get("payer_ein", "")
            box2 = get_val(fields, "box2_rental_real_estate")

            if box2 is None or box2 == 0:
                cont_key = k1_ein if k1_ein else k1_name.upper().strip()
                for ci in continuation_data.get(cont_key, []):
                    ref = str(ci.get("line_reference", "")).lower()
                    desc = str(ci.get("description", "")).lower()
                    amt = ci.get("amount")
                    if amt is not None and amt != 0 and (
                        "8825" in ref or "rental" in ref or "rental" in desc or "box 2" in ref
                    ):
                        print(f"  K-1 cross-ref: {k1_name} Box 2 = {amt} (from continuation)")
                        fields["box2_rental_real_estate"] = {
                            "value": amt, "confidence": "from_continuation",
                            "label_on_form": f"From: {ci.get('line_reference','')}"
                        }
                        break

            # Roll up interest
            box5 = get_val(fields, "box5_interest")
            if box5 and box5 != 0:
                rollups.append({
                    "document_type": "_interest_rollup",
                    "payer_or_entity": k1_name,
                    "fields": {
                        "interest_income": {"value": box5, "confidence": "from_k1_box5"},
                        "us_savings_bonds_and_treasury": {"value": 0, "confidence": "n/a"},
                    },
                    "_source_name": f"{k1_name} (K-1 Box 5)",
                })
                print(f"  K-1 → Interest: {k1_name} Box 5 = {box5}")

    normalized.extend(rollups)
    return normalized


# ─── PHASE 5: VALIDATE ───────────────────────────────────────────────────────

def validate(extractions, prior_year_context=None):
    """Validate extracted data using accounting arithmetic and rules.

    This is deterministic — no LLM calls. Pure math and accounting logic.
    Checks fall into three categories:
      1. Arithmetic: do the numbers on the document add up?
      2. Relational: do related fields follow known constraints?
      3. Cross-document: do related documents agree?
    """
    print("\n── [PASSION] Phase 5: Validate ──")
    warnings = []

    # Per-extraction checks
    for ext in extractions:
        dtype = ext.get("document_type", "")
        entity = ext.get("payer_or_entity", "")
        fields = ext.get("fields", {})
        label = f"{dtype} ({entity})" if entity else dtype

        # ─── Confidence warnings (existing) ───
        for fname, fdata in fields.items():
            if isinstance(fdata, dict):
                conf = fdata.get("confidence", "")
                if conf == "verified_corrected":
                    warnings.append(f"CORRECTED[CONF_FIELD_CORRECTED]: {label} — {fname}: {fdata.get('original_value')} → {fdata.get('value')}")
                elif conf == "low":
                    warnings.append(f"LOW CONFIDENCE[CONF_FIELD_LOW]: {label} — {fname} = {fdata.get('value')}")
                elif conf == "found_in_verification":
                    warnings.append(f"FOUND IN VERIFY[CONF_FIELD_FOUND_VERIFY]: {label} — {fname} = {fdata.get('value')}")

        # ─── W-2: wage relationship checks ───
        if dtype == "W-2":
            wages = get_val(fields, "wages")
            ss_wages = get_val(fields, "ss_wages")
            medicare_wages = get_val(fields, "medicare_wages")
            fed_wh = get_val(fields, "federal_wh")
            state_wh = get_val(fields, "state_wh")
            ss_wh = get_val(fields, "ss_wh")
            med_wh = get_val(fields, "medicare_wh")

            if wages and medicare_wages and medicare_wages < wages * 0.8:
                warnings.append(f"CHECK[REL_W2_MEDICARE_WAGES_LOW]: {label} — Medicare wages ({medicare_wages:,.2f}) << Box 1 wages ({wages:,.2f})")
            # SS wages capped at wage base ($168,600 for 2024, $176,100 for 2025)
            if wages and ss_wages and ss_wages > wages * 1.01:
                warnings.append(f"ARITH[ACC_W2_SS_WAGES_EXCEEDS_BOX1]: {label} — SS wages ({ss_wages:,.2f}) > Box 1 wages ({wages:,.2f})")
            # SS withholding ≈ 6.2% of SS wages
            if ss_wages and ss_wh:
                expected_ss = ss_wages * 0.062
                if abs(ss_wh - expected_ss) > max(1.0, expected_ss * 0.02):
                    warnings.append(f"CHECK[REL_W2_SS_WH_RATE]: {label} — SS WH ({ss_wh:,.2f}) ≠ 6.2% of SS wages ({expected_ss:,.2f})")
            # Medicare withholding ≈ 1.45% of Medicare wages (plus 0.9% above $200k)
            if medicare_wages and med_wh:
                expected_med = medicare_wages * 0.0145
                if medicare_wages > 200000:
                    expected_med += (medicare_wages - 200000) * 0.009
                if abs(med_wh - expected_med) > max(1.0, expected_med * 0.05):
                    warnings.append(f"CHECK[REL_W2_MEDICARE_WH_RATE]: {label} — Medicare WH ({med_wh:,.2f}) vs expected ({expected_med:,.2f})")
            # Federal WH sanity: shouldn't exceed wages
            if wages and fed_wh and fed_wh > wages:
                warnings.append(f"ARITH[ACC_W2_FED_WH_EXCEEDS_WAGES]: {label} — Federal WH ({fed_wh:,.2f}) > wages ({wages:,.2f})")

        # ─── 1099-DIV: qualified ≤ ordinary ───
        if "1099-DIV" in dtype:
            ordinary = get_val(fields, "ordinary_dividends") or get_val(fields, "div_ordinary_dividends")
            qualified = get_val(fields, "qualified_dividends") or get_val(fields, "div_qualified_dividends")
            if ordinary and qualified and qualified > ordinary + 0.01:
                warnings.append(f"ARITH[ACC_1099DIV_QUALIFIED_EXCEEDS_ORDINARY]: {label} — Qualified ({qualified:,.2f}) > Ordinary ({ordinary:,.2f})")

        # ─── 1099-R: taxable ≤ gross ───
        if "1099-R" in dtype:
            gross = get_val(fields, "gross_distribution")
            taxable = get_val(fields, "taxable_amount")
            fed_wh = get_val(fields, "federal_wh")
            if gross and taxable and taxable > gross + 0.01:
                warnings.append(f"ARITH[ACC_1099R_TAXABLE_EXCEEDS_GROSS]: {label} — Taxable ({taxable:,.2f}) > Gross distribution ({gross:,.2f})")
            if gross and fed_wh and fed_wh > gross:
                warnings.append(f"ARITH[ACC_1099R_FED_WH_EXCEEDS_GROSS]: {label} — Federal WH ({fed_wh:,.2f}) > Gross ({gross:,.2f})")

        # ─── K-1: Box 2 / Box 15 confusion ───
        if "K-1" in dtype:
            box2 = get_val(fields, "box2_rental_real_estate")
            box15 = get_val(fields, "box15_credits")
            box1 = get_val(fields, "box1_ordinary_income")
            if box2 and box15 and box2 == box15 and box2 != 0:
                warnings.append(f"CHECK[REL_K1_BOX2_BOX15_MISMATCH]: {label} — Box 2 ({box2:,.2f}) = Box 15 ({box15:,.2f}), possible misassignment")
            if box2 and box2 > 0 and not box1:
                warnings.append(f"CHECK[REL_K1_BOX2_NO_BOX1]: {label} — Box 2 positive ({box2:,.2f}) with no Box 1; verify not credits")

        # ─── 1099-K: monthly totals ≈ gross ───
        if "1099-K" in dtype:
            gross = get_val(fields, "gross_amount")
            months = ["jan", "feb", "mar", "apr", "may", "jun",
                      "jul", "aug", "sep", "oct", "nov", "dec"]
            monthly_sum = sum(get_val(fields, m) or 0 for m in months)
            if gross and monthly_sum > 0 and abs(gross - monthly_sum) > 1.0:
                warnings.append(f"ARITH[ACC_1099K_GROSS_MONTHLY_MISMATCH]: {label} — Gross ({gross:,.2f}) ≠ sum of monthly ({monthly_sum:,.2f})")

        # ─── Bank statement reconciliation ───
        if "bank_statement" in dtype:
            begin = get_val(fields, "beginning_balance")
            end = get_val(fields, "ending_balance")
            deposits = get_val(fields, "total_deposits") or 0
            withdrawals = get_val(fields, "total_withdrawals") or 0
            fees = get_val(fields, "fees_charged") or 0
            interest = get_val(fields, "interest_earned") or 0

            if begin is not None and end is not None:
                # Fundamental accounting equation for bank statements:
                # ending = beginning + deposits - withdrawals - fees + interest
                expected_end = begin + deposits - withdrawals - fees + interest
                diff = abs(end - expected_end)
                if diff > 1.0:
                    warnings.append(
                        f"ARITH[ACC_BANK_BALANCE_RECONCILIATION]: {label} — Balance doesn't reconcile: "
                        f"begin ({begin:,.2f}) + dep ({deposits:,.2f}) - wdl ({withdrawals:,.2f}) "
                        f"- fees ({fees:,.2f}) + int ({interest:,.2f}) = {expected_end:,.2f}, "
                        f"but ending = {end:,.2f} (diff: {diff:,.2f})")

            # Transaction-level reconciliation
            txn_nums = sorted(set(int(m.group(1)) for k in fields
                                  for m in [re.match(r"txn_(\d+)_", k)] if m))
            if txn_nums and deposits:
                txn_deposits = sum(abs(get_val(fields, f"txn_{n}_amount") or 0)
                                   for n in txn_nums
                                   if (get_str(fields, f"txn_{n}_type") or "").lower() in
                                   ("deposit", "transfer in", "credit"))
                if txn_deposits > 0 and abs(txn_deposits - deposits) > 1.0:
                    warnings.append(
                        f"CHECK[REL_BANK_TXN_DEPOSIT_MISMATCH]: {label} — Sum of deposit txns ({txn_deposits:,.2f}) "
                        f"≠ total deposits ({deposits:,.2f})")

        # ─── Credit card statement reconciliation ───
        if "credit_card" in dtype:
            prev = get_val(fields, "previous_balance") or 0
            payments = get_val(fields, "payments") or 0
            credits = get_val(fields, "credits") or 0
            purchases = get_val(fields, "purchases") or 0
            fees = get_val(fields, "fees_charged") or 0
            interest = get_val(fields, "interest_charged") or 0
            new_bal = get_val(fields, "new_balance")

            if new_bal is not None and prev:
                expected = prev - payments - credits + purchases + fees + interest
                diff = abs(new_bal - expected)
                if diff > 1.0:
                    warnings.append(
                        f"ARITH[ACC_CC_BALANCE_RECONCILIATION]: {label} — CC balance doesn't reconcile: "
                        f"expected {expected:,.2f}, got {new_bal:,.2f} (diff: {diff:,.2f})")

        # ─── Invoice arithmetic ───
        if "invoice" in dtype:
            subtotal = get_val(fields, "subtotal")
            tax = get_val(fields, "tax_amount") or 0
            total = get_val(fields, "total_amount")
            if subtotal is not None and total is not None:
                expected = subtotal + tax
                if abs(total - expected) > 0.05:
                    warnings.append(
                        f"ARITH[ACC_INVOICE_ARITHMETIC]: {label} — subtotal ({subtotal:,.2f}) + tax ({tax:,.2f}) "
                        f"= {expected:,.2f}, but total = {total:,.2f}")

        # ─── Receipt arithmetic ───
        if "receipt" in dtype:
            subtotal = get_val(fields, "subtotal")
            tax = get_val(fields, "tax_amount") or 0
            total = get_val(fields, "total_amount")
            if subtotal is not None and total is not None:
                expected = subtotal + tax
                if abs(total - expected) > 0.10:
                    warnings.append(
                        f"ARITH[ACC_RECEIPT_ARITHMETIC]: {label} — subtotal ({subtotal:,.2f}) + tax ({tax:,.2f}) "
                        f"= {expected:,.2f}, but total = {total:,.2f}")

        # ─── Payroll arithmetic ───
        if "check_stub" in dtype:
            gross = get_val(fields, "gross_pay")
            net = get_val(fields, "net_pay")
            fed = get_val(fields, "federal_wh") or 0
            state = get_val(fields, "state_wh") or 0
            ss = get_val(fields, "social_security") or 0
            med = get_val(fields, "medicare") or 0
            if gross and net:
                total_deductions = fed + state + ss + med
                expected_net = gross - total_deductions
                # Allow 10% tolerance because there may be other deductions
                # (401k, insurance, garnishments) we didn't extract
                if expected_net > 0 and abs(net - expected_net) > expected_net * 0.20:
                    warnings.append(
                        f"CHECK[REL_CHECKSTUB_PAYROLL_DEDUCTIONS]: {label} — gross ({gross:,.2f}) - known deductions ({total_deductions:,.2f}) "
                        f"= {expected_net:,.2f}, but net = {net:,.2f} "
                        f"(possible unextracted deductions: {expected_net - net:,.2f})")

        # ─── Loan statement check ───
        if "loan" in dtype or "mortgage" in dtype:
            principal = get_val(fields, "principal_paid") or 0
            interest = get_val(fields, "interest_paid") or 0
            escrow = get_val(fields, "escrow_paid") or 0
            payment = get_val(fields, "payment_amount")
            if payment and (principal + interest + escrow) > 0:
                component_sum = principal + interest + escrow
                if abs(payment - component_sum) > 1.0:
                    warnings.append(
                        f"CHECK[REL_LOAN_PAYMENT_COMPONENTS]: {label} — Payment ({payment:,.2f}) ≠ principal ({principal:,.2f}) "
                        f"+ interest ({interest:,.2f}) + escrow ({escrow:,.2f}) = {component_sum:,.2f}")

    # ─── Cross-document checks ───
    # Bank statements: consecutive months should have ending = next beginning
    bank_stmts = [e for e in extractions if "bank_statement" in str(e.get("document_type", ""))]
    if len(bank_stmts) > 1:
        for i in range(len(bank_stmts) - 1):
            e1 = bank_stmts[i]
            e2 = bank_stmts[i + 1]
            end1 = get_val(e1.get("fields", {}), "ending_balance")
            begin2 = get_val(e2.get("fields", {}), "beginning_balance")
            if end1 is not None and begin2 is not None and abs(end1 - begin2) > 0.01:
                ent1 = e1.get("payer_or_entity", "")
                ent2 = e2.get("payer_or_entity", "")
                if ent1.upper() == ent2.upper():
                    warnings.append(
                        f"CROSS-DOC[XDOC_BANK_CONTINUITY]: {ent1} — ending balance ({end1:,.2f}) ≠ "
                        f"next beginning balance ({begin2:,.2f})")

    # ─── Duplicate document detection ───
    # Flag when same payer + same doc type + same key amounts appear twice
    seen_docs = {}  # key → (entity, page, key_amount)
    for ext in extractions:
        dtype = str(ext.get("document_type", ""))
        entity = ext.get("payer_or_entity", "")
        ein = ext.get("payer_ein", "")
        fields = ext.get("fields", {})
        page = ext.get("_page", "?")

        # Build a fingerprint from key amounts depending on doc type
        key_vals = []
        if "W-2" in dtype:
            key_vals = [get_val(fields, "wages"), get_val(fields, "federal_wh")]
        elif "1099-DIV" in dtype:
            key_vals = [get_val(fields, "ordinary_dividends")]
        elif "1099-INT" in dtype:
            key_vals = [get_val(fields, "interest_income")]
        elif "1099-R" in dtype:
            key_vals = [get_val(fields, "gross_distribution")]
        elif "1099-NEC" in dtype:
            key_vals = [get_val(fields, "nonemployee_compensation")]
        elif "SSA-1099" in dtype:
            key_vals = [get_val(fields, "net_benefits")]
        elif "1098" in dtype:
            key_vals = [get_val(fields, "mortgage_interest")]

        # Only check if we have real values
        key_vals = [v for v in key_vals if v is not None and v != 0]
        if key_vals and (ein or entity):
            ident = ein or entity.upper().strip()
            fingerprint = f"{dtype}|{ident}|{'|'.join(f'{v:.2f}' for v in key_vals)}"
            if fingerprint in seen_docs:
                prev_entity, prev_page, _ = seen_docs[fingerprint]
                warnings.append(
                    f"CROSS-DOC[XDOC_DUPLICATE_DOCUMENT]: Possible duplicate — {dtype} from {entity} "
                    f"(page {page}) has same amounts as page {prev_page}. "
                    f"Check if scanned twice.")
            else:
                seen_docs[fingerprint] = (entity, page, key_vals)

    # ─── Prior-year variance detection ───
    if prior_year_context:
        prior_payers = {}
        for doc in prior_year_context.get("prior_year_data", {}).get("documents", []):
            for p in doc.get("payers", []):
                pein = p.get("ein", "")
                pname = p.get("name", "").upper().strip()
                pform = p.get("form_type", "")
                pamounts = p.get("amounts", [])
                key = pein if pein else pname
                if key and pamounts:
                    prior_payers[f"{pform}|{key}"] = {
                        "name": p.get("name", ""),
                        "form": pform,
                        "amount": max(pamounts),  # Use largest amount for comparison
                    }

        for ext in extractions:
            dtype = str(ext.get("document_type", ""))
            entity = ext.get("payer_or_entity", "")
            ein = ext.get("payer_ein", "")
            fields = ext.get("fields", {})

            # Get the primary dollar amount for this extraction
            current_amt = None
            if "W-2" in dtype:
                current_amt = get_val(fields, "wages")
            elif "1099-DIV" in dtype:
                current_amt = get_val(fields, "ordinary_dividends")
            elif "1099-INT" in dtype:
                current_amt = get_val(fields, "interest_income")
            elif "1099-R" in dtype:
                current_amt = get_val(fields, "gross_distribution")
            elif "1099-NEC" in dtype:
                current_amt = get_val(fields, "nonemployee_compensation")
            elif "K-1" in dtype:
                current_amt = get_val(fields, "box1_ordinary_income")

            if current_amt is not None and current_amt != 0:
                ident = ein if ein else entity.upper().strip()
                # Try to match against prior year
                for form_prefix in [dtype, dtype.split("-")[0] if "-" in dtype else dtype]:
                    prior_key = f"{form_prefix}|{ident}"
                    if prior_key in prior_payers:
                        prior = prior_payers[prior_key]
                        prior_amt = prior["amount"]
                        if prior_amt and prior_amt != 0:
                            pct_change = abs(current_amt - prior_amt) / abs(prior_amt) * 100
                            if pct_change > 50:
                                direction = "↑" if current_amt > prior_amt else "↓"
                                warnings.append(
                                    f"VARIANCE[VAR_PRIOR_YEAR_THRESHOLD]: {entity} {dtype} — {direction} "
                                    f"{pct_change:.0f}% vs prior year "
                                    f"(${current_amt:,.2f} vs PY ${prior_amt:,.2f})")
                        break

    if warnings:
        arith = sum(1 for w in warnings if w.startswith("ARITH"))
        checks = sum(1 for w in warnings if w.startswith("CHECK"))
        cross = sum(1 for w in warnings if w.startswith("CROSS"))
        variance = sum(1 for w in warnings if w.startswith("VARIANCE"))
        other = len(warnings) - arith - checks - cross - variance
        print(f"  ⚠ {len(warnings)} items: {arith} arithmetic, {checks} checks, "
              f"{cross} cross-doc, {variance} variance, {other} other")
        for w in warnings: print(f"    {w}")
    else:
        print("  ✓ No warnings")
    return warnings


# ─── PHASE 6: EXCEL OUTPUT ───────────────────────────────────────────────────

BOLD = Font(bold=True)
SECTION_FONT = Font(bold=True, size=11, color="000000")
SECTION_FILL = PatternFill("solid", fgColor="D9D9D9")     # Light gray — section header rows
COL_HEADER_FONT = Font(size=11, color="000000")            # Normal weight, matches section row
COL_HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")   # Same gray as section header
MONEY_FMT = '#,##0.00_);(#,##0.00)'                        # Accounting: parentheses for negatives
PCT_FMT = '0.00%'
DATE_FMT = 'MM/DD/YYYY'
SUM_FONT = Font(bold=True, size=11, color="000000")
SUM_FILL = PatternFill()                                    # No fill on total rows
FLAG_FILL = PatternFill("solid", fgColor="FFFDE7")         # Soft yellow — low confidence
CORRECTED_FILL = PatternFill("solid", fgColor="C8E6C9")   # Green — corrected
REVIEW_FILL = PatternFill("solid", fgColor="FFE0B2")      # Orange — needs human
CONFIRMED_FILL = PatternFill("solid", fgColor="E8F5E9")   # Light green — confirmed
DUAL_FILL = PatternFill("solid", fgColor="A5D6A7")        # Darker green — OCR + image agree
FLAG_FONT = Font(italic=True, color="CC0000")
ALT_ROW_FILL = PatternFill()                               # No alternating rows
DARK_HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
DARK_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
THIN_BORDER = openpyxl.styles.Border()                     # No cell borders on data rows
SECTION_BORDER = openpyxl.styles.Border()                  # No border on section headers
SUM_BORDER = openpyxl.styles.Border()                      # No border on totals

def populate_template(extractions, template_path, output_path, year, output_format="tax_review"):
    """Router: create workbook, delegate to format-specific function, save."""
    fmt_labels = {
        "tax_review": "Tax Review", "journal_entries": "Journal Entries",
        "account_balances": "Account Balances", "trial_balance": "Trial Balance",
        "transaction_register": "Transaction Register",
    }
    print(f"\n── [PASSION] Phase 6: Excel ({fmt_labels.get(output_format, output_format)}) ──")

    if template_path and os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
    else:
        wb = openpyxl.Workbook()

    sheet_name = str(year)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name, 0)

    if output_format == "journal_entries":
        _populate_journal_entries(ws, extractions, year)
    elif output_format == "account_balances":
        _populate_account_balances(ws, extractions, year)
    elif output_format == "trial_balance":
        _populate_trial_balance(ws, extractions, year)
    elif output_format == "transaction_register":
        _populate_transaction_register(ws, extractions, year)
    else:
        _populate_tax_review(ws, extractions, year)

    ws.freeze_panes = "A4"

    # Print setup for all formats
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.oddHeader.center.text = f"&B{fmt_labels.get(output_format, 'Document Intake')} — {year}"
    ws.oddHeader.center.size = 10
    ws.oddFooter.left.text = "Bearden Accounting — Document Intake v5"
    ws.oddFooter.left.size = 8
    ws.oddFooter.right.text = "Page &P of &N"
    ws.oddFooter.right.size = 8
    ws.print_options.gridLines = True

    # Remove default sheet if it exists and is empty
    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row <= 1:
        del wb["Sheet"]

    wb.save(output_path)
    print(f"  ✓ Saved: {output_path}")


# ─── Shared helpers for section-based rendering ──────────────────────────────

def _match_exts(extractions, match_types):
    matched = []
    for ext in extractions:
        dtype = str(ext.get("document_type", ""))
        for mt in match_types:
            if mt in dtype:
                matched.append(ext)
                break
    return matched

def _dedup_by_ein(exts):
    """Deduplicate extractions by EIN/entity, keeping the BEST version.

    When two copies of the same document exist (common with scanned returns),
    keep the one with more high-confidence fields instead of first-seen.
    Missing fields from the discarded copy are merged into the winner.
    """
    seen = {}       # key → (ext, score)
    order = []      # preserve insertion order of keys

    def _confidence_score(ext):
        """Score an extraction by field confidence. Higher = better."""
        score = 0
        for fdata in ext.get("fields", {}).values():
            if not isinstance(fdata, dict):
                score += 1  # bare value = ok
                continue
            c = fdata.get("confidence", "")
            if c in ("dual_confirmed", "verified_confirmed"):
                score += 3
            elif c in ("verified_corrected", "found_in_verification"):
                score += 2
            elif c in ("high", "ocr_accepted"):
                score += 2
            elif c in ("medium", "from_continuation", "from_k1_box5"):
                score += 1
            elif c in ("low",):
                score += 0
            else:
                score += 1
        return score

    for ext in exts:
        ein = ext.get("payer_ein", "")
        fields = ext.get("fields", {})
        name = get_str(fields, "partnership_name") or get_str(fields, "employer_name") or ext.get("payer_or_entity") or ""
        recip = ext.get("recipient", "")
        if ein:
            key = f"EIN:{ein}|{recip}".upper()
        else:
            norm = re.sub(r'\s*\(\d{4}\)\s*', '', name).upper().strip()
            key = f"NAME:{norm}|{recip}".upper()

        score = _confidence_score(ext)

        if key not in seen:
            seen[key] = (ext, score)
            order.append(key)
        else:
            existing_ext, existing_score = seen[key]
            if score > existing_score:
                # New copy is better — swap, merge missing fields from old
                for fname, fdata in existing_ext.get("fields", {}).items():
                    if fname not in ext.get("fields", {}):
                        ext.setdefault("fields", {})[fname] = fdata
                seen[key] = (ext, score)
                print(f"  Dedup: kept better copy of {name or ein} (score {score} > {existing_score})")
            else:
                # Existing is better — merge missing fields from new
                for fname, fdata in ext.get("fields", {}).items():
                    if fname not in existing_ext.get("fields", {}):
                        existing_ext.setdefault("fields", {})[fname] = fdata

    return [seen[k][0] for k in order]

def _write_title(ws, title, year):
    """Write title rows, return next row number."""
    ws["A1"] = f"{title} — {year}"
    ws["A1"].font = Font(bold=True, size=14, color="000000")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:F1")
    ws["A2"] = f"Extracted {datetime.now().strftime('%m/%d/%Y %I:%M %p')}"
    ws["A2"].font = Font(italic=True, color="999999", size=9)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:F2")
    return 4

def _write_cell_value(ws, col, row, fields, field_name, ext, matched):
    """Write a single field cell with formatting and confidence coloring."""
    if field_name == "_display_name":
        name = get_str(fields, "partnership_name") or ext.get("payer_or_entity", "")
        recip = ext.get("recipient", "")
        same = [e for e in matched if (e.get("payer_ein","") == ext.get("payer_ein","") and ext.get("payer_ein",""))
                or (get_str(e.get("fields",{}), "partnership_name") or e.get("payer_or_entity","")).upper() == name.upper()]
        if len(same) > 1 and recip:
            name = f"{name} - {recip}"
        ws[f"{col}{row}"] = name
    elif field_name == "_source_name":
        ws[f"{col}{row}"] = ext.get("_source_name", ext.get("payer_or_entity", ""))
    elif field_name == "_carryover_prior":
        cell = ws[f"{col}{row}"]
        cell.value = 0
        cell.number_format = MONEY_FMT
        cell.alignment = Alignment(horizontal="right")
        cell.fill = REVIEW_FILL
        cell.comment = Comment("REQUIRES PRIOR YEAR DATA", "System")
    elif field_name in ("_allowed", "_carryover_next"):
        ws[f"{col}{row}"] = None
        ws[f"{col}{row}"].fill = REVIEW_FILL
        ws[f"{col}{row}"].comment = Comment("REQUIRES PREPARER JUDGMENT", "System")
    elif field_name == "employer_name":
        name = get_str(fields, "employer_name") or ext.get("payer_or_entity", "")
        name = re.sub(r'\s*\(\d{4}\)\s*$', '', name)
        ws[f"{col}{row}"] = name
    elif field_name in ("payer_or_entity", "institution_name"):
        ws[f"{col}{row}"] = get_str(fields, field_name) or ext.get("payer_or_entity", "")
    else:
        val = get_val(fields, field_name)
        if val is None:
            val = get_str(fields, field_name)
        cell = ws[f"{col}{row}"]
        cell.value = val
        if isinstance(val, (int, float)):
            cell.number_format = MONEY_FMT
            cell.alignment = Alignment(horizontal="right")
        # Confidence comments (no cell coloring — keeps spreadsheet clean)
        fdata = fields.get(field_name)
        if isinstance(fdata, dict):
            conf = fdata.get("confidence", "")
            if conf == "verified_corrected":
                cell.comment = Comment(f"Corrected: was {fdata.get('original_value','?')}. {fdata.get('correction_note','')}", "System")
            elif conf == "low":
                cell.comment = Comment("Low confidence — check source", "System")
            elif conf == "operator_corrected":
                cell.comment = Comment(f"Operator corrected (was {fdata.get('_original_value','?')})", "Operator")


# ─── TAX REVIEW (original format) ────────────────────────────────────────────

def _populate_tax_review(ws, extractions, year):
    row = _write_title(ws, "Document Intake", year)
    k1_extras = []

    for section in TEMPLATE_SECTIONS:
        sid = section["id"]

        if section.get("special") == "k1_extras":
            if k1_extras:
                ws[f"A{row}"] = section["header"]
                ws[f"A{row}"].font = SECTION_FONT
                ws[f"A{row}"].fill = SECTION_FILL
                for hcol, hlabel in [("B", "Line Ref"), ("C", "Description"), ("D", "Amount")]:
                    cell = ws[f"{hcol}{row}"]
                    cell.value = hlabel
                    cell.font = COL_HEADER_FONT
                    cell.fill = COL_HEADER_FILL
                    cell.alignment = Alignment(horizontal="center")
                for c in ["E", "F"]:
                    ws[f"{c}{row}"].fill = SECTION_FILL
                row += 1
                for item in k1_extras:
                    ws[f"A{row}"] = item.get("entity", "")
                    ws[f"B{row}"] = item.get("line_reference", "")
                    ws[f"C{row}"] = item.get("description", "")
                    amt_cell = ws[f"D{row}"]
                    amt_cell.value = item.get("amount")
                    if isinstance(amt_cell.value, (int, float)):
                        amt_cell.number_format = MONEY_FMT
                        amt_cell.alignment = Alignment(horizontal="right")
                    for c in ["A", "B", "C", "D"]:
                        ws[f"{c}{row}"].border = THIN_BORDER
                    row += 1
                row += 1
            continue

        match_types = section.get("match_types", [])
        matched = _match_exts(extractions, match_types) if match_types else []
        col_headers = section.get("col_headers", {})
        columns = section.get("columns", {})
        sum_cols = section.get("sum_cols", [])
        flags = section.get("flags", [])

        if not matched and sid not in ALWAYS_SHOW:
            continue

        ws[f"A{row}"] = section["header"]
        ws[f"A{row}"].font = SECTION_FONT
        ws[f"A{row}"].fill = SECTION_FILL
        for col, label in col_headers.items():
            cell = ws[f"{col}{row}"]
            cell.value = label
            cell.font = COL_HEADER_FONT
            cell.fill = COL_HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        # Fill remaining columns in header row with gray
        for c in ["A", "B", "C", "D", "E", "F"]:
            if c not in col_headers:
                ws[f"{c}{row}"].fill = SECTION_FILL
        row += 1

        if not matched:
            ws[f"A{row}"] = "(no documents found)"
            ws[f"A{row}"].font = Font(italic=True, color="BBBBBB")
            row += 2
            continue

        matched = _dedup_by_ein(matched)

        # Filter out zero-value entries for interest/dividend sections
        if sid in ("interest", "dividends"):
            matched = [e for e in matched if any(
                (get_val(e.get("fields", {}), fn) or 0) != 0
                for fn in columns.values() if not fn.startswith("_")
            )]
            if not matched:
                ws[f"A{row}"] = "(no documents found)"
                ws[f"A{row}"].font = Font(italic=True, color="BBBBBB")
                row += 2
                continue

        data_start = row
        all_cols = list(columns.keys())

        field_aliases = section.get("field_aliases", {})

        for ext_idx, ext in enumerate(matched):
            fields = ext.get("fields", {})
            # Resolve field aliases: if primary field not found, try alternates
            if field_aliases:
                for primary, alternates in field_aliases.items():
                    if primary not in fields:
                        for alt in alternates:
                            if alt in fields:
                                fields[primary] = fields[alt]
                                break
            for col, field_name in columns.items():
                _write_cell_value(ws, col, row, fields, field_name, ext, matched)

            # K-1 extras
            if sid == "k1":
                entity_name = get_str(fields, "partnership_name") or ext.get("payer_or_entity", "")
                extras_map = {
                    "box5_interest": "Box 5 (Interest)",
                    "box6a_ordinary_dividends": "Box 6a (Ordinary Dividends)",
                    "box6b_qualified_dividends": "Box 6b (Qualified Dividends)",
                    "box7_royalties": "Box 7 (Royalties)",
                    "box8_short_term_capital_gain": "Box 8 (ST Cap Gain)",
                    "box9a_long_term_capital_gain": "Box 9a (LT Cap Gain)",
                    "box9c_unrecaptured_1250": "Box 9c (Unrec 1250)",
                    "box10_net_1231_gain": "Box 10 (1231 Gain)",
                    "box11_other_income": "Box 11 (Other Income)",
                    "box12_section_179": "Box 12 (Sec 179)",
                    "box13_other_deductions": "Box 13 (Other Ded)",
                    "box14_self_employment": "Box 14 (SE Earnings)",
                    "box15_credits": "Box 15 (Credits)",
                    "box17_alt_min_tax": "Box 17 (AMT)",
                    "box18_tax_exempt_income": "Box 18 (Tax Exempt)",
                    "box19_distributions": "Box 19 (Distributions)",
                    "box20_other_info": "Box 20 (Other Info)",
                }
                for fkey, label in extras_map.items():
                    val = get_val(fields, fkey)
                    if val and val != 0:
                        k1_extras.append({"entity": entity_name, "line_reference": label, "description": "", "amount": val})
                for ci in ext.get("continuation_items", []):
                    if ci.get("amount") and ci["amount"] != 0:
                        k1_extras.append({"entity": entity_name, "line_reference": ci.get("line_reference", ""), "description": ci.get("description", ""), "amount": ci["amount"]})

            for bcol in list(columns.keys()):
                ws[f"{bcol}{row}"].border = THIN_BORDER
            row += 1

        data_end = row - 1
        if data_end >= data_start and sum_cols:
            for col in sum_cols:
                cell = ws[f"{col}{row}"]
                cell.value = f"=SUM({col}{data_start}:{col}{data_end})"
                cell.font = SUM_FONT
                cell.fill = SUM_FILL
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right")
                cell.border = SUM_BORDER
            ws[f"A{row}"].font = SUM_FONT
            ws[f"A{row}"].fill = SUM_FILL
            ws[f"A{row}"].value = "TOTAL"
            ws[f"A{row}"].border = SUM_BORDER
            if section.get("total_formula_col"):
                for col, formula in section["total_formula_col"].items():
                    parts = formula.split("+")
                    cell = ws[f"{col}{row}"]
                    cell.value = "=" + "+".join([f"{p}{row}" for p in parts])
                    cell.font = SUM_FONT
                    cell.number_format = MONEY_FMT
                    cell.alignment = Alignment(horizontal="right")
            row += 1

        for flag in flags:
            ws[f"A{row}"] = flag
            ws[f"A{row}"].font = FLAG_FONT
            row += 1
        row += 1

    # Schedule A — cross-references extracted data
    ws[f"A{row}"] = "Schedule A:"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    # Column headers for Schedule A
    ws[f"C{row}"] = "Total"
    ws[f"C{row}"].font = COL_HEADER_FONT
    ws[f"C{row}"].fill = COL_HEADER_FILL
    ws[f"C{row}"].alignment = Alignment(horizontal="center")
    ws[f"E{row}"] = "Allowed"
    ws[f"E{row}"].font = COL_HEADER_FONT
    ws[f"E{row}"].fill = COL_HEADER_FILL
    ws[f"E{row}"].alignment = Alignment(horizontal="center")
    for c in ["B", "D", "F"]:
        ws[f"{c}{row}"].fill = SECTION_FILL
    row += 1

    # Medical
    ws[f"A{row}"] = "Medical:"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="333333")
    row += 1
    total_medical = 0
    medical_exts = [e for e in extractions if e.get("document_type") in ("receipt",)
                    and get_str(e.get("fields", {}), "category") == "medical"]
    for mext in medical_exts:
        mfields = mext.get("fields", {})
        mamt = get_val(mfields, "total_amount") or 0
        ws[f"A{row}"] = get_str(mfields, "vendor_name") or mext.get("payer_or_entity", "")
        mcell = ws[f"C{row}"]
        mcell.value = mamt
        mcell.number_format = MONEY_FMT
        mcell.alignment = Alignment(horizontal="right")
        total_medical += mamt
        row += 1
    if not medical_exts:
        ws[f"A{row}"] = "(none found)"
        ws[f"A{row}"].font = Font(italic=True, color="BBBBBB")
        row += 1
    ws[f"A{row}"] = "Total Medical"
    ws[f"A{row}"].font = SUM_FONT
    ws[f"A{row}"].fill = SUM_FILL
    mcell = ws[f"D{row}"]
    mcell.value = total_medical
    mcell.number_format = MONEY_FMT
    mcell.alignment = Alignment(horizontal="right")
    mcell.font = SUM_FONT
    mcell.fill = SUM_FILL
    row += 1

    # Taxes
    ws[f"A{row}"] = "Taxes:"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="333333")
    row += 1

    # Income Taxes — State WH from W-2s
    total_state_wh = sum(get_val(e.get("fields", {}), "state_wh") or 0
                         for e in extractions if e.get("document_type") == "W-2")
    ws[f"A{row}"] = "Income Taxes:"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="333333")
    row += 1
    ws[f"A{row}"] = "State Withholding"
    tcell = ws[f"C{row}"]
    tcell.value = total_state_wh
    tcell.number_format = MONEY_FMT
    tcell.alignment = Alignment(horizontal="right")
    row += 1
    ws[f"A{row}"] = "Total State Tax"
    ws[f"A{row}"].font = SUM_FONT
    ws[f"A{row}"].fill = SUM_FILL
    tcell = ws[f"D{row}"]
    tcell.value = total_state_wh
    tcell.number_format = MONEY_FMT
    tcell.alignment = Alignment(horizontal="right")
    tcell.font = SUM_FONT
    tcell.fill = SUM_FILL
    row += 1

    # Real Estate Taxes — from 1098 property_tax + property_tax_bill
    ws[f"A{row}"] = "Real Estate Taxes:"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="333333")
    row += 1
    total_re_tax = 0
    for rext in extractions:
        if rext.get("document_type") == "1098":
            prop_tax = (get_val(rext.get("fields", {}), "property_tax")
                        or get_val(rext.get("fields", {}), "real_estate_tax") or 0)
            if prop_tax:
                rentity = rext.get("payer_or_entity", "1098")
                ws[f"A{row}"] = f"{rentity} (1098)"
                rcell = ws[f"C{row}"]
                rcell.value = prop_tax
                rcell.number_format = MONEY_FMT
                rcell.alignment = Alignment(horizontal="right")
                total_re_tax += prop_tax
                row += 1
    for rext in extractions:
        if rext.get("document_type") == "property_tax_bill":
            tax_amt = (get_val(rext.get("fields", {}), "tax_amount")
                       or get_val(rext.get("fields", {}), "total_due")
                       or get_val(rext.get("fields", {}), "total_tax")
                       or get_val(rext.get("fields", {}), "total_estimated_tax") or 0)
            if tax_amt:
                addr = get_str(rext.get("fields", {}), "property_address") or "Property"
                ws[f"A{row}"] = addr
                rcell = ws[f"C{row}"]
                rcell.value = tax_amt
                rcell.number_format = MONEY_FMT
                rcell.alignment = Alignment(horizontal="right")
                total_re_tax += tax_amt
                row += 1
    if total_re_tax == 0:
        ws[f"A{row}"] = "(none found)"
        ws[f"A{row}"].font = Font(italic=True, color="BBBBBB")
        row += 1
    ws[f"A{row}"] = "Total Real Estate Tax"
    ws[f"A{row}"].font = SUM_FONT
    ws[f"A{row}"].fill = SUM_FILL
    rcell = ws[f"D{row}"]
    rcell.value = total_re_tax
    rcell.number_format = MONEY_FMT
    rcell.alignment = Alignment(horizontal="right")
    rcell.font = SUM_FONT
    rcell.fill = SUM_FILL
    row += 1

    # Total Taxes
    total_taxes = total_state_wh + total_re_tax
    ws[f"A{row}"] = "Total Taxes:"
    ws[f"A{row}"].font = SUM_FONT
    tcell = ws[f"D{row}"]
    tcell.value = total_taxes
    tcell.number_format = MONEY_FMT
    tcell.alignment = Alignment(horizontal="right")
    tcell.font = SUM_FONT
    ecell = ws[f"E{row}"]
    ecell.value = total_taxes
    ecell.number_format = MONEY_FMT
    ecell.alignment = Alignment(horizontal="right")
    ecell.font = SUM_FONT
    row += 1

    # Mortgage Interest — from 1098
    ws[f"A{row}"] = "Mortgage Interest:"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="333333")
    row += 1
    total_mortgage = 0
    for mext in extractions:
        if mext.get("document_type") == "1098":
            mort_int = (get_val(mext.get("fields", {}), "mortgage_interest")
                        or get_val(mext.get("fields", {}), "mortgage_interest_received") or 0)
            if mort_int:
                mentity = mext.get("payer_or_entity", "1098")
                ws[f"A{row}"] = mentity
                mcell = ws[f"C{row}"]
                mcell.value = mort_int
                mcell.number_format = MONEY_FMT
                mcell.alignment = Alignment(horizontal="right")
                total_mortgage += mort_int
                row += 1
    if total_mortgage == 0:
        ws[f"A{row}"] = "(none found)"
        ws[f"A{row}"].font = Font(italic=True, color="BBBBBB")
        row += 1
    ws[f"A{row}"] = "Total Mortgage Interest"
    ws[f"A{row}"].font = SUM_FONT
    mcell = ws[f"D{row}"]
    mcell.value = total_mortgage
    mcell.number_format = MONEY_FMT
    mcell.alignment = Alignment(horizontal="right")
    mcell.font = SUM_FONT
    ecell = ws[f"E{row}"]
    ecell.value = total_mortgage
    ecell.number_format = MONEY_FMT
    ecell.alignment = Alignment(horizontal="right")
    ecell.font = SUM_FONT
    row += 1

    # Charitable Contributions
    ws[f"A{row}"] = "Donations:"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="333333")
    row += 1
    total_donations = 0
    for dext in extractions:
        if dext.get("document_type") == "charitable_receipt":
            damt = get_val(dext.get("fields", {}), "donation_amount") or 0
            if damt:
                dorg = get_str(dext.get("fields", {}), "organization_name") or dext.get("payer_or_entity", "")
                ws[f"A{row}"] = dorg
                dcell = ws[f"C{row}"]
                dcell.value = damt
                dcell.number_format = MONEY_FMT
                dcell.alignment = Alignment(horizontal="right")
                total_donations += damt
                row += 1
    if total_donations == 0:
        ws[f"A{row}"] = "(none found)"
        ws[f"A{row}"].font = Font(italic=True, color="BBBBBB")
        row += 1
    ws[f"A{row}"] = "Total Donations"
    ws[f"A{row}"].font = SUM_FONT
    dcell = ws[f"D{row}"]
    dcell.value = total_donations
    dcell.number_format = MONEY_FMT
    dcell.alignment = Alignment(horizontal="right")
    dcell.font = SUM_FONT
    ecell = ws[f"E{row}"]
    ecell.value = total_donations
    ecell.number_format = MONEY_FMT
    ecell.alignment = Alignment(horizontal="right")
    ecell.font = SUM_FONT
    row += 1

    # Schedule A grand total
    sched_a_total = total_taxes + total_mortgage + total_donations
    row += 1
    gcell = ws[f"E{row}"]
    gcell.value = sched_a_total
    gcell.number_format = MONEY_FMT
    gcell.alignment = Alignment(horizontal="right")
    gcell.font = SUM_FONT
    row += 1

    # Column widths
    ws.column_dimensions["A"].width = 33
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 15


# ─── JOURNAL ENTRIES FORMAT ──────────────────────────────────────────────────

# Fields on tax documents that are informational, not dollar amounts to post.
# These must never become journal entry lines.
_NON_POSTING_FIELDS = {
    # K-1 informational
    "partnership_name", "partnership_ein", "partner_name", "entity_type",
    "partner_type", "profit_share_begin", "profit_share_end",
    "loss_share_begin", "loss_share_end", "capital_share_begin",
    "capital_share_end", "beginning_capital_account", "ending_capital_account",
    "current_year_net_income", "withdrawals_distributions", "capital_contributed",
    # Identifiers on any form
    "employer_ein", "payer_ein", "recipient_ssn_last4", "tax_year",
    "state_id", "account_number", "account_number_last4",
    # Statement period / dates
    "statement_period_start", "statement_period_end",
    "pay_period_start", "pay_period_end",
    # Counts / metadata
    "num_deposits", "num_withdrawals", "number_of_transactions",
    "hours_regular", "hours_overtime", "rate_regular", "rate_overtime",
    "distribution_code", "identifiable_event_code", "type_of_wager",
}


def _build_journal_entries(extractions, year):
    """Transform extractions into balanced double-entry journal entries.

    ACCOUNTING RULES:
    1. Every entry MUST balance: total debits == total credits.
       Entries that fail are flagged "UNBALANCED" and still included
       (so the operator sees them) but marked for review.
    2. Tax information documents (W-2, 1099-*, K-1, SSA-1099, 1098-*,
       5498-*, 1095-*) are NOT source transactions. They report income
       for tax return preparation. They do NOT generate journal entries.
    3. A bank deposit is not revenue until classified. Posts to
       "Unclassified Deposits" pending operator review.
    4. A withdrawal is not a known expense until classified. Posts to
       "Unclassified Expenses" pending operator review.
    5. Credit cards are liabilities (credit-normal). Purchases increase
       the liability (credit CC, debit expense). Payments decrease it.
    6. Payroll: debit Wages Expense for gross, credit each tax payable,
       credit Cash for net. If net is missing, compute it.

    Returns list of:
      {"date", "description", "lines": [{"account", "debit", "credit"}],
       "balanced": bool, "source_type": str}
    """
    journal = []

    def _post(date, description, lines, source_type=""):
        """Validate balance and append entry. Flag if unbalanced."""
        # Round everything to 2 decimals
        for line in lines:
            if line.get("debit"):
                line["debit"] = round(line["debit"], 2)
            else:
                line["debit"] = None
            if line.get("credit"):
                line["credit"] = round(line["credit"], 2)
            else:
                line["credit"] = None

        total_dr = round(sum(l["debit"] or 0 for l in lines), 2)
        total_cr = round(sum(l["credit"] or 0 for l in lines), 2)
        diff = round(total_dr - total_cr, 2)
        balanced = abs(diff) < 0.02

        # Auto-fix small rounding differences (< $1.00)
        if not balanced and 0.02 <= abs(diff) < 1.00 and total_dr > 0 and total_cr > 0:
            if diff > 0:
                lines.append({"account": "Rounding", "debit": None, "credit": abs(diff)})
            else:
                lines.append({"account": "Rounding", "debit": abs(diff), "credit": None})
            balanced = True

        entry = {
            "date": date,
            "description": description,
            "lines": lines,
            "balanced": balanced,
            "source_type": source_type,
        }
        if not balanced:
            entry["description"] = (
                f"\u26a0 UNBALANCED \u2014 {description} "
                f"(DR {total_dr:,.2f} \u2260 CR {total_cr:,.2f})"
            )
        journal.append(entry)

    def _get_cat(fields, field_name, default=""):
        """Read operator-assigned category from a field, if present."""
        fdata = fields.get(field_name)
        if isinstance(fdata, dict):
            return fdata.get("_operator_category", default)
        return default

    for ext in extractions:
        dtype = str(ext.get("document_type", ""))
        fields = ext.get("fields", {})
        entity = ext.get("payer_or_entity", "")

        # ═══════════════════════════════════════════════════════════════
        # TAX INFORMATION DOCUMENTS — DO NOT POST
        # ═══════════════════════════════════════════════════════════════
        # W-2, 1099-*, K-1, SSA-1099, 1098-*, 5498-*, 1095-*, W-2G
        # are information returns. They tell the IRS (and the taxpayer)
        # what happened, but they are NOT the transactions themselves.
        # The actual transactions are on bank statements, checks, etc.
        # These belong in tax_review format, not journal entries.
        if any(tag in dtype for tag in (
            "W-2", "1099", "K-1", "SSA", "1098", "5498", "1095", "W-2G",
        )):
            continue

        # ═══════════════════════════════════════════════════════════════
        # BANK STATEMENTS
        # ═══════════════════════════════════════════════════════════════
        # Cash is an ASSET (debit-normal).
        #   Deposit:    DR Cash / CR Unclassified Deposits
        #   Withdrawal: DR Unclassified Expenses / CR Cash
        #   Fee:        DR Bank Service Charges / CR Cash
        #   Interest:   DR Cash / CR Interest Income
        if "bank_statement" in dtype:
            stmt_date = get_str(fields, "statement_period_end") or f"12/31/{year}"
            bank = get_str(fields, "bank_name") or entity
            acct = get_str(fields, "account_number_last4") or ""
            cash_acct = f"Cash \u2014 {bank}" + (f" (...{acct[-4:]})" if acct else "")

            txn_nums = sorted(set(
                int(m.group(1)) for k in fields
                for m in [re.match(r"txn_(\d+)_", k)] if m
            ))

            if txn_nums:
                for n in txn_nums:
                    tdate = get_str(fields, f"txn_{n}_date") or stmt_date
                    tdesc = get_str(fields, f"txn_{n}_desc") or f"Transaction #{n}"
                    tamt = get_val(fields, f"txn_{n}_amount")
                    ttype = (get_str(fields, f"txn_{n}_type") or "").lower()
                    if not tamt:
                        continue
                    amt = round(abs(tamt), 2)
                    # Operator may have categorized this transaction
                    cat = _get_cat(fields, f"txn_{n}_amount")

                    if ttype in ("deposit", "transfer in", "credit"):
                        cr_acct = cat or "Unclassified Deposits"
                        _post(tdate, tdesc, [
                            {"account": cash_acct, "debit": amt, "credit": None},
                            {"account": cr_acct, "debit": None, "credit": amt},
                        ], "bank_txn")
                    elif ttype == "fee":
                        _post(tdate, f"{bank}: {tdesc}", [
                            {"account": cat or "Bank Service Charges", "debit": amt, "credit": None},
                            {"account": cash_acct, "debit": None, "credit": amt},
                        ], "bank_txn")
                    elif ttype == "interest":
                        _post(tdate, f"{bank}: {tdesc}", [
                            {"account": cash_acct, "debit": amt, "credit": None},
                            {"account": cat or "Interest Income", "debit": None, "credit": amt},
                        ], "bank_txn")
                    else:
                        dr_acct = cat or "Unclassified Expenses"
                        _post(tdate, tdesc, [
                            {"account": dr_acct, "debit": amt, "credit": None},
                            {"account": cash_acct, "debit": None, "credit": amt},
                        ], "bank_txn")
            else:
                # Summary-level entries
                deposits = get_val(fields, "total_deposits")
                if deposits:
                    _post(stmt_date, f"{cash_acct} \u2014 Total deposits", [
                        {"account": cash_acct, "debit": abs(deposits), "credit": None},
                        {"account": "Unclassified Deposits", "debit": None, "credit": abs(deposits)},
                    ], "bank_summary")
                withdrawals = get_val(fields, "total_withdrawals")
                if withdrawals:
                    _post(stmt_date, f"{cash_acct} \u2014 Total withdrawals", [
                        {"account": "Unclassified Expenses", "debit": abs(withdrawals), "credit": None},
                        {"account": cash_acct, "debit": None, "credit": abs(withdrawals)},
                    ], "bank_summary")

            # Fees / interest as separate entries (known categories)
            fees = get_val(fields, "fees_charged")
            if fees and not txn_nums:
                _post(stmt_date, f"{bank} \u2014 Service charges", [
                    {"account": "Bank Service Charges", "debit": abs(fees), "credit": None},
                    {"account": cash_acct, "debit": None, "credit": abs(fees)},
                ], "bank_summary")
            interest = get_val(fields, "interest_earned")
            if interest and not txn_nums:
                _post(stmt_date, f"{bank} \u2014 Interest earned", [
                    {"account": cash_acct, "debit": abs(interest), "credit": None},
                    {"account": "Interest Income", "debit": None, "credit": abs(interest)},
                ], "bank_summary")

        # ═══════════════════════════════════════════════════════════════
        # CREDIT CARD STATEMENTS
        # ═══════════════════════════════════════════════════════════════
        # Credit card is a LIABILITY (credit-normal).
        #   Purchase: DR Unclassified Expenses / CR Credit Card Payable
        #   Payment:  DR Credit Card Payable / CR Cash
        #   Interest: DR Interest Expense / CR Credit Card Payable
        elif "credit_card" in dtype:
            stmt_date = get_str(fields, "statement_period_end") or f"12/31/{year}"
            issuer = get_str(fields, "card_issuer") or entity
            acct_num = get_str(fields, "account_number_last4") or ""
            cc_acct = f"Credit Card Payable \u2014 {issuer}" + (
                f" (...{acct_num[-4:]})" if acct_num else "")

            txn_nums = sorted(set(
                int(m.group(1)) for k in fields
                for m in [re.match(r"txn_(\d+)_", k)] if m
            ))

            if txn_nums:
                for n in txn_nums:
                    tdate = get_str(fields, f"txn_{n}_date") or stmt_date
                    tdesc = get_str(fields, f"txn_{n}_desc") or f"CC Transaction #{n}"
                    tamt = get_val(fields, f"txn_{n}_amount")
                    if not tamt:
                        continue
                    amt = round(abs(tamt), 2)
                    cat = _get_cat(fields, f"txn_{n}_amount")
                    if tamt < 0:
                        # Payment / credit / return
                        cr_acct = cat or "Cash"
                        _post(tdate, f"{issuer}: {tdesc}", [
                            {"account": cc_acct, "debit": amt, "credit": None},
                            {"account": cr_acct, "debit": None, "credit": amt},
                        ], "cc_txn")
                    else:
                        # Purchase / charge
                        dr_acct = cat or "Unclassified Expenses"
                        _post(tdate, tdesc, [
                            {"account": dr_acct, "debit": amt, "credit": None},
                            {"account": cc_acct, "debit": None, "credit": amt},
                        ], "cc_txn")
            else:
                purchases = get_val(fields, "purchases")
                if purchases:
                    _post(stmt_date, f"{cc_acct} \u2014 Purchases", [
                        {"account": "Unclassified Expenses", "debit": abs(purchases), "credit": None},
                        {"account": cc_acct, "debit": None, "credit": abs(purchases)},
                    ], "cc_summary")
                payments = get_val(fields, "payments")
                if payments:
                    _post(stmt_date, f"{cc_acct} \u2014 Payment", [
                        {"account": cc_acct, "debit": abs(payments), "credit": None},
                        {"account": "Cash", "debit": None, "credit": abs(payments)},
                    ], "cc_summary")

            interest_chg = get_val(fields, "interest_charged")
            if interest_chg and not txn_nums:
                _post(stmt_date, f"{issuer} \u2014 Finance charges", [
                    {"account": "Interest Expense", "debit": abs(interest_chg), "credit": None},
                    {"account": cc_acct, "debit": None, "credit": abs(interest_chg)},
                ], "cc_summary")

        # ═══════════════════════════════════════════════════════════════
        # CHECKS
        # ═══════════════════════════════════════════════════════════════
        elif dtype == "check":
            check_date = get_str(fields, "check_date") or f"12/31/{year}"
            check_num = get_str(fields, "check_number") or ""
            payee = get_str(fields, "payee") or get_str(fields, "pay_to") or ""
            check_amt = get_val(fields, "check_amount")
            memo = get_str(fields, "memo_line") or ""
            cat = _get_cat(fields, "check_amount")
            desc = f"Check #{check_num}" if check_num else "Check"
            if payee:
                desc += f" to {payee}"
            if memo:
                desc += f" ({memo})"
            if check_amt:
                dr_acct = cat or "Unclassified Expenses"
                _post(check_date, desc, [
                    {"account": dr_acct, "debit": abs(check_amt), "credit": None},
                    {"account": "Cash", "debit": None, "credit": abs(check_amt)},
                ], "check")

        # ═══════════════════════════════════════════════════════════════
        # INVOICES
        # ═══════════════════════════════════════════════════════════════
        # DR Expense / CR Accounts Payable
        elif "invoice" in dtype:
            inv_date = get_str(fields, "invoice_date") or f"12/31/{year}"
            vendor = get_str(fields, "vendor_name") or entity
            inv_num = get_str(fields, "invoice_number") or ""
            total = get_val(fields, "total_amount")
            cat = _get_cat(fields, "total_amount")
            desc_text = get_str(fields, "description") or ""
            desc = f"{vendor}" + (f" Inv #{inv_num}" if inv_num else "")
            if desc_text:
                desc += f" \u2014 {desc_text}"
            if total:
                dr_acct = cat or "Unclassified Expenses"
                _post(inv_date, desc, [
                    {"account": dr_acct, "debit": abs(total), "credit": None},
                    {"account": "Accounts Payable", "debit": None, "credit": abs(total)},
                ], "invoice")

        # ═══════════════════════════════════════════════════════════════
        # RECEIPTS
        # ═══════════════════════════════════════════════════════════════
        # DR Expense (category if known) / CR Cash or Credit Card
        elif "receipt" in dtype:
            r_date = get_str(fields, "receipt_date") or f"12/31/{year}"
            vendor = get_str(fields, "vendor_name") or entity
            total = get_val(fields, "total_amount")
            cat = _get_cat(fields, "total_amount")
            category = get_str(fields, "category") or ""
            pay_method = (get_str(fields, "payment_method") or "").lower()

            # Priority: operator category > AI-extracted category > unclassified
            if cat:
                expense_acct = cat
            else:
                cat_map = {
                    "meals": "Meals & Entertainment",
                    "supplies": "Office Supplies",
                    "travel": "Auto & Travel",
                    "equipment": "Equipment",
                    "utilities": "Utilities",
                    "rent": "Rent",
                    "insurance": "Insurance",
                    "professional_services": "Legal & Professional",
                }
                expense_acct = cat_map.get(category.lower(), "Unclassified Expenses")

            if "credit" in pay_method or "card" in pay_method:
                contra = "Credit Card Payable"
            elif "check" in pay_method:
                contra = "Cash"
            else:
                contra = "Cash"

            if total:
                _post(r_date, vendor, [
                    {"account": expense_acct, "debit": abs(total), "credit": None},
                    {"account": contra, "debit": None, "credit": abs(total)},
                ], "receipt")

        # ═══════════════════════════════════════════════════════════════
        # PAYROLL (Check Stubs)
        # ═══════════════════════════════════════════════════════════════
        # DR Wages Expense (gross)
        # CR Federal WH Payable, State WH Payable, SS Payable, Medicare Payable
        # CR Cash (net pay)
        # If net is missing: compute as gross - sum(known deductions)
        elif "check_stub" in dtype:
            pay_date = get_str(fields, "pay_date") or f"12/31/{year}"
            employer = get_str(fields, "employer_name") or entity
            employee = get_str(fields, "employee_name") or ""
            label = f"Payroll \u2014 {employer}: {employee}" if employee else f"Payroll \u2014 {employer}"

            gross = get_val(fields, "gross_pay")
            if not gross:
                continue

            lines = [{"account": "Wages Expense", "debit": round(abs(gross), 2), "credit": None}]
            known_deductions = 0.0

            for acct_name, field_name in [
                ("Federal WH Payable", "federal_wh"),
                ("State WH Payable", "state_wh"),
                ("Social Security Payable", "social_security"),
                ("Medicare Payable", "medicare"),
            ]:
                val = get_val(fields, field_name)
                if val:
                    amt = round(abs(val), 2)
                    lines.append({"account": acct_name, "debit": None, "credit": amt})
                    known_deductions += amt

            net = get_val(fields, "net_pay")
            if net:
                lines.append({"account": "Cash", "debit": None, "credit": round(abs(net), 2)})
            else:
                # Compute net to force balance
                computed_net = round(abs(gross) - known_deductions, 2)
                if computed_net > 0:
                    lines.append({"account": "Cash (net \u2014 computed)", "debit": None, "credit": computed_net})

            _post(pay_date, label, lines, "payroll")

        # ═══════════════════════════════════════════════════════════════
        # LOAN / MORTGAGE PAYMENTS
        # ═══════════════════════════════════════════════════════════════
        # DR Loan Payable (principal) + DR Interest Expense + DR Escrow
        # CR Cash (total payment)
        elif "loan_statement" in dtype or "mortgage_statement" in dtype:
            pay_date = (get_str(fields, "payment_date")
                        or get_str(fields, "next_due_date")
                        or f"12/31/{year}")
            lender = get_str(fields, "lender") or entity
            principal = get_val(fields, "principal_paid")
            interest = get_val(fields, "interest_paid")
            escrow = get_val(fields, "escrow_paid")

            if principal or interest:
                lines = []
                if principal:
                    lines.append({"account": f"Loan Payable \u2014 {lender}",
                                  "debit": round(abs(principal), 2), "credit": None})
                if interest:
                    lines.append({"account": "Interest Expense",
                                  "debit": round(abs(interest), 2), "credit": None})
                if escrow:
                    lines.append({"account": "Escrow Deposit",
                                  "debit": round(abs(escrow), 2), "credit": None})
                total = round(sum(l["debit"] or 0 for l in lines), 2)
                lines.append({"account": "Cash", "debit": None, "credit": total})
                _post(pay_date, f"Loan payment \u2014 {lender}", lines, "loan")

        # ═══════════════════════════════════════════════════════════════
        # PROFIT & LOSS STATEMENTS (summary-level)
        # ═══════════════════════════════════════════════════════════════
        elif "profit_loss" in dtype:
            period_end = get_str(fields, "period_end") or f"12/31/{year}"
            revenue = get_val(fields, "total_revenue")
            expenses = (get_val(fields, "total_operating_expenses")
                        or get_val(fields, "total_expenses"))
            if revenue:
                _post(period_end, f"Revenue \u2014 {entity}", [
                    {"account": "Accounts Receivable", "debit": abs(revenue), "credit": None},
                    {"account": "Revenue", "debit": None, "credit": abs(revenue)},
                ], "pnl")
            if expenses:
                _post(period_end, f"Operating expenses \u2014 {entity}", [
                    {"account": "Operating Expenses", "debit": abs(expenses), "credit": None},
                    {"account": "Accounts Payable", "debit": None, "credit": abs(expenses)},
                ], "pnl")

        # All other document types (payroll_register, balance_sheet, etc.)
        # are reporting documents, not source transactions. Skip.

    journal.sort(key=lambda e: e.get("date", ""))
    return journal


def _populate_journal_entries(ws, extractions, year):
    """Write proper double-entry journal entries.

    Layout per entry:
      Row 1: Date | Account (debit) | Description | Debit amount |
      Row 2:      |   Account (credit, indented) |             |  | Credit amount
      ...repeat for multi-line entries
      [blank row separator between entries]

    Columns: A=Date, B=Account, C=Description, D=Debit, E=Credit
    Credits are indented with leading spaces in column B.
    """
    row = _write_title(ws, "Journal Entries", year)

    # Column headers
    for col, label in [("A", "Date"), ("B", "Account"), ("C", "Description"), ("D", "Debit"), ("E", "Credit")]:
        cell = ws[f"{col}{row}"]
        cell.value = label
        cell.font = COL_HEADER_FONT
        cell.fill = COL_HEADER_FILL
        cell.border = SECTION_BORDER
        if col in ("D", "E"):
            cell.alignment = Alignment(horizontal="right")
    row += 1

    journal = _build_journal_entries(extractions, year)

    if not journal:
        ws[f"A{row}"] = "(no journal entries generated)"
        ws[f"A{row}"].font = Font(italic=True, color="BBBBBB")
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 40
        for col in ["D", "E"]:
            ws.column_dimensions[col].width = 16
        return

    # Styles for journal entries
    credit_font = Font(size=10, color="444444")
    entry_sep_border = openpyxl.styles.Border(
        bottom=openpyxl.styles.Side(style="thin", color="CCCCCC"))
    entry_num_font = Font(size=8, color="999999")

    data_start = row
    entry_count = 0
    unbalanced_count = 0

    for entry in journal:
        entry_count += 1
        date_str = entry.get("date", "")
        desc = entry.get("description", "")
        lines = entry.get("lines", [])
        balanced = entry.get("balanced", True)
        entry_start_row = row

        if not balanced:
            unbalanced_count += 1

        # Separate debit lines (listed first) and credit lines
        debit_lines = [l for l in lines if l.get("debit")]
        credit_lines = [l for l in lines if l.get("credit")]

        # Write debit lines first
        first_line = True
        for dl in debit_lines:
            if first_line:
                ws[f"A{row}"] = date_str
                ws[f"C{row}"] = desc
                first_line = False
            ws[f"B{row}"] = dl["account"]
            ws[f"B{row}"].font = Font(bold=True, size=10)
            cell_d = ws[f"D{row}"]
            cell_d.value = dl["debit"]
            cell_d.number_format = MONEY_FMT
            cell_d.alignment = Alignment(horizontal="right")
            for c in ["A", "B", "C", "D", "E"]:
                ws[f"{c}{row}"].border = THIN_BORDER
            row += 1

        # Write credit lines (indented)
        for cl in credit_lines:
            if first_line:
                ws[f"A{row}"] = date_str
                ws[f"C{row}"] = desc
                first_line = False
            ws[f"B{row}"] = f"    {cl['account']}"
            ws[f"B{row}"].font = credit_font
            cell_e = ws[f"E{row}"]
            cell_e.value = cl["credit"]
            cell_e.number_format = MONEY_FMT
            cell_e.alignment = Alignment(horizontal="right")
            for c in ["A", "B", "C", "D", "E"]:
                ws[f"{c}{row}"].border = THIN_BORDER
            row += 1

        # Highlight unbalanced entries in orange
        if not balanced:
            for r in range(entry_start_row, row):
                for c in ["A", "B", "C", "D", "E"]:
                    ws[f"{c}{r}"].fill = REVIEW_FILL

        # Blank separator row between entries
        for c in ["A", "B", "C", "D", "E"]:
            ws[f"{c}{row}"].border = entry_sep_border
        row += 1

    # Grand totals
    data_end = row - 2  # skip last separator
    ws[f"B{row}"] = "TOTALS"
    ws[f"B{row}"].font = SUM_FONT
    ws[f"B{row}"].fill = SUM_FILL
    for col in ["D", "E"]:
        cell = ws[f"{col}{row}"]
        cell.value = f"=SUM({col}{data_start}:{col}{data_end})"
        cell.font = SUM_FONT
        cell.fill = SUM_FILL
        cell.number_format = MONEY_FMT
        cell.alignment = Alignment(horizontal="right")
    for c in ["A", "B", "C", "D", "E"]:
        ws[f"{c}{row}"].border = openpyxl.styles.Border(
            top=openpyxl.styles.Side(style="double", color="333333"))
    ws[f"A{row}"].fill = SUM_FILL
    ws[f"C{row}"].fill = SUM_FILL
    row += 1

    # Balance check row
    ws[f"B{row}"] = "BALANCE CHECK (should be zero):"
    ws[f"B{row}"].font = Font(bold=True, size=9, color="666666")
    bal_cell = ws[f"D{row}"]
    bal_cell.value = f"=D{row-1}-E{row-1}"
    bal_cell.number_format = MONEY_FMT
    bal_cell.font = Font(bold=True, size=10, color="CC0000")
    bal_cell.alignment = Alignment(horizontal="right")
    row += 1

    # Entry count
    count_text = f"{entry_count} journal entries"
    if unbalanced_count:
        count_text += f" ({unbalanced_count} UNBALANCED — highlighted orange)"
    ws[f"B{row}"] = count_text
    ws[f"B{row}"].font = Font(italic=True, size=9, color="CC0000" if unbalanced_count else "999999")

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 42
    for col in ["D", "E"]:
        ws.column_dimensions[col].width = 16


# ─── ACCOUNT BALANCES FORMAT ─────────────────────────────────────────────────

ACCT_BAL_TEMPLATE_SECTIONS = [
    {
        "id": "acct_balances",
        "header": "Bank Account Balances:",
        "match_types": ["bank_statement", "bank_statement_deposit_slip"],
        "columns": {
            "A": "bank_name", "B": "account_number_last4",
            "C": "statement_period_start", "D": "statement_period_end",
            "E": "beginning_balance", "F": "total_deposits",
            "G": "total_withdrawals", "H": "fees_charged",
            "I": "interest_earned", "J": "ending_balance",
        },
        "col_headers": {
            "B": "Acct #", "C": "Period Start", "D": "Period End",
            "E": "Begin Bal", "F": "Deposits", "G": "Withdrawals",
            "H": "Fees", "I": "Interest", "J": "End Bal",
        },
        "sum_cols": ["E", "F", "G", "H", "I", "J"],
        "flags": [],
    },
    {
        "id": "credit_card_balances",
        "header": "Credit Card Balances:",
        "match_types": ["credit_card_statement"],
        "columns": {
            "A": "card_issuer", "B": "account_number_last4",
            "C": "statement_period_start", "D": "statement_period_end",
            "E": "previous_balance", "F": "purchases",
            "G": "payments", "H": "interest_charged", "I": "new_balance",
        },
        "col_headers": {
            "B": "Acct #", "C": "Period Start", "D": "Period End",
            "E": "Prev Bal", "F": "Purchases", "G": "Payments",
            "H": "Interest", "I": "New Bal",
        },
        "sum_cols": ["E", "F", "G", "H", "I"],
        "flags": [],
    },
]

ACCT_BAL_ALWAYS_SHOW = ["acct_balances", "credit_card_balances"]

def _populate_account_balances(ws, extractions, year):
    """Write account balance format: one row per bank statement."""
    row = _write_title(ws, "Account Balances", year)

    for section in ACCT_BAL_TEMPLATE_SECTIONS:
        sid = section["id"]
        match_types = section.get("match_types", [])
        matched = _match_exts(extractions, match_types) if match_types else []
        col_headers = section.get("col_headers", {})
        columns = section.get("columns", {})
        sum_cols = section.get("sum_cols", [])

        if not matched and sid not in ACCT_BAL_ALWAYS_SHOW:
            continue

        ws[f"A{row}"] = section["header"]
        ws[f"A{row}"].font = SECTION_FONT
        ws[f"A{row}"].fill = SECTION_FILL
        for col, label in col_headers.items():
            cell = ws[f"{col}{row}"]
            cell.value = label
            cell.font = COL_HEADER_FONT
            cell.fill = COL_HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        # Fill remaining columns in header row with gray
        for c in ["A", "B", "C", "D", "E", "F"]:
            if c not in col_headers:
                ws[f"{c}{row}"].fill = SECTION_FILL
        row += 1

        if not matched:
            ws[f"A{row}"] = "(no bank statements found)"
            ws[f"A{row}"].font = Font(italic=True, color="BBBBBB")
            row += 2
            continue

        matched = _dedup_by_ein(matched)
        data_start = row
        all_cols = list(columns.keys())

        for ext_idx, ext in enumerate(matched):
            fields = ext.get("fields", {})
            for col, field_name in columns.items():
                if field_name in ("bank_name",):
                    ws[f"{col}{row}"] = get_str(fields, field_name) or ext.get("payer_or_entity", "")
                else:
                    val = get_val(fields, field_name)
                    if val is None:
                        val = get_str(fields, field_name)
                    cell = ws[f"{col}{row}"]
                    cell.value = val
                    if isinstance(val, (int, float)):
                        cell.number_format = MONEY_FMT
                        cell.alignment = Alignment(horizontal="right")
            # Alternating row fill
            if ext_idx % 2 == 1:
                for bcol in all_cols:
                    cell = ws[f"{bcol}{row}"]
                    if cell.fill == PatternFill() or cell.fill is None:
                        cell.fill = ALT_ROW_FILL
            for bcol in all_cols:
                ws[f"{bcol}{row}"].border = THIN_BORDER
            row += 1

        data_end = row - 1
        if data_end >= data_start and sum_cols:
            ws[f"A{row}"] = "TOTAL"
            ws[f"A{row}"].font = SUM_FONT
            ws[f"A{row}"].fill = SUM_FILL
            for col in sum_cols:
                cell = ws[f"{col}{row}"]
                cell.value = f"=SUM({col}{data_start}:{col}{data_end})"
                cell.font = SUM_FONT
                cell.fill = SUM_FILL
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right")
                cell.border = SUM_BORDER
            ws[f"A{row}"].border = SUM_BORDER
            row += 1

    # ─── Transaction Detail Section ───
    row += 1
    all_txns = []
    for ext in extractions:
        if ext.get("document_type") not in ("bank_statement", "bank_statement_deposit_slip", "check", "check_stub", "credit_card_statement"):
            continue
        fields = ext.get("fields", {})
        entity = ext.get("payer_or_entity", "")
        # Gather numbered transaction fields: txn_N_date, txn_N_desc, txn_N_amount, txn_N_type
        txn_nums = set()
        for k in fields:
            m = __import__("re").match(r"txn_(\d+)_", k)
            if m:
                txn_nums.add(int(m.group(1)))
        for n in sorted(txn_nums):
            txn_date = get_str(fields, f"txn_{n}_date") or ""
            txn_desc = get_str(fields, f"txn_{n}_desc") or ""
            txn_amt = get_val(fields, f"txn_{n}_amount")
            txn_type = get_str(fields, f"txn_{n}_type") or ""
            all_txns.append((entity, txn_date, txn_desc, txn_amt, txn_type))

    if all_txns:
        ws[f"A{row}"] = "Transaction Detail:"
        ws[f"A{row}"].font = SECTION_FONT
        ws[f"A{row}"].border = SECTION_BORDER
        for col, label in [("B", "Date"), ("C", "Description"), ("D", "Amount"), ("E", "Type")]:
            cell = ws[f"{col}{row}"]
            cell.value = label
            cell.font = COL_HEADER_FONT
            cell.fill = COL_HEADER_FILL
            cell.alignment = Alignment(horizontal="right") if col == "D" else Alignment(horizontal="left")
            cell.border = SECTION_BORDER
        row += 1
        txn_data_start = row
        for txn_idx, (entity, tdate, tdesc, tamt, ttype) in enumerate(all_txns):
            ws[f"A{row}"] = entity
            ws[f"B{row}"] = tdate
            ws[f"C{row}"] = tdesc
            cell = ws[f"D{row}"]
            cell.value = tamt
            if isinstance(tamt, (int, float)):
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right")
            ws[f"E{row}"] = ttype
            if txn_idx % 2 == 1:
                for bcol in ["A", "B", "C", "D", "E"]:
                    cell = ws[f"{bcol}{row}"]
                    if cell.fill == PatternFill() or cell.fill is None:
                        cell.fill = ALT_ROW_FILL
            for bcol in ["A", "B", "C", "D", "E"]:
                ws[f"{bcol}{row}"].border = THIN_BORDER
            row += 1
        txn_data_end = row - 1
        # Total row for transaction amounts
        ws[f"A{row}"] = "TOTAL"
        ws[f"A{row}"].font = SUM_FONT
        ws[f"A{row}"].fill = SUM_FILL
        cell = ws[f"D{row}"]
        cell.value = f"=SUM(D{txn_data_start}:D{txn_data_end})"
        cell.font = SUM_FONT
        cell.fill = SUM_FILL
        cell.number_format = MONEY_FMT
        cell.alignment = Alignment(horizontal="right")
        cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style="thin", color="999999"))
        row += 1

    ws.column_dimensions["A"].width = 30
    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        ws.column_dimensions[col].width = 16


# ─── TRIAL BALANCE FORMAT ───────────────────────────────────────────────────

def _populate_trial_balance(ws, extractions, year):
    """Write a trial balance: accounts with debit and credit totals.

    Builds journal entries first, then aggregates by account into
    a standard trial balance (Account | Debit | Credit) with totals.
    """
    row = _write_title(ws, "Trial Balance", year)

    # Column headers
    for col, label, align in [("A", "Account", "left"), ("B", "Debit", "right"),
                               ("C", "Credit", "right"), ("D", "Net Balance", "right")]:
        cell = ws[f"{col}{row}"]
        cell.value = label
        cell.font = DARK_HEADER_FONT
        cell.fill = DARK_HEADER_FILL
        cell.border = SECTION_BORDER
        cell.alignment = Alignment(horizontal=align)
    row += 1

    # Build journal entries and aggregate by account
    journal = _build_journal_entries(extractions, year)
    account_totals = {}  # account_name → {"debit": float, "credit": float}

    for entry in journal:
        for line in entry.get("lines", []):
            acct = line.get("account", "Uncategorized")
            dr = line.get("debit") or 0
            cr = line.get("credit") or 0
            if acct not in account_totals:
                account_totals[acct] = {"debit": 0.0, "credit": 0.0}
            account_totals[acct]["debit"] += dr
            account_totals[acct]["credit"] += cr

    if not account_totals:
        ws[f"A{row}"] = "(no account data — upload bank statements, invoices, or receipts)"
        ws[f"A{row}"].font = Font(italic=True, color="BBBBBB")
        _tb_col_widths(ws)
        return

    # Sort accounts: Assets/Expenses (debit-normal) first, then Liabilities/Revenue (credit-normal)
    # Simple heuristic: sort by net balance direction, then alphabetically
    def sort_key(item):
        acct, tots = item
        net = tots["debit"] - tots["credit"]
        # Debit-normal accounts first (positive net = debit heavy)
        return (0 if net >= 0 else 1, acct.lower())

    data_start = row
    for idx, (acct, tots) in enumerate(sorted(account_totals.items(), key=sort_key)):
        ws[f"A{row}"] = acct
        dr_cell = ws[f"B{row}"]
        cr_cell = ws[f"C{row}"]
        net_cell = ws[f"D{row}"]

        if tots["debit"] > 0:
            dr_cell.value = round(tots["debit"], 2)
            dr_cell.number_format = MONEY_FMT
            dr_cell.alignment = Alignment(horizontal="right")
        if tots["credit"] > 0:
            cr_cell.value = round(tots["credit"], 2)
            cr_cell.number_format = MONEY_FMT
            cr_cell.alignment = Alignment(horizontal="right")

        net = round(tots["debit"] - tots["credit"], 2)
        net_cell.value = net
        net_cell.number_format = MONEY_FMT
        net_cell.alignment = Alignment(horizontal="right")
        if net < 0:
            net_cell.font = Font(color="CC0000")

        # Alternating rows
        if idx % 2 == 1:
            for c in ["A", "B", "C", "D"]:
                ws[f"{c}{row}"].fill = ALT_ROW_FILL
        for c in ["A", "B", "C", "D"]:
            ws[f"{c}{row}"].border = THIN_BORDER
        row += 1

    data_end = row - 1

    # Totals row
    for c in ["A", "B", "C", "D"]:
        ws[f"{c}{row}"].fill = SUM_FILL
        ws[f"{c}{row}"].font = SUM_FONT
        ws[f"{c}{row}"].border = openpyxl.styles.Border(
            top=openpyxl.styles.Side(style="double", color="333333"))
    ws[f"A{row}"] = "TOTALS"
    for col in ["B", "C", "D"]:
        cell = ws[f"{col}{row}"]
        cell.value = f"=SUM({col}{data_start}:{col}{data_end})"
        cell.number_format = MONEY_FMT
        cell.alignment = Alignment(horizontal="right")
    row += 1

    # Balance check
    ws[f"A{row}"] = "BALANCE CHECK (Debits − Credits, should be zero):"
    ws[f"A{row}"].font = Font(bold=True, size=9, color="666666")
    bal_cell = ws[f"B{row}"]
    bal_cell.value = f"=B{row-1}-C{row-1}"
    bal_cell.number_format = MONEY_FMT
    bal_cell.font = Font(bold=True, size=10, color="CC0000")
    bal_cell.alignment = Alignment(horizontal="right")
    row += 2

    # Account count
    ws[f"A{row}"] = f"{len(account_totals)} accounts from {len(journal)} journal entries"
    ws[f"A{row}"].font = Font(italic=True, size=9, color="999999")

    _tb_col_widths(ws)


def _tb_col_widths(ws):
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18


# ─── TRANSACTION REGISTER FORMAT ────────────────────────────────────────────

def _populate_transaction_register(ws, extractions, year):
    """Write a chronological transaction register with running balance.

    Like a checkbook register: Date | Description | Source | Debit | Credit | Balance
    Every transaction from all documents in one flat, time-ordered list.
    """
    row = _write_title(ws, "Transaction Register", year)

    # Column headers
    headers = [("A", "Date", "left"), ("B", "Description", "left"), ("C", "Source", "left"),
               ("D", "Type", "left"), ("E", "Debit (In)", "right"), ("F", "Credit (Out)", "right"),
               ("G", "Balance", "right")]
    for col, label, align in headers:
        cell = ws[f"{col}{row}"]
        cell.value = label
        cell.font = DARK_HEADER_FONT
        cell.fill = DARK_HEADER_FILL
        cell.border = SECTION_BORDER
        cell.alignment = Alignment(horizontal=align)
    row += 1

    # Gather all transactions from all extractions
    txns = []  # list of (date_str, description, source, type, amount_signed)

    for ext in extractions:
        dtype = str(ext.get("document_type", ""))
        fields = ext.get("fields", {})
        entity = ext.get("payer_or_entity", "")

        # Bank statement / credit card individual transactions
        if "bank_statement" in dtype or "credit_card" in dtype:
            bank = get_str(fields, "bank_name") or get_str(fields, "card_issuer") or entity
            acct = get_str(fields, "account_number_last4") or ""
            source = f"{bank}" + (f" (...{acct[-4:]})" if acct else "")

            txn_nums = sorted(set(int(m.group(1)) for k in fields
                                  for m in [re.match(r"txn_(\d+)_", k)] if m))
            for n in txn_nums:
                tdate = get_str(fields, f"txn_{n}_date") or ""
                tdesc = get_str(fields, f"txn_{n}_desc") or ""
                tamt = get_val(fields, f"txn_{n}_amount")
                ttype = get_str(fields, f"txn_{n}_type") or "transaction"
                if tamt is not None:
                    txns.append((tdate, tdesc, source, ttype, tamt))

            # If no individual txns, add summary lines
            if not txn_nums:
                stmt_date = get_str(fields, "statement_period_end") or f"12/31/{year}"
                deposits = get_val(fields, "total_deposits")
                if deposits:
                    txns.append((stmt_date, f"{source} — Total deposits", source, "deposit", abs(deposits)))
                withdrawals = get_val(fields, "total_withdrawals")
                if withdrawals:
                    txns.append((stmt_date, f"{source} — Total withdrawals", source, "withdrawal", -abs(withdrawals)))
                fees = get_val(fields, "fees_charged")
                if fees:
                    txns.append((stmt_date, f"{source} — Fees", source, "fee", -abs(fees)))
                interest = get_val(fields, "interest_earned")
                if interest:
                    txns.append((stmt_date, f"{source} — Interest", source, "interest", abs(interest)))

        elif dtype == "check":
            check_date = get_str(fields, "check_date") or f"12/31/{year}"
            check_num = get_str(fields, "check_number") or ""
            payee = get_str(fields, "payee") or get_str(fields, "pay_to") or ""
            check_amt = get_val(fields, "check_amount")
            desc = f"Check #{check_num}" if check_num else "Check"
            if payee:
                desc += f" to {payee}"
            if check_amt:
                txns.append((check_date, desc, "Check", "check", -abs(check_amt)))

        elif "invoice" in dtype:
            inv_date = get_str(fields, "invoice_date") or f"12/31/{year}"
            vendor = get_str(fields, "vendor_name") or entity
            inv_num = get_str(fields, "invoice_number") or ""
            total = get_val(fields, "total_amount")
            desc = f"{vendor}" + (f" Inv #{inv_num}" if inv_num else "")
            if total:
                txns.append((inv_date, desc, "Invoice", "expense", -abs(total)))

        elif "receipt" in dtype:
            r_date = get_str(fields, "receipt_date") or f"12/31/{year}"
            vendor = get_str(fields, "vendor_name") or entity
            total = get_val(fields, "total_amount")
            category = get_str(fields, "category") or "expense"
            if total:
                txns.append((r_date, vendor, "Receipt", category, -abs(total)))

        elif "check_stub" in dtype:
            pay_date = get_str(fields, "pay_date") or f"12/31/{year}"
            employer = get_str(fields, "employer_name") or entity
            net = get_val(fields, "net_pay")
            if net:
                txns.append((pay_date, f"Payroll — {employer}", "Payroll", "payroll", abs(net)))

    if not txns:
        ws[f"A{row}"] = "(no transactions found — upload bank statements, invoices, or receipts)"
        ws[f"A{row}"].font = Font(italic=True, color="BBBBBB")
        _tr_col_widths(ws)
        return

    # Sort by date
    def date_sort_key(t):
        d = t[0]
        # Try to parse common date formats for proper sorting
        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y"):
            try:
                return datetime.strptime(d, fmt)
            except (ValueError, TypeError):
                continue
        return datetime(int(year), 12, 31)  # fallback: end of year

    txns.sort(key=date_sort_key)

    # Start running balance from earliest beginning balance (if available)
    # This is fundamental — a register without a starting point is meaningless
    opening_balance = 0.0
    for ext in extractions:
        if "bank_statement" in str(ext.get("document_type", "")):
            begin = get_val(ext.get("fields", {}), "beginning_balance")
            if begin is not None:
                opening_balance = begin
                break  # Use the first one found (earliest in document order)

    data_start = row
    running_balance = opening_balance

    # Write opening balance row if we have one
    if opening_balance != 0.0:
        ws[f"A{row}"] = ""
        ws[f"B{row}"] = "Opening Balance"
        ws[f"B{row}"].font = Font(bold=True, italic=True, size=10)
        ws[f"G{row}"] = round(opening_balance, 2)
        ws[f"G{row}"].number_format = MONEY_FMT
        ws[f"G{row}"].alignment = Alignment(horizontal="right")
        ws[f"G{row}"].font = Font(bold=True)
        for c in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{c}{row}"].fill = SUM_FILL
            ws[f"{c}{row}"].border = THIN_BORDER
        row += 1

    debit_font = Font(color="006600")
    credit_font = Font(color="CC0000")

    for idx, (tdate, tdesc, source, ttype, amount) in enumerate(txns):
        ws[f"A{row}"] = tdate
        ws[f"B{row}"] = tdesc
        ws[f"C{row}"] = source
        ws[f"D{row}"] = ttype.title()

        if amount >= 0:
            # Debit (money in)
            dr_cell = ws[f"E{row}"]
            dr_cell.value = round(amount, 2)
            dr_cell.number_format = MONEY_FMT
            dr_cell.alignment = Alignment(horizontal="right")
            dr_cell.font = debit_font
        else:
            # Credit (money out)
            cr_cell = ws[f"F{row}"]
            cr_cell.value = round(abs(amount), 2)
            cr_cell.number_format = MONEY_FMT
            cr_cell.alignment = Alignment(horizontal="right")
            cr_cell.font = credit_font

        running_balance += amount
        bal_cell = ws[f"G{row}"]
        bal_cell.value = round(running_balance, 2)
        bal_cell.number_format = MONEY_FMT
        bal_cell.alignment = Alignment(horizontal="right")
        if running_balance < 0:
            bal_cell.font = Font(bold=True, color="CC0000")

        # Alternating rows
        if idx % 2 == 1:
            for c in ["A", "B", "C", "D", "E", "F", "G"]:
                cell = ws[f"{c}{row}"]
                if cell.fill == PatternFill():
                    cell.fill = ALT_ROW_FILL
        for c in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{c}{row}"].border = THIN_BORDER
        row += 1

    data_end = row - 1

    # Totals row
    for c in ["A", "B", "C", "D", "E", "F", "G"]:
        ws[f"{c}{row}"].fill = SUM_FILL
        ws[f"{c}{row}"].font = SUM_FONT
        ws[f"{c}{row}"].border = openpyxl.styles.Border(
            top=openpyxl.styles.Side(style="double", color="333333"))
    ws[f"A{row}"] = "TOTALS"
    for col in ["E", "F"]:
        cell = ws[f"{col}{row}"]
        cell.value = f"=SUM({col}{data_start}:{col}{data_end})"
        cell.number_format = MONEY_FMT
        cell.alignment = Alignment(horizontal="right")
    ws[f"G{row}"] = round(running_balance, 2)
    ws[f"G{row}"].number_format = MONEY_FMT
    ws[f"G{row}"].alignment = Alignment(horizontal="right")
    if running_balance < 0:
        ws[f"G{row}"].font = Font(bold=True, color="CC0000")
    row += 2

    # Summary
    ws[f"A{row}"] = f"{len(txns)} transactions"
    ws[f"A{row}"].font = Font(italic=True, size=9, color="999999")

    _tr_col_widths(ws)


def _tr_col_widths(ws):
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16


# ─── LOG + SUMMARY ───────────────────────────────────────────────────────────

def save_log(extractions, classifications, warnings, output_path, output_format="tax_review", user_notes="", ai_instructions="", cost_data=None, text_layer_stats=None, page_preprocessing=None, routing_plan=None, consensus_data=None, sections_by_page=None, timing_data=None, throughput_stats=None, streaming_stats=None, ardent_result=None, ardent_diff=None, ardent_summary=None):
    global _cost_tracker
    log_path = output_path.replace(".xlsx", "_log.json")

    # Build per-page method summary
    per_page_method = {}
    methods_used = set()
    for e in extractions:
        page = e.get("_page")
        method = e.get("_extraction_method", "unknown")
        methods_used.add(method)
        if page is not None:
            per_page_method[str(page)] = {
                "method": method,
                "text_source": e.get("_text_source", "unknown"),
            }

    log = {
        "version": "v6",
        "architecture": "text_layer_ocr_vision",
        "output_format": output_format,
        "user_notes": user_notes,
        "ai_instructions": ai_instructions,
        "timestamp": datetime.now().isoformat(),
        "model": MODEL,
        "extraction_methods_used": sorted(methods_used),
        "per_page_method": per_page_method,
        "classifications": classifications,
        "extractions": [{k: v for k, v in e.items() if not k.startswith("_")} | {
            "_page": e.get("_page"),
            "_extraction_method": e.get("_extraction_method"),
            "_text_source": e.get("_text_source"),
            "_overall_confidence": e.get("_overall_confidence"),
        } for e in extractions],
        "warnings": warnings,
        "human_review_required": REQUIRES_HUMAN_REVIEW,
    }
    if text_layer_stats:
        log["text_layer_stats"] = {
            "text_chars_total": text_layer_stats.get("total_chars", 0),
            "meaningful_pages": text_layer_stats.get("meaningful_pages", 0),
            "total_pages": text_layer_stats.get("total_pages", 0),
            "result": text_layer_stats.get("reason", "unknown"),
        }
    if page_preprocessing:
        blank_pages = [m["page_num"] for m in page_preprocessing if m.get("is_blank")]
        quality_scores = [m["quality_score"] for m in page_preprocessing if not m.get("is_blank")]
        log["preprocessing"] = {
            "total_pages": len(page_preprocessing),
            "blank_pages": blank_pages,
            "blank_count": len(blank_pages),
            "pages_rotated": sum(1 for m in page_preprocessing if m.get("was_rotated")),
            "pages_deskewed": sum(1 for m in page_preprocessing if m.get("deskew_angle", 0) != 0),
            "pages_contrast_enhanced": sum(1 for m in page_preprocessing if m.get("contrast_enhanced")),
            "avg_quality_score": round(sum(quality_scores) / len(quality_scores), 3) if quality_scores else 0,
            "min_quality_score": round(min(quality_scores), 3) if quality_scores else 0,
            "per_page": [{
                "page": m.get("page_num"),
                "is_blank": m.get("is_blank", False),
                "blank_reason": m.get("blank_reason", "not_blank"),
                "quality_score": m.get("quality_score", 0),
                "dpi": m.get("dpi", DPI),
                "original_size": m.get("original_size"),
                "processed_size": m.get("processed_size"),
                "rotated": m.get("was_rotated", False),
                "deskew_angle": m.get("deskew_angle", 0),
                "contrast_enhanced": m.get("contrast_enhanced", False),
            } for m in page_preprocessing],
        }
    if routing_plan:
        method_counts = {}
        for r in routing_plan:
            m = r.get("method", "unknown")
            method_counts[m] = method_counts.get(m, 0) + 1
        log["routing"] = {
            "total_pages": len(routing_plan),
            "text_layer_pages": method_counts.get("text_layer", 0),
            "ocr_pages": method_counts.get("ocr", 0),
            "vision_pages": method_counts.get("vision", 0),
            "blank_pages": method_counts.get("skip_blank", 0),
            "per_page": routing_plan,
        }
    if sections_by_page:
        label_counts = {}
        for labels in sections_by_page.values():
            for lbl in labels:
                label_counts[lbl] = label_counts.get(lbl, 0) + 1
        log["sections"] = {
            "total_pages_labeled": len(sections_by_page),
            "label_counts": label_counts,
            "per_page": sections_by_page,
        }
    if consensus_data:
        log["consensus"] = consensus_data
    if timing_data:
        log["timing"] = timing_data
    if throughput_stats:
        log["throughput"] = throughput_stats
    if streaming_stats:
        log["streaming"] = streaming_stats
    if cost_data:
        log["cost"] = cost_data
    elif _cost_tracker:
        log["cost"] = _cost_tracker.to_dict()
    # ── Lite: Ardent result + diff artifact ──
    if ardent_result is not None:
        try:
            log["ardent_result"] = ardent_result.model_dump(mode="json")
        except Exception:
            log["ardent_result"] = str(ardent_result)
    if ardent_diff is not None:
        log["ardent_diff"] = ardent_diff
    if ardent_summary is not None:
        try:
            log["ardent_summary"] = ardent_summary.model_dump(mode="json")
        except Exception:
            log["ardent_summary"] = str(ardent_summary)
    with open(log_path, "w") as f:
        json.dump(log, f, indent=2, default=str)
    print(f"  Log: {log_path}")


def save_checkpoint(output_path, phase, classifications=None, extractions=None, groups=None):
    """Save partial results after each phase for crash recovery.

    Checkpoint file is {output}_checkpoint.json. If the run completes
    successfully, the checkpoint is deleted. If it crashes, the operator
    can resume from the last completed phase with --resume.
    """
    ckpt_path = output_path.replace(".xlsx", "_checkpoint.json")
    data = {
        "timestamp": datetime.now().isoformat(),
        "completed_phase": phase,
        "model": MODEL,
    }
    if classifications is not None:
        data["classifications"] = classifications
    if groups is not None:
        data["groups"] = [{k: v for k, v in g.items()} for g in groups]
    if extractions is not None:
        data["extractions"] = [{k: v for k, v in e.items() if not k.startswith("_") or k in (
            "_page", "_extraction_method", "_text_source", "_overall_confidence", "_is_brokerage", "_ambiguous_fields"
        )} for e in extractions]
    with open(ckpt_path, "w") as f:
        json.dump(data, f, indent=2, default=str)
    print(f"  Checkpoint saved: {phase}")
    return ckpt_path


def load_checkpoint(output_path):
    """Load checkpoint if it exists. Returns dict or None."""
    ckpt_path = output_path.replace(".xlsx", "_checkpoint.json")
    if os.path.exists(ckpt_path):
        try:
            with open(ckpt_path) as f:
                data = json.load(f)
            print(f"  Found checkpoint: completed through {data.get('completed_phase', '?')}")
            return data
        except (json.JSONDecodeError, IOError) as e:
            print(f"  Checkpoint corrupt: {e} — starting fresh")
    return None


def clear_checkpoint(output_path):
    """Remove checkpoint file after successful completion."""
    ckpt_path = output_path.replace(".xlsx", "_checkpoint.json")
    if os.path.exists(ckpt_path):
        os.remove(ckpt_path)


# ─── T1.5: PARTIAL RESULTS (for mid-run review) ─────────────────────────────

def write_partial_results(output_path, extractions, batch_num, total_batches,
                          time_to_first_values_s=None, sections_by_page=None):
    """Write accumulated extraction results for mid-run review.

    File: {stem}_partial_results.json
    Written atomically (tmp file + os.replace) after each batch.
    app.py reads this file to serve partial results during extraction.
    """
    partial_path = output_path.replace(".xlsx", "_partial_results.json")
    tmp_path = partial_path + ".tmp"
    try:
        # Count total fields across all extractions
        total_fields = 0
        for ext in extractions:
            for fname, fdata in (ext.get("fields") or {}).items():
                val = fdata.get("value") if isinstance(fdata, dict) else fdata
                if val is not None:
                    total_fields += 1

        data = {
            "version": "v6",
            "partial": True,
            "batch_num": batch_num,
            "total_batches": total_batches,
            "timestamp": datetime.now().isoformat(),
            "time_to_first_values_s": time_to_first_values_s,
            "fields_count": total_fields,
            "extractions": [{k: v for k, v in e.items()
                             if not k.startswith("_") or k in (
                                 "_page", "_extraction_method", "_text_source",
                                 "_overall_confidence", "_is_brokerage",
                                 "_ambiguous_fields", "_batch",
                             )} for e in extractions],
        }
        if sections_by_page:
            data["sections_by_page"] = sections_by_page

        with open(tmp_path, "w") as f:
            json.dump(data, f, indent=2, default=str)
        os.replace(tmp_path, partial_path)
    except (IOError, OSError) as e:
        print(f"  Partial results write failed (non-fatal): {e}")
        # Clean up tmp file if it exists
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass


def clear_partial_results(output_path):
    """Remove partial results file after successful completion."""
    partial_path = output_path.replace(".xlsx", "_partial_results.json")
    if os.path.exists(partial_path):
        os.remove(partial_path)


def print_summary(extractions):
    print("\n── [PASSION] Summary ──")
    methods = {}
    confs = {}
    for ext in extractions:
        m = ext.get("_extraction_method", "unknown")
        methods[m] = methods.get(m, 0) + 1
        for fdata in ext.get("fields", {}).values():
            if isinstance(fdata, dict):
                c = fdata.get("confidence", "unknown")
                confs[c] = confs.get(c, 0) + 1

    total_f = sum(confs.values())
    print(f"  Extraction methods: {dict(methods)}")
    print(f"  Total fields: {total_f}")
    for c, n in sorted(confs.items(), key=lambda x: -x[1]):
        print(f"    {c}: {n} ({n/total_f*100:.0f}%)")


# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Document Intake Extractor v6")
    parser.add_argument("pdf", nargs="?", default=None, help="Scanned PDF path")
    parser.add_argument("--year", type=int, default=2024)
    parser.add_argument("--template", default=None)
    parser.add_argument("--output", default=None)
    parser.add_argument("--dpi", type=int, default=DPI)
    parser.add_argument("--skip-verify", action="store_true", help="Skip verification (faster, cheaper)")
    parser.add_argument("--log-only", action="store_true")
    parser.add_argument("--no-pii", action="store_true", help="Disable PII tokenization (send raw data)")
    parser.add_argument("--regen-excel", action="store_true",
                        help="Regenerate Excel from a log JSON (no extraction). Requires --log-input and --output.")
    parser.add_argument("--log-input", default=None,
                        help="Path to extraction log JSON (used with --regen-excel)")
    parser.add_argument("--doc-type", default="tax_returns",
                        choices=["tax_returns", "bank_statements", "trust_documents", "bookkeeping"],
                        help="Document type category (affects classification and extraction)")
    parser.add_argument("--output-format", default="tax_review",
                        choices=["tax_review", "journal_entries", "account_balances", "trial_balance", "transaction_register"],
                        help="Output Excel format")
    parser.add_argument("--user-notes", default="",
                        help="Operator notes providing additional context for extraction")
    parser.add_argument("--ai-instructions", default="",
                        help="Direct instructions to the AI for classification and extraction")
    parser.add_argument("--context-file", default=None,
                        help="Path to client context index JSON (prior-year data, payer info)")
    parser.add_argument("--resume", action="store_true",
                        help="Resume from checkpoint if available (crash recovery)")
    parser.add_argument("--no-ocr-first", action="store_true",
                        help="Disable OCR-first optimization (force vision-only, more expensive)")
    parser.add_argument("--no-cache", action="store_true",
                        help="Skip page cache (force full reprocessing)")
    args = parser.parse_args()

    # ─── Regen-only mode: read log JSON → populate_template → done ───
    if args.regen_excel:
        if not args.log_input or not args.output:
            sys.exit("ERROR: --regen-excel requires --log-input <log.json> and --output <file.xlsx>")
        if not os.path.exists(args.log_input):
            sys.exit(f"ERROR: Log file not found: {args.log_input}")
        with open(args.log_input) as f:
            log_data = json.load(f)
        extractions = log_data.get("extractions", [])
        if not extractions:
            sys.exit("ERROR: No extractions found in log file")
        # Normalize brokerage data (in case corrections changed composite fields)
        extractions = normalize_brokerage_data(extractions)
        fmt = log_data.get("output_format", args.output_format)
        populate_template(extractions, args.template, args.output, args.year, output_format=fmt)
        print(f"  Regenerated: {args.output}")
        return

    # ─── Normal extraction mode ───
    if not args.pdf:
        sys.exit("ERROR: PDF path is required (unless using --regen-excel)")
    if not os.path.exists(args.pdf):
        sys.exit(f"ERROR: File not found: {args.pdf}")
    if not os.environ.get("ANTHROPIC_API_KEY"):
        sys.exit("ERROR: export ANTHROPIC_API_KEY=sk-ant-...")

    output = args.output or args.pdf.replace(".pdf", "_intake.xlsx").replace(".PDF", "_intake.xlsx")

    # Initialize PII tokenizer
    tokenizer = None if args.no_pii else PIITokenizer()

    # Initialize cost tracker + pipeline timer
    global _cost_tracker, _pipeline_timer
    _cost_tracker = CostTracker()
    _pipeline_timer = PipelineTimer()

    print("=" * 60)
    print("  Document Intake Extractor v6")
    print("  Text-layer | OCR-first | Vision fallback | Checkpointing")
    print("=" * 60)
    print(f"  PDF:      {args.pdf}")
    print(f"  Year:     {args.year}")
    print(f"  Output:   {output}")
    print(f"  PyMuPDF:  {'YES (text layer extraction)' if HAS_PYMUPDF else 'NOT INSTALLED'}")
    print(f"  OCR-first:{'YES (cheap text when readable)' if not args.no_ocr_first else 'DISABLED (vision-only)'}")
    print(f"  PII:      {'TOKENIZED (Tesseract for SSN detection)' if tokenizer and HAS_TESSERACT else 'TOKENIZED (no Tesseract — text only)' if tokenizer else 'DISABLED (raw data sent)'}")
    print(f"  Verify:   {'YES' if not args.skip_verify else 'SKIPPED'}")
    print(f"  Cache:    {'ON' if not args.no_cache else 'DISABLED'}")

    client = anthropic.Anthropic()

    # ─── Page cache check (T1.4) ───
    cache_hit = False
    page_texts = None
    text_layer_stats = {}
    b64_images = None
    page_preprocessing = []
    ocr_texts = None
    ocr_confidences = None
    routing_plan = []
    blank_pages = set()
    ocr_skipped_tl = 0

    if not args.no_cache:
        _pipeline_timer.start("cache_check")
        cached = _load_cache(args.pdf, args.dpi)
        if cached:
            print(f"\n  Cache HIT — skipping preprocessing, OCR, routing")
            page_texts = cached["page_texts"]
            text_layer_stats = cached["text_layer_stats"]
            b64_images = cached["b64_images"]
            page_preprocessing = cached["page_preprocessing"]
            ocr_texts = cached["ocr_texts"]
            ocr_confidences = cached["ocr_confidences"]
            routing_plan = cached["routing_plan"]
            blank_pages = {m["page_num"] for m in page_preprocessing if m.get("is_blank")}
            cache_hit = True
            # Print routing summary from cache (so app.py progress parsing works)
            _print_routing_summary(routing_plan)
        _pipeline_timer.stop()

    if not cache_hit:
        # ─── Phase 0a: Text-layer extraction (PyMuPDF — instant, no OCR) ───
        _pipeline_timer.start("text_layer")
        # Always extract text layer per page (never discard — routing decides per-page)
        if HAS_PYMUPDF:
            print("\n── [PASSION] Text-Layer Extraction (PyMuPDF) ──")
            page_texts = extract_text_per_page(args.pdf)
            if page_texts:
                _tl_usable, text_layer_stats = has_meaningful_text(page_texts)
                mp = text_layer_stats.get("meaningful_pages", 0)
                tp = text_layer_stats.get("total_pages", 0)
                tc = text_layer_stats.get("total_chars", 0)
                print(f"  Text layer: {mp}/{tp} pages have ≥{TEXT_MIN_CHARS_PER_PAGE} chars, {tc:,} total chars")
            else:
                print(f"  ✗ No text layer found")
        else:
            print("\n  PyMuPDF not installed — no text layer available")

        # ─── Phase 0b: PDF → images + preprocessing ───
        _pipeline_timer.start("images_preprocess")
        b64_images, page_preprocessing = pdf_to_images(args.pdf, args.dpi, page_texts=page_texts)
        blank_pages = {m["page_num"] for m in page_preprocessing if m.get("is_blank")}
        if blank_pages:
            print(f"  Blank pages detected: {sorted(blank_pages)} — will skip OCR/vision")

        # ─── Phase 0c: OCR — lazy per-page skip (T1.4) ───
        _pipeline_timer.start("ocr")
        if not args.no_ocr_first and HAS_TESSERACT:
            n_total = len(b64_images)
            n_non_blank = n_total - len(blank_pages)

            # Count pages with good text-layer coverage
            tl_good_pages = set()
            if page_texts:
                for i, pt in enumerate(page_texts):
                    if pt and len(pt.strip()) >= ROUTE_TEXT_MIN_CHARS and (i + 1) not in blank_pages:
                        tl_good_pages.add(i + 1)

            # Build PIL images for OCR — skip blank AND per-page text-layer-good pages
            # (T1.4: each page decides independently, no global threshold)
            pil_for_ocr = []
            ocr_skipped_tl = 0
            for i, b in enumerate(b64_images):
                pnum = i + 1
                if pnum in blank_pages:
                    pil_for_ocr.append(None)  # skip blank
                elif pnum in tl_good_pages:
                    pil_for_ocr.append(None)  # skip — this page has good text layer
                    ocr_skipped_tl += 1
                else:
                    pil_for_ocr.append(Image.open(BytesIO(base64.b64decode(b))))

            if ocr_skipped_tl > 0:
                pages_to_ocr = n_non_blank - ocr_skipped_tl
                print(f"\n  Text-layer fast-path: {ocr_skipped_tl}/{n_non_blank} pages have good text layer — OCR skipped")
                if pages_to_ocr > 0:
                    print(f"  OCR running on {pages_to_ocr} remaining pages")

            ocr_texts, ocr_confidences = ocr_all_pages(pil_for_ocr)
        elif args.no_ocr_first:
            print("\n  OCR-first: DISABLED (--no-ocr-first, all pages use vision)")
        else:
            print("\n  OCR: unavailable (Tesseract not installed)")

        # ─── Phase 0d: Per-page routing ───
        _pipeline_timer.start("routing")
        routing_plan = route_pages(page_texts, ocr_texts, ocr_confidences, page_preprocessing)

        # ─── Save cache for future runs ───
        if not args.no_cache:
            _save_cache(args.pdf, args.dpi, page_texts, text_layer_stats,
                        b64_images, page_preprocessing, ocr_texts, ocr_confidences, routing_plan)

    # Load prior-year context if provided
    context_summary = ""
    context_data = None
    if args.context_file and os.path.exists(args.context_file):
        try:
            with open(args.context_file) as cf:
                context_data = json.load(cf)
            prior_docs = context_data.get("prior_year_data", {}).get("documents", [])
            if prior_docs:
                payer_lines = []
                for doc in prior_docs:
                    for p in doc.get("payers", []):
                        name = p.get("name", "Unknown")
                        ein = p.get("ein", "")
                        ftype = p.get("form_type", "")
                        amounts = p.get("amounts", [])
                        amt_str = ", ".join(f"${a:,.2f}" for a in amounts[:3]) if amounts else "amounts unknown"
                        payer_lines.append(f"  {ftype or 'Unknown form'}: {name} (EIN {ein}) — {amt_str}")
                if payer_lines:
                    context_summary = (
                        "PRIOR-YEAR REFERENCE DATA (use for sanity-checking, NOT for values):\n"
                        + "\n".join(payer_lines[:30])  # cap at 30 payers
                    )
                    print(f"  Context: {len(payer_lines)} prior-year payers loaded")
        except (json.JSONDecodeError, IOError) as e:
            print(f"  Warning: Could not load context file: {e}")

    # Merge context into ai_instructions
    effective_instructions = args.ai_instructions
    if context_summary:
        effective_instructions = (effective_instructions + "\n\n" + context_summary
                                  if effective_instructions else context_summary)

    # ─── Check for checkpoint (crash recovery) ───
    checkpoint = load_checkpoint(output) if args.resume else None
    resume_phase = checkpoint.get("completed_phase", "") if checkpoint else ""

    if resume_phase:
        print(f"\n  ⟳ Resuming from checkpoint (completed: {resume_phase})")

    # Phase 1: Classify (vision — need to see layout)
    _pipeline_timer.start("classify")
    if resume_phase in ("classify", "group", "extract", "verify"):
        classifications = checkpoint.get("classifications", [])
        print(f"\n── [PASSION] Phase 1: Classification (restored {len(classifications)} from checkpoint) ──")
    else:
        classifications = classify_pages(client, b64_images, tokenizer=tokenizer, doc_type=args.doc_type, user_notes=args.user_notes, ai_instructions=effective_instructions)
        save_checkpoint(output, "classify", classifications=classifications)

    # Phase 1.3: Section / form detection (keyword-based, no API calls)
    _pipeline_timer.start("sections")
    sections_by_page = detect_sections(page_texts, routing_plan, ocr_texts=ocr_texts)

    # Phase 1.5: Group
    if resume_phase in ("group", "extract", "verify"):
        groups = checkpoint.get("groups", [])
        print(f"\n  {len(groups)} document groups (from checkpoint)")
    else:
        groups = group_pages(classifications)
        save_checkpoint(output, "group", classifications=classifications, groups=groups)

    print(f"\n  {len(groups)} document groups:")
    for g in groups:
        cont = g.get("continuation_pages", [])
        brok = " [BROKERAGE]" if g.get("is_consolidated_brokerage") else ""
        print(f"    {g['document_type']}: {g['payer_or_entity']} (EIN: {g.get('payer_ein','?')}) pp. {g['pages']}{f' + {cont}' if cont else ''}{brok}")

    # Phase 2: Extract (OCR-first with vision fallback)
    _pipeline_timer.start("extract")
    streaming_meta = {"time_to_first_values_s": None, "batches_processed": 0, "fields_streamed": 0}
    if resume_phase in ("extract", "verify"):
        extractions = checkpoint.get("extractions", [])
        print(f"\n── [PASSION] Phase 2: Extraction (restored {len(extractions)} from checkpoint) ──")
    else:
        extractions, streaming_meta = extract_data(client, b64_images, groups, tokenizer=tokenizer,
                                    doc_type=args.doc_type, user_notes=args.user_notes,
                                    ai_instructions=effective_instructions, ocr_texts=ocr_texts,
                                    page_texts=page_texts, routing_plan=routing_plan,
                                    sections_by_page=sections_by_page, output_path=output)
        save_checkpoint(output, "extract", classifications=classifications, groups=groups, extractions=extractions)

    # Phase 2.5: Consensus verification (brokerage + K-1 only)
    _pipeline_timer.start("consensus")
    consensus_data = None
    if not args.skip_verify and resume_phase not in ("verify",):
        extractions, consensus_data = build_consensus(
            extractions, page_texts, ocr_texts, ocr_confidences=ocr_confidences)

    # Phase 3: Verify (cross-check critical fields against image)
    # Fields already auto_verified by consensus are skipped
    _pipeline_timer.start("verify")
    verify_stats = {"pages_verified": 0, "pages_skipped": 0}
    if resume_phase == "verify":
        print(f"\n── [PASSION] Phase 3: Verification (already done in checkpoint) ──")
        verify_stats = {"pages_verified": 0, "pages_skipped": len(extractions)}
    elif not args.skip_verify:
        extractions, verify_stats = verify_extractions(client, b64_images, extractions, tokenizer=tokenizer)
        save_checkpoint(output, "verify", classifications=classifications, groups=groups, extractions=extractions)
    else:
        verify_stats = {"pages_verified": 0, "pages_skipped": len(extractions)}

    # Phases 4-6: Normalize, Validate, Output
    _pipeline_timer.start("normalize_validate_export")

    # Phase 4: Normalize
    extractions = normalize_brokerage_data(extractions)

    # Phase 5: Validate
    warnings = validate(extractions, prior_year_context=context_data)

    # ─── [ARDENT] Shadow evaluation (feature-flagged) ────────────────────
    _ardent_result = None
    _ardent_diff = None
    _ardent_summary = None
    if os.environ.get("LITE_ARDENT_ENABLED"):
        try:
            from lite.adapters.oathledger import (
                extractions_to_candidates, diff_ardent_vs_warnings,
            )
            from lite.ardent.engine import evaluate as ardent_evaluate
            from lite.ardent.summary import build_ardent_summary
            from lite.lens import Lens

            _ardent_candidates = extractions_to_candidates(
                extractions,
                job_id=os.path.basename(args.pdf or ""),
                client_id="",
            )
            _ardent_context = Lens.build_bundle_from_files(
                context_file=args.context_file,
                tax_year=str(args.year),
            )
            _ardent_result = ardent_evaluate(_ardent_candidates, _ardent_context)
            _ardent_diff = diff_ardent_vs_warnings(_ardent_result, warnings)

            # Build UI-facing summary
            _ardent_summary = build_ardent_summary(
                _ardent_result,
                evaluated_at_iso=_ardent_result.evaluated_at.isoformat() if _ardent_result.evaluated_at else "",
                deterministic_match_pct=_ardent_diff.get("deterministic_match_pct") if _ardent_diff else None,
            )

            print(f"\n── [ARDENT] Shadow evaluation ──")
            print(f"  Ruleset: {_ardent_result.ruleset_id} "
                  f"({_ardent_result.ruleset_hash[:12]})")
            print(f"  Rules: {_ardent_result.total_rules_evaluated} evaluated, "
                  f"{_ardent_result.rules_passed} passed, "
                  f"{_ardent_result.rules_failed} failed")
            if _ardent_result.evaluation_duration_ms:
                print(f"  Duration: {_ardent_result.evaluation_duration_ms:.1f}ms")
            if _ardent_diff:
                n_match = len(_ardent_diff.get("matches", []))
                n_disc = len(_ardent_diff.get("discrepancies", []))
                print(f"  Diff: {n_match} matches, {n_disc} discrepancies")
            if _ardent_summary:
                _status = "BLOCKED" if _ardent_summary.blocked else ("REVIEW" if _ardent_summary.needs_review else "OK")
                print(f"  Summary: {_status} ({len(_ardent_summary.findings)} findings)")
        except Exception as _ardent_err:
            print(f"\n── [ARDENT] Shadow evaluation failed "
                  f"(non-fatal): {_ardent_err} ──")
    # ─── End ARDENT shadow ───────────────────────────────────────────────

    # Summary
    print_summary(extractions)

    # Cost summary
    print("\n── [PASSION] Cost ──")
    print(_cost_tracker.summary())

    # PII tokenization summary
    if tokenizer:
        stats = tokenizer.get_stats()
        if stats["ssns_tokenized"] > 0:
            print(f"\n  🔒 PII: {stats['ssns_tokenized']} SSN(s) tokenized before API calls")
        else:
            print(f"\n  🔒 PII: tokenizer active, no SSN patterns detected")

    _pipeline_timer.stop()

    # Throughput stats for logging
    n_total = len(b64_images) if b64_images else 0
    n_blank = len(blank_pages)
    n_ocr_run = n_total - n_blank - ocr_skipped_tl if not cache_hit else 0
    throughput = {
        "pages_total": n_total,
        "pages_blank": n_blank,
        "pages_ocr": max(n_ocr_run, 0),
        "pages_ocr_skipped": ocr_skipped_tl,
        "pages_vision_verified": verify_stats.get("pages_verified", 0),
        "pages_vision_skipped": verify_stats.get("pages_skipped", 0),
        "cache_hit": cache_hit,
    }

    # Save
    save_log(extractions, classifications, warnings, output,
             output_format=args.output_format, user_notes=args.user_notes,
             ai_instructions=effective_instructions, cost_data=_cost_tracker.to_dict(),
             text_layer_stats=text_layer_stats if text_layer_stats else None,
             page_preprocessing=page_preprocessing,
             routing_plan=routing_plan, consensus_data=consensus_data,
             sections_by_page=sections_by_page,
             timing_data=_pipeline_timer.to_dict(),
             throughput_stats=throughput,
             streaming_stats=streaming_meta,
             ardent_result=_ardent_result,
             ardent_diff=_ardent_diff,
             ardent_summary=_ardent_summary)
    if not args.log_only:
        populate_template(extractions, args.template, output, args.year, output_format=args.output_format)

    # Clean up checkpoint + partial results on success
    clear_checkpoint(output)
    clear_partial_results(output)

    # Timing summary
    print("\n── [PASSION] Timing ──")
    print(_pipeline_timer.summary())

    print("\n" + "=" * 60)
    print("  COMPLETE")
    tl_pages = sum(1 for e in extractions if e.get("_text_source") == "text_layer")
    if tl_pages > 0:
        print(f"  📄 Text layer: {tl_pages}/{len(extractions)} pages (OCR skipped)")
    if warnings:
        print(f"  ⚠ {len(warnings)} items flagged")
    print(f"  ⚠ {len(REQUIRES_HUMAN_REVIEW)} PY/judgment items need manual entry")
    print(f"  💰 Est. cost: ${_cost_tracker.total_cost():.4f}")
    print("=" * 60)

if __name__ == "__main__":
    main()
