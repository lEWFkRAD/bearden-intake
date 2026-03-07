"""
Renderer that writes to the same Tax Review worksheet layout as inkspren._populate_tax_review,
but driven by an explicit payload (see oathledger.rules_engine.build_tax_review_payload).

This allows:
- deterministic payload generation (rules)
- deterministic Excel rendering (presentation)
- swap templates later without rewriting rules

Now uses workpaper_styles (Pi for Excel–inspired) for composable named styles
instead of reaching into inkspren module constants.
"""

from __future__ import annotations

from typing import Any, Dict, List, Optional

from openpyxl.comments import Comment

# Shared style system — single source of truth for all renderers
from workpaper_styles import (
    apply_styles,
    write_section_header,
    write_total_row,
    write_flag_rows,
    write_title_block,
    set_standard_widths,
    MONEY_FMT,
    BORDERS,
    COLORS,
    NAMED_STYLES,
)


def _write_cell_from_payload(ws, col: str, row: int, field_name: str, value: Any, meta: Dict[str, Any] | None):
    """Write a single cell with value, formatting, and optional confidence metadata."""
    cell = ws[f"{col}{row}"]
    cell.value = value

    # Numeric formatting using named style
    if isinstance(value, (int, float)):
        apply_styles(cell, "currency")

    if meta and meta.get("comment"):
        cell.comment = Comment(str(meta["comment"]), "System")
        # Apply review fill for fields needing preparer judgment
        if meta.get("requires_preparer_judgment") or meta.get("requires_prior_year_data"):
            apply_styles(cell, "review")


def populate_tax_review_from_payload(ws, payload: Dict[str, Any], year: int):
    """
    Writes the Tax Review worksheet using shared workpaper_styles.

    Uses composable named styles (Pi for Excel pattern) instead of raw
    openpyxl objects for consistent, maintainable formatting.
    """
    client_name = payload.get("client_name", "") or ""
    row = write_title_block(ws, "Document Intake", year, client_name=client_name)

    # Set standard column widths (A=35 chars, B-F=15 chars)
    set_standard_widths(ws, num_cols=6)

    sections: List[Dict[str, Any]] = payload.get("sections", []) or []

    for section in sections:
        sid = section.get("id")

        # Special k1_detail block (mirrors inkspren special section)
        if section.get("special") == "k1_extras":
            rows = section.get("rows", []) or []
            if not rows:
                continue

            k1_headers = {"B": "Line Ref", "C": "Description", "D": "Amount"}
            row = write_section_header(ws, row, section.get("header", ""), k1_headers)

            for item in rows:
                ws[f"A{row}"].value = item.get("entity", "")
                ws[f"B{row}"].value = item.get("line_reference", "")
                ws[f"C{row}"].value = item.get("description", "")
                amt_cell = ws[f"D{row}"]
                amt_cell.value = item.get("amount")
                if isinstance(amt_cell.value, (int, float)):
                    apply_styles(amt_cell, "currency")
                row += 1

            row += 1  # blank spacer
            continue

        rows = section.get("rows", []) or []
        if not rows:
            continue

        columns = section.get("columns", {}) or {}
        col_headers = section.get("col_headers", {}) or {}
        sum_cols = section.get("sum_cols", []) or []

        # Section header row using shared helper
        row = write_section_header(ws, row, section.get("header", ""), col_headers)

        data_start = row

        for r in rows:
            f = (r.get("fields", {}) or {})
            fmeta = ((r.get("meta", {}) or {}).get("field_meta", {}) or {})

            for col, field_name in columns.items():
                _write_cell_from_payload(ws, col, row, field_name, f.get(field_name), fmeta.get(field_name))
            row += 1

        data_end = row - 1

        # Totals using shared helper
        if data_end >= data_start and sum_cols:
            row = write_total_row(
                ws, row, data_start, data_end, sum_cols,
                total_formula_col=section.get("total_formula_col"),
            )

        # Flags using shared helper
        flags = section.get("flags") or []
        if flags:
            row = write_flag_rows(ws, row, flags)

        row += 1  # blank spacer between sections

    # NOTE: inkspren._populate_tax_review continues with Schedule A derived blocks etc.
    # This renderer intentionally stops after TEMPLATE_SECTIONS + k1_detail.
    # Keep the existing Schedule A logic in inkspren for now (lowest risk), or migrate it later
    # into payload-driven rendering once you decide the contract you want for those blocks.
