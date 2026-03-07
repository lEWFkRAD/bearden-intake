#!/usr/bin/env python3
# ============================================================
# URITHIRU — Central Command of OathLedger
# ============================================================
"""
Bearden Document Intake Platform v5.1
==================================
Full-featured local web app wrapping extract.py.

Features:
  - Drag-and-drop PDF upload
  - Live extraction progress with console output
  - Side-by-side review: source PDF page ↔ extracted values
  - Job history with client name search
  - Excel + JSON log download
  - Audit trail generation

Run:
    python3 app.py

Open:
    http://localhost:5000
"""

import os
import sys
import json
import threading
import collections
import atexit
import uuid
import re
import glob
import subprocess
import sqlite3
from pathlib import Path
from datetime import datetime, timedelta
from io import BytesIO

# ── Load .env file for persistent secrets ────────────────────────────
def _load_env_file(base: Path = None):
    """Read .env from project root and inject into os.environ (skip if missing)."""
    env_path = (base or Path(__file__).resolve().parent) / ".env"
    if not env_path.exists():
        return
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, val = line.partition("=")
            key = key.strip()
            val = val.strip().strip("\"'")
            if key and val and not os.environ.get(key):
                os.environ[key] = val

_load_env_file()

try:
    from flask import (Flask, render_template_string, render_template, request,
                       jsonify, send_file, abort, session, redirect)
except ImportError:
    sys.exit("Install Flask: pip3 install flask")

try:
    from werkzeug.security import generate_password_hash, check_password_hash
except ImportError:
    sys.exit("Install Werkzeug: pip3 install werkzeug")

try:
    from pdf2image import convert_from_path
except ImportError:
    sys.exit("Install pdf2image: pip3 install pdf2image")

try:
    from PIL import Image
except ImportError:
    sys.exit("Install Pillow: pip3 install Pillow")

# Auto-detect poppler on Windows (pdf2image needs it for PDF -> image conversion)
_POPPLER_PATH = None
if sys.platform == "win32":
    import glob as _glob
    for _p in (_glob.glob(r"C:\tools\poppler*\Library\bin") +
               _glob.glob(r"C:\tools\poppler*\bin") +
               _glob.glob(r"C:\Program Files\poppler*\Library\bin") +
               _glob.glob(r"C:\Program Files\poppler*\bin")):
        if os.path.isfile(os.path.join(_p, "pdftoppm.exe")):
            _POPPLER_PATH = _p
            break

# Auto-detect Tesseract (needed for page-word OCR / highlighting)
# Priority: TESSERACT_CMD env var > PATH > common install locations
_tesseract_found = False
if os.environ.get("TESSERACT_CMD"):
    try:
        import pytesseract as _pt
        _pt.pytesseract.tesseract_cmd = os.environ["TESSERACT_CMD"]
        _tesseract_found = True
    except ImportError:
        pass
if not _tesseract_found:
    # Check if tesseract is already in PATH
    import shutil
    if shutil.which("tesseract"):
        _tesseract_found = True  # pytesseract will find it automatically
    else:
        # Search common install locations (Windows + Mac + Linux)
        _tess_search = []
        if sys.platform == "win32":
            import glob as _tg
            _tess_search = (
                _tg.glob(r"C:\Program Files\Tesseract-OCR\tesseract.exe") +
                _tg.glob(r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe") +
                _tg.glob(r"C:\tools\Tesseract*\tesseract.exe") +
                _tg.glob(os.path.expanduser(r"~\AppData\Local\Programs\Tesseract*\tesseract.exe"))
            )
        else:
            _tess_search = ["/usr/local/bin/tesseract", "/opt/homebrew/bin/tesseract",
                            "/usr/bin/tesseract"]
        for _tp in _tess_search:
            if os.path.isfile(_tp):
                try:
                    import pytesseract as _pt
                    _pt.pytesseract.tesseract_cmd = _tp
                    _tesseract_found = True
                except ImportError:
                    pass
                break
        if not _tesseract_found:
            print("  ⚠ Tesseract not found — PDF highlighting will be unavailable.")
            print("    Install: https://github.com/tesseract-ocr/tesseract")
            print("    Or set: TESSERACT_CMD=/path/to/tesseract")

# ─── CAS: Telemetry Store (lazy init) ─────────────────────────────────────────
_telemetry_store = None

def _get_telemetry_store():
    """Lazy-init the CAS telemetry store. Returns TelemetryStore or None."""
    global _telemetry_store
    if _telemetry_store is None:
        try:
            from telemetry_store import TelemetryStore
            _telemetry_store = TelemetryStore(str(Path(__file__).parent / "data" / "bearden.db"))
        except Exception:
            pass
    return _telemetry_store

# ─── T-TXN-LEDGER-1: Transaction Store (lazy init) ───────────────────────────
_transaction_store = None

def _get_transaction_store():
    """Lazy-init the transaction store. Returns TransactionStore or None."""
    global _transaction_store
    if _transaction_store is None:
        try:
            from transaction_store import TransactionStore
            _transaction_store = TransactionStore(str(Path(__file__).parent / "data" / "bearden.db"))
        except Exception:
            pass
    return _transaction_store

# ─── Configuration ────────────────────────────────────────────────────────────

BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
UPLOAD_DIR = DATA_DIR / "uploads"
OUTPUT_DIR = DATA_DIR / "outputs"
CLIENTS_DIR = BASE_DIR / "clients"
PAGES_DIR = DATA_DIR / "page_images"
EVIDENCE_DIR = DATA_DIR / "evidence"
VERIFY_DIR = BASE_DIR / "verifications"
JOBS_FILE = DATA_DIR / "jobs_history.json"

for d in [DATA_DIR, UPLOAD_DIR, OUTPUT_DIR, CLIENTS_DIR, PAGES_DIR, EVIDENCE_DIR, VERIFY_DIR]:
    d.mkdir(exist_ok=True)

# Guided review lock timeout (seconds)
REVIEW_LOCK_TIMEOUT_SECONDS = 300

VENDOR_CATEGORIES_FILE = DATA_DIR / "vendor_categories.json"

# GAP-009: DB path configurable via env var to avoid OneDrive sync corruption
DB_PATH = Path(os.environ.get("BEARDEN_DB_PATH", str(DATA_DIR / "bearden.db")))

# GAP-009: Warn if SQLite DB lives under a cloud-synced directory
_CLOUD_SYNC_MARKERS = ("OneDrive", "Dropbox", "Google Drive", "iCloudDrive")
_db_path_str = str(DB_PATH).replace("\\", "/")
for _marker in _CLOUD_SYNC_MARKERS:
    if _marker.lower() in _db_path_str.lower():
        import warnings
        warnings.warn(
            f"[GAP-009] SQLite database '{DB_PATH}' is under a cloud-synced "
            f"directory ({_marker}). This risks corruption if the .db or .db-wal "
            f"files sync mid-transaction. Set BEARDEN_DB_PATH to a local path "
            f"(e.g., C:/data/bearden.db) to mitigate.",
            RuntimeWarning,
            stacklevel=1,
        )
        break

def _secure_file(path):
    """Set restrictive permissions on sensitive files (owner-only read/write)."""
    try:
        os.chmod(str(path), 0o600)
    except OSError:
        pass  # Non-fatal: may fail on some filesystems


# ─── SQLite Database ──────────────────────────────────────────────────────────

def _get_db():
    """Get a SQLite connection. Each call creates a new connection (thread-safe)."""
    conn = sqlite3.connect(str(DB_PATH), timeout=10)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    return conn

def _safe_create_index(conn, sql):
    """Create index, silently skip if column doesn't exist (older schema)."""
    try:
        conn.execute(sql)
    except sqlite3.OperationalError:
        pass

def _init_db():
    """Create tables if needed. Migrate from JSON files on first run."""
    conn = _get_db()
    try:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS jobs (
                id TEXT PRIMARY KEY,
                data TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT '',
                client_name TEXT DEFAULT '',
                created TEXT DEFAULT '',
                updated TEXT DEFAULT ''
            );
            CREATE INDEX IF NOT EXISTS idx_jobs_status ON jobs(status);
            CREATE INDEX IF NOT EXISTS idx_jobs_client ON jobs(client_name);

            CREATE TABLE IF NOT EXISTS verifications (
                job_id TEXT PRIMARY KEY,
                data TEXT NOT NULL,
                updated TEXT DEFAULT ''
            );

            CREATE TABLE IF NOT EXISTS vendor_categories (
                vendor TEXT PRIMARY KEY,
                category TEXT NOT NULL,
                data TEXT NOT NULL,
                updated TEXT DEFAULT ''
            );

            CREATE TABLE IF NOT EXISTS verified_fields (
                job_id TEXT NOT NULL,
                field_key TEXT NOT NULL,
                canonical_value TEXT,
                original_value TEXT,
                status TEXT NOT NULL DEFAULT 'confirmed',
                category TEXT DEFAULT '',
                vendor_desc TEXT DEFAULT '',
                note TEXT DEFAULT '',
                reviewer TEXT DEFAULT '',
                verified_at TEXT NOT NULL DEFAULT '',
                PRIMARY KEY (job_id, field_key)
            );
            CREATE INDEX IF NOT EXISTS idx_vf_job ON verified_fields(job_id);

            CREATE TABLE IF NOT EXISTS client_canonical_values (
                client_name TEXT NOT NULL,
                year TEXT NOT NULL,
                document_type TEXT NOT NULL,
                payer_key TEXT NOT NULL,
                payer_display TEXT DEFAULT '',
                field_name TEXT NOT NULL,
                canonical_value TEXT,
                original_value TEXT,
                status TEXT NOT NULL DEFAULT 'confirmed',
                source_job_id TEXT NOT NULL,
                reviewer TEXT DEFAULT '',
                verified_at TEXT NOT NULL DEFAULT '',
                PRIMARY KEY (client_name, year, document_type, payer_key, field_name)
            );
            CREATE INDEX IF NOT EXISTS idx_ccv_client_year
                ON client_canonical_values(client_name, year);
        """)
        # T1.6: Extend client_canonical_values for workpaper support
        for col, col_type in [("evidence_ref", "TEXT DEFAULT ''"),
                               ("source_doc", "TEXT DEFAULT ''"),
                               ("page_number", "INTEGER")]:
            try:
                conn.execute(
                    f"ALTER TABLE client_canonical_values ADD COLUMN {col} {col_type}"
                )
            except sqlite3.OperationalError:
                pass  # Column already exists

        # T1.6.2: Unified facts table — single source of truth
        conn.execute("""
            CREATE TABLE IF NOT EXISTS facts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_id TEXT NOT NULL,
                client_id TEXT NOT NULL,
                tax_year INTEGER,
                fact_key TEXT NOT NULL,
                value_num REAL,
                value_text TEXT,
                status TEXT NOT NULL DEFAULT 'extracted',
                confidence REAL,
                source_method TEXT,
                source_doc TEXT,
                source_page INTEGER,
                evidence_ref TEXT,
                locked INTEGER NOT NULL DEFAULT 0,
                updated_at TEXT NOT NULL,
                UNIQUE(job_id, tax_year, fact_key)
            )
        """)
        # Indexes on facts — wrapped in try/except for older DBs with different schema
        _safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_facts_job ON facts(job_id)")
        _safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_facts_client_year ON facts(client_id, tax_year)")
        _safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_facts_fact_key ON facts(fact_key)")

        # Sprint 2: Users table — PIN-based auth, role-based access
        conn.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                display_name TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'reviewer',
                pin_hash TEXT NOT NULL DEFAULT '',
                must_reset_pin INTEGER NOT NULL DEFAULT 0,
                is_active INTEGER NOT NULL DEFAULT 1,
                last_login TEXT DEFAULT '',
                created_at TEXT NOT NULL DEFAULT '',
                updated_at TEXT NOT NULL DEFAULT ''
            )
        """)
        _safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_users_username ON users(username)")
        _safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_users_role ON users(role)")

        # Migrate: ensure users table has 'id' column (old DBs may lack it)
        try:
            cols = [r[1] for r in conn.execute("PRAGMA table_info(users)").fetchall()]
            if "id" not in cols:
                conn.execute("ALTER TABLE users RENAME TO users_old")
                conn.execute("""
                    CREATE TABLE users (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        username TEXT NOT NULL UNIQUE,
                        display_name TEXT NOT NULL,
                        role TEXT NOT NULL DEFAULT 'reviewer',
                        pin_hash TEXT NOT NULL DEFAULT '',
                        must_reset_pin INTEGER NOT NULL DEFAULT 0,
                        is_active INTEGER NOT NULL DEFAULT 1,
                        last_login TEXT DEFAULT '',
                        created_at TEXT NOT NULL DEFAULT '',
                        updated_at TEXT NOT NULL DEFAULT ''
                    )
                """)
                conn.execute("""INSERT INTO users (username, display_name, role, pin_hash, must_reset_pin, is_active, last_login, created_at, updated_at)
                                SELECT username, display_name, role, pin_hash, 0, is_active, last_login, created_at, updated_at FROM users_old""")
                conn.execute("DROP TABLE users_old")
                conn.commit()
        except Exception:
            pass  # Table already has id or doesn't exist yet

        # SEC-006: Migrate — add must_reset_pin column if missing
        try:
            cols = [r[1] for r in conn.execute("PRAGMA table_info(users)").fetchall()]
            if "must_reset_pin" not in cols:
                conn.execute("ALTER TABLE users ADD COLUMN must_reset_pin INTEGER NOT NULL DEFAULT 1")
                conn.commit()
                print("  Migrated: added must_reset_pin column (all existing users flagged for reset)")
        except Exception:
            pass  # Column already exists

        # Sprint 2: Audit events — immutable by default, no delete endpoint
        conn.execute("""
            CREATE TABLE IF NOT EXISTS app_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ts TEXT NOT NULL,
                level TEXT NOT NULL DEFAULT 'info',
                event_type TEXT NOT NULL,
                message TEXT NOT NULL DEFAULT '',
                user_id INTEGER,
                user_display TEXT DEFAULT '',
                job_id TEXT DEFAULT '',
                details_json TEXT DEFAULT '',
                ip_addr TEXT DEFAULT ''
            )
        """)
        _safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_events_ts ON app_events(ts)")
        _safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_events_type ON app_events(event_type)")
        _safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_events_level ON app_events(level)")
        _safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_events_job ON app_events(job_id)")
        _safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_events_user ON app_events(user_id)")

        # Guided review: concurrency locks
        conn.execute("""
            CREATE TABLE IF NOT EXISTS review_locks (
                job_id TEXT NOT NULL,
                field_id TEXT NOT NULL,
                locked_by TEXT NOT NULL,
                locked_at TEXT NOT NULL,
                expires_at TEXT NOT NULL,
                PRIMARY KEY (job_id, field_id)
            )
        """)

        # B1: Review chain columns on jobs table
        for col, col_type in [("review_stage", "TEXT DEFAULT 'draft'"),
                               ("stage_owner_role", "TEXT DEFAULT ''"),
                               ("stage_updated", "TEXT DEFAULT ''")]:
            try:
                conn.execute(f"ALTER TABLE jobs ADD COLUMN {col} {col_type}")
            except sqlite3.OperationalError:
                pass

        # B1: Review chain columns on verified_fields table
        for col, col_type in [("review_stage", "TEXT DEFAULT ''"),
                               ("reviewer_id", "INTEGER")]:
            try:
                conn.execute(f"ALTER TABLE verified_fields ADD COLUMN {col} {col_type}")
            except sqlite3.OperationalError:
                pass

        # B9: Time tracking — per-field duration on verified_fields
        for col, col_type in [("field_duration_ms", "INTEGER")]:
            try:
                conn.execute(f"ALTER TABLE verified_fields ADD COLUMN {col} {col_type}")
            except sqlite3.OperationalError:
                pass

        # B9: Review sessions — tracks wall-clock review time per job
        conn.execute("""
            CREATE TABLE IF NOT EXISTS review_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_id TEXT NOT NULL,
                reviewer TEXT NOT NULL DEFAULT '',
                reviewer_id INTEGER,
                session_start TEXT NOT NULL,
                session_end TEXT,
                duration_seconds INTEGER,
                fields_reviewed INTEGER DEFAULT 0,
                created_at TEXT NOT NULL DEFAULT ''
            )
        """)
        _safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_rsess_job ON review_sessions(job_id)")

        # B1: Auto-set existing complete jobs to preparer_review if still draft
        conn.execute("""
            UPDATE jobs SET review_stage = 'preparer_review', stage_updated = datetime('now')
            WHERE status = 'complete'
              AND (review_stage IS NULL OR review_stage = '' OR review_stage = 'draft')
        """)

        # ─── CAS: Operational Telemetry Tables (T-CAS-1) ─────────────────────
        # These op_* tables store ONLY operational metrics — never financial data.
        # Dropping all op_* tables leaves the financial pipeline fully intact.

        conn.execute("""
            CREATE TABLE IF NOT EXISTS op_runs (
                id INTEGER PRIMARY KEY,
                job_id TEXT UNIQUE,
                client_name TEXT,
                doc_type TEXT,
                status TEXT DEFAULT 'running',
                started_at TEXT,
                finished_at TEXT,
                total_s REAL,
                cost_usd REAL,
                total_pages INTEGER,
                pages_ocr INTEGER,
                pages_vision INTEGER,
                pages_blank INTEGER,
                cache_hit INTEGER DEFAULT 0,
                total_fields INTEGER,
                fields_high_conf INTEGER,
                fields_low_conf INTEGER,
                fields_needs_review INTEGER,
                total_api_calls INTEGER,
                vision_calls INTEGER,
                text_calls INTEGER,
                input_tokens INTEGER,
                output_tokens INTEGER,
                time_to_first_values_s REAL,
                batches_total INTEGER,
                fields_streamed INTEGER,
                app_version TEXT,
                extract_version TEXT,
                log_path TEXT,
                error_message TEXT
            )
        """)
        _safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_op_runs_job ON op_runs(job_id)")
        _safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_op_runs_started ON op_runs(started_at)")
        _safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_op_runs_status ON op_runs(status)")

        conn.execute("""
            CREATE TABLE IF NOT EXISTS op_phases (
                id INTEGER PRIMARY KEY,
                run_id INTEGER REFERENCES op_runs(id),
                job_id TEXT,
                phase_name TEXT,
                duration_s REAL,
                UNIQUE(run_id, phase_name)
            )
        """)

        conn.execute("""
            CREATE TABLE IF NOT EXISTS op_drift (
                id INTEGER PRIMARY KEY,
                job_id TEXT UNIQUE,
                measured_at TEXT,
                edit_rate REAL,
                missing_evidence_rate REAL,
                needs_review_rate REAL,
                audit_pass_rate REAL,
                low_confidence_rate REAL
            )
        """)

        conn.execute("""
            CREATE TABLE IF NOT EXISTS op_smoke_results (
                id INTEGER PRIMARY KEY,
                run_at TEXT,
                passed INTEGER,
                total_checks INTEGER,
                results_json TEXT,
                duration_s REAL
            )
        """)

        conn.execute("""
            CREATE TABLE IF NOT EXISTS op_golden_results (
                id INTEGER PRIMARY KEY,
                run_at TEXT,
                golden_name TEXT,
                passed INTEGER,
                total_checks INTEGER,
                fields_matched INTEGER,
                fields_mismatched INTEGER,
                fields_missing INTEGER,
                fields_extra INTEGER,
                duration_s REAL,
                details_json TEXT
            )
        """)

        conn.execute("""
            CREATE TABLE IF NOT EXISTS op_backups (
                id INTEGER PRIMARY KEY,
                created_at TEXT,
                backup_path TEXT,
                db_size_bytes INTEGER,
                sha256 TEXT,
                row_counts_json TEXT,
                verified INTEGER DEFAULT 0,
                verify_sha256 TEXT,
                verify_at TEXT
            )
        """)
        # ─── CAS: Change Request Tables (T-CAS-2B) ─────────────────────────

        conn.execute("""
            CREATE TABLE IF NOT EXISTS op_change_requests (
                id INTEGER PRIMARY KEY,
                cr_id TEXT UNIQUE NOT NULL,
                status TEXT NOT NULL DEFAULT 'open',
                severity TEXT NOT NULL DEFAULT 'WARNING',
                source TEXT NOT NULL,
                trigger_summary TEXT,
                trigger_snapshot TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT,
                closed_at TEXT,
                closed_by TEXT,
                folder_path TEXT
            )
        """)
        _safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_op_cr_id ON op_change_requests(cr_id)")
        _safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_op_cr_status ON op_change_requests(status)")
        _safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_op_cr_created ON op_change_requests(created_at)")

        conn.execute("""
            CREATE TABLE IF NOT EXISTS op_cr_findings (
                id INTEGER PRIMARY KEY,
                cr_id TEXT NOT NULL,
                finding_id TEXT NOT NULL,
                severity TEXT NOT NULL DEFAULT 'WARNING',
                source TEXT NOT NULL,
                check_name TEXT,
                details TEXT,
                measured_value TEXT,
                threshold TEXT,
                recommended_action TEXT,
                UNIQUE(cr_id, finding_id)
            )
        """)
        _safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_op_findings_cr ON op_cr_findings(cr_id)")

        conn.execute("""
            CREATE TABLE IF NOT EXISTS op_post_fix_gates (
                id INTEGER PRIMARY KEY,
                cr_id TEXT NOT NULL,
                run_at TEXT NOT NULL,
                gate_result TEXT NOT NULL,
                checks_run INTEGER,
                checks_passed INTEGER,
                before_snapshot TEXT,
                after_snapshot TEXT,
                details_json TEXT
            )
        """)
        _safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_op_gates_cr ON op_post_fix_gates(cr_id)")

        # ─── End CAS Tables ──────────────────────────────────────────────────

        # ─── T-TXN-LEDGER-1: Transaction Ledger Tables ─────────────────────
        # These tables are owned by transaction_store.py but mirrored here
        # for consistency with the codebase pattern.
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS txn_values (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                txn_id TEXT NOT NULL UNIQUE,
                job_id TEXT NOT NULL,
                client_name TEXT NOT NULL,
                year INTEGER NOT NULL,
                document_type TEXT NOT NULL,
                payer_key TEXT NOT NULL DEFAULT '',
                txn_index INTEGER NOT NULL,
                txn_date TEXT DEFAULT '',
                description TEXT DEFAULT '',
                amount REAL,
                txn_type TEXT DEFAULT '',
                category TEXT DEFAULT '',
                category_group TEXT DEFAULT '',
                vendor_norm TEXT DEFAULT '',
                status TEXT NOT NULL DEFAULT 'staged',
                source_page INTEGER,
                confidence TEXT DEFAULT '',
                evidence_ref TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_txn_values_job ON txn_values(job_id);
            CREATE INDEX IF NOT EXISTS idx_txn_values_client_year ON txn_values(client_name, year);
            CREATE INDEX IF NOT EXISTS idx_txn_values_status ON txn_values(status);
            CREATE INDEX IF NOT EXISTS idx_txn_values_category ON txn_values(category);
            CREATE INDEX IF NOT EXISTS idx_txn_values_vendor ON txn_values(vendor_norm);
            CREATE INDEX IF NOT EXISTS idx_txn_values_date ON txn_values(txn_date);

            CREATE TABLE IF NOT EXISTS txn_events (
                event_id INTEGER PRIMARY KEY AUTOINCREMENT,
                txn_id TEXT NOT NULL,
                event_type TEXT NOT NULL,
                old_value TEXT DEFAULT '',
                new_value TEXT DEFAULT '',
                reviewer TEXT DEFAULT '',
                event_at TEXT NOT NULL,
                details_json TEXT DEFAULT ''
            );
            CREATE INDEX IF NOT EXISTS idx_txn_events_txn ON txn_events(txn_id);
            CREATE INDEX IF NOT EXISTS idx_txn_events_type ON txn_events(event_type);

            CREATE TABLE IF NOT EXISTS txn_evidence (
                evidence_id INTEGER PRIMARY KEY AUTOINCREMENT,
                txn_id TEXT NOT NULL,
                job_id TEXT NOT NULL,
                page_number INTEGER,
                crop_coords TEXT DEFAULT '',
                ocr_text TEXT DEFAULT ''
            );
            CREATE INDEX IF NOT EXISTS idx_txn_evidence_txn ON txn_evidence(txn_id);

            CREATE TABLE IF NOT EXISTS vendor_rules (
                rule_id INTEGER PRIMARY KEY AUTOINCREMENT,
                vendor_pattern TEXT NOT NULL,
                match_type TEXT NOT NULL DEFAULT 'exact',
                category TEXT NOT NULL,
                category_group TEXT DEFAULT '',
                source TEXT NOT NULL DEFAULT 'manual',
                confidence REAL DEFAULT 1.0,
                usage_count INTEGER DEFAULT 0,
                created_by TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE UNIQUE INDEX IF NOT EXISTS idx_vendor_rules_unique ON vendor_rules(vendor_pattern, match_type);
            CREATE INDEX IF NOT EXISTS idx_vendor_rules_pattern ON vendor_rules(vendor_pattern);
            CREATE INDEX IF NOT EXISTS idx_vendor_rules_category ON vendor_rules(category);

            CREATE TABLE IF NOT EXISTS category_rules (
                rule_id INTEGER PRIMARY KEY AUTOINCREMENT,
                keyword TEXT NOT NULL,
                category TEXT NOT NULL,
                category_group TEXT DEFAULT '',
                priority INTEGER DEFAULT 100,
                source TEXT NOT NULL DEFAULT 'manual',
                created_by TEXT DEFAULT '',
                created_at TEXT NOT NULL
            );
            CREATE UNIQUE INDEX IF NOT EXISTS idx_category_rules_unique ON category_rules(keyword, category);
            CREATE INDEX IF NOT EXISTS idx_category_rules_keyword ON category_rules(keyword);
        """)
        # ─── End Transaction Ledger Tables ──────────────────────────────────

        # ─── Lite Platform: Event Store ───────────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS lite_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_id TEXT NOT NULL,
                event_type TEXT NOT NULL,
                event_data TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
        """)
        _safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_lite_events_job ON lite_events(job_id)")
        _safe_create_index(conn, "CREATE INDEX IF NOT EXISTS idx_lite_events_type ON lite_events(event_type)")
        # ─── End Lite Platform Tables ─────────────────────────────────────────

        conn.commit()
        _secure_file(DB_PATH)
        _migrate_from_json(conn)
    finally:
        conn.close()

def _migrate_from_json(conn):
    """One-time import from legacy JSON files into SQLite."""
    # Migrate jobs_history.json
    row_count = conn.execute("SELECT COUNT(*) FROM jobs").fetchone()[0]
    if row_count == 0 and JOBS_FILE.exists():
        try:
            with open(JOBS_FILE) as f:
                legacy_jobs = json.load(f)
            for jid, jdata in legacy_jobs.items():
                conn.execute(
                    "INSERT OR IGNORE INTO jobs (id, data, status, client_name, created, updated) VALUES (?, ?, ?, ?, ?, ?)",
                    (jid, json.dumps(jdata, default=str), jdata.get("status", ""),
                     jdata.get("client_name", ""), jdata.get("created", ""),
                     datetime.now().isoformat())
                )
            conn.commit()
            print(f"  Migrated {len(legacy_jobs)} jobs from jobs_history.json to SQLite")
            os.rename(str(JOBS_FILE), str(JOBS_FILE) + ".migrated")
        except (json.JSONDecodeError, IOError, OSError) as e:
            print(f"  Warning: Could not migrate jobs_history.json: {e}")

    # Migrate verifications/*.json
    vcount = conn.execute("SELECT COUNT(*) FROM verifications").fetchone()[0]
    if vcount == 0:
        verify_files = list(VERIFY_DIR.glob("*.json"))
        migrated = 0
        for vf in verify_files:
            try:
                job_id = vf.stem
                with open(vf) as f:
                    vdata = json.load(f)
                conn.execute(
                    "INSERT OR IGNORE INTO verifications (job_id, data, updated) VALUES (?, ?, ?)",
                    (job_id, json.dumps(vdata, default=str), vdata.get("updated", ""))
                )
                migrated += 1
            except (json.JSONDecodeError, IOError):
                continue
        conn.commit()
        if migrated:
            print(f"  Migrated {migrated} verification files to SQLite")
            backup_dir = VERIFY_DIR / "_migrated"
            backup_dir.mkdir(exist_ok=True)
            for vf in verify_files:
                try:
                    os.rename(str(vf), str(backup_dir / vf.name))
                except OSError:
                    pass

    # Migrate vendor_categories.json
    vccount = conn.execute("SELECT COUNT(*) FROM vendor_categories").fetchone()[0]
    if vccount == 0 and VENDOR_CATEGORIES_FILE.exists():
        try:
            with open(VENDOR_CATEGORIES_FILE) as f:
                vc_data = json.load(f)
            for vendor, info in vc_data.items():
                cat = info.get("category", "") if isinstance(info, dict) else str(info)
                conn.execute(
                    "INSERT OR IGNORE INTO vendor_categories (vendor, category, data, updated) VALUES (?, ?, ?, ?)",
                    (vendor, cat, json.dumps(info, default=str), datetime.now().isoformat())
                )
            conn.commit()
            print(f"  Migrated {len(vc_data)} vendor categories to SQLite")
            os.rename(str(VENDOR_CATEGORIES_FILE), str(VENDOR_CATEGORIES_FILE) + ".migrated")
        except (json.JSONDecodeError, IOError, OSError) as e:
            print(f"  Warning: Could not migrate vendor_categories.json: {e}")


def _persist_lite_event(job_id, event_type, event_data_json):
    """Store a Lite platform event (VerificationAction, ArdentResult, etc.).

    Feature-flagged — only called when LITE_VERIFICATION_ENABLED is set.
    Non-fatal: swallows all exceptions to avoid disrupting the review flow.
    """
    conn = _get_db()
    try:
        conn.execute(
            "INSERT INTO lite_events (job_id, event_type, event_data, created_at) VALUES (?,?,?,?)",
            (job_id, event_type, event_data_json, datetime.now().isoformat())
        )
        conn.commit()
    except Exception:
        pass  # Non-fatal — Lite event persistence should never break production
    finally:
        conn.close()


def _client_dir(client_name, doc_type, year):
    """Build a per-client output directory:  clients/<Client Name>/<doc_type>/<year>/"""
    safe_client = re.sub(r'[^\w\s\-\.,()]', '', client_name).strip() or "Unknown Client"
    safe_client = safe_client.title()
    type_labels = {
        "tax_returns": "Tax Returns",
        "bank_statements": "Bank Statements",
        "trust_documents": "Trust Documents",
        "bookkeeping": "Bookkeeping",
        "payroll": "Payroll",
        "other": "Other Documents",
    }
    type_folder = type_labels.get(doc_type, "Other")
    client_dir = CLIENTS_DIR / safe_client / type_folder / str(year)
    client_dir.mkdir(parents=True, exist_ok=True)
    return client_dir

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 150 * 1024 * 1024  # 150MB

# Sprint 2: Session secret — persisted so sessions survive restarts
_secret_path = DATA_DIR / ".flask_secret"
if _secret_path.exists():
    app.secret_key = _secret_path.read_bytes()
else:
    import secrets as _secrets
    _sk = _secrets.token_bytes(32)
    _secret_path.write_bytes(_sk)
    _secure_file(_secret_path)
    app.secret_key = _sk
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"

_start_time = datetime.now()
_app_version = "5.2"

# ─── Sprint 2: Auth Constants ────────────────────────────────────────────────

SESSION_IDLE_SECONDS = 45 * 60       # 45 minute idle timeout
LOGIN_LOCKOUT_SECONDS = 120           # 2 minute lockout after max failures
MAX_FAILED_ATTEMPTS = 5               # lockout threshold
VALID_ROLES = frozenset({"admin", "preparer", "partner", "reviewer"})

# ─── Review Chain Stage Model ────────────────────────────────────────────────
# draft → preparer_review → reviewer_review → partner_review → final
REVIEW_STAGES = ("draft", "preparer_review", "reviewer_review", "partner_review", "final")

STAGE_ROLE_MAP = {
    "preparer_review": {"preparer", "admin"},
    "reviewer_review": {"reviewer", "admin"},
    "partner_review":  {"partner", "admin"},
}

STAGE_NEXT = {
    "draft":             "preparer_review",
    "preparer_review":   "reviewer_review",
    "reviewer_review":   "partner_review",
    "partner_review":    "final",
}

STAGE_PREV = {
    "reviewer_review":   "preparer_review",
    "partner_review":    "reviewer_review",
}

STAGE_DISPLAY = {
    "draft":             "Draft",
    "preparer_review":   "Preparer Review",
    "reviewer_review":   "Reviewer Review",
    "partner_review":    "Partner Review",
    "final":             "Final",
}

def can_act_at_stage(user_role, stage):
    """Check if a user role is authorized to act at a given review stage."""
    if user_role == "admin":
        return True
    return user_role in STAGE_ROLE_MAP.get(stage, set())

_failed_logins = {}  # key=(username, ip) -> {"count": int, "locked_until": epoch}


# ─── Sprint 2: User DB Functions ─────────────────────────────────────────────

def _seed_default_users():
    """Create default firm users if users table is empty. Called once at startup.

    SEC-006: Each user gets a unique random PIN (not a shared default).
    All seeded users have must_reset_pin=1 so they're forced to change on first login.
    """
    conn = _get_db()
    try:
        count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        if count > 0:
            return  # Already seeded
        now = datetime.now().isoformat()
        seed_users = [
            ("jeff",    "Jeffrey Watts",  "admin"),
            ("susan",   "Susan",          "reviewer"),
            ("charles", "Charles",        "partner"),
            ("chris",   "Chris",          "partner"),
            ("ashley",  "Ashley",         "preparer"),
            ("leigh",   "Leigh",          "preparer"),
            ("molly",   "Molly",          "reviewer"),
        ]
        temp_pins = {}
        for username, display_name, role in seed_users:
            temp_pin = generate_6_digit_pin()
            pin_hash = generate_password_hash(temp_pin)
            temp_pins[username] = temp_pin
            conn.execute(
                """INSERT OR IGNORE INTO users
                   (username, display_name, role, pin_hash, must_reset_pin, is_active, created_at, updated_at)
                   VALUES (?, ?, ?, ?, 1, 1, ?, ?)""",
                (username, display_name, role, pin_hash, now, now)
            )
        conn.commit()
        # Print temp PINs to console for admin to distribute securely
        print(f"  Seeded {len(seed_users)} default users with unique PINs (must reset on first login):")
        for username, pin in temp_pins.items():
            print(f"    {username}: {pin}")
    finally:
        conn.close()

    # B1: Migrate existing user roles for review chain
    conn2 = _get_db()
    try:
        # Susan should be reviewer (second pass), not partner
        conn2.execute("UPDATE users SET role='reviewer' WHERE username='susan' AND role='partner'")
        # Ashley/Leigh should be preparers
        conn2.execute("UPDATE users SET role='preparer' WHERE username='ashley' AND role='reviewer'")
        conn2.execute("UPDATE users SET role='preparer' WHERE username='leigh' AND role='reviewer'")
        conn2.commit()
    finally:
        conn2.close()


def get_user_by_id(user_id):
    """Fetch a single user by ID. Returns dict or None."""
    conn = _get_db()
    try:
        row = conn.execute(
            """SELECT id, username, display_name, role, pin_hash,
                      must_reset_pin, is_active, last_login, created_at, updated_at
               FROM users WHERE id = ?""", (user_id,)
        ).fetchone()
        if not row:
            return None
        return _user_row_to_dict(row)
    finally:
        conn.close()


def get_user_by_username(username):
    """Fetch a single user by username. Returns dict or None."""
    conn = _get_db()
    try:
        row = conn.execute(
            """SELECT id, username, display_name, role, pin_hash,
                      must_reset_pin, is_active, last_login, created_at, updated_at
               FROM users WHERE username = ?""", (username,)
        ).fetchone()
        if not row:
            return None
        return _user_row_to_dict(row)
    finally:
        conn.close()


def list_all_users():
    """List all users (for admin user management)."""
    conn = _get_db()
    try:
        rows = conn.execute(
            """SELECT id, username, display_name, role, pin_hash,
                      must_reset_pin, is_active, last_login, created_at, updated_at
               FROM users ORDER BY role, username"""
        ).fetchall()
        return [_user_row_to_dict(r) for r in rows]
    finally:
        conn.close()


def list_active_users():
    """List active users (for login dropdown)."""
    conn = _get_db()
    try:
        rows = conn.execute(
            """SELECT id, username, display_name, role, pin_hash,
                      must_reset_pin, is_active, last_login, created_at, updated_at
               FROM users WHERE is_active = 1 ORDER BY display_name"""
        ).fetchall()
        return [_user_row_to_dict(r) for r in rows]
    finally:
        conn.close()


def update_last_login(user_id):
    """Update last_login timestamp for a user."""
    conn = _get_db()
    try:
        conn.execute(
            "UPDATE users SET last_login = ?, updated_at = ? WHERE id = ?",
            (datetime.now().isoformat(), datetime.now().isoformat(), user_id)
        )
        conn.commit()
    finally:
        conn.close()


def set_user_pin_hash(user_id, pin_hash):
    """Set the pin_hash for a user (admin reset)."""
    conn = _get_db()
    try:
        conn.execute(
            "UPDATE users SET pin_hash = ?, updated_at = ? WHERE id = ?",
            (pin_hash, datetime.now().isoformat(), user_id)
        )
        conn.commit()
    finally:
        conn.close()


def set_user_active(user_id, is_active):
    """Enable or disable a user."""
    conn = _get_db()
    try:
        conn.execute(
            "UPDATE users SET is_active = ?, updated_at = ? WHERE id = ?",
            (1 if is_active else 0, datetime.now().isoformat(), user_id)
        )
        conn.commit()
    finally:
        conn.close()


def create_user(username, display_name, role, pin_hash):
    """Create a new user. Returns user ID."""
    if role not in VALID_ROLES:
        raise ValueError(f"Invalid role: {role}")
    conn = _get_db()
    try:
        now = datetime.now().isoformat()
        cursor = conn.execute(
            """INSERT INTO users
               (username, display_name, role, pin_hash, is_active, created_at, updated_at)
               VALUES (?, ?, ?, ?, 1, ?, ?)""",
            (username, display_name, role, pin_hash, now, now)
        )
        conn.commit()
        return cursor.lastrowid
    finally:
        conn.close()


def generate_6_digit_pin():
    """Generate a random 6-digit PIN string."""
    import secrets as _secrets
    return str(_secrets.randbelow(900000) + 100000)


# SEC-006: PIN complexity requirements
_PIN_REJECT_PATTERNS = {
    "000000", "111111", "222222", "333333", "444444",
    "555555", "666666", "777777", "888888", "999999",
    "123456", "654321", "012345", "543210",
}


def validate_pin_complexity(pin: str) -> str | None:
    """Check PIN meets complexity requirements.

    Returns error message or None if valid.
    Requirements:
      - Exactly 6 digits
      - Not all same digit (e.g., 111111)
      - Not a simple sequence (e.g., 123456)
    """
    if not pin or not isinstance(pin, str):
        return "PIN is required"
    if not pin.isdigit() or len(pin) != 6:
        return "PIN must be exactly 6 digits"
    if pin in _PIN_REJECT_PATTERNS:
        return "PIN is too simple — avoid repeated digits or sequences"
    # At least 2 unique digits
    if len(set(pin)) < 2:
        return "PIN must contain at least 2 different digits"
    return None


def clear_must_reset_pin(user_id: int):
    """Clear the must_reset_pin flag after a successful PIN change."""
    conn = _get_db()
    try:
        conn.execute(
            "UPDATE users SET must_reset_pin = 0, updated_at = ? WHERE id = ?",
            (datetime.now().isoformat(), user_id),
        )
        conn.commit()
    finally:
        conn.close()


def _user_row_to_dict(row):
    """Convert a users table row to a dict.

    Expects columns: id, username, display_name, role, pin_hash,
    must_reset_pin, is_active, last_login, created_at, updated_at
    """
    return {
        "id": row[0],
        "username": row[1],
        "display_name": row[2],
        "role": row[3],
        "pin_hash": row[4],
        "must_reset_pin": bool(row[5]) if len(row) > 9 else False,
        "is_active": bool(row[6] if len(row) > 9 else row[5]),
        "last_login": (row[7] if len(row) > 9 else row[6]) or "",
        "created_at": row[8] if len(row) > 9 else row[7],
        "updated_at": row[9] if len(row) > 9 else row[8],
    }


# ─── Sprint 2: Auth Helpers ──────────────────────────────────────────────────

def current_user():
    """Get the currently logged-in user from session. Returns dict or None."""
    uid = session.get("user_id")
    if not uid:
        return None
    return get_user_by_id(uid)


def _is_api_request():
    """Return True if the current request is an API/JSON call (not a browser page)."""
    return (request.path.startswith("/api/") or
            request.accept_mimetypes.best == "application/json" or
            "X-Scope-Id" in request.headers)


def require_login(fn):
    """Decorator: redirect to /login if not authenticated or session expired.
    For API requests, return JSON 401 instead of a redirect."""
    import functools
    @functools.wraps(fn)
    def wrapper(*args, **kwargs):
        import time as _time
        uid = session.get("user_id")
        last = session.get("last_seen")
        now = int(_time.time())
        if not uid:
            if _is_api_request():
                return jsonify({"error": "Authentication required"}), 401
            return redirect("/login")
        if last and now - last > SESSION_IDLE_SECONDS:
            log_event("info", "session_expired", "Session timed out",
                      user_id=uid)
            session.clear()
            if _is_api_request():
                return jsonify({"error": "Session expired"}), 401
            return redirect("/login")
        session["last_seen"] = now
        # SEC-006: Force PIN reset — block navigation to anything except /change-pin and /logout
        if session.get("force_pin_reset") and request.path not in ("/change-pin", "/logout"):
            if _is_api_request():
                return jsonify({"error": "PIN reset required before API access"}), 403
            return redirect("/change-pin")
        return fn(*args, **kwargs)
    return wrapper


def require_role(*roles):
    """Decorator: abort 403 if current user's role not in allowed roles."""
    def deco(fn):
        import functools
        @functools.wraps(fn)
        def wrapper(*args, **kwargs):
            u = current_user()
            if not u or u["role"] not in roles:
                abort(403)
            return fn(*args, **kwargs)
        return wrapper
    return deco


# ─── Sprint 2: Event Logging ─────────────────────────────────────────────────

def log_event(level, event_type, message, user_id=None, job_id=None,
              details=None, ip_addr=None):
    """Write an immutable audit event to app_events. Thread-safe."""
    conn = _get_db()
    try:
        now = datetime.now().isoformat()
        user_display = ""
        if user_id:
            u = get_user_by_id(user_id)
            if u:
                user_display = u["display_name"]
        details_json = json.dumps(details, default=str) if details else ""
        conn.execute(
            """INSERT INTO app_events
               (ts, level, event_type, message, user_id, user_display,
                job_id, details_json, ip_addr)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (now, level, event_type, message, user_id, user_display,
             job_id or "", details_json, ip_addr or "")
        )
        conn.commit()
    except Exception:
        pass  # Never let logging failure crash the app
    finally:
        conn.close()


def query_events(level=None, event_type=None, job_id=None, user_id=None,
                 limit=200):
    """Query audit events with optional filters. Returns list of dicts."""
    conn = _get_db()
    try:
        sql = "SELECT id, ts, level, event_type, message, user_id, user_display, job_id, details_json, ip_addr FROM app_events WHERE 1=1"
        params = []
        if level:
            sql += " AND level = ?"
            params.append(level)
        if event_type:
            sql += " AND event_type = ?"
            params.append(event_type)
        if job_id:
            sql += " AND job_id = ?"
            params.append(job_id)
        if user_id:
            sql += " AND user_id = ?"
            params.append(int(user_id))
        sql += " ORDER BY id DESC LIMIT ?"
        params.append(limit)
        rows = conn.execute(sql, params).fetchall()
        return [_event_row_to_dict(r) for r in rows]
    finally:
        conn.close()


def _event_row_to_dict(row):
    """Convert an app_events row to a dict."""
    return {
        "id": row[0],
        "ts": row[1],
        "level": row[2],
        "event_type": row[3],
        "message": row[4],
        "user_id": row[5],
        "user_display": row[6] or "",
        "job_id": row[7] or "",
        "details_json": row[8] or "",
        "ip_addr": row[9] or "",
    }


# ─── Sprint 2: Admin Summary Builder ─────────────────────────────────────────

def build_admin_summary():
    """Build the admin dashboard summary: health, KPIs, recent jobs, recent events."""
    import shutil as _shutil
    import time as _time

    now = datetime.now()
    uptime_s = (now - _start_time).total_seconds()
    today_str = now.strftime("%Y-%m-%d")

    # Health
    tesseract_ok = _shutil.which("tesseract") is not None
    extract_ok = (BASE_DIR / "extract.py").exists()
    api_key_set = bool(os.environ.get("ANTHROPIC_API_KEY"))
    all_ok = tesseract_ok and extract_ok and api_key_set
    health_state = "good" if all_ok else ("warn" if (tesseract_ok and extract_ok) else "bad")
    health_label = "Healthy" if all_ok else ("Degraded" if health_state == "warn" else "Unhealthy")

    # KPIs from jobs
    jobs_today = 0
    failures_today = 0
    runtimes = []
    ttfv_values = []
    running_count = 0
    for j in jobs.values():
        created = j.get("created", "")
        if created and created[:10] == today_str:
            jobs_today += 1
            if j.get("status") in ("error", "failed"):
                failures_today += 1
        if j.get("status") == "running":
            running_count += 1
        # Compute runtimes from log metadata
        log_path_str = j.get("log_path", "")
        if log_path_str and os.path.exists(log_path_str):
            try:
                with open(log_path_str) as _lf:
                    log_data = json.load(_lf)
                rt = log_data.get("timing", {}).get("total_s")
                if rt:
                    runtimes.append(rt)
                tf = log_data.get("streaming", {}).get("time_to_first_values_s")
                if tf:
                    ttfv_values.append(tf)
            except (json.JSONDecodeError, IOError, KeyError):
                pass

    avg_runtime = round(sum(runtimes) / len(runtimes), 1) if runtimes else 0
    avg_ttfv = round(sum(ttfv_values) / len(ttfv_values), 1) if ttfv_values else 0

    # Disk free
    try:
        usage = _shutil.disk_usage(str(DATA_DIR))
        disk_free_gb = round(usage.free / (1024**3), 1)
    except Exception:
        disk_free_gb = 0

    # Recent jobs (last 10)
    sorted_jobs = sorted(jobs.items(), key=lambda kv: kv[1].get("created", ""), reverse=True)[:10]
    recent_jobs = []
    for jid, j in sorted_jobs:
        status = j.get("status", "unknown")
        status_class = {"done": "good", "error": "bad", "failed": "bad", "running": "warn"}.get(status, "")
        recent_jobs.append({
            "job_id": jid[:12],
            "client_name": j.get("client_name", ""),
            "status": status,
            "status_class": status_class,
            "started": j.get("created", "")[:16].replace("T", " "),
            "runtime_s": j.get("runtime_s", "-"),
        })

    # Recent events (last 15)
    recent_events = query_events(limit=15)

    return {
        "health": {
            "version": _app_version,
            "uptime_h": round(uptime_s / 3600, 1),
            "state": health_state,
            "label": health_label,
        },
        "kpis": {
            "jobs_today": jobs_today,
            "failures_today": failures_today,
            "avg_runtime_s": avg_runtime,
            "time_to_first_values_s": avg_ttfv,
            "disk_free_gb": disk_free_gb,
            "running_jobs": running_count,
        },
        "recent_jobs": recent_jobs,
        "recent_events": recent_events,
    }

VALID_DOC_TYPES = {"tax_returns", "bank_statements", "trust_documents", "bookkeeping", "payroll", "other"}

# In-memory job tracking (persisted to jobs_history.json)
jobs = {}
_jobs_lock = threading.Lock()
_active_procs = {}  # job_id -> subprocess.Popen for cancellation

# ── Aftercare: background task queue for deferred post-confirm work (T-UX-CONFIRM-FASTPATH) ──
_aftercare_queue = collections.deque()
_aftercare_event = threading.Event()
_aftercare_running = True


def _aftercare_worker():
    """Single background thread processing deferred confirm tasks.

    Consumes from _aftercare_queue (FIFO). Sleeps when idle, wakes on enqueue
    via _aftercare_event. Runs as a daemon thread — exits when main process exits.
    """
    while _aftercare_running or _aftercare_queue:
        _aftercare_event.wait(timeout=1.0)
        _aftercare_event.clear()
        while _aftercare_queue:
            try:
                task = _aftercare_queue.popleft()
                _process_aftercare(task)
            except Exception as e:
                print(f"  Aftercare error: {e}")


def _process_aftercare(task):
    """Run deferred work for a single confirm action.

    Operations: canonical promotion, FactStore update, verify summary, audit log.
    All wrapped in try/except — failures logged, never block the user.
    """
    job_id = task["job_id"]
    incoming = task["incoming"]
    reviewer = task.get("reviewer", "")
    action = task.get("action", "confirm")
    field_id = task.get("field_id", "")
    reviewer_id = task.get("reviewer_id")
    field_count = task.get("field_count", 1)
    statuses = task.get("statuses", {})
    mode = task.get("mode", "guided")

    # 1. Canonical promotion + FactStore
    _aftercare_promote_facts(job_id, incoming)

    # 2. Verify summary + save_jobs
    try:
        vdata = _load_verifications(job_id)
        _update_verify_summary(job_id, vdata)
    except Exception:
        pass

    # 3. Audit log
    try:
        if mode == "guided":
            log_event("info", "fact_verified",
                      f"Guided review: {action} {field_id} on job {job_id[:12]}",
                      user_id=reviewer_id, job_id=job_id,
                      details={"reviewer": reviewer, "action": action,
                               "field_id": field_id, "mode": "guided"})
        else:
            log_event("info", "fact_verified",
                      f"Verified {field_count} field(s) on job {job_id[:12]}",
                      user_id=reviewer_id, job_id=job_id,
                      details={"reviewer": reviewer, "field_count": field_count,
                               "statuses": statuses})
    except Exception:
        pass


def _aftercare_promote_facts(job_id, incoming_fields):
    """Promote confirmed/corrected values to client_canonical + FactStore.

    This is the expensive part of _upsert_verified_fields extracted for
    background execution. Opens its own DB connection.
    """
    if not incoming_fields:
        return
    job = jobs.get(job_id)
    log_data = _load_extraction_log(job_id)
    if not job or not log_data:
        return

    client_name = job.get("client_name", "")
    year = job.get("year", "")
    if not client_name or not year:
        return

    conn = _get_db()
    try:
        from fact_store import FactStore
        fs = FactStore(str(DB_PATH))
        tax_year = int(year) if year.isdigit() else None
    except Exception:
        fs = None
        tax_year = None

    try:
        now = datetime.now().isoformat()
        for field_key, decision in incoming_fields.items():
            status = decision.get("status", "")
            if status not in ("confirmed", "corrected"):
                continue
            ext, fn = _resolve_extraction_for_field(log_data, field_key)
            if not ext or not fn:
                continue
            doc_type = ext.get("document_type", "")
            if not doc_type:
                continue
            payer_key = _normalize_payer_key(ext)
            payer_display = ext.get("payer_or_entity", "")

            canonical = None
            original = _resolve_field_value(log_data, field_key)
            if status == "corrected":
                canonical = decision.get("corrected_value")
            elif status == "confirmed":
                canonical = original

            if canonical is not None:
                _upsert_client_canonical(
                    conn, client_name, year, doc_type, payer_key,
                    payer_display, fn, canonical, original,
                    status, job_id,
                    decision.get("reviewer", ""),
                    decision.get("timestamp", now)
                )

            # FactStore update
            if fs and tax_year is not None:
                fact_key = FactStore.fact_key(doc_type, payer_key, fn)
                try:
                    if status == "corrected" and canonical is not None:
                        corr_num = None
                        corr_text = None
                        try:
                            corr_num = float(str(canonical).replace(",", "").replace("$", ""))
                        except (ValueError, TypeError):
                            corr_text = str(canonical)
                        fs.apply_correction(
                            job_id, tax_year, fact_key,
                            value_num=corr_num, value_text=corr_text,
                            reviewer=decision.get("reviewer", "")
                        )
                    elif status == "confirmed":
                        fs.upgrade_fact_status(
                            job_id, tax_year, fact_key, "confirmed"
                        )
                except Exception:
                    pass
        conn.commit()
    except Exception as e:
        print(f"  Aftercare canonical promotion error: {e}")
    finally:
        conn.close()


def _enqueue_aftercare(task):
    """Add a task to the aftercare queue and wake the worker."""
    _aftercare_queue.append(task)
    _aftercare_event.set()


# Start aftercare worker thread
_aftercare_thread = threading.Thread(target=_aftercare_worker, daemon=True, name="aftercare")
_aftercare_thread.start()


def load_jobs():
    """Load all jobs from SQLite into the in-memory dict."""
    global jobs
    conn = _get_db()
    try:
        rows = conn.execute("SELECT id, data, review_stage FROM jobs").fetchall()
        jobs = {}
        for row in rows:
            jid = row[0]
            data_json = row[1]
            review_stage = row[2] if len(row) > 2 else "draft"
            try:
                jdata = json.loads(data_json)
                # Clear stale "running" or "queued" jobs from previous sessions
                if jdata.get("status") in ("running", "queued"):
                    jdata["status"] = "interrupted"
                # Sync review_stage from DB column into in-memory dict
                jdata["review_stage"] = review_stage or "draft"
                jobs[jid] = jdata
            except json.JSONDecodeError:
                print(f"  Warning: Could not parse job {jid}")
    except sqlite3.Error as e:
        print(f"  Warning: Could not load jobs from database: {e}")
        jobs = {}
    finally:
        conn.close()

def save_jobs():
    """Persist all jobs to SQLite (upsert)."""
    with _jobs_lock:
        conn = _get_db()
        try:
            for jid, j in jobs.items():
                safe = {k: v for k, v in j.items() if k != "log"}
                conn.execute(
                    """INSERT INTO jobs (id, data, status, client_name, created, updated, review_stage, stage_updated)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                       ON CONFLICT(id) DO UPDATE SET
                           data = excluded.data, status = excluded.status,
                           client_name = excluded.client_name, updated = excluded.updated,
                           review_stage = excluded.review_stage, stage_updated = excluded.stage_updated""",
                    (jid, json.dumps(safe, default=str), j.get("status", ""),
                     j.get("client_name", ""), j.get("created", ""),
                     datetime.now().isoformat(),
                     j.get("review_stage", "draft"),
                     j.get("stage_updated", ""))
                )
            conn.commit()
        except sqlite3.Error as e:
            print(f"  Warning: Could not save jobs to database: {e}")
        finally:
            conn.close()

def _set_review_stage(job_id, new_stage, user=None):
    """Update review stage for a job in both memory and DB. Logs the transition."""
    job = jobs.get(job_id)
    if not job:
        return False
    old_stage = job.get("review_stage", "draft")
    now = datetime.now().isoformat()
    job["review_stage"] = new_stage
    job["stage_updated"] = now
    save_jobs()

    # Log the transition
    log_event("info", "review_stage_advanced",
              f"Job {job_id[:12]}: {STAGE_DISPLAY.get(old_stage, old_stage)} → {STAGE_DISPLAY.get(new_stage, new_stage)}",
              user_id=user["id"] if user else None,
              job_id=job_id,
              details={"old_stage": old_stage, "new_stage": new_stage})

    # CAS: Compute drift when job reaches final stage (never crashes workflow)
    if new_stage == "final":
        try:
            ts = _get_telemetry_store()
            if ts:
                ts.compute_drift_for_job(job_id)
                # T-CAS-2B: Check drift thresholds after computation
                try:
                    drift_check = ts.check_drift_thresholds()
                    if drift_check and drift_check.get("triggered"):
                        _maybe_create_cr(
                            source="drift", severity="WARNING",
                            trigger_summary=f"Drift threshold exceeded: {len(drift_check['violations'])} violations",
                            trigger_snapshot={"drift_latest": ts.get_drift_summary(limit=1)},
                            findings=drift_check["violations"],
                        )
                except Exception:
                    pass
        except Exception:
            pass

    return True

def _backfill_verified_fields():
    """One-time backfill: populate verified_fields from existing verifications blobs.

    Runs once on first startup after the verified_fields table is added.
    Resolves original values from extraction logs on disk.
    """
    conn = _get_db()
    try:
        vf_count = conn.execute("SELECT COUNT(*) FROM verified_fields").fetchone()[0]
        if vf_count > 0:
            return  # Already populated

        rows = conn.execute("SELECT job_id, data FROM verifications").fetchall()
        if not rows:
            return

        migrated = 0
        for job_id, data_json in rows:
            try:
                vdata = json.loads(data_json)
            except json.JSONDecodeError:
                continue

            fields_dict = vdata.get("fields", {})
            if not fields_dict:
                continue

            # Load extraction log for this job to get original values
            job = jobs.get(job_id)
            log_data = None
            if job:
                log_path = job.get("output_log")
                if log_path and os.path.exists(log_path):
                    try:
                        with open(log_path) as f:
                            log_data = json.load(f)
                    except (json.JSONDecodeError, IOError):
                        pass

            for field_key, decision in fields_dict.items():
                status = decision.get("status", "")
                if status not in ("confirmed", "corrected", "flagged"):
                    continue

                canonical = None
                original = None

                if log_data:
                    # Inline resolution of field value from extraction log
                    parts = field_key.split(":")
                    if len(parts) == 3:
                        try:
                            pn = int(parts[0])
                            ei = int(parts[1])
                            fn = parts[2]
                            page_exts = [e for e in log_data.get("extractions", []) if e.get("_page") == pn]
                            if ei < len(page_exts):
                                fdata = page_exts[ei].get("fields", {}).get(fn)
                                if fdata is not None:
                                    original = fdata.get("value") if isinstance(fdata, dict) else fdata
                        except (ValueError, TypeError):
                            pass

                if status == "corrected":
                    canonical = decision.get("corrected_value")
                elif status == "confirmed":
                    canonical = original
                # flagged: canonical stays None

                conn.execute(
                    """INSERT OR IGNORE INTO verified_fields
                       (job_id, field_key, canonical_value, original_value, status,
                        category, vendor_desc, note, reviewer, verified_at)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (job_id, field_key,
                     json.dumps(canonical) if canonical is not None else None,
                     json.dumps(original) if original is not None else None,
                     status,
                     decision.get("category", ""),
                     decision.get("vendor_desc", ""),
                     decision.get("note", ""),
                     decision.get("reviewer", vdata.get("reviewer", "")),
                     decision.get("timestamp", vdata.get("updated", datetime.now().isoformat())))
                )
                migrated += 1

        conn.commit()
        if migrated:
            print(f"  Backfilled {migrated} verified fields from verifications blob")
    except sqlite3.Error as e:
        print(f"  Warning: Could not backfill verified_fields: {e}")
    finally:
        conn.close()


def _normalize_payer_key(ext):
    """Derive a stable payer key from an extraction dict.

    Uses EIN when available (prefixed 'ein:'), else normalized payer name (prefixed 'name:').
    The prefix prevents collisions between EIN strings and name strings.
    """
    # Try top-level payer_ein first
    ein = ext.get("payer_ein") or ""
    if not ein:
        # Try inside fields
        fields = ext.get("fields", {})
        for ek in ("employer_ein", "partnership_ein", "payer_ein"):
            fdata = fields.get(ek)
            if fdata:
                ein = fdata.get("value") if isinstance(fdata, dict) else fdata
                if ein:
                    break
    # Clean EIN: keep digits and dashes
    if ein:
        ein = re.sub(r'[^0-9\-]', '', str(ein))
        if len(ein) >= 5:
            return f"ein:{ein}"

    # Fallback: normalize payer name
    name = str(ext.get("payer_or_entity") or "unknown").upper().strip()
    # Strip common suffixes
    for suffix in (", LLC", " LLC", ", INC", " INC", ", CORP", " CORP",
                   ", LP", " LP", ", LLP", " LLP", " CO", ", CO",
                   " COMPANY", ", COMPANY", " CORPORATION", ", CORPORATION"):
        if name.endswith(suffix):
            name = name[:-len(suffix)].rstrip(" ,")
    # Strip trailing store numbers like #1234
    name = re.sub(r'\s*#\d+$', '', name)
    name = re.sub(r'\s+', ' ', name).strip()
    return f"name:{name}" if name else "name:UNKNOWN"


def _backfill_client_canonicals():
    """One-time backfill: populate client_canonical_values from existing verified_fields.

    For each verified field with status confirmed/corrected, resolves its extraction's
    document_type, payer, and field name, then upserts into the client-level store.
    """
    conn = _get_db()
    try:
        cc_count = conn.execute("SELECT COUNT(*) FROM client_canonical_values").fetchone()[0]
        if cc_count > 0:
            return  # Already populated

        # Get all verified fields that have a canonical value
        rows = conn.execute(
            "SELECT job_id, field_key, canonical_value, original_value, status, reviewer, verified_at "
            "FROM verified_fields WHERE status IN ('confirmed', 'corrected') AND canonical_value IS NOT NULL"
        ).fetchall()
        if not rows:
            return

        migrated = 0
        # Cache loaded logs to avoid re-reading the same file for each field
        log_cache = {}

        for job_id, field_key, canon_json, orig_json, status, reviewer, verified_at in rows:
            job = jobs.get(job_id)
            if not job:
                continue
            client_name = job.get("client_name", "")
            year = job.get("year", "")
            if not client_name or not year:
                continue

            # Load log (cached) — inline to avoid dependency on _load_extraction_log
            if job_id not in log_cache:
                log_path = job.get("output_log")
                if log_path and os.path.exists(log_path):
                    try:
                        with open(log_path) as f:
                            log_cache[job_id] = json.load(f)
                    except (json.JSONDecodeError, IOError):
                        log_cache[job_id] = None
                else:
                    log_cache[job_id] = None
            log_data = log_cache[job_id]
            if not log_data:
                continue

            # Resolve extraction for this field
            parts = field_key.split(":")
            if len(parts) != 3:
                continue
            try:
                page_num = int(parts[0])
                ext_idx = int(parts[1])
                fn = parts[2]
            except ValueError:
                continue

            page_exts = [e for e in log_data.get("extractions", []) if e.get("_page") == page_num]
            if ext_idx >= len(page_exts):
                continue
            ext = page_exts[ext_idx]
            doc_type = ext.get("document_type", "")
            if not doc_type:
                continue

            payer_key = _normalize_payer_key(ext)
            payer_display = ext.get("payer_or_entity", "")

            canonical = None
            if canon_json is not None:
                try:
                    canonical = json.loads(canon_json)
                except (json.JSONDecodeError, ValueError):
                    canonical = canon_json

            original = None
            if orig_json is not None:
                try:
                    original = json.loads(orig_json)
                except (json.JSONDecodeError, ValueError):
                    original = orig_json

            conn.execute(
                """INSERT OR IGNORE INTO client_canonical_values
                   (client_name, year, document_type, payer_key, payer_display,
                    field_name, canonical_value, original_value, status,
                    source_job_id, reviewer, verified_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (client_name, year, doc_type, payer_key, payer_display,
                 fn,
                 json.dumps(canonical) if canonical is not None else None,
                 json.dumps(original) if original is not None else None,
                 status, job_id, reviewer or "", verified_at or "")
            )
            migrated += 1

        conn.commit()
        if migrated:
            print(f"  Backfilled {migrated} client canonical values from verified fields")
    except sqlite3.Error as e:
        print(f"  Warning: Could not backfill client_canonical_values: {e}")
    finally:
        conn.close()


_init_db()
_seed_default_users()
load_jobs()
_backfill_verified_fields()
_backfill_client_canonicals()

# ─── Chart of Accounts + Vendor Memory ────────────────────────────────────────

CHART_OF_ACCOUNTS = {
    "Expense": [
        "Advertising & Marketing",
        "Auto & Travel",
        "Bank Service Charges",
        "Computer & Internet",
        "Depreciation",
        "Dues & Subscriptions",
        "Equipment",
        "Insurance",
        "Interest Expense",
        "Legal & Professional",
        "Meals & Entertainment",
        "Office Supplies",
        "Payroll Expenses",
        "Rent",
        "Repairs & Maintenance",
        "Taxes & Licenses",
        "Telephone",
        "Utilities",
        "Miscellaneous Expense",
    ],
    "Income": [
        "Service Revenue",
        "Product Sales",
        "Interest Income",
        "Rental Income",
        "Refund / Rebate",
        "Other Income",
    ],
    "Other": [
        "Owner Draw / Distribution",
        "Owner Contribution / Investment",
        "Loan Proceeds",
        "Loan Payment",
        "Transfer Between Accounts",
    ],
}

# Flat list for validation
ALL_ACCOUNTS = []
for _grp in CHART_OF_ACCOUNTS.values():
    ALL_ACCOUNTS.extend(_grp)

def _normalize_vendor(desc):
    """Normalize a vendor/payee name for matching.
    'GEORGIA POWER COMPANY #12345' → 'GEORGIA POWER'
    'WAL-MART SUPER CENTER 0423' → 'WAL-MART SUPER CENTER'
    """
    if not desc:
        return ""
    s = str(desc).upper().strip()
    # Strip trailing reference/store numbers
    s = re.sub(r'[\s#*]+\d{2,}$', '', s)
    # Strip common suffixes
    s = re.sub(r'\s+(LLC|INC|CORP|CO|COMPANY|LTD|LP|NA|N\.A\.)\.?\s*$', '', s, flags=re.IGNORECASE)
    # Strip trailing punctuation
    s = s.rstrip(' .,;:*#-')
    return s.strip()

def _load_vendor_categories():
    conn = _get_db()
    try:
        rows = conn.execute("SELECT vendor, data FROM vendor_categories").fetchall()
        result = {}
        for vendor, data_json in rows:
            try:
                result[vendor] = json.loads(data_json)
            except json.JSONDecodeError:
                pass
        return result
    except sqlite3.Error:
        return {}
    finally:
        conn.close()

def _save_vendor_categories(data):
    conn = _get_db()
    try:
        for vendor, info in data.items():
            cat = info.get("category", "") if isinstance(info, dict) else str(info)
            conn.execute(
                """INSERT INTO vendor_categories (vendor, category, data, updated) VALUES (?, ?, ?, ?)
                   ON CONFLICT(vendor) DO UPDATE SET
                       category = excluded.category, data = excluded.data, updated = excluded.updated""",
                (vendor, cat, json.dumps(info, default=str), datetime.now().isoformat())
            )
        conn.commit()
    except sqlite3.Error as e:
        print(f"  Warning: Could not save vendor categories: {e}")
    finally:
        conn.close()

def _learn_vendor_category(vendor_desc, category):
    """Record that vendor_desc was categorized as category."""
    if not vendor_desc or not category:
        return
    norm = _normalize_vendor(vendor_desc)
    if not norm or len(norm) < 2:
        return
    data = _load_vendor_categories()
    existing = data.get(norm, {})
    data[norm] = {
        "category": category,
        "count": existing.get("count", 0) + 1,
        "last_used": datetime.now().isoformat(),
        "original": vendor_desc,  # keep one raw example
    }
    _save_vendor_categories(data)

def _suggest_category(vendor_desc):
    """Look up a vendor in the learned map. Returns category or ''."""
    if not vendor_desc:
        return ""
    norm = _normalize_vendor(vendor_desc)
    if not norm:
        return ""
    data = _load_vendor_categories()
    entry = data.get(norm)
    if entry:
        return entry.get("category", "")
    # Try prefix match (e.g., "WALMART SUPERCENTER" matches "WALMART")
    for known, info in data.items():
        if norm.startswith(known) or known.startswith(norm):
            return info.get("category", "")
    return ""

# ─── Prior-Year Context Engine ────────────────────────────────────────────────

def _safe_client_name(name):
    """Sanitize client name for filesystem use."""
    safe = re.sub(r'[^\w\s\-\.,()]', '', name).strip() or "Unknown Client"
    return safe.title()

def _client_info_path(name):
    """Path to client metadata JSON."""
    return CLIENTS_DIR / _safe_client_name(name) / "client_info.json"

def _load_client_info(name):
    """Load client metadata or return None."""
    p = _client_info_path(name)
    if p.exists():
        try:
            with open(p) as f:
                return json.load(f)
        except Exception:
            pass
    return None

def _save_client_info(name, info):
    """Save client metadata JSON."""
    p = _client_info_path(name)
    p.parent.mkdir(parents=True, exist_ok=True)
    info["updated"] = datetime.now().isoformat()
    with open(p, "w") as f:
        json.dump(info, f, indent=2)

def _context_dir(client_name):
    """Get or create the context directory for a client."""
    d = CLIENTS_DIR / _safe_client_name(client_name) / "context"
    d.mkdir(parents=True, exist_ok=True)
    return d

def _context_index_path(client_name):
    return _context_dir(client_name) / "index.json"

def _load_context_index(client_name):
    p = _context_index_path(client_name)
    if p.exists():
        try:
            with open(p) as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    return {"documents": [], "prior_year_data": {}, "updated": None}

def _save_context_index(client_name, data):
    data["updated"] = datetime.now().isoformat()
    with open(_context_index_path(client_name), "w") as f:
        json.dump(data, f, indent=2, default=str)

def _parse_context_document(file_path, doc_label=""):
    """Parse a context document (PDF, XLSX, TXT) into structured payer/amount data.

    Returns a dict with:
      payers: [{"name": ..., "ein": ..., "form_type": ..., "amounts": {...}}]
      raw_text: str (for instructions/notes)
      year: str or None
    """
    ext = Path(file_path).suffix.lower()
    result = {"payers": [], "raw_text": "", "year": None, "source": doc_label or Path(file_path).name}

    if ext == ".txt":
        try:
            with open(file_path, encoding="utf-8", errors="replace") as f:
                result["raw_text"] = f.read()
        except IOError:
            pass
        return result

    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            rows = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    row_data = [str(c).strip() if c is not None else "" for c in row]
                    if any(cell for cell in row_data):
                        rows.append(row_data)
            wb.close()

            # Try to extract payer/amount pairs from tabular data
            text = "\n".join(["\t".join(r) for r in rows])
            result["raw_text"] = text
            result["payers"] = _extract_payers_from_text(text)
        except Exception as e:
            result["raw_text"] = f"(Error reading xlsx: {e})"
        return result

    if ext == ".pdf":
        try:
            # Use Tesseract OCR via the same method as extract.py
            _cfp_kwargs = {"dpi": 200, "fmt": "jpeg"}
            if _POPPLER_PATH:
                _cfp_kwargs["poppler_path"] = _POPPLER_PATH
            images = convert_from_path(str(file_path), **_cfp_kwargs)
            import pytesseract
            text_parts = []
            for img in images:
                text_parts.append(pytesseract.image_to_string(img))
            full_text = "\n".join(text_parts)
            result["raw_text"] = full_text
            result["payers"] = _extract_payers_from_text(full_text)
        except Exception as e:
            result["raw_text"] = f"(Error reading PDF: {e})"
        return result

    return result

def _extract_payers_from_text(text):
    """Extract payer names, EINs, form types, and dollar amounts from OCR text.

    Uses pattern matching — no LLM call. Catches common tax form patterns.
    """
    payers = []
    seen_eins = set()

    # EIN pattern: XX-XXXXXXX
    ein_pattern = re.compile(r'\b(\d{2}-\d{7})\b')
    # Dollar amounts: $1,234.56 or 1,234.56
    money_pattern = re.compile(r'\$?([\d,]+\.\d{2})\b')
    # Form type indicators
    form_patterns = {
        "W-2": re.compile(r'\bW[\s-]*2\b', re.I),
        "1099-INT": re.compile(r'\b1099[\s-]*INT\b', re.I),
        "1099-DIV": re.compile(r'\b1099[\s-]*DIV\b', re.I),
        "1099-R": re.compile(r'\b1099[\s-]*R\b', re.I),
        "1099-NEC": re.compile(r'\b1099[\s-]*NEC\b', re.I),
        "1099-MISC": re.compile(r'\b1099[\s-]*MISC\b', re.I),
        "1099-B": re.compile(r'\b1099[\s-]*B\b', re.I),
        "1099-K": re.compile(r'\b1099[\s-]*K\b', re.I),
        "K-1": re.compile(r'\bK[\s-]*1\b', re.I),
        "SSA-1099": re.compile(r'\bSSA[\s-]*1099\b', re.I),
        "1098": re.compile(r'\b1098\b'),
    }

    # Process line by line looking for EINs near entity names
    lines = text.split("\n")
    for i, line in enumerate(lines):
        eins = ein_pattern.findall(line)
        for ein in eins:
            if ein in seen_eins:
                continue
            seen_eins.add(ein)

            # Look for entity name near this EIN (same line or adjacent)
            context = line
            if i > 0:
                context = lines[i-1] + " " + context
            if i < len(lines) - 1:
                context = context + " " + lines[i+1]

            # Detect form type
            form_type = ""
            for ftype, fpat in form_patterns.items():
                if fpat.search(context):
                    form_type = ftype
                    break

            # Extract dollar amounts from context
            amounts = [float(m.replace(",", "")) for m in money_pattern.findall(context)]

            # Try to get entity name (text before or after EIN, not a number)
            parts = re.split(ein_pattern, line)
            name_candidates = [p.strip() for p in parts
                               if p.strip() and p.strip() != ein
                               and not money_pattern.match(p.strip())]
            entity_name = name_candidates[0] if name_candidates else ""
            # Clean up
            entity_name = re.sub(r'^[\s\d\-:]+', '', entity_name).strip()
            entity_name = entity_name[:80]

            payers.append({
                "name": entity_name,
                "ein": ein,
                "form_type": form_type,
                "amounts": amounts[:10],  # cap
            })

    return payers

def _build_completeness_report(client_name, current_extractions, year):
    """Compare current extractions against prior-year context.

    Returns:
      matched: [{payer, form, status: "received", current_amounts, prior_amounts}]
      missing: [{payer, form, status: "expected"}]
      new: [{payer, form, status: "new"}]
      variances: [{payer, form, field, prior, current, pct_change, severity}]
    """
    ctx = _load_context_index(client_name)
    prior_data = ctx.get("prior_year_data", {})
    if not prior_data:
        return {"matched": [], "missing": [], "new": [], "variances": []}

    # Build sets: prior payers by (ein, form_type)
    prior_set = {}
    for doc in prior_data.get("documents", []):
        for payer in doc.get("payers", []):
            key = (payer.get("ein", ""), payer.get("form_type", ""))
            if key[0]:  # only track if we have an EIN
                prior_set[key] = payer

    # Build current set from extractions
    current_set = {}
    for ext in current_extractions:
        dtype = ext.get("document_type", "")
        ein = ""
        fields = ext.get("fields", {})
        for ek in ["payer_ein", "employer_ein", "partnership_ein"]:
            v = fields.get(ek)
            if isinstance(v, dict):
                v = v.get("value", "")
            if v:
                ein = str(v)
                break
        entity = ext.get("payer_or_entity", "")
        key = (ein, dtype)
        current_set[key] = {"name": entity, "ein": ein, "form_type": dtype, "fields": fields}

    matched = []
    missing = []
    new_items = []
    variances = []

    for key, prior in prior_set.items():
        if key in current_set:
            cur = current_set[key]
            matched.append({
                "payer": cur.get("name") or prior.get("name", ""),
                "ein": key[0],
                "form": key[1],
                "status": "received",
            })
            # Check for variances on key amounts
            for pa in prior.get("amounts", []):
                # Find closest matching current amount
                for fname, fdata in cur.get("fields", {}).items():
                    cv = fdata.get("value") if isinstance(fdata, dict) else fdata
                    if isinstance(cv, (int, float)) and cv > 0 and pa > 0:
                        pct = abs(cv - pa) / pa * 100
                        if pct > 25:
                            variances.append({
                                "payer": cur.get("name", ""),
                                "form": key[1],
                                "field": fname,
                                "prior": pa,
                                "current": cv,
                                "pct_change": round(pct, 1),
                                "severity": "red" if pct > 50 else "yellow",
                            })
                        break  # one comparison per prior amount
        else:
            missing.append({
                "payer": prior.get("name", "Unknown"),
                "ein": key[0],
                "form": key[1] or "Unknown form",
                "status": "expected",
            })

    for key, cur in current_set.items():
        if key not in prior_set and key[0]:
            new_items.append({
                "payer": cur.get("name", ""),
                "ein": key[0],
                "form": key[1],
                "status": "new",
            })

    return {"matched": matched, "missing": missing, "new": new_items, "variances": variances}


# ─── Client Instruction Memory ────────────────────────────────────────────────

def _instructions_path(client_name):
    return CLIENTS_DIR / _safe_client_name(client_name) / "instructions.json"

def _load_instructions(client_name):
    p = _instructions_path(client_name)
    if p.exists():
        try:
            with open(p) as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    return {"rules": [], "updated": None}

def _save_instructions(client_name, data):
    # Ensure parent dir exists
    p = _instructions_path(client_name)
    p.parent.mkdir(parents=True, exist_ok=True)
    data["updated"] = datetime.now().isoformat()
    with open(p, "w") as f:
        json.dump(data, f, indent=2, default=str)

def _instructions_text(client_name):
    """Get all instructions as a single string for prompt injection."""
    data = _load_instructions(client_name)
    rules = data.get("rules", [])
    if not rules:
        return ""
    lines = [f"- {r['text']}" for r in rules if r.get("text")]
    return "CLIENT-SPECIFIC INSTRUCTIONS:\n" + "\n".join(lines)


# ─── Batch Categorization Helpers ─────────────────────────────────────────────

def _gather_uncategorized(job_ids=None, client_name=None):
    """Gather all transactions needing categorization across jobs.

    Returns list of:
      {job_id, page, ext_idx, txn_num, date, desc, amount, type, source,
       current_category, suggested_category, vendor_norm}
    """
    target_jobs = {}
    for jid, j in jobs.items():
        if j.get("status") != "complete":
            continue
        if job_ids and jid not in job_ids:
            continue
        if client_name and j.get("client_name", "").lower() != client_name.lower():
            continue
        target_jobs[jid] = j

    items = []
    for jid, j in target_jobs.items():
        log_path = j.get("output_log")
        if not log_path or not os.path.exists(log_path):
            continue
        try:
            with open(log_path) as f:
                log_data = json.load(f)
        except (json.JSONDecodeError, IOError):
            continue

        vdata = _load_verifications(jid)
        vfields = vdata.get("fields", {})

        for ext in log_data.get("extractions", []):
            dtype = str(ext.get("document_type", ""))
            fields = ext.get("fields", {})
            entity = ext.get("payer_or_entity", "")
            page = ext.get("_page", 0)

            # Find ext index for this page
            page_exts = [e for e in log_data.get("extractions", []) if e.get("_page") == page]
            ext_idx = 0
            for ei, pe in enumerate(page_exts):
                if pe is ext:
                    ext_idx = ei
                    break

            # Bank/CC transactions
            if "bank_statement" in dtype or "credit_card" in dtype:
                bank = ""
                for k in ["bank_name", "card_issuer"]:
                    v = fields.get(k)
                    bank = (v.get("value", "") if isinstance(v, dict) else str(v or "")) if v else ""
                    if bank:
                        break
                source = bank or entity

                txn_nums = sorted(set(
                    int(m.group(1)) for k in fields
                    for m in [re.match(r"txn_(\d+)_", k)] if m
                ))
                for n in txn_nums:
                    amt_key = f"txn_{n}_amount"
                    vk = f"{page}:{ext_idx}:{amt_key}"
                    vstate = vfields.get(vk, {})
                    current_cat = vstate.get("category", "")

                    date_f = fields.get(f"txn_{n}_date")
                    desc_f = fields.get(f"txn_{n}_desc")
                    amt_f = fields.get(amt_key)
                    type_f = fields.get(f"txn_{n}_type")

                    date_v = (date_f.get("value", "") if isinstance(date_f, dict) else str(date_f or "")) if date_f else ""
                    desc_v = (desc_f.get("value", "") if isinstance(desc_f, dict) else str(desc_f or "")) if desc_f else ""
                    amt_v = (amt_f.get("value") if isinstance(amt_f, dict) else amt_f) if amt_f else None
                    type_v = (type_f.get("value", "") if isinstance(type_f, dict) else str(type_f or "")) if type_f else ""

                    if amt_v is None:
                        continue

                    norm = _normalize_vendor(desc_v)
                    suggested = _suggest_category(desc_v) if not current_cat else ""

                    items.append({
                        "job_id": jid, "page": page, "ext_idx": ext_idx,
                        "field_key": vk, "date": date_v, "desc": desc_v,
                        "amount": amt_v, "type": type_v, "source": source,
                        "doc_type": dtype,
                        "current_category": current_cat,
                        "suggested_category": suggested,
                        "vendor_norm": norm,
                        "client_name": j.get("client_name", ""),
                    })

            # Checks
            elif dtype == "check":
                check_amt_f = fields.get("check_amount")
                amt_v = (check_amt_f.get("value") if isinstance(check_amt_f, dict) else check_amt_f) if check_amt_f else None
                if amt_v is None:
                    continue
                vk = f"{page}:{ext_idx}:check_amount"
                vstate = vfields.get(vk, {})
                current_cat = vstate.get("category", "")

                payee_f = fields.get("payee") or fields.get("pay_to")
                payee = (payee_f.get("value", "") if isinstance(payee_f, dict) else str(payee_f or "")) if payee_f else ""
                date_f = fields.get("check_date")
                date_v = (date_f.get("value", "") if isinstance(date_f, dict) else str(date_f or "")) if date_f else ""
                num_f = fields.get("check_number")
                num_v = (num_f.get("value", "") if isinstance(num_f, dict) else str(num_f or "")) if num_f else ""

                norm = _normalize_vendor(payee)
                suggested = _suggest_category(payee) if not current_cat else ""

                items.append({
                    "job_id": jid, "page": page, "ext_idx": ext_idx,
                    "field_key": vk, "date": date_v,
                    "desc": f"Check #{num_v} to {payee}" if num_v else payee,
                    "amount": amt_v, "type": "check", "source": "Check",
                    "doc_type": dtype,
                    "current_category": current_cat,
                    "suggested_category": suggested,
                    "vendor_norm": norm,
                    "client_name": j.get("client_name", ""),
                })

    return items

def auto_rotate_page(img):
    """Detect and fix rotated pages for the review viewer.

    - Landscape pages: always use OSD rotation
    - Portrait + 90/270 OSD: rotate (content is sideways in portrait frame)
    - Portrait + 180 OSD: only rotate if conf >= 10 (180 is usually false positive)"""
    w, h = img.size
    is_landscape = w > h * 1.15
    try:
        import pytesseract
        osd = pytesseract.image_to_osd(img, output_type=pytesseract.Output.DICT)
        angle = osd.get("rotate", 0)
        conf = float(osd.get("orientation_conf", 0))
        if angle != 0:
            if is_landscape or angle in (90, 270) or conf >= 10:
                return img.rotate(-angle, expand=True)
        return img
    except Exception as e:
        if is_landscape:
            print(f"  Auto-rotate: Tesseract OSD failed ({e}), rotating 90 CW as fallback")
            return img.rotate(-90, expand=True)
    return img

def generate_page_images(job_id, pdf_path):
    """Convert PDF pages to JPEG images for the side-by-side viewer. Auto-rotates sideways pages."""
    job_pages_dir = PAGES_DIR / job_id
    job_pages_dir.mkdir(exist_ok=True)
    try:
        _cfp_kw = {"dpi": 150}
        if _POPPLER_PATH:
            _cfp_kw["poppler_path"] = _POPPLER_PATH
        images = convert_from_path(str(pdf_path), **_cfp_kw)
        for i, img in enumerate(images):
            img = auto_rotate_page(img)
            page_path = job_pages_dir / f"page_{i+1}.jpg"
            img.save(str(page_path), "JPEG", quality=80)
        return len(images)
    except Exception as e:
        print(f"Page image generation error: {e}")
        return 0

# ─── Job Runner ───────────────────────────────────────────────────────────────

def run_extraction(job_id, pdf_path, year, skip_verify, doc_type="tax_returns", output_format="tax_review", user_notes="", ai_instructions="", disable_pii=False, resume=False, use_ocr_first=False):
    """Run extract.py in a background thread, capturing progress line by line."""
    import subprocess
    job = jobs[job_id]
    job["status"] = "running"
    job["log"] = []
    job["stage"] = "starting"
    job["progress"] = 0
    job["start_time"] = datetime.now().isoformat()

    # Sprint 2: Log job start
    log_event("info", "job_started",
              f"Job started: {job.get('filename', '')} for {job.get('client_name', '')}",
              job_id=job_id,
              details={"doc_type": doc_type, "year": year})

    # CAS Hook 1: Record run start (never crashes extraction)
    try:
        ts = _get_telemetry_store()
        if ts:
            ts.record_run_start(job_id, client_name=job.get("client_name", ""),
                                doc_type=doc_type, app_version=_app_version,
                                extract_version="v6")
    except Exception:
        pass

    # Generate page images for the side-by-side viewer
    job["stage"] = "rendering"
    job["progress"] = 2
    job["log"].append("Rendering PDF pages for review...")
    num_pages = generate_page_images(job_id, pdf_path)
    job["total_pages"] = num_pages
    job["log"].append(f"  {num_pages} pages rendered")

    # Build command
    output_name = Path(pdf_path).stem + "_intake.xlsx"
    output_path = OUTPUT_DIR / output_name
    log_name = Path(pdf_path).stem + "_intake_log.json"
    log_path = OUTPUT_DIR / log_name

    # Map doc types to extract.py-compatible values
    # (extract.py only accepts: tax_returns, bank_statements, trust_documents, bookkeeping)
    EXTRACT_DOC_TYPE_MAP = {
        "payroll": "bookkeeping",
        "other": "bookkeeping",
    }
    extract_doc_type = EXTRACT_DOC_TYPE_MAP.get(doc_type, doc_type)

    cmd = [sys.executable, str(BASE_DIR / "extract.py"), str(pdf_path),
           "--year", str(year), "--output", str(output_path),
           "--doc-type", extract_doc_type, "--output-format", output_format]
    if skip_verify:
        cmd.append("--skip-verify")
    if disable_pii:
        cmd.append("--no-pii")
    if resume:
        cmd.append("--resume")
    if not use_ocr_first:
        cmd.append("--no-ocr-first")

    # Inject client instructions into ai_instructions
    client_name = job.get("client_name", "")
    instr_text = _instructions_text(client_name) if client_name else ""
    combined_instructions = ai_instructions
    if instr_text:
        combined_instructions = (combined_instructions + "\n\n" + instr_text) if combined_instructions else instr_text

    if user_notes:
        cmd.extend(["--user-notes", user_notes])
    if combined_instructions:
        cmd.extend(["--ai-instructions", combined_instructions])

    # Pass prior-year context if available
    if client_name:
        ctx_idx_path = _context_index_path(client_name)
        if ctx_idx_path.exists():
            cmd.extend(["--context-file", str(ctx_idx_path)])

    try:
        proc = subprocess.Popen(
            cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, bufsize=1, cwd=str(BASE_DIR)
        )
        _active_procs[job_id] = proc

        for line in proc.stdout:
            line = line.rstrip()
            if not line:
                continue
            job["log"].append(line)

            # Parse progress from extract.py output
            ll = line.lower()
            if "converting pdf" in ll:
                job["stage"] = "scanning"
                job["progress"] = 5
            elif "preprocessing" in ll and "pages" in ll:
                job["stage"] = "preprocessing"
                job["progress"] = 6
            elif "blank pages detected" in ll:
                job["progress"] = 7
            elif "pages converted" in ll:
                job["progress"] = 7
            elif "ocr pass" in ll:
                job["stage"] = "ocr"
                job["progress"] = 8
            elif "ocr success" in ll:
                job["progress"] = 11
            elif "cache hit" in ll:
                job["stage"] = "cache_restore"
                job["progress"] = 12
            elif "per-page routing" in ll:
                job["stage"] = "routing"
                job["progress"] = 12
            elif "routing summary" in ll:
                job["progress"] = 13
            elif "phase 1:" in ll or "classification" in ll or "classify" in ll:
                job["stage"] = "classifying"
                job["progress"] = 15
            elif "section detection" in ll:
                job["stage"] = "section_detect"
                job["progress"] = 17
            elif "section summary" in ll:
                job["progress"] = 18
            elif "document group" in ll:
                job["progress"] = 25
            elif "phase 2:" in ll or "── extraction" in ll:
                job["stage"] = "extracting"
                job["progress"] = 30
            elif "ocr sufficient" in ll or "text extraction" in ll:
                # OCR-first path saved a vision call
                job["progress"] = min(job["progress"] + 3, 72)
            elif ("page" in ll and "extracted" in ll) or ("multi-page extracted" in ll):
                job["progress"] = min(job["progress"] + 3, 72)
            elif ll.startswith("batch_complete:"):
                parts = ll.split(":")
                if len(parts) >= 4:
                    try:
                        bn, tb, fc = int(parts[1]), int(parts[2]), int(parts[3])
                        job["stage"] = "extracting"
                        job["progress"] = 30 + int(45 * bn / max(tb, 1))
                        job["batches_complete"] = bn
                        job["total_batches"] = tb
                        job["fields_streamed"] = fc
                        job["partial_results_ready"] = True
                    except (ValueError, IndexError):
                        pass
            elif ll.startswith("first_values_ready:"):
                try:
                    job["time_to_first_values_s"] = float(ll.split(":")[1])
                    job["partial_results_ready"] = True
                except (ValueError, IndexError):
                    pass
            elif ll.startswith("finalize_complete"):
                job["progress"] = 75
                job["stage"] = "finalizing"
            elif "extraction stats" in ll:
                job["progress"] = 75
            elif "phase 2.5" in ll or ("consensus" in ll and "verification" in ll):
                job["stage"] = "consensus"
                job["progress"] = 76
            elif "consensus:" in ll and ("auto_verified" in ll or "needs_review" in ll):
                job["progress"] = 77
            elif "phase 3:" in ll or "── verification" in ll:
                job["stage"] = "verifying"
                job["progress"] = 78
            elif "corrected:" in ll:
                job["progress"] = min(job["progress"] + 1, 88)
            elif "phase 4:" in ll or "normalize" in ll:
                job["stage"] = "normalizing"
                job["progress"] = 90
            elif "phase 5:" in ll or "validate" in ll:
                job["progress"] = 93
            elif "phase 6:" in ll or "excel" in ll:
                job["stage"] = "writing"
                job["progress"] = 96
            elif "est. cost:" in ll:
                # Capture cost from final summary line
                import re as _re
                m = _re.search(r'\$(\d+\.\d+)', line)
                if m:
                    job["cost_usd"] = float(m.group(1))
            elif ll.strip().startswith("complete") or ll.strip().endswith("complete"):
                job["progress"] = 100

        proc.wait()
        _active_procs.pop(job_id, None)

        if proc.returncode == 0:
            job["status"] = "complete"
            job["progress"] = 100
            job["stage"] = "done"
            job["end_time"] = datetime.now().isoformat()
            # B1: Auto-advance to preparer_review on extraction complete
            _set_review_stage(job_id, "preparer_review")
            # Sprint 2: Log job completion
            log_event("info", "job_completed",
                      f"Job completed: {job.get('client_name', '')}",
                      job_id=job_id)
            job["output_xlsx"] = str(output_path) if output_path.exists() else None
            job["output_log"] = str(log_path) if log_path.exists() else None
            if job["output_xlsx"]:
                _secure_file(job["output_xlsx"])
            if job["output_log"]:
                _secure_file(job["output_log"])

            # Copy outputs to client directory
            import shutil
            client_folder = job.get("client_folder")
            if client_folder:
                client_dir = Path(client_folder)
                client_dir.mkdir(parents=True, exist_ok=True)
                try:
                    if output_path.exists():
                        dst_xlsx = client_dir / output_path.name
                        shutil.copy2(str(output_path), str(dst_xlsx))
                        _secure_file(dst_xlsx)
                        job["client_xlsx"] = str(dst_xlsx)
                        job["log"].append(f"  Saved to: {dst_xlsx}")
                    if log_path.exists():
                        dst_log = client_dir / log_path.name
                        shutil.copy2(str(log_path), str(dst_log))
                        _secure_file(dst_log)
                        job["client_log"] = str(dst_log)
                    # Copy the original PDF too (use original filename for client folder)
                    src_pdf = Path(pdf_path)
                    if src_pdf.exists():
                        original_name = job.get("filename", src_pdf.name)
                        safe_original = re.sub(r'[^\w\s\-\.,()]', '', original_name).strip() or src_pdf.name
                        dst_pdf = client_dir / safe_original
                        if not dst_pdf.exists():
                            shutil.copy2(str(src_pdf), str(dst_pdf))
                            _secure_file(dst_pdf)
                except Exception as e:
                    job["log"].append(f"  Warning: Could not copy to client folder: {e}")

            # Parse the JSON log for summary stats
            if log_path.exists():
                try:
                    with open(log_path) as f:
                        log_data = json.load(f)
                    exts = log_data.get("extractions", [])
                    methods = {}
                    confs = {}
                    for e in exts:
                        m = e.get("_extraction_method") or "unknown"
                        methods[m] = methods.get(m, 0) + 1
                        for fv in (e.get("fields") or {}).values():
                            if isinstance(fv, dict):
                                c = fv.get("confidence") or "unknown"
                                confs[c] = confs.get(c, 0) + 1
                    job["stats"] = {
                        "documents": len(exts),
                        "methods": methods,
                        "confidences": confs,
                        "warnings": len(log_data.get("warnings", [])),
                        "total_fields": sum(confs.values()),
                    }
                    # Include cost data if present in log
                    cost = log_data.get("cost")
                    if cost:
                        job["stats"]["cost"] = cost
                        job["cost_usd"] = cost.get("estimated_cost_usd", 0)

                    # T1.6: Populate fact store from extraction results
                    try:
                        _populate_facts_from_extraction(job)
                    except Exception as fact_err:
                        job["log"].append(f"  Warning: Fact population failed: {fact_err}")

                    # T-TXN-LEDGER-1: Ingest transactions into ledger
                    try:
                        txn_store = _get_transaction_store()
                        if txn_store:
                            txn_result = txn_store.ingest_from_extraction(
                                job_id, log_data,
                                client_name=job.get("client_name", ""),
                                year=int(year) if str(year).isdigit() else 0,
                            )
                            if txn_result["inserted"] > 0:
                                job["log"].append(
                                    f"  Ledger: {txn_result['inserted']} transactions ingested"
                                )
                                # Auto-apply vendor rules
                                txn_store.apply_vendor_rules(
                                    client_name=job.get("client_name", ""),
                                    year=int(year) if str(year).isdigit() else 0,
                                )
                    except Exception as txn_err:
                        job["log"].append(f"  Warning: Transaction ingest failed: {txn_err}")

                    # CAS Hook 2: Record run completion + phase timing (never crashes extraction)
                    try:
                        ts = _get_telemetry_store()
                        if ts:
                            log_data["log_path"] = str(log_path)
                            ts.record_run_complete(job_id, log_data)
                            # Record phase timing if available
                            phase_timing = log_data.get("timing", {}).get("phases")
                            if phase_timing:
                                ts.record_phases(job_id, phase_timing)
                    except Exception:
                        pass

                except (json.JSONDecodeError, KeyError, TypeError, IOError) as e:
                    job["log"].append(f"  Warning: Could not parse log stats: {e}")
        else:
            job["status"] = "error"
            job["end_time"] = datetime.now().isoformat()
            job["error"] = f"extract.py exited with code {proc.returncode}"
            # Sprint 2: Log job failure
            log_event("error", "job_failed",
                      f"Job failed: {job.get('client_name', '')} (exit code {proc.returncode})",
                      job_id=job_id)
            # CAS Hook 3a: Record run error (never crashes extraction)
            try:
                ts = _get_telemetry_store()
                if ts:
                    ts.record_run_error(job_id, f"exit code {proc.returncode}")
                    # T-CAS-2B: Check 24h error rate after error
                    try:
                        err_data = ts.get_error_rate_24h()
                        if err_data and err_data.get("error_rate", 0) > 0.20:
                            _maybe_create_cr(
                                source="error_rate", severity="CRITICAL",
                                trigger_summary=f"Extraction error rate {err_data['error_rate']*100:.1f}% exceeds 20% (last 24h)",
                                trigger_snapshot=err_data,
                                findings=[{
                                    "severity": "CRITICAL",
                                    "source": "error_rate",
                                    "check_name": "extraction_error_rate_24h",
                                    "details": f"{err_data['error_runs']}/{err_data['total_runs']} runs errored in last 24h",
                                    "measured_value": f"{err_data['error_rate']*100:.1f}%",
                                    "threshold": "20%",
                                    "recommended_action": "Review recent error logs and fix extraction failures",
                                }],
                            )
                    except Exception:
                        pass
            except Exception:
                pass

    except Exception as e:
        job["status"] = "error"
        job["end_time"] = datetime.now().isoformat()
        job["error"] = str(e)
        _active_procs.pop(job_id, None)
        # Sprint 2: Log job error
        log_event("error", "job_failed",
                  f"Job error: {job.get('client_name', '')} — {e}",
                  job_id=job_id)
        # CAS Hook 3b: Record run error (never crashes extraction)
        try:
            ts = _get_telemetry_store()
            if ts:
                ts.record_run_error(job_id, str(e)[:500])
        except Exception:
            pass

    save_jobs()


# ─── Sprint 2: Login / Logout Routes ─────────────────────────────────────────

@app.route("/login", methods=["GET"])
def login_page():
    users = list_active_users()
    return render_template("login.html", users=users, error=None)


@app.route("/login", methods=["POST"])
def login_post():
    import time as _time
    username = request.form.get("username", "").strip().lower()
    pin = request.form.get("pin", "").strip()
    ip = request.remote_addr or "unknown"
    key = (username, ip)
    st = _failed_logins.get(key, {"count": 0, "locked_until": 0})
    now_epoch = int(_time.time())

    # Lockout check
    if st["locked_until"] > now_epoch:
        log_event("warn", "login_lockout",
                  f"Login locked out for {username}", ip_addr=ip)
        return render_template("login.html", users=list_active_users(),
                               error="Too many attempts. Try again shortly.")

    u = get_user_by_username(username)
    if not u or not u["is_active"] or not pin.isdigit() or len(pin) != 6:
        st["count"] += 1
        if st["count"] >= MAX_FAILED_ATTEMPTS:
            st["locked_until"] = now_epoch + LOGIN_LOCKOUT_SECONDS
        _failed_logins[key] = st
        log_event("warn", "login_failed",
                  f"Login failed for {username}",
                  user_id=u["id"] if u else None, ip_addr=ip)
        return render_template("login.html", users=list_active_users(),
                               error="Invalid credentials.")

    if not check_password_hash(u["pin_hash"], pin):
        st["count"] += 1
        if st["count"] >= MAX_FAILED_ATTEMPTS:
            st["locked_until"] = now_epoch + LOGIN_LOCKOUT_SECONDS
        _failed_logins[key] = st
        log_event("warn", "login_failed",
                  f"Login failed for {username}",
                  user_id=u["id"], ip_addr=ip)
        return render_template("login.html", users=list_active_users(),
                               error="Invalid credentials.")

    # Successful login
    _failed_logins.pop(key, None)
    session.clear()
    session["user_id"] = u["id"]
    session["last_seen"] = now_epoch
    update_last_login(u["id"])
    log_event("info", "login_success",
              f"{u['display_name']} logged in",
              user_id=u["id"], ip_addr=ip)

    # SEC-006: Force PIN reset if flagged
    if u.get("must_reset_pin"):
        session["force_pin_reset"] = True
        return redirect("/change-pin", code=303)

    return redirect("/admin", code=303)


@app.route("/logout", methods=["POST"])
@require_login
def logout():
    u = current_user()
    if u:
        log_event("info", "logout",
                  f"{u['display_name']} logged out",
                  user_id=u["id"],
                  ip_addr=request.remote_addr or "")
    session.clear()
    return redirect("/login", code=303)


# ─── SEC-006: Forced PIN Change ──────────────────────────────────────────────

@app.route("/change-pin", methods=["GET"])
@require_login
def change_pin_page():
    """Show PIN change form. Users flagged with must_reset_pin are redirected here."""
    u = current_user()
    forced = session.get("force_pin_reset", False)
    return render_template_string(_CHANGE_PIN_TEMPLATE,
                                  user=u, forced=forced, error=None, success=None)


@app.route("/change-pin", methods=["POST"])
@require_login
def change_pin_submit():
    """Process PIN change. Validates complexity and clears must_reset_pin flag."""
    u = current_user()
    forced = session.get("force_pin_reset", False)
    current_pin = request.form.get("current_pin", "").strip()
    new_pin = request.form.get("new_pin", "").strip()
    confirm_pin = request.form.get("confirm_pin", "").strip()

    # Verify current PIN
    if not check_password_hash(u["pin_hash"], current_pin):
        return render_template_string(_CHANGE_PIN_TEMPLATE,
                                      user=u, forced=forced,
                                      error="Current PIN is incorrect.", success=None)

    # Check new PINs match
    if new_pin != confirm_pin:
        return render_template_string(_CHANGE_PIN_TEMPLATE,
                                      user=u, forced=forced,
                                      error="New PINs don't match.", success=None)

    # SEC-006: Complexity check
    complexity_err = validate_pin_complexity(new_pin)
    if complexity_err:
        return render_template_string(_CHANGE_PIN_TEMPLATE,
                                      user=u, forced=forced,
                                      error=complexity_err, success=None)

    # Don't allow reuse of same PIN
    if check_password_hash(u["pin_hash"], new_pin):
        return render_template_string(_CHANGE_PIN_TEMPLATE,
                                      user=u, forced=forced,
                                      error="New PIN must be different from current PIN.", success=None)

    # Apply change
    new_hash = generate_password_hash(new_pin)
    set_user_pin_hash(u["id"], new_hash)
    clear_must_reset_pin(u["id"])
    session.pop("force_pin_reset", None)
    log_event("info", "pin_change",
              f"{u['display_name']} changed their PIN",
              user_id=u["id"], ip_addr=request.remote_addr or "")

    if forced:
        return redirect("/admin", code=303)
    return render_template_string(_CHANGE_PIN_TEMPLATE,
                                  user=u, forced=False,
                                  error=None, success="PIN changed successfully.")


# Inline template for PIN change page (avoids needing a separate HTML file)
_CHANGE_PIN_TEMPLATE = """<!DOCTYPE html>
<html><head><title>Change PIN</title>
<style>
body { font-family: 'Georgia', serif; background: #1a1a2e; color: #e8e0d0; display: flex; justify-content: center; align-items: center; min-height: 100vh; margin: 0; }
.card { background: #16213e; border: 1px solid #333; border-radius: 8px; padding: 2rem; max-width: 400px; width: 100%; }
h2 { color: #e2a73b; margin-top: 0; }
.forced { background: #5c3a08; border: 1px solid #e2a73b; padding: 0.75rem; border-radius: 4px; margin-bottom: 1rem; font-size: 0.9rem; }
.error { color: #ff6b6b; margin-bottom: 1rem; }
.success { color: #51cf66; margin-bottom: 1rem; }
label { display: block; margin-top: 1rem; font-size: 0.85rem; color: #a0a0a0; }
input[type=password] { width: 100%; padding: 0.5rem; margin-top: 0.25rem; background: #0f3460; color: #e8e0d0; border: 1px solid #444; border-radius: 4px; font-size: 1.1rem; letter-spacing: 0.3em; text-align: center; box-sizing: border-box; }
button { margin-top: 1.5rem; width: 100%; padding: 0.6rem; background: #e2a73b; color: #1a1a2e; border: none; border-radius: 4px; font-weight: bold; cursor: pointer; font-size: 1rem; }
button:hover { background: #d4953a; }
.hint { font-size: 0.75rem; color: #777; margin-top: 0.25rem; }
</style></head>
<body><div class="card">
<h2>Change PIN</h2>
{% if forced %}<div class="forced">Your PIN must be changed before continuing.</div>{% endif %}
{% if error %}<div class="error">{{ error }}</div>{% endif %}
{% if success %}<div class="success">{{ success }}</div>{% endif %}
<form method="POST">
<label>Current PIN</label><input type="password" name="current_pin" maxlength="6" pattern="[0-9]{6}" required autofocus>
<label>New PIN</label><input type="password" name="new_pin" maxlength="6" pattern="[0-9]{6}" required>
<div class="hint">6 digits, no repeated or sequential patterns</div>
<label>Confirm New PIN</label><input type="password" name="confirm_pin" maxlength="6" pattern="[0-9]{6}" required>
<button type="submit">Change PIN</button>
</form>
{% if not forced %}<p style="text-align:center;margin-top:1rem;"><a href="/admin" style="color:#e2a73b;">Back to Dashboard</a></p>{% endif %}
</div></body></html>"""


# ─── Sprint 2: Admin Routes ─────────────────────────────────────────────────

@app.route("/admin")
@require_login
def admin_home():
    u = current_user()
    if u["role"] not in ("admin", "partner"):
        abort(403)
    summary = build_admin_summary()
    return render_template("admin_home.html",
        active="overview", header="Overview",
        current_user=u,
        version=summary["health"]["version"],
        uptime_h=summary["health"]["uptime_h"],
        health_state=summary["health"]["state"],
        health_label=summary["health"]["label"],
        kpis=summary["kpis"],
        recent_jobs=summary["recent_jobs"],
        recent_events=summary["recent_events"],
    )


@app.route("/admin/events")
@require_login
def admin_events():
    u = current_user()
    if u["role"] not in ("admin", "partner"):
        abort(403)
    level = request.args.get("level") or None
    job_id = request.args.get("job_id") or None
    user_id = request.args.get("user_id") or None
    events = query_events(level=level, job_id=job_id, user_id=user_id, limit=200)
    summary = build_admin_summary()
    return render_template("admin_events.html",
        active="events", header="Events",
        current_user=u,
        level=level, job_id=job_id, user_id=user_id,
        events=events,
        version=summary["health"]["version"],
        uptime_h=summary["health"]["uptime_h"],
        health_state=summary["health"]["state"],
        health_label=summary["health"]["label"],
    )


@app.route("/admin/users")
@require_login
@require_role("admin")
def admin_users():
    u = current_user()
    users = list_all_users()
    summary = build_admin_summary()
    return render_template("admin_users.html",
        active="users", header="Users",
        current_user=u, users=users, temp_pin=None,
        version=summary["health"]["version"],
        uptime_h=summary["health"]["uptime_h"],
        health_state=summary["health"]["state"],
        health_label=summary["health"]["label"],
    )


@app.route("/admin/users/create", methods=["POST"])
@require_login
@require_role("admin")
def admin_create_user():
    username = request.form.get("username", "").strip().lower()
    display_name = request.form.get("display_name", "").strip()
    role = request.form.get("role", "reviewer").strip()

    if not username or not display_name:
        abort(400)
    if role not in VALID_ROLES:
        abort(400)

    # Check for duplicate
    existing = get_user_by_username(username)
    if existing:
        summary = build_admin_summary()
        return render_template("admin_users.html",
            active="users", header="Users",
            current_user=current_user(), users=list_all_users(),
            temp_pin=None,
            version=summary["health"]["version"],
            uptime_h=summary["health"]["uptime_h"],
            health_state=summary["health"]["state"],
            health_label=summary["health"]["label"],
            error=f"Username '{username}' already exists.",
        )

    temp_pin = generate_6_digit_pin()
    pin_hash = generate_password_hash(temp_pin)
    new_id = create_user(username, display_name, role, pin_hash)
    log_event("info", "user_create",
              f"Created user '{display_name}' ({username}, {role})",
              user_id=current_user()["id"],
              details={"new_user_id": new_id, "username": username, "role": role})

    summary = build_admin_summary()
    return render_template("admin_users.html",
        active="users", header="Users",
        current_user=current_user(), users=list_all_users(),
        temp_pin=temp_pin,
        version=summary["health"]["version"],
        uptime_h=summary["health"]["uptime_h"],
        health_state=summary["health"]["state"],
        health_label=summary["health"]["label"],
    )


@app.route("/admin/users/<int:user_id>/reset_pin", methods=["POST"])
@require_login
@require_role("admin")
def admin_reset_pin(user_id):
    temp_pin = generate_6_digit_pin()
    set_user_pin_hash(user_id, generate_password_hash(temp_pin))
    # SEC-006: Flag user for forced PIN change on next login
    conn = _get_db()
    try:
        conn.execute("UPDATE users SET must_reset_pin = 1 WHERE id = ?", (user_id,))
        conn.commit()
    finally:
        conn.close()
    target = get_user_by_id(user_id)
    target_name = target["display_name"] if target else f"user_id={user_id}"
    log_event("warn", "pin_reset",
              f"PIN reset for {target_name}",
              user_id=current_user()["id"],
              details={"target_user_id": user_id})

    summary = build_admin_summary()
    return render_template("admin_users.html",
        active="users", header="Users",
        current_user=current_user(), users=list_all_users(),
        temp_pin=temp_pin,
        version=summary["health"]["version"],
        uptime_h=summary["health"]["uptime_h"],
        health_state=summary["health"]["state"],
        health_label=summary["health"]["label"],
    )


@app.route("/admin/users/<int:user_id>/toggle", methods=["POST"])
@require_login
@require_role("admin")
def admin_toggle_user(user_id):
    target = get_user_by_id(user_id)
    if not target:
        abort(404)
    new_state = not target["is_active"]
    set_user_active(user_id, new_state)
    action = "enabled" if new_state else "disabled"
    log_event("warn", f"user_{action}",
              f"User {target['display_name']} {action}",
              user_id=current_user()["id"],
              details={"target_user_id": user_id})
    return redirect("/admin/users", code=303)


@app.route("/api/admin/summary")
@require_login
def api_admin_summary():
    u = current_user()
    if u["role"] not in ("admin", "partner"):
        abort(403)
    return jsonify(build_admin_summary())


# ─── CAS: Continuous Assurance System Routes (T-CAS-1) ──────────────────────

@app.route("/api/cas/health")
@require_login
def api_cas_health():
    """Aggregated CAS health status."""
    u = current_user()
    if u["role"] not in ("admin", "partner"):
        abort(403)
    ts = _get_telemetry_store()
    if not ts:
        return jsonify({"error": "CAS not available"}), 503
    return jsonify(ts.cas_health_summary())


@app.route("/api/cas/smoke", methods=["POST"])
@require_login
def api_cas_smoke():
    """Trigger smoke tests (admin only)."""
    u = current_user()
    if u["role"] != "admin":
        abort(403)
    try:
        from assurance_smoke import run_smoke_tests
        result = run_smoke_tests(str(DB_PATH), str(BASE_DIR))
        # Store result in telemetry
        ts = _get_telemetry_store()
        if ts:
            ts.record_smoke_result(result["passed"], result["total"],
                                   result["results"], result["duration_s"])
        # T-CAS-2B: Auto-generate CR if smoke fails
        try:
            if result["passed"] < result["total"]:
                failed_checks = [r for r in result["results"] if not r.get("passed")]
                findings = []
                for fc in failed_checks:
                    findings.append({
                        "severity": "CRITICAL" if fc["name"] in ("db_writable", "op_tables_exist") else "WARNING",
                        "source": "smoke",
                        "check_name": fc["name"],
                        "details": fc.get("message", "Check failed"),
                        "measured_value": "FAIL",
                        "threshold": "PASS",
                        "recommended_action": f"Investigate and fix {fc['name']}",
                    })
                sev = "CRITICAL" if any(f["severity"] == "CRITICAL" for f in findings) else "WARNING"
                _maybe_create_cr(
                    source="smoke", severity=sev,
                    trigger_summary=f"Smoke test failed {len(failed_checks)}/{result['total']} checks",
                    trigger_snapshot={"smoke_result": result},
                    findings=findings,
                )
        except Exception:
            pass
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/cas/goldens/run", methods=["POST"])
@require_login
def api_cas_goldens_run():
    """Trigger golden regression tests (admin only). WARNING: costs API tokens."""
    u = current_user()
    if u["role"] != "admin":
        abort(403)
    try:
        from assurance_goldens import run_all_goldens
        results = run_all_goldens(goldens_dir=DATA_DIR / "goldens", base_dir=BASE_DIR)
        # Store each result
        ts = _get_telemetry_store()
        if ts:
            for r in results:
                ts.record_golden_result(
                    r["golden_name"], 1 if r["passed"] else 0,
                    r["total_checks"], r.get("matched", 0),
                    r.get("mismatched", 0), r.get("missing", 0),
                    r.get("extra", 0), r.get("duration_s", 0),
                    r.get("details"))
        # T-CAS-2B: Auto-generate CR if any golden fails
        try:
            failed_goldens = [r for r in results if not r.get("passed")]
            if failed_goldens:
                findings = []
                for fg in failed_goldens:
                    findings.append({
                        "severity": "CRITICAL",
                        "source": "golden",
                        "check_name": fg.get("golden_name", "unknown"),
                        "details": f"Mismatched: {fg.get('mismatched', 0)}, Missing: {fg.get('missing', 0)}",
                        "measured_value": f"{fg.get('mismatched', 0)} mismatches",
                        "threshold": "0 mismatches",
                        "recommended_action": f"Review golden case {fg.get('golden_name', '')} extraction output",
                    })
                _maybe_create_cr(
                    source="golden", severity="CRITICAL",
                    trigger_summary=f"Golden regression: {len(failed_goldens)}/{len(results)} cases failed",
                    trigger_snapshot={"golden_results": results},
                    findings=findings,
                )
        except Exception:
            pass
        return jsonify({"results": results, "count": len(results)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/cas/backup", methods=["POST"])
@require_login
def api_cas_backup():
    """Create a database backup (admin only)."""
    u = current_user()
    if u["role"] != "admin":
        abort(403)
    try:
        from assurance_backup import create_backup, cleanup_old_backups
        result = create_backup(str(DB_PATH), str(DATA_DIR / "backups"))
        # Store in telemetry
        ts = _get_telemetry_store()
        if ts:
            ts.record_backup(result["path"], result["size_bytes"],
                             result["sha256"], result["row_counts"])
        # Cleanup old backups
        cleanup_old_backups(str(DATA_DIR / "backups"), keep=30)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/cas/backup/verify/<int:backup_id>", methods=["POST"])
@require_login
def api_cas_backup_verify(backup_id):
    """Verify a backup (admin only)."""
    u = current_user()
    if u["role"] != "admin":
        abort(403)
    ts = _get_telemetry_store()
    if not ts:
        return jsonify({"error": "CAS not available"}), 503
    try:
        backups = ts.get_recent_backups(limit=100)
        target = None
        for b in backups:
            if b["id"] == backup_id:
                target = b
                break
        if not target:
            return jsonify({"error": "Backup not found"}), 404

        from assurance_backup import verify_backup
        result = verify_backup(target["backup_path"], target["sha256"])
        ts.record_backup_verify(backup_id, result["verified"], result["sha256"])
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/cas/run/<job_id>")
@require_login
def api_cas_run_detail(job_id):
    """Get run detail for a specific job."""
    u = current_user()
    if u["role"] not in ("admin", "partner"):
        abort(403)
    ts = _get_telemetry_store()
    if not ts:
        return jsonify({"error": "CAS not available"}), 503
    run = ts.get_run(job_id)
    if not run:
        return jsonify({"error": "Run not found"}), 404
    return jsonify(run)


@app.route("/api/cas/runs")
@require_login
def api_cas_runs():
    """Get recent extraction runs."""
    u = current_user()
    if u["role"] not in ("admin", "partner"):
        abort(403)
    ts = _get_telemetry_store()
    if not ts:
        return jsonify({"error": "CAS not available"}), 503
    limit = request.args.get("limit", 50, type=int)
    runs = ts.get_recent_runs(limit=min(limit, 200))
    return jsonify({"runs": runs, "count": len(runs)})


@app.route("/api/cas/drift")
@require_login
def api_cas_drift():
    """Get drift metrics history."""
    u = current_user()
    if u["role"] not in ("admin", "partner"):
        abort(403)
    ts = _get_telemetry_store()
    if not ts:
        return jsonify({"error": "CAS not available"}), 503
    limit = request.args.get("limit", 20, type=int)
    drift = ts.get_drift_summary(limit=min(limit, 100))
    return jsonify({"drift": drift, "count": len(drift)})


@app.route("/admin/cas")
@require_login
def admin_cas():
    """CAS System Health dashboard page."""
    u = current_user()
    if u["role"] not in ("admin", "partner"):
        abort(403)
    ts = _get_telemetry_store()
    cas_health = ts.cas_health_summary() if ts else {"state": "unknown", "label": "Unavailable"}
    summary = build_admin_summary()
    return render_template("admin_cas.html",
        active="cas", header="System Health",
        current_user=u,
        version=summary["health"]["version"],
        uptime_h=summary["health"]["uptime_h"],
        health_state=summary["health"]["state"],
        health_label=summary["health"]["label"],
        cas=cas_health,
    )


@app.route("/admin/cas/runs")
@require_login
def admin_cas_runs():
    """CAS Run Inspector dashboard page."""
    u = current_user()
    if u["role"] not in ("admin", "partner"):
        abort(403)
    ts = _get_telemetry_store()
    runs = ts.get_recent_runs(limit=50) if ts else []
    summary = build_admin_summary()
    return render_template("admin_cas_runs.html",
        active="cas_runs", header="Run Inspector",
        current_user=u,
        version=summary["health"]["version"],
        uptime_h=summary["health"]["uptime_h"],
        health_state=summary["health"]["state"],
        health_label=summary["health"]["label"],
        runs=runs,
    )


# ─── CAS: Report Endpoints (T-CAS-1R) ───────────────────────────────────────

def _get_cas_report_gen():
    """Lazy-init the CAS report generator."""
    ts = _get_telemetry_store()
    if not ts:
        return None
    try:
        from cas_reports import CASReportGenerator
        return CASReportGenerator(ts, app_version=_app_version)
    except Exception:
        return None


@app.route("/api/cas/report/daily")
@require_login
def api_cas_report_daily():
    """Generate Daily Health report (R1)."""
    u = current_user()
    if u["role"] not in ("admin", "partner"):
        abort(403)
    gen = _get_cas_report_gen()
    if not gen:
        return jsonify({"error": "CAS report generator not available"}), 503
    try:
        result = gen.render_daily_health()
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/cas/report/runs")
@require_login
def api_cas_report_runs():
    """Generate Runs report (R2)."""
    u = current_user()
    if u["role"] not in ("admin", "partner"):
        abort(403)
    gen = _get_cas_report_gen()
    if not gen:
        return jsonify({"error": "CAS report generator not available"}), 503
    try:
        limit = request.args.get("limit", 50, type=int)
        result = gen.render_runs(limit=min(limit, 200))
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/cas/report/regressions")
@require_login
def api_cas_report_regressions():
    """Generate Regressions report (R3)."""
    u = current_user()
    if u["role"] not in ("admin", "partner"):
        abort(403)
    gen = _get_cas_report_gen()
    if not gen:
        return jsonify({"error": "CAS report generator not available"}), 503
    try:
        days = request.args.get("days", 7, type=int)
        result = gen.render_regressions(days=days)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/cas/report/backups")
@require_login
def api_cas_report_backups():
    """Generate Backups report (R4)."""
    u = current_user()
    if u["role"] not in ("admin", "partner"):
        abort(403)
    gen = _get_cas_report_gen()
    if not gen:
        return jsonify({"error": "CAS report generator not available"}), 503
    try:
        days = request.args.get("days", 30, type=int)
        result = gen.render_backups(days=days)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/cas/report/agent-pack")
@require_login
def api_cas_report_agent_pack():
    """Generate and download Agent Pack (zip with all reports)."""
    u = current_user()
    if u["role"] not in ("admin", "partner"):
        abort(403)
    gen = _get_cas_report_gen()
    if not gen:
        return jsonify({"error": "CAS report generator not available"}), 503
    try:
        zip_bytes = gen.build_agent_pack()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return send_file(
            BytesIO(zip_bytes),
            mimetype="application/zip",
            as_attachment=True,
            download_name=f"cas_agent_pack_{timestamp}.zip",
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── CAS: Change Request Routes (T-CAS-2B) ──────────────────────────────────

def _maybe_create_cr(source, severity, trigger_summary, trigger_snapshot, findings):
    """Create a CR if findings are non-empty. Writes folder + files. Never crashes caller.

    Includes dedup guard: skips if there's already an open CR for the same source.
    All wrapped in try/except: pass to maintain the safety contract.

    Returns:
        cr_id (str) or None.
    """
    try:
        ts = _get_telemetry_store()
        if not ts or not findings:
            return None

        # Dedup guard: skip if open CR exists for same source
        existing_open = ts.get_open_change_requests()
        for ocr in existing_open:
            if ocr.get("source") == source and ocr.get("status") in ("open", "fix_submitted"):
                return None  # Already tracked

        result = ts.create_change_request(source, severity, trigger_summary,
                                          trigger_snapshot, findings)
        cr_id = result["cr_id"]

        # Create CR folder and write findings files
        cr_folder = DATA_DIR / "reports" / "change_requests" / cr_id
        os.makedirs(str(cr_folder), exist_ok=True)

        # Write findings.json
        findings_json_path = cr_folder / "findings.json"
        with open(str(findings_json_path), "w") as f:
            json.dump({"cr_id": cr_id, "findings": findings,
                       "trigger_snapshot": trigger_snapshot}, f, indent=2, default=str)

        # Write findings.md
        gen = _get_cas_report_gen()
        if gen:
            try:
                md_result = gen.render_cr_findings(cr_id)
                findings_md_path = cr_folder / "findings.md"
                with open(str(findings_md_path), "w") as f:
                    f.write(md_result["markdown"])
            except Exception:
                pass

        return cr_id
    except Exception:
        pass
    return None


def _run_post_fix_gate(cr_id):
    """Re-run the checks that triggered a CR and compare before/after.

    Returns:
        dict with keys: gate_result, checks_run, checks_passed,
                        before_snapshot, after_snapshot, details, error.
    """
    ts = _get_telemetry_store()
    if not ts:
        return {"gate_result": "REJECTED", "error": "TelemetryStore not available"}

    cr = ts.get_change_request(cr_id)
    if not cr:
        return {"gate_result": "REJECTED", "error": "CR not found"}

    source = cr["source"]
    before_snapshot = json.loads(cr.get("trigger_snapshot") or "{}")
    details = []
    checks_run = 0
    checks_passed = 0

    try:
        if source == "smoke":
            from assurance_smoke import run_smoke_tests
            result = run_smoke_tests(str(DB_PATH), str(BASE_DIR))
            ts.record_smoke_result(result["passed"], result["total"],
                                   result["results"], result["duration_s"])
            checks_run = result["total"]
            checks_passed = result["passed"]
            after_snapshot = {"smoke_result": result}
            for r in result["results"]:
                details.append({"check": r["name"], "passed": r["passed"],
                               "message": r.get("message", "")})

        elif source == "golden":
            from assurance_goldens import run_all_goldens
            results = run_all_goldens(goldens_dir=DATA_DIR / "goldens",
                                      base_dir=BASE_DIR)
            for r in results:
                ts.record_golden_result(
                    r["golden_name"], 1 if r["passed"] else 0,
                    r["total_checks"], r.get("matched", 0),
                    r.get("mismatched", 0), r.get("missing", 0),
                    r.get("extra", 0), r.get("duration_s", 0),
                    r.get("details"))
            checks_run = len(results)
            checks_passed = sum(1 for r in results if r.get("passed"))
            after_snapshot = {"golden_results": results}
            for r in results:
                details.append({"check": r["golden_name"], "passed": r.get("passed", False),
                               "mismatched": r.get("mismatched", 0),
                               "missing": r.get("missing", 0)})

        elif source == "drift":
            drift_check = ts.check_drift_thresholds()
            checks_run = 3  # edit_rate, needs_review_rate, audit_pass_rate
            violation_count = len(drift_check.get("violations", []))
            checks_passed = checks_run - violation_count
            after_snapshot = {"drift_latest": ts.get_drift_summary(limit=1),
                             "drift_check": drift_check}
            details = drift_check.get("violations", [])

        elif source == "error_rate":
            err_data = ts.get_error_rate_24h()
            checks_run = 1
            rate = err_data.get("error_rate", 0) if err_data else 0
            checks_passed = 1 if rate <= 0.20 else 0
            after_snapshot = err_data or {}
            details = [{"check": "error_rate_24h", "passed": checks_passed == 1,
                        "measured": f"{rate*100:.1f}%", "threshold": "20%"}]
        else:
            return {"gate_result": "REJECTED", "error": f"Unknown source: {source}"}

    except Exception as e:
        return {"gate_result": "REJECTED", "error": str(e)}

    # Determine gate result
    if checks_run > 0 and checks_passed == checks_run:
        gate_result = "ACCEPTED"
    elif checks_passed > 0:
        gate_result = "NEEDS_REVIEW"
    else:
        gate_result = "REJECTED"

    # Record gate result
    ts.record_gate_result(cr_id, gate_result, checks_run, checks_passed,
                          before_snapshot, after_snapshot, details)

    # Write gate_result.json to CR folder
    try:
        cr_folder = DATA_DIR / "reports" / "change_requests" / cr_id
        os.makedirs(str(cr_folder), exist_ok=True)
        gate_path = cr_folder / "gate_result.json"
        with open(str(gate_path), "w") as f:
            json.dump({
                "gate_result": gate_result,
                "checks_run": checks_run,
                "checks_passed": checks_passed,
                "details": details,
            }, f, indent=2, default=str)
    except Exception:
        pass

    return {
        "gate_result": gate_result,
        "checks_run": checks_run,
        "checks_passed": checks_passed,
        "before_snapshot": before_snapshot,
        "after_snapshot": after_snapshot,
        "details": details,
    }


@app.route("/api/cas/cr")
@require_login
def api_cas_cr_list():
    """List Change Requests. Optional ?status=open filter."""
    u = current_user()
    if u["role"] not in ("admin", "partner"):
        abort(403)
    ts = _get_telemetry_store()
    if not ts:
        return jsonify({"error": "CAS not available"}), 503
    status_filter = request.args.get("status")
    if status_filter == "open":
        crs = ts.get_open_change_requests()
    elif status_filter == "closed":
        all_crs = ts.get_all_change_requests(limit=100)
        crs = [c for c in all_crs if c.get("status") == "closed"]
    else:
        crs = ts.get_all_change_requests(limit=100)
    return jsonify({"change_requests": crs, "count": len(crs)})


@app.route("/api/cas/cr/<cr_id>")
@require_login
def api_cas_cr_detail(cr_id):
    """Get CR detail including findings and gate result."""
    u = current_user()
    if u["role"] not in ("admin", "partner"):
        abort(403)
    ts = _get_telemetry_store()
    if not ts:
        return jsonify({"error": "CAS not available"}), 503
    cr = ts.get_change_request(cr_id)
    if not cr:
        return jsonify({"error": "CR not found"}), 404
    # Add merge guard info
    cr["merge_guard"] = ts.can_merge_fix(cr_id)
    return jsonify(cr)


@app.route("/api/cas/cr/<cr_id>/manifest", methods=["POST"])
@require_login
def api_cas_cr_submit_manifest(cr_id):
    """Submit fix manifest. Body: JSON with files_changed, tests_added, config_changed, description, author."""
    u = current_user()
    if u["role"] != "admin":
        abort(403)
    ts = _get_telemetry_store()
    if not ts:
        return jsonify({"error": "CAS not available"}), 503

    data = request.get_json(silent=True) or {}
    # Auto-fill author and timestamp if not provided
    if not data.get("author"):
        data["author"] = u.get("display_name", u.get("username", "admin"))
    if not data.get("timestamp"):
        data["timestamp"] = datetime.now().isoformat()

    result = ts.submit_fix_manifest(cr_id, data)
    if not result["success"]:
        return jsonify(result), 400
    return jsonify(result)


@app.route("/api/cas/cr/<cr_id>/gate", methods=["POST"])
@require_login
def api_cas_cr_run_gate(cr_id):
    """Run post-fix gate verification."""
    u = current_user()
    if u["role"] != "admin":
        abort(403)
    ts = _get_telemetry_store()
    if not ts:
        return jsonify({"error": "CAS not available"}), 503

    # Verify CR exists and has manifest
    cr = ts.get_change_request(cr_id)
    if not cr:
        return jsonify({"error": "CR not found"}), 404
    if cr["status"] not in ("fix_submitted", "gate_failed", "gate_passed"):
        return jsonify({"error": f"CR must have fix manifest submitted first (status: {cr['status']})"}), 400

    result = _run_post_fix_gate(cr_id)
    if result.get("error"):
        return jsonify(result), 500
    return jsonify(result)


@app.route("/api/cas/cr/<cr_id>/close", methods=["POST"])
@require_login
def api_cas_cr_close(cr_id):
    """Close a CR. Enforces merge guard (can_merge_fix must return True)."""
    u = current_user()
    if u["role"] != "admin":
        abort(403)
    ts = _get_telemetry_store()
    if not ts:
        return jsonify({"error": "CAS not available"}), 503

    merge = ts.can_merge_fix(cr_id)
    if not merge["can_merge"]:
        return jsonify({"error": f"Cannot close CR: {merge['reason']}"}), 400

    ok = ts.update_cr_status(cr_id, "closed", closed_by=u.get("display_name", u.get("username", "admin")))
    if not ok:
        return jsonify({"error": "Failed to close CR"}), 500
    return jsonify({"status": "closed", "cr_id": cr_id})


@app.route("/api/cas/cr/<cr_id>/agent-pack")
@require_login
def api_cas_cr_agent_pack(cr_id):
    """Download CR-specific agent pack zip."""
    u = current_user()
    if u["role"] not in ("admin", "partner"):
        abort(403)
    gen = _get_cas_report_gen()
    if not gen:
        return jsonify({"error": "CAS report generator not available"}), 503
    try:
        zip_bytes = gen.build_cr_agent_pack(cr_id)
        return send_file(
            BytesIO(zip_bytes),
            mimetype="application/zip",
            as_attachment=True,
            download_name=f"cas_cr_{cr_id}.zip",
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/admin/cas/cr")
@require_login
def admin_cas_cr_list():
    """CR list page."""
    u = current_user()
    if u["role"] not in ("admin", "partner"):
        abort(403)
    ts = _get_telemetry_store()
    status_filter = request.args.get("status", "all")
    if status_filter == "open":
        crs = ts.get_open_change_requests() if ts else []
    elif status_filter == "closed":
        all_crs = ts.get_all_change_requests(limit=100) if ts else []
        crs = [c for c in all_crs if c.get("status") == "closed"]
    else:
        crs = ts.get_all_change_requests(limit=100) if ts else []
    summary = build_admin_summary()
    return render_template("admin_cas_cr.html",
        active="cas_cr", header="Change Requests",
        current_user=u,
        version=summary["health"]["version"],
        uptime_h=summary["health"]["uptime_h"],
        health_state=summary["health"]["state"],
        health_label=summary["health"]["label"],
        change_requests=crs,
        filter=status_filter,
    )


@app.route("/admin/cas/cr/<cr_id>")
@require_login
def admin_cas_cr_detail(cr_id):
    """CR detail page."""
    u = current_user()
    if u["role"] not in ("admin", "partner"):
        abort(403)
    ts = _get_telemetry_store()
    if not ts:
        abort(503)
    cr = ts.get_change_request(cr_id)
    if not cr:
        abort(404)
    cr["merge_guard"] = ts.can_merge_fix(cr_id)
    summary = build_admin_summary()
    return render_template("admin_cas_cr_detail.html",
        active="cas_cr", header=f"CR {cr_id}",
        current_user=u,
        version=summary["health"]["version"],
        uptime_h=summary["health"]["uptime_h"],
        health_state=summary["health"]["state"],
        health_label=summary["health"]["label"],
        cr=cr,
    )


# ─── Transaction Ledger Routes (T-TXN-LEDGER-1) ────────────────────────────

@app.route("/admin/ledger")
@require_login
def admin_ledger():
    """Transaction ledger dashboard: client picker + monthly summary grid."""
    u = current_user()
    if u["role"] not in ("admin", "partner", "preparer"):
        abort(403)
    summary = build_admin_summary()
    txn_store = _get_transaction_store()
    clients = txn_store.get_clients_with_transactions() if txn_store else []
    return render_template("admin_ledger.html",
        active="ledger", header="Transaction Ledger",
        current_user=u, clients=clients,
        version=summary["health"]["version"],
        uptime_h=summary["health"]["uptime_h"],
        health_state=summary["health"]["state"],
        health_label=summary["health"]["label"],
    )

@app.route("/admin/ledger/review")
@require_login
def admin_ledger_review():
    """Transaction review page: table with inline category picker."""
    u = current_user()
    if u["role"] not in ("admin", "partner", "preparer", "reviewer"):
        abort(403)
    summary = build_admin_summary()
    txn_store = _get_transaction_store()
    clients = txn_store.get_clients_with_transactions() if txn_store else []
    return render_template("admin_ledger_review.html",
        active="ledger", header="Transaction Review",
        current_user=u, clients=clients,
        version=summary["health"]["version"],
        uptime_h=summary["health"]["uptime_h"],
        health_state=summary["health"]["state"],
        health_label=summary["health"]["label"],
    )

@app.route("/admin/ledger/rules")
@require_login
def admin_ledger_rules():
    """Category and vendor rules management."""
    u = current_user()
    if u["role"] not in ("admin", "partner"):
        abort(403)
    summary = build_admin_summary()
    return render_template("admin_ledger_rules.html",
        active="ledger", header="Ledger Rules",
        current_user=u,
        version=summary["health"]["version"],
        uptime_h=summary["health"]["uptime_h"],
        health_state=summary["health"]["state"],
        health_label=summary["health"]["label"],
    )


# ─── Transaction Ledger API ─────────────────────────────────────────────────

@app.route("/api/ledger/taxonomy")
@require_login
def api_ledger_taxonomy():
    """Return category taxonomy for UI dropdowns."""
    from transaction_store import CATEGORY_TAXONOMY
    return jsonify({"ok": True, "taxonomy": CATEGORY_TAXONOMY})

@app.route("/api/ledger/summary")
@require_login
def api_ledger_summary():
    """Get monthly category summary for a client/year."""
    client = request.args.get("client", "")
    year = request.args.get("year", "", type=str)
    if not client or not year:
        return jsonify({"error": "client and year required"}), 400

    txn_store = _get_transaction_store()
    if not txn_store:
        return jsonify({"error": "Transaction store not available"}), 503

    summary = txn_store.get_monthly_summary(client, int(year) if year.isdigit() else 0)
    stats = txn_store.count_by_status(client, int(year) if year.isdigit() else 0)
    return jsonify({"ok": True, "summary": summary, "stats": stats})

@app.route("/api/ledger/transactions")
@require_login
def api_ledger_transactions():
    """Get transactions with pagination and filters."""
    client = request.args.get("client", "")
    year = request.args.get("year", "", type=str)
    if not client or not year:
        return jsonify({"error": "client and year required"}), 400

    txn_store = _get_transaction_store()
    if not txn_store:
        return jsonify({"error": "Transaction store not available"}), 503

    page = request.args.get("page", 1, type=int)
    per_page = min(request.args.get("per_page", 100, type=int), 500)

    filters = {}
    for key in ("status", "category", "category_group", "month",
                 "txn_type", "search", "date_from", "date_to", "vendor_norm"):
        val = request.args.get(key, "")
        if val:
            filters[key] = int(val) if key == "month" else val

    result = txn_store.get_transactions(
        client, int(year) if year.isdigit() else 0,
        filters=filters, page=page, per_page=per_page,
    )
    return jsonify({"ok": True, **result})

@app.route("/api/ledger/categorize", methods=["POST"])
@require_login
def api_ledger_categorize():
    """Set category on one or more transactions."""
    payload = request.get_json(silent=True) or {}
    txn_ids = payload.get("txn_ids", [])
    category = payload.get("category", "")
    learn = payload.get("learn", False)

    if not txn_ids or not category:
        return jsonify({"error": "txn_ids and category required"}), 400

    txn_store = _get_transaction_store()
    if not txn_store:
        return jsonify({"error": "Transaction store not available"}), 503

    u = current_user()
    reviewer = u["display_name"] if u else ""

    count = txn_store.bulk_categorize(txn_ids, category, reviewer=reviewer)

    # Learn vendor rule if requested
    if learn and count > 0:
        for tid in txn_ids:
            txn = txn_store.get_transaction(tid)
            if txn and txn.get("vendor_norm"):
                txn_store.learn_vendor_rule(txn["vendor_norm"], category)
                break  # Learn from first transaction with a vendor

    return jsonify({"ok": True, "updated": count})

@app.route("/api/ledger/verify", methods=["POST"])
@require_login
def api_ledger_verify():
    """Mark transactions as verified."""
    payload = request.get_json(silent=True) or {}
    txn_ids = payload.get("txn_ids", [])

    if not txn_ids:
        return jsonify({"error": "txn_ids required"}), 400

    txn_store = _get_transaction_store()
    if not txn_store:
        return jsonify({"error": "Transaction store not available"}), 503

    u = current_user()
    reviewer = u["display_name"] if u else ""

    verified = 0
    skipped = 0
    for tid in txn_ids:
        if txn_store.verify(tid, reviewer=reviewer):
            verified += 1
        else:
            skipped += 1

    return jsonify({"ok": True, "verified": verified, "skipped": skipped})

@app.route("/api/ledger/correct", methods=["POST"])
@require_login
def api_ledger_correct():
    """Apply corrections to a single transaction."""
    payload = request.get_json(silent=True) or {}
    txn_id = payload.get("txn_id", "")
    corrections = payload.get("corrections", {})

    if not txn_id or not corrections:
        return jsonify({"error": "txn_id and corrections required"}), 400

    txn_store = _get_transaction_store()
    if not txn_store:
        return jsonify({"error": "Transaction store not available"}), 503

    u = current_user()
    reviewer = u["display_name"] if u else ""

    ok = txn_store.correct(txn_id, corrections, reviewer=reviewer)
    if not ok:
        return jsonify({"error": "Transaction not found or correction failed"}), 404

    return jsonify({"ok": True, "txn_id": txn_id})

@app.route("/api/ledger/apply-rules", methods=["POST"])
@require_login
def api_ledger_apply_rules():
    """Run vendor/category rules engine on uncategorized transactions."""
    payload = request.get_json(silent=True) or {}
    client = payload.get("client", "")
    year = payload.get("year")

    txn_store = _get_transaction_store()
    if not txn_store:
        return jsonify({"error": "Transaction store not available"}), 503

    result = txn_store.apply_vendor_rules(
        client_name=client or None,
        year=int(year) if year else None,
    )
    return jsonify({"ok": True, **result})

@app.route("/api/ledger/report", methods=["POST"])
@require_login
def api_ledger_report():
    """Generate and download monthly summary Excel report."""
    payload = request.get_json(silent=True) or {}
    client = payload.get("client", "")
    year = payload.get("year")

    if not client or not year:
        return jsonify({"error": "client and year required"}), 400

    txn_store = _get_transaction_store()
    if not txn_store:
        return jsonify({"error": "Transaction store not available"}), 503

    try:
        from transaction_reports import TransactionReportBuilder
        safe_client = re.sub(r'[^\w\s-]', '', client).strip().replace(' ', '_')
        filename = f"{safe_client}-txn-summary-{year}.xlsx"
        output_path = OUTPUT_DIR / filename

        builder = TransactionReportBuilder(txn_store, client, int(year))
        builder.build(str(output_path))

        return send_file(str(output_path), as_attachment=True,
                         download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/ledger/events/<txn_id>")
@require_login
def api_ledger_events(txn_id):
    """Get audit trail for a specific transaction."""
    txn_store = _get_transaction_store()
    if not txn_store:
        return jsonify({"error": "Transaction store not available"}), 503

    events = txn_store.get_events(txn_id)
    txn = txn_store.get_transaction(txn_id)
    return jsonify({"ok": True, "events": events, "transaction": txn})


# ─── Ledger: Vendor/Category Rules API ──────────────────────────────────────

@app.route("/api/ledger/vendor-rules")
@require_login
def api_ledger_vendor_rules_list():
    """List all vendor rules."""
    txn_store = _get_transaction_store()
    if not txn_store:
        return jsonify({"error": "Transaction store not available"}), 503

    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 50, type=int)
    result = txn_store.get_vendor_rules(page=page, per_page=per_page)
    return jsonify({"ok": True, **result})

@app.route("/api/ledger/vendor-rules", methods=["POST"])
@require_login
def api_ledger_vendor_rules_create():
    """Create a vendor rule."""
    payload = request.get_json(silent=True) or {}
    pattern = payload.get("vendor_pattern", "").strip()
    match_type = payload.get("match_type", "exact")
    category = payload.get("category", "")

    if not pattern or not category:
        return jsonify({"error": "vendor_pattern and category required"}), 400

    txn_store = _get_transaction_store()
    if not txn_store:
        return jsonify({"error": "Transaction store not available"}), 503

    u = current_user()
    try:
        rule_id = txn_store.add_vendor_rule(
            pattern, match_type, category,
            created_by=u["display_name"] if u else "",
        )
        return jsonify({"ok": True, "rule_id": rule_id})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

@app.route("/api/ledger/vendor-rules/<int:rule_id>", methods=["DELETE"])
@require_login
def api_ledger_vendor_rules_delete(rule_id):
    """Delete a vendor rule."""
    txn_store = _get_transaction_store()
    if not txn_store:
        return jsonify({"error": "Transaction store not available"}), 503

    ok = txn_store.delete_vendor_rule(rule_id)
    if not ok:
        return jsonify({"error": "Rule not found"}), 404
    return jsonify({"ok": True})

@app.route("/api/ledger/category-rules")
@require_login
def api_ledger_category_rules_list():
    """List all category rules."""
    txn_store = _get_transaction_store()
    if not txn_store:
        return jsonify({"error": "Transaction store not available"}), 503

    rules = txn_store.get_category_rules()
    return jsonify({"ok": True, "rules": rules})

@app.route("/api/ledger/category-rules", methods=["POST"])
@require_login
def api_ledger_category_rules_create():
    """Create a category rule."""
    payload = request.get_json(silent=True) or {}
    keyword = payload.get("keyword", "").strip()
    category = payload.get("category", "")
    priority = payload.get("priority", 100)

    if not keyword or not category:
        return jsonify({"error": "keyword and category required"}), 400

    txn_store = _get_transaction_store()
    if not txn_store:
        return jsonify({"error": "Transaction store not available"}), 503

    u = current_user()
    try:
        rule_id = txn_store.add_category_rule(
            keyword, category, priority=priority,
            created_by=u["display_name"] if u else "",
        )
        return jsonify({"ok": True, "rule_id": rule_id})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

@app.route("/api/ledger/category-rules/<int:rule_id>", methods=["DELETE"])
@require_login
def api_ledger_category_rules_delete(rule_id):
    """Delete a category rule."""
    txn_store = _get_transaction_store()
    if not txn_store:
        return jsonify({"error": "Transaction store not available"}), 503

    ok = txn_store.delete_category_rule(rule_id)
    if not ok:
        return jsonify({"error": "Rule not found"}), 404
    return jsonify({"ok": True})


# ─── Routes ───────────────────────────────────────────────────────────────────

@app.route("/")
@require_login
def index():
    return render_template_string(MAIN_HTML)

@app.route("/api/login", methods=["POST"])
def api_login():
    """JSON login endpoint for headless clients (Excel add-in)."""
    import time as _time
    data = request.get_json(silent=True) or {}
    username = data.get("username", "").strip().lower()
    pin = data.get("pin", "").strip()
    ip = request.remote_addr or "unknown"
    key = (username, ip)
    st = _failed_logins.get(key, {"count": 0, "locked_until": 0})
    now_epoch = int(_time.time())

    if st["locked_until"] > now_epoch:
        return jsonify({"error": "Too many attempts. Try again shortly."}), 429

    u = get_user_by_username(username)
    if not u or not u["is_active"] or not pin.isdigit() or len(pin) != 6:
        st["count"] += 1
        if st["count"] >= MAX_FAILED_ATTEMPTS:
            st["locked_until"] = now_epoch + LOGIN_LOCKOUT_SECONDS
        _failed_logins[key] = st
        return jsonify({"error": "Invalid credentials."}), 401

    if not check_password_hash(u["pin_hash"], pin):
        st["count"] += 1
        if st["count"] >= MAX_FAILED_ATTEMPTS:
            st["locked_until"] = now_epoch + LOGIN_LOCKOUT_SECONDS
        _failed_logins[key] = st
        return jsonify({"error": "Invalid credentials."}), 401

    _failed_logins.pop(key, None)
    session.clear()
    session["user_id"] = u["id"]
    session["last_seen"] = now_epoch
    update_last_login(u["id"])
    log_event("info", "login_success",
              f"{u['display_name']} logged in via API",
              user_id=u["id"], ip_addr=ip)

    # SEC-006: Inform API clients if PIN reset is required
    must_reset = bool(u.get("must_reset_pin"))
    if must_reset:
        session["force_pin_reset"] = True

    return jsonify({
        "ok": True,
        "user_id": u["id"],
        "username": u["username"],
        "display_name": u["display_name"],
        "role": u["role"],
        "must_reset_pin": must_reset,
    })


@app.route("/api/users-list")
def api_users_list():
    """Public list of active usernames for login dropdown (no secrets)."""
    try:
        users = list_active_users()
        return jsonify([{"username": u["username"], "display_name": u["display_name"]} for u in users])
    except Exception:
        return jsonify([])


@app.route("/api/me")
def api_me():
    """Return current logged-in user info (or anonymous fallback)."""
    u = current_user()
    if u:
        return jsonify({
            "logged_in": True,
            "user_id": u["id"],
            "username": u["username"],
            "display_name": u["display_name"],
            "role": u["role"],
            "initials": "".join(w[0] for w in u["display_name"].split() if w).upper()[:3],
        })
    return jsonify({
        "logged_in": False,
        "username": "operator",
        "display_name": "Operator",
        "role": "reviewer",
        "initials": "",
    })

@app.route("/api/upload", methods=["POST"])
@require_login
def upload():
    if "pdf" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["pdf"]
    if not f.filename.lower().endswith(".pdf"):
        return jsonify({"error": "File must be a PDF"}), 400

    year = request.form.get("year", str(datetime.now().year))
    # Validate year is a 4-digit number in reasonable range
    try:
        year_int = int(year)
        if year_int < 2000 or year_int > datetime.now().year + 1:
            return jsonify({"error": f"Invalid year: {year}"}), 400
    except ValueError:
        return jsonify({"error": f"Invalid year: {year}"}), 400

    skip_verify = request.form.get("skip_verify") == "true"
    disable_pii = request.form.get("disable_pii") == "true"
    use_ocr_first = request.form.get("use_ocr_first") == "true"
    client_name = request.form.get("client_name", "").strip()
    if not client_name:
        return jsonify({"error": "Please select a client"}), 400
    doc_type = request.form.get("doc_type", "tax_returns")
    if doc_type not in VALID_DOC_TYPES:
        doc_type = "tax_returns"  # Safe default
    output_format = request.form.get("output_format", "tax_review")
    VALID_OUTPUT_FORMATS = {"tax_review", "journal_entries", "account_balances", "trial_balance", "transaction_register"}
    if output_format not in VALID_OUTPUT_FORMATS:
        output_format = "tax_review"
    user_notes = request.form.get("user_notes", "").strip()[:2000]  # Cap at 2000 chars
    ai_instructions = request.form.get("ai_instructions", "").strip()[:2000]

    # Generate job ID first (needed for unique filename)
    job_id = datetime.now().strftime("%m%d") + "-" + str(uuid.uuid4())[:6]

    # Save with unique name to prevent overwrites
    pdf_path = UPLOAD_DIR / (job_id + ".pdf")
    f.save(str(pdf_path))
    _secure_file(pdf_path)

    # Build client folder path
    resolved_client = _safe_client_name(client_name)
    client_dir = _client_dir(resolved_client, doc_type, year)

    jobs[job_id] = {
        "id": job_id,
        "filename": f.filename,
        "client_name": resolved_client,
        "doc_type": doc_type,
        "output_format": output_format,
        "user_notes": user_notes,
        "ai_instructions": ai_instructions,
        "year": year,
        "status": "queued",
        "stage": "queued",
        "progress": 0,
        "log": [],
        "created": datetime.now().isoformat(),
        "pdf_path": str(pdf_path),
        "client_folder": str(client_dir),
        "disable_pii": disable_pii,
        "use_ocr_first": use_ocr_first,
    }
    save_jobs()

    # Sprint 2: Log upload event
    upload_user = current_user()
    log_event("info", "upload_received",
              f"Upload: {f.filename} for {resolved_client} ({doc_type})",
              user_id=upload_user["id"] if upload_user else None,
              job_id=job_id,
              details={"filename": f.filename, "client": resolved_client,
                       "doc_type": doc_type, "year": year})

    t = threading.Thread(target=run_extraction, args=(job_id, pdf_path, year, skip_verify, doc_type, output_format, user_notes, ai_instructions, disable_pii, False, use_ocr_first))
    t.daemon = True
    t.start()

    return jsonify({"job_id": job_id})

@app.route("/api/status/<job_id>")
def status(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    out = _sanitize_job(job)
    out["recent_log"] = job.get("log", [])[-40:]
    out["log_length"] = len(job.get("log", []))
    return jsonify(out)

# ─── T1.5: Helpers for partial/progressive results ──────────────────────────

_NON_TAX_DOC_TYPES = frozenset({
    "supplemental_detail", "continuation_statement",
    "auto_insurance_policy", "insurance_card", "insurance_notice",
    "privacy_policy", "loan_statement", "mortgage_statement",
    "receipt",  # retail receipts (Belk, Walmart etc) are not tax forms
})

# Field names that indicate a non-tax "other" page (insurance docs misclassified as "other")
_INSURANCE_FIELD_SIGNALS = frozenset({
    "policy_number", "agency_code", "form_number", "coverage_type",
    "deductible", "premium", "vin", "vehicle_year",
})


def _build_page_map(extractions):
    """Build page_map dict from extractions for the review UI.

    Filters out non-tax document types (insurance, privacy policies, supplemental
    breakdowns) that pollute field counts and audit checks. These pages still
    appear in the review (images render) but their fields are excluded.
    Also flags percentage values that were misread as dollars.
    """
    page_map = {}
    for ext in extractions:
        p = ext.get("_page")
        if not p:
            continue
        doc_type = ext.get("document_type", "")
        # Detect "other" pages that are really insurance (misclassified)
        if doc_type == "other":
            field_names = set((ext.get("fields") or {}).keys())
            if field_names & _INSURANCE_FIELD_SIGNALS:
                doc_type = "insurance_other"  # reclassify
        # Skip non-tax document types — they pollute field counts and audit
        if doc_type in _NON_TAX_DOC_TYPES or doc_type == "insurance_other":
            # Still create the page entry so the image renders, but with zero fields
            if p not in page_map:
                page_map[p] = []
            page_map[p].append({
                "document_type": doc_type,
                "entity": ext.get("payer_or_entity", ""),
                "method": ext.get("_extraction_method", ""),
                "confidence": "",
                "fields": {},
                "_non_tax": True,
            })
            continue
        if p not in page_map:
            page_map[p] = []
        fields_out = {}
        for k, v in (ext.get("fields") or {}).items():
            val = v.get("value") if isinstance(v, dict) else v
            if val is None:
                continue
            conf = v.get("confidence", "") if isinstance(v, dict) else ""
            label = v.get("label_on_form", "") if isinstance(v, dict) else ""
            # Detect percentage values misread as dollar amounts:
            # if the label_on_form contains "percent" or "%" it's not a dollar field
            val_str = str(val)
            if label and ("percent" in label.lower() or "%" in label.lower()):
                if not val_str.endswith("%"):
                    val_str = val_str + "%"
                    val = val_str
                    conf = "low"  # flag as uncertain
            fields_out[k] = {
                "value": val,
                "confidence": conf,
                "label": label,
            }
        page_map[p].append({
            "document_type": doc_type,
            "entity": ext.get("payer_or_entity", ""),
            "method": ext.get("_extraction_method", ""),
            "confidence": ext.get("_overall_confidence", ""),
            "fields": fields_out,
        })
    return page_map


def _populate_facts(job_id, client_name, year, page_map):
    """Flatten all extracted fields into the facts table for cross-document queries.

    Uses FactStore with monotonic trust — corrected/confirmed facts are never
    overwritten by new extractions. Only 'extracted' status facts get refreshed.
    """
    try:
        from fact_store import FactStore
        fs = FactStore(str(DB_PATH))
        tax_year = None
        try:
            tax_year = int(year) if year else None
        except (ValueError, TypeError):
            pass
        count = 0
        for page_str, exts in page_map.items():
            page_num = int(page_str)
            for ext_idx, ext in enumerate(exts):
                doc_type = ext.get("document_type", "")
                entity = ext.get("entity", "") or ext.get("payer_or_entity", "")
                method = ext.get("method", "")
                for field_name, field_data in (ext.get("fields") or {}).items():
                    if isinstance(field_data, dict):
                        value = field_data.get("value")
                        conf = field_data.get("confidence", "")
                    else:
                        value = field_data
                        conf = ""
                    if value is None:
                        continue
                    value_num = None
                    value_text = str(value)
                    try:
                        value_num = float(str(value).replace(",", "").replace("$", ""))
                    except (ValueError, TypeError):
                        pass
                    fact_key = f"{doc_type}|{entity}|{field_name}"
                    evidence_ref = f"{page_num}:{ext_idx}:{field_name}"
                    # upsert respects monotonic trust — won't overwrite confirmed/corrected
                    fs.upsert_candidate_fact(
                        job_id=job_id,
                        client_id=client_name or "",
                        tax_year=tax_year or 0,
                        fact_key=fact_key,
                        value_num=value_num,
                        value_text=value_text,
                        status="extracted",
                        confidence=conf,
                        source_method=method,
                        source_doc=doc_type,
                        source_page=page_num,
                        evidence_ref=evidence_ref,
                    )
                    count += 1
        if count:
            print(f"  [facts] populated {count} facts for {job_id}")
    except Exception as e:
        print(f"  [facts] populate error: {e}")


def _get_partial_path(job_id):
    """Derive partial results file path for a job."""
    job = jobs.get(job_id)
    if not job:
        return None
    pdf_path = job.get("pdf_path", "")
    stem = Path(pdf_path).stem if pdf_path else job_id
    return str(OUTPUT_DIR / (stem + "_intake_partial_results.json"))


def _apply_locks(partial_data, vdata):
    """Overlay user corrections onto partial extraction data.

    Locked fields (user-edited) keep the user's value, not the extraction's.
    Later batches may attach more evidence but cannot change locked values.
    """
    corrections = vdata.get("fields", {})
    if not corrections:
        return partial_data

    for ext in partial_data.get("extractions", []):
        page = ext.get("_page")
        if not page:
            continue
        fields = ext.get("fields", {})
        for field_key, decision in corrections.items():
            parts = field_key.split(":")
            if len(parts) != 3:
                continue
            try:
                fpage = int(parts[0])
            except (ValueError, TypeError):
                continue
            fname = parts[2]
            if fpage != page:
                continue
            if fname in fields and decision.get("status") == "corrected":
                if isinstance(fields[fname], dict):
                    fields[fname]["value"] = decision.get("corrected_value", fields[fname].get("value"))
                    fields[fname]["_locked"] = True

    return partial_data


@app.route("/api/results/<job_id>")
def results(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    # T1.5: Support partial results during extraction
    if job.get("status") == "running":
        partial_path = _get_partial_path(job_id)
        if partial_path and os.path.exists(partial_path):
            try:
                with open(partial_path) as f:
                    data = json.load(f)
                # Merge user corrections (locked fields)
                vdata = _load_verifications(job_id)
                data = _apply_locks(data, vdata)
                data["page_map"] = _build_page_map(data.get("extractions", []))
                data["total_pages"] = job.get("total_pages") or max(
                    (int(k) for k in data["page_map"].keys()), default=1)
                return jsonify(data)
            except (IOError, json.JSONDecodeError):
                pass
        return jsonify({"error": "Results not ready", "partial": False}), 404

    # Complete results from log file
    log_path = job.get("output_log")
    if log_path and os.path.exists(log_path):
        with open(log_path) as f:
            data = json.load(f)
        data["page_map"] = _build_page_map(data.get("extractions", []))
        data["total_pages"] = job.get("total_pages") or max((int(k) for k in data["page_map"].keys()), default=1)
        # ── Populate facts table for cross-document queries ──
        _populate_facts(job_id, job.get("client_name", ""), job.get("year", 0), data["page_map"])
        # ── Doctrine drift detection ──
        try:
            from lite.doctrine.registry import get_current_manifest
            from lite.doctrine.drift import doctrine_drift_status
            manifest = get_current_manifest()
            # Extract doctrine fields from stored ardent_result or ardent_summary
            ar = data.get("ardent_result") or {}
            ars = data.get("ardent_summary") or {}
            log_dv = ar.get("doctrine_version") or ars.get("doctrine_version")
            log_dh = ar.get("doctrine_hash") or ars.get("doctrine_hash")
            drift = doctrine_drift_status(log_dv, log_dh, manifest.doctrine_version, manifest.doctrine_hash)
            data["doctrine_drift"] = drift
            data["doctrine_current"] = {
                "version": manifest.doctrine_version,
                "hash_short": manifest.doctrine_hash[:8],
            }
            if drift["status"] not in ("ok", "legacy"):
                print(f"  [DOCTRINE] DRIFT: {drift['message']}")
        except Exception:
            pass  # Doctrine unavailable — non-fatal
        # B7-UX: Include user's doc type selection for mismatch transparency
        data["user_doc_type"] = job.get("doc_type", "")
        return jsonify(data)
    return jsonify({"error": "Results not ready"}), 404

@app.route("/api/reextract-page/<job_id>/<int:page_num>", methods=["POST"])
def reextract_page(job_id, page_num):
    """Re-extract a single page with custom AI instructions."""
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    instructions = request.json.get("instructions", "").strip()
    if not instructions:
        return jsonify({"error": "No instructions provided"}), 400

    # Load the page image
    img_path = PAGES_DIR / job_id / f"page_{page_num}.jpg"
    if not img_path.exists():
        return jsonify({"error": f"Page {page_num} image not found"}), 404

    import base64
    try:
        import anthropic as _anthropic
    except ImportError:
        return jsonify({"error": "Anthropic library not available"}), 500

    with open(img_path, "rb") as f:
        img_b64 = base64.b64encode(f.read()).decode("utf-8")

    # Build extraction prompt with operator instructions
    # Import the vision prompt template from extract.py
    import importlib.util
    spec = importlib.util.spec_from_file_location("extract", str(BASE_DIR / "extract.py"))
    ext_mod = importlib.util.module_from_spec(spec)
    # Only load the constants we need (not the whole module execution)
    try:
        spec.loader.exec_module(ext_mod)
        vision_prompt = ext_mod.VISION_EXTRACTION_PROMPT
        model = ext_mod.MODEL
    except Exception:
        # Fallback if module load fails
        model = "claude-sonnet-4-20250514"
        vision_prompt = "Extract all data from this document page. Return JSON with document_type, payer_or_entity, fields (each with value, label_on_form, confidence)."

    context = f"The operator has provided these specific instructions for this page:\n{instructions}"
    doc_type = job.get("doc_type", "")
    if doc_type:
        context = f"Document type: {doc_type}\n{context}"
    prompt = vision_prompt.replace("{context}", context)

    # Call Claude (with PII guard — GAP-001)
    try:
        from pii_guard import guard_messages as _guard_messages
    except ImportError:
        _guard_messages = None

    try:
        client = _anthropic.Anthropic()
        raw_messages = [{"role": "user", "content": [
            {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": img_b64}},
            {"type": "text", "text": prompt}
        ]}]

        pii_tok = None
        if _guard_messages:
            safe_messages, pii_tok = _guard_messages(
                raw_messages, job_id=job_id, model=model,
                caller="reextract_page",
            )
        else:
            safe_messages = raw_messages

        msg = client.messages.create(
            model=model, max_tokens=8000,
            messages=safe_messages
        )
        raw = msg.content[0].text

        # Parse JSON from response
        import re as _re
        result = None
        # Try full response first
        try:
            result = json.loads(raw)
        except json.JSONDecodeError:
            # Try extracting JSON block
            m = _re.search(r'\{[\s\S]*\}', raw)
            if m:
                try:
                    result = json.loads(m.group(0))
                except json.JSONDecodeError:
                    pass

        # Detokenize PII back into result
        if pii_tok and result:
            result = pii_tok.detokenize_json(result)

        if not result or "fields" not in result:
            return jsonify({"error": "AI returned invalid response", "raw": raw[:500]}), 500

    except Exception as e:
        return jsonify({"error": f"AI call failed: {str(e)}"}), 500

    # Update the log file with new extraction
    log_path = job.get("output_log")
    if log_path and os.path.exists(log_path):
        with open(log_path) as f:
            log_data = json.load(f)

        # Find and replace extraction for this page, or add new one
        exts = log_data.get("extractions", [])
        replaced = False
        for i, ext in enumerate(exts):
            if ext.get("_page") == page_num:
                # Preserve metadata, replace content
                result["_page"] = page_num
                result["_extraction_method"] = "vision_reextract"
                result["_overall_confidence"] = ext.get("_overall_confidence")
                result["_reextract_instructions"] = instructions
                exts[i] = result
                replaced = True
                break
        if not replaced:
            result["_page"] = page_num
            result["_extraction_method"] = "vision_reextract"
            result["_reextract_instructions"] = instructions
            exts.append(result)
            exts.sort(key=lambda e: e.get("_page", 0))

        log_data["extractions"] = exts
        with open(log_path, "w") as f:
            json.dump(log_data, f, indent=2, default=str)

        # Also update client folder copy if it exists
        client_log = job.get("client_folder")
        if client_log:
            client_log_path = Path(client_log) / Path(log_path).name
            if client_log_path.exists():
                with open(client_log_path, "w") as f:
                    json.dump(log_data, f, indent=2, default=str)

    # Return the new extraction data in page_map format
    fields_out = {}
    for k, v in (result.get("fields") or {}).items():
        val = v.get("value") if isinstance(v, dict) else v
        if val is not None:
            fields_out[k] = {
                "value": val,
                "confidence": v.get("confidence", "") if isinstance(v, dict) else "",
                "label": v.get("label_on_form", "") if isinstance(v, dict) else "",
            }

    return jsonify({
        "success": True,
        "page": page_num,
        "document_type": result.get("document_type", ""),
        "entity": result.get("payer_or_entity", ""),
        "method": "vision_reextract",
        "fields": fields_out,
    })

@app.route("/api/ai-chat/<job_id>", methods=["POST"])
def ai_chat(job_id):
    """Chat with AI about the current extraction / page."""
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    message = request.json.get("message", "").strip()
    page_num = request.json.get("page")
    if not message:
        return jsonify({"error": "No message provided"}), 400

    import base64
    try:
        import anthropic as _anthropic
    except ImportError:
        return jsonify({"error": "Anthropic library not available"}), 500

    # Build context from verified extraction data (includes confirmed/corrected values)
    extraction_context = ""
    verified_log = get_verified_extractions(job_id)
    if verified_log:
        exts = verified_log.get("extractions", [])
        if page_num:
            page_exts = [e for e in exts if e.get("_page") == page_num]
            if page_exts:
                extraction_context = f"Extracted data for page {page_num}:\n{json.dumps(page_exts, indent=2, default=str)[:4000]}"
        if not extraction_context:
            summary = []
            for e in exts:
                p = e.get("_page", "?")
                dt = e.get("document_type", "?")
                ent = e.get("payer_or_entity", "?")
                summary.append(f"Page {p}: {dt} — {ent}")
            extraction_context = "Document summary:\n" + "\n".join(summary)

    # Include page image if available
    content = []
    if page_num:
        img_path = PAGES_DIR / job_id / f"page_{page_num}.jpg"
        if img_path.exists():
            with open(img_path, "rb") as f:
                img_b64 = base64.b64encode(f.read()).decode("utf-8")
            content.append({"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": img_b64}})

    prompt = f"""You are an assistant helping a CPA firm review tax document extractions.
The operator is reviewing extracted data and has a question.

{extraction_context}

Operator's question: {message}

Be concise and helpful. If the operator asks about a specific value, reference the extracted data. If they ask you to look at the page image, describe what you see."""

    content.append({"type": "text", "text": prompt})

    # PII guard — GAP-001
    try:
        from pii_guard import guard_messages as _guard_chat
    except ImportError:
        _guard_chat = None

    try:
        client = _anthropic.Anthropic()
        raw_messages = [{"role": "user", "content": content}]

        pii_tok = None
        if _guard_chat:
            safe_messages, pii_tok = _guard_chat(
                raw_messages, job_id=job_id,
                model="claude-sonnet-4-20250514", caller="ai_chat",
            )
        else:
            safe_messages = raw_messages

        msg = client.messages.create(
            model="claude-sonnet-4-20250514", max_tokens=1500,
            messages=safe_messages
        )
        reply = msg.content[0].text

        # Detokenize PII in the reply
        if pii_tok:
            reply = pii_tok.detokenize_text(reply)

        return jsonify({"reply": reply})
    except Exception as e:
        return jsonify({"error": f"AI call failed: {str(e)}"}), 500


@app.route("/api/page-image/<job_id>/<int:page_num>")
def page_image(job_id, page_num):
    img_path = PAGES_DIR / job_id / f"page_{page_num}.jpg"
    if img_path.exists():
        resp = send_file(str(img_path), mimetype="image/jpeg")
        resp.headers["Cache-Control"] = "public, max-age=86400"
        return resp
    abort(404)


@app.route("/api/page-words/<job_id>/<int:page_num>")
@require_login
def page_words(job_id, page_num):
    """Return OCR word data for a page (used by grid review highlighting)."""
    words = _get_page_word_data(job_id, page_num)
    resp = jsonify(words)
    resp.headers["Cache-Control"] = "public, max-age=86400"
    return resp


# ─── Guided Review: Backend Functions ────────────────────────────────────────

def _get_page_word_data(job_id, page_num):
    """Get Tesseract word-level bounding boxes for a review page image.

    Runs OCR once and caches results as JSON alongside the page image.
    Returns list of {text, left, top, width, height, conf} dicts.
    """
    cache_path = PAGES_DIR / job_id / f"page_{page_num}_words.json"
    if cache_path.exists():
        try:
            with open(cache_path) as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    img_path = PAGES_DIR / job_id / f"page_{page_num}.jpg"
    if not img_path.exists():
        return []
    try:
        import pytesseract
        img = Image.open(str(img_path))
        data = pytesseract.image_to_data(img, output_type=pytesseract.Output.DICT,
                                          config='--oem 3 --psm 6')
        words = []
        for i in range(len(data['text'])):
            txt = data['text'][i].strip()
            if txt and int(data['conf'][i]) >= 0:
                words.append({
                    'text': txt,
                    'left': int(data['left'][i]),
                    'top': int(data['top'][i]),
                    'width': int(data['width'][i]),
                    'height': int(data['height'][i]),
                    'conf': int(data['conf'][i]),
                })
        with open(cache_path, 'w') as f:
            json.dump(words, f)
        return words
    except Exception:
        return []


def _normalize_num_str(s):
    """Normalize a string to a canonical numeric form for matching.
    Strips $, commas, spaces. Returns '1234.56' form or None if not numeric."""
    s = re.sub(r'[$,\s()]', '', str(s).strip())
    # Handle negatives in parens like (1234.56)
    neg = s.startswith('-')
    s = s.lstrip('-')
    try:
        val = float(s)
        if neg:
            val = -val
        return f"{val:.2f}"
    except (ValueError, TypeError):
        return None


def _find_value_bboxes(words, value):
    """Find bounding boxes of OCR words matching a field value.

    Strategies:
    1. Exact single-word numeric match
    2. Multi-word numeric assembly (consecutive words forming a number)
    3. Case-insensitive text sequence match
    Returns list of word dicts with bounding box info.
    """
    if value is None:
        return []
    value_str = str(value).strip()
    if not value_str:
        return []

    norm_val = _normalize_num_str(value_str)

    # Strategy 1: Exact single-word numeric match
    if norm_val:
        for w in words:
            w_norm = _normalize_num_str(w['text'])
            if w_norm and w_norm == norm_val:
                return [w]

    # Strategy 2: Multi-word numeric assembly
    if norm_val:
        for i in range(len(words)):
            running = ""
            group = []
            for j in range(i, min(i + 10, len(words))):
                running += words[j]['text']
                group.append(words[j])
                r_norm = _normalize_num_str(running)
                if r_norm and r_norm == norm_val:
                    return group

    # Strategy 3: Case-insensitive text sequence match
    val_upper = value_str.upper()
    val_words = val_upper.split()
    if val_words:
        for i in range(len(words)):
            if words[i]['text'].upper().startswith(val_words[0]) or val_words[0].startswith(words[i]['text'].upper()):
                if len(val_words) == 1 and words[i]['text'].upper() == val_words[0]:
                    return [words[i]]
                match = True
                group = [words[i]]
                for k in range(1, len(val_words)):
                    if i + k < len(words) and words[i + k]['text'].upper() == val_words[k]:
                        group.append(words[i + k])
                    else:
                        match = False
                        break
                if match and len(group) == len(val_words):
                    return group

    return []


def _generate_uncertain_evidence(img, cache_path):
    """Generate full-page evidence with 'location uncertain' banner."""
    from PIL import ImageDraw, ImageFont
    result = img.convert("RGB")
    draw = ImageDraw.Draw(result)
    banner_h = 36
    draw.rectangle([0, 0, result.width, banner_h], fill=(255, 200, 0))
    try:
        font = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 16)
    except Exception:
        font = ImageFont.load_default()
    draw.text((12, 8), "EXACT LOCATION UNCERTAIN", fill=(120, 0, 0), font=font)
    result.save(str(cache_path), "PNG", optimize=True)
    return str(cache_path)


def _generate_evidence_image(job_id, page_num, field_value, field_id):
    """Generate a highlighted evidence PNG showing where a value appears on the page.

    Returns the path to the generated PNG, or None if generation fails.
    Caches results in data/evidence/<job_id>/<field_id>.png.
    """
    from PIL import ImageDraw
    evidence_dir = EVIDENCE_DIR / job_id
    evidence_dir.mkdir(parents=True, exist_ok=True)
    safe_field_id = re.sub(r'[^\w\-.]', '_', str(field_id))
    cache_path = evidence_dir / f"{safe_field_id}.png"
    if cache_path.exists():
        return str(cache_path)

    img_path = PAGES_DIR / job_id / f"page_{page_num}.jpg"
    if not img_path.exists():
        return None

    img = Image.open(str(img_path)).convert("RGBA")
    words = _get_page_word_data(job_id, page_num)
    if not words:
        return _generate_uncertain_evidence(img, cache_path)

    bboxes = _find_value_bboxes(words, field_value)
    if not bboxes:
        return _generate_uncertain_evidence(img, cache_path)

    # Draw highlight overlay
    overlay = Image.new("RGBA", img.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)
    PADDING = 5
    MIN_SIZE = 20
    for bbox in bboxes:
        x0 = max(0, bbox['left'] - PADDING)
        y0 = max(0, bbox['top'] - PADDING)
        x1 = min(img.width, bbox['left'] + bbox['width'] + PADDING)
        y1 = min(img.height, bbox['top'] + bbox['height'] + PADDING)
        # Enforce minimum highlight size
        if (x1 - x0) < MIN_SIZE:
            cx = (x0 + x1) // 2
            x0, x1 = cx - MIN_SIZE // 2, cx + MIN_SIZE // 2
        if (y1 - y0) < MIN_SIZE:
            cy = (y0 + y1) // 2
            y0, y1 = cy - MIN_SIZE // 2, cy + MIN_SIZE // 2
        draw.rectangle([x0, y0, x1, y1], fill=(255, 230, 0, 80))
        draw.rectangle([x0, y0, x1, y1], outline=(255, 0, 0), width=3)

    result = Image.alpha_composite(img, overlay).convert("RGB")
    result.save(str(cache_path), "PNG", optimize=True)
    return str(cache_path)


# ── Guided Review: Queue Builder ──

# Field type priority for review ordering (lower = reviewed first)
_FIELD_TYPE_PRIORITY = {
    'payer_or_entity': 10, 'payer_ein': 10, 'recipient_name': 10,
    'employer_name': 10, 'employer_ein': 10,
    'total_income': 20, 'total_wages': 20, 'taxable_income': 20,
    'total_tax': 20, 'total_deposits': 20, 'total_debits': 20,
    'total_credits': 20, 'total_withdrawals': 20,
    'ending_balance': 20, 'beginning_balance': 20,
    'wages': 25, 'federal_tax_withheld': 25, 'federal_wh': 25,
    'state_tax_withheld': 25, 'state_wh': 25,
    'social_security_wages': 25, 'ss_wages': 25,
    'medicare_wages': 25, 'medicare_wh': 25,
    'ordinary_dividends': 30, 'qualified_dividends': 30,
    'interest_income': 30, 'capital_gain_distributions': 30,
    'total_gain_loss': 30, 'b_total_gain_loss': 30,
    'gross_distribution': 30, 'taxable_amount': 30,
}

# Fields to skip in guided review (better in grid view)
_GUIDED_REVIEW_SKIP_PATTERNS = re.compile(
    r'^(txn_\d+|_|line_\d+|continuation_)'
)


def _field_display_name(doc_type, field_name):
    """Generate a human-readable display name for a field."""
    # Convert snake_case to Title Case
    display = field_name.replace('_', ' ').title()
    # Add box numbers for well-known W-2/1099 fields
    box_labels = {
        'wages': 'Box 1 - Wages',
        'federal_wh': 'Box 2 - Federal Tax Withheld',
        'federal_tax_withheld': 'Box 2 - Federal Tax Withheld',
        'ss_wages': 'Box 3 - Social Security Wages',
        'social_security_wages': 'Box 3 - Social Security Wages',
        'ss_tax': 'Box 4 - Social Security Tax',
        'medicare_wages': 'Box 5 - Medicare Wages',
        'medicare_wh': 'Box 6 - Medicare Tax Withheld',
        'ordinary_dividends': 'Box 1a - Ordinary Dividends',
        'qualified_dividends': 'Box 1b - Qualified Dividends',
        'capital_gain_distributions': 'Box 2a - Capital Gain Distributions',
        'interest_income': 'Box 1 - Interest Income',
        'gross_distribution': 'Box 1 - Gross Distribution',
        'taxable_amount': 'Box 2a - Taxable Amount',
        'beginning_balance': 'Beginning Balance',
        'ending_balance': 'Ending Balance',
        'total_deposits': 'Total Deposits',
        'total_withdrawals': 'Total Withdrawals',
    }
    return box_labels.get(field_name, display)


def _build_guided_queue(job_id):
    """Build prioritized review queue for guided review.

    Reads from the extraction log (same source as grid review), filters out
    already-verified fields, and orders by importance.
    """
    log_data = _load_extraction_log(job_id)
    if not log_data:
        return [], 0

    # Load already-verified fields
    conn = _get_db()
    verified = set()
    try:
        rows = conn.execute(
            "SELECT field_key FROM verified_fields WHERE job_id = ? AND status IN ('confirmed', 'corrected')",
            (job_id,)
        ).fetchall()
        verified = {r[0] for r in rows}
    finally:
        conn.close()

    extractions = log_data.get("extractions", [])
    queue = []
    total_fields = 0

    # Group by page for stable ordering
    page_groups = {}
    for ext in extractions:
        p = ext.get("_page")
        if p is not None:
            page_groups.setdefault(p, []).append(ext)

    for page_num in sorted(page_groups.keys()):
        for ext_idx, ext in enumerate(page_groups[page_num]):
            fields = ext.get("fields") or {}
            doc_type = ext.get("document_type", "")
            entity = ext.get("payer_or_entity", "") or ext.get("employer_name", "") or ""

            # Skip non-tax document types (same filter as _build_page_map)
            if doc_type in _NON_TAX_DOC_TYPES:
                continue
            if doc_type == "other" and set(fields.keys()) & _INSURANCE_FIELD_SIGNALS:
                continue

            for field_name, fdata in fields.items():
                if not isinstance(fdata, dict):
                    continue
                if field_name.startswith("_"):
                    continue
                if _GUIDED_REVIEW_SKIP_PATTERNS.match(field_name):
                    continue

                field_id = f"{page_num}:{ext_idx}:{field_name}"
                total_fields += 1

                if field_id in verified:
                    continue

                value = fdata.get("value")
                confidence = fdata.get("confidence", "")

                # Priority scoring (lower = first)
                priority = _FIELD_TYPE_PRIORITY.get(field_name, 50)
                if confidence in ('low', 'needs_review', 'unverified'):
                    priority -= 20

                queue.append({
                    'field_id': field_id,
                    'field_name': field_name,
                    'display_name': _field_display_name(doc_type, field_name),
                    'value': value,
                    'page_num': page_num,
                    'document_type': doc_type,
                    'entity': entity,
                    'confidence': confidence,
                    'method': ext.get("_extraction_method", ""),
                    'status': 'needs_review' if confidence in ('low', 'needs_review', 'unverified') else 'extracted',
                    '_priority': priority,
                })

    queue.sort(key=lambda x: (x['_priority'], x['page_num'], x['field_name']))
    # Strip internal priority from output
    for item in queue:
        item.pop('_priority', None)

    reviewed_count = total_fields - len(queue)

    # Cache total for fast-path response (T-UX-CONFIRM-FASTPATH)
    _job = jobs.get(job_id)
    if _job is not None:
        _job["_guided_total_fields"] = total_fields

    return queue, reviewed_count


# ── Guided Review: Concurrency Locks ──

def _acquire_review_lock(job_id, field_id, reviewer):
    """Acquire or extend a review lock. Returns (success, lock_holder)."""
    conn = _get_db()
    try:
        now = datetime.now().isoformat()
        # Clean expired locks
        conn.execute("DELETE FROM review_locks WHERE expires_at < ?", (now,))
        # Check existing lock by different reviewer
        existing = conn.execute(
            "SELECT locked_by, expires_at FROM review_locks WHERE job_id = ? AND field_id = ?",
            (job_id, field_id)
        ).fetchone()
        if existing and existing[0] != reviewer:
            return False, existing[0]
        # Acquire or extend
        expires = (datetime.now() + timedelta(seconds=REVIEW_LOCK_TIMEOUT_SECONDS)).isoformat()
        conn.execute(
            """INSERT INTO review_locks (job_id, field_id, locked_by, locked_at, expires_at)
               VALUES (?, ?, ?, ?, ?)
               ON CONFLICT(job_id, field_id) DO UPDATE SET
                   locked_by = excluded.locked_by,
                   locked_at = excluded.locked_at,
                   expires_at = excluded.expires_at""",
            (job_id, field_id, reviewer, now, expires)
        )
        conn.commit()
        return True, reviewer
    finally:
        conn.close()


def _release_review_lock(job_id, field_id):
    """Release a review lock."""
    conn = _get_db()
    try:
        conn.execute("DELETE FROM review_locks WHERE job_id = ? AND field_id = ?",
                      (job_id, field_id))
        conn.commit()
    finally:
        conn.close()


# ── Guided Review: API Routes ──

@app.route("/api/guided-review/queue/<job_id>")
@require_login
def guided_review_queue(job_id):
    """Return ordered review queue for guided review."""
    queue, reviewed_count = _build_guided_queue(job_id)
    return jsonify({
        "queue": queue,
        "total": len(queue) + reviewed_count,
        "reviewed": reviewed_count,
        "remaining": len(queue),
    })


@app.route("/api/guided-review/item/<job_id>/<path:field_id>")
@require_login
def guided_review_item(job_id, field_id):
    """Return full detail for one field, triggering evidence generation."""
    # Parse field_id: "page:extIdx:fieldName"
    parts = field_id.split(":")
    if len(parts) < 3:
        return jsonify({"error": "Invalid field_id"}), 400
    try:
        page_num = int(parts[0])
    except ValueError:
        return jsonify({"error": "Invalid page number"}), 400
    field_name = ":".join(parts[2:])

    # Load field data from extraction log
    log_data = _load_extraction_log(job_id)
    if not log_data:
        return jsonify({"error": "No extraction data"}), 404

    extractions = log_data.get("extractions", [])
    ext_idx = int(parts[1])
    target_ext = None
    # Find extraction for this page + index
    page_exts = [e for e in extractions if e.get("_page") == page_num]
    if ext_idx < len(page_exts):
        target_ext = page_exts[ext_idx]

    if not target_ext:
        return jsonify({"error": "Extraction not found"}), 404

    fdata = (target_ext.get("fields") or {}).get(field_name)
    if not fdata or not isinstance(fdata, dict):
        return jsonify({"error": "Field not found"}), 404

    value = fdata.get("value")
    entity = target_ext.get("payer_or_entity", "") or target_ext.get("employer_name", "") or ""
    doc_type = target_ext.get("document_type", "")

    # Generate evidence image
    evidence_path = _generate_evidence_image(job_id, page_num, value, field_id)
    safe_field_id = re.sub(r'[^\w\-.]', '_', str(field_id))
    evidence_url = f"/api/guided-review/evidence/{job_id}/{safe_field_id}.png" if evidence_path else None

    # Queue position removed — client tracks position locally from queue fetch
    return jsonify({
        "field_id": field_id,
        "field_name": field_name,
        "display_name": _field_display_name(doc_type, field_name),
        "value": value,
        "page_num": page_num,
        "document_type": doc_type,
        "entity": entity,
        "confidence": fdata.get("confidence", ""),
        "method": target_ext.get("_extraction_method", ""),
        "evidence_url": evidence_url,
        "evidence_available": evidence_path is not None,
        "page_url": f"/api/page-image/{job_id}/{page_num}",
        "position_in_queue": None,
        "queue_total": None,
    })


@app.route("/api/guided-review/evidence/<job_id>/<filename>")
@require_login
def guided_review_evidence(job_id, filename):
    """Serve a cached evidence highlight image."""
    # Sanitize filename to prevent path traversal
    safe_name = re.sub(r'[^\w\-.]', '_', filename)
    path = EVIDENCE_DIR / job_id / safe_name
    if path.exists():
        resp = send_file(str(path), mimetype="image/png")
        resp.headers["Cache-Control"] = "public, max-age=86400"
        return resp
    abort(404)


@app.route("/api/guided-review/action/<job_id>/<path:field_id>", methods=["POST"])
@require_login
def guided_review_action(job_id, field_id):
    """Process a guided review action on a single field.

    Body: { action: confirm|correct|not_present|skip, corrected_value?, note?, reviewer? }
    Returns: { ok: true, next: <next_item_or_null> }
    """
    payload = request.get_json(silent=True) or {}
    action = payload.get("action", "")

    # B1: Session-enforced reviewer identity
    u = current_user()
    reviewer = u["display_name"] if u else payload.get("reviewer", "")
    reviewer_id = u["id"] if u else None

    # B1: Check stage permission
    job = jobs.get(job_id)
    if job:
        stage = job.get("review_stage", "draft")
        user_role = u["role"] if u else "admin"
        if stage in STAGE_ROLE_MAP and not can_act_at_stage(user_role, stage):
            return jsonify({"error": f"Your role ({user_role}) cannot act at stage: {STAGE_DISPLAY.get(stage, stage)}"}), 403

    if action not in ("confirm", "correct", "not_present", "skip"):
        return jsonify({"error": "Invalid action"}), 400

    # Release the lock on this field
    _release_review_lock(job_id, field_id)

    if action != "skip":
        # Map guided review action to the existing verification format
        status_map = {
            "confirm": "confirmed",
            "correct": "corrected",
            "not_present": "flagged",
        }
        status = status_map[action]
        field_decision = {
            "status": status,
            "reviewer": reviewer,
            "reviewer_id": reviewer_id,
            "review_stage": job.get("review_stage", "") if job else "",
            "note": payload.get("note", ""),
        }
        # B9: Per-field time tracking
        field_dur = payload.get("field_duration_ms")
        if field_dur is not None:
            try:
                field_decision["field_duration_ms"] = int(field_dur)
            except (ValueError, TypeError):
                pass
        if action == "correct":
            field_decision["corrected_value"] = payload.get("corrected_value")
            # Delete cached evidence for this field (value changed)
            safe_fid = re.sub(r'[^\w\-.]', '_', str(field_id))
            evidence_file = EVIDENCE_DIR / job_id / f"{safe_fid}.png"
            if evidence_file.exists():
                try:
                    evidence_file.unlink()
                except OSError:
                    pass
        if action == "not_present":
            field_decision["note"] = payload.get("note", "") or "Value not present on document"

        incoming = {field_id: field_decision}

        # ── FAST PATH (T-UX-CONFIRM-FASTPATH): minimal synchronous writes ──

        # 1. Write verifications table (skip summary — deferred to aftercare)
        data = _load_verifications(job_id)
        data["reviewer"] = reviewer
        field_decision["timestamp"] = datetime.now().isoformat()
        data["fields"][field_id] = field_decision
        _save_verifications(job_id, data, skip_summary=True)

        # 2. Write verified_fields row only (skip canonical + FactStore — deferred)
        _upsert_verified_fields_fast(job_id, incoming)

        # Excel regen deferred to Finish Review (no auto-regen per field)

        # ── Lite: VerificationAction capture (feature-flagged) ──
        if os.environ.get("LITE_VERIFICATION_ENABLED"):
            try:
                from lite.adapters.oathledger import review_payload_to_action
                _va = review_payload_to_action(
                    payload, field_id=field_id, job_id=job_id,
                    review_stage=stage if job else "draft",
                )
                _va.reviewer_id = reviewer_id
                _persist_lite_event(job_id, "verification_action", _va.model_dump_json())
            except Exception:
                pass  # Non-fatal — Lite capture should never break production

        # ── AFTERCARE: defer heavy work to background thread ──
        _enqueue_aftercare({
            "job_id": job_id,
            "incoming": incoming,
            "reviewer": reviewer,
            "reviewer_id": reviewer_id,
            "action": action,
            "field_id": field_id,
            "mode": "guided",
        })

    # Return counts using cached total (no queue rebuild — client manages queue locally)
    cached_total = (job or {}).get("_guided_total_fields", 0)
    conn = _get_db()
    try:
        row = conn.execute(
            "SELECT COUNT(*) FROM verified_fields WHERE job_id = ? AND status IN ('confirmed','corrected')",
            (job_id,)).fetchone()
        reviewed = row[0] if row else 0
    finally:
        conn.close()
    total = max(cached_total, reviewed)

    return jsonify({
        "ok": True,
        "next": None,  # Client manages queue locally via splice
        "remaining": total - reviewed,
        "reviewed": reviewed,
        "total": total,
    })


@app.route("/api/guided-review/lock/<job_id>/<path:field_id>", methods=["POST"])
@require_login
def guided_review_lock(job_id, field_id):
    """Acquire or extend a review lock on a field."""
    payload = request.get_json(silent=True) or {}
    reviewer = payload.get("reviewer", "")
    if not reviewer:
        return jsonify({"error": "Reviewer required"}), 400
    success, holder = _acquire_review_lock(job_id, field_id, reviewer)
    if success:
        return jsonify({"ok": True, "locked_by": holder})
    return jsonify({"error": f"Field locked by {holder}"}), 409


@app.route("/api/guided-review/lock/<job_id>/<path:field_id>", methods=["DELETE"])
@require_login
def guided_review_unlock(job_id, field_id):
    """Release a review lock."""
    _release_review_lock(job_id, field_id)
    return jsonify({"ok": True})


# ─── B9: Review Time Tracking ────────────────────────────────────────────────

@app.route("/api/review-session/<job_id>/start", methods=["POST"])
@require_login
def review_session_start(job_id):
    """Start a review session for time tracking."""
    u = current_user()
    reviewer = u["display_name"] if u else "operator"
    reviewer_id = u["id"] if u else None
    now = datetime.now().isoformat()
    conn = _get_db()
    try:
        conn.execute(
            """INSERT INTO review_sessions
               (job_id, reviewer, reviewer_id, session_start, created_at)
               VALUES (?, ?, ?, ?, ?)""",
            (job_id, reviewer, reviewer_id, now, now)
        )
        conn.commit()
        session_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    except sqlite3.Error as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()
    return jsonify({"ok": True, "session_id": session_id, "session_start": now})


@app.route("/api/review-session/<job_id>/end", methods=["POST"])
@require_login
def review_session_end(job_id):
    """End a review session. Computes duration from session start."""
    payload = request.get_json(silent=True) or {}
    session_id = payload.get("session_id")
    fields_reviewed = payload.get("fields_reviewed", 0)
    now = datetime.now().isoformat()
    conn = _get_db()
    try:
        if session_id:
            row = conn.execute(
                "SELECT session_start FROM review_sessions WHERE id = ? AND job_id = ?",
                (session_id, job_id)
            ).fetchone()
        else:
            # Find most recent open session for this job
            row = conn.execute(
                "SELECT id, session_start FROM review_sessions WHERE job_id = ? AND session_end IS NULL ORDER BY id DESC LIMIT 1",
                (job_id,)
            ).fetchone()
            if row:
                session_id = row[0]
                row = (row[1],)
        if row:
            try:
                start_dt = datetime.fromisoformat(row[0])
                end_dt = datetime.fromisoformat(now)
                duration = int((end_dt - start_dt).total_seconds())
            except (ValueError, TypeError):
                duration = None
            conn.execute(
                """UPDATE review_sessions SET session_end = ?, duration_seconds = ?, fields_reviewed = ?
                   WHERE id = ?""",
                (now, duration, fields_reviewed, session_id)
            )
            conn.commit()
            return jsonify({"ok": True, "duration_seconds": duration})
        return jsonify({"ok": False, "error": "No open session found"}), 404
    except sqlite3.Error as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/review-session/<job_id>/timing", methods=["GET"])
@require_login
def review_session_timing(job_id):
    """Return aggregate review timing for a job — total time, per-field durations."""
    conn = _get_db()
    try:
        # Session totals
        rows = conn.execute(
            "SELECT SUM(duration_seconds), SUM(fields_reviewed), COUNT(*) FROM review_sessions WHERE job_id = ? AND duration_seconds IS NOT NULL",
            (job_id,)
        ).fetchone()
        total_seconds = rows[0] or 0
        total_fields = rows[1] or 0
        session_count = rows[2] or 0

        # Per-field durations (top 10 slowest)
        field_rows = conn.execute(
            """SELECT field_key, field_duration_ms, status, reviewer, verified_at
               FROM verified_fields
               WHERE job_id = ? AND field_duration_ms IS NOT NULL AND field_duration_ms > 0
               ORDER BY field_duration_ms DESC LIMIT 10""",
            (job_id,)
        ).fetchall()
        slowest = [{"field_key": r[0], "duration_ms": r[1], "status": r[2],
                     "reviewer": r[3], "verified_at": r[4]} for r in field_rows]

        # Average per-field
        avg_row = conn.execute(
            "SELECT AVG(field_duration_ms), COUNT(*) FROM verified_fields WHERE job_id = ? AND field_duration_ms IS NOT NULL AND field_duration_ms > 0",
            (job_id,)
        ).fetchone()
        avg_ms = int(avg_row[0]) if avg_row[0] else 0
        timed_fields = avg_row[1] or 0
    except sqlite3.Error:
        total_seconds = 0
        total_fields = 0
        session_count = 0
        slowest = []
        avg_ms = 0
        timed_fields = 0
    finally:
        conn.close()

    return jsonify({
        "total_review_seconds": total_seconds,
        "total_fields_reviewed": total_fields,
        "session_count": session_count,
        "avg_field_ms": avg_ms,
        "timed_fields": timed_fields,
        "slowest_fields": slowest,
    })


# ─── Post-Run Audit Sampling ─────────────────────────────────────────────────

def _compute_audit_sample_size(n):
    """Compute audit sample size from meaningful field count. Zero-inflation aware."""
    if n <= 20:
        return min(3, n)
    elif n <= 50:
        return 5
    elif n <= 120:
        return 8
    elif n <= 250:
        return 10
    else:
        return 12

def _is_meaningful_field(field_name, value, status):
    """Determine if a field is meaningful (not zero-inflated placeholder)."""
    # Always meaningful if explicitly confirmed or edited
    if status in ("confirmed", "corrected"):
        return True
    # Skip empty/null/placeholder values
    if value is None or value == "" or value == "(empty)":
        return False
    # Skip zero values for numeric fields (zero inflation)
    try:
        numval = float(str(value).replace(",", "").replace("$", "").replace("(", "-").replace(")", ""))
        if numval == 0.0:
            return False
    except (ValueError, TypeError):
        pass  # Non-numeric values like names, EINs — keep them
    return True


@app.route("/api/post-run-audit/sample/<job_id>")
@require_login
def post_run_audit_sample(job_id):
    """Generate a random audit sample from meaningful verified fields.

    Returns a list of field_ids to audit, plus metadata.
    Zero-inflation aware: excludes $0, null, placeholder fields unless explicitly confirmed/edited.
    """
    import random

    job = jobs.get(job_id)
    if not job or job.get("status") != "complete":
        return jsonify({"error": "Job not found or not complete"}), 404

    # Get all verified fields for this job
    vdata = _load_verifications(job_id)
    verified = vdata.get("fields", {})

    # Get extraction data to know field values
    log_path = job.get("output_log")
    extractions = []
    if log_path and os.path.exists(log_path):
        try:
            with open(log_path) as f:
                log_data = json.load(f)
            extractions = log_data.get("extractions", [])
        except (json.JSONDecodeError, IOError):
            pass

    # Build population of meaningful fields
    meaningful = []
    for ext_idx, ext in enumerate(extractions):
        page_num = ext.get("page_num", ext_idx + 1)
        fields = ext.get("fields", {})
        for fname, fval in fields.items():
            if fname.startswith("_"):
                continue
            field_id = f"{page_num}:{ext_idx}:{fname}"
            # Get verification status
            vfield = verified.get(field_id, {})
            status = vfield.get("status", "unverified")
            value = vfield.get("corrected_value", fval) if vfield.get("corrected_value") else fval
            # Get the display value
            if isinstance(fval, dict):
                value = fval.get("value", fval)
            if _is_meaningful_field(fname, value, status):
                meaningful.append({
                    "field_id": field_id,
                    "field_name": fname,
                    "value": value,
                    "status": status,
                    "page_num": page_num,
                })

    n = len(meaningful)
    sample_size = _compute_audit_sample_size(n)

    if n == 0:
        return jsonify({
            "sample": [],
            "sample_size": 0,
            "population_size": 0,
            "message": "No meaningful fields to audit",
        })

    sample = random.sample(meaningful, min(sample_size, n))

    return jsonify({
        "sample": sample,
        "sample_size": len(sample),
        "population_size": n,
    })


@app.route("/api/post-run-audit/result/<job_id>", methods=["POST"])
@require_login
def post_run_audit_result(job_id):
    """Store post-run audit results.

    Expects JSON: {
        "sample_size": N,
        "pass_count": N,
        "fail_count": N,
        "results": [{"field_id": "...", "outcome": "pass"|"flag", "note": "..."}],
        "reviewer": "..."
    }
    """
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    data = request.json or {}
    audit_record = {
        "job_id": job_id,
        "timestamp": datetime.now().isoformat(),
        "sample_size": data.get("sample_size", 0),
        "pass_count": data.get("pass_count", 0),
        "fail_count": data.get("fail_count", 0),
        "results": data.get("results", []),
        "reviewer": data.get("reviewer", ""),
    }

    # Store in job metadata
    job["post_run_audit"] = audit_record
    _save_jobs()

    # Any flagged items get routed back into the review queue
    flagged_ids = [r["field_id"] for r in audit_record["results"] if r.get("outcome") == "flag"]
    if flagged_ids:
        # Mark these fields as needing re-review by setting status to 'flagged'
        vdata = _load_verifications(job_id)
        for fid in flagged_ids:
            if fid in vdata.get("fields", {}):
                vdata["fields"][fid]["status"] = "flagged"
                vdata["fields"][fid]["note"] = (vdata["fields"][fid].get("note", "") +
                    " [AUDIT FLAG: " + next((r.get("note", "") for r in audit_record["results"] if r["field_id"] == fid), "") + "]").strip()
        _save_verifications(job_id, vdata)

    # Log event
    log_event("info", "post_run_audit",
              f"Audit complete: {audit_record['pass_count']}/{audit_record['sample_size']} passed",
              job_id=job_id,
              details_json=json.dumps(audit_record, default=str))

    return jsonify({
        "ok": True,
        "pass_count": audit_record["pass_count"],
        "fail_count": audit_record["fail_count"],
        "flagged_field_ids": flagged_ids,
    })


# ─── B1: Review Chain — Stage Transitions ────────────────────────────────────

@app.route("/api/jobs/<job_id>/submit-review", methods=["POST"])
@require_login
def submit_review(job_id):
    """Advance job to the next review stage.

    Validates current user's role can act at the current stage.
    Advances review_stage to next stage in chain.
    """
    u = current_user()
    if not u:
        return jsonify({"error": "Not authenticated"}), 401

    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    stage = job.get("review_stage", "draft")

    # Check user can act at this stage
    if not can_act_at_stage(u["role"], stage):
        return jsonify({"error": f"Your role ({u['role']}) cannot submit from stage: {STAGE_DISPLAY.get(stage, stage)}"}), 403

    next_stage = STAGE_NEXT.get(stage)
    if not next_stage:
        return jsonify({"error": f"No next stage from {STAGE_DISPLAY.get(stage, stage)}"}), 400

    _set_review_stage(job_id, next_stage, user=u)

    return jsonify({
        "ok": True,
        "old_stage": stage,
        "new_stage": next_stage,
        "display": STAGE_DISPLAY.get(next_stage, next_stage),
    })


@app.route("/api/jobs/<job_id>/send-back", methods=["POST"])
@require_login
def send_back_review(job_id):
    """Send job back to the previous review stage.

    Only reviewer/partner (or admin) can send back. Requires a reason.
    """
    u = current_user()
    if not u:
        return jsonify({"error": "Not authenticated"}), 401

    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    stage = job.get("review_stage", "draft")
    payload = request.get_json(silent=True) or {}
    reason = (payload.get("reason") or "").strip()

    if not reason:
        return jsonify({"error": "A reason is required when sending back"}), 400

    # Check user can act at this stage
    if not can_act_at_stage(u["role"], stage):
        return jsonify({"error": f"Your role ({u['role']}) cannot send back from stage: {STAGE_DISPLAY.get(stage, stage)}"}), 403

    prev_stage = STAGE_PREV.get(stage)
    if not prev_stage:
        return jsonify({"error": f"Cannot send back from {STAGE_DISPLAY.get(stage, stage)}"}), 400

    _set_review_stage(job_id, prev_stage, user=u)

    # Log the send-back with reason
    log_event("info", "review_stage_returned",
              f"Job {job_id[:12]}: sent back to {STAGE_DISPLAY.get(prev_stage, prev_stage)} — {reason}",
              user_id=u["id"],
              job_id=job_id,
              details={"old_stage": stage, "new_stage": prev_stage, "reason": reason})

    return jsonify({
        "ok": True,
        "old_stage": stage,
        "new_stage": prev_stage,
        "display": STAGE_DISPLAY.get(prev_stage, prev_stage),
    })


@app.route("/api/jobs/<job_id>/stage")
@require_login
def get_job_stage(job_id):
    """Return the current review stage for a job."""
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    stage = job.get("review_stage", "draft")
    u = current_user()
    user_role = u["role"] if u else "admin"
    return jsonify({
        "stage": stage,
        "display": STAGE_DISPLAY.get(stage, stage),
        "can_act": can_act_at_stage(user_role, stage),
        "can_submit": stage in STAGE_NEXT,
        "can_send_back": stage in STAGE_PREV,
    })


# ─── B1: Inbox Endpoint ─────────────────────────────────────────────────────

@app.route("/api/inbox")
@require_login
def api_inbox():
    """Return jobs assigned to the current user's review stage.

    admin sees all non-final jobs.
    preparer sees preparer_review jobs.
    reviewer sees reviewer_review jobs.
    partner sees partner_review jobs.
    """
    u = current_user()
    if not u:
        return jsonify({"inbox": []})

    role = u["role"]
    inbox = []

    for jid, j in jobs.items():
        if j.get("status") != "complete":
            continue
        stage = j.get("review_stage", "draft")
        if stage == "final":
            continue

        # Filter by role
        if role == "admin":
            pass  # admin sees everything
        elif role == "preparer" and stage != "preparer_review":
            continue
        elif role == "reviewer" and stage != "reviewer_review":
            continue
        elif role == "partner" and stage != "partner_review":
            continue

        inbox.append({
            "job_id": jid,
            "filename": j.get("filename", ""),
            "client_name": j.get("client_name", ""),
            "doc_type": j.get("doc_type", ""),
            "review_stage": stage,
            "stage_display": STAGE_DISPLAY.get(stage, stage),
            "created": j.get("start_time", ""),
            "stage_updated": j.get("stage_updated", ""),
        })

    # Sort by stage_updated descending (newest first)
    inbox.sort(key=lambda x: x.get("stage_updated") or x.get("created") or "", reverse=True)

    return jsonify({"inbox": inbox, "count": len(inbox)})


# ─── End Guided Review Backend ───────────────────────────────────────────────

@app.route("/api/download/<job_id>")
@require_login
def download_xlsx(job_id):
    job = jobs.get(job_id)
    if not job or not job.get("output_xlsx"):
        abort(404)
    p = job["output_xlsx"]
    if os.path.exists(p):
        original_stem = Path(job.get("filename", "")).stem or job_id
        friendly = re.sub(r'[^\w\s\-\.,()]', '', original_stem).strip() or job_id
        return send_file(p, as_attachment=True, download_name=friendly + "_intake.xlsx")
    abort(404)

@app.route("/api/download-log/<job_id>")
@require_login
def download_log(job_id):
    job = jobs.get(job_id)
    if not job or not job.get("output_log"):
        abort(404)
    p = job["output_log"]
    if os.path.exists(p):
        original_stem = Path(job.get("filename", "")).stem or job_id
        friendly = re.sub(r'[^\w\s\-\.,()]', '', original_stem).strip() or job_id
        return send_file(p, as_attachment=True, download_name=friendly + "_intake_log.json")
    abort(404)

def _sanitize_job(j):
    """Make a JSON-safe copy of a job dict (no None keys, no log/pdf_path)."""
    safe = {}
    for k, v in j.items():
        if k in ("log", "pdf_path"):
            continue
        if k == "stats" and isinstance(v, dict):
            v = dict(v)
            for sk in ("methods", "confidences"):
                if sk in v and isinstance(v[sk], dict):
                    v[sk] = {str(dk) if dk is None else dk: dv for dk, dv in v[sk].items()}
        safe[k] = v
    return safe

@app.route("/api/jobs")
@require_login
def list_jobs():
    q = request.args.get("q", "").strip().lower()
    dtype = request.args.get("doc_type", "").strip()

    # B9: Batch-fetch review session totals for all jobs
    review_times = {}
    try:
        conn = _get_db()
        rows = conn.execute(
            "SELECT job_id, SUM(duration_seconds) FROM review_sessions WHERE duration_seconds IS NOT NULL GROUP BY job_id"
        ).fetchall()
        for r in rows:
            review_times[r[0]] = r[1]
        conn.close()
    except sqlite3.Error:
        pass

    out = []
    for j in sorted(jobs.values(), key=lambda x: x.get("created", ""), reverse=True):
        if q and q not in j.get("client_name", "").lower() and q not in j.get("filename", "").lower():
            continue
        if dtype and j.get("doc_type", "") != dtype:
            continue
        job_out = _sanitize_job(j)
        # Compute duration in seconds
        st = j.get("start_time")
        et = j.get("end_time")
        if st and et:
            try:
                start_dt = datetime.fromisoformat(st)
                end_dt = datetime.fromisoformat(et)
                job_out["duration_seconds"] = int((end_dt - start_dt).total_seconds())
            except (ValueError, TypeError):
                job_out["duration_seconds"] = None
        elif st and j.get("status") == "running":
            try:
                start_dt = datetime.fromisoformat(st)
                job_out["duration_seconds"] = int((datetime.now() - start_dt).total_seconds())
            except (ValueError, TypeError):
                job_out["duration_seconds"] = None
        else:
            job_out["duration_seconds"] = None
        # B9: Include review time
        job_out["review_time_seconds"] = review_times.get(j.get("id"))
        out.append(job_out)
    return jsonify(out)

@app.route("/api/delete/<job_id>", methods=["POST"])
def delete_job(job_id):
    if job_id in jobs:
        del jobs[job_id]
        save_jobs()
        # Remove from SQLite (verifications)
        conn = _get_db()
        try:
            conn.execute("DELETE FROM jobs WHERE id = ?", (job_id,))
            conn.execute("DELETE FROM verifications WHERE job_id = ?", (job_id,))
            conn.execute("DELETE FROM verified_fields WHERE job_id = ?", (job_id,))
            conn.commit()
        except sqlite3.Error:
            pass
        finally:
            conn.close()
        # Clean review locks + B9 review sessions
        try:
            conn2 = _get_db()
            conn2.execute("DELETE FROM review_locks WHERE job_id = ?", (job_id,))
            conn2.execute("DELETE FROM review_sessions WHERE job_id = ?", (job_id,))
            conn2.commit()
            conn2.close()
        except sqlite3.Error:
            pass
        # Clean up page images
        job_pages = PAGES_DIR / job_id
        if job_pages.exists():
            import shutil
            shutil.rmtree(str(job_pages), ignore_errors=True)
        # Clean up evidence images
        job_evidence = EVIDENCE_DIR / job_id
        if job_evidence.exists():
            import shutil
            shutil.rmtree(str(job_evidence), ignore_errors=True)
    return jsonify({"ok": True})

# ─── Verification ────────────────────────────────────────────────────────────

def _verify_path(job_id):
    return VERIFY_DIR / f"{job_id}.json"

def _load_verifications(job_id):
    conn = _get_db()
    try:
        row = conn.execute("SELECT data FROM verifications WHERE job_id = ?", (job_id,)).fetchone()
        if row:
            return json.loads(row[0])
    except (sqlite3.Error, json.JSONDecodeError):
        pass
    finally:
        conn.close()
    return {"fields": {}, "updated": None, "reviewer": ""}

def _save_verifications(job_id, data, skip_summary=False):
    data["updated"] = datetime.now().isoformat()
    conn = _get_db()
    try:
        conn.execute(
            """INSERT INTO verifications (job_id, data, updated) VALUES (?, ?, ?)
               ON CONFLICT(job_id) DO UPDATE SET data = excluded.data, updated = excluded.updated""",
            (job_id, json.dumps(data, default=str), data["updated"])
        )
        conn.commit()
    except sqlite3.Error as e:
        print(f"  Warning: Could not save verifications for {job_id}: {e}")
    finally:
        conn.close()
    # Update job-level verification summary (skipped in fast path — deferred to aftercare)
    if not skip_summary:
        _update_verify_summary(job_id, data)

def _update_verify_summary(job_id, vdata):
    job = jobs.get(job_id)
    if not job:
        return
    fields = vdata.get("fields", {})
    total = len(fields)
    confirmed = sum(1 for f in fields.values() if f.get("status") == "confirmed")
    corrected = sum(1 for f in fields.values() if f.get("status") == "corrected")
    flagged = sum(1 for f in fields.values() if f.get("status") == "flagged")
    job["verification"] = {
        "reviewed": confirmed + corrected + flagged,
        "confirmed": confirmed,
        "corrected": corrected,
        "flagged": flagged,
        "reviewer": vdata.get("reviewer", ""),
        "updated": vdata.get("updated"),
    }
    save_jobs()


def _resolve_field_value(log_data, field_key):
    """Given an extraction log and a field key like '3:0:wages', return the extracted value."""
    parts = field_key.split(":")
    if len(parts) != 3:
        return None
    page_str, ext_idx_str, field_name = parts
    try:
        page_num = int(page_str)
        ext_idx = int(ext_idx_str)
    except ValueError:
        return None
    extractions = log_data.get("extractions", [])
    page_exts = [e for e in extractions if e.get("_page") == page_num]
    if ext_idx >= len(page_exts):
        return None
    ext = page_exts[ext_idx]
    fields = ext.get("fields", {})
    fdata = fields.get(field_name)
    if fdata is None:
        return None
    return fdata.get("value") if isinstance(fdata, dict) else fdata


_extraction_log_cache = {}  # job_id -> (mtime, data)

def _load_extraction_log(job_id):
    """Load the raw extraction log JSON for a job. Returns dict or None.
    Caches in memory by mtime to avoid redundant disk reads."""
    job = jobs.get(job_id)
    if not job:
        return None
    log_path = job.get("output_log")
    if not log_path or not os.path.exists(log_path):
        return None
    try:
        mtime = os.path.getmtime(log_path)
        cached = _extraction_log_cache.get(job_id)
        if cached and cached[0] == mtime:
            return cached[1]
        with open(log_path) as f:
            data = json.load(f)
        _extraction_log_cache[job_id] = (mtime, data)
        return data
    except (json.JSONDecodeError, IOError, OSError):
        return None


def _resolve_extraction_for_field(log_data, field_key):
    """Given a field_key like '3:0:wages', return the parent extraction dict.

    Returns (extraction_dict, field_name) or (None, None).
    """
    parts = field_key.split(":")
    if len(parts) != 3:
        return None, None
    page_str, ext_idx_str, field_name = parts
    try:
        page_num = int(page_str)
        ext_idx = int(ext_idx_str)
    except ValueError:
        return None, None
    extractions = log_data.get("extractions", [])
    page_exts = [e for e in extractions if e.get("_page") == page_num]
    if ext_idx >= len(page_exts):
        return None, None
    return page_exts[ext_idx], field_name


def _upsert_client_canonical(conn, client_name, year, document_type, payer_key,
                              payer_display, field_name, canonical_value,
                              original_value, status, job_id, reviewer, verified_at,
                              evidence_ref='', source_doc='', page_number=None):
    """Upsert a single field into the client-level canonical store.

    Newer verifications always overwrite older ones.
    T1.6: Extended with evidence_ref, source_doc, page_number for workpaper support.
    """
    conn.execute(
        """INSERT INTO client_canonical_values
           (client_name, year, document_type, payer_key, payer_display,
            field_name, canonical_value, original_value, status,
            source_job_id, reviewer, verified_at,
            evidence_ref, source_doc, page_number)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
           ON CONFLICT(client_name, year, document_type, payer_key, field_name)
           DO UPDATE SET
               canonical_value = excluded.canonical_value,
               original_value = excluded.original_value,
               payer_display = excluded.payer_display,
               status = excluded.status,
               source_job_id = excluded.source_job_id,
               reviewer = excluded.reviewer,
               verified_at = excluded.verified_at,
               evidence_ref = excluded.evidence_ref,
               source_doc = excluded.source_doc,
               page_number = excluded.page_number""",
        (client_name, year, document_type, payer_key, payer_display,
         field_name,
         json.dumps(canonical_value) if canonical_value is not None else None,
         json.dumps(original_value) if original_value is not None else None,
         status, job_id, reviewer, verified_at,
         evidence_ref, source_doc, page_number)
    )


def _populate_facts_from_extraction(job):
    """T1.6.2: Populate unified facts table from extraction results.

    Chain of custody: extraction results are written to the unified facts
    table immediately. The FactStore enforces monotonic trust — corrected and
    confirmed facts are never overwritten by automation.

    Also syncs to the legacy client_canonical_values table for workpaper compat.

    Only runs when client_name and year are available.
    """
    from fact_store import FactStore

    client_name = job.get("client_name") or ""
    year = job.get("year") or ""
    job_id = job.get("id") or ""
    filename = job.get("filename") or ""

    if not client_name or not year:
        return

    log_data = _load_extraction_log(job_id)
    if not log_data:
        return

    exts = log_data.get("extractions", [])
    if not exts:
        return

    # Status mapping: high-confidence extractions get auto_verified
    AUTO_VERIFIED_CONFS = frozenset({
        "dual_confirmed", "verified_confirmed", "auto_verified",
        "consensus_accepted", "multipage_verified",
    })

    fs = FactStore(str(DB_PATH))
    try:
        tax_year = int(year) if year.isdigit() else None
    except (ValueError, AttributeError):
        tax_year = None

    count = 0
    for ext in exts:
        payer_key = _normalize_payer_key(ext)
        doc_type = ext.get("document_type") or "unknown"
        page = ext.get("_page")
        extraction_method = ext.get("_extraction_method") or ""

        fields = ext.get("fields") or {}
        for fname, fdata in fields.items():
            if fname.startswith("_"):
                continue
            if not isinstance(fdata, dict):
                continue

            value = fdata.get("value")
            confidence_str = fdata.get("confidence") or ""

            # Determine status
            if confidence_str in AUTO_VERIFIED_CONFS:
                status = "auto_verified"
            else:
                status = "extracted"

            # Build fact_key: "doc_type.payer_key.field_name"
            fact_key = FactStore.fact_key(doc_type, payer_key, fname)

            # Parse numeric value
            value_num = None
            value_text = None
            if value is not None:
                try:
                    value_num = float(str(value).replace(",", "").replace("$", ""))
                except (ValueError, TypeError):
                    value_text = str(value)

            # Determine source_method from extraction method
            source_method = extraction_method
            if "ocr" in extraction_method.lower():
                source_method = "ocr"
            elif "vision" in extraction_method.lower():
                source_method = "vision"
            elif "text" in extraction_method.lower():
                source_method = "text_layer"

            # Parse confidence to float
            conf_num = None
            if confidence_str:
                conf_map = {
                    "high": 0.9, "medium": 0.7, "low": 0.4,
                    "auto_verified": 0.95, "dual_confirmed": 0.99,
                    "verified_confirmed": 0.98, "consensus_accepted": 0.97,
                    "multipage_verified": 0.96,
                }
                conf_num = conf_map.get(confidence_str)

            # Write to unified facts table (respects lock rules automatically)
            try:
                fs.upsert_candidate_fact(
                    job_id=job_id, client_id=client_name,
                    tax_year=tax_year, fact_key=fact_key,
                    value_num=value_num, value_text=value_text,
                    status=status, confidence=conf_num,
                    source_method=source_method,
                    source_doc=filename, source_page=page,
                    evidence_ref=confidence_str
                )
                count += 1
            except (ValueError, sqlite3.Error):
                pass  # Skip invalid values silently

    # Sync to legacy table for workpaper compatibility
    try:
        fs.sync_to_legacy(job_id, client_name, str(year))
    except Exception:
        pass  # Non-fatal

    if count > 0:
        job.setdefault("log", []).append(
            f"  T1.6.2: Populated {count} facts for {client_name} / {year}"
        )


def _upsert_verified_fields(job_id, incoming_fields):
    """Write individual field decisions to the normalized verified_fields table.

    For each field:
      confirmed → canonical_value = original extracted value
      corrected → canonical_value = corrected value
      flagged   → canonical_value = NULL
      _remove   → DELETE from table
    """
    if not incoming_fields:
        return
    conn = _get_db()
    try:
        # Load extraction log once for resolving confirmed values
        log_data = _load_extraction_log(job_id)
        now = datetime.now().isoformat()

        for field_key, decision in incoming_fields.items():
            status = decision.get("status", "")

            if status == "_remove":
                conn.execute(
                    "DELETE FROM verified_fields WHERE job_id = ? AND field_key = ?",
                    (job_id, field_key)
                )
                continue

            if status not in ("confirmed", "corrected", "flagged"):
                continue

            canonical = None
            original = None

            if status == "corrected":
                canonical = decision.get("corrected_value")
                if log_data:
                    original = _resolve_field_value(log_data, field_key)
            elif status == "confirmed":
                if log_data:
                    original = _resolve_field_value(log_data, field_key)
                    canonical = original
            # flagged: canonical stays None

            conn.execute(
                """INSERT INTO verified_fields
                   (job_id, field_key, canonical_value, original_value, status,
                    category, vendor_desc, note, reviewer, verified_at,
                    review_stage, reviewer_id, field_duration_ms)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                   ON CONFLICT(job_id, field_key) DO UPDATE SET
                       canonical_value = excluded.canonical_value,
                       original_value = excluded.original_value,
                       status = excluded.status,
                       category = excluded.category,
                       vendor_desc = excluded.vendor_desc,
                       note = excluded.note,
                       reviewer = excluded.reviewer,
                       verified_at = excluded.verified_at,
                       review_stage = excluded.review_stage,
                       reviewer_id = excluded.reviewer_id,
                       field_duration_ms = excluded.field_duration_ms""",
                (job_id, field_key,
                 json.dumps(canonical) if canonical is not None else None,
                 json.dumps(original) if original is not None else None,
                 status,
                 decision.get("category", ""),
                 decision.get("vendor_desc", ""),
                 decision.get("note", ""),
                 decision.get("reviewer", decision.get("reviewer", "")),
                 decision.get("timestamp", now),
                 decision.get("review_stage", ""),
                 decision.get("reviewer_id"),
                 decision.get("field_duration_ms"))
            )

        # Promote to client-level canonical store + unified facts table
        job = jobs.get(job_id)
        if job and log_data:
            client_name = job.get("client_name", "")
            year = job.get("year", "")
            if client_name and year:
                # T1.6.2: Also update unified facts table
                from fact_store import FactStore
                try:
                    fs = FactStore(str(DB_PATH))
                    tax_year = int(year) if year.isdigit() else None
                except Exception:
                    fs = None
                    tax_year = None

                for field_key, decision in incoming_fields.items():
                    status = decision.get("status", "")
                    if status not in ("confirmed", "corrected"):
                        continue
                    ext, fn = _resolve_extraction_for_field(log_data, field_key)
                    if not ext or not fn:
                        continue
                    doc_type = ext.get("document_type", "")
                    if not doc_type:
                        continue
                    payer_key = _normalize_payer_key(ext)
                    payer_display = ext.get("payer_or_entity", "")

                    canonical = None
                    original = _resolve_field_value(log_data, field_key)
                    if status == "corrected":
                        canonical = decision.get("corrected_value")
                    elif status == "confirmed":
                        canonical = original

                    if canonical is not None:
                        _upsert_client_canonical(
                            conn, client_name, year, doc_type, payer_key,
                            payer_display, fn, canonical, original,
                            status, job_id,
                            decision.get("reviewer", ""),
                            decision.get("timestamp", now)
                        )

                    # T1.6.2: Update unified facts table
                    if fs and tax_year is not None:
                        fact_key = FactStore.fact_key(doc_type, payer_key, fn)
                        try:
                            if status == "corrected" and canonical is not None:
                                # Parse corrected value
                                corr_num = None
                                corr_text = None
                                try:
                                    corr_num = float(str(canonical).replace(",", "").replace("$", ""))
                                except (ValueError, TypeError):
                                    corr_text = str(canonical)
                                fs.apply_correction(
                                    job_id, tax_year, fact_key,
                                    value_num=corr_num, value_text=corr_text,
                                    reviewer=decision.get("reviewer", "")
                                )
                            elif status == "confirmed":
                                fs.upgrade_fact_status(
                                    job_id, tax_year, fact_key, "confirmed"
                                )
                        except Exception:
                            pass  # Non-fatal — legacy table is still updated

        conn.commit()
    except sqlite3.Error as e:
        print(f"  Warning: Could not upsert verified_fields for {job_id}: {e}")
    finally:
        conn.close()


def _upsert_verified_fields_fast(job_id, incoming_fields):
    """Fast-path: write verified_fields rows only. Skips canonical + FactStore (deferred to aftercare).

    This is the synchronous hot-path version of _upsert_verified_fields. It writes only the
    verified_fields table (the source of truth for the review UI) and returns immediately.
    Canonical promotion and FactStore updates happen in the aftercare background thread.
    """
    if not incoming_fields:
        return
    conn = _get_db()
    try:
        log_data = _load_extraction_log(job_id)
        now = datetime.now().isoformat()

        for field_key, decision in incoming_fields.items():
            status = decision.get("status", "")
            if status == "_remove":
                conn.execute(
                    "DELETE FROM verified_fields WHERE job_id = ? AND field_key = ?",
                    (job_id, field_key))
                continue
            if status not in ("confirmed", "corrected", "flagged"):
                continue

            canonical = None
            original = None
            if status == "corrected":
                canonical = decision.get("corrected_value")
                if log_data:
                    original = _resolve_field_value(log_data, field_key)
            elif status == "confirmed":
                if log_data:
                    original = _resolve_field_value(log_data, field_key)
                    canonical = original
            # flagged: canonical stays None

            conn.execute(
                """INSERT INTO verified_fields
                   (job_id, field_key, canonical_value, original_value, status,
                    category, vendor_desc, note, reviewer, verified_at,
                    review_stage, reviewer_id)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                   ON CONFLICT(job_id, field_key) DO UPDATE SET
                       canonical_value = excluded.canonical_value,
                       original_value = excluded.original_value,
                       status = excluded.status,
                       category = excluded.category,
                       vendor_desc = excluded.vendor_desc,
                       note = excluded.note,
                       reviewer = excluded.reviewer,
                       verified_at = excluded.verified_at,
                       review_stage = excluded.review_stage,
                       reviewer_id = excluded.reviewer_id""",
                (job_id, field_key,
                 json.dumps(canonical) if canonical is not None else None,
                 json.dumps(original) if original is not None else None,
                 status,
                 decision.get("category", ""),
                 decision.get("vendor_desc", ""),
                 decision.get("note", ""),
                 decision.get("reviewer", ""),
                 decision.get("timestamp", now),
                 decision.get("review_stage", ""),
                 decision.get("reviewer_id")))
        conn.commit()
    except sqlite3.Error as e:
        print(f"  Warning: Could not fast-upsert verified_fields for {job_id}: {e}")
    finally:
        conn.close()


def get_verified_extractions(job_id):
    """Load extraction log and overlay all verified field values.

    Two-tier resolution:
      1. Job-level: verified_fields table (per-job confirmations/corrections)
      2. Client-level: client_canonical_values table (cross-job canonical truth)

    Job-level always takes precedence. Client-level fills in unverified fields
    when the same client/year/payer/field was verified in another job.

    Returns the full log_data dict with verified values replacing raw values.
    Returns None if no extraction log exists.
    """
    import copy

    job = jobs.get(job_id)
    log_data = _load_extraction_log(job_id)
    if not log_data:
        return None

    extractions = log_data.get("extractions", [])
    if not extractions:
        return log_data

    conn = _get_db()
    try:
        # Load job-level verified fields
        job_rows = conn.execute(
            "SELECT field_key, canonical_value, status, category, vendor_desc "
            "FROM verified_fields WHERE job_id = ?",
            (job_id,)
        ).fetchall()

        # Load client-level canonical values
        client_canonicals = {}
        if job:
            client_name = job.get("client_name", "")
            year = job.get("year", "")
            if client_name and year:
                cc_rows = conn.execute(
                    "SELECT document_type, payer_key, field_name, canonical_value, status "
                    "FROM client_canonical_values WHERE client_name = ? AND year = ?",
                    (client_name, year)
                ).fetchall()
                for doc_type, payer_key, field_name, canon_json, status in cc_rows:
                    cc_key = (doc_type, payer_key, field_name)
                    canonical = None
                    if canon_json is not None:
                        try:
                            canonical = json.loads(canon_json)
                        except (json.JSONDecodeError, ValueError):
                            canonical = canon_json
                    client_canonicals[cc_key] = {
                        "canonical": canonical,
                        "status": status,
                    }
    except sqlite3.Error:
        job_rows = []
        client_canonicals = {}
    finally:
        conn.close()

    if not job_rows and not client_canonicals:
        return log_data  # No verifications at all — return raw

    # Build job-level lookup: field_key → {canonical, status, category, vendor_desc}
    verified = {}
    for field_key, canonical_json, status, category, vendor_desc in job_rows:
        canonical = None
        if canonical_json is not None:
            try:
                canonical = json.loads(canonical_json)
            except (json.JSONDecodeError, ValueError):
                canonical = canonical_json
        verified[field_key] = {
            "canonical": canonical,
            "status": status,
            "category": category or "",
            "vendor_desc": vendor_desc or "",
        }

    # Deep copy extractions so we don't mutate the cached log data
    extractions = copy.deepcopy(extractions)

    # Build page-index mapping
    page_groups = {}
    for ext in extractions:
        p = ext.get("_page")
        if p is not None:
            page_groups.setdefault(p, []).append(ext)

    # Pass 1: Overlay job-level verified values
    # Track which fields were handled at job level
    job_verified_fields = set()

    for page, page_exts in page_groups.items():
        for ext_idx, ext in enumerate(page_exts):
            fields = ext.get("fields")
            if not fields:
                continue
            for field_name in list(fields.keys()):
                fk = f"{page}:{ext_idx}:{field_name}"
                vf = verified.get(fk)
                if not vf:
                    continue

                job_verified_fields.add(fk)
                fdata = fields[field_name]
                canonical = vf["canonical"]

                if canonical is not None:
                    try:
                        canonical = float(str(canonical).replace(",", ""))
                    except (ValueError, TypeError):
                        pass

                    if isinstance(fdata, dict):
                        fdata["_original_value"] = fdata.get("value")
                        fdata["value"] = canonical
                        fdata["confidence"] = "operator_corrected" if vf["status"] == "corrected" else "operator_confirmed"
                    else:
                        fields[field_name] = {
                            "value": canonical,
                            "_original_value": fdata,
                            "confidence": "operator_corrected" if vf["status"] == "corrected" else "operator_confirmed",
                        }

                if vf.get("category"):
                    if isinstance(fields[field_name], dict):
                        fields[field_name]["_operator_category"] = vf["category"]
                if vf.get("vendor_desc"):
                    if isinstance(fields[field_name], dict):
                        fields[field_name]["_vendor_desc"] = vf["vendor_desc"]

    # Pass 2: Overlay client-level canonicals for unverified fields
    if client_canonicals:
        for page, page_exts in page_groups.items():
            for ext_idx, ext in enumerate(page_exts):
                fields = ext.get("fields")
                if not fields:
                    continue
                doc_type = ext.get("document_type", "")
                if not doc_type:
                    continue
                payer_key = _normalize_payer_key(ext)

                for field_name in list(fields.keys()):
                    fk = f"{page}:{ext_idx}:{field_name}"
                    if fk in job_verified_fields:
                        continue  # Job-level takes precedence

                    cc_key = (doc_type, payer_key, field_name)
                    cc = client_canonicals.get(cc_key)
                    if not cc:
                        continue

                    canonical = cc["canonical"]
                    if canonical is None:
                        continue

                    try:
                        canonical = float(str(canonical).replace(",", ""))
                    except (ValueError, TypeError):
                        pass

                    fdata = fields[field_name]
                    if isinstance(fdata, dict):
                        fdata["_original_value"] = fdata.get("value")
                        fdata["value"] = canonical
                        fdata["confidence"] = "client_canonical"
                    else:
                        fields[field_name] = {
                            "value": canonical,
                            "_original_value": fdata,
                            "confidence": "client_canonical",
                        }

    log_data["extractions"] = extractions
    return log_data


@app.route("/api/verify/<job_id>", methods=["GET"])
@require_login
def get_verifications(job_id):
    return jsonify(_load_verifications(job_id))

@app.route("/api/verify/<job_id>", methods=["POST"])
@require_login
def save_verification(job_id):
    """Save one or more field verification decisions.

    Body: { "fields": { "page:extIdx:fieldName": { "status": "confirmed"|"corrected"|"flagged",
                                                     "corrected_value": ..., "note": ...,
                                                     "category": "Utilities",
                                                     "vendor_desc": "Georgia Power" } },
            "reviewer": "JW" }

    If a category and vendor_desc are present, also learns the mapping
    for future auto-suggest.
    """
    payload = request.get_json(silent=True) or {}
    data = _load_verifications(job_id)

    # B1: Session-enforced reviewer identity
    u = current_user()
    reviewer = u["display_name"] if u else payload.get("reviewer", data.get("reviewer", ""))
    reviewer_id = u["id"] if u else None
    data["reviewer"] = reviewer

    # B1: Check stage permission
    job = jobs.get(job_id)
    if job:
        stage = job.get("review_stage", "draft")
        user_role = u["role"] if u else "admin"
        if stage in STAGE_ROLE_MAP and not can_act_at_stage(user_role, stage):
            return jsonify({"error": f"Your role ({user_role}) cannot act at stage: {STAGE_DISPLAY.get(stage, stage)}"}), 403

    for key, decision in (payload.get("fields") or {}).items():
        if decision.get("status") == "_remove":
            data["fields"].pop(key, None)
        else:
            decision["timestamp"] = datetime.now().isoformat()
            decision["reviewer"] = reviewer
            decision["reviewer_id"] = reviewer_id
            decision["review_stage"] = job.get("review_stage", "") if job else ""
            data["fields"][key] = decision

            # Learn vendor → category mapping
            cat = decision.get("category", "")
            vendor = decision.get("vendor_desc", "")
            if cat and vendor:
                _learn_vendor_category(vendor, cat)

    # ── FAST PATH (T-UX-CONFIRM-FASTPATH): minimal synchronous writes ──
    incoming_fields = payload.get("fields") or {}
    _save_verifications(job_id, data, skip_summary=True)
    _upsert_verified_fields_fast(job_id, incoming_fields)

    # Excel regen deferred to Finish Review (no auto-regen per field)

    # ── AFTERCARE: defer heavy work to background thread ──
    statuses = {}
    for _fk, _fd in incoming_fields.items():
        s = _fd.get("status", "unknown")
        statuses[s] = statuses.get(s, 0) + 1

    _enqueue_aftercare({
        "job_id": job_id,
        "incoming": incoming_fields,
        "reviewer": reviewer,
        "reviewer_id": reviewer_id,
        "field_count": len(incoming_fields),
        "statuses": statuses,
        "mode": "grid",
    })

    return jsonify({"ok": True, "total_reviewed": len(data["fields"])})


@app.route("/api/vendor-categories", methods=["GET"])
def get_vendor_categories():
    """Return the learned vendor → category map and the chart of accounts."""
    return jsonify({
        "vendors": _load_vendor_categories(),
        "chart_of_accounts": CHART_OF_ACCOUNTS,
    })


@app.route("/api/suggest-categories", methods=["POST"])
def suggest_categories():
    """Given a list of vendor descriptions, return category suggestions.

    Body: { "descriptions": ["GEORGIA POWER #1234", "WALMART SUPERCENTER", ...] }
    Returns: { "suggestions": {"GEORGIA POWER #1234": "Utilities", ...} }
    """
    payload = request.get_json(silent=True) or {}
    descriptions = payload.get("descriptions", [])
    suggestions = {}
    for desc in descriptions:
        cat = _suggest_category(desc)
        if cat:
            suggestions[desc] = cat
    return jsonify({"suggestions": suggestions})


# ─── Client Management Routes ────────────────────────────────────────────────

@app.route("/api/clients")
def list_clients():
    """List all known clients with basic stats."""
    clients = {}
    # Gather from jobs
    for jid, j in jobs.items():
        cn = j.get("client_name", "")
        if not cn:
            continue
        safe = _safe_client_name(cn)
        if safe not in clients:
            clients[safe] = {"name": safe, "jobs": 0, "latest": "", "years": set(),
                             "has_context": False, "has_instructions": False}
        clients[safe]["jobs"] += 1
        clients[safe]["years"].add(str(j.get("year", "")))
        ts = j.get("created", "")
        if ts > clients[safe]["latest"]:
            clients[safe]["latest"] = ts

    # Check for context and instructions files
    if CLIENTS_DIR.exists():
        for d in CLIENTS_DIR.iterdir():
            if d.is_dir():
                name = d.name
                if name not in clients:
                    clients[name] = {"name": name, "jobs": 0, "latest": "",
                                     "years": set(), "has_context": False, "has_instructions": False}
                ctx_idx = d / "context" / "index.json"
                if ctx_idx.exists():
                    clients[name]["has_context"] = True
                instr = d / "instructions.json"
                if instr.exists():
                    clients[name]["has_instructions"] = True

    result = []
    for c in sorted(clients.values(), key=lambda x: x.get("latest", ""), reverse=True):
        c["years"] = sorted(c["years"])
        # Include client metadata if available
        info = _load_client_info(c["name"])
        if info:
            c["ein_last4"] = info.get("ein_last4", "")
            c["contact"] = info.get("contact", "")
            c["notes"] = info.get("notes", "")
        result.append(c)
    return jsonify(result)


@app.route("/api/clients/create", methods=["POST"])
def create_client():
    """Create a new client with metadata."""
    data = request.get_json(force=True)
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Client name is required"}), 400
    safe = _safe_client_name(name)
    # Create directory
    client_dir = CLIENTS_DIR / safe
    client_dir.mkdir(parents=True, exist_ok=True)
    info = {
        "name": safe,
        "ein_last4": (data.get("ein_last4") or "").strip()[:4],
        "contact": (data.get("contact") or "").strip()[:200],
        "notes": (data.get("notes") or "").strip()[:1000],
        "created": datetime.now().isoformat(),
    }
    _save_client_info(safe, info)
    return jsonify({"ok": True, "name": safe})


@app.route("/api/clients/<path:client_name>/info", methods=["GET"])
def get_client_info(client_name):
    """Get client metadata."""
    info = _load_client_info(client_name)
    if not info:
        return jsonify({"name": _safe_client_name(client_name)})
    return jsonify(info)


@app.route("/api/clients/<path:client_name>/info", methods=["PUT"])
def update_client_info(client_name):
    """Update client metadata."""
    data = request.get_json(force=True)
    safe = _safe_client_name(client_name)
    info = _load_client_info(safe) or {"name": safe, "created": datetime.now().isoformat()}
    if "ein_last4" in data:
        info["ein_last4"] = (data["ein_last4"] or "").strip()[:4]
    if "contact" in data:
        info["contact"] = (data["contact"] or "").strip()[:200]
    if "notes" in data:
        info["notes"] = (data["notes"] or "").strip()[:1000]
    _save_client_info(safe, info)
    return jsonify({"ok": True})


@app.route("/api/clients/<path:client_name>", methods=["DELETE"])
def delete_client(client_name):
    """Delete a client folder and all its data. Requires typing 'delete' to confirm."""
    import shutil
    payload = request.get_json(silent=True) or {}
    confirm_text = (payload.get("confirm") or "").strip().lower()
    if confirm_text != "delete":
        return jsonify({"error": "Type 'delete' to confirm"}), 400

    safe = _safe_client_name(client_name)
    client_dir = CLIENTS_DIR / safe
    if not client_dir.exists():
        return jsonify({"error": f"Client folder '{safe}' not found"}), 404

    # Block if any job is running/queued for this client
    for jid, j in jobs.items():
        if _safe_client_name(j.get("client_name", "")) == safe and j.get("status") in ("running", "queued"):
            return jsonify({"error": f"Cannot delete: job '{jid}' is currently running for this client"}), 409

    try:
        shutil.rmtree(str(client_dir))
    except OSError as e:
        return jsonify({"error": f"Failed to delete client folder: {e}"}), 500

    # Clean up client canonical values from SQLite
    conn = _get_db()
    try:
        conn.execute("DELETE FROM client_canonical_values WHERE client_name = ?", (safe,))
        conn.commit()
    except sqlite3.Error:
        pass
    finally:
        conn.close()

    # Clear file paths from jobs (keep records for audit trail)
    updated_count = 0
    for jid, j in jobs.items():
        if _safe_client_name(j.get("client_name", "")) == safe:
            j["client_folder"] = None
            updated_count += 1
    if updated_count:
        save_jobs()

    return jsonify({"ok": True, "name": safe, "jobs_updated": updated_count})


@app.route("/api/clients/merge", methods=["POST"])
def merge_clients():
    """Merge source client into target client. Moves all data."""
    import shutil
    payload = request.get_json(force=True)
    source_name = (payload.get("source") or "").strip()
    target_name = (payload.get("target") or "").strip()

    if not source_name or not target_name:
        return jsonify({"error": "Both source and target are required"}), 400

    safe_source = _safe_client_name(source_name)
    safe_target = _safe_client_name(target_name)
    if safe_source == safe_target:
        return jsonify({"error": "Source and target cannot be the same client"}), 400

    source_dir = CLIENTS_DIR / safe_source
    target_dir = CLIENTS_DIR / safe_target
    if not source_dir.exists():
        return jsonify({"error": f"Source client '{safe_source}' not found"}), 404

    # Block if source has running/queued jobs
    for jid, j in jobs.items():
        if _safe_client_name(j.get("client_name", "")) == safe_source and j.get("status") in ("running", "queued"):
            return jsonify({"error": f"Cannot merge: job '{jid}' is running for source client"}), 409

    target_dir.mkdir(parents=True, exist_ok=True)
    merge_log = []

    # 1. Merge client_info.json
    source_info = _load_client_info(safe_source)
    target_info = _load_client_info(safe_target) or {"name": safe_target, "created": datetime.now().isoformat()}
    if source_info:
        source_notes = source_info.get("notes", "")
        if source_notes:
            existing = target_info.get("notes", "") or ""
            target_info["notes"] = (existing + "\n[Merged from " + safe_source + "] " + source_notes).strip()
        for field in ("ein_last4", "contact"):
            if not target_info.get(field) and source_info.get(field):
                target_info[field] = source_info[field]
        target_info["updated"] = datetime.now().isoformat()
        _save_client_info(safe_target, target_info)
        merge_log.append("Merged client info")

    # 2. Merge instructions
    source_instr = _load_instructions(safe_source)
    if source_instr.get("rules"):
        target_instr = _load_instructions(safe_target)
        existing_texts = {r.get("text", "").lower() for r in target_instr.get("rules", [])}
        added = 0
        for rule in source_instr["rules"]:
            if rule.get("text", "").lower() not in existing_texts:
                rule["id"] = datetime.now().strftime("%Y%m%d%H%M%S") + str(len(target_instr.get("rules", [])) + added)
                target_instr.setdefault("rules", []).append(rule)
                added += 1
        if added:
            _save_instructions(safe_target, target_instr)
            merge_log.append(f"Merged {added} instructions")

    # 3. Merge context documents
    source_ctx_dir = CLIENTS_DIR / safe_source / "context"
    if source_ctx_dir.exists():
        target_ctx_dir = _context_dir(safe_target)
        source_ctx_idx = _load_context_index(safe_source)
        target_ctx_idx = _load_context_index(safe_target)

        moved_docs = 0
        for doc in source_ctx_idx.get("documents", []):
            src_file = source_ctx_dir / doc.get("filename", "")
            if src_file.exists():
                dst_file = target_ctx_dir / src_file.name
                if dst_file.exists():
                    dst_file = target_ctx_dir / f"{dst_file.stem}_from_{safe_source}{dst_file.suffix}"
                    doc["filename"] = dst_file.name
                shutil.move(str(src_file), str(dst_file))
            target_ctx_idx.setdefault("documents", []).append(doc)
            moved_docs += 1

        # Merge prior_year_data
        source_pyd = source_ctx_idx.get("prior_year_data", {})
        target_pyd = target_ctx_idx.get("prior_year_data", {})
        for year, payers in source_pyd.items():
            if year not in target_pyd:
                target_pyd[year] = payers
            else:
                existing_eins = {p.get("ein", "") for p in target_pyd[year] if p.get("ein")}
                for payer in payers:
                    if payer.get("ein") and payer["ein"] in existing_eins:
                        continue
                    target_pyd[year].append(payer)
        target_ctx_idx["prior_year_data"] = target_pyd
        _save_context_index(safe_target, target_ctx_idx)
        if moved_docs:
            merge_log.append(f"Merged {moved_docs} context documents")

    # 4. Move output directories (DocType/Year folders)
    for item in source_dir.iterdir():
        if item.is_dir() and item.name not in ("context",):
            target_doctype_dir = target_dir / item.name
            target_doctype_dir.mkdir(parents=True, exist_ok=True)
            for sub in item.iterdir():
                if sub.is_dir():
                    target_sub = target_doctype_dir / sub.name
                    if target_sub.exists():
                        for f in sub.iterdir():
                            dst = target_sub / f.name
                            if dst.exists():
                                dst = target_sub / f"{f.stem}_from_{safe_source}{f.suffix}"
                            shutil.move(str(f), str(dst))
                    else:
                        shutil.move(str(sub), str(target_sub))
            merge_log.append(f"Moved output folder: {item.name}")

    # 5. Update all jobs: source → target
    jobs_updated = 0
    for jid, j in jobs.items():
        if _safe_client_name(j.get("client_name", "")) == safe_source:
            j["client_name"] = safe_target
            old_cf = j.get("client_folder") or ""
            if old_cf:
                j["client_folder"] = old_cf.replace(str(source_dir), str(target_dir))
            jobs_updated += 1
    if jobs_updated:
        save_jobs()
        merge_log.append(f"Updated {jobs_updated} job records")

    # 6. Delete source folder
    try:
        shutil.rmtree(str(source_dir))
        merge_log.append(f"Deleted source folder: {safe_source}")
    except OSError as e:
        merge_log.append(f"Warning: Could not fully delete source folder: {e}")

    return jsonify({"ok": True, "source": safe_source, "target": safe_target, "log": merge_log, "jobs_updated": jobs_updated})


@app.route("/api/clients/<path:client_name>/documents", methods=["GET"])
def get_client_documents(client_name):
    """List all extraction jobs for a client, grouped by document type."""
    safe = _safe_client_name(client_name)
    docs = []
    for jid, j in jobs.items():
        jclient = _safe_client_name(j.get("client_name", ""))
        if jclient != safe:
            continue
        # Check for output files
        has_xlsx = False
        has_log = False
        if j.get("status") == "complete":
            xlsx_path = OUTPUT_DIR / f"{jid}.xlsx"
            log_path = OUTPUT_DIR / f"{jid}_log.json"
            has_xlsx = xlsx_path.exists()
            has_log = log_path.exists()
        docs.append({
            "job_id": jid,
            "filename": j.get("filename", ""),
            "doc_type": j.get("doc_type", ""),
            "year": j.get("year", ""),
            "status": j.get("status", ""),
            "cost_usd": j.get("cost_usd"),
            "created": j.get("created", ""),
            "has_xlsx": has_xlsx,
            "has_log": has_log,
        })
    # Sort by created descending
    docs.sort(key=lambda d: d.get("created", ""), reverse=True)
    # Group by doc_type
    grouped = {}
    for d in docs:
        dt = d["doc_type"] or "other"
        if dt not in grouped:
            grouped[dt] = []
        grouped[dt].append(d)
    return jsonify({"documents": docs, "grouped": grouped})


@app.route("/api/clients/<path:client_name>/files", methods=["GET"])
def get_client_files(client_name):
    """List actual files on disk for a client — source PDFs, outputs, context docs."""
    safe = _safe_client_name(client_name)
    client_base = CLIENTS_DIR / safe
    files = {"source": [], "outputs": [], "context": []}

    if not client_base.exists():
        return jsonify({"files": files, "client_path": str(client_base)})

    # Walk through all subdirectories
    for root, dirs, filenames in os.walk(str(client_base)):
        rel_root = os.path.relpath(root, str(client_base))
        for fname in sorted(filenames):
            if fname.startswith(".") or fname.endswith(".migrated"):
                continue
            full_path = os.path.join(root, fname)
            finfo = {
                "name": fname,
                "path": full_path,
                "rel_path": os.path.join(rel_root, fname) if rel_root != "." else fname,
                "size_kb": round(os.path.getsize(full_path) / 1024, 1),
                "modified": datetime.fromtimestamp(os.path.getmtime(full_path)).isoformat(),
            }
            lower = fname.lower()
            if lower.endswith(".pdf"):
                files["source"].append(finfo)
            elif lower.endswith((".xlsx", ".json", ".csv")):
                files["outputs"].append(finfo)
            elif rel_root.startswith("context"):
                files["context"].append(finfo)
            else:
                files["outputs"].append(finfo)

    return jsonify({
        "files": files,
        "client_path": str(client_base),
        "total_files": sum(len(v) for v in files.values()),
    })


@app.route("/api/clients/<path:client_name>/export-zip", methods=["GET"])
def export_client_zip(client_name):
    """Export entire client folder as a zip download."""
    import zipfile
    safe = _safe_client_name(client_name)
    client_base = CLIENTS_DIR / safe

    if not client_base.exists():
        return jsonify({"error": "Client folder not found"}), 404

    # Create zip in memory
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, filenames in os.walk(str(client_base)):
            for fname in filenames:
                if fname.startswith(".") or fname.endswith(".migrated"):
                    continue
                full_path = os.path.join(root, fname)
                arcname = os.path.relpath(full_path, str(client_base))
                zf.write(full_path, arcname)

    zip_buffer.seek(0)
    friendly = re.sub(r'[^\w\s\-]', '', safe).strip() or "client"
    return send_file(
        zip_buffer,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"{friendly}_documents.zip",
    )


# ─── Prior-Year Context Routes ────────────────────────────────────────────────

@app.route("/api/context/<path:client_name>", methods=["GET"])
def get_context(client_name):
    """Get the context index for a client."""
    idx = _load_context_index(client_name)
    return jsonify(idx)


@app.route("/api/context/<path:client_name>/upload", methods=["POST"])
def upload_context(client_name):
    """Upload a context document (prior-year return, workbook, notes)."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    fname = f.filename or "document"
    ext = Path(fname).suffix.lower()
    if ext not in (".pdf", ".xlsx", ".xls", ".txt", ".csv"):
        return jsonify({"error": "Supported formats: PDF, XLSX, XLS, TXT, CSV"}), 400

    doc_label = request.form.get("label", "").strip() or fname
    doc_year = request.form.get("year", "").strip()

    ctx_dir = _context_dir(client_name)
    safe_fname = re.sub(r'[^\w\s\-\.,()]', '', fname).strip() or "document" + ext
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    saved_name = f"{ts}_{safe_fname}"
    saved_path = ctx_dir / saved_name
    f.save(str(saved_path))

    # Parse the document
    parsed = _parse_context_document(str(saved_path), doc_label)
    if doc_year:
        parsed["year"] = doc_year

    # Update the index
    idx = _load_context_index(client_name)
    doc_entry = {
        "id": ts,
        "filename": saved_name,
        "original_name": fname,
        "label": doc_label,
        "year": doc_year,
        "uploaded": datetime.now().isoformat(),
        "payer_count": len(parsed.get("payers", [])),
        "has_text": bool(parsed.get("raw_text", "").strip()),
    }
    idx["documents"].append(doc_entry)

    # Merge payers into prior_year_data
    if not idx.get("prior_year_data"):
        idx["prior_year_data"] = {"documents": []}
    idx["prior_year_data"]["documents"].append(parsed)

    _save_context_index(client_name, idx)

    # Save parsed data separately for quick access
    parsed_path = ctx_dir / f"{ts}_parsed.json"
    with open(parsed_path, "w") as pf:
        json.dump(parsed, pf, indent=2, default=str)

    return jsonify({"ok": True, "document": doc_entry, "payers_found": len(parsed.get("payers", []))})


@app.route("/api/context/<path:client_name>/<doc_id>", methods=["DELETE"])
def delete_context(client_name, doc_id):
    """Delete a context document."""
    idx = _load_context_index(client_name)
    ctx_dir = _context_dir(client_name)

    new_docs = []
    removed = False
    for doc in idx.get("documents", []):
        if doc.get("id") == doc_id:
            # Delete file
            fpath = ctx_dir / doc.get("filename", "")
            if fpath.exists():
                try:
                    os.remove(str(fpath))
                except OSError:
                    pass
            # Delete parsed JSON
            parsed = ctx_dir / f"{doc_id}_parsed.json"
            if parsed.exists():
                try:
                    os.remove(str(parsed))
                except OSError:
                    pass
            removed = True
        else:
            new_docs.append(doc)

    idx["documents"] = new_docs

    # Rebuild prior_year_data from remaining parsed files
    idx["prior_year_data"] = {"documents": []}
    for doc in new_docs:
        parsed_path = ctx_dir / f"{doc['id']}_parsed.json"
        if parsed_path.exists():
            try:
                with open(parsed_path) as pf:
                    idx["prior_year_data"]["documents"].append(json.load(pf))
            except (json.JSONDecodeError, IOError):
                pass

    _save_context_index(client_name, idx)
    return jsonify({"ok": removed})


@app.route("/api/context/<path:client_name>/completeness", methods=["GET"])
def completeness_report(client_name):
    """Generate a completeness report comparing current extractions to prior year."""
    # Gather current extractions for this client
    current_exts = []
    for jid, j in jobs.items():
        if j.get("status") != "complete":
            continue
        if _safe_client_name(j.get("client_name", "")) != _safe_client_name(client_name):
            continue
        log_path = j.get("output_log")
        if not log_path or not os.path.exists(log_path):
            continue
        try:
            with open(log_path) as f:
                log_data = json.load(f)
            current_exts.extend(log_data.get("extractions", []))
        except (json.JSONDecodeError, IOError):
            pass

    year = str(datetime.now().year)
    report = _build_completeness_report(client_name, current_exts, year)
    return jsonify(report)


# ─── Client Instructions Routes ──────────────────────────────────────────────

@app.route("/api/instructions/<path:client_name>", methods=["GET"])
@require_login
def get_instructions(client_name):
    """Get client instructions."""
    return jsonify(_load_instructions(client_name))


@app.route("/api/instructions/<path:client_name>", methods=["POST"])
@require_login
def save_instruction(client_name):
    """Add or update a client instruction.

    Body: { "text": "...", "id": "..." (optional, for update) }
    """
    payload = request.get_json(silent=True) or {}
    text = payload.get("text", "").strip()
    if not text:
        return jsonify({"error": "Instruction text is required"}), 400
    if len(text) > 500:
        return jsonify({"error": "Instruction too long (max 500 chars)"}), 400

    data = _load_instructions(client_name)
    rule_id = payload.get("id", "")

    if rule_id:
        # Update existing
        for rule in data["rules"]:
            if rule.get("id") == rule_id:
                rule["text"] = text
                rule["updated"] = datetime.now().isoformat()
                break
    else:
        # Add new
        rule_id = datetime.now().strftime("%Y%m%d%H%M%S") + str(len(data["rules"]))
        data["rules"].append({
            "id": rule_id,
            "text": text,
            "created": datetime.now().isoformat(),
        })

    _save_instructions(client_name, data)
    return jsonify({"ok": True, "id": rule_id, "total": len(data["rules"])})


@app.route("/api/instructions/<path:client_name>/<rule_id>", methods=["DELETE"])
@require_login
def delete_instruction(client_name, rule_id):
    """Delete a client instruction."""
    data = _load_instructions(client_name)
    data["rules"] = [r for r in data["rules"] if r.get("id") != rule_id]
    _save_instructions(client_name, data)
    return jsonify({"ok": True, "total": len(data["rules"])})


# ─── Batch Categorization Routes ─────────────────────────────────────────────

@app.route("/api/batch-categories", methods=["GET"])
def batch_categories():
    """Get all uncategorized transactions for batch categorization."""
    client_name = request.args.get("client", "")
    show_all = request.args.get("all", "false") == "true"
    items = _gather_uncategorized(client_name=client_name if client_name else None)
    if not show_all:
        items = [i for i in items if not i.get("current_category")]
    # Group by normalized vendor
    groups = {}
    for item in items:
        key = item.get("vendor_norm") or item.get("desc", "")[:40]
        if key not in groups:
            groups[key] = {
                "vendor": key,
                "display_name": item.get("desc", key),
                "suggested": item.get("suggested_category", ""),
                "current": item.get("current_category", ""),
                "count": 0,
                "total_amount": 0,
                "items": [],
            }
        groups[key]["count"] += 1
        amt = item.get("amount", 0)
        if isinstance(amt, (int, float)):
            groups[key]["total_amount"] += abs(amt)
        groups[key]["items"].append(item)

    sorted_groups = sorted(groups.values(), key=lambda g: g["count"], reverse=True)
    total = len(items)
    categorized = sum(1 for i in items if i.get("current_category"))
    return jsonify({
        "groups": sorted_groups,
        "total": total,
        "categorized": categorized,
        "uncategorized": total - categorized,
        "chart_of_accounts": CHART_OF_ACCOUNTS,
    })


@app.route("/api/batch-categories/apply", methods=["POST"])
def apply_batch_categories():
    """Apply a category to a batch of transactions.

    Body: { "vendor": "VENDOR_NORM", "category": "Utilities",
            "items": [{job_id, field_key, desc}, ...],
            "learn": true }
    """
    payload = request.get_json(silent=True) or {}
    category = payload.get("category", "").strip()
    items = payload.get("items", [])
    learn = payload.get("learn", True)
    vendor = payload.get("vendor", "")

    if not category or not items:
        return jsonify({"error": "category and items required"}), 400

    applied = 0
    for item in items:
        jid = item.get("job_id")
        field_key = item.get("field_key")
        desc = item.get("desc", "")
        if not jid or not field_key:
            continue

        vdata = _load_verifications(jid)
        decision = vdata["fields"].get(field_key, {})
        decision["category"] = category
        decision["vendor_desc"] = desc
        decision["timestamp"] = datetime.now().isoformat()
        decision["reviewer"] = "BATCH"
        if not decision.get("status"):
            decision["status"] = "confirmed"
        vdata["fields"][field_key] = decision
        _save_verifications(jid, vdata)
        applied += 1

    if learn and vendor:
        _learn_vendor_category(vendor, category)

    # Regen Excel for affected jobs
    affected_jobs = set(item.get("job_id") for item in items if item.get("job_id"))
    for jid in affected_jobs:
        _regen_excel(jid)

    return jsonify({"ok": True, "applied": applied})


# ─── Excel Regeneration ──────────────────────────────────────────────────────

def _regen_excel(job_id, output_format_override=None):
    """Regenerate the Excel file with operator verification corrections applied.

    Uses get_verified_extractions() to overlay all confirmed/corrected values
    onto the raw extraction data, then calls extract.py --regen-excel to rebuild.

    Args:
        output_format_override: If set, pass --output-format to extract.py subprocess.
    """
    job = jobs.get(job_id)
    if not job or job.get("status") != "complete":
        return False

    log_path = job.get("output_log")
    xlsx_path = job.get("output_xlsx")
    if not log_path or not xlsx_path or not os.path.exists(log_path):
        return False

    # Load verifications for audit trail (decision metadata: status, notes, timestamps)
    vdata = _load_verifications(job_id)
    corrections = vdata.get("fields", {})

    # Get canonical data with verified values overlaid
    verified_log = get_verified_extractions(job_id)
    if not verified_log:
        return False

    extractions = verified_log.get("extractions", [])
    if not extractions:
        return False

    try:
        # Re-run populate_template via extract.py as a subprocess
        import subprocess
        year = job.get("year", "2024")
        cmd = [
            sys.executable, str(BASE_DIR / "extract.py"),
            "--regen-excel",
            "--log-input", log_path,
            "--output", xlsx_path,
            "--year", str(year),
        ]

        if output_format_override:
            cmd.extend(["--output-format", output_format_override])

        # Write verified log to a temp file for extract.py to read
        corrected_log_path = log_path.replace("_log.json", "_corrected_log.json")
        with open(corrected_log_path, "w") as f:
            json.dump(verified_log, f, indent=2, default=str)

        cmd[4] = corrected_log_path  # --log-input uses the verified version

        proc = subprocess.run(cmd, capture_output=True, text=True, cwd=str(BASE_DIR), timeout=30)

        # Clean up temp file
        try:
            os.remove(corrected_log_path)
        except OSError:
            pass

        if proc.returncode != 0:
            print(f"  Regen subprocess failed (rc={proc.returncode}): {proc.stderr[:500] if proc.stderr else 'no stderr'}")
            # Fallback: apply corrections directly to existing Excel
            _apply_corrections_to_excel(xlsx_path, corrections, vdata.get("reviewer", ""))
        else:
            corrected_count = sum(1 for d in corrections.values() if d.get("status") == "corrected")
            if corrected_count > 0:
                print(f"  ✓ Regen complete: {corrected_count} corrections applied via extract.py")

        # Always add the audit trail worksheet (primary path doesn't create one)
        _add_audit_trail_worksheet(xlsx_path, corrections, vdata.get("reviewer", ""))

        # Copy updated Excel to client folder
        if job.get("client_folder"):
            import shutil
            client_dir = Path(job["client_folder"])
            if client_dir.exists():
                try:
                    dst = client_dir / Path(xlsx_path).name
                    shutil.copy2(xlsx_path, str(dst))
                except Exception:
                    pass

        return True

    except Exception as e:
        print(f"  Excel regen error: {e}")
        # Fallback: try direct Excel patching
        try:
            _apply_corrections_to_excel(xlsx_path, corrections, vdata.get("reviewer", ""))
        except Exception:
            pass
        return False


def _add_audit_trail_worksheet(xlsx_path, corrections, reviewer=""):
    """Add or replace the Audit Trail worksheet with all verification decisions."""
    if not os.path.exists(xlsx_path) or not corrections:
        return
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.load_workbook(xlsx_path)
        audit_name = "Audit Trail"
        if audit_name in wb.sheetnames:
            del wb[audit_name]
        ws = wb.create_sheet(audit_name)

        # Title
        ws["A1"] = "Operator Verification Audit Trail"
        ws["A1"].font = Font(bold=True, size=14, color="1A252F")
        ws.merge_cells("A1:G1")
        ws["A2"] = f"Generated {datetime.now().strftime('%m/%d/%Y %I:%M %p')}"
        ws["A2"].font = Font(italic=True, color="888888", size=9)
        ws["A3"] = f"Reviewer: {reviewer}" if reviewer else "Reviewer: (not specified)"
        ws["A3"].font = Font(bold=True, size=11)

        # Headers
        row = 5
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        headers = ["Page:Field", "Status", "Original Value", "Corrected Value", "Reviewer", "Timestamp", "Note"]
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=i+1, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center" if i == 1 else "left")
        row += 1

        # Data rows with alternating colors
        alt_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        status_fills = {
            "corrected": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
            "confirmed": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
            "flagged": PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
        }
        thin_border = Border(bottom=Side(style="thin", color="E0E0E0"))

        for idx, key in enumerate(sorted(corrections.keys())):
            decision = corrections[key]
            parts = key.split(":")
            if len(parts) == 3:
                page_label = f"Pg {parts[0]}"
                field_label = parts[2].replace("_", " ").title()
                field_display = f"{page_label}: {field_label}"
            else:
                field_display = key

            status = decision.get("status", "")
            original = decision.get("original_value", "")
            corrected = decision.get("corrected_value", "")
            rev = decision.get("reviewer", reviewer)
            ts = decision.get("timestamp", "")
            note = decision.get("note", "")

            ws.cell(row=row, column=1, value=field_display)
            status_cell = ws.cell(row=row, column=2, value=status.upper())
            status_cell.alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=3, value=str(original) if original else "")
            ws.cell(row=row, column=4, value=str(corrected) if corrected else "")
            ws.cell(row=row, column=5, value=rev)
            ws.cell(row=row, column=6, value=ts)
            ws.cell(row=row, column=7, value=note)

            # Status color
            if status in status_fills:
                status_cell.fill = status_fills[status]
            # Alternating row background
            if idx % 2 == 1:
                for c in range(1, 8):
                    cell = ws.cell(row=row, column=c)
                    if cell.fill == PatternFill():  # only if not already colored
                        cell.fill = alt_fill
            # Border
            for c in range(1, 8):
                ws.cell(row=row, column=c).border = thin_border
            row += 1

        # Summary
        row += 1
        confirmed = sum(1 for d in corrections.values() if d.get("status") == "confirmed")
        corrected = sum(1 for d in corrections.values() if d.get("status") == "corrected")
        flagged = sum(1 for d in corrections.values() if d.get("status") == "flagged")
        ws.cell(row=row, column=1, value="Summary:").font = Font(bold=True, size=11)
        row += 1
        summary_items = [
            ("Confirmed", confirmed, "C8E6C9"),
            ("Corrected", corrected, "FFF9C4"),
            ("Flagged", flagged, "FFE0B2"),
            ("Total Reviewed", len(corrections), "E0E0E0"),
        ]
        for label, count, color in summary_items:
            ws.cell(row=row, column=1, value=label)
            ct_cell = ws.cell(row=row, column=2, value=count)
            ct_cell.font = Font(bold=True, size=11)
            ct_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ct_cell.alignment = Alignment(horizontal="center")
            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 24
        ws.column_dimensions["G"].width = 32

        # Print setup
        ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.orientation = "landscape"

        wb.save(xlsx_path)
    except Exception as e:
        print(f"  Audit trail worksheet error: {e}")


def _apply_corrections_to_excel(xlsx_path, corrections, reviewer=""):
    """Direct fallback: patch existing Excel cells with corrected values + add audit sheet."""
    if not os.path.exists(xlsx_path):
        return

    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.comments import Comment

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Color for operator-corrected cells
    op_fill = PatternFill(start_color="B3E5FC", end_color="B3E5FC", fill_type="solid")

    # Build a list of corrections to apply: (original_value, corrected_value, field_name)
    pending = []
    for key, decision in corrections.items():
        if decision.get("status") != "corrected" or decision.get("corrected_value") is None:
            continue
        parts = key.split(":")
        field_name = parts[2] if len(parts) == 3 else key
        orig = decision.get("original_value")
        corr = decision.get("corrected_value")
        # Normalize to number if possible
        try:
            corr_num = float(str(corr).replace(",", ""))
        except (ValueError, TypeError):
            corr_num = None
        try:
            orig_num = float(str(orig).replace(",", "")) if orig is not None else None
        except (ValueError, TypeError):
            orig_num = None
        pending.append({
            "field": field_name,
            "orig": orig, "orig_num": orig_num,
            "corr": corr, "corr_num": corr_num,
            "reviewer": decision.get("reviewer", reviewer),
            "applied": False,
        })

    # Scan all data cells and match by original value
    # Strategy: for each cell with a value, check if it matches any pending correction's
    # original value. Apply the first match and mark it done. This is best-effort but
    # far better than the previous no-op.
    patched = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is None:
                continue
            for p in pending:
                if p["applied"]:
                    continue
                matched = False
                # Numeric match (within tolerance for floating point)
                if p["orig_num"] is not None and isinstance(cell.value, (int, float)):
                    if abs(cell.value - p["orig_num"]) < 0.005:
                        matched = True
                # String match
                elif p["orig"] is not None and isinstance(cell.value, str):
                    if str(cell.value).strip() == str(p["orig"]).strip():
                        matched = True
                # String-to-number: cell has number but orig was stored as string
                elif p["orig_num"] is not None and isinstance(cell.value, (int, float)):
                    pass  # already handled above
                elif p["orig"] is not None and isinstance(cell.value, (int, float)):
                    try:
                        if abs(cell.value - float(str(p["orig"]).replace(",", ""))) < 0.005:
                            matched = True
                    except (ValueError, TypeError):
                        pass

                if matched:
                    # Apply correction
                    if p["corr_num"] is not None and isinstance(cell.value, (int, float)):
                        cell.value = p["corr_num"]
                    else:
                        cell.value = p["corr"]
                    cell.fill = op_fill
                    old_display = str(p["orig"]) if p["orig"] is not None else "?"
                    cell.comment = Comment(
                        f"Operator corrected (was {old_display}) — {p['reviewer']}",
                        "Operator"
                    )
                    p["applied"] = True
                    patched += 1
                    break  # move to next cell

    wb.save(xlsx_path)

    # Add audit trail worksheet
    _add_audit_trail_worksheet(xlsx_path, corrections, reviewer)


@app.route("/api/regen-excel/<job_id>", methods=["POST"])
@require_login
def regen_excel(job_id):
    """Manually trigger Excel regeneration with verification corrections.
    Accepts optional JSON body: { "output_format": "tax_review" } to override format."""
    job = jobs.get(job_id)
    if not job or job.get("status") != "complete":
        return jsonify({"error": "Job not found or not complete"}), 404
    fmt_override = None
    if request.is_json and request.json:
        fmt_override = request.json.get("output_format")
    ok = _regen_excel(job_id, output_format_override=fmt_override)
    return jsonify({"ok": ok})


@app.route("/api/clients/<path:client_name>/generate-report", methods=["POST"])
def generate_report(client_name):
    """Generate a combined Excel report from multiple extraction jobs."""
    data = request.get_json(force=True)
    job_ids = data.get("job_ids", [])
    output_format = data.get("output_format", "tax_review")
    year = data.get("year", str(datetime.now().year))

    if not job_ids:
        return jsonify({"error": "No jobs selected"}), 400

    # Gather extractions from all selected jobs (using verified values where available)
    combined_extractions = []
    for jid in job_ids:
        job = jobs.get(jid)
        if not job or job.get("status") != "complete":
            continue
        try:
            verified_log = get_verified_extractions(jid)
            if verified_log:
                combined_extractions.extend(verified_log.get("extractions", []))
        except Exception:
            continue

    if not combined_extractions:
        return jsonify({"error": "No extraction data found in selected jobs"}), 400

    # Write combined log
    safe = _safe_client_name(client_name)
    report_id = f"report-{safe}-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    combined_log = {
        "extractions": combined_extractions,
        "output_format": output_format,
        "year": year,
    }
    combined_path = OUTPUT_DIR / f"{report_id}_log.json"
    with open(combined_path, "w") as f:
        json.dump(combined_log, f)

    # Generate Excel via extract.py
    output_path = OUTPUT_DIR / f"{report_id}.xlsx"
    try:
        cmd = [
            sys.executable, str(BASE_DIR / "extract.py"),
            "--regen-excel",
            "--log-input", str(combined_path),
            "--output", str(output_path),
            "--year", str(year),
            "--output-format", output_format,
        ]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        if result.returncode != 0:
            return jsonify({"error": f"Report generation failed: {result.stderr[:500]}"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    if not output_path.exists():
        return jsonify({"error": "Report file was not created"}), 500

    return jsonify({"ok": True, "filename": report_id, "download_url": f"/api/download-report/{report_id}"})


@app.route("/api/download-report/<report_id>")
def download_report(report_id):
    """Download a generated report."""
    safe_id = re.sub(r'[^\w\-]', '', report_id)
    path = OUTPUT_DIR / f"{safe_id}.xlsx"
    if not path.exists():
        return jsonify({"error": "Report not found"}), 404
    return send_file(str(path), as_attachment=True, download_name=f"{safe_id}.xlsx")


# ─── T1.6.2: Facts API (DB-backed) ──────────────────────────────────────────

@app.route("/api/facts/<job_id>")
def get_facts(job_id):
    """Return all facts for a job from the unified facts table.

    This is the DB-backed source of truth. The review UI can use this
    instead of (or in addition to) /api/results/<job_id>.

    Query params:
        tax_year: optional filter by tax year
    """
    from fact_store import FactStore

    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    tax_year = request.args.get("tax_year", type=int)

    try:
        fs = FactStore(str(DB_PATH))
        facts = fs.get_facts_for_job(job_id, tax_year=tax_year)
        counts = fs.count_facts(job_id, tax_year=tax_year)
        review_queue = fs.get_review_queue(job_id)

        return jsonify({
            "ok": True,
            "job_id": job_id,
            "client_id": job.get("client_name", ""),
            "tax_year": tax_year or job.get("year"),
            "facts": facts,
            "counts": counts,
            "total": sum(counts.values()),
            "needs_review": len(review_queue),
            "source": "db",
        })
    except Exception as e:
        return jsonify({"error": f"Could not read facts: {e}"}), 500


# ─── T1.6: Workpaper Generation ──────────────────────────────────────────────

@app.route("/api/workpaper/<path:client_name>", methods=["POST"])
def generate_workpaper(client_name):
    """Generate a professional workpaper Excel file for a client.

    POST body: {"year": "2025", "mode": "assisted"|"safe"}
    Returns:   {"ok": true, "download_url": "/api/download-report/{id}"}
    """
    from fact_store import FactStore
    from workpaper_export import WorkpaperBuilder

    data = request.get_json(silent=True) or {}
    year = data.get("year") or ""
    mode = data.get("mode") or "assisted"
    job_id = data.get("job_id")  # WORKPAPER-001: optional canonical facts path

    if not year:
        return jsonify({"error": "year is required"}), 400
    if mode not in ("assisted", "safe"):
        return jsonify({"error": "mode must be 'assisted' or 'safe'"}), 400

    try:
        fs = FactStore(str(DB_PATH))

        # WORKPAPER-001: If job_id provided, check canonical facts first
        fact_count = 0
        fact_source = "legacy"
        if job_id:
            try:
                canonical = fs.get_workpaper_facts(job_id, year)
                if canonical:
                    fact_count = len(canonical)
                    fact_source = "canonical"
            except Exception:
                pass  # Fall through to legacy

        if fact_count == 0:
            fact_list = fs.list_legacy_facts(client_name, year)
            fact_count = len(fact_list)
            fact_source = "legacy"
            if not fact_list:
                return jsonify({
                    "error": f"No facts found for {client_name} / {year}. "
                             "Run extraction and review first."
                }), 404

        # Generate filename
        safe_name = re.sub(r'[^\w\-]', '_', client_name)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_id = f"{safe_name}-workpaper-{year}-{timestamp}"
        output_path = OUTPUT_DIR / f"{report_id}.xlsx"

        builder = WorkpaperBuilder(fs, client_name, year, mode=mode,
                                   job_id=job_id)
        builder.build(str(output_path))

        if not output_path.exists():
            return jsonify({"error": "Workpaper file was not created"}), 500

        # Sprint 2: Log workpaper generation
        wp_user = current_user()
        log_event("info", "workpaper_generated",
                  f"Workpaper generated: {client_name} / {year} ({mode} mode, {fact_source} facts)",
                  user_id=wp_user["id"] if wp_user else None,
                  details={"client": client_name, "year": year, "mode": mode,
                           "facts_count": fact_count,
                           "fact_source": fact_source,
                           "job_id": job_id,
                           "filename": f"{report_id}.xlsx"})

        return jsonify({
            "ok": True,
            "filename": f"{report_id}.xlsx",
            "download_url": f"/api/download-report/{report_id}",
            "facts_count": fact_count,
            "mode": mode,
            "fact_source": fact_source,  # WORKPAPER-001: "canonical" or "legacy"
        })

    except (ValueError, TypeError) as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"Workpaper generation failed: {e}"}), 500


# ─── Retry Failed/Interrupted Jobs ──────────────────────────────────────────

@app.route("/api/retry/<job_id>", methods=["POST"])
def retry_job(job_id):
    """Re-run extraction for a failed or interrupted job using the original PDF."""
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    if job.get("status") not in ("error", "interrupted", "failed"):
        return jsonify({"error": f"Cannot retry a job with status '{job.get('status')}'. Only failed, error, or interrupted jobs can be retried."}), 400

    pdf_path = job.get("pdf_path", "")
    if not pdf_path or not os.path.exists(pdf_path):
        return jsonify({"error": "Original PDF no longer exists. Please re-upload."}), 410

    # Reset job state for re-run
    job["status"] = "queued"
    job["stage"] = "queued"
    job["progress"] = 0
    job["log"] = []
    job["error"] = ""
    job.pop("end_time", None)
    job["retry_count"] = job.get("retry_count", 0) + 1
    job["last_retry"] = datetime.now().isoformat()
    save_jobs()

    # Rebuild client folder if needed
    client_dir = job.get("client_folder")
    if client_dir:
        Path(client_dir).mkdir(parents=True, exist_ok=True)

    year = job.get("year", "2024")
    skip_verify = False  # Always verify on retry
    doc_type = job.get("doc_type", "tax_returns")
    output_format = job.get("output_format", "tax_review")
    user_notes = job.get("user_notes", "")
    ai_instructions = job.get("ai_instructions", "")
    disable_pii = job.get("disable_pii", False)
    use_ocr_first = job.get("use_ocr_first", False)

    t = threading.Thread(target=run_extraction, kwargs=dict(
        job_id=job_id, pdf_path=pdf_path, year=year, skip_verify=skip_verify,
        doc_type=doc_type, output_format=output_format, user_notes=user_notes,
        ai_instructions=ai_instructions, disable_pii=disable_pii, resume=True,
        use_ocr_first=use_ocr_first,
    ))
    t.daemon = True
    t.start()

    return jsonify({"job_id": job_id, "retry_count": job["retry_count"]})


@app.route("/api/cancel/<job_id>", methods=["POST"])
def cancel_job(job_id):
    """Cancel a running extraction job."""
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    if job.get("status") not in ("queued", "running"):
        return jsonify({"error": "Job is not running"}), 400

    proc = _active_procs.get(job_id)
    if proc:
        try:
            proc.terminate()
        except Exception:
            pass
        _active_procs.pop(job_id, None)

    job["status"] = "interrupted"
    job["end_time"] = datetime.now().isoformat()
    job["error"] = "Cancelled by user"
    job["log"].append("── Cancelled by user ──")
    save_jobs()

    return jsonify({"success": True})


# ─── Health Check ─────────────────────────────────────────────────────────────

@app.route("/api/health")
def health_check():
    """System health check: version, uptime, job counts, dependency status, disk usage."""
    import shutil as _shutil

    now = datetime.now()
    uptime_seconds = (now - _start_time).total_seconds()

    # Job counts by status
    status_counts = {}
    for j in jobs.values():
        s = j.get("status", "unknown")
        status_counts[s] = status_counts.get(s, 0) + 1

    # Dependency checks
    tesseract_ok = _shutil.which("tesseract") is not None
    extract_ok = (BASE_DIR / "extract.py").exists()
    api_key_set = bool(os.environ.get("ANTHROPIC_API_KEY"))

    # Directory writability
    dirs_ok = {}
    for name, d in [("uploads", UPLOAD_DIR), ("outputs", OUTPUT_DIR), ("clients", CLIENTS_DIR), ("verifications", VERIFY_DIR)]:
        dirs_ok[name] = os.access(str(d), os.W_OK)

    # Disk usage
    try:
        usage = _shutil.disk_usage(str(DATA_DIR))
        disk = {
            "total_gb": round(usage.total / (1024**3), 2),
            "free_gb": round(usage.free / (1024**3), 2),
            "percent_used": round(usage.used / usage.total * 100, 1),
        }
    except Exception:
        disk = None

    # Data directory size
    data_size_mb = 0
    try:
        for dirpath, dirnames, filenames in os.walk(str(DATA_DIR)):
            for fname in filenames:
                try:
                    data_size_mb += os.path.getsize(os.path.join(dirpath, fname))
                except OSError:
                    pass
        data_size_mb = round(data_size_mb / (1024 * 1024), 2)
    except Exception:
        pass

    return jsonify({
        "status": "ok",
        "version": _app_version,
        "uptime_hours": round(uptime_seconds / 3600, 2),
        "started": _start_time.isoformat(),
        "jobs": {"total": len(jobs), "by_status": status_counts},
        "dependencies": {"extract_py": extract_ok, "tesseract": tesseract_ok, "api_key_set": api_key_set},
        "directories": dirs_ok,
        "disk": disk,
        "data_size_mb": data_size_mb,
    })


# ─── API Key Management ──────────────────────────────────────────────────────

@app.route("/api/config/api-key", methods=["GET"])
def get_api_key_status():
    """Check whether the Anthropic API key is configured."""
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    return jsonify({
        "configured": bool(key),
        "hint": (key[:7] + "..." + key[-4:]) if len(key) > 15 else "",
        "source": "env" if key else "none",
    })


@app.route("/api/config/api-key", methods=["POST"])
def set_api_key():
    """Persist the Anthropic API key to .env and set in current process."""
    data = request.get_json(force=True) or {}
    key = (data.get("key") or "").strip()
    if not key:
        return jsonify({"error": "key is required"}), 400
    if not key.startswith("sk-ant-"):
        return jsonify({"error": "Key must start with sk-ant-"}), 400

    env_path = BASE_DIR / ".env"

    # Read existing .env lines (preserve other vars)
    existing_lines = []
    replaced = False
    if env_path.exists():
        with open(env_path) as f:
            for line in f:
                if line.strip().startswith("ANTHROPIC_API_KEY"):
                    existing_lines.append(f"ANTHROPIC_API_KEY={key}\n")
                    replaced = True
                else:
                    existing_lines.append(line)
    if not replaced:
        existing_lines.append(f"ANTHROPIC_API_KEY={key}\n")

    with open(env_path, "w") as f:
        f.writelines(existing_lines)

    # Set in current process so subprocesses inherit it immediately
    os.environ["ANTHROPIC_API_KEY"] = key

    hint = key[:7] + "..." + key[-4:]
    print(f"  ✓ API key saved to .env ({hint})")
    return jsonify({"ok": True, "hint": hint})


# ─── HTML ─────────────────────────────────────────────────────────────────────


MAIN_HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>OathLedger — Document Intake</title>
<script>
// Apply saved theme before paint to prevent flash
(function(){var t=localStorage.getItem('oathledger-theme');if(t)document.documentElement.setAttribute('data-theme',t);})();
</script>
<style>
/* ═══ DESIGN SYSTEM ═══ */
:root {
  --bg: #F7F6F3;
  --bg-card: #FFFFFF;
  --bg-sidebar: #1E2A38;
  --bg-sidebar-hover: #2A3A4C;
  --bg-sidebar-active: #344C64;
  --navy: #2C3E50;
  --navy-light: #3D566E;
  --accent: #3498DB;
  --accent-hover: #2980B9;
  --green: #27AE60;
  --green-bg: #E8F8F0;
  --yellow: #F39C12;
  --yellow-bg: #FFF8E8;
  --red: #E74C3C;
  --red-bg: #FDECEC;
  --purple: #8E44AD;
  --purple-bg: #F5EEFA;
  --text: #2C3E50;
  --text-secondary: #7F8C8D;
  --text-light: #95A5A6;
  --border: #E5E5E0;
  --border-light: #F0EFEC;
  --shadow-sm: 0 1px 3px rgba(0,0,0,0.06);
  --shadow-md: 0 4px 12px rgba(0,0,0,0.08);
  --shadow-lg: 0 8px 24px rgba(0,0,0,0.1);
  --radius: 8px;
  --radius-lg: 12px;
  --mono: 'SF Mono', 'Menlo', 'Consolas', monospace;
  --sans: -apple-system, BlinkMacSystemFont, 'Segoe UI', 'Inter', sans-serif;
  --transition: 0.2s ease;
  --hover-bg: #FAFAF8;
  --focus-bg: #EBF5FB;
  --confirmed-bg: #F0FBF4;
  --corrected-bg: #FFF8E8;
  --flagged-bg: #FFF0E0;
  --entity-bg: #F8F8F6;
  --pdf-bg: #3D3D3D;
  --input-bg: #FFFFFF;
}

* { margin: 0; padding: 0; box-sizing: border-box; }
body { font-family: var(--sans); background: var(--bg); color: var(--text); font-size: 14px; line-height: 1.5; }
::selection { background: var(--accent); color: white; }

/* ═══ LAYOUT ═══ */
.app { display: flex; min-height: 100vh; }
.sidebar { width: 220px; background: var(--bg-sidebar); color: white; display: flex; flex-direction: column; position: fixed; top: 0; left: 0; bottom: 0; z-index: 100; transition: var(--transition); }
.main { margin-left: 220px; flex: 1; min-height: 100vh; padding: 0; }

/* ═══ SIDEBAR ═══ */
.sidebar-brand { padding: 20px 16px 12px; border-bottom: 1px solid rgba(255,255,255,0.08); }
.sidebar-brand h1 { font-size: 16px; font-weight: 700; letter-spacing: 0.02em; }
.sidebar-brand p { font-size: 11px; color: rgba(255,255,255,0.5); margin-top: 2px; }
.sidebar-nav { flex: 1; padding: 8px 0; }
.nav-item { display: flex; align-items: center; gap: 10px; padding: 10px 16px; color: rgba(255,255,255,0.65); cursor: pointer; transition: var(--transition); font-size: 13px; font-weight: 500; border-left: 3px solid transparent; text-decoration: none; }
.nav-item:hover { background: var(--bg-sidebar-hover); color: rgba(255,255,255,0.9); }
.nav-item.active { background: var(--bg-sidebar-active); color: white; border-left-color: var(--accent); }
.nav-item svg { width: 18px; height: 18px; flex-shrink: 0; opacity: 0.7; }
.nav-item.active svg { opacity: 1; }
.nav-badge { background: var(--accent); color: white; font-size: 10px; font-weight: 700; padding: 1px 6px; border-radius: 10px; margin-left: auto; }
.sidebar-footer { padding: 12px 16px; border-top: 1px solid rgba(255,255,255,0.08); font-size: 11px; color: rgba(255,255,255,0.35); }
.sidebar-footer label { display: flex; align-items: center; gap: 6px; font-weight: 600; color: rgba(255,255,255,0.6); }
.sidebar-footer input { background: rgba(255,255,255,0.1); border: 1px solid rgba(255,255,255,0.15); color: white; font-size: 12px; padding: 4px 8px; border-radius: 4px; width: 50px; font-weight: 600; }

/* ═══ PAGE HEADER ═══ */
.page-header { padding: 24px 32px 16px; border-bottom: 1px solid var(--border); background: var(--bg-card); }
.page-header h2 { font-size: 20px; font-weight: 700; color: var(--navy); }
.page-header p { font-size: 13px; color: var(--text-secondary); margin-top: 2px; }
.page-content { padding: 24px 32px; }

/* ═══ SECTIONS (show/hide) ═══ */
.section { display: none; }
.section.active { display: block; }

/* ═══ CARDS ═══ */
.card { background: var(--bg-card); border-radius: var(--radius-lg); box-shadow: var(--shadow-sm); border: 1px solid var(--border-light); }
.card-header { padding: 16px 20px; border-bottom: 1px solid var(--border-light); display: flex; align-items: center; justify-content: space-between; }
.card-header h3 { font-size: 14px; font-weight: 700; color: var(--navy); }
.card-body { padding: 20px; }
.card + .card { margin-top: 16px; }

/* ═══ BUTTONS ═══ */
.btn { display: inline-flex; align-items: center; gap: 6px; padding: 8px 16px; border-radius: 6px; font-size: 13px; font-weight: 600; cursor: pointer; border: none; transition: var(--transition); font-family: var(--sans); }
.btn-primary { background: var(--accent); color: white; }
.btn-primary:hover { background: var(--accent-hover); box-shadow: var(--shadow-sm); }
.btn-secondary { background: var(--bg); color: var(--text); border: 1px solid var(--border); }
.btn-secondary:hover { background: white; border-color: var(--navy-light); }
.btn-success { background: var(--green); color: white; }
.btn-success:hover { opacity: 0.9; }
.btn-danger { background: var(--red); color: white; }
.btn-danger:hover { opacity: 0.9; }
.btn-sm { padding: 5px 10px; font-size: 12px; }
.btn-ghost { background: none; color: var(--text-secondary); padding: 4px 8px; }
.btn-ghost:hover { color: var(--text); background: var(--bg); }
.btn:disabled { opacity: 0.5; cursor: not-allowed; }

/* ═══ FORMS ═══ */
.form-group { margin-bottom: 16px; }
.form-label { display: block; font-size: 12px; font-weight: 600; color: var(--text-secondary); text-transform: uppercase; letter-spacing: 0.04em; margin-bottom: 4px; }
.form-input { width: 100%; padding: 8px 12px; border: 1px solid var(--border); border-radius: 6px; font-size: 13px; font-family: var(--sans); transition: var(--transition); background: white; }
.form-input:focus { outline: none; border-color: var(--accent); box-shadow: 0 0 0 3px rgba(52,152,219,0.12); }
.form-select { appearance: none; background: white url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 12 12'%3E%3Cpath d='M6 8L1 3h10z' fill='%237F8C8D'/%3E%3C/svg%3E") right 10px center no-repeat; padding-right: 28px; }
.form-row { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }
.form-row-3 { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 16px; }
textarea.form-input { resize: vertical; min-height: 60px; }
.form-hint { font-size: 11px; color: var(--text-light); margin-top: 2px; }

/* ═══ TABLES ═══ */
.table-wrap { overflow-x: auto; }
table.data-table { width: 100%; border-collapse: collapse; font-size: 13px; }
table.data-table thead th { font-size: 11px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.04em; color: var(--text-light); padding: 8px 12px; text-align: left; border-bottom: 2px solid var(--border); white-space: nowrap; }
table.data-table tbody td { padding: 10px 12px; border-bottom: 1px solid var(--border-light); vertical-align: middle; }
table.data-table tbody tr:hover { background: #FAFAF8; }
table.data-table tbody tr.row-success { background: var(--green-bg); }
table.data-table tbody tr.row-warning { background: var(--yellow-bg); }
table.data-table tbody tr.row-danger { background: var(--red-bg); }
td.mono { font-family: var(--mono); font-size: 12px; }
td.amount { text-align: right; font-family: var(--mono); font-weight: 600; }
td.actions { white-space: nowrap; text-align: right; }

/* ═══ BADGES / PILLS ═══ */
.badge { display: inline-flex; align-items: center; padding: 2px 8px; border-radius: 10px; font-size: 11px; font-weight: 600; letter-spacing: 0.02em; }
.badge-green { background: var(--green-bg); color: var(--green); }
.badge-yellow { background: var(--yellow-bg); color: #B7791F; }
.badge-red { background: var(--red-bg); color: var(--red); }
.badge-blue { background: #EBF5FB; color: var(--accent); }
.badge-purple { background: var(--purple-bg); color: var(--purple); }
.badge-gray { background: #ECF0F1; color: var(--text-secondary); }
.pill { display: inline-flex; align-items: center; gap: 4px; padding: 4px 12px; border-radius: 20px; font-size: 12px; font-weight: 600; cursor: pointer; border: 1px solid var(--border); background: white; transition: var(--transition); }
.pill:hover { border-color: var(--accent); }
.pill.active { background: var(--accent); color: white; border-color: var(--accent); }

/* ═══ TOAST ═══ */
#toast-container { position: fixed; top: 20px; right: 20px; z-index: 9999; display: flex; flex-direction: column; gap: 8px; }
.toast { padding: 10px 16px; border-radius: 8px; font-size: 13px; font-weight: 500; box-shadow: var(--shadow-md); animation: toastIn 0.3s ease; max-width: 360px; display: flex; align-items: center; gap: 8px; }
.toast-success { background: var(--green); color: white; }
.toast-error { background: var(--red); color: white; }
.toast-info { background: var(--navy); color: white; }
@keyframes toastIn { from { opacity: 0; transform: translateY(-10px); } to { opacity: 1; transform: translateY(0); } }

/* ═══ UPLOAD SECTION ═══ */
.upload-area { border: 2px dashed var(--border); border-radius: var(--radius-lg); padding: 48px 24px; text-align: center; cursor: pointer; transition: var(--transition); background: var(--upload-bg, #FAFAF8); }
.upload-area:hover, .upload-area.dragover { border-color: var(--accent); background: var(--upload-hover-bg, #F0F8FF); }
.upload-area svg { width: 48px; height: 48px; color: var(--text-light); margin-bottom: 12px; }
.upload-area h3 { font-size: 16px; color: var(--text); margin-bottom: 4px; }
.upload-area p { font-size: 13px; color: var(--text-secondary); }
.upload-form { display: none; margin-top: 20px; }
.upload-form.visible { display: block; }
.upload-file-name { font-size: 14px; font-weight: 600; color: var(--accent); margin-bottom: 16px; display: flex; align-items: center; gap: 8px; }

/* ═══ DOC TYPE + OUTPUT FORMAT PILLS ═══ */
.pill-group { display: flex; flex-wrap: wrap; gap: 6px; }

/* ═══ FILTER BAR (cross-document category filter) ═══ */
.filter-bar { position: sticky; top: 0; z-index: 10; background: var(--bg); border-bottom: 1px solid var(--border-light); padding: 10px 16px 8px; margin: -20px -20px 12px; }
.filter-bar-label { font-size: 10px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.06em; color: var(--text-light); margin-bottom: 6px; }
.filter-pills { display: flex; flex-wrap: wrap; gap: 4px; margin-bottom: 6px; }
.filter-pill { display: inline-flex; align-items: center; gap: 3px; padding: 3px 10px; border-radius: 16px; font-size: 11px; font-weight: 600; cursor: pointer; border: 1px solid var(--border); background: var(--bg-card); color: var(--text-secondary); transition: all 0.15s ease; user-select: none; white-space: nowrap; }
.filter-pill:hover { border-color: var(--accent); color: var(--accent); background: rgba(52,152,219,0.04); }
.filter-pill.active { background: var(--accent); color: white; border-color: var(--accent); box-shadow: 0 1px 4px rgba(52,152,219,0.25); }
.filter-pill .pill-count { font-size: 10px; font-weight: 700; background: rgba(0,0,0,0.08); padding: 1px 5px; border-radius: 8px; min-width: 16px; text-align: center; line-height: 1.4; }
.filter-pill.active .pill-count { background: rgba(255,255,255,0.25); }
.filter-search { width: 100%; padding: 6px 10px 6px 28px; border: 1px solid var(--border); border-radius: 8px; font-size: 12px; font-family: var(--sans); background: var(--input-bg); color: var(--text); outline: none; transition: border-color 0.15s ease, box-shadow 0.15s ease; }
.filter-search:focus { border-color: var(--accent); box-shadow: 0 0 0 2px rgba(52,152,219,0.1); }
.filter-search-wrap { position: relative; }
.filter-search-wrap::before { content: '\1F50D'; position: absolute; left: 8px; top: 50%; transform: translateY(-50%); font-size: 12px; opacity: 0.4; pointer-events: none; }
.filter-active-hint { font-size: 11px; color: var(--accent); font-weight: 600; margin-top: 4px; display: flex; align-items: center; gap: 4px; }
.filter-active-hint .filter-clear { cursor: pointer; font-size: 10px; background: var(--accent); color: white; border: none; border-radius: 10px; padding: 1px 6px; font-weight: 700; }
.filter-active-hint .filter-clear:hover { background: var(--accent-hover); }

/* ═══ CROSS-DOCUMENT VIEW ═══ */
.xdoc-container { padding: 0; }
.xdoc-group { background: var(--bg-card); border-radius: var(--radius-lg); margin-bottom: 12px; box-shadow: var(--shadow-sm); border: 1px solid var(--border-light); overflow: hidden; }
.xdoc-group-header { font-size: 12px; font-weight: 700; padding: 10px 16px; background: var(--navy); color: white; display: flex; align-items: center; justify-content: space-between; letter-spacing: 0.01em; }
.xdoc-group-header .xdoc-count { font-size: 10px; font-weight: 600; opacity: 0.7; background: rgba(255,255,255,0.15); padding: 2px 8px; border-radius: 10px; }
.xdoc-field-row { display: flex; align-items: center; padding: 8px 16px; border-bottom: 1px solid var(--border-light); cursor: pointer; transition: background 0.15s ease; min-height: 38px; }
.xdoc-field-row:last-child { border-bottom: none; }
.xdoc-field-row:hover { background: var(--hover-bg); }
.xdoc-field-left { flex: 1; min-width: 0; }
.xdoc-field-name { font-size: 13px; font-weight: 500; color: var(--text-secondary); line-height: 1.3; }
.xdoc-field-source { font-size: 10px; color: var(--text-light); margin-top: 1px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
.xdoc-field-right { display: flex; align-items: center; gap: 8px; flex-shrink: 0; margin-left: 12px; }
.xdoc-field-value { font-size: 14px; font-weight: 600; font-family: var(--mono); color: var(--text); font-variant-numeric: tabular-nums; }
.xdoc-field-icon { font-size: 12px; width: 20px; text-align: center; }
.xdoc-field-icon.vf-ok { color: var(--green); }
.xdoc-field-icon.vf-flag { color: var(--yellow); }
.xdoc-field-icon.vf-none { color: var(--text-light); opacity: 0.3; }
.xdoc-nav-hint { font-size: 10px; color: var(--accent); opacity: 0; transition: opacity 0.15s; margin-left: 4px; white-space: nowrap; }
.xdoc-field-row:hover .xdoc-nav-hint { opacity: 1; }

/* ═══ ROLLUP SUMMARY PANEL ═══ */
.rollup-panel { background: var(--bg-card); border-radius: var(--radius-lg); margin-bottom: 12px; box-shadow: var(--shadow-sm); border: 1px solid var(--border-light); overflow: hidden; }
.rollup-header { display: flex; align-items: center; justify-content: space-between; padding: 10px 16px; cursor: pointer; user-select: none; background: linear-gradient(135deg, var(--accent), #2980b9); color: white; font-size: 13px; font-weight: 700; transition: opacity 0.15s; }
.rollup-header:hover { opacity: 0.92; }
.rollup-header .rollup-chevron { font-size: 11px; transition: transform 0.2s ease; }
.rollup-header .rollup-chevron.collapsed { transform: rotate(-90deg); }
.rollup-body { padding: 0; max-height: 300px; overflow: hidden; transition: max-height 0.3s ease, padding 0.3s ease; }
.rollup-body.collapsed { max-height: 0; }
.rollup-body.expanded { padding: 8px 0; }
.rollup-row { display: flex; align-items: center; justify-content: space-between; padding: 5px 16px; font-size: 13px; }
.rollup-row:hover { background: var(--hover-bg); }
.rollup-label { color: var(--text-secondary); font-weight: 500; }
.rollup-value { font-family: var(--mono); font-weight: 700; color: var(--text); font-variant-numeric: tabular-nums; }
.rollup-total { border-top: 2px solid var(--border); margin-top: 4px; padding-top: 6px; }
.rollup-total .rollup-label { color: var(--text); font-weight: 700; }
.rollup-total .rollup-value { color: var(--accent); font-size: 15px; }
.rollup-empty { font-size: 12px; color: var(--text-light); padding: 8px 16px; font-style: italic; }

/* ═══ PROCESSING ═══ */
.processing-card { max-width: 640px; margin: 0 auto; }
.progress-bar { width: 100%; height: 6px; background: var(--border); border-radius: 3px; overflow: hidden; margin: 12px 0; }
.progress-fill { height: 100%; background: linear-gradient(90deg, var(--accent), #5DADE2); border-radius: 3px; transition: width 0.4s ease; }
.progress-label { display: flex; justify-content: space-between; font-size: 12px; color: var(--text-secondary); }
.console-output { background: #1E2A38; color: #BDC3C7; font-family: var(--mono); font-size: 11px; padding: 12px; border-radius: 6px; max-height: 320px; overflow-y: auto; margin-top: 12px; line-height: 1.6; }
.console-output .line-highlight { color: #5DADE2; }

/* ═══ REVIEW (modernized) ═══ */
.review-header { display: flex; align-items: center; justify-content: space-between; padding: 10px 24px; background: var(--bg-card); border-bottom: 1px solid var(--border); gap: 16px; min-height: 56px; }
.review-nav { display: flex; align-items: center; gap: 6px; flex-shrink: 0; }
.review-nav-btn { width: 34px; height: 34px; padding: 0; display: flex; align-items: center; justify-content: center; border-radius: 8px; border: 1px solid var(--border); background: var(--bg-card); cursor: pointer; transition: all 0.15s ease; color: var(--text-secondary); }
.review-nav-btn:hover { background: var(--hover-bg); border-color: var(--accent); color: var(--accent); box-shadow: var(--shadow-sm); }
.review-pager { font-size: 13px; font-weight: 700; color: var(--text); min-width: 70px; text-align: center; font-variant-numeric: tabular-nums; }
.review-center { flex: 1; display: flex; flex-direction: column; align-items: center; gap: 6px; max-width: 400px; }
.review-actions { display: flex; gap: 8px; align-items: center; flex-shrink: 0; }
.btn-accent { background: var(--accent); color: white; border: none; }
.btn-accent:hover { background: var(--accent-hover); box-shadow: var(--shadow-sm); }
.review-split { display: grid; grid-template-columns: 1fr 1fr; height: calc(100vh - 112px); }
.review-pdf { background: var(--pdf-bg); overflow: auto; display: flex; align-items: center; justify-content: center; padding: 16px; }
.review-pdf-wrap { position: relative; display: inline-block; max-width: 95%; max-height: calc(100vh - 160px); }
.review-pdf img { max-width: 100%; max-height: calc(100vh - 170px); width: auto; height: auto; object-fit: contain; box-shadow: 0 4px 20px rgba(0,0,0,0.35), 0 0 1px rgba(0,0,0,0.2); border-radius: 6px; background: white; transition: box-shadow 0.3s ease; display: block; }
.review-pdf img:hover { box-shadow: 0 8px 32px rgba(0,0,0,0.45), 0 0 2px rgba(0,0,0,0.25); }
.pdf-highlight { position: absolute; background: rgba(255, 230, 0, 0.30); border: 2px solid rgba(255, 60, 0, 0.8); border-radius: 3px; pointer-events: none; transition: opacity 0.25s ease; z-index: 2; }
.pdf-highlight-pulse { animation: highlightPulse 1.5s ease-in-out infinite; }
@keyframes highlightPulse { 0%,100% { box-shadow: 0 0 6px rgba(255,60,0,0.3); } 50% { box-shadow: 0 0 16px rgba(255,60,0,0.6); } }
.review-fields { overflow-y: auto; padding: 20px; background: var(--bg); scroll-behavior: smooth; }
.verify-progress { height: 6px; background: var(--border); border-radius: 3px; overflow: hidden; width: 100%; max-width: 300px; }
.verify-progress-fill { height: 100%; background: linear-gradient(90deg, var(--green), var(--accent)); border-radius: 3px; transition: width 0.4s cubic-bezier(0.4, 0, 0.2, 1); box-shadow: 0 0 8px rgba(39, 174, 96, 0.3); }
.verify-stats { display: flex; gap: 12px; font-size: 12px; color: var(--text-secondary); font-variant-numeric: tabular-nums; }
.verify-stats span { font-weight: 600; }

/* ─── Field rendering (modernized) ─── */
.field-group { background: var(--bg-card); border-radius: var(--radius-lg); margin-bottom: 16px; box-shadow: var(--shadow-sm); border: 1px solid var(--border-light); overflow: hidden; transition: box-shadow 0.2s ease; }
.field-group:hover { box-shadow: var(--shadow-md); }
.field-group-title { font-size: 13px; font-weight: 700; padding: 12px 16px; background: var(--navy); color: white; display: flex; align-items: center; justify-content: space-between; letter-spacing: 0.01em; }
.field-entity { font-size: 12px; color: var(--text-secondary); padding: 8px 16px; background: var(--entity-bg); border-bottom: 1px solid var(--border-light); display: flex; align-items: center; justify-content: space-between; }
.field-entity .all-done { color: var(--green); font-size: 11px; font-weight: 600; }
.field-row { display: flex; align-items: center; padding: 10px 16px; border-bottom: 1px solid var(--border-light); transition: background 0.2s ease, box-shadow 0.2s ease; min-height: 44px; cursor: pointer; position: relative; }
.field-row:last-child { border-bottom: none; }
.field-row:hover { background: var(--hover-bg); }
.field-row.focused { background: var(--focus-bg); box-shadow: inset 3px 0 0 var(--accent); }
.field-row.vf-confirmed { background: var(--confirmed-bg); box-shadow: inset 3px 0 0 var(--green); }
.field-row.vf-corrected { background: var(--corrected-bg); box-shadow: inset 3px 0 0 var(--yellow); }
.field-row.vf-flagged { background: var(--flagged-bg); box-shadow: inset 3px 0 0 var(--red); }
.field-name { flex: 0 0 42%; font-size: 13px; color: var(--text-secondary); font-weight: 500; padding-right: 12px; line-height: 1.4; }
.field-val-wrap { flex: 1; display: flex; align-items: center; gap: 8px; }
.field-val { font-size: 14px; font-weight: 600; font-family: var(--mono); color: var(--text); cursor: pointer; padding: 2px 6px; border-radius: 4px; transition: color 0.15s ease, background 0.15s ease; }
.field-val:hover { color: var(--accent); background: rgba(52, 152, 219, 0.06); }
.field-actions { display: flex; gap: 4px; margin-left: auto; opacity: 0.7; transition: opacity 0.2s ease; }
.field-row:hover .field-actions, .field-row.focused .field-actions { opacity: 1; }
.field-entity-hint { display: block; font-size: 10px; color: var(--text-light); font-weight: 400; margin-top: 1px; opacity: 0.7; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; max-width: 200px; }
.field-name { cursor: default; }
.field-name .relabel-icon { display: none; font-size: 10px; margin-left: 4px; cursor: pointer; color: var(--accent); opacity: 0.5; }
.field-row:hover .field-name .relabel-icon { display: inline; }
.field-name .relabel-icon:hover { opacity: 1; }
.field-relabel-input { font-size: 12px; padding: 2px 6px; border: 1px solid var(--accent); border-radius: 4px; background: var(--bg-card); color: var(--text); width: 90%; font-weight: 500; }
.field-edit-input { font-size: 14px; font-family: var(--mono); padding: 4px 10px; border: 2px solid var(--accent); border-radius: 6px; width: 140px; background: var(--input-bg); color: var(--text); outline: none; box-shadow: 0 0 0 3px rgba(52, 152, 219, 0.15); }
.field-edit-input:focus { box-shadow: 0 0 0 4px rgba(52, 152, 219, 0.25); }

/* Confidence dots */
.conf-dot { width: 8px; height: 8px; border-radius: 50%; display: inline-block; flex-shrink: 0; transition: transform 0.2s ease; }
.conf-dual { background: #1A8C42; box-shadow: 0 0 4px rgba(26, 140, 66, 0.4); }
.conf-confirmed { background: #5CB85C; box-shadow: 0 0 4px rgba(92, 184, 92, 0.3); }
.conf-corrected { background: #FFCC00; box-shadow: 0 0 4px rgba(255, 204, 0, 0.3); }
.conf-low { background: #FF9800; box-shadow: 0 0 4px rgba(255, 152, 0, 0.3); }
.conf-other { background: #BDC3C7; }

/* Verify buttons */
.vf-btn { width: 30px; height: 30px; border-radius: 8px; border: 1px solid transparent; background: transparent; cursor: pointer; font-size: 14px; display: flex; align-items: center; justify-content: center; transition: all 0.2s ease; color: var(--text-light); }
.vf-btn:hover { background: var(--hover-bg); border-color: var(--border); color: var(--text); transform: scale(1.1); }
.vf-btn-confirm.active { background: var(--green); color: white; border-color: var(--green); box-shadow: 0 2px 8px rgba(39, 174, 96, 0.3); }
.vf-btn-flag.active { background: var(--yellow); color: white; border-color: var(--yellow); box-shadow: 0 2px 8px rgba(243, 156, 18, 0.3); }
.vf-btn-note.has-note { background: rgba(52, 152, 219, 0.1); border-color: var(--accent); color: var(--accent); }
.vf-note { font-size: 11px; color: var(--text-secondary); padding: 4px 16px 6px 16px; line-height: 1.4; }
.vf-note-input { display:flex; align-items:center; gap:6px; padding:6px 16px; background: var(--entity-bg); border-bottom: 1px solid var(--border-light); }
.vf-note-input input { font-size:12px; padding:5px 10px; border:1px solid var(--border); border-radius:6px; flex:1; font-family:var(--sans); background: var(--input-bg); color: var(--text); transition: border-color 0.15s ease; }
.vf-note-input input:focus { border-color: var(--accent); outline: none; box-shadow: 0 0 0 2px rgba(52, 152, 219, 0.1); }
.vf-note-input button { font-size:11px; padding:4px 12px; border-radius:6px; border:none; background:var(--accent); color:white; cursor:pointer; transition: background 0.15s ease; }
.vf-note-input button:hover { background: var(--accent-hover); }
.vf-original { text-decoration: line-through; color: var(--red); }

/* ─── Confirm/Flag flash animations ─── */
@keyframes confirmFlash { 0% { box-shadow: inset 3px 0 0 var(--green), 0 0 0 0 rgba(39,174,96,0.4); } 50% { box-shadow: inset 3px 0 0 var(--green), 0 0 12px 2px rgba(39,174,96,0.15); } 100% { box-shadow: inset 3px 0 0 var(--green), 0 0 0 0 rgba(39,174,96,0); } }
@keyframes flagFlash { 0% { box-shadow: inset 3px 0 0 var(--red), 0 0 0 0 rgba(231,76,60,0.4); } 50% { box-shadow: inset 3px 0 0 var(--red), 0 0 12px 2px rgba(231,76,60,0.15); } 100% { box-shadow: inset 3px 0 0 var(--red), 0 0 0 0 rgba(231,76,60,0); } }
.field-row.vf-just-confirmed { animation: confirmFlash 0.5s ease; }
.field-row.vf-just-flagged { animation: flagFlash 0.5s ease; }

/* ─── Transaction table ─── */
.txn-section { margin: 8px 0; }
.txn-header { font-size: 11px; font-weight: 700; color: var(--text-light); text-transform: uppercase; letter-spacing: 0.04em; padding: 6px 16px; }
.txn-table { width: 100%; border-collapse: collapse; font-size: 12px; }
.txn-table thead th { font-size: 10px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.03em; color: var(--text-light); padding: 4px 6px; text-align: left; border-bottom: 1px solid var(--border); }
.txn-table tbody tr { border-bottom: 1px solid var(--border-light); transition: background 0.15s ease; }
.txn-table tbody tr:hover { background: var(--hover-bg); }
.txn-table tbody tr.vf-confirmed { background: var(--confirmed-bg); }
.txn-table tbody tr.vf-flagged { background: var(--flagged-bg); }
.txn-table td { padding: 6px 6px; vertical-align: middle; }
.txn-amt { text-align: right; font-family: var(--mono); font-weight: 600; white-space: nowrap; }
.txn-type { font-size: 10px; font-weight: 600; padding: 2px 8px; border-radius: 4px; text-transform: uppercase; }
.txn-type-deposit { background: #D5F5E3; color: #1B7A3D; }
.txn-type-withdrawal { background: #FADBD8; color: #A93226; }
.txn-type-check { background: #FFF3CD; color: #856404; }
.txn-type-fee { background: #F5CBA7; color: #7E5109; }
.txn-type-transfer { background: #D6EAF8; color: #1F618D; }

/* Category dropdown */
.cat-select { font-size: 11px; padding: 2px 4px; border: 1px solid var(--border); border-radius: 4px; background: var(--input-bg); color: var(--text); max-width: 150px; cursor: pointer; }
.cat-select:focus { border-color: var(--accent); outline: none; }
.cat-select.cat-set { background: var(--green-bg); border-color: var(--green); font-weight: 600; }
.cat-select.cat-suggested { background: var(--yellow-bg); border-color: #D4B95E; }
.cat-learned-badge { font-size: 9px; padding: 1px 5px; border-radius: 3px; background: var(--purple-bg); color: var(--purple); font-weight: 600; }
.field-cat-row { display: flex; align-items: center; gap: 8px; padding: 2px 16px 4px; font-size: 11px; color: var(--text-light); }
.field-cat-row label { font-weight: 600; text-transform: uppercase; letter-spacing: 0.03em; font-size: 10px; }

/* ─── Info section (collapsible) ─── */
.info-section { margin: 4px 0; border: 1px solid var(--border-light); border-radius: 8px; overflow: hidden; }
.info-toggle { display: flex; align-items: center; gap: 8px; padding: 8px 16px; background: var(--entity-bg); cursor: pointer; user-select: none; font-size: 11px; font-weight: 700; color: var(--text-light); text-transform: uppercase; letter-spacing: 0.04em; transition: background 0.15s ease; }
.info-toggle:hover { background: var(--hover-bg); }
.info-toggle-arrow { font-size: 10px; transition: transform 0.2s; }
.info-toggle-arrow.open { transform: rotate(90deg); }
.info-field { display: flex; padding: 4px 12px; font-size: 12px; border-bottom: 1px solid var(--border-light); }
.info-field-name { flex: 0 0 45%; color: var(--text-secondary); }
.info-field-val { flex: 1; font-weight: 500; }

/* ═══ CLIENTS SECTION ═══ */
.client-list { display: grid; grid-template-columns: repeat(auto-fill, minmax(280px, 1fr)); gap: 16px; }
.client-card { background: var(--bg-card); border: 1px solid var(--border-light); border-radius: var(--radius); padding: 16px; cursor: pointer; transition: var(--transition); }
.client-card:hover { border-color: var(--accent); box-shadow: var(--shadow-md); transform: translateY(-1px); }
.client-card h4 { font-size: 15px; font-weight: 700; color: var(--navy); margin-bottom: 4px; }
.client-card .client-meta { font-size: 12px; color: var(--text-secondary); display: flex; gap: 12px; flex-wrap: wrap; }
.client-card .client-badges { margin-top: 8px; display: flex; gap: 6px; flex-wrap: wrap; }
.client-detail { display: none; }
.client-detail.visible { display: block; }
.client-back { font-size: 13px; color: var(--accent); cursor: pointer; display: flex; align-items: center; gap: 4px; margin-bottom: 16px; font-weight: 500; }
.client-back:hover { text-decoration: underline; }

/* Client tabs */
.client-tabs { display: flex; gap: 0; border-bottom: 2px solid var(--border); margin-bottom: 16px; }
.client-tab { padding: 10px 20px; font-size: 13px; font-weight: 600; color: var(--text-secondary); cursor: pointer; border-bottom: 2px solid transparent; margin-bottom: -2px; transition: var(--transition); }
.client-tab:hover { color: var(--text); }
.client-tab.active { color: var(--accent); border-bottom-color: var(--accent); }
.client-tab-content { display: none; }
.client-tab-content.active { display: block; }

/* Context uploads */
.context-doc { display: flex; align-items: center; gap: 12px; padding: 10px 12px; border: 1px solid var(--border-light); border-radius: 6px; margin-bottom: 8px; }
.context-doc-icon { width: 36px; height: 36px; background: #EBF5FB; border-radius: 6px; display: flex; align-items: center; justify-content: center; font-size: 16px; }
.context-doc-info { flex: 1; }
.context-doc-info .name { font-size: 13px; font-weight: 600; }
.context-doc-info .meta { font-size: 11px; color: var(--text-light); }

/* Instructions */
.instruction-item { display: flex; align-items: flex-start; gap: 10px; padding: 10px 12px; border: 1px solid var(--border-light); border-radius: 6px; margin-bottom: 6px; }
.instruction-item .inst-text { flex: 1; font-size: 13px; }
.instruction-item .inst-date { font-size: 11px; color: var(--text-light); white-space: nowrap; }

/* Completeness */
.completeness-item { display: flex; align-items: center; gap: 10px; padding: 8px 12px; border-bottom: 1px solid var(--border-light); font-size: 13px; }
.completeness-icon { width: 24px; text-align: center; font-size: 16px; }
.completeness-info { flex: 1; }
.completeness-info .ci-form { font-weight: 600; }
.completeness-info .ci-payer { color: var(--text-secondary); font-size: 12px; }

/* ═══ BATCH CATEGORIZE ═══ */
.batch-stats { display: grid; grid-template-columns: repeat(3, 1fr); gap: 12px; margin-bottom: 20px; }
.batch-stat { background: var(--bg-card); border: 1px solid var(--border-light); border-radius: var(--radius); padding: 16px; text-align: center; }
.batch-stat .stat-num { font-size: 28px; font-weight: 700; color: var(--navy); }
.batch-stat .stat-label { font-size: 12px; color: var(--text-secondary); margin-top: 2px; }
.vendor-group { border: 1px solid var(--border-light); border-radius: var(--radius); margin-bottom: 8px; overflow: hidden; }
.vendor-group-header { display: flex; align-items: center; gap: 12px; padding: 10px 14px; background: #FAFAF8; cursor: pointer; transition: background 0.1s; }
.vendor-group-header:hover { background: #F0EFEC; }
.vendor-group-header .vg-name { font-weight: 600; flex: 1; }
.vendor-group-header .vg-count { font-size: 12px; color: var(--text-secondary); }
.vendor-group-header .vg-amount { font-family: var(--mono); font-weight: 600; font-size: 13px; }
.vendor-group-items { display: none; padding: 0 14px 8px; }
.vendor-group-items.open { display: block; }

/* ═══ HISTORY ═══ */
.history-filters { display: flex; gap: 8px; margin-bottom: 16px; flex-wrap: wrap; align-items: center; }
.history-filters .form-input { width: 200px; }
.history-filters .form-select { width: 160px; }
.job-status { font-size: 11px; font-weight: 700; text-transform: uppercase; }
.job-status.complete { color: var(--green); }
.job-status.running { color: var(--accent); }
.job-status.failed { color: var(--red); }
.job-status.interrupted { color: var(--yellow); }

/* ═══ EMPTY STATES ═══ */
.empty-state { text-align: center; padding: 48px 24px; color: var(--text-secondary); }
.empty-state svg { width: 48px; height: 48px; color: var(--border); margin-bottom: 12px; }
.empty-state h3 { font-size: 16px; color: var(--text); margin-bottom: 4px; }
.empty-state p { font-size: 13px; }

/* ═══ KEYBOARD HELP ═══ */
.kbd-overlay { display: none; position: fixed; inset: 0; background: rgba(0,0,0,0.5); z-index: 1000; align-items: center; justify-content: center; }
.kbd-overlay.visible { display: flex; }
.kbd-card { background: white; border-radius: var(--radius-lg); padding: 24px; max-width: 400px; box-shadow: var(--shadow-lg); }
.kbd-card h3 { margin-bottom: 12px; }
.kbd-row { display: flex; justify-content: space-between; padding: 4px 0; font-size: 13px; }
kbd { background: var(--bg); border: 1px solid var(--border); border-radius: 4px; padding: 2px 8px; font-size: 12px; font-family: var(--mono); }

/* ═══ RESPONSIVE ═══ */
@media (max-width: 900px) {
  .sidebar { width: 60px; }
  .sidebar-brand h1, .sidebar-brand p, .nav-item span, .sidebar-footer label, .sidebar-footer .theme-picker, .sidebar-footer p { display: none; }
  .nav-item { justify-content: center; padding: 12px; }
  .main { margin-left: 60px; }
  .review-split { grid-template-columns: 1fr; }
  .form-row { grid-template-columns: 1fr; }
}

/* Modal */
.modal-overlay { position:fixed; inset:0; background:rgba(0,0,0,0.4); z-index:9999; display:none; align-items:center; justify-content:center; }
.modal-overlay.visible { display:flex; }
.modal-content { background:white; border-radius:12px; padding:24px; width:420px; max-width:90vw; box-shadow:0 20px 60px rgba(0,0,0,0.3); }

/* ═══ GUIDED REVIEW ═══ */
.guided-header { padding: 12px 24px; background: var(--bg-card); border-bottom: 1px solid var(--border); display: flex; align-items: center; justify-content: space-between; gap: 16px; min-height: 56px; }
.guided-progress { display: flex; align-items: center; gap: 12px; }
.guided-progress-text { font-size: 14px; font-weight: 700; color: var(--text); white-space: nowrap; font-variant-numeric: tabular-nums; }
.guided-progress-bar { width: 220px; height: 8px; background: var(--border); border-radius: 4px; overflow: hidden; }
.guided-progress-fill { height: 100%; background: linear-gradient(90deg, var(--green), var(--accent)); border-radius: 4px; transition: width 0.4s cubic-bezier(0.4, 0, 0.2, 1); box-shadow: 0 0 8px rgba(39, 174, 96, 0.3); }
.guided-actions-top { display: flex; gap: 8px; align-items: center; }
.guided-split { display: grid; grid-template-columns: 3fr 2fr; height: calc(100vh - 68px); }
.guided-evidence { background: var(--pdf-bg); overflow: auto; display: flex; align-items: flex-start; justify-content: center; padding: 24px; }
.guided-evidence img { max-width: 95%; height: auto; box-shadow: 0 4px 20px rgba(0,0,0,0.35), 0 0 1px rgba(0,0,0,0.2); border-radius: 6px; background: white; transition: box-shadow 0.3s ease; }
.guided-detail { padding: 36px; display: flex; flex-direction: column; gap: 20px; background: var(--bg-card); overflow-y: auto; }
.guided-field-label { font-size: 14px; font-weight: 600; color: var(--text-secondary); text-transform: uppercase; letter-spacing: 0.05em; }
.guided-field-dest { font-size: 13px; color: var(--text-light); margin-top: -8px; }
.guided-field-value { font-size: 36px; font-weight: 700; color: var(--text); font-family: var(--mono); padding: 28px; background: var(--bg); border-radius: var(--radius-lg); border: 2px solid var(--border); text-align: center; word-break: break-all; transition: border-color 0.3s ease, box-shadow 0.3s ease; box-shadow: var(--shadow-sm); }
.guided-field-value:hover { border-color: var(--accent); box-shadow: var(--shadow-md); }
.guided-field-meta { display: flex; gap: 16px; font-size: 12px; color: var(--text-secondary); flex-wrap: wrap; }
.guided-field-meta .badge { font-size: 11px; }
.guided-actions { display: grid; grid-template-columns: 1fr 1fr; gap: 12px; margin-top: 16px; }
.guided-actions .btn { padding: 16px 20px; font-size: 15px; font-weight: 700; justify-content: center; border-radius: 10px; transition: all 0.2s ease; box-shadow: var(--shadow-sm); }
.guided-actions .btn:hover { transform: translateY(-1px); box-shadow: var(--shadow-md); }
.guided-actions .btn:active { transform: translateY(0); box-shadow: var(--shadow-sm); }
.guided-actions .btn kbd { background: rgba(255,255,255,0.2); border: 1px solid rgba(255,255,255,0.3); padding: 2px 8px; border-radius: 4px; font-size: 11px; margin-left: 8px; }
.guided-edit-area { background: var(--bg); border-radius: var(--radius-lg); padding: 20px; display: flex; flex-direction: column; gap: 12px; border: 1px solid var(--border-light); }
.guided-edit-input { font-size: 24px; padding: 14px; text-align: center; font-family: var(--mono); font-weight: 700; border: 2px solid var(--accent); border-radius: 10px; background: var(--input-bg); color: var(--text); transition: box-shadow 0.2s ease; }
.guided-edit-input:focus { outline: none; box-shadow: 0 0 0 4px rgba(52,152,219,0.2); }
.guided-edit-btns { display: flex; gap: 8px; justify-content: center; }
.guided-complete { text-align: center; padding: 60px 24px; }
.guided-complete h2 { font-size: 24px; color: var(--green); margin-bottom: 8px; }
.guided-complete p { font-size: 15px; color: var(--text-secondary); margin-bottom: 24px; }
.guided-lock-banner { background: var(--yellow-bg); color: #7E5109; padding: 8px 16px; border-radius: 6px; font-size: 13px; font-weight: 500; text-align: center; }
@media (max-width: 768px) {
  .guided-split { grid-template-columns: 1fr; }
  .guided-evidence { max-height: 40vh; }
  .guided-field-value { font-size: 24px; padding: 16px; }
  .guided-actions { grid-template-columns: 1fr; }
}

/* ═══ DARK THEME ═══ */
[data-theme="dark"] {
  --bg: #0D1117;
  --bg-card: #161B22;
  --bg-sidebar: #010409;
  --bg-sidebar-hover: #1C2333;
  --bg-sidebar-active: #253044;
  --navy: #E6EDF3;
  --navy-light: #B1BAC4;
  --accent: #58A6FF;
  --accent-hover: #79C0FF;
  --green: #3FB950;
  --green-bg: #0D2818;
  --yellow: #D29922;
  --yellow-bg: #2A1F00;
  --red: #F85149;
  --red-bg: #2D0B0E;
  --purple: #BC8CFF;
  --purple-bg: #1C0F2B;
  --text: #E6EDF3;
  --text-secondary: #8B949E;
  --text-light: #6E7681;
  --border: #30363D;
  --border-light: #21262D;
  --shadow-sm: 0 1px 3px rgba(0,0,0,0.3);
  --shadow-md: 0 4px 12px rgba(0,0,0,0.4);
  --shadow-lg: 0 8px 24px rgba(0,0,0,0.5);
  --hover-bg: #1C2128;
  --focus-bg: #1A2332;
  --confirmed-bg: #0D2818;
  --corrected-bg: #2A1F00;
  --flagged-bg: #2D1500;
  --entity-bg: #1C2128;
  --console-bg: #010409;
  --console-text: #8B949E;
  --console-hl: #58A6FF;
  --pdf-bg: #010409;
  --upload-bg: #161B22;
  --upload-hover-bg: #1A2332;
  --badge-blue-bg: #1A2332;
  --badge-gray-bg: #21262D;
  --pill-bg: #21262D;
  --btn-bg: #21262D;
  --input-bg: #0D1117;
}
[data-theme="dark"] ::selection { background: var(--accent); color: #0D1117; }
[data-theme="dark"] .sidebar { border-right: 1px solid var(--border); }

/* ═══ SYNTHWAVE THEME ═══ */
[data-theme="synthwave"] {
  --bg: #13081E;
  --bg-card: #1A0F2E;
  --bg-sidebar: #0A0514;
  --bg-sidebar-hover: #241542;
  --bg-sidebar-active: #331D5C;
  --navy: #F0E6FF;
  --navy-light: #C4A8FF;
  --accent: #FF2D95;
  --accent-hover: #FF5CAF;
  --green: #00F0FF;
  --green-bg: #001A1F;
  --yellow: #FFD700;
  --yellow-bg: #2A2200;
  --red: #FF3860;
  --red-bg: #2D0A14;
  --purple: #BD93F9;
  --purple-bg: #1C0D35;
  --text: #F0E6FF;
  --text-secondary: #A78BCC;
  --text-light: #7B5EAA;
  --border: #2D1B4E;
  --border-light: #1F1336;
  --shadow-sm: 0 1px 4px rgba(255,45,149,0.1);
  --shadow-md: 0 4px 14px rgba(255,45,149,0.15);
  --shadow-lg: 0 8px 28px rgba(255,45,149,0.2);
  --hover-bg: #221340;
  --focus-bg: #261850;
  --confirmed-bg: #001A1F;
  --corrected-bg: #2A2200;
  --flagged-bg: #2D0A14;
  --entity-bg: #1F1336;
  --console-bg: #0A0514;
  --console-text: #A78BCC;
  --console-hl: #FF2D95;
  --pdf-bg: #0A0514;
  --upload-bg: #1A0F2E;
  --upload-hover-bg: #261850;
  --badge-blue-bg: #1C0D35;
  --badge-gray-bg: #1F1336;
  --pill-bg: #1F1336;
  --btn-bg: #1F1336;
  --input-bg: #13081E;
}
[data-theme="synthwave"] ::selection { background: #FF2D95; color: #13081E; }
[data-theme="synthwave"] .sidebar { border-right: 1px solid #2D1B4E; }
[data-theme="synthwave"] .card { border-color: #2D1B4E; box-shadow: 0 0 20px rgba(255,45,149,0.08), 0 0 60px rgba(189,147,249,0.04), var(--shadow-sm); }
[data-theme="synthwave"] .card:hover { box-shadow: 0 0 30px rgba(255,45,149,0.15), 0 0 80px rgba(189,147,249,0.08); }
[data-theme="synthwave"] .nav-item.active { border-left-color: #FF2D95; }
[data-theme="synthwave"] .btn.primary { background: linear-gradient(135deg, #FF2D95, #BD93F9); border: none; }
[data-theme="synthwave"] .btn.primary:hover { background: linear-gradient(135deg, #FF5CAF, #D4B0FF); }
/* Synthwave glow effects */
[data-theme="synthwave"] .page-header h2 { text-shadow: 0 0 20px rgba(255,45,149,0.5), 0 0 40px rgba(255,45,149,0.2); }
[data-theme="synthwave"] .dash-kpi-value { text-shadow: 0 0 16px rgba(0,240,255,0.4), 0 0 40px rgba(0,240,255,0.15); }
[data-theme="synthwave"] .dash-kpi-card { border: 1px solid rgba(189,147,249,0.2); }
[data-theme="synthwave"] .dash-kpi-card:hover { border-color: rgba(255,45,149,0.4); box-shadow: 0 0 30px rgba(255,45,149,0.15), 0 4px 20px rgba(0,0,0,0.3); }
[data-theme="synthwave"] .dash-pipeline-count { text-shadow: 0 0 12px rgba(255,45,149,0.4); }
[data-theme="synthwave"] .card-header h3 { text-shadow: 0 0 10px rgba(189,147,249,0.3); }
[data-theme="synthwave"] .dash-kpi-icon { box-shadow: 0 0 16px rgba(255,45,149,0.25); }
[data-theme="synthwave"] .badge { box-shadow: 0 0 8px rgba(255,45,149,0.2); }
@keyframes synthBorderGlow {
  0%, 100% { border-color: rgba(189,147,249,0.2); }
  50% { border-color: rgba(255,45,149,0.35); }
}
[data-theme="synthwave"] .dash-chart-card { animation: synthBorderGlow 4s ease-in-out infinite; }

/* Synthwave review enhancements */
[data-theme="synthwave"] .field-row.vf-confirmed { box-shadow: inset 3px 0 0 var(--green), 0 0 8px rgba(0, 240, 255, 0.08); }
[data-theme="synthwave"] .vf-btn-confirm.active { box-shadow: 0 0 12px rgba(0, 240, 255, 0.3); }
[data-theme="synthwave"] .verify-progress-fill, [data-theme="synthwave"] .guided-progress-fill { background: linear-gradient(90deg, var(--green), #FF2D95); box-shadow: 0 0 12px rgba(255, 45, 149, 0.4); }
[data-theme="synthwave"] .guided-field-value { border-color: #2D1B4E; box-shadow: 0 0 20px rgba(189, 147, 249, 0.08); }
[data-theme="synthwave"] .guided-field-value:hover { border-color: #FF2D95; box-shadow: 0 0 30px rgba(255, 45, 149, 0.15); }
[data-theme="synthwave"] .review-pdf img, [data-theme="synthwave"] .guided-evidence img { box-shadow: 0 4px 20px rgba(0,0,0,0.5), 0 0 20px rgba(189, 147, 249, 0.06); }

/* ═══ RETRO THEME (Amber Terminal) ═══ */
[data-theme="retro"] {
  --bg: #0C0C0C;
  --bg-card: #141410;
  --bg-sidebar: #080804;
  --bg-sidebar-hover: #1A1A10;
  --bg-sidebar-active: #2A2A18;
  --navy: #FFB000;
  --navy-light: #CC8C00;
  --accent: #FFB000;
  --accent-hover: #FFC84D;
  --green: #33FF33;
  --green-bg: #0A1A0A;
  --yellow: #FFB000;
  --yellow-bg: #1A1400;
  --red: #FF3333;
  --red-bg: #1A0808;
  --purple: #FF8C00;
  --purple-bg: #1A1000;
  --text: #FFB000;
  --text-secondary: #CC8C00;
  --text-light: #8A6000;
  --border: #332B00;
  --border-light: #1F1A00;
  --shadow-sm: 0 0 4px rgba(255,176,0,0.08);
  --shadow-md: 0 0 10px rgba(255,176,0,0.1);
  --shadow-lg: 0 0 20px rgba(255,176,0,0.12);
  --sans: 'SF Mono', 'Menlo', 'Consolas', 'Courier New', monospace;
  --hover-bg: #1A1A10;
  --focus-bg: #1F1F0A;
  --confirmed-bg: #0A1A0A;
  --corrected-bg: #1A1400;
  --flagged-bg: #1A0808;
  --entity-bg: #1A1A10;
  --console-bg: #080804;
  --console-text: #CC8C00;
  --console-hl: #FFB000;
  --pdf-bg: #080804;
  --upload-bg: #141410;
  --upload-hover-bg: #1F1F0A;
  --badge-blue-bg: #1A1400;
  --badge-gray-bg: #1A1A10;
  --pill-bg: #1A1A10;
  --btn-bg: #1A1A10;
  --input-bg: #0C0C0C;
}
[data-theme="retro"] ::selection { background: #FFB000; color: #0C0C0C; }
[data-theme="retro"] .sidebar { border-right: 1px solid #332B00; }
[data-theme="retro"] .card { border-color: #332B00; box-shadow: 0 0 12px rgba(255,176,0,0.06), inset 0 0 30px rgba(255,176,0,0.02); }
[data-theme="retro"] .card:hover { box-shadow: 0 0 20px rgba(255,176,0,0.12), inset 0 0 30px rgba(255,176,0,0.03); }
[data-theme="retro"] .nav-item.active { border-left-color: #FFB000; }
[data-theme="retro"] body::after {
  content: ''; position: fixed; top: 0; left: 0; right: 0; bottom: 0;
  background: repeating-linear-gradient(0deg, rgba(255,176,0,0.02) 0px, rgba(255,176,0,0.02) 1px, transparent 1px, transparent 3px);
  pointer-events: none; z-index: 9999;
}
[data-theme="retro"] *:focus { box-shadow: 0 0 8px rgba(255,176,0,0.3); }
/* Retro phosphor glow */
[data-theme="retro"] .page-header h2 { text-shadow: 0 0 8px rgba(255,176,0,0.6), 0 0 20px rgba(255,176,0,0.3); }
[data-theme="retro"] .dash-kpi-value { text-shadow: 0 0 10px rgba(255,176,0,0.5), 0 0 30px rgba(255,176,0,0.2); }
[data-theme="retro"] .dash-pipeline-count { text-shadow: 0 0 8px rgba(255,176,0,0.4); }
[data-theme="retro"] .card-header h3 { text-shadow: 0 0 6px rgba(255,176,0,0.3); }
[data-theme="retro"] .dash-kpi-icon { box-shadow: 0 0 12px rgba(255,176,0,0.2); }
[data-theme="retro"] .dash-kpi-card { border-color: rgba(255,176,0,0.15); }
[data-theme="retro"] .dash-kpi-card:hover { border-color: rgba(255,176,0,0.35); box-shadow: 0 0 20px rgba(255,176,0,0.1); }
[data-theme="retro"] .badge { text-shadow: 0 0 4px rgba(255,176,0,0.3); }
@keyframes retroFlicker {
  0%, 100% { opacity: 1; }
  92% { opacity: 1; }
  93% { opacity: 0.8; }
  94% { opacity: 1; }
  96% { opacity: 0.9; }
  97% { opacity: 1; }
}
[data-theme="retro"] .dash-kpi-value { animation: retroFlicker 8s linear infinite; }

/* Retro review enhancements */
[data-theme="retro"] .review-pdf img, [data-theme="retro"] .guided-evidence img { box-shadow: 0 0 20px rgba(255, 176, 0, 0.1), 0 4px 16px rgba(0,0,0,0.6); border: 1px solid rgba(255, 176, 0, 0.15); }
[data-theme="retro"] .verify-progress-fill, [data-theme="retro"] .guided-progress-fill { background: var(--green); box-shadow: 0 0 8px rgba(51, 255, 51, 0.4); }
[data-theme="retro"] .vf-btn-confirm.active { box-shadow: 0 0 8px rgba(51, 255, 51, 0.3); }
[data-theme="retro"] .field-row.vf-confirmed { box-shadow: inset 3px 0 0 var(--green), 0 0 6px rgba(51, 255, 51, 0.06); }

/* ═══ THEME PICKER ═══ */
.theme-picker { display: flex; gap: 6px; margin-top: 10px; justify-content: center; }
.theme-dot { width: 20px; height: 20px; border-radius: 50%; cursor: pointer; border: 2px solid transparent; transition: var(--transition); position: relative; }
.theme-dot:hover { transform: scale(1.2); }
.theme-dot.active { border-color: white; box-shadow: 0 0 6px rgba(255,255,255,0.4); }
.theme-dot[data-theme="light"] { background: linear-gradient(135deg, #F7F6F3, #3498DB); }
.theme-dot[data-theme="dark"] { background: linear-gradient(135deg, #0D1117, #58A6FF); }
.theme-dot[data-theme="synthwave"] { background: linear-gradient(135deg, #13081E, #FF2D95); }
.theme-dot[data-theme="retro"] { background: linear-gradient(135deg, #0C0C0C, #FFB000); }
</style>
</head>
<body>
<div class="app">

<!-- ═══ SIDEBAR ═══ -->
<aside class="sidebar">
  <div class="sidebar-brand">
    <h1>OathLedger</h1>
    <p>Deterministic Accounting Intelligence</p>
  </div>
  <nav class="sidebar-nav">
    <a class="nav-item active" onclick="showSection('upload')" data-section="upload">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg>
      <span>Upload</span>
    </a>
    <a class="nav-item" onclick="showSection('inbox')" data-section="inbox">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M22 12h-6l-2 3h-4l-2-3H2"/><path d="M5.45 5.11L2 12v6a2 2 0 002 2h16a2 2 0 002-2v-6l-3.45-6.89A2 2 0 0016.76 4H7.24a2 2 0 00-1.79 1.11z"/></svg>
      <span>Inbox</span>
      <span class="nav-badge" id="inboxCount" style="display:none">0</span>
    </a>
    <a class="nav-item" onclick="openGridReview()" data-section="review" id="navReview" style="display:none">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M14 2H6a2 2 0 00-2 2v16a2 2 0 002 2h12a2 2 0 002-2V8z"/><polyline points="14 2 14 8 20 8"/><line x1="16" y1="13" x2="8" y2="13"/><line x1="16" y1="17" x2="8" y2="17"/></svg>
      <span>Review</span>
    </a>
    <a class="nav-item" onclick="openGuidedReview()" data-section="guided-review" id="navGuidedReview" style="display:none">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M9 11l3 3L22 4"/><path d="M21 12v7a2 2 0 01-2 2H5a2 2 0 01-2-2V5a2 2 0 012-2h11"/></svg>
      <span>Audit Check</span>
    </a>
    <a class="nav-item" onclick="showSection('clients')" data-section="clients">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M20 21v-2a4 4 0 00-4-4H8a4 4 0 00-4 4v2"/><circle cx="12" cy="7" r="4"/></svg>
      <span>Clients &amp; PY Docs</span>
    </a>
    <a class="nav-item" onclick="showSection('batch')" data-section="batch" style="display:none">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="3" y="3" width="7" height="7"/><rect x="14" y="3" width="7" height="7"/><rect x="3" y="14" width="7" height="7"/><rect x="14" y="14" width="7" height="7"/></svg>
      <span>Categorize</span>
    </a>
    <a class="nav-item" onclick="showSection('history')" data-section="history">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="12" cy="12" r="10"/><polyline points="12 6 12 12 16 14"/></svg>
      <span>History</span>
      <span class="nav-badge" id="historyCount">0</span>
    </a>
  </nav>
  <div class="sidebar-footer">
    <label>Reviewer
      <input type="text" id="reviewerInitials" maxlength="4" placeholder="JW" value="">
    </label>
    <div class="theme-picker">
      <div class="theme-dot active" data-theme="light" title="Light" onclick="setTheme('light')"></div>
      <div class="theme-dot" data-theme="dark" title="Dark" onclick="setTheme('dark')"></div>
      <div class="theme-dot" data-theme="synthwave" title="Synthwave" onclick="setTheme('synthwave')"></div>
      <div class="theme-dot" data-theme="retro" title="Retro" onclick="setTheme('retro')"></div>
    </div>
  </div>
</aside>

<!-- ═══ MAIN CONTENT ═══ -->
<div class="main">
<div id="toast-container"></div>

<!-- ═══ UPLOAD SECTION ═══ -->
<div class="section active" id="sec-upload">
  <div class="page-header"><h2>Upload Document</h2><p>Scan a PDF to extract structured data</p></div>
  <div class="page-content">
    <div id="apiKeyBanner" style="display:none;background:var(--danger,#e74c3c);color:#fff;padding:14px 20px;border-radius:8px;margin-bottom:16px;font-size:13px;">
      <strong>&#x26A0; Anthropic API Key Not Set</strong>
      <p style="margin:6px 0 10px;opacity:0.9">Processing requires an Anthropic API key. Paste yours below to save it permanently.</p>
      <div style="display:flex;gap:8px;align-items:center">
        <input type="password" id="apiKeyInput" placeholder="sk-ant-api03-..." style="flex:1;padding:8px 12px;border:none;border-radius:4px;font-size:13px;font-family:monospace;background:rgba(255,255,255,0.95);color:#222;">
        <button onclick="saveApiKey()" style="padding:8px 16px;border:none;border-radius:4px;background:#fff;color:var(--danger,#e74c3c);font-weight:600;cursor:pointer;font-size:13px;white-space:nowrap">Save Key</button>
      </div>
    </div>
    <div class="card">
      <div class="card-body">
        <div class="upload-area" id="dropZone" onclick="document.getElementById('fileInput').click()">
          <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.5"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg>
          <h3>Drop PDF here or click to browse</h3>
          <p>Supports scanned tax documents, bank statements, invoices, checks, and more</p>
        </div>
        <input type="file" id="fileInput" accept=".pdf" style="display:none" onchange="handleFile(this)">

        <div class="upload-form" id="uploadForm">
          <div class="upload-file-name" id="fileName">
            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" width="16" height="16"><path d="M14 2H6a2 2 0 00-2 2v16a2 2 0 002 2h12a2 2 0 002-2V8z"/><polyline points="14 2 14 8 20 8"/></svg>
            <span id="fileNameText"></span>
            <button class="btn btn-ghost btn-sm" onclick="resetUpload()">Change</button>
          </div>

          <div class="form-row">
            <div class="form-group">
              <label class="form-label">Client</label>
              <div style="display:flex;gap:8px;align-items:center">
                <select id="clientName" class="form-input" style="flex:1">
                  <option value="">— Select client —</option>
                </select>
                <button class="btn btn-secondary btn-sm" onclick="openNewClientModal()" title="Create new client" style="white-space:nowrap">+ New</button>
              </div>
              <a href="#" onclick="event.preventDefault(); const cn=document.getElementById('clientName').value.trim(); if(cn){showSection('clients');setTimeout(()=>openClientDetail(cn),200);} else {showToast('Select a client first','error');}" style="font-size:11px; color:var(--accent); text-decoration:none; margin-top:4px; display:inline-block;">&#x1F4C2; Upload prior-year docs / manage instructions</a>
            </div>
            <div class="form-group">
              <label class="form-label">Tax Year</label>
              <input type="number" id="taxYear" class="form-input" value="2025" min="2000" max="2030">
            </div>
          </div>

          <div class="form-group">
            <label class="form-label">Document Type <span style="font-size:11px;color:var(--text-muted);font-weight:400">&mdash; helps AI focus; final type detected automatically</span></label>
            <div class="pill-group" id="docTypePills"></div>
          </div>

          <div class="form-group">
            <label class="form-label">Output Format</label>
            <div class="pill-group" id="outputFormatPills"></div>
          </div>

          <details style="margin-top:8px">
            <summary style="cursor:pointer;font-size:13px;color:var(--text-muted);user-select:none">&#x2699; Advanced Options</summary>
            <div style="margin-top:8px">
              <div class="form-group">
                <label class="form-label">AI Instructions</label>
                <textarea id="aiInstructions" class="form-input" rows="3" placeholder="Tell the AI how to handle this document. Example: 'This is a trust return — extract K-1 box 1-14 only' or 'Combine all 1099-DIV pages into one entry per payer'"></textarea>
              </div>
            </div>
          </details>

          <details style="margin-bottom:16px">
            <summary style="font-size:12px; font-weight:600; color:var(--text-secondary); cursor:pointer; padding:4px 0;">Advanced Options</summary>
            <div style="padding-top:12px">
              <div class="form-group">
                <label class="form-label">Notes for Extraction</label>
                <textarea id="userNotes" class="form-input" rows="2" placeholder="Optional context about this document..."></textarea>
              </div>
              <div class="form-group">
                <label style="display:flex;align-items:center;gap:6px;font-size:13px;cursor:pointer">
                  <input type="checkbox" id="skipVerify"> Skip AI verification (faster, lower cost)
                </label>
              </div>
              <div class="form-group">
                <label style="display:flex;align-items:center;gap:6px;font-size:13px;cursor:pointer">
                  <input type="checkbox" id="disablePii"> Disable PII tokenization
                </label>
              </div>
              <div class="form-group">
                <label style="display:flex;align-items:center;gap:6px;font-size:13px;cursor:pointer">
                  <input type="checkbox" id="useOcrFirst"> Use OCR-first mode (lower cost, less accurate)
                </label>
              </div>
            </div>
          </details>

          <button class="btn btn-primary" id="startBtn" onclick="startExtraction()" style="width:100%;justify-content:center;padding:12px;">
            Start Extraction
          </button>
        </div>
      </div>
    </div>
  </div>
</div>

<!-- ═══ INBOX SECTION (B1: Review Chain) ═══ -->
<div class="section" id="sec-inbox">
  <div class="page-header">
    <h2>Inbox</h2>
    <p>Documents assigned to your review stage</p>
  </div>
  <div class="page-content">
    <div id="inboxContent" style="display:flex;flex-direction:column;gap:12px">
      <p style="color:var(--text-muted)">Loading...</p>
    </div>
  </div>
</div>

<!-- ═══ PROCESSING SECTION ═══ -->
<div class="section" id="sec-processing">
  <div class="page-header"><h2>Processing</h2><p id="processingFile"></p></div>
  <div class="page-content">
    <div class="card processing-card">
      <div class="card-body">
        <div class="progress-label">
          <span id="procStage">Starting...</span>
          <span id="procPct">0%</span>
          <span id="procElapsed" style="display:none; font-size:11px; color:var(--text-light); margin-left:8px"></span>
        </div>
        <div class="progress-bar"><div class="progress-fill" id="procBar" style="width:0%"></div></div>
        <div class="console-output" id="procConsole"></div>
        <div style="margin-top:16px; text-align:center">
          <button class="btn btn-secondary btn-sm" id="procCancelBtn" onclick="cancelJob()" style="display:none">Cancel</button>
          <button class="btn btn-secondary btn-sm" id="procReviewEarlyBtn" style="display:none;margin-left:8px" onclick="openEarlyReview()">Review extracted fields</button>
        </div>
      </div>
    </div>
  </div>
</div>

<!-- ═══ REVIEW SECTION ═══ -->
<div class="section" id="sec-review">
  <div id="reviewPartialBanner" style="display:none;padding:8px 16px;background:#FFF3CD;color:#856404;text-align:center;font-size:13px;border-bottom:1px solid #E0C96B"></div>
  <div class="review-header">
    <div class="review-nav">
      <button class="review-nav-btn" onclick="prevPage()" title="Previous page (\u2190)"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><path d="M15 18l-6-6 6-6"/></svg></button>
      <span class="review-pager" id="reviewPager">1 / 1</span>
      <button class="review-nav-btn" onclick="nextPage()" title="Next page (\u2192)"><svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><path d="M9 18l6-6-6-6"/></svg></button>
    </div>
    <div class="review-center">
      <div class="verify-stats" id="verifyStats"></div>
      <div class="verify-progress"><div class="verify-progress-fill" id="verifyBar" style="width:0%"></div></div>
    </div>
    <div class="review-actions">
      <button class="btn btn-secondary btn-sm" onclick="reextractPage()" title="Re-extract this page with AI instructions">&#x21BB; Re-extract</button>
      <button class="btn btn-secondary btn-sm" onclick="toggleAiChat()" title="Ask AI about this page" id="aiChatToggle">&#x1F4AC; Ask AI</button>
      <button class="btn btn-ghost btn-sm" title="Keyboard shortcuts (?)" onclick="toggleKbdHelp()">&#x2328;</button>
      <button class="btn btn-accent btn-sm" onclick="openGuidedReview()" title="Audit check — review one field at a time with evidence">&#x1F50D; Audit Check</button>
    </div>
  </div>
  <!-- Client instructions banner (if any) -->
  <div id="reviewInstructionsBanner" style="display:none; padding:8px 20px; background:#FFF8E8; border-bottom:1px solid #F5E6C8; font-size:12px;"></div>
  <!-- B7-UX: Doc type mismatch banner -->
  <div id="reviewDocTypeBanner" style="display:none; padding:6px 20px; background:#FEF3E8; border-bottom:1px solid #F5D8B8; font-size:12px; color:#92610A;"></div>
  <div class="review-split">
    <div class="review-pdf" id="pdfViewer"></div>
    <div class="review-fields" id="fieldsPanel"></div>
  </div>
  <!-- AI Chat Panel -->
  <div id="aiChatPanel" style="display:none; border-top:2px solid var(--accent); background:var(--bg-card);">
    <div style="padding:12px 20px; display:flex; align-items:center; justify-content:space-between; border-bottom:1px solid var(--border);">
      <span style="font-size:13px; font-weight:700; color:var(--navy);">&#x1F4AC; AI Assistant — Page <span id="aiChatPage">1</span></span>
      <button class="btn btn-ghost btn-sm" onclick="toggleAiChat()" title="Close">&#x2716;</button>
    </div>
    <div id="aiChatMessages" style="padding:12px 20px; max-height:200px; overflow-y:auto; font-size:13px;"></div>
    <div style="padding:8px 20px 12px; display:flex; gap:8px;">
      <input type="text" id="aiChatInput" class="form-input" placeholder="Ask about this page... e.g. 'What's in box 14?' or 'Is this K-1 or K-3?'" style="flex:1; font-size:13px;" onkeydown="if(event.key==='Enter')sendAiChat()">
      <button class="btn btn-primary btn-sm" onclick="sendAiChat()">Send</button>
    </div>
  </div>
  <!-- Lite Findings Panel (data-guarded: only visible when ardent_summary exists) -->
  <div id="liteFindingsPanel" style="display:none; border-top:2px solid var(--purple); background:var(--bg-card);">
    <details>
      <summary style="padding:10px 20px; cursor:pointer; font-size:13px; font-weight:700; color:var(--navy); user-select:none; display:flex; align-items:center; gap:10px;">
        <span>Lite Findings</span>
        <span id="liteFindingsStatusBadge" style="font-size:11px; font-weight:700; padding:2px 10px; border-radius:10px; color:#fff;"></span>
        <span id="liteFindingsCounts" style="font-size:11px; color:var(--text-secondary); margin-left:auto;"></span>
      </summary>
      <div style="padding:0 20px 16px;">
        <!-- Severity count pills -->
        <div id="liteFindingsSeverityRow" style="display:flex; gap:8px; margin-bottom:12px; flex-wrap:wrap;"></div>
        <!-- Findings list -->
        <div id="liteFindingsList" style="max-height:320px; overflow-y:auto;"></div>
        <!-- Provenance -->
        <div id="liteFindingsProvenance" style="margin-top:8px; font-size:11px; color:var(--text-light);"></div>
      </div>
    </details>
  </div>
</div>

<!-- ═══ GUIDED REVIEW SECTION ═══ -->
<div class="section" id="sec-guided-review">
  <div class="guided-header">
    <div style="display:flex;align-items:center;gap:12px">
      <button class="btn btn-secondary btn-sm" onclick="guidedGoBack()" id="guidedBackBtn" disabled title="Go to previous field (Backspace)">&#9664; Back</button>
      <div class="guided-progress">
        <span class="guided-progress-text" id="guidedProgressText">0 of 0</span>
        <div class="guided-progress-bar">
          <div class="guided-progress-fill" id="guidedProgressBar" style="width:0%"></div>
        </div>
      </div>
    </div>
    <div class="guided-actions-top">
      <span id="guidedStageBadge" style="font-size:12px;font-weight:600;color:#fff;padding:3px 10px;border-radius:12px;margin-right:8px;display:none"></span>
      <span id="guidedReviewerBadge" style="font-size:12px;color:var(--text-muted);margin-right:8px"></span>
      <button class="btn btn-secondary btn-sm" onclick="openGridReview()">&#x2630; List View</button>
      <button class="btn btn-ghost btn-sm" title="Keyboard shortcuts (?)" onclick="toggleKbdHelp()">&#x2328;</button>
    </div>
  </div>
  <div class="guided-split">
    <div class="guided-evidence" id="guidedEvidence">
      <div class="empty-state" style="color:rgba(255,255,255,0.5)"><p>Loading evidence...</p></div>
    </div>
    <div class="guided-detail" id="guidedDetail">
      <div id="guidedLockBanner" class="guided-lock-banner" style="display:none"></div>
      <div class="guided-field-label" id="guidedLabel"></div>
      <div class="guided-field-dest" id="guidedDest"></div>
      <div class="guided-field-value" id="guidedValue"></div>
      <div class="guided-field-meta" id="guidedMeta"></div>
      <div class="guided-actions" id="guidedBtns">
        <button class="btn btn-success" onclick="guidedAction('confirm')">&#x2714; Confirm <kbd>Y</kbd></button>
        <button class="btn btn-primary" onclick="guidedStartEdit()">&#x270F; Edit <kbd>E</kbd></button>
        <button class="btn btn-danger" onclick="guidedAction('not_present')">&#x2716; Not Present <kbd>N</kbd></button>
        <button class="btn btn-secondary" onclick="guidedAction('skip')">&#x23ED; Skip <kbd>S</kbd></button>
      </div>
      <div class="guided-note-area" id="guidedNoteArea">
        <div style="display:flex;gap:8px;align-items:center;margin-top:12px">
          <input type="text" class="form-input" id="guidedNoteInput" placeholder="Add a note (optional)..." style="flex:1;font-size:13px">
        </div>
      </div>
      <div class="guided-edit-area" id="guidedEditArea" style="display:none">
        <input type="text" class="guided-edit-input" id="guidedEditInput" placeholder="Enter corrected value...">
        <div class="guided-edit-btns">
          <button class="btn btn-success" onclick="guidedFinishEdit()">Save Correction <kbd>Enter</kbd></button>
          <button class="btn btn-secondary" onclick="guidedCancelEdit()">Cancel <kbd>Esc</kbd></button>
        </div>
      </div>
    </div>
  </div>
</div>

<!-- ═══ CLIENTS SECTION ═══ -->
<div class="section" id="sec-clients">
  <div class="page-header"><h2>Client Manager</h2><p>Upload prior-year returns &amp; workpapers, set extraction instructions, track document completeness</p></div>
  <div class="page-content">
    <div id="clientListView">
      <div style="margin-bottom:16px; display:flex; gap:8px; align-items:center">
        <input type="text" class="form-input" id="clientSearch" placeholder="Search clients..." style="max-width:300px" oninput="filterClients()">
      </div>
      <div class="client-list" id="clientGrid"></div>
    </div>
    <div class="client-detail" id="clientDetailView">
      <div class="client-back" onclick="closeClientDetail()">&#9664; Back to all clients</div>
      <div style="display:flex;align-items:center;gap:12px;margin-bottom:4px">
        <h2 id="clientDetailName" style="font-size:20px;font-weight:700;color:var(--navy);margin:0"></h2>
        <button class="btn btn-secondary btn-sm" onclick="copyClientPath()" title="Copy folder path to clipboard">&#x1F4CB; Copy Path</button>
        <button class="btn btn-secondary btn-sm" onclick="exportClientZip()" title="Download client folder as zip">&#x1F4E6; Export Zip</button>
        <button class="btn btn-ghost btn-sm" onclick="openMergeClientModal()" title="Merge into another client" style="color:var(--purple)">&#x21C4; Merge</button>
        <button class="btn btn-ghost btn-sm" onclick="openDeleteClientModal()" title="Delete this client" style="color:var(--red)">&#x1F5D1; Delete</button>
      </div>
      <div id="clientDetailMeta" style="font-size:13px;color:var(--text-secondary);margin-bottom:16px"></div>
      <div id="clientFilePath" style="font-size:11px;color:var(--text-muted);margin-bottom:8px;display:none"></div>
      <div class="client-tabs">
        <div class="client-tab active" data-tab="documents" onclick="showClientTab('documents')">Documents</div>
        <div class="client-tab" data-tab="context" onclick="showClientTab('context')">Prior-Year Context</div>
        <div class="client-tab" data-tab="instructions" onclick="showClientTab('instructions')">Instructions</div>
        <div class="client-tab" data-tab="completeness" onclick="showClientTab('completeness')">Completeness</div>
      </div>
      <!-- Documents Tab -->
      <div class="client-tab-content active" id="tab-documents">
        <div id="clientDocGroups">
          <div class="empty-state"><p>No documents yet. Upload a PDF from the Upload section to get started.</p></div>
        </div>
      </div>
      <!-- Context Tab -->
      <div class="client-tab-content" id="tab-context">
        <div class="card" style="margin-bottom:16px">
          <div class="card-header"><h3>Upload Context Document</h3></div>
          <div class="card-body">
            <p style="font-size:13px;color:var(--text-secondary);margin-bottom:12px">Upload a prior-year return, workbook, or notes. The system will extract payer information for completeness tracking and variance checking.</p>
            <div class="form-row">
              <div class="form-group">
                <label class="form-label">File</label>
                <input type="file" id="contextFile" accept=".pdf,.xlsx,.xls,.txt,.csv" class="form-input" style="padding:6px">
              </div>
              <div class="form-group">
                <label class="form-label">Year</label>
                <input type="number" id="contextYear" class="form-input" value="2024" min="2000" max="2030">
              </div>
            </div>
            <div class="form-group">
              <label class="form-label">Label (optional)</label>
              <input type="text" id="contextLabel" class="form-input" placeholder="e.g. 2024 Filed Return">
            </div>
            <button class="btn btn-primary" onclick="uploadContext()">Upload Context</button>
          </div>
        </div>
        <div id="contextDocList"></div>
      </div>
      <!-- Instructions Tab -->
      <div class="client-tab-content" id="tab-instructions">
        <div class="card" style="margin-bottom:16px">
          <div class="card-header"><h3>Add Instruction</h3></div>
          <div class="card-body">
            <p style="font-size:13px;color:var(--text-secondary);margin-bottom:12px">Client-specific rules that apply to every extraction. These are injected into the AI prompts automatically.</p>
            <div class="form-group">
              <textarea id="newInstruction" class="form-input" rows="2" placeholder="e.g. All payments from X Corp are commissions, not regular income."></textarea>
            </div>
            <button class="btn btn-primary" onclick="addInstruction()">Add Instruction</button>
          </div>
        </div>
        <div id="instructionsList"></div>
      </div>
      <!-- Completeness Tab -->
      <div class="client-tab-content" id="tab-completeness">
        <div id="completenessReport">
          <div class="empty-state">
            <p>Upload prior-year context to enable completeness tracking.</p>
          </div>
        </div>
      </div>
    </div>
  </div>
</div>

<!-- ═══ BATCH CATEGORIZE ═══ -->
<div class="section" id="sec-batch" style="display:none">
  <div class="page-header"><h2>Batch Categorize</h2><p>Classify transactions across all documents at once</p></div>
  <div class="page-content">
    <div style="margin-bottom:16px; display:flex; gap:8px; align-items:center; flex-wrap:wrap">
      <input type="text" class="form-input" id="batchClientFilter" placeholder="Filter by client..." style="max-width:220px" oninput="loadBatchData()">
      <label style="font-size:13px; display:flex; align-items:center; gap:4px; cursor:pointer">
        <input type="checkbox" id="batchShowAll" onchange="loadBatchData()"> Show categorized
      </label>
      <input type="text" class="form-input" id="batchSearch" placeholder="Search vendors..." style="max-width:220px" oninput="filterBatchVendors()">
    </div>
    <div class="batch-stats" id="batchStats"></div>
    <div id="batchVendorGroups"></div>
  </div>
</div>

<!-- ═══ HISTORY SECTION ═══ -->
<div class="section" id="sec-history">
  <div class="page-header"><h2>Job History</h2><p>All extractions and their status</p></div>
  <div class="page-content">
    <div class="history-filters">
      <input type="text" class="form-input" id="historySearch" placeholder="Search by client or filename..." oninput="filterHistory()">
      <select class="form-input form-select" id="historyStatusFilter" onchange="filterHistory()" style="width:140px">
        <option value="">All statuses</option>
        <option value="complete">Complete</option>
        <option value="running">Running</option>
        <option value="failed">Failed</option>
        <option value="interrupted">Interrupted</option>
      </select>
    </div>
    <div class="card">
      <div class="table-wrap">
        <table class="data-table" id="historyTable">
          <thead><tr><th>Client</th><th>File</th><th>Type</th><th>Year</th><th>Status</th><th>Cost</th><th>Duration</th><th>Review Time</th><th>Date</th><th></th></tr></thead>
          <tbody id="historyBody"></tbody>
        </table>
      </div>
    </div>
  </div>
</div>

<!-- ═══ KEYBOARD HELP ═══ -->
<div class="kbd-overlay" id="kbdOverlay" onclick="if(event.target===this)toggleKbdHelp()">
  <div class="kbd-card">
    <h3>Keyboard Shortcuts</h3>
    <div style="font-size:11px;font-weight:700;color:var(--text-light);text-transform:uppercase;margin-bottom:6px">Grid Review</div>
    <div class="kbd-row"><span>Confirm field</span><kbd>Enter</kbd></div>
    <div class="kbd-row"><span>Flag field</span><kbd>F</kbd></div>
    <div class="kbd-row"><span>Edit value</span><kbd>E</kbd></div>
    <div class="kbd-row"><span>Add note</span><kbd>N</kbd></div>
    <div class="kbd-row"><span>Next field</span><kbd>&#x2193; / Tab</kbd></div>
    <div class="kbd-row"><span>Prev field</span><kbd>&#x2191; / Shift+Tab</kbd></div>
    <div class="kbd-row"><span>Next page</span><kbd>&#x2192;</kbd></div>
    <div class="kbd-row"><span>Prev page</span><kbd>&#x2190;</kbd></div>
    <div style="font-size:11px;font-weight:700;color:var(--text-light);text-transform:uppercase;margin:10px 0 6px;padding-top:8px;border-top:1px solid var(--border)">Audit Check</div>
    <div class="kbd-row"><span>Confirm</span><kbd>Y</kbd></div>
    <div class="kbd-row"><span>Edit</span><kbd>E</kbd></div>
    <div class="kbd-row"><span>Not Present</span><kbd>N</kbd></div>
    <div class="kbd-row"><span>Skip</span><kbd>S</kbd></div>
    <div class="kbd-row"><span>Save correction</span><kbd>Enter</kbd></div>
    <div class="kbd-row"><span>Cancel edit</span><kbd>Esc</kbd></div>
    <div class="kbd-row" style="margin-top:8px"><span>This help</span><kbd>?</kbd></div>
  </div>
</div>

</div><!-- /main -->
</div><!-- /app -->

<!-- ═══════════════════════════════════════════════════════════════════════════ -->
<!-- JAVASCRIPT -->
<!-- ═══════════════════════════════════════════════════════════════════════════ -->
<script>
// ─── Theme ───
function setTheme(t) {
  if (t === 'light') { document.documentElement.removeAttribute('data-theme'); }
  else { document.documentElement.setAttribute('data-theme', t); }
  localStorage.setItem('oathledger-theme', t);
  document.querySelectorAll('.theme-dot').forEach(d => d.classList.toggle('active', d.dataset.theme === t));
  if (typeof loadDashboard === 'function' && document.getElementById('sec-dashboard') && document.getElementById('sec-dashboard').classList.contains('active')) {
    setTimeout(loadDashboard, 50);
  }
}
// Mark active dot on load
(function(){
  var t = localStorage.getItem('oathledger-theme') || 'light';
  document.querySelectorAll('.theme-dot').forEach(d => d.classList.toggle('active', d.dataset.theme === t));
})();

// ─── State ───
let currentJobId = null;
let pollTimer = null;
let startTime = null;
let elapsedTimer = null;
let reviewData = null;
let currentPage = 1;
let totalPages = 1;
let verifications = {};
let totalFieldCount = 0;
let focusedFieldIdx = -1;
let pageFieldKeys = [];
let selectedDocType = 'tax_returns';
let earlyReviewActive = false;
let _loadedImagePage = null;
let _pageWordData = null;    // OCR word bboxes for current page
let _pageWordPage = null;    // which page _pageWordData is for
let _pageImgNatW = 0;        // natural width of page image (for scaling bboxes)
let _pageImgNatH = 0;        // natural height of page image

// ─── Performance Instrumentation ───
const _perfTimers = {};
const _perfLog = {};  // label -> [ms, ms, ...]
function _perf(label) { _perfTimers[label] = performance.now(); }
function _perfEnd(label) {
  const t0 = _perfTimers[label];
  if (t0 === undefined) return;
  const ms = performance.now() - t0;
  delete _perfTimers[label];
  if (!_perfLog[label]) _perfLog[label] = [];
  _perfLog[label].push(ms);
  if (_perfLog[label].length > 200) _perfLog[label].shift();
  console.log('[PERF] ' + label + ': ' + ms.toFixed(1) + 'ms');
  return ms;
}
function _perfSummary() {
  const out = {};
  for (const [label, times] of Object.entries(_perfLog)) {
    if (times.length === 0) continue;
    const sorted = [...times].sort((a, b) => a - b);
    const avg = sorted.reduce((s, v) => s + v, 0) / sorted.length;
    const p95 = sorted[Math.floor(sorted.length * 0.95)] || sorted[sorted.length - 1];
    const max = sorted[sorted.length - 1];
    out[label] = { n: sorted.length, avg: avg.toFixed(1) + 'ms', p95: p95.toFixed(1) + 'ms', max: max.toFixed(1) + 'ms' };
  }
  console.table(out);
  return out;
}

// Field display order by document type (matches extract.py TEMPLATE_SECTIONS)
const FIELD_ORDER = {
  'W-2': ['wages','federal_wh','ss_wages','ss_wh','medicare_wages','medicare_wh','state_wages','state_wh','local_wages','local_wh','nonqualified_plans_12a'],
  '1099-INT': ['interest_income','early_withdrawal_penalty','us_savings_bonds_and_treasury','federal_wh','state_wh'],
  '1099-DIV': ['ordinary_dividends','qualified_dividends','capital_gain_distributions','nondividend_distributions','federal_wh','state_wh','foreign_tax_paid','exempt_interest_dividends'],
  '1099-R': ['gross_distribution','taxable_amount','federal_wh','state_wh','distribution_code','employee_contributions'],
  'K-1': ['ordinary_income','net_rental_income','guaranteed_payments','interest_income','dividends','royalties','net_short_term_capital_gain','net_long_term_capital_gain','net_section_1231_gain','other_income','section_179_deduction','other_deductions','self_employment_earnings'],
  '1099-NEC': ['nonemployee_compensation','federal_wh'],
  '1099-MISC': ['rents','royalties','other_income','federal_wh','fishing_boat_proceeds','medical_payments','nonemployee_compensation'],
  'SSA-1099': ['net_benefits','federal_wh','repaid_benefits'],
};
const FIELD_BOX_LABELS = {
  'W-2': {wages:'Box 1',federal_wh:'Box 2',ss_wages:'Box 3',ss_wh:'Box 4',medicare_wages:'Box 5',medicare_wh:'Box 6',state_wages:'Box 16',state_wh:'Box 17',local_wages:'Box 18',local_wh:'Box 19'},
  '1099-DIV': {ordinary_dividends:'Box 1a',qualified_dividends:'Box 1b',capital_gain_distributions:'Box 2a',nondividend_distributions:'Box 3',federal_wh:'Box 4',foreign_tax_paid:'Box 7',exempt_interest_dividends:'Box 12'},
  '1099-INT': {interest_income:'Box 1',early_withdrawal_penalty:'Box 2',us_savings_bonds_and_treasury:'Box 3',federal_wh:'Box 4'},
  '1099-R': {gross_distribution:'Box 1',taxable_amount:'Box 2a',federal_wh:'Box 4',distribution_code:'Box 7',employee_contributions:'Box 5',state_wh:'Box 12'},
  'K-1': {ordinary_income:'Line 1',net_rental_income:'Line 2',guaranteed_payments:'Line 4c',interest_income:'Line 5',dividends:'Line 6a',royalties:'Line 7',net_short_term_capital_gain:'Line 8',net_long_term_capital_gain:'Line 9a',net_section_1231_gain:'Line 10',other_income:'Line 11',section_179_deduction:'Line 12',other_deductions:'Line 13',self_employment_earnings:'Line 14a'},
  '1099-NEC': {nonemployee_compensation:'Box 1',federal_wh:'Box 4'},
  'SSA-1099': {net_benefits:'Box 5',federal_wh:'Box 6'},
};

// ─── Field Taxonomy — maps field keys to tax-meaningful categories ───
const FIELD_TAXONOMY = {
  // Wages & Compensation
  wages:'wages', nonemployee_compensation:'wages', gross_pay:'wages', net_benefits:'wages',
  net_pay:'wages', total_gross:'wages', total_net_pay:'wages', gross_winnings:'wages',
  gross_amount:'wages', unemployment:'wages',
  // Federal Withholding
  federal_wh:'federal_wh',
  // FICA (Social Security + Medicare)
  ss_wages:'fica', ss_wh:'fica', social_security:'fica',
  medicare_wages:'fica', medicare_wh:'fica', medicare:'fica',
  total_social_security:'fica', total_medicare:'fica',
  total_social_security_tax:'fica', total_medicare_tax:'fica',
  // State & Local Tax
  state_wages:'state_local', state_wh:'state_local', state_income:'state_local',
  local_wages:'state_local', local_wh:'state_local',
  total_state_wh:'state_local', state_local_refund:'state_local',
  state_amount:'state_local',
  // Interest Income
  interest_income:'interest', us_savings_bonds_and_treasury:'interest',
  tax_exempt_interest:'interest', interest_earned:'interest',
  box5_interest:'interest', student_loan_interest:'interest',
  // Dividends
  ordinary_dividends:'dividends', qualified_dividends:'dividends',
  section_199a:'dividends', foreign_tax_paid:'dividends',
  nondividend_distributions:'dividends', exempt_interest_dividends:'dividends',
  box6a_ordinary_dividends:'dividends', box6b_qualified_dividends:'dividends',
  // Capital Gains
  capital_gain_distributions:'capital_gains',
  total_proceeds:'capital_gains', total_basis:'capital_gains',
  total_gain_loss:'capital_gains', wash_sale_loss:'capital_gains',
  short_term_gain_loss:'capital_gains', long_term_gain_loss:'capital_gains',
  gross_proceeds:'capital_gains',
  box8_short_term_capital_gain:'capital_gains',
  box9a_long_term_capital_gain:'capital_gains',
  box9c_unrecaptured_1250:'capital_gains',
  box10_net_1231_gain:'capital_gains',
  // Retirement & Distributions
  gross_distribution:'retirement', taxable_amount:'retirement',
  distribution_code:'retirement', employee_contributions:'retirement',
  ira_contributions:'retirement', rollover_contributions:'retirement',
  roth_conversion:'retirement', rmd_amount:'retirement',
  repaid_benefits:'retirement',
  // Deductions
  mortgage_interest:'deductions', property_tax:'deductions',
  mortgage_insurance_premiums:'deductions', tax_amount:'deductions',
  donation_amount:'deductions', donation_type:'deductions',
  box12_section_179:'deductions', box13_other_deductions:'deductions',
  federal_amount:'deductions',
  // K-1 / Partnership Income
  box1_ordinary_income:'k1_income', box2_rental_real_estate:'k1_income',
  box3_other_rental:'k1_income', box4a_guaranteed_services:'k1_income',
  box7_royalties:'k1_income', box11_other_income:'k1_income',
  box14_self_employment:'k1_income', box17_alt_min_tax:'k1_income',
  box18_tax_exempt_income:'k1_income', box19_distributions:'k1_income',
  box20_other_info:'k1_income',
  net_rental_income:'k1_income', guaranteed_payments:'k1_income',
  // Other Income
  rents:'other_income', royalties:'other_income', other_income:'other_income',
  debt_cancelled:'other_income', gross_farm_income:'other_income',
  net_farm_income:'other_income', gross_income:'other_income',
  total_revenue:'other_income', net_income:'other_income', net_profit:'other_income',
  // Credits
  box15_credits:'credits', scholarships_grants:'credits', payments_received:'credits',
};

const CATEGORY_LABELS = {
  wages:'Wages', federal_wh:'Fed WH', fica:'FICA', state_local:'State/Local',
  interest:'Interest', dividends:'Dividends', capital_gains:'Cap Gains',
  retirement:'Retirement', deductions:'Deductions', k1_income:'K-1',
  other_income:'Other Income', credits:'Credits',
};

const CATEGORY_ORDER = [
  'wages','federal_wh','fica','state_local','interest','dividends',
  'capital_gains','retirement','deductions','k1_income','other_income','credits'
];

function getFieldCategory(fieldName) {
  var bare = fieldName.indexOf(':') >= 0 ? fieldName.split(':').pop() : fieldName;
  return FIELD_TAXONOMY[bare] || null;
}

// ═══ CROSS-DOCUMENT FILTER STATE & LOGIC ═══
var filterState = { activeCategory: null, searchText: '' };
var _filterSearchTimer = null;

function _initFilterBar() {
  if (!reviewData || !reviewData.page_map) return;
  var fp = document.getElementById('fieldsPanel');
  if (!fp) return;
  // Remove old bar if exists
  var old = document.getElementById('filterBar');
  if (old) old.remove();
  // Count fields per category
  var counts = {};
  CATEGORY_ORDER.forEach(function(c) { counts[c] = 0; });
  var totalCat = 0;
  for (var pg in reviewData.page_map) {
    reviewData.page_map[pg].forEach(function(ext) {
      Object.keys(ext.fields || {}).forEach(function(k) {
        if (REVIEW_SKIP_FIELDS.has(k)) return;
        var f = ext.fields[k];
        var v = f.value != null ? f.value : (typeof f !== 'object' ? f : null);
        if (v == null) return;
        var isNum = typeof v === 'number' || (typeof v === 'string' && /^\-?\$?[\d,]+\.?\d*$/.test(v.trim()));
        if (!isNum) return;
        var cat = getFieldCategory(k);
        if (cat && counts[cat] !== undefined) { counts[cat]++; totalCat++; }
        else { if (!counts['other_income']) counts['other_income'] = 0; counts['other_income']++; totalCat++; }
      });
    });
  }
  // Build HTML
  var h = '<div id="filterBar" class="filter-bar">';
  h += '<div class="filter-bar-label">Filter by Category</div>';
  h += '<div class="filter-pills">';
  h += '<div class="filter-pill' + (!filterState.activeCategory ? ' active' : '') + '" data-cat="" onclick="_setFilter(null)">All <span class="pill-count">' + totalCat + '</span></div>';
  CATEGORY_ORDER.forEach(function(cat) {
    if (!counts[cat]) return;
    var active = filterState.activeCategory === cat ? ' active' : '';
    h += '<div class="filter-pill' + active + '" data-cat="' + cat + '" onclick="_setFilter(\'' + cat + '\')">'
       + esc(CATEGORY_LABELS[cat] || cat) + ' <span class="pill-count">' + counts[cat] + '</span></div>';
  });
  h += '</div>';
  h += '<div class="filter-search-wrap"><input class="filter-search" id="filterSearchInput" type="text" placeholder="Search fields across all pages\u2026" value="' + esc(filterState.searchText) + '"></div>';
  if (filterState.activeCategory || filterState.searchText) {
    h += '<div class="filter-active-hint">\u26A1 Showing cross-document view <button class="filter-clear" onclick="_clearFilter()">✕ Clear</button></div>';
  }
  h += '</div>';
  fp.insertAdjacentHTML('afterbegin', h);
  // Debounced search
  var si = document.getElementById('filterSearchInput');
  if (si) {
    si.addEventListener('input', function() {
      clearTimeout(_filterSearchTimer);
      var val = si.value;
      _filterSearchTimer = setTimeout(function() {
        filterState.searchText = val.trim();
        _applyFilter();
      }, 200);
    });
  }
}

function _setFilter(category) {
  filterState.activeCategory = (filterState.activeCategory === category) ? null : category;
  _applyFilter();
}

function _clearFilter() {
  filterState.activeCategory = null;
  filterState.searchText = '';
  var si = document.getElementById('filterSearchInput');
  if (si) si.value = '';
  _applyFilter();
}

function _applyFilter() {
  if (!filterState.activeCategory && !filterState.searchText) {
    // No filter — reload normal page view (loadPage first, then filter bar on top)
    loadPage(currentPage);
    _initFilterBar();
    return;
  }
  // Collect filtered fields and render cross-doc view
  var matched = _collectFilteredFields(filterState.activeCategory, filterState.searchText);
  _renderCrossDocView(matched);
}

function _collectFilteredFields(category, search) {
  var results = [];
  if (!reviewData || !reviewData.page_map) return results;
  var searchLower = (search || '').toLowerCase();
  for (var pg in reviewData.page_map) {
    var pageNum = parseInt(pg);
    reviewData.page_map[pg].forEach(function(ext, extIdx) {
      var docType = ext.document_type || '';
      var entity = ext.entity || ext.payer_or_entity || '';
      var fields = ext.fields || {};
      Object.keys(fields).forEach(function(k) {
        if (REVIEW_SKIP_FIELDS.has(k)) return;
        var f = fields[k];
        var value = (typeof f === 'object' && f !== null) ? f.value : f;
        if (value == null) return;
        var isNum = typeof value === 'number' || (typeof value === 'string' && /^\-?\$?[\d,]+\.?\d*$/.test(String(value).trim()));
        if (!isNum) return;
        var fCat = getFieldCategory(k);
        if (!fCat) fCat = 'other_income';
        // Category filter
        if (category && fCat !== category) return;
        // Search filter
        if (searchLower) {
          var haystack = (k + ' ' + docType + ' ' + entity + ' ' + String(value)).toLowerCase();
          if (haystack.indexOf(searchLower) < 0) return;
        }
        var conf = (typeof f === 'object' && f !== null) ? (f.confidence || '') : '';
        var vk = fieldKey(pageNum, extIdx, k);
        var vstate = verifications[vk] || null;
        var displayVal = value;
        if (vstate && vstate.corrected_value !== undefined) displayVal = vstate.corrected_value;
        results.push({
          page: pageNum, extIdx: extIdx, fieldKey: k, fieldName: k,
          value: displayVal, rawValue: value, category: fCat,
          docType: docType, entity: entity, confidence: conf,
          vk: vk, vstate: vstate
        });
      });
    });
  }
  // Sort by category order, then by page
  var catIdx = {};
  CATEGORY_ORDER.forEach(function(c, i) { catIdx[c] = i; });
  results.sort(function(a, b) {
    var ca = catIdx[a.category] != null ? catIdx[a.category] : 99;
    var cb = catIdx[b.category] != null ? catIdx[b.category] : 99;
    if (ca !== cb) return ca - cb;
    if (a.page !== b.page) return a.page - b.page;
    return a.fieldKey.localeCompare(b.fieldKey);
  });
  return results;
}

// ═══ CROSS-DOCUMENT VIEW RENDERING ═══
function _renderCrossDocView(matched) {
  var fp = document.getElementById('fieldsPanel');
  if (!fp) return;
  // Rebuild filter bar first (to update active states)
  fp.innerHTML = '';
  _initFilterBar();
  var container = document.createElement('div');
  container.className = 'xdoc-container';

  if (matched.length === 0) {
    container.innerHTML = '<div style="padding:30px 20px;text-align:center;color:var(--text-light)">'
      + '<div style="font-size:28px;margin-bottom:8px">\uD83D\uDD0D</div>'
      + '<div style="font-size:13px;font-weight:600">No matching fields found</div>'
      + '<div style="font-size:12px;margin-top:4px">Try a different category or search term</div></div>';
    fp.appendChild(container);
    return;
  }

  // Build rollup summary
  container.innerHTML += _buildRollupSummary(matched);

  // Group fields
  var grouped;
  if (filterState.activeCategory && !filterState.searchText) {
    grouped = _groupByEntity(matched);
  } else {
    grouped = _groupByCategory(matched);
  }

  grouped.forEach(function(group) {
    var gh = '<div class="xdoc-group">';
    gh += '<div class="xdoc-group-header"><span>' + esc(group.label) + '</span><span class="xdoc-count">' + group.fields.length + ' fields</span></div>';
    group.fields.forEach(function(f) {
      gh += _buildCrossDocRow(f);
    });
    gh += '</div>';
    container.innerHTML += gh;
  });
  fp.appendChild(container);
}

function _buildCrossDocRow(f) {
  var displayStr = typeof f.value === 'number' ? f.value.toLocaleString('en-US',{minimumFractionDigits:2,maximumFractionDigits:2}) : String(f.value||'');
  var boxLabel = (FIELD_BOX_LABELS[f.docType] || {})[f.fieldKey];
  var displayName = (boxLabel ? boxLabel + ' \u2014 ' : '') + f.fieldKey.replace(/_/g,' ').replace(/\b\w/g, function(c){return c.toUpperCase();});
  var iconClass = 'xdoc-field-icon vf-none';
  var iconChar = '\u25CB';
  if (f.vstate) {
    if (f.vstate.status === 'confirmed' || f.vstate.status === 'corrected') { iconClass = 'xdoc-field-icon vf-ok'; iconChar = '\u2713'; }
    else if (f.vstate.status === 'flagged') { iconClass = 'xdoc-field-icon vf-flag'; iconChar = '\u26A0'; }
  }
  var source = esc(f.docType) + ' \u00B7 ' + esc(f.entity || 'Unknown') + ' \u00B7 p.' + f.page;
  return '<div class="xdoc-field-row" data-page="' + f.page + '" data-ext="' + f.extIdx + '" data-field="' + esc(f.fieldKey) + '" onclick="_xdocNavigate(this)">'
    + '<div class="xdoc-field-left"><div class="xdoc-field-name">' + esc(displayName) + '</div><div class="xdoc-field-source">' + source + '</div></div>'
    + '<div class="xdoc-field-right"><span class="' + iconClass + '">' + iconChar + '</span><span class="xdoc-field-value">' + esc(displayStr) + '</span>'
    + '<span class="xdoc-nav-hint">\u2192 go</span></div></div>';
}

function _groupByCategory(fields) {
  var groups = {};
  var order = [];
  fields.forEach(function(f) {
    var label = CATEGORY_LABELS[f.category] || f.category || 'Other';
    if (!groups[label]) { groups[label] = []; order.push(label); }
    groups[label].push(f);
  });
  return order.map(function(l) { return { label: l, fields: groups[l] }; });
}

function _groupByEntity(fields) {
  var groups = {};
  var order = [];
  fields.forEach(function(f) {
    var label = (f.entity || 'Unknown') + ' (' + f.docType + ')';
    if (!groups[label]) { groups[label] = []; order.push(label); }
    groups[label].push(f);
  });
  return order.map(function(l) { return { label: l, fields: groups[l] }; });
}

function _xdocNavigate(el) {
  var page = parseInt(el.getAttribute('data-page'));
  var field = el.getAttribute('data-field');
  var extIdx = parseInt(el.getAttribute('data-ext'));
  // Clear filter, go to page, highlight field
  filterState.activeCategory = null;
  filterState.searchText = '';
  loadPage(page);
  // After DOM updates, find and focus the field
  setTimeout(function() {
    var vk = fieldKey(page, extIdx, field);
    var idx = pageFieldKeys.indexOf(vk);
    if (idx >= 0) {
      moveFocus(idx);
      var row = document.querySelector('.field-row[data-key="' + vk.replace(/"/g, '\\\\"') + '"]');
      if (row) {
        row.scrollIntoView({ behavior: 'smooth', block: 'center' });
        row.style.transition = 'background 0.3s';
        row.style.background = 'rgba(52,152,219,0.15)';
        setTimeout(function() { row.style.background = ''; }, 1500);
      }
    }
    _initFilterBar();
  }, 100);
}

// ═══ ROLLUP SUMMARY PANEL ═══
function _buildRollupSummary(matched) {
  var numericFields = matched.filter(function(f) {
    var v = f.value;
    if (typeof v === 'number') return true;
    if (typeof v === 'string') {
      var cleaned = v.replace(/[$,\s]/g, '');
      return /^\-?\d+\.?\d*$/.test(cleaned);
    }
    return false;
  });
  if (numericFields.length === 0) return '<div class="rollup-panel"><div class="rollup-empty">No numeric values to summarize</div></div>';

  var h = '<div class="rollup-panel">';
  // Compute totals
  var grandTotal = 0;
  var subTotals = {};
  var subOrder = [];
  numericFields.forEach(function(f) {
    var v = typeof f.value === 'number' ? f.value : parseFloat(String(f.value).replace(/[$,]/g, ''));
    if (isNaN(v)) return;
    grandTotal += v;
    var groupKey;
    if (filterState.activeCategory && !filterState.searchText) {
      groupKey = (f.entity || 'Unknown') + ' (' + f.docType + ')';
    } else {
      groupKey = CATEGORY_LABELS[f.category] || f.category || 'Other';
    }
    if (!subTotals[groupKey]) { subTotals[groupKey] = 0; subOrder.push(groupKey); }
    subTotals[groupKey] += v;
  });

  var fmt = function(n) { return (n < 0 ? '-' : '') + '$' + Math.abs(n).toLocaleString('en-US',{minimumFractionDigits:2,maximumFractionDigits:2}); };
  h += '<div class="rollup-header" onclick="_toggleRollup(this)"><span>\u03A3 Total: ' + fmt(grandTotal) + ' (' + numericFields.length + ' fields)</span><span class="rollup-chevron">\u25BC</span></div>';
  h += '<div class="rollup-body expanded">';
  subOrder.forEach(function(key) {
    h += '<div class="rollup-row"><span class="rollup-label">' + esc(key) + '</span><span class="rollup-value">' + fmt(subTotals[key]) + '</span></div>';
  });
  if (subOrder.length > 1) {
    h += '<div class="rollup-row rollup-total"><span class="rollup-label">Grand Total</span><span class="rollup-value">' + fmt(grandTotal) + '</span></div>';
  }
  h += '</div></div>';
  return h;
}

function _toggleRollup(el) {
  var body = el.parentElement.querySelector('.rollup-body');
  var chev = el.querySelector('.rollup-chevron');
  if (!body) return;
  if (body.classList.contains('expanded')) {
    body.classList.remove('expanded');
    body.classList.add('collapsed');
    if (chev) chev.classList.add('collapsed');
  } else {
    body.classList.remove('collapsed');
    body.classList.add('expanded');
    if (chev) chev.classList.remove('collapsed');
  }
}

let selectedOutputFormat = 'tax_review';
let vendorMap = {};
let chartOfAccounts = {};
let currentClientName = '';
let batchData = null;
let allJobs = [];

const DOC_TYPES = [
  {id:'tax_returns', label:'Tax Returns', icon:'&#x1F4CB;'},
  {id:'bank_statements', label:'Bank Statements', icon:'&#x1F3E6;'},
  {id:'bookkeeping', label:'Bookkeeping', icon:'&#x1F4D2;'},
  {id:'trust_documents', label:'Trust Documents', icon:'&#x1F512;'},
  {id:'payroll', label:'Payroll', icon:'&#x1F4B5;'},
  {id:'other', label:'Other', icon:'&#x1F4C4;'},
];
const OUTPUT_FORMATS = [
  {id:'tax_review', label:'Tax Review'},
  {id:'journal_entries', label:'Journal Entries'},
  {id:'account_balances', label:'Account Balances'},
  {id:'trial_balance', label:'Trial Balance'},
  {id:'transaction_register', label:'Transaction Register'},
];

// ─── Init ───
(function init() {
  buildPills();
  loadJobs();
  loadClientSuggestions();
  _checkApiKey();
})();

// ─── API Key Check ───
function _checkApiKey() {
  fetch('/api/config/api-key').then(r=>r.json()).then(d=>{
    var banner = document.getElementById('apiKeyBanner');
    if (!banner) return;
    if (d.configured) {
      banner.style.display = 'none';
    } else {
      banner.style.display = 'block';
    }
  }).catch(()=>{});
}
function saveApiKey() {
  var inp = document.getElementById('apiKeyInput');
  var key = (inp ? inp.value : '').trim();
  if (!key) { showToast('Enter your API key', 'error'); return; }
  if (!key.startsWith('sk-ant-')) { showToast('Key must start with sk-ant-', 'error'); return; }
  fetch('/api/config/api-key', {method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({key:key})})
    .then(r=>r.json()).then(d=>{
      if (d.ok) {
        showToast('API key saved! (' + d.hint + ')', 'success');
        document.getElementById('apiKeyBanner').style.display = 'none';
      } else {
        showToast(d.error || 'Failed to save key', 'error');
      }
    }).catch(()=>showToast('Network error','error'));
}

function buildPills() {
  let dh = '';
  DOC_TYPES.forEach(dt => {
    dh += '<div class="pill' + (dt.id === selectedDocType ? ' active' : '') + '" onclick="selectDocType(\'' + dt.id + '\')">' + dt.icon + ' ' + dt.label + '</div>';
  });
  var dtp = document.getElementById('docTypePills');
  if (dtp) dtp.innerHTML = dh;
  let oh = '';
  OUTPUT_FORMATS.forEach(of_ => {
    oh += '<div class="pill' + (of_.id === selectedOutputFormat ? ' active' : '') + '" onclick="selectOutputFormat(\'' + of_.id + '\')">' + of_.label + '</div>';
  });
  var ofp = document.getElementById('outputFormatPills');
  if (ofp) ofp.innerHTML = oh;
}

function selectDocType(id) { selectedDocType = id; buildPills(); }
function selectOutputFormat(id) { selectedOutputFormat = id; buildPills(); }

// ─── Toast ───
function showToast(msg, type) {
  type = type || 'info';
  const el = document.createElement('div');
  el.className = 'toast toast-' + type;
  el.textContent = msg;
  document.getElementById('toast-container').appendChild(el);
  setTimeout(() => { el.style.opacity = '0'; setTimeout(() => el.remove(), 300); }, 3000);
}

// ─── Navigation ───
function showSection(id) {
  document.querySelectorAll('.section').forEach(s => s.classList.remove('active'));
  document.querySelectorAll('.nav-item').forEach(n => n.classList.remove('active'));
  document.getElementById('sec-' + id).classList.add('active');
  const nav = document.querySelector('[data-section="' + id + '"]');
  if (nav) nav.classList.add('active');
  if (id === 'history') loadJobs();
  if (id === 'clients') loadClients();
  if (id === 'batch') loadBatchData();
  if (id === 'inbox') loadInbox();
}

function esc(s) { const d = document.createElement('div'); d.textContent = s; return d.innerHTML; }
function getReviewer() { return (document.getElementById('reviewerInitials').value || '').trim(); }

// ─── B1: Inbox ───
var _currentUserRole = 'admin';
var _currentUserName = '';

function loadInbox() {
  fetch('/api/inbox').then(r => r.json()).then(data => {
    var items = data.inbox || [];
    var el = document.getElementById('inboxContent');
    var badge = document.getElementById('inboxCount');

    if (badge) {
      if (items.length > 0) {
        badge.textContent = items.length;
        badge.style.display = '';
      } else {
        badge.style.display = 'none';
      }
    }

    if (items.length === 0) {
      el.innerHTML = '<div class="card"><div class="card-body" style="text-align:center;padding:48px;color:var(--text-muted)">' +
        '<div style="font-size:48px;margin-bottom:12px">&#x2705;</div>' +
        '<h3 style="margin:0 0 8px;color:var(--navy)">Inbox Empty</h3>' +
        '<p style="margin:0">No documents waiting for your review.</p>' +
        '</div></div>';
      return;
    }

    var stageColors = {
      'preparer_review': '#3B82F6',
      'reviewer_review': '#F59E0B',
      'partner_review': '#8B5CF6',
      'draft': '#94A3B8',
      'final': '#10B981'
    };

    var html = '<div class="card"><div class="card-body" style="padding:0"><table class="data-table" style="width:100%">' +
      '<thead><tr><th>Client</th><th>Document</th><th>Stage</th><th>Updated</th><th></th></tr></thead><tbody>';

    items.forEach(function(item) {
      var color = stageColors[item.review_stage] || '#94A3B8';
      var updated = item.stage_updated ? new Date(item.stage_updated).toLocaleDateString() : '';
      html += '<tr>' +
        '<td style="padding:10px 12px;font-weight:600">' + esc(item.client_name || 'Unknown') + '</td>' +
        '<td style="padding:10px 12px">' + esc(item.filename || '') + '</td>' +
        '<td style="padding:10px 12px"><span style="background:' + color + ';color:#fff;padding:3px 10px;border-radius:12px;font-size:12px;font-weight:600">' + esc(item.stage_display) + '</span></td>' +
        '<td style="padding:10px 12px;color:var(--text-muted);font-size:13px">' + updated + '</td>' +
        '<td style="padding:10px 12px"><button class="btn btn-sm btn-primary" onclick=\'openReviewFromInbox(' + JSON.stringify(item.job_id) + ',' + JSON.stringify(item.client_name || '') + ')\'>Review</button></td>' +
        '</tr>';
    });

    html += '</tbody></table></div></div>';
    el.innerHTML = html;
  }).catch(function(e) {
    document.getElementById('inboxContent').innerHTML = '<p style="color:var(--error)">Failed to load inbox: ' + e + '</p>';
  });
}

function openReviewFromInbox(jobId, clientName) {
  var job = { id: jobId, client_name: clientName || '' };
  _loadReviewData(job, function() {
    openGuidedReview();
  });
}

function refreshInboxBadge() {
  fetch('/api/inbox').then(r => r.json()).then(data => {
    var badge = document.getElementById('inboxCount');
    var count = (data.inbox || []).length;
    if (badge) {
      badge.textContent = count;
      badge.style.display = count > 0 ? '' : 'none';
    }
  }).catch(function() {});
}

// Load user info and inbox badge on startup
fetch('/api/me').then(r => r.json()).then(data => {
  _currentUserRole = data.role || 'admin';
  _currentUserName = data.display_name || '';
  refreshInboxBadge();
}).catch(function() {});

// ─── Upload ───
const dropZone = document.getElementById('dropZone');
dropZone.addEventListener('dragover', e => { e.preventDefault(); dropZone.classList.add('dragover'); });
dropZone.addEventListener('dragleave', () => dropZone.classList.remove('dragover'));
dropZone.addEventListener('drop', e => { e.preventDefault(); dropZone.classList.remove('dragover'); if (e.dataTransfer.files.length) handleFileObj(e.dataTransfer.files[0]); });

let uploadedFile = null;
function handleFile(input) { if (input.files.length) handleFileObj(input.files[0]); }
function handleFileObj(f) {
  if (!f.name.toLowerCase().endsWith('.pdf')) { showToast('Please upload a PDF file', 'error'); return; }
  uploadedFile = f;
  document.getElementById('fileNameText').textContent = f.name;
  document.getElementById('uploadForm').classList.add('visible');
  dropZone.style.display = 'none';
  // Auto-match client from filename
  if (!document.getElementById('clientName').value) {
    const stem = f.name.replace(/\.pdf$/i, '').replace(/[_-]/g, ' ').toLowerCase();
    const sel = document.getElementById('clientName');
    for (let i = 0; i < sel.options.length; i++) {
      if (sel.options[i].value && stem.includes(sel.options[i].value.toLowerCase())) {
        sel.value = sel.options[i].value;
        break;
      }
    }
  }
}
function resetUpload() {
  uploadedFile = null;
  document.getElementById('fileInput').value = '';
  document.getElementById('uploadForm').classList.remove('visible');
  dropZone.style.display = '';
}

function startExtraction() {
  if (!uploadedFile) return;
  const cn = document.getElementById('clientName').value;
  if (!cn) { showToast('Please select a client', 'error'); return; }
  const fd = new FormData();
  fd.append('pdf', uploadedFile);
  fd.append('year', document.getElementById('taxYear').value);
  fd.append('client_name', cn);
  fd.append('doc_type', selectedDocType);
  fd.append('output_format', 'tax_review');  // Default; user picks format at Finish Review
  fd.append('user_notes', document.getElementById('userNotes').value);
  fd.append('ai_instructions', document.getElementById('aiInstructions').value);
  fd.append('skip_verify', document.getElementById('skipVerify').checked ? 'true' : 'false');
  fd.append('disable_pii', document.getElementById('disablePii').checked ? 'true' : 'false');
  fd.append('use_ocr_first', document.getElementById('useOcrFirst').checked ? 'true' : 'false');

  document.getElementById('startBtn').disabled = true;
  fetch('/api/upload', { method: 'POST', body: fd })
    .then(r => r.json())
    .then(data => {
      if (data.error) { showToast(data.error, 'error'); document.getElementById('startBtn').disabled = false; return; }
      currentJobId = data.job_id;
      document.getElementById('processingFile').textContent = uploadedFile.name;
      showSection('processing');
      document.getElementById('startBtn').disabled = false;
      resetUpload();
      startPolling();
    })
    .catch(e => { showToast('Upload failed: ' + e, 'error'); document.getElementById('startBtn').disabled = false; });
}

// ─── Polling ───
function startPolling() {
  if (pollTimer) clearInterval(pollTimer);
  pollTimer = setInterval(pollStatus, 800);
  pollStatus();
  document.getElementById('procCancelBtn').style.display = '';
}

function cancelJob() {
  if (!currentJobId) return;
  if (!confirm('Cancel this extraction?')) return;
  fetch('/api/cancel/' + currentJobId, { method: 'POST' })
    .then(r => r.json())
    .then(data => {
      if (data.error) { showToast(data.error, 'error'); return; }
      clearInterval(pollTimer); pollTimer = null;
      document.getElementById('procCancelBtn').style.display = 'none';
      showToast('Extraction cancelled', 'info');
      showSection('upload');
    })
    .catch(e => { showToast('Cancel failed: ' + e, 'error'); });
}

function pollStatus() {
  if (!currentJobId) return;
  fetch('/api/status/' + currentJobId).then(r => r.json()).then(data => {
    document.getElementById('procStage').textContent = (data.stage || 'starting').replace(/_/g, ' ');
    const pct = data.progress || 0;
    document.getElementById('procPct').textContent = pct + '%';
    document.getElementById('procBar').style.width = pct + '%';

    // Elapsed time
    const elapsedEl = document.getElementById('procElapsed');
    if (data.start_time && data.status === 'running') {
      const elapsed = Math.floor((Date.now() - new Date(data.start_time).getTime()) / 1000);
      const mins = Math.floor(elapsed / 60); const secs = elapsed % 60;
      elapsedEl.textContent = mins + 'm ' + secs + 's elapsed';
      elapsedEl.style.display = '';
    } else if (data.status === 'complete' && data.start_time && data.end_time) {
      const elapsed = Math.floor((new Date(data.end_time).getTime() - new Date(data.start_time).getTime()) / 1000);
      const mins = Math.floor(elapsed / 60); const secs = elapsed % 60;
      elapsedEl.textContent = 'Completed in ' + mins + 'm ' + secs + 's';
      elapsedEl.style.display = '';
    } else {
      elapsedEl.style.display = 'none';
    }

    // Console output
    const log = data.recent_log || data.log || [];
    const console_el = document.getElementById('procConsole');
    console_el.innerHTML = log.slice(-50).map(l => '<div' + (/phase|complete|error|warning/i.test(l) ? ' class="line-highlight"' : '') + '>' + esc(l) + '</div>').join('');
    console_el.scrollTop = console_el.scrollHeight;

    // T1.5: Show early review button when partial results are ready
    if (data.partial_results_ready && data.status === 'running') {
      const btn = document.getElementById('procReviewEarlyBtn');
      if (btn) {
        btn.style.display = '';
        btn.textContent = 'Review ' + (data.fields_streamed || 0) + ' fields now';
      }
    }

    if (data.status === 'complete') {
      clearInterval(pollTimer); pollTimer = null;
      document.getElementById('procCancelBtn').style.display = 'none';
      document.getElementById('procReviewEarlyBtn').style.display = 'none';
      const costStr = data.cost_usd ? ' ($' + data.cost_usd.toFixed(4) + ')' : '';
      showToast('Extraction complete!' + costStr, 'success');
      document.getElementById('navReview').style.display = '';
      document.getElementById('navGuidedReview').style.display = '';
      if (earlyReviewActive) { earlyReviewActive = false; return; }
      openReview(data);
    } else if (data.status === 'failed' || data.status === 'interrupted') {
      clearInterval(pollTimer); pollTimer = null;
      document.getElementById('procCancelBtn').style.display = 'none';
      document.getElementById('procReviewEarlyBtn').style.display = 'none';
      earlyReviewActive = false;
      showToast(data.status === 'interrupted' ? 'Extraction cancelled' : 'Extraction failed', 'error');
    }
  }).catch(() => {});
}

// ─── Review ───
function openReview(job) {
  currentJobId = job.id || job.job_id || currentJobId;
  document.getElementById('navReview').style.display = '';
  document.getElementById('navGuidedReview').style.display = '';
  // Load review data first, then open grid (list) review as default
  _loadReviewData(job, function() { openGridReview(); });
  return;
}
// Load review data without switching section (used by both guided and grid)
function _loadReviewData(job, callback) {
  currentJobId = job.id || job.job_id || currentJobId;
  _loadedImagePage = null;  // reset image cache on job change
  _pageWordData = null; _pageWordPage = null;  // reset highlight data

  Promise.all([
    fetch('/api/results/' + currentJobId).then(r => r.json()),
    fetch('/api/verify/' + currentJobId).then(r => r.json()),
    fetch('/api/vendor-categories').then(r => r.json()),
  ]).then(([data, vdata, vcdata]) => {
    reviewData = data;
    verifications = (vdata && vdata.fields) ? vdata.fields : {};
    vendorMap = (vcdata && vcdata.vendors) ? vcdata.vendors : {};
    chartOfAccounts = (vcdata && vcdata.chart_of_accounts) ? vcdata.chart_of_accounts : {};
    if (vdata && vdata.reviewer && !getReviewer()) {
      document.getElementById('reviewerInitials').value = vdata.reviewer;
    }
    // Show client instructions banner
    const clientName = (job.client_name || '');
    if (clientName) {
      fetch('/api/instructions/' + encodeURIComponent(clientName)).then(r => r.json()).then(idata => {
        const rules = (idata.rules || []).filter(r => r.text);
        const banner = document.getElementById('reviewInstructionsBanner');
        if (rules.length) {
          banner.innerHTML = '<strong style="color:#B7791F">&#x26A0; Client Instructions:</strong> ' + rules.map(r => esc(r.text)).join(' &bull; ');
          banner.style.display = '';
        } else {
          banner.style.display = 'none';
        }
      }).catch(() => {});
    }
    // B7-UX: Show doc type mismatch banner if user selection differs from detected types
    var dtBanner = document.getElementById('reviewDocTypeBanner');
    if (data.user_doc_type && data.user_doc_type !== 'other') {
      var exts = data.extractions || [];
      var detectedTypes = [...new Set(exts.map(function(e) { return e.document_type; }).filter(Boolean))];
      if (detectedTypes.length > 0) {
        var labels = { tax_returns: 'Tax Returns', bank_statements: 'Bank Statements', bookkeeping: 'Bookkeeping', trust_documents: 'Trust Documents', payroll: 'Payroll', other: 'Other' };
        dtBanner.innerHTML = '&#128270; <strong>Selected:</strong> ' + esc(labels[data.user_doc_type] || data.user_doc_type) + ' &nbsp;&bull;&nbsp; <strong>Detected:</strong> ' + detectedTypes.map(esc).join(', ') + ' <span style="margin-left:8px;font-size:11px;color:#B08030">(Final classification is automatic)</span>';
        dtBanner.style.display = '';
      } else {
        dtBanner.style.display = 'none';
      }
    } else if (dtBanner) {
      dtBanner.style.display = 'none';
    }
    countTotalFields();
    updateVerifyBar();
    loadPage(1);
    _initFilterBar();
    renderLiteFindings();
    if (callback) callback();
  }).catch(() => { reviewData = null; verifications = {}; loadPage(1); _initFilterBar(); renderLiteFindings(); if (callback) callback(); });
}

// Open grid view explicitly (from "List View" button in guided review)
function openGridReview() {
  if (!reviewData && currentJobId) {
    _loadReviewData({ id: currentJobId }, function() { showSection('review'); });
  } else {
    showSection('review');
  }
}

// ─── Early Review (T1.5) ───
function openEarlyReview() {
  earlyReviewActive = true;
  fetch('/api/results/' + currentJobId).then(r => r.json()).then(data => {
    if (data.error) { showToast('Partial results not ready yet', 'info'); return; }
    reviewData = data;
    // Also load verifications + vendor categories
    Promise.all([
      fetch('/api/verify/' + currentJobId).then(r => r.json()),
      fetch('/api/vendor-categories').then(r => r.json()),
    ]).then(([vdata, vcdata]) => {
      verifications = (vdata && vdata.fields) ? vdata.fields : {};
      vendorMap = (vcdata && vcdata.vendors) ? vcdata.vendors : {};
      chartOfAccounts = (vcdata && vcdata.chart_of_accounts) ? vcdata.chart_of_accounts : {};
    }).catch(() => {});
    showSection('review');
    const banner = document.getElementById('reviewPartialBanner');
    banner.style.display = '';
    banner.textContent = 'Partial review \u2014 extraction in progress (' +
      (data.batch_num || '?') + '/' + (data.total_batches || '?') + ' batches)';
    countTotalFields();
    updateVerifyBar();
    loadPage(1);
    _initFilterBar();
    renderLiteFindings();
    // Switch to slower polling while reviewing
    clearInterval(pollTimer);
    pollTimer = setInterval(pollPartialStatus, 2000);
  }).catch(() => { showToast('Could not load partial results', 'error'); });
}

function pollPartialStatus() {
  if (!currentJobId) return;
  fetch('/api/status/' + currentJobId).then(r => r.json()).then(data => {
    if (data.status === 'complete') {
      // Extraction finished — transition to full review
      earlyReviewActive = false;
      clearInterval(pollTimer); pollTimer = null;
      document.getElementById('reviewPartialBanner').style.display = 'none';
      const costStr = data.cost_usd ? ' ($' + data.cost_usd.toFixed(4) + ')' : '';
      showToast('Extraction complete!' + costStr, 'success');
      document.getElementById('navReview').style.display = '';
      document.getElementById('navGuidedReview').style.display = '';
      // Reload full results
      const savedPage = currentPage;
      fetch('/api/results/' + currentJobId).then(r => r.json()).then(rdata => {
        reviewData = rdata;
        countTotalFields();
        updateVerifyBar();
        loadPage(savedPage);
        renderLiteFindings();
      }).catch(() => {});
      return;
    }
    if (data.status === 'failed' || data.status === 'interrupted') {
      earlyReviewActive = false;
      clearInterval(pollTimer); pollTimer = null;
      document.getElementById('reviewPartialBanner').style.display = 'none';
      showToast(data.status === 'interrupted' ? 'Extraction cancelled' : 'Extraction failed', 'error');
      showSection('processing');
      return;
    }
    // Still running — refresh partial data if new batches arrived
    if (data.batches_complete && reviewData &&
        data.batches_complete > (reviewData.batch_num || 0)) {
      const savedPage = currentPage;
      fetch('/api/results/' + currentJobId).then(r => r.json()).then(rdata => {
        if (!rdata.error) {
          reviewData = rdata;
          const banner = document.getElementById('reviewPartialBanner');
          banner.textContent = 'Partial review \u2014 extraction in progress (' +
            (rdata.batch_num || '?') + '/' + (rdata.total_batches || '?') + ' batches)';
          countTotalFields();
          updateVerifyBar();
          loadPage(savedPage);
          renderLiteFindings();
        }
      }).catch(() => {});
    }
  }).catch(() => {});
}

// Fields to hide from review panel — metadata, PII, and non-tax-relevant details
const REVIEW_SKIP_FIELDS = new Set([
  // PII / identifiers
  'payer_ein','recipient_ssn_last4','tax_year','entity_type','partner_type','state_id',
  'account_number_last4','account_number','employee_ssn','card_number','card_number_last_four',
  'card_number_last_4','card_type','card_holder_name','auth_code','auth_number','authorization_code',
  'cc_auth','response_message_code','confirmation_number','receipt_number','register_number',
  'check_number','guarantor_id','guarantor_name','guarantor_number','patient_id','patient_name',
  'patient_address','patient_phone',
  // Property assessment metadata (not dollar amounts for tax return)
  'property_id','map_code','district','tax_district','homestead','exemptions','acreage','acres',
  'building_value','land_value','total_fair_market_value','appraised_value_100_percent',
  'assessed_value_40_percent','current_year_assessed_value','homestead_exemption_value',
  'net_taxable_value','county_millage_rate','school_millage_rate','previous_year_fair_market_value',
  'current_year_fair_market_value','assessment_notice_date','appeal_deadline','current_year_other_value',
  'other_exemption_value','covenant_year','taxpayer_returned_value','taxing_authority_county',
  'taxing_authority_school','property_owner_name','property_description',
  // 1098 metadata
  'outstanding_mortgage_principal','mortgage_origination_date','mortgage_acquisition_date',
  'number_of_properties','box_7_checkbox','points_paid_on_purchase',
  // Medical/pharmacy metadata
  'pharmacy_name','pharmacy_phone','rx_number','ndc_number','dea_number','prescriber',
  'prescriber_name','prescriber_dea','prescribing_doctor','medication','medication_name',
  'syrup_strength','manufacturer','quantity','date_filled','prescription_date','prescription_number',
  'copay_amount','copay','insurance','insurance_adjustments','insurance_paid','insurance_adjustment',
  'claim_adjustments','previously_paid','practice_address','payment_website','payment_code',
  'payment_location','payment_source','payment_clerk','payment_method','payment_type',
  'payment_status','reference','appointment_date','procedure_description',
  'previous_statement_amount','new_charges','payments_and_adjustments',
  'patient_payments_since_last_statement','phone_numbers','service_date',
  // Receipt/payment metadata
  'receipt_date','transaction_date','billing_date','print_time','paid_by','mh_decal',
  // Course/tuition metadata
  'course_code','course_description','credit_hours','semester_term','book_type',
  // Vehicle tag metadata
  'tag_title','dmv_amount','service_fee_amount',
  // Other metadata
  'bill_number','statement_date','foreign_country','name','date','due_date',
]);

function countTotalFields() {
  totalFieldCount = 0;
  const skipFields = REVIEW_SKIP_FIELDS;
  if (!reviewData || !reviewData.page_map) return;
  for (const pg in reviewData.page_map) {
    reviewData.page_map[pg].forEach((ext, extIdx) => {
      Object.keys(ext.fields || {}).forEach(k => {
        if (skipFields.has(k)) return;
        if (/^txn_\d+_(date|desc|type)$/.test(k)) return;
        const f = ext.fields[k];
        const v = f.value;
        if (typeof v === 'number' || (typeof v === 'string' && /^\-?\$?[\d,]+\.?\d*$/.test(v.trim()))) {
          totalFieldCount++;
        }
      });
    });
  }
}

function updateVerifyBar() {
  const reviewed = Object.keys(verifications).length;
  const pct = totalFieldCount > 0 ? Math.min(100, Math.round(reviewed / totalFieldCount * 100)) : 0;
  var bar = document.getElementById('verifyBar');
  if (bar) bar.style.width = pct + '%';
  var stats = document.getElementById('verifyStats');
  if (stats) stats.innerHTML = '<span>' + reviewed + '</span> of <span>' + totalFieldCount + '</span> fields verified (' + pct + '%)';
  // When all fields are reviewed, show completion panel
  var sec = document.getElementById('sec-review');
  if (totalFieldCount > 0 && reviewed >= totalFieldCount && sec && sec.style.display !== 'none') {
    _showGridComplete(reviewed, totalFieldCount);
  }
}
function _showGridComplete(reviewed, total) {
  // Show completion overlay in the fields panel
  var panel = document.getElementById('fieldsPanel');
  if (!panel || panel.querySelector('.grid-complete-banner')) return; // already showing
  fetch('/api/jobs/' + currentJobId + '/stage').then(function(r) { return r.json(); }).then(function(stageInfo) {
    _renderGridComplete(panel, reviewed, total, stageInfo);
  }).catch(function() {
    _renderGridComplete(panel, reviewed, total, { stage: 'preparer_review', can_act: true, can_submit: true, display: 'Preparer Review' });
  });
}
function _renderGridComplete(panel, reviewed, total, stageInfo) {
  var stage = stageInfo.stage || 'preparer_review';
  var stageDisplay = stageInfo.display || stage;
  var stageColors = { 'preparer_review': '#3B82F6', 'reviewer_review': '#F59E0B', 'partner_review': '#8B5CF6', 'final': '#10B981', 'draft': '#94A3B8' };
  var stageColor = stageColors[stage] || '#94A3B8';
  var banner = document.createElement('div');
  banner.className = 'grid-complete-banner';
  banner.innerHTML =
    '<div style="text-align:center;padding:24px 16px;background:var(--bg-card);border:2px solid var(--green);border-radius:var(--radius-lg);margin:12px 0">' +
    '<div style="font-size:40px;margin-bottom:8px">&#x2714;</div>' +
    '<h3 style="margin:0 0 4px;color:var(--green)">All Fields Reviewed</h3>' +
    '<p style="color:var(--text-muted);margin:0 0 12px;font-size:13px">' + reviewed + ' of ' + total + ' fields verified</p>' +
    '<div style="display:inline-block;background:' + stageColor + ';color:#fff;padding:4px 12px;border-radius:12px;font-size:12px;font-weight:600;margin-bottom:16px">' + esc(stageDisplay) + '</div>' +
    '<div style="display:flex;flex-wrap:wrap;gap:8px;justify-content:center">' +
    '<button class="btn btn-accent" onclick="openGuidedReview()" style="font-size:13px;padding:8px 16px">&#x1F50D; Audit Check</button>' +
    '<button class="btn btn-primary" onclick="submitToNextStage()" style="font-size:13px;padding:8px 16px">&#x27A1; Submit</button>' +
    '<button class="btn btn-secondary" onclick="finishReviewGenerate()" style="font-size:13px;padding:8px 16px">&#x1F4C4; Generate Report</button>' +
    '<button class="btn btn-secondary" onclick="window.open(\'/api/download/\'+currentJobId,\'_blank\')" style="font-size:13px;padding:8px 16px">&#x2B73; Download Excel</button>' +
    '</div></div>';
  panel.insertBefore(banner, panel.firstChild);
  banner.scrollIntoView({ behavior: 'smooth', block: 'start' });
}

function fieldKey(page, extIdx, fieldName) { return page + ':' + extIdx + ':' + fieldName; }

// ─── Save Verification ───
function saveVerification(key, status, correctedValue, note, category, vendorDesc, opts) {
  const decision = { status: status };
  if (correctedValue !== undefined && correctedValue !== null) decision.corrected_value = correctedValue;
  if (note) decision.note = note;
  if (category) decision.category = category;
  if (vendorDesc) decision.vendor_desc = vendorDesc;
  if (opts && opts.relabeled_name) decision.relabeled_name = opts.relabeled_name;
  const existing = verifications[key];
  if (existing && existing.category && !category) {
    decision.category = existing.category;
    if (existing.vendor_desc) decision.vendor_desc = existing.vendor_desc;
  }
  if (existing && existing.relabeled_name && !(opts && opts.relabeled_name)) {
    decision.relabeled_name = existing.relabeled_name;
  }
  verifications[key] = decision;
  fetch('/api/verify/' + currentJobId, {
    method: 'POST', headers: { 'Content-Type': 'application/json' },
    body: JSON.stringify({ fields: { [key]: decision }, reviewer: getReviewer() })
  }).catch(() => {});
  updateVerifyBar();
}

function confirmField(key) {
  _perf('confirmField');
  // If an edit input is active for this field, finish the edit first
  const activeInput = document.querySelector('.field-edit-input[data-key="' + key.replace(/"/g, '\\"') + '"]');
  if (activeInput) {
    const nv = activeInput.value.trim();
    const orig = activeInput.dataset.original;
    // Remove listeners so finishEdit won't double-fire
    activeInput.removeEventListener('blur', activeInput._finishEdit);
    activeInput.removeEventListener('keydown', activeInput._onKey);
    if (nv !== '' && nv !== orig) {
      const nextIdx = focusedFieldIdx + 1;
      saveVerification(key, 'corrected', nv);
      showToast('\u2713 ' + key.split(':').pop().replace(/_/g,' ') + ' corrected', 'success');
      // Correction changes displayed value — need full row rebuild
      loadPage(currentPage, nextIdx);
      _perfEnd('confirmField');
      return;
    }
  }
  const current = verifications[key];
  if (current && current.status === 'confirmed') {
    // Toggle off — un-confirm
    delete verifications[key];
    fetch('/api/verify/' + currentJobId, {
      method: 'POST', headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ fields: { [key]: { status: '_remove' } }, reviewer: getReviewer() })
    }).catch(() => {});
    _updateRowVerification(key, null);
    updateVerifyBar();
    _perfEnd('confirmField');
    return;
  }
  if (current && current.status === 'corrected' && current._justCorrected) {
    // Just corrected by finishEdit blur race — don't overwrite with plain confirmed.
    // Clear the flag and advance to next field.
    delete current._justCorrected;
    const nextIdx = focusedFieldIdx + 1;
    showToast('\u2713 ' + key.split(':').pop().replace(/_/g,' ') + ' corrected', 'success');
    _updateRowVerification(key, 'corrected');
    moveFocus(nextIdx);
    _perfEnd('confirmField');
    return;
  }
  const nextIdx = focusedFieldIdx + 1;
  saveVerification(key, 'confirmed', current && current.corrected_value !== undefined ? current.corrected_value : null);
  showToast('\u2713 ' + key.split(':').pop().replace(/_/g,' '), 'success');
  _updateRowVerification(key, 'confirmed');
  moveFocus(nextIdx);
  _perfEnd('confirmField');
}

function flagField(key) {
  _perf('flagField');
  const curIdx = focusedFieldIdx;
  const current = verifications[key];
  if (current && current.status === 'flagged') {
    // Toggle off — un-flag
    delete verifications[key];
    fetch('/api/verify/' + currentJobId, {
      method: 'POST', headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ fields: { [key]: { status: '_remove' } }, reviewer: getReviewer() })
    }).catch(() => {});
    _updateRowVerification(key, null);
    updateVerifyBar();
  } else {
    const note = prompt('Flag note (optional):') || '';
    saveVerification(key, 'flagged', null, note);
    _updateRowVerification(key, 'flagged');
    showToast('\u26A0 Flagged: ' + key.split(':').pop().replace(/_/g,' '), 'error');
  }
  _perfEnd('flagField');
}

function startEdit(key, currentVal) {
  const row = document.querySelector('[data-key="' + key.replace(/"/g, '\\"') + '"]');
  if (!row) return;
  const valSpan = row.querySelector('.field-val');
  if (!valSpan) return;
  const input = document.createElement('input');
  input.type = 'text'; input.className = 'field-edit-input';
  input.value = currentVal; input.dataset.key = key; input.dataset.original = String(currentVal);
  valSpan.innerHTML = ''; valSpan.appendChild(input); input.focus(); input.select();
  let finished = false;
  function finishEdit() {
    if (finished) return;  // Prevent double-fire from blur + confirmField
    finished = true;
    const nv = input.value.trim();
    input.removeEventListener('blur', finishEdit); input.removeEventListener('keydown', onKey);
    if (nv !== '' && nv !== String(currentVal)) {
      saveVerification(key, 'corrected', nv);
      if (verifications[key]) verifications[key]._justCorrected = true;
      showToast('\u2713 ' + key.split(':').pop().replace(/_/g,' ') + ' corrected', 'success');
    }
    loadPage(currentPage, focusedFieldIdx);
  }
  function onKey(e) {
    if (e.key === 'Enter') { e.preventDefault(); e.stopPropagation(); finishEdit(); }
    else if (e.key === 'Escape') { finished = true; loadPage(currentPage, focusedFieldIdx); }
  }
  // Store references so confirmField can remove them
  input._finishEdit = finishEdit;
  input._onKey = onKey;
  input.addEventListener('blur', finishEdit);
  input.addEventListener('keydown', onKey);
}

function startRelabel(key, currentName) {
  const escapedKey = key.replace(/"/g, '\\"');
  const row = document.querySelector('[data-key="' + escapedKey + '"]');
  if (!row) return;
  const nameSpan = row.querySelector('.field-name');
  if (!nameSpan) return;
  // Replace field name with editable input
  const input = document.createElement('input');
  input.type = 'text';
  input.className = 'field-relabel-input';
  input.value = currentName;
  input.dataset.key = key;
  input.dataset.original = currentName;
  nameSpan.innerHTML = '';
  nameSpan.appendChild(input);
  input.focus();
  input.select();
  let finished = false;
  function finishRelabel() {
    if (finished) return;
    finished = true;
    const nv = input.value.trim();
    input.removeEventListener('blur', finishRelabel);
    input.removeEventListener('keydown', onKey);
    if (nv && nv !== currentName) {
      // Save relabeled name — preserve existing verification status
      const existing = verifications[key];
      const curStatus = (existing && existing.status) ? existing.status : 'confirmed';
      const curCorrected = existing ? existing.corrected_value : undefined;
      const curNote = existing ? existing.note : undefined;
      const curCat = existing ? existing.category : undefined;
      const curVendor = existing ? existing.vendor_desc : undefined;
      saveVerification(key, curStatus, curCorrected, curNote, curCat, curVendor, {relabeled_name: nv});
      showToast('\u270E Label renamed to: ' + nv, 'success');
    }
    loadPage(currentPage, focusedFieldIdx);
  }
  function onKey(e) {
    if (e.key === 'Enter') { e.preventDefault(); e.stopPropagation(); finishRelabel(); }
    else if (e.key === 'Escape') { finished = true; loadPage(currentPage, focusedFieldIdx); }
  }
  input.addEventListener('blur', finishRelabel);
  input.addEventListener('keydown', onKey);
}

function toggleNoteInput(key) {
  const escapedKey = key.replace(/"/g, '\\"');
  const row = document.querySelector('[data-key="' + escapedKey + '"]');
  if (!row) return;
  // Check if note input already exists after this row
  const existing = row.nextElementSibling;
  if (existing && existing.classList.contains('vf-note-input')) {
    existing.remove();
    return;
  }
  const current = verifications[key];
  const currentNote = (current && current.note) || '';
  const div = document.createElement('div');
  div.className = 'vf-note-input';
  div.innerHTML = '<input type="text" placeholder="Add a review note..." value="' + esc(currentNote) + '" onkeydown="if(event.key===\'Enter\')saveFieldNote(\'' + esc(key) + '\',this.value)">'
    + '<button onclick="saveFieldNote(\'' + esc(key) + '\',this.previousElementSibling.value)">Save</button>';
  row.after(div);
  div.querySelector('input').focus();
}

function saveFieldNote(key, note) {
  note = (note || '').trim();
  const current = verifications[key] || {};
  const status = current.status || 'confirmed';
  saveVerification(key, status, current.corrected_value !== undefined ? current.corrected_value : null, note);
  showToast(note ? 'Note saved' : 'Note removed', 'success');
  loadPage(currentPage, focusedFieldIdx);
}

// ─── Category Handling ───
function normalizeVendor(desc) {
  if (!desc) return '';
  var s = String(desc).toUpperCase().trim();
  s = s.replace(/[\s#*]+\d{2,}$/, '');
  s = s.replace(/\s+(LLC|INC|CORP|CO|COMPANY|LTD|LP|NA|N\.A\.)\s*$/i, '');
  s = s.replace(/[\s.,;:*#\-]+$/, '');
  return s.trim();
}
function suggestCategory(desc) {
  if (!desc || !vendorMap) return '';
  var norm = normalizeVendor(desc);
  if (!norm) return '';
  if (vendorMap[norm]) return vendorMap[norm].category || '';
  for (var k in vendorMap) { if (norm.indexOf(k)===0 || k.indexOf(norm)===0) return vendorMap[k].category||''; }
  return '';
}
function buildCategorySelect(vk, vendorDesc, compact) {
  var existing = verifications[vk];
  var currentCat = (existing && existing.category) ? existing.category : '';
  var suggested = !currentCat && vendorDesc ? suggestCategory(vendorDesc) : '';
  var activeCat = currentCat || suggested;
  var cls = 'cat-select' + (currentCat ? ' cat-set' : suggested ? ' cat-suggested' : '');
  var h = '<select class="'+cls+'" onchange="saveFieldCategory(\''+esc(vk)+'\',this,\''+esc(String(vendorDesc||'').replace(/'/g,"\\'"))+'\')" title="'+(activeCat?esc(activeCat):'Assign category')+'">';
  h += '<option value="">'+(compact?'\u2014':'— Category —')+'</option>';
  for (var g in chartOfAccounts) {
    h += '<optgroup label="'+esc(g)+'">';
    (chartOfAccounts[g]||[]).forEach(function(a) { h += '<option value="'+esc(a)+'"'+(a===activeCat?' selected':'')+'>'+esc(a)+'</option>'; });
    h += '</optgroup>';
  }
  h += '</select>';
  if (suggested && !currentCat) h += ' <span class="cat-learned-badge">auto</span>';
  return h;
}
function saveFieldCategory(vk, sel, vendorDesc) {
  var cat = sel.value;
  sel.className = cat ? 'cat-select cat-set' : 'cat-select';
  var badge = sel.parentElement ? sel.parentElement.querySelector('.cat-learned-badge') : null;
  if (badge) badge.remove();
  var ex = verifications[vk] || {};
  saveVerification(vk, ex.status||'confirmed', ex.corrected_value||undefined, ex.note||undefined, cat, vendorDesc);
  if (cat && vendorDesc) { var n = normalizeVendor(vendorDesc); if(n) vendorMap[n] = {category:cat,count:1}; }
  if (cat) showToast('\uD83D\uDCC1 ' + cat, 'success');
}
function needsCategoryPicker(fn, dt) {
  if (!dt) return false;
  if (/^txn_\d+_amount$/.test(fn)) return false;
  if (dt === 'check' && fn === 'check_amount') return true;
  if (/invoice/.test(dt) && fn === 'total_amount') return true;
  if (/receipt/.test(dt) && fn === 'total_amount') return true;
  return false;
}

function toggleInfoSection(id, toggle) {
  const el = document.getElementById(id);
  if (!el) return;
  const showing = el.style.display !== 'none';
  el.style.display = showing ? 'none' : 'block';
  const arrow = toggle.querySelector('.info-toggle-arrow');
  if (arrow) arrow.classList.toggle('open', !showing);
}

// ─── Page Rendering ───
function loadPage(page, focusIdx) {
  _perf('loadPage');
  if (!reviewData || !reviewData.page_map) return;
  totalPages = reviewData.total_pages || Object.keys(reviewData.page_map).length;
  if (page < 1) page = 1;
  if (page > totalPages) page = totalPages;
  currentPage = page;
  focusedFieldIdx = (focusIdx !== undefined && focusIdx !== null) ? focusIdx : 0;
  pageFieldKeys = [];

  document.getElementById('reviewPager').textContent = page + ' / ' + totalPages;
  if (_loadedImagePage !== page) {
    document.getElementById('pdfViewer').innerHTML = '<div class="review-pdf-wrap"><img src="/api/page-image/' + currentJobId + '/' + page + '" alt="Page ' + page + '" onload="_onPageImgLoad(this)"></div>';
    _loadedImagePage = page;
    _pageWordData = null;
    _pageWordPage = null;
  }
  // Always ensure word data is available (even if image was cached from prior session)
  _fetchPageWords(page);

  const pageExts = reviewData.page_map[currentPage];
  let html = '';

  if (!pageExts || pageExts.length === 0) {
    html = '<div style="padding:40px 20px;text-align:center;color:var(--text-light)">'
      + '<div style="font-size:32px;margin-bottom:12px">&#128196;</div>'
      + '<div style="font-size:14px;font-weight:600;margin-bottom:6px">No extracted data for this page</div>'
      + '<div style="font-size:12px">This page may be a continuation, composite summary, or supplemental info that was processed as part of another document.</div>'
      + '<div style="margin-top:12px"><button class="btn btn-secondary btn-sm" onclick="reextractPage()">&#x21BB; Re-extract this page</button></div>'
      + '</div>';
    document.getElementById('fieldsPanel').innerHTML = html;
    updateVerifyBar();
    return;
  }

  const skipFields = REVIEW_SKIP_FIELDS;

  pageExts.forEach((ext, extIdx) => {
    const fields = ext.fields || {};
    // Sort fields by document-type order (box/line number), then alphabetical for unknowns
    const docOrder = FIELD_ORDER[ext.document_type] || [];
    const allKeys = Object.keys(fields).sort((a, b) => {
      const ai = docOrder.indexOf(a), bi = docOrder.indexOf(b);
      if (ai !== -1 && bi !== -1) return ai - bi;
      if (ai !== -1) return -1;
      if (bi !== -1) return 1;
      return a.localeCompare(b);
    });

    html += '<div class="field-group">';
    html += '<div class="field-group-title">' + esc(ext.document_type) + '</div>';
    html += '<div class="field-entity"><span>' + esc(ext.entity) + '</span></div>';

    // Separate txn fields from summary
    const txnRegex = /^txn_(\d+)_(date|desc|amount|type)$/;
    const summaryKeys = allKeys.filter(k => !txnRegex.test(k));
    const txnKeys = allKeys.filter(k => txnRegex.test(k));
    const txnGroups = {};
    txnKeys.forEach(k => { const m = k.match(txnRegex); if(m) { if(!txnGroups[m[1]]) txnGroups[m[1]]={}; txnGroups[m[1]][m[2]]=k; }});
    const txnNums = Object.keys(txnGroups).sort((a,b)=>parseInt(a)-parseInt(b));

    // Split monetary vs info, filter out $0.00 unless significant
    const monetaryKeys = summaryKeys.filter(k => {
      if (skipFields.has(k)) return false;
      const v = fields[k].value;
      const isNumeric = typeof v === 'number' || (typeof v === 'string' && /^\-?\$?[\d,]+\.?\d*$/.test(v.trim()));
      if (!isNumeric) return false;
      // Filter out zero values unless field name suggests significance
      const numVal = typeof v === 'number' ? v : parseFloat(String(v).replace(/[$,]/g, ''));
      if (numVal === 0 && !/(balance|total|net)/i.test(k)) return false;
      return true;
    });
    const infoKeys = summaryKeys.filter(k => {
      if (skipFields.has(k)) return false;  // Hide skipped fields entirely
      const v = fields[k].value;
      return !(typeof v === 'number' || (typeof v === 'string' && /^\-?\$?[\d,]+\.?\d*$/.test(v.trim())));
    });

    // Info section (collapsible)
    if (infoKeys.length > 0) {
      const colId = 'info-' + currentPage + '-' + extIdx;
      html += '<div class="info-section">';
      html += '<div class="info-toggle" onclick="toggleInfoSection(\'' + colId + '\',this)">';
      html += '<span class="info-toggle-arrow">\u25B6</span> Document Info (' + infoKeys.length + ')</div>';
      html += '<div class="info-fields" id="' + colId + '" style="display:none">';
      infoKeys.forEach(k => {
        const v = fields[k].value;
        html += '<div class="info-field"><span class="info-field-name">' + esc(k.replace(/_/g,' ').replace(/\b\w/g,c=>c.toUpperCase())) + '</span><span class="info-field-val">' + esc(v==null?'\u2014':String(v)) + '</span></div>';
      });
      html += '</div></div>';
    }

    // Monetary fields
    monetaryKeys.forEach(k => {
      const f = fields[k];
      const vk = fieldKey(currentPage, extIdx, k);
      const vstate = verifications[vk] || null;
      pageFieldKeys.push(vk);
      const idx = pageFieldKeys.length - 1;

      const rawVal = f.value;
      let displayVal = rawVal;
      if (vstate && vstate.corrected_value !== undefined) displayVal = vstate.corrected_value;
      const displayStr = typeof displayVal === 'number' ? displayVal.toLocaleString('en-US',{minimumFractionDigits:2,maximumFractionDigits:2}) : String(displayVal||'');

      let rowClass = 'field-row';
      if (idx === focusedFieldIdx) rowClass += ' focused';
      if (vstate) { if (vstate.status==='confirmed') rowClass+=' vf-confirmed'; else if (vstate.status==='corrected') rowClass+=' vf-corrected'; else if (vstate.status==='flagged') rowClass+=' vf-flagged'; }

      const conf = f.confidence || '';
      let dotClass = 'conf-other';
      if (conf.includes('dual')) dotClass='conf-dual'; else if (conf.includes('confirmed')||conf==='ocr_accepted') dotClass='conf-confirmed'; else if (conf.includes('corrected')) dotClass='conf-corrected'; else if (conf==='low') dotClass='conf-low';

      html += '<div class="' + rowClass + '" data-key="' + esc(vk) + '" data-idx="' + idx + '" onclick="setFocus(' + idx + ')">';
      const boxLabel = (FIELD_BOX_LABELS[ext.document_type] || {})[k];
      const defaultName = (boxLabel ? boxLabel + ' \u2014 ' : '') + k.replace(/_/g,' ').replace(/\b\w/g,c=>c.toUpperCase());
      const displayName = (vstate && vstate.relabeled_name) ? vstate.relabeled_name : defaultName;
      const entityHint = ext.entity ? '<span class="field-entity-hint">' + esc(ext.entity) + '</span>' : '';
      html += '<span class="field-name">' + esc(displayName) + '<span class="relabel-icon" onclick="event.stopPropagation();startRelabel(\'' + esc(vk) + '\',' + JSON.stringify(displayName).replace(/&/g,'&amp;').replace(/"/g,'&quot;') + ')" title="Rename field label">\u270E</span>' + entityHint + '</span>';
      html += '<span class="field-val-wrap"><span class="conf-dot ' + dotClass + '"></span>';
      const safeVal = JSON.stringify(displayStr).replace(/&/g,'&amp;').replace(/"/g,'&quot;');
      html += '<span class="field-val" onclick="event.stopPropagation();startEdit(\'' + esc(vk) + '\',' + safeVal + ')">' + esc(displayStr) + '</span>';
      html += '<span class="field-actions">';
      html += '<button class="vf-btn" onclick="event.stopPropagation();startEdit(\'' + esc(vk) + '\',' + safeVal + ')" title="Edit value (E)">&#x270F;</button>';
      html += '<button class="vf-btn vf-btn-confirm' + (vstate&&vstate.status==='confirmed'?' active':'') + '" onclick="event.stopPropagation();confirmField(\'' + esc(vk) + '\')" title="Confirm (Enter)">\u2713</button>';
      html += '<button class="vf-btn vf-btn-flag' + (vstate&&vstate.status==='flagged'?' active':'') + '" onclick="event.stopPropagation();flagField(\'' + esc(vk) + '\')" title="Flag (F)">\u2691</button>';
      html += '<button class="vf-btn vf-btn-note' + (vstate&&vstate.note?' has-note':'') + '" onclick="event.stopPropagation();toggleNoteInput(\'' + esc(vk) + '\')" title="Add note (N)">&#x270E;</button>';
      html += '</span></span></div>';

      if (vstate && vstate.status==='corrected') { html += '<div class="vf-note"><span class="vf-original">' + esc(String(rawVal)) + '</span> \u2192 ' + esc(displayStr) + '</div>'; }
      if (vstate && vstate.note) { html += '<div class="vf-note" id="note-' + esc(vk) + '">' + esc(vstate.note) + '</div>'; }

      if (needsCategoryPicker(k, ext.document_type)) {
        var cv = '';
        if (ext.document_type==='check') { cv = (fields.payee&&fields.payee.value)||(fields.pay_to&&fields.pay_to.value)||ext.entity||''; }
        else { cv = (fields.vendor_name&&fields.vendor_name.value)||ext.entity||''; }
        html += '<div class="field-cat-row"><label>Account:</label>' + buildCategorySelect(vk, String(cv), false) + '</div>';
      }
    });

    // Transaction table
    if (txnNums.length > 0) {
      html += '<div class="txn-section"><div class="txn-header">Transactions (' + txnNums.length + ')</div>';
      html += '<table class="txn-table"><thead><tr><th>Date</th><th>Description</th><th class="txn-amt">Amount</th><th>Type</th><th>Category</th><th></th></tr></thead><tbody>';
      txnNums.forEach(num => {
        const grp = txnGroups[num];
        const pk = grp.amount||grp.desc||grp.date;
        if (!pk) return;
        const vk = fieldKey(currentPage, extIdx, pk);
        const vstate = verifications[vk]||null;
        pageFieldKeys.push(vk);
        let trCls = '';
        if (vstate) { if (vstate.status==='confirmed') trCls=' class="vf-confirmed"'; else if (vstate.status==='flagged') trCls=' class="vf-flagged"'; }
        const dv = grp.date ? (fields[grp.date].value||'') : '';
        const descV = grp.desc ? (fields[grp.desc].value||'') : '';
        const amtF = grp.amount ? fields[grp.amount] : null;
        const amtV = amtF ? (typeof amtF.value==='number'?amtF.value.toLocaleString('en-US',{minimumFractionDigits:2,maximumFractionDigits:2}):(amtF.value||'')) : '';
        const tpV = grp.type ? (fields[grp.type].value||'') : '';
        const conf = amtF ? (amtF.confidence||'') : '';
        let dc = 'conf-other';
        if (conf.includes('dual')) dc='conf-dual'; else if (conf.includes('confirmed')||conf==='ocr_accepted') dc='conf-confirmed'; else if (conf.includes('corrected')) dc='conf-corrected'; else if (conf==='low') dc='conf-low';
        html += '<tr' + trCls + ' data-key="' + esc(vk) + '">';
        html += '<td>'+esc(String(dv))+'</td><td>'+esc(String(descV))+'</td>';
        html += '<td class="txn-amt"><span class="conf-dot '+dc+'"></span>'+esc(String(amtV))+'</td>';
        html += '<td><span class="txn-type txn-type-'+esc(String(tpV).toLowerCase())+'">'+esc(String(tpV))+'</span></td>';
        html += '<td>'+buildCategorySelect(vk,String(descV),true)+'</td>';
        html += '<td><button class="vf-btn vf-btn-confirm'+(vstate&&vstate.status==='confirmed'?' active':'')+'" onclick="confirmField(\''+esc(vk)+'\')">\u2713</button></td>';
        html += '</tr>';
      });
      html += '</tbody></table></div>';
    }

    html += '</div>'; // field-group
  });

  document.getElementById('fieldsPanel').innerHTML = html;
  if (focusedFieldIdx >= pageFieldKeys.length) focusedFieldIdx = Math.max(0, pageFieldKeys.length - 1);
  // Trigger highlight for focused field (word data may still be loading)
  _highlightFocusedField();
  _perfEnd('loadPage');
}

function moveFocus(newIdx) {
  // Lightweight focus change: swap CSS classes only, no DOM rebuild
  if (newIdx < 0 || newIdx >= pageFieldKeys.length) return;
  _perf('moveFocus');
  const oldRow = document.querySelector('.field-row.focused');
  if (oldRow) oldRow.classList.remove('focused');
  focusedFieldIdx = newIdx;
  const newRow = document.querySelector('.field-row[data-idx="' + newIdx + '"]');
  if (newRow) {
    newRow.classList.add('focused');
    newRow.scrollIntoView({ block: 'nearest' });
  }
  _highlightFocusedField();
  _perfEnd('moveFocus');
}

function _updateRowVerification(key, status) {
  // Update a single row's visual state after confirm/flag without full rebuild
  const row = document.querySelector('.field-row[data-key="' + key.replace(/"/g, '\\\\"') + '"]');
  if (!row) return;
  row.classList.remove('vf-confirmed', 'vf-corrected', 'vf-flagged');
  if (status) row.classList.add('vf-' + status);
  // Toggle button active states
  const confirmBtn = row.querySelector('.vf-btn-confirm');
  if (confirmBtn) confirmBtn.classList.toggle('active', status === 'confirmed');
  const flagBtn = row.querySelector('.vf-btn-flag');
  if (flagBtn) flagBtn.classList.toggle('active', status === 'flagged');
  // Flash animation on confirm/flag
  row.classList.remove('vf-just-confirmed', 'vf-just-flagged');
  if (status === 'confirmed') {
    row.classList.add('vf-just-confirmed');
    setTimeout(function() { row.classList.remove('vf-just-confirmed'); }, 500);
  } else if (status === 'flagged') {
    row.classList.add('vf-just-flagged');
    setTimeout(function() { row.classList.remove('vf-just-flagged'); }, 500);
  }
}

// ─── PDF Highlight Overlay (grid review) ───
function _normalizeNumJS(s) {
  // Strip $, commas, spaces — keep digits, dots, minus
  if (s == null) return '';
  return String(s).replace(/[$,\s]/g, '').replace(/^0+(\d)/, '$1');
}
function _findValueBboxesJS(words, value) {
  if (value == null) return [];
  var vs = String(value).trim();
  if (!vs) return [];
  var norm = _normalizeNumJS(vs);
  // Strategy 1: exact single-word numeric match
  if (norm && /[\d.]/.test(norm)) {
    for (var i = 0; i < words.length; i++) {
      var wn = _normalizeNumJS(words[i].text);
      if (wn && wn === norm) return [words[i]];
    }
  }
  // Strategy 2: multi-word numeric assembly
  if (norm && /[\d.]/.test(norm)) {
    for (var i = 0; i < words.length; i++) {
      var running = '', group = [];
      for (var j = i; j < Math.min(i + 10, words.length); j++) {
        running += words[j].text;
        group.push(words[j]);
        var rn = _normalizeNumJS(running);
        if (rn && rn === norm) return group;
      }
    }
  }
  // Strategy 3: case-insensitive text sequence
  var upper = vs.toUpperCase(), parts = upper.split(/\s+/);
  if (parts.length) {
    for (var i = 0; i < words.length; i++) {
      if (words[i].text.toUpperCase() === parts[0] || parts[0].indexOf(words[i].text.toUpperCase()) === 0) {
        if (parts.length === 1 && words[i].text.toUpperCase() === parts[0]) return [words[i]];
        var ok = true, grp = [words[i]];
        for (var k = 1; k < parts.length; k++) {
          if (i + k < words.length && words[i + k].text.toUpperCase() === parts[k]) { grp.push(words[i + k]); }
          else { ok = false; break; }
        }
        if (ok && grp.length === parts.length) return grp;
      }
    }
  }
  return [];
}
function _clearHighlights() {
  var old = document.querySelectorAll('.pdf-highlight');
  for (var i = 0; i < old.length; i++) old[i].remove();
}
function _drawHighlights(bboxes) {
  _clearHighlights();
  var wrap = document.querySelector('.review-pdf-wrap');
  if (!wrap || !bboxes.length) return;
  var img = wrap.querySelector('img');
  if (!img || !_pageImgNatW) return;
  var scaleX = img.clientWidth / _pageImgNatW;
  var scaleY = img.clientHeight / _pageImgNatH;
  var PAD = 6;
  bboxes.forEach(function(b) {
    var div = document.createElement('div');
    div.className = 'pdf-highlight pdf-highlight-pulse';
    div.style.left = Math.max(0, b.left * scaleX - PAD) + 'px';
    div.style.top = Math.max(0, b.top * scaleY - PAD) + 'px';
    div.style.width = (b.width * scaleX + PAD * 2) + 'px';
    div.style.height = (b.height * scaleY + PAD * 2) + 'px';
    wrap.appendChild(div);
  });
  // Scroll highlight into view within the PDF panel
  var first = wrap.querySelector('.pdf-highlight');
  if (first) first.scrollIntoView({ block: 'nearest', behavior: 'smooth' });
}
function _onPageImgLoad(img) {
  _pageImgNatW = img.naturalWidth;
  _pageImgNatH = img.naturalHeight;
  // If words already loaded, draw highlight now
  if (_pageWordData) _highlightFocusedField();
}
function _highlightFocusedField() {
  if (!_pageWordData || focusedFieldIdx < 0 || focusedFieldIdx >= pageFieldKeys.length) {
    _clearHighlights();
    return;
  }
  var key = pageFieldKeys[focusedFieldIdx];
  // Extract the field value from the DOM
  var row = document.querySelector('.field-row[data-idx="' + focusedFieldIdx + '"]');
  if (!row) { _clearHighlights(); return; }
  var valEl = row.querySelector('.field-val');
  if (!valEl) { _clearHighlights(); return; }
  var val = valEl.textContent.trim();
  var bboxes = _findValueBboxesJS(_pageWordData, val);
  _drawHighlights(bboxes);
}
function _fetchPageWords(page) {
  if (_pageWordPage === page && _pageWordData !== null) return; // already have it
  _pageWordData = null;
  _pageWordPage = page;
  fetch('/api/page-words/' + currentJobId + '/' + page)
    .then(function(r) { return r.json(); })
    .then(function(words) {
      if (_pageWordPage === page) { // still on same page?
        _pageWordData = words;
        _highlightFocusedField(); // draw highlight now that words are loaded
      }
    })
    .catch(function() { _pageWordData = []; });
}

function setFocus(idx) {
  // Don't change focus if an edit input is active (would destroy it)
  if (document.querySelector('.field-edit-input')) return;
  moveFocus(idx);
}
function prevPage() { _loadedImagePage = null; _pageWordData = null; _pageWordPage = null; if (currentPage > 1) loadPage(currentPage - 1); }
function nextPage() { _loadedImagePage = null; _pageWordData = null; _pageWordPage = null; if (currentPage < totalPages) loadPage(currentPage + 1); }

function reextractPage() {
  if (!currentJobId || !currentPage) return;
  const instructions = prompt('Enter instructions for re-extracting this page (e.g., "This is a K-1 Schedule, focus on box 1-3"):');
  if (!instructions) return;
  showToast('Re-extracting page ' + currentPage + '...', 'info');
  fetch('/api/reextract-page/' + currentJobId + '/' + currentPage, {
    method: 'POST',
    headers: { 'Content-Type': 'application/json' },
    body: JSON.stringify({ instructions: instructions })
  })
  .then(r => r.json())
  .then(data => {
    if (data.error) { showToast('Re-extract failed: ' + data.error, 'error'); return; }
    showToast('Page ' + currentPage + ' re-extracted successfully', 'success');
    // Reload extractions + verifications (must fetch both separately)
    Promise.all([
      fetch('/api/results/' + currentJobId).then(r => r.json()),
      fetch('/api/verify/' + currentJobId).then(r => r.json()),
    ]).then(function([rd, vd]) {
      if (rd.error) return;
      reviewData = rd;
      verifications = (vd && vd.fields) ? vd.fields : verifications;
      totalPages = rd.total_pages || 1;
      countTotalFields();
      updateVerifyBar();
      loadPage(currentPage);
      _initFilterBar();
    });
  })
  .catch(e => { showToast('Re-extract failed: ' + e, 'error'); });
}

// ─── Downloads ───
function downloadFile(type) {
  if (!currentJobId) return;
  window.location = '/api/download' + (type==='log'?'-log':'') + '/' + currentJobId;
}

function regenExcel() {
  if (!currentJobId) return;
  showToast('Regenerating Excel...', 'info');
  fetch('/api/regen-excel/' + currentJobId, { method: 'POST' })
    .then(r => r.json())
    .then(data => {
      if (data.error) { showToast('Regen failed: ' + data.error, 'error'); return; }
      showToast('Excel regenerated successfully', 'success');
    })
    .catch(e => { showToast('Regen failed: ' + e, 'error'); });
}

// ─── T1.6: Workpaper Generation ───
function generateWorkpaper() {
  if (!currentJobId) { showToast('No job selected', 'error'); return; }
  const job = jobs.find(j => j.id === currentJobId);
  if (!job) { showToast('Job not found', 'error'); return; }
  const clientName = job.client_name;
  if (!clientName) { showToast('No client name set for this job', 'error'); return; }

  const year = job.year || prompt('Enter tax year:', '2025');
  if (!year) return;

  const mode = confirm('Use Safe mode? (only verified values)\\n\\nOK = Safe mode\\nCancel = Assisted mode (all values, flagged)') ? 'safe' : 'assisted';

  showToast('Generating workpaper (' + mode + ' mode)...', 'info');
  fetch('/api/workpaper/' + encodeURIComponent(clientName), {
    method: 'POST',
    headers: { 'Content-Type': 'application/json' },
    body: JSON.stringify({ year: year, mode: mode })
  })
  .then(r => r.json())
  .then(data => {
    if (data.error) {
      showToast('Workpaper failed: ' + data.error, 'error');
      return;
    }
    showToast('Workpaper generated (' + data.facts_count + ' facts, ' + mode + ' mode)', 'success');
    window.open(data.download_url, '_blank');
  })
  .catch(e => { showToast('Workpaper failed: ' + e, 'error'); });
}

// ─── AI Chat ───
function toggleAiChat() {
  const panel = document.getElementById('aiChatPanel');
  const isVisible = panel.style.display !== 'none';
  panel.style.display = isVisible ? 'none' : '';
  if (!isVisible) {
    document.getElementById('aiChatPage').textContent = currentPage;
    document.getElementById('aiChatInput').focus();
  }
}

function sendAiChat() {
  const input = document.getElementById('aiChatInput');
  const message = input.value.trim();
  if (!message || !currentJobId) return;
  input.value = '';

  const msgs = document.getElementById('aiChatMessages');
  msgs.innerHTML += '<div style="margin-bottom:8px;"><strong style="color:var(--navy);">You:</strong> ' + esc(message) + '</div>';
  msgs.innerHTML += '<div id="aiTyping" style="margin-bottom:8px;color:var(--text-light);font-style:italic;">AI is thinking...</div>';
  msgs.scrollTop = msgs.scrollHeight;

  fetch('/api/ai-chat/' + currentJobId, {
    method: 'POST',
    headers: { 'Content-Type': 'application/json' },
    body: JSON.stringify({ message: message, page: currentPage })
  })
  .then(r => r.json())
  .then(data => {
    const typing = document.getElementById('aiTyping');
    if (typing) typing.remove();
    if (data.error) {
      msgs.innerHTML += '<div style="margin-bottom:8px;color:var(--red);">Error: ' + esc(data.error) + '</div>';
    } else {
      msgs.innerHTML += '<div style="margin-bottom:8px;"><strong style="color:var(--accent);">AI:</strong> ' + esc(data.reply).replace(/\n/g, '<br>') + '</div>';
    }
    msgs.scrollTop = msgs.scrollHeight;
  })
  .catch(e => {
    const typing = document.getElementById('aiTyping');
    if (typing) typing.remove();
    msgs.innerHTML += '<div style="margin-bottom:8px;color:var(--red);">Error: ' + esc(String(e)) + '</div>';
  });
}

// ─── History ───
function loadJobs() {
  fetch('/api/jobs').then(r=>r.json()).then(data => {
    allJobs = data;
    document.getElementById('historyCount').textContent = data.length;
    renderHistory(data);
  }).catch(()=>{});
}

function formatDuration(secs) {
  if (secs === null || secs === undefined) return '\u2014';
  if (secs < 60) return secs + 's';
  const m = Math.floor(secs / 60); const s = secs % 60;
  if (m >= 60) { const h = Math.floor(m / 60); return h + 'h ' + (m % 60) + 'm'; }
  return m + 'm ' + s + 's';
}

function renderHistory(data) {
  const body = document.getElementById('historyBody');
  if (!data.length) { body.innerHTML = '<tr><td colspan="10" style="text-align:center;padding:24px;color:var(--text-light)">No jobs yet</td></tr>'; return; }
  body.innerHTML = data.map(j => {
    const dt = j.created ? new Date(j.created).toLocaleDateString('en-US',{month:'short',day:'numeric',hour:'numeric',minute:'2-digit'}) : '';
    const typeLabel = DOC_TYPES.find(d=>d.id===j.doc_type);
    const costStr = j.cost_usd ? '$' + j.cost_usd.toFixed(4) : '—';
    return '<tr>' +
      '<td><strong>' + esc(j.client_name||'—') + '</strong></td>' +
      '<td>' + esc(j.filename||'') + '</td>' +
      '<td><span class="badge badge-blue">' + esc(typeLabel?typeLabel.label:j.doc_type||'') + '</span></td>' +
      '<td>' + esc(j.year||'') + '</td>' +
      '<td><span class="job-status ' + (j.status||'') + '">' + esc(j.status||'') + '</span></td>' +
      '<td style="font-size:12px;font-family:var(--mono);color:var(--text-secondary)">' + costStr + '</td>' +
      '<td style="font-size:12px;font-family:var(--mono);color:var(--text-secondary)">' + formatDuration(j.duration_seconds) + '</td>' +
      '<td style="font-size:12px;font-family:var(--mono);color:var(--text-secondary)">' + formatDuration(j.review_time_seconds) + '</td>' +
      '<td style="font-size:12px;color:var(--text-secondary)">' + dt + '</td>' +
      '<td class="actions">' +
        (j.status==='complete'?'<button class="btn btn-sm btn-secondary" onclick=\'openReview('+JSON.stringify({id:j.id,client_name:j.client_name})+')\'>\u{1F50D} Review</button> ':'') +
        (j.status==='running'||j.status==='queued'?'<button class="btn btn-sm btn-primary" onclick="monitorJob(\''+j.id+'\')">Monitor</button> ':'') +
        (j.status==='failed'||j.status==='interrupted'||j.status==='error'?'<button class="btn btn-sm btn-secondary" onclick="retryJob(\''+j.id+'\')">Retry</button> ':'') +
        '<button class="btn btn-ghost btn-sm" onclick="deleteJob(\''+j.id+'\')" title="Delete">\u2716</button>' +
      '</td></tr>';
  }).join('');
}

function filterHistory() {
  const q = (document.getElementById('historySearch').value||'').toLowerCase();
  const s = document.getElementById('historyStatusFilter').value;
  const filtered = allJobs.filter(j => {
    if (s && j.status !== s) return false;
    if (q && !(j.client_name||'').toLowerCase().includes(q) && !(j.filename||'').toLowerCase().includes(q)) return false;
    return true;
  });
  renderHistory(filtered);
}

function retryJob(id) { fetch('/api/retry/'+id,{method:'POST'}).then(r=>r.json()).then(d=>{if(d.job_id){currentJobId=d.job_id;showSection('processing');startPolling();}}).catch(()=>{}); }
function deleteJob(id) { if(!confirm('Delete this job?')) return; fetch('/api/delete/'+id,{method:'POST'}).then(()=>loadJobs()).catch(()=>{}); }
function monitorJob(id) {
  currentJobId = id;
  const job = allJobs.find(j => j.id === id);
  if (job) document.getElementById('processingFile').textContent = job.filename || '';
  showSection('processing');
  startPolling();
}

// ─── Clients ───
function loadClientSuggestions() {
  fetch('/api/clients').then(r=>r.json()).then(data => {
    const sel = document.getElementById('clientName');
    const current = sel.value;
    sel.innerHTML = '<option value="">\u2014 Select client \u2014</option>' +
      data.map(c => {
        const label = c.ein_last4 ? c.name + ' (' + c.ein_last4 + ')' : c.name;
        return '<option value="' + esc(c.name) + '">' + esc(label) + '</option>';
      }).join('');
    if (current) sel.value = current;
  }).catch(()=>{});
}

let allClientsData = [];
function loadClients() {
  fetch('/api/clients').then(r=>r.json()).then(data => {
    allClientsData = data;
    renderClientGrid(data);
  }).catch(()=>{});
}
function filterClients() {
  const q = (document.getElementById('clientSearch').value||'').toLowerCase();
  renderClientGrid(allClientsData.filter(c => c.name.toLowerCase().includes(q)));
}
function renderClientGrid(clients) {
  const g = document.getElementById('clientGrid');
  if (!clients.length) { g.innerHTML = '<div class="empty-state"><h3>No clients yet</h3><p>Upload a document to create a client record.</p></div>'; return; }
  g.innerHTML = clients.map(c => {
    let badges = '';
    if (c.ein_last4) badges += '<span class="badge badge-purple">EIN \u2026'+esc(c.ein_last4)+'</span>';
    if (c.has_context) badges += '<span class="badge badge-purple">Context</span>';
    if (c.has_instructions) badges += '<span class="badge badge-blue">Instructions</span>';
    return '<div class="client-card" onclick="openClientDetail(\''+esc(c.name)+'\')">' +
      '<h4>'+esc(c.name)+'</h4>' +
      '<div class="client-meta"><span>'+c.jobs+' job'+(c.jobs!==1?'s':'')+'</span>' +
      (c.years.length?'<span>'+c.years.join(', ')+'</span>':'') + '</div>' +
      (badges?'<div class="client-badges">'+badges+'</div>':'') +
      '</div>';
  }).join('');
}

function openClientDetail(name) {
  currentClientName = name;
  document.getElementById('clientListView').style.display = 'none';
  document.getElementById('clientDetailView').classList.add('visible');
  document.getElementById('clientDetailName').textContent = name;
  showClientTab('documents');
  loadClientDocuments(name);
  loadClientInfo(name);
  loadContextDocs(name);
  loadInstructions(name);
  // Load file path info
  fetch('/api/clients/'+encodeURIComponent(name)+'/files').then(function(r){return r.json();}).then(function(data) {
    var pathEl = document.getElementById('clientFilePath');
    if (data.client_path) {
      pathEl.textContent = data.client_path;
      pathEl.style.display = '';
      pathEl.dataset.path = data.client_path;
    }
  }).catch(function(){});
}
function copyClientPath() {
  var pathEl = document.getElementById('clientFilePath');
  var path = pathEl ? pathEl.dataset.path : '';
  if (!path && currentClientName) { path = 'clients/' + currentClientName; }
  if (navigator.clipboard) {
    navigator.clipboard.writeText(path).then(function() {
      showToast('Path copied: ' + path, 'success');
    });
  } else {
    // Fallback
    var ta = document.createElement('textarea');
    ta.value = path; document.body.appendChild(ta);
    ta.select(); document.execCommand('copy');
    document.body.removeChild(ta);
    showToast('Path copied: ' + path, 'success');
  }
}
function exportClientZip() {
  if (!currentClientName) { showToast('No client selected', 'error'); return; }
  window.open('/api/clients/' + encodeURIComponent(currentClientName) + '/export-zip', '_blank');
}
function closeClientDetail() {
  document.getElementById('clientListView').style.display = '';
  document.getElementById('clientDetailView').classList.remove('visible');
  currentClientName = '';
}
function showClientTab(tab) {
  document.querySelectorAll('.client-tab').forEach(t => t.classList.remove('active'));
  document.querySelectorAll('.client-tab-content').forEach(t => t.classList.remove('active'));
  const tabContent = document.querySelector('.client-tab-content#tab-'+tab);
  if (tabContent) tabContent.classList.add('active');
  const tabBtn = document.querySelector('.client-tab[data-tab="'+tab+'"]');
  if (tabBtn) tabBtn.classList.add('active');
  if (tab === 'completeness') loadCompleteness(currentClientName);
  if (tab === 'documents') loadClientDocuments(currentClientName);
}

// Client Info
function loadClientInfo(name) {
  fetch('/api/clients/'+encodeURIComponent(name)+'/info').then(r=>r.json()).then(info => {
    const el = document.getElementById('clientDetailMeta');
    let parts = [];
    if (info.ein_last4) parts.push('<span class="badge badge-purple">EIN \u2026'+esc(info.ein_last4)+'</span>');
    if (info.contact) parts.push('<span>'+esc(info.contact)+'</span>');
    if (info.notes) parts.push('<span style="color:var(--text-light)">'+esc(info.notes)+'</span>');
    el.innerHTML = parts.join(' &middot; ');
  }).catch(()=>{});
}

// Client Documents
let clientCompletedJobs = [];
function loadClientDocuments(name) {
  fetch('/api/clients/'+encodeURIComponent(name)+'/documents').then(r=>r.json()).then(data => {
    const el = document.getElementById('clientDocGroups');
    const docs = data.documents || [];
    clientCompletedJobs = docs.filter(d => d.status === 'complete');
    if (!docs.length) {
      el.innerHTML = '<div class="empty-state"><p>No documents yet. Upload a PDF from the Upload section.</p></div>';
      return;
    }
    const grouped = data.grouped || {};
    const hasComplete = clientCompletedJobs.length > 0;
    let html = '<div style="margin-bottom:16px;display:flex;justify-content:space-between;align-items:center">';
    html += '<span style="font-size:13px;color:var(--text-secondary)">' + docs.length + ' document' + (docs.length!==1?'s':'') + '</span>';
    if (hasComplete) html += '<button class="btn btn-primary btn-sm" onclick="openReportModal()">\u{1F4CA} Generate Report</button>';
    html += '</div>';
    const typeLabels = {tax_returns:'Tax Returns',bank_statements:'Bank Statements',trust_documents:'Trust Documents',bookkeeping:'Bookkeeping',payroll:'Payroll',other:'Other'};
    for (const [dtype, items] of Object.entries(grouped)) {
      const label = typeLabels[dtype] || dtype;
      html += '<div class="card" style="margin-bottom:12px"><div class="card-header"><h3>'+esc(label)+' ('+items.length+')</h3></div>';
      html += '<div class="card-body" style="padding:0"><table style="width:100%;font-size:13px;border-collapse:collapse">';
      html += '<tr style="background:var(--bg);border-bottom:1px solid var(--border)"><th style="padding:8px 12px;text-align:left">File</th><th style="padding:8px 12px;text-align:left">Year</th><th style="padding:8px 12px;text-align:left">Status</th><th style="padding:8px 12px;text-align:left">Cost</th><th style="padding:8px 12px;text-align:right">Actions</th></tr>';
      items.forEach(d => {
        const statusClass = d.status==='complete'?'badge-green':d.status==='running'?'badge-blue':'badge-yellow';
        html += '<tr style="border-bottom:1px solid var(--border)">';
        html += '<td style="padding:8px 12px">'+esc(d.filename)+'</td>';
        html += '<td style="padding:8px 12px">'+esc(d.year)+'</td>';
        html += '<td style="padding:8px 12px"><span class="badge '+statusClass+'">'+esc(d.status)+'</span></td>';
        html += '<td style="padding:8px 12px">'+(d.cost_usd != null && d.status==='complete' ? '$'+Number(d.cost_usd).toFixed(4) : '\u2014')+'</td>';
        html += '<td style="padding:8px 12px;text-align:right">';
        if (d.status === 'complete') {
          html += '<button class="btn btn-secondary btn-sm" onclick="openReview({id:\''+esc(d.job_id)+'\'})">Review</button> ';
          if (d.has_xlsx) html += '<a class="btn btn-ghost btn-sm" href="/api/download/'+esc(d.job_id)+'" title="Download Excel">\u{1F4CA}</a> ';
          if (d.has_log) html += '<a class="btn btn-ghost btn-sm" href="/api/download-log/'+esc(d.job_id)+'" title="Download JSON log">\u{1F4CB}</a> ';
        }
        html += '</td></tr>';
      });
      html += '</table></div></div>';
    }
    el.innerHTML = html;
  }).catch(()=>{});
}

// Context
function loadContextDocs(name) {
  fetch('/api/context/'+encodeURIComponent(name)).then(r=>r.json()).then(data => {
    const docs = data.documents || [];
    const el = document.getElementById('contextDocList');
    if (!docs.length) { el.innerHTML = '<div class="empty-state" style="padding:24px"><p>No context documents uploaded yet.</p></div>'; return; }
    el.innerHTML = docs.map(d => '<div class="context-doc">' +
      '<div class="context-doc-icon">\uD83D\uDCC4</div>' +
      '<div class="context-doc-info"><div class="name">'+esc(d.label||d.original_name)+'</div>' +
      '<div class="meta">'+esc(d.year||'')+' &bull; '+d.payer_count+' payers found &bull; '+esc(d.uploaded||'').split('T')[0]+'</div></div>' +
      '<button class="btn btn-ghost btn-sm" onclick="deleteContext(\''+esc(currentClientName)+'\',\''+esc(d.id)+'\')">&#x2716;</button>' +
    '</div>').join('');
  }).catch(()=>{});
}
function uploadContext() {
  const file = document.getElementById('contextFile').files[0];
  if (!file) { showToast('Select a file', 'error'); return; }
  const fd = new FormData();
  fd.append('file', file);
  fd.append('year', document.getElementById('contextYear').value);
  fd.append('label', document.getElementById('contextLabel').value);
  fetch('/api/context/'+encodeURIComponent(currentClientName)+'/upload', {method:'POST', body:fd})
    .then(r=>r.json()).then(d => {
      if (d.error) { showToast(d.error,'error'); return; }
      showToast('Context uploaded — '+d.payers_found+' payers found', 'success');
      document.getElementById('contextFile').value = '';
      document.getElementById('contextLabel').value = '';
      loadContextDocs(currentClientName);
    }).catch(e => showToast('Upload failed','error'));
}
function deleteContext(client, docId) {
  if (!confirm('Delete this context document?')) return;
  fetch('/api/context/'+encodeURIComponent(client)+'/'+docId, {method:'DELETE'}).then(()=>loadContextDocs(client)).catch(()=>{});
}

// Instructions
function loadInstructions(name) {
  fetch('/api/instructions/'+encodeURIComponent(name)).then(r=>r.json()).then(data => {
    const rules = data.rules || [];
    const el = document.getElementById('instructionsList');
    if (!rules.length) { el.innerHTML = '<div class="empty-state" style="padding:24px"><p>No instructions set.</p></div>'; return; }
    el.innerHTML = rules.map(r => '<div class="instruction-item">' +
      '<div class="inst-text">'+esc(r.text)+'</div>' +
      '<div class="inst-date">'+esc((r.created||'').split('T')[0])+'</div>' +
      '<button class="btn btn-ghost btn-sm" onclick="deleteInstruction(\''+esc(currentClientName)+'\',\''+esc(r.id)+'\')">&#x2716;</button>' +
    '</div>').join('');
  }).catch(()=>{});
}
function addInstruction() {
  const text = document.getElementById('newInstruction').value.trim();
  if (!text) { showToast('Enter an instruction','error'); return; }
  fetch('/api/instructions/'+encodeURIComponent(currentClientName), {
    method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({text:text})
  }).then(r=>r.json()).then(d => {
    if (d.error) { showToast(d.error,'error'); return; }
    showToast('Instruction added','success');
    document.getElementById('newInstruction').value = '';
    loadInstructions(currentClientName);
  }).catch(()=>{});
}
function deleteInstruction(client, ruleId) {
  fetch('/api/instructions/'+encodeURIComponent(client)+'/'+ruleId, {method:'DELETE'}).then(()=>loadInstructions(client)).catch(()=>{});
}

// Completeness
function loadCompleteness(name) {
  fetch('/api/context/'+encodeURIComponent(name)+'/completeness').then(r=>r.json()).then(data => {
    const el = document.getElementById('completenessReport');
    const matched = data.matched || [];
    const missing = data.missing || [];
    const newI = data.new || [];
    if (!matched.length && !missing.length && !newI.length) {
      el.innerHTML = '<div class="empty-state"><p>Upload prior-year context to enable completeness tracking. Then process current-year documents to compare.</p></div>';
      return;
    }
    let h = '';
    if (missing.length) {
      h += '<div class="card" style="margin-bottom:12px"><div class="card-header"><h3 style="color:var(--red)">\u26A0 Missing ('+missing.length+')</h3></div><div class="card-body" style="padding:0">';
      missing.forEach(m => { h += '<div class="completeness-item"><div class="completeness-icon">\u23F3</div><div class="completeness-info"><div class="ci-form">'+esc(m.form)+'</div><div class="ci-payer">'+esc(m.payer)+' (EIN '+esc(m.ein)+')</div></div><span class="badge badge-red">Expected</span></div>'; });
      h += '</div></div>';
    }
    if (matched.length) {
      h += '<div class="card" style="margin-bottom:12px"><div class="card-header"><h3 style="color:var(--green)">\u2705 Received ('+matched.length+')</h3></div><div class="card-body" style="padding:0">';
      matched.forEach(m => { h += '<div class="completeness-item"><div class="completeness-icon">\u2705</div><div class="completeness-info"><div class="ci-form">'+esc(m.form)+'</div><div class="ci-payer">'+esc(m.payer)+'</div></div><span class="badge badge-green">Received</span></div>'; });
      h += '</div></div>';
    }
    if (newI.length) {
      h += '<div class="card"><div class="card-header"><h3 style="color:var(--accent)">\u2728 New This Year ('+newI.length+')</h3></div><div class="card-body" style="padding:0">';
      newI.forEach(m => { h += '<div class="completeness-item"><div class="completeness-icon">\uD83C\uDD95</div><div class="completeness-info"><div class="ci-form">'+esc(m.form)+'</div><div class="ci-payer">'+esc(m.payer)+'</div></div><span class="badge badge-blue">New</span></div>'; });
      h += '</div></div>';
    }
    el.innerHTML = h;
  }).catch(()=>{});
}

// ─── Batch Categorize ───
function loadBatchData() {
  const client = (document.getElementById('batchClientFilter').value||'').trim();
  const showAll = document.getElementById('batchShowAll').checked;
  fetch('/api/batch-categories?client='+encodeURIComponent(client)+'&all='+(showAll?'true':'false'))
    .then(r=>r.json()).then(data => {
      batchData = data;
      chartOfAccounts = data.chart_of_accounts || chartOfAccounts;
      renderBatchStats(data);
      renderBatchGroups(data.groups || []);
    }).catch(()=>{});
}
function renderBatchStats(data) {
  document.getElementById('batchStats').innerHTML =
    '<div class="batch-stat"><div class="stat-num">' + (data.total||0) + '</div><div class="stat-label">Total Transactions</div></div>' +
    '<div class="batch-stat"><div class="stat-num" style="color:var(--green)">' + (data.categorized||0) + '</div><div class="stat-label">Categorized</div></div>' +
    '<div class="batch-stat"><div class="stat-num" style="color:var(--yellow)">' + (data.uncategorized||0) + '</div><div class="stat-label">Uncategorized</div></div>';
}
function renderBatchGroups(groups) {
  const el = document.getElementById('batchVendorGroups');
  if (!groups.length) { el.innerHTML = '<div class="empty-state"><h3>No transactions found</h3><p>Process some bank statements or credit card statements first.</p></div>'; return; }
  el.innerHTML = groups.map((g, gi) => {
    let catSel = '<select class="cat-select'+(g.current?' cat-set':g.suggested?' cat-suggested':'')+'" id="bcat-'+gi+'">';
    catSel += '<option value="">— Category —</option>';
    for (const grp in chartOfAccounts) {
      catSel += '<optgroup label="'+esc(grp)+'">';
      (chartOfAccounts[grp]||[]).forEach(a => { catSel += '<option value="'+esc(a)+'"'+((a===(g.current||g.suggested))?' selected':'')+'>'+esc(a)+'</option>'; });
      catSel += '</optgroup>';
    }
    catSel += '</select>';
    return '<div class="vendor-group">' +
      '<div class="vendor-group-header" onclick="toggleBatchGroup('+gi+')">' +
      '<span class="vg-name">'+esc(g.display_name||g.vendor)+'</span>' +
      '<span class="vg-count">'+g.count+' txn'+(g.count!==1?'s':'')+'</span>' +
      '<span class="vg-amount">$'+Math.abs(g.total_amount).toLocaleString('en-US',{minimumFractionDigits:2})+'</span>' +
      catSel +
      ' <button class="btn btn-sm btn-primary" onclick="event.stopPropagation();applyBatchCategory('+gi+')">Apply</button>' +
      (g.suggested&&!g.current?' <span class="cat-learned-badge">auto</span>':'') +
      '</div>' +
      '<div class="vendor-group-items" id="bg-'+gi+'">' +
      '<table class="data-table" style="font-size:12px"><thead><tr><th>Date</th><th>Description</th><th style="text-align:right">Amount</th><th>Source</th></tr></thead><tbody>' +
      (g.items||[]).map(it => '<tr><td>'+esc(it.date||'')+'</td><td>'+esc(it.desc||'')+'</td><td class="amount">$'+Math.abs(it.amount||0).toLocaleString('en-US',{minimumFractionDigits:2})+'</td><td>'+esc(it.source||'')+'</td></tr>').join('') +
      '</tbody></table></div></div>';
  }).join('');
}
function toggleBatchGroup(i) {
  const el = document.getElementById('bg-'+i);
  if (el) el.classList.toggle('open');
}
function filterBatchVendors() {
  const q = (document.getElementById('batchSearch').value||'').toLowerCase();
  if (!batchData) return;
  const filtered = (batchData.groups||[]).filter(g => (g.vendor||'').toLowerCase().includes(q) || (g.display_name||'').toLowerCase().includes(q));
  renderBatchGroups(filtered);
}
function applyBatchCategory(gi) {
  if (!batchData || !batchData.groups || !batchData.groups[gi]) return;
  const g = batchData.groups[gi];
  const sel = document.getElementById('bcat-'+gi);
  const cat = sel ? sel.value : '';
  if (!cat) { showToast('Select a category first','error'); return; }
  const items = (g.items||[]).map(it => ({job_id:it.job_id, field_key:it.field_key, desc:it.desc}));
  fetch('/api/batch-categories/apply', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({vendor:g.vendor, category:cat, items:items, learn:true})
  }).then(r=>r.json()).then(d => {
    if (d.error) { showToast(d.error,'error'); return; }
    showToast(d.applied + ' transactions \u2192 ' + cat, 'success');
    loadBatchData();
  }).catch(e => showToast('Failed: '+e,'error'));
}

// ─── Keyboard Shortcuts ───
function toggleKbdHelp() { document.getElementById('kbdOverlay').classList.toggle('visible'); }
document.addEventListener('keydown', function(e) {
  if (e.target.tagName === 'INPUT' || e.target.tagName === 'TEXTAREA' || e.target.tagName === 'SELECT') {
    // Allow Enter/Escape in guided edit input
    if (e.target.id === 'guidedEditInput') {
      if (e.key === 'Enter') { guidedFinishEdit(); e.preventDefault(); }
      else if (e.key === 'Escape') { guidedCancelEdit(); e.preventDefault(); }
    }
    // Allow normal typing in guided note input
    if (e.target.id === 'guidedNoteInput') {
      return;
    }
    return;
  }
  const sec = document.querySelector('.section.active');
  // Guided review shortcuts
  if (sec && sec.id === 'sec-guided-review') {
    // Skip shortcuts when note input is focused
    if (e.target.id === 'guidedNoteInput') return;
    if (e.key === '?') { toggleKbdHelp(); return; }
    if (e.key === 'y' || e.key === 'Y') { guidedAction('confirm'); e.preventDefault(); }
    else if (e.key === 'e' || e.key === 'E') { guidedStartEdit(); e.preventDefault(); }
    else if (e.key === 'n' || e.key === 'N') { guidedAction('not_present'); e.preventDefault(); }
    else if (e.key === 's' || e.key === 'S') { guidedAction('skip'); e.preventDefault(); }
    else if (e.key === 'Backspace') { guidedGoBack(); e.preventDefault(); }
    else if (e.key === 'Escape') { guidedCancelEdit(); e.preventDefault(); }
    return;
  }
  if (!sec || sec.id !== 'sec-review') {
    if (e.key === '?') toggleKbdHelp();
    return;
  }
  if (e.key === '?') { toggleKbdHelp(); return; }
  // Escape clears filter mode
  if (e.key === 'Escape' && (filterState.activeCategory || filterState.searchText)) { _clearFilter(); e.preventDefault(); return; }
  // Suppress page nav arrows when in cross-doc filter view
  if ((filterState.activeCategory || filterState.searchText) && (e.key === 'ArrowRight' || e.key === 'ArrowLeft')) { e.preventDefault(); return; }
  if (e.key === 'ArrowRight') { nextPage(); e.preventDefault(); }
  else if (e.key === 'ArrowLeft') { prevPage(); e.preventDefault(); }
  else if (e.key === 'ArrowDown' || e.key === 'Tab' && !e.shiftKey) {
    e.preventDefault();
    if (focusedFieldIdx < pageFieldKeys.length - 1) moveFocus(focusedFieldIdx + 1);
  }
  else if (e.key === 'ArrowUp' || (e.key === 'Tab' && e.shiftKey)) {
    e.preventDefault();
    if (focusedFieldIdx > 0) moveFocus(focusedFieldIdx - 1);
  }
  else if (e.key === 'Enter') { if (pageFieldKeys[focusedFieldIdx]) confirmField(pageFieldKeys[focusedFieldIdx]); }
  else if (e.key === 'f' || e.key === 'F') { if (pageFieldKeys[focusedFieldIdx]) flagField(pageFieldKeys[focusedFieldIdx]); }
  else if (e.key === 'n' || e.key === 'N') { if (pageFieldKeys[focusedFieldIdx]) toggleNoteInput(pageFieldKeys[focusedFieldIdx]); }
  else if (e.key === 'e' || e.key === 'E') {
    const vk = pageFieldKeys[focusedFieldIdx];
    if (vk) {
      const row = document.querySelector('[data-key="'+vk.replace(/"/g, '\\\\"')+'"]');
      const valEl = row ? row.querySelector('.field-val') : null;
      if (valEl) startEdit(vk, valEl.textContent);
    }
  }
});

// ─── New Client Modal ───
function openNewClientModal() {
  document.getElementById('newClientOverlay').classList.add('visible');
  document.getElementById('newClientName').value = '';
  document.getElementById('newClientEin').value = '';
  document.getElementById('newClientContact').value = '';
  document.getElementById('newClientNotes').value = '';
  setTimeout(() => document.getElementById('newClientName').focus(), 100);
}
function closeNewClientModal() {
  document.getElementById('newClientOverlay').classList.remove('visible');
}
function createNewClient() {
  const name = document.getElementById('newClientName').value.trim();
  if (!name) { showToast('Client name is required', 'error'); return; }
  const payload = {
    name: name,
    ein_last4: document.getElementById('newClientEin').value.trim(),
    contact: document.getElementById('newClientContact').value.trim(),
    notes: document.getElementById('newClientNotes').value.trim()
  };
  fetch('/api/clients/create', {
    method: 'POST', headers: {'Content-Type':'application/json'},
    body: JSON.stringify(payload)
  }).then(r => r.json()).then(d => {
    if (d.error) { showToast(d.error, 'error'); return; }
    showToast('Client "' + d.name + '" created', 'success');
    closeNewClientModal();
    loadClientSuggestions();
    // Auto-select the new client after dropdown refreshes
    setTimeout(() => { document.getElementById('clientName').value = d.name; }, 300);
    // Refresh clients list if on that section
    if (document.getElementById('sec-clients').classList.contains('active')) loadClients();
  }).catch(e => showToast('Failed: ' + e, 'error'));
}

// ─── Generate Report Modal ───
function openReportModal() {
  const el = document.getElementById('reportJobList');
  if (!clientCompletedJobs.length) { showToast('No completed jobs to report on', 'error'); return; }
  el.innerHTML = clientCompletedJobs.map(d =>
    '<label style="display:flex;align-items:center;gap:8px;padding:6px 0;font-size:13px;cursor:pointer">' +
    '<input type="checkbox" class="report-job-cb" value="'+esc(d.job_id)+'" checked> ' +
    esc(d.filename) + ' <span style="color:var(--text-light)">('+esc(d.year)+')</span></label>'
  ).join('');
  document.getElementById('reportOverlay').classList.add('visible');
}
function closeReportModal() {
  document.getElementById('reportOverlay').classList.remove('visible');
}
function generateReport() {
  const cbs = document.querySelectorAll('.report-job-cb:checked');
  const jobIds = Array.from(cbs).map(cb => cb.value);
  if (!jobIds.length) { showToast('Select at least one job', 'error'); return; }
  const fmt = document.getElementById('reportFormat').value;
  const year = document.getElementById('reportYear').value;
  const btn = document.querySelector('#reportOverlay .btn-primary');
  btn.disabled = true; btn.textContent = 'Generating...';
  fetch('/api/clients/'+encodeURIComponent(currentClientName)+'/generate-report', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({job_ids:jobIds, output_format:fmt, year:year})
  }).then(r=>r.json()).then(d => {
    btn.disabled = false; btn.textContent = 'Generate';
    if (d.error) { showToast(d.error, 'error'); return; }
    showToast('Report generated!', 'success');
    closeReportModal();
    window.open(d.download_url, '_blank');
  }).catch(e => { btn.disabled = false; btn.textContent = 'Generate'; showToast('Failed: '+e, 'error'); });
}

// ─── Delete Client Modal ───
function openDeleteClientModal() {
  document.getElementById('deleteClientTarget').textContent = currentClientName;
  document.getElementById('deleteClientConfirm').value = '';
  document.getElementById('deleteClientOverlay').classList.add('visible');
  setTimeout(() => document.getElementById('deleteClientConfirm').focus(), 100);
}
function closeDeleteClientModal() {
  document.getElementById('deleteClientOverlay').classList.remove('visible');
}
function confirmDeleteClient() {
  const val = document.getElementById('deleteClientConfirm').value.trim().toLowerCase();
  if (val !== 'delete') { showToast('Type "delete" to confirm', 'error'); return; }
  const btn = document.getElementById('deleteClientBtn');
  btn.disabled = true; btn.textContent = 'Deleting...';
  fetch('/api/clients/' + encodeURIComponent(currentClientName), {
    method: 'DELETE',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({confirm: 'delete'})
  }).then(r => r.json()).then(d => {
    btn.disabled = false; btn.textContent = 'Delete Client';
    if (d.error) { showToast(d.error, 'error'); return; }
    showToast('Client "' + currentClientName + '" deleted', 'success');
    closeDeleteClientModal();
    closeClientDetail();
    loadClients();
    loadClientSuggestions();
  }).catch(e => { btn.disabled = false; btn.textContent = 'Delete Client'; showToast('Failed: ' + e, 'error'); });
}

// ─── Merge Client Modal ───
function openMergeClientModal() {
  document.getElementById('mergeSourceName').textContent = currentClientName;
  const sel = document.getElementById('mergeTargetSelect');
  sel.innerHTML = '<option value="">-- Select target client --</option>';
  allClientsData.filter(c => c.name !== currentClientName).forEach(c => {
    sel.innerHTML += '<option value="' + esc(c.name) + '">' + esc(c.name) + (c.ein_last4 ? ' (' + esc(c.ein_last4) + ')' : '') + '</option>';
  });
  document.getElementById('mergeClientOverlay').classList.add('visible');
}
function closeMergeClientModal() {
  document.getElementById('mergeClientOverlay').classList.remove('visible');
}
function confirmMergeClient() {
  const target = document.getElementById('mergeTargetSelect').value;
  if (!target) { showToast('Select a target client', 'error'); return; }
  if (!confirm('Merge "' + currentClientName + '" into "' + target + '"? The source folder will be deleted. This cannot be undone.')) return;
  const btn = document.getElementById('mergeClientBtn');
  btn.disabled = true; btn.textContent = 'Merging...';
  fetch('/api/clients/merge', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({source: currentClientName, target: target})
  }).then(r => r.json()).then(d => {
    btn.disabled = false; btn.textContent = 'Merge';
    if (d.error) { showToast(d.error, 'error'); return; }
    showToast('Merged "' + currentClientName + '" into "' + target + '"', 'success');
    closeMergeClientModal();
    closeClientDetail();
    loadClients();
    loadClientSuggestions();
  }).catch(e => { btn.disabled = false; btn.textContent = 'Merge'; showToast('Failed: ' + e, 'error'); });
}

// ═══ GUIDED REVIEW ═══

var guidedQueue = [];
var guidedIdx = 0;
var guidedCurrentItem = null;
var guidedHeartbeatTimer = null;
var guidedJobId = null;
var guidedHistory = [];  // Stack of previously viewed field_ids for Back navigation
var guidedReviewer = '';  // Will be populated from /api/me
// B9: Time tracking state
var _fieldLoadTime = null;      // Timestamp when current field was rendered
var _reviewSessionId = null;    // Current review session ID from backend
var _reviewSessionStart = null; // Timestamp when session started
var _fieldsReviewedCount = 0;   // Count of fields reviewed in this session

// Auto-populate reviewer from session on page load
(function() {
  fetch('/api/me').then(function(r) { return r.json(); }).then(function(u) {
    guidedReviewer = u.display_name || u.username || '';
    var badge = document.getElementById('guidedReviewerBadge');
    if (badge && guidedReviewer) badge.textContent = 'Reviewing as: ' + guidedReviewer;
    // Also fill legacy initials field if empty
    var ini = document.getElementById('reviewerInitials');
    if (ini && !ini.value && u.initials) ini.value = u.initials;
  }).catch(function() {});
})();

function _getGuidedReviewer() {
  return guidedReviewer || getReviewer() || 'operator';
}

function openGuidedReview() {
  guidedJobId = currentJobId;
  if (!guidedJobId) { showToast('No job loaded', 'error'); return; }
  showSection('guided-review');
  guidedHistory = [];
  _fieldLoadTime = null;
  _fieldsReviewedCount = 0;
  document.getElementById('guidedBackBtn').disabled = true;
  document.getElementById('guidedDetail').style.opacity = '0.5';
  document.getElementById('guidedEvidence').innerHTML = '<div class="empty-state" style="color:rgba(255,255,255,0.5)"><p>Loading queue...</p></div>';

  // B1: Show stage badge
  _updateStageBadge();

  // B9: Start review session for time tracking
  _reviewSessionStart = Date.now();
  fetch('/api/review-session/' + guidedJobId + '/start', { method: 'POST',
    headers: {'Content-Type': 'application/json'}, body: '{}' })
    .then(function(r) { return r.json(); })
    .then(function(d) { if (d.session_id) _reviewSessionId = d.session_id; })
    .catch(function() {});

  fetch('/api/guided-review/queue/' + guidedJobId)
    .then(function(r) { return r.json(); })
    .then(function(data) {
      guidedQueue = data.queue || [];
      guidedIdx = 0;
      updateGuidedProgress(data.reviewed || 0, data.total || 0);
      if (guidedQueue.length === 0) {
        showGuidedComplete(data.reviewed || 0, data.total || 0);
      } else {
        loadGuidedItem();
      }
    })
    .catch(function(e) { showToast('Failed to load queue: ' + e, 'error'); });
}

function _updateStageBadge() {
  var badge = document.getElementById('guidedStageBadge');
  if (!badge || !guidedJobId) return;
  fetch('/api/jobs/' + guidedJobId + '/stage').then(r => r.json()).then(function(info) {
    var colors = {'preparer_review':'#3B82F6','reviewer_review':'#F59E0B','partner_review':'#8B5CF6','final':'#10B981','draft':'#94A3B8'};
    badge.textContent = info.display || info.stage;
    badge.style.background = colors[info.stage] || '#94A3B8';
    badge.style.display = '';
    if (!info.can_act && info.stage !== 'final') {
      badge.title = 'Read-only: your role cannot act at this stage';
    } else {
      badge.title = '';
    }
  }).catch(function() { badge.style.display = 'none'; });
}

function updateGuidedProgress(reviewed, total) {
  var pct = total > 0 ? Math.round(reviewed / total * 100) : 0;
  document.getElementById('guidedProgressText').textContent = reviewed + ' of ' + total + ' reviewed';
  document.getElementById('guidedProgressBar').style.width = pct + '%';
}

function loadGuidedItem(skipHistory) {
  if (guidedIdx >= guidedQueue.length) {
    // Reload queue to check for remaining items
    fetch('/api/guided-review/queue/' + guidedJobId)
      .then(function(r) { return r.json(); })
      .then(function(data) {
        guidedQueue = data.queue || [];
        guidedIdx = 0;
        updateGuidedProgress(data.reviewed || 0, data.total || 0);
        if (guidedQueue.length === 0) {
          showGuidedComplete(data.reviewed || 0, data.total || 0);
        } else {
          loadGuidedItem();
        }
      });
    return;
  }

  var item = guidedQueue[guidedIdx];
  // Push to history for Back navigation (unless we're going back)
  if (!skipHistory && guidedCurrentItem) {
    guidedHistory.push(guidedCurrentItem.field_id);
    document.getElementById('guidedBackBtn').disabled = false;
  }

  document.getElementById('guidedDetail').style.opacity = '0.5';
  document.getElementById('guidedEditArea').style.display = 'none';
  document.getElementById('guidedBtns').style.display = '';
  // Clear note field
  document.getElementById('guidedNoteInput').value = '';

  fetch('/api/guided-review/item/' + guidedJobId + '/' + encodeURIComponent(item.field_id))
    .then(function(r) { return r.json(); })
    .then(function(data) {
      if (data.error) { showToast(data.error, 'error'); return; }
      guidedCurrentItem = data;
      renderGuidedItem(data);
      // Prefetch next item's evidence (fire-and-forget)
      if (guidedIdx + 1 < guidedQueue.length) {
        var nextItem = guidedQueue[guidedIdx + 1];
        fetch('/api/guided-review/item/' + guidedJobId + '/' + encodeURIComponent(nextItem.field_id))
          .catch(function() {});
      }
      // Acquire lock
      var reviewer = _getGuidedReviewer();
      if (reviewer) {
        fetch('/api/guided-review/lock/' + guidedJobId + '/' + encodeURIComponent(item.field_id), {
          method: 'POST',
          headers: {'Content-Type': 'application/json'},
          body: JSON.stringify({ reviewer: reviewer })
        }).then(function(r) { return r.json(); }).then(function(lockData) {
          if (lockData.error) {
            document.getElementById('guidedLockBanner').textContent = lockData.error;
            document.getElementById('guidedLockBanner').style.display = '';
          } else {
            document.getElementById('guidedLockBanner').style.display = 'none';
          }
        });
      }
      // Heartbeat
      if (guidedHeartbeatTimer) clearInterval(guidedHeartbeatTimer);
      guidedHeartbeatTimer = setInterval(function() {
        if (reviewer && guidedCurrentItem) {
          fetch('/api/guided-review/lock/' + guidedJobId + '/' + encodeURIComponent(guidedCurrentItem.field_id), {
            method: 'POST',
            headers: {'Content-Type': 'application/json'},
            body: JSON.stringify({ reviewer: reviewer })
          });
        }
      }, 120000);
    })
    .catch(function(e) { showToast('Failed to load item: ' + e, 'error'); });
}

function guidedGoBack() {
  if (guidedHistory.length === 0) return;
  var prevFieldId = guidedHistory.pop();
  if (guidedHistory.length === 0) document.getElementById('guidedBackBtn').disabled = true;
  // Find the field in current queue, or fetch it directly
  var foundIdx = -1;
  for (var i = 0; i < guidedQueue.length; i++) {
    if (guidedQueue[i].field_id === prevFieldId) { foundIdx = i; break; }
  }
  if (foundIdx >= 0) {
    guidedIdx = foundIdx;
    loadGuidedItem(true);
  } else {
    // Field was already confirmed/removed from queue — load directly
    document.getElementById('guidedDetail').style.opacity = '0.5';
    document.getElementById('guidedEditArea').style.display = 'none';
    document.getElementById('guidedBtns').style.display = '';
    document.getElementById('guidedNoteInput').value = '';
    fetch('/api/guided-review/item/' + guidedJobId + '/' + encodeURIComponent(prevFieldId))
      .then(function(r) { return r.json(); })
      .then(function(data) {
        if (data.error) { showToast(data.error, 'error'); return; }
        guidedCurrentItem = data;
        renderGuidedItem(data);
      });
  }
}

function renderGuidedItem(data) {
  document.getElementById('guidedDetail').style.opacity = '1';
  // B9: Record when this field was rendered for per-field timing
  _fieldLoadTime = Date.now();
  // Evidence image — cropped highlight preferred, full page with banner as fallback
  var evEl = document.getElementById('guidedEvidence');
  if (data.evidence_url) {
    evEl.innerHTML = '<img src="' + esc(data.evidence_url) + '" alt="Evidence for ' + esc(data.field_name) + '" style="max-width:100%;height:auto">';
  } else if (data.page_url) {
    evEl.innerHTML = '<div style="background:#FFF3CD;color:#856404;padding:6px 12px;font-size:12px;text-align:center;border-radius:4px;margin-bottom:4px">' +
      '&#x26A0; Exact location uncertain — showing full page</div>' +
      '<img src="' + esc(data.page_url) + '" alt="Page ' + data.page_num + '" style="max-width:100%;height:auto">';
  }
  // Field label + destination
  document.getElementById('guidedLabel').textContent = data.display_name || data.field_name;
  document.getElementById('guidedDest').textContent =
    (data.document_type || '') + (data.entity ? ' \u2022 ' + data.entity : '') + ' \u2022 Page ' + data.page_num;
  // Big value
  var displayVal = data.value;
  if (typeof data.value === 'number') {
    displayVal = data.value.toLocaleString('en-US', { minimumFractionDigits: 2, maximumFractionDigits: 2 });
  } else {
    displayVal = String(data.value || '(empty)');
  }
  document.getElementById('guidedValue').textContent = displayVal;
  // Meta badges
  var confClass = 'badge-gray';
  var conf = String(data.confidence || '');
  if (conf.includes('dual') || conf.includes('verified_confirmed') || conf === 'auto_verified') confClass = 'badge-green';
  else if (conf === 'high' || conf === 'ocr_accepted') confClass = 'badge-blue';
  else if (conf === 'medium') confClass = 'badge-yellow';
  else if (conf === 'low' || conf === 'needs_review') confClass = 'badge-red';
  var metaHtml = '<span class="badge ' + confClass + '">' + esc(conf || 'unknown') + '</span>';
  metaHtml += ' <span class="badge badge-gray">' + esc(data.method || '') + '</span>';
  if (!data.evidence_available) {
    metaHtml += ' <span class="badge badge-yellow">location uncertain</span>';
  }
  document.getElementById('guidedMeta').innerHTML = metaHtml;
}

function guidedAction(action) {
  _perf('guidedAction');
  if (!guidedCurrentItem || !guidedJobId) return;
  var noteVal = (document.getElementById('guidedNoteInput').value || '').trim();
  var body = { action: action, reviewer: _getGuidedReviewer(), note: noteVal };
  // B9: Per-field time tracking — compute how long user spent on this field
  if (_fieldLoadTime && action !== 'skip') {
    body.field_duration_ms = Date.now() - _fieldLoadTime;
  }
  _fieldLoadTime = null;
  if (action !== 'skip') _fieldsReviewedCount++;
  if (action === 'correct') {
    body.corrected_value = document.getElementById('guidedEditInput').value.trim();
    if (!body.corrected_value) { showToast('Enter a corrected value', 'error'); return; }
  }
  // Disable buttons during save
  var btns = document.querySelectorAll('#guidedBtns .btn');
  btns.forEach(function(b) { b.disabled = true; });

  fetch('/api/guided-review/action/' + guidedJobId + '/' + encodeURIComponent(guidedCurrentItem.field_id), {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify(body)
  }).then(function(r) { return r.json(); })
    .then(function(data) {
      btns.forEach(function(b) { b.disabled = false; });
      if (!data.ok) { showToast(data.error || 'Action failed', 'error'); return; }
      var msg = action === 'confirm' ? 'Confirmed' : action === 'correct' ? 'Corrected' : action === 'not_present' ? 'Marked not present' : 'Skipped';
      showToast(msg, action === 'skip' ? 'info' : 'success');
      _perfEnd('guidedAction');
      updateGuidedProgress(data.reviewed || 0, data.total || 0);
      if (action === 'skip') {
        guidedIdx++;
      } else {
        guidedQueue.splice(guidedIdx, 1);
      }
      if (guidedQueue.length === 0 || guidedIdx >= guidedQueue.length) {
        // Refetch queue for any remaining
        fetch('/api/guided-review/queue/' + guidedJobId)
          .then(function(r) { return r.json(); })
          .then(function(qdata) {
            guidedQueue = qdata.queue || [];
            guidedIdx = 0;
            updateGuidedProgress(qdata.reviewed || 0, qdata.total || 0);
            if (guidedQueue.length === 0) {
              showGuidedComplete(qdata.reviewed || 0, qdata.total || 0);
            } else {
              loadGuidedItem();
            }
          });
      } else {
        loadGuidedItem();
      }
    })
    .catch(function(e) {
      btns.forEach(function(b) { b.disabled = false; });
      showToast('Error: ' + e, 'error');
    });
}

function guidedStartEdit() {
  if (!guidedCurrentItem) return;
  document.getElementById('guidedEditInput').value = String(guidedCurrentItem.value || '');
  document.getElementById('guidedEditArea').style.display = '';
  document.getElementById('guidedBtns').style.display = 'none';
  document.getElementById('guidedEditInput').focus();
  document.getElementById('guidedEditInput').select();
}

function guidedFinishEdit() {
  guidedAction('correct');
}

function guidedCancelEdit() {
  document.getElementById('guidedEditArea').style.display = 'none';
  document.getElementById('guidedBtns').style.display = '';
}

function showGuidedComplete(reviewed, total) {
  if (guidedHeartbeatTimer) { clearInterval(guidedHeartbeatTimer); guidedHeartbeatTimer = null; }
  guidedCurrentItem = null;

  // B9: End review session — compute elapsed time
  var sessionElapsed = _reviewSessionStart ? Math.round((Date.now() - _reviewSessionStart) / 1000) : null;
  if (_reviewSessionId && guidedJobId) {
    fetch('/api/review-session/' + guidedJobId + '/end', {
      method: 'POST', headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({ session_id: _reviewSessionId, fields_reviewed: _fieldsReviewedCount })
    }).catch(function() {});
  }
  _reviewSessionId = null;
  _reviewSessionStart = null;

  // Fetch current stage info to decide what buttons to show
  fetch('/api/jobs/' + guidedJobId + '/stage').then(r => r.json()).then(function(stageInfo) {
    _renderGuidedComplete(reviewed, total, stageInfo, sessionElapsed);
  }).catch(function() {
    _renderGuidedComplete(reviewed, total, { stage: 'preparer_review', can_act: true, can_submit: true, can_send_back: false, display: 'Preparer Review' }, sessionElapsed);
  });
}

function _renderGuidedComplete(reviewed, total, stageInfo, sessionElapsed) {
  var stage = stageInfo.stage || 'preparer_review';
  var canAct = stageInfo.can_act;
  var stageDisplay = stageInfo.display || stage;

  var stageColors = {
    'preparer_review': '#3B82F6',
    'reviewer_review': '#F59E0B',
    'partner_review': '#8B5CF6',
    'final': '#10B981',
    'draft': '#94A3B8'
  };
  var stageColor = stageColors[stage] || '#94A3B8';

  // Left panel — completion summary with stage badge + B9 session timing
  var timingHtml = '';
  if (sessionElapsed !== null && sessionElapsed !== undefined) {
    timingHtml = '<div style="margin-top:16px;padding:12px 20px;background:rgba(255,255,255,0.1);border-radius:10px;font-size:14px">' +
      '<div style="font-size:11px;text-transform:uppercase;letter-spacing:1px;opacity:0.7;margin-bottom:4px">Session Time</div>' +
      '<div style="font-size:24px;font-weight:700;font-family:var(--mono)">' + formatDuration(sessionElapsed) + '</div>' +
      (_fieldsReviewedCount > 0 ? '<div style="font-size:12px;opacity:0.7;margin-top:4px">' +
        Math.round(sessionElapsed / _fieldsReviewedCount) + 's avg per field</div>' : '') +
      '</div>';
  }
  document.getElementById('guidedEvidence').innerHTML =
    '<div style="display:flex;flex-direction:column;align-items:center;justify-content:center;height:100%;color:rgba(255,255,255,0.8);text-align:center;padding:40px">' +
    '<div style="font-size:64px;margin-bottom:16px">&#x2714;</div>' +
    '<h2 style="color:#fff;margin:0 0 8px">Review Complete</h2>' +
    '<p style="font-size:18px;margin:0 0 16px">' + reviewed + ' of ' + total + ' fields reviewed</p>' +
    '<div style="background:' + stageColor + ';color:#fff;padding:6px 16px;border-radius:16px;font-size:14px;font-weight:600">' + esc(stageDisplay) + '</div>' +
    timingHtml +
    '</div>';

  // Right panel — stage-dependent actions
  var html = '<div class="guided-complete" style="padding:32px">';

  if (stage === 'final') {
    // Already finalized — just show download buttons
    html += '<h2 style="margin:0 0 8px">Review Finalized</h2>' +
      '<p style="color:#10B981;margin:0 0 24px">This document has been finalized through all review stages.</p>' +
      '<div style="display:flex;flex-wrap:wrap;gap:8px">' +
      '<button class="btn btn-success" onclick="window.open(\'/api/download/' + guidedJobId + '\',\'_blank\')">&#x2B73; Download Excel</button>' +
      '<button class="btn btn-secondary" onclick="window.open(\'/api/download-log/' + guidedJobId + '\',\'_blank\')">&#x2B73; Download JSON</button>' +
      '<button class="btn btn-ghost" onclick="showSection(\'inbox\')">&#x1F4E5; Back to Inbox</button>' +
      '</div>';
  } else if (stage === 'preparer_review') {
    html += '<h2 style="margin:0 0 8px">Finish Preparer Review</h2>' +
      '<p style="color:var(--text-muted);margin:0 0 24px">All fields reviewed. Submit to the next reviewer, or generate reports.</p>';
    // Report format selection
    html += _reportFormatCheckboxes();
    html += '<div style="display:flex;flex-wrap:wrap;gap:8px;margin-top:16px">' +
      '<button class="btn btn-primary" onclick="submitToNextStage()" style="font-size:14px;padding:10px 20px">&#x27A1; Submit to Reviewer</button>' +
      '<button class="btn btn-secondary" onclick="finishReviewGenerate()" style="font-size:14px;padding:10px 20px">&#x1F4C4; Generate Reports</button>' +
      '<button class="btn btn-ghost" onclick="openGuidedReview()">&#x21BB; Re-audit</button>' +
      '<button class="btn btn-ghost" onclick="openGridReview()">&#x2630; List View</button>' +
      '</div>';
  } else if (stage === 'reviewer_review') {
    html += '<h2 style="margin:0 0 8px">Finish Reviewer Review</h2>' +
      '<p style="color:var(--text-muted);margin:0 0 24px">All fields reviewed. Approve to Partner, send back to Preparer, or generate reports.</p>';
    html += _reportFormatCheckboxes();
    html += '<div style="display:flex;flex-wrap:wrap;gap:8px;margin-top:16px">' +
      '<button class="btn btn-primary" onclick="submitToNextStage()" style="font-size:14px;padding:10px 20px">&#x2705; Approve &rarr; Partner</button>' +
      '<button class="btn btn-warning" onclick="sendBackToPrev()" style="font-size:14px;padding:10px 20px;background:#F59E0B;color:#fff;border:none;border-radius:6px;cursor:pointer">&#x21A9; Send Back to Preparer</button>' +
      '<button class="btn btn-secondary" onclick="finishReviewGenerate()">&#x1F4C4; Generate Reports</button>' +
      '<button class="btn btn-ghost" onclick="openGuidedReview()">&#x21BB; Re-audit</button>' +
      '</div>';
  } else if (stage === 'partner_review') {
    html += '<h2 style="margin:0 0 8px">Partner Final Review</h2>' +
      '<p style="color:var(--text-muted);margin:0 0 24px">All fields reviewed. Finalize to complete the review chain, or send back.</p>';
    html += _reportFormatCheckboxes();
    html += '<div style="display:flex;flex-wrap:wrap;gap:8px;margin-top:16px">' +
      '<button class="btn btn-success" onclick="submitToNextStage()" style="font-size:14px;padding:10px 20px">&#x1F3C6; Finalize</button>' +
      '<button class="btn btn-warning" onclick="sendBackToPrev()" style="font-size:14px;padding:10px 20px;background:#F59E0B;color:#fff;border:none;border-radius:6px;cursor:pointer">&#x21A9; Send Back to Reviewer</button>' +
      '<button class="btn btn-secondary" onclick="finishReviewGenerate()">&#x1F4C4; Generate Reports</button>' +
      '<button class="btn btn-ghost" onclick="openGuidedReview()">&#x21BB; Re-audit</button>' +
      '</div>';
  } else {
    // draft or unknown — generic finish
    html += '<h2 style="margin:0 0 8px">Finish Review</h2>' +
      '<p style="color:var(--text-muted);margin:0 0 24px">All fields reviewed.</p>';
    html += _reportFormatCheckboxes();
    html += '<div style="display:flex;flex-wrap:wrap;gap:8px;margin-top:16px">' +
      '<button class="btn btn-primary" onclick="finishReviewGenerate()" style="font-size:14px;padding:10px 20px">&#x1F4C4; Generate Selected Reports</button>' +
      '<button class="btn btn-ghost" onclick="openGuidedReview()">&#x21BB; Re-audit</button>' +
      '<button class="btn btn-ghost" onclick="openGridReview()">&#x2630; List View</button>' +
      '</div>';
  }

  html += '</div>';
  document.getElementById('guidedDetail').innerHTML = html;
  document.getElementById('guidedDetail').style.opacity = '1';
}

function _reportFormatCheckboxes() {
  return '<div style="margin-bottom:8px">' +
    '<h3 style="font-size:14px;margin:0 0 12px;color:var(--navy)">Generate Reports</h3>' +
    '<p style="font-size:12px;color:var(--text-muted);margin:0 0 12px">Select report formats to generate from verified data:</p>' +
    '<label style="display:flex;align-items:center;gap:8px;margin-bottom:8px;font-size:14px;cursor:pointer">' +
    '<input type="checkbox" id="frFmtTaxReview" checked> Tax Review Worksheet</label>' +
    '<label style="display:flex;align-items:center;gap:8px;margin-bottom:8px;font-size:14px;cursor:pointer">' +
    '<input type="checkbox" id="frFmtJournal"> Journal Entries <span style="font-size:11px;color:var(--text-muted)">(bank/payroll only)</span></label>' +
    '<label style="display:flex;align-items:center;gap:8px;margin-bottom:8px;font-size:14px;cursor:pointer">' +
    '<input type="checkbox" id="frFmtAcctBal"> Account Balances <span style="font-size:11px;color:var(--text-muted)">(bank/payroll only)</span></label>' +
    '<label style="display:flex;align-items:center;gap:8px;margin-bottom:8px;font-size:14px;cursor:pointer">' +
    '<input type="checkbox" id="frFmtTrialBal"> Trial Balance <span style="font-size:11px;color:var(--text-muted)">(bank/payroll only)</span></label>' +
    '<label style="display:flex;align-items:center;gap:8px;margin-bottom:8px;font-size:14px;cursor:pointer">' +
    '<input type="checkbox" id="frFmtTxnReg"> Transaction Register <span style="font-size:11px;color:var(--text-muted)">(bank/payroll only)</span></label>' +
    '</div>';
}

function submitToNextStage() {
  if (!guidedJobId) { showToast('No job loaded', 'error'); return; }
  fetch('/api/jobs/' + guidedJobId + '/submit-review', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({})
  }).then(r => r.json()).then(function(data) {
    if (data.error) { showToast(data.error, 'error'); return; }
    showToast('Submitted! Moved to: ' + data.display, 'success');
    refreshInboxBadge();
    // If moving to final, chain into report generation + audit
    if (data.new_stage === 'final') {
      finishReviewGenerate();
    } else {
      // Go back to inbox
      setTimeout(function() { showSection('inbox'); }, 1200);
    }
  }).catch(function(e) { showToast('Submit failed: ' + e, 'error'); });
}

function sendBackToPrev() {
  if (!guidedJobId) { showToast('No job loaded', 'error'); return; }
  var reason = prompt('Reason for sending back:');
  if (!reason || !reason.trim()) { showToast('A reason is required', 'error'); return; }
  fetch('/api/jobs/' + guidedJobId + '/send-back', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({ reason: reason.trim() })
  }).then(r => r.json()).then(function(data) {
    if (data.error) { showToast(data.error, 'error'); return; }
    showToast('Sent back to: ' + data.display, 'success');
    refreshInboxBadge();
    setTimeout(function() { showSection('inbox'); }, 1200);
  }).catch(function(e) { showToast('Send back failed: ' + e, 'error'); });
}

function finishReviewGenerate() {
  var jobId = guidedJobId || currentJobId;
  if (!jobId) { showToast('No job loaded', 'error'); return; }
  var formats = [];
  // Checkboxes may not exist in grid view — default to tax_review
  var cb = document.getElementById('frFmtTaxReview');
  if (cb) {
    if (cb.checked) formats.push('tax_review');
    if (document.getElementById('frFmtJournal') && document.getElementById('frFmtJournal').checked) formats.push('journal_entries');
    if (document.getElementById('frFmtAcctBal') && document.getElementById('frFmtAcctBal').checked) formats.push('account_balances');
    if (document.getElementById('frFmtTrialBal') && document.getElementById('frFmtTrialBal').checked) formats.push('trial_balance');
    if (document.getElementById('frFmtTxnReg') && document.getElementById('frFmtTxnReg').checked) formats.push('transaction_register');
  } else {
    // Grid view: default to tax_review
    formats.push('tax_review');
  }
  if (formats.length === 0) { showToast('Select at least one report format', 'error'); return; }

  showToast('Generating ' + formats.length + ' report(s)...', 'info');

  // Generate each selected format
  var completed = 0;
  var failed = 0;
  formats.forEach(function(fmt) {
    fetch('/api/regen-excel/' + jobId, {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({ output_format: fmt })
    }).then(function(r) { return r.json(); })
      .then(function(data) {
        completed++;
        if (data.error) { failed++; showToast('Failed: ' + fmt + ' \u2014 ' + data.error, 'error'); }
        if (completed === formats.length) {
          if (failed === 0) {
            showToast('\u2713 Reports generated!', 'success');
            // Open download in new tab
            window.open('/api/download/' + jobId, '_blank');
          }
          // Chain into post-run audit if in guided view
          if (guidedJobId) startPostRunAudit();
        }
      })
      .catch(function() { completed++; failed++; });
  });
}

// ─── Post-Run Audit UI ───────────────────────────────────────────────────
var auditSample = [];
var auditIdx = 0;
var auditResults = [];

function startPostRunAudit() {
  if (!guidedJobId) return;
  fetch('/api/post-run-audit/sample/' + guidedJobId)
    .then(function(r) { return r.json(); })
    .then(function(data) {
      auditSample = data.sample || [];
      auditIdx = 0;
      auditResults = [];
      if (auditSample.length === 0) {
        showAuditComplete();
        return;
      }
      showAuditItem();
    })
    .catch(function(e) { showToast('Audit sample failed: ' + e, 'error'); });
}

function showAuditItem() {
  if (auditIdx >= auditSample.length) {
    submitAuditResults();
    return;
  }
  var item = auditSample[auditIdx];
  var pct = Math.round((auditIdx / auditSample.length) * 100);

  // Update left panel with evidence
  var evEl = document.getElementById('guidedEvidence');
  evEl.innerHTML = '<div style="display:flex;flex-direction:column;align-items:center;justify-content:center;height:100%;color:rgba(255,255,255,0.7);text-align:center;padding:20px">' +
    '<div style="font-size:13px;margin-bottom:12px;color:rgba(255,255,255,0.5)">POST-RUN AUDIT</div>' +
    '<div style="font-size:14px;margin-bottom:20px">' + (auditIdx + 1) + ' of ' + auditSample.length + '</div>' +
    '</div>';

  // Load full item detail with evidence
  fetch('/api/guided-review/item/' + guidedJobId + '/' + encodeURIComponent(item.field_id))
    .then(function(r) { return r.json(); })
    .then(function(detail) {
      // Show evidence in left panel
      if (detail.evidence_url) {
        evEl.innerHTML = '<div style="background:#E8F5E9;color:#2E7D32;padding:6px 12px;font-size:12px;text-align:center;border-radius:4px;margin-bottom:4px">' +
          '&#x1F50D; AUDIT CHECK ' + (auditIdx + 1) + '/' + auditSample.length + '</div>' +
          '<img src="' + esc(detail.evidence_url) + '" style="max-width:100%;height:auto">';
      } else if (detail.page_url) {
        evEl.innerHTML = '<div style="background:#FFF3CD;color:#856404;padding:6px 12px;font-size:12px;text-align:center;border-radius:4px;margin-bottom:4px">' +
          '&#x1F50D; AUDIT CHECK ' + (auditIdx + 1) + '/' + auditSample.length + ' — location uncertain</div>' +
          '<img src="' + esc(detail.page_url) + '" style="max-width:100%;height:auto">';
      }

      // Right panel — audit check card
      var displayName = detail.display_name || item.field_name;
      var displayVal = item.value;
      if (typeof displayVal === 'number') {
        displayVal = displayVal.toLocaleString('en-US', { minimumFractionDigits: 2, maximumFractionDigits: 2 });
      } else {
        displayVal = String(displayVal || '(empty)');
      }

      var statusBadge = '';
      if (item.status === 'confirmed') statusBadge = '<span class="badge badge-green">confirmed</span>';
      else if (item.status === 'corrected') statusBadge = '<span class="badge badge-yellow">corrected</span>';
      else statusBadge = '<span class="badge badge-gray">' + esc(item.status) + '</span>';

      var html = '<div style="padding:32px">' +
        '<div style="display:flex;align-items:center;gap:8px;margin-bottom:24px">' +
        '<span style="background:#E8F5E9;color:#2E7D32;padding:4px 10px;border-radius:12px;font-size:12px;font-weight:600">AUDIT CHECK</span>' +
        '<span style="font-size:12px;color:var(--text-muted)">' + (auditIdx + 1) + ' of ' + auditSample.length + '</span>' +
        '</div>' +
        '<div style="font-size:13px;color:var(--text-muted);margin-bottom:4px">' + esc(detail.document_type || '') + (detail.entity ? ' \u2022 ' + esc(detail.entity) : '') + ' \u2022 Page ' + item.page_num + '</div>' +
        '<div style="font-size:18px;font-weight:600;color:var(--navy);margin-bottom:8px">' + esc(displayName) + '</div>' +
        '<div style="font-size:36px;font-family:monospace;font-weight:700;color:var(--navy);margin-bottom:12px;word-break:break-all">' + esc(displayVal) + '</div>' +
        '<div style="margin-bottom:16px">' + statusBadge + '</div>' +
        '<div style="margin-bottom:12px"><input type="text" class="form-input" id="auditNoteInput" placeholder="Audit note (optional)..." style="width:100%;font-size:13px"></div>' +
        '<div style="display:flex;gap:8px">' +
        '<button class="btn btn-success" onclick="auditAction(\'pass\')" style="font-size:14px;padding:10px 24px">&#x2714; Confirm</button>' +
        '<button class="btn btn-danger" onclick="auditAction(\'flag\')" style="font-size:14px;padding:10px 24px">&#x1F6A9; Flag</button>' +
        '</div>' +
        '<div style="margin-top:16px">' +
        '<div style="background:var(--border);border-radius:4px;height:6px;overflow:hidden">' +
        '<div style="background:#4CAF50;height:100%;width:' + pct + '%;transition:width 0.3s"></div>' +
        '</div></div>' +
        '</div>';

      document.getElementById('guidedDetail').innerHTML = html;
      document.getElementById('guidedDetail').style.opacity = '1';
    })
    .catch(function() {
      // Fallback without evidence
      showToast('Could not load evidence for audit item', 'info');
      auditIdx++;
      showAuditItem();
    });
}

function auditAction(outcome) {
  var item = auditSample[auditIdx];
  var note = (document.getElementById('auditNoteInput') || {}).value || '';
  auditResults.push({
    field_id: item.field_id,
    field_name: item.field_name,
    outcome: outcome,
    note: note.trim(),
  });
  auditIdx++;
  if (outcome === 'flag') {
    showToast('Flagged — will route back for re-review', 'error');
  } else {
    showToast('Confirmed', 'success');
  }
  showAuditItem();
}

function submitAuditResults() {
  var passCount = auditResults.filter(function(r) { return r.outcome === 'pass'; }).length;
  var failCount = auditResults.filter(function(r) { return r.outcome === 'flag'; }).length;

  fetch('/api/post-run-audit/result/' + guidedJobId, {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({
      sample_size: auditResults.length,
      pass_count: passCount,
      fail_count: failCount,
      results: auditResults,
      reviewer: _getGuidedReviewer(),
    })
  }).then(function(r) { return r.json(); })
    .then(function(data) {
      showAuditComplete(passCount, failCount, data.flagged_field_ids || []);
    })
    .catch(function() { showAuditComplete(passCount, failCount, []); });
}

function showAuditComplete(passCount, failCount, flaggedIds) {
  passCount = passCount || 0;
  failCount = failCount || 0;

  var badgeColor = failCount === 0 ? '#4CAF50' : '#E53935';
  var badgeText = failCount === 0
    ? 'Audit Passed: ' + passCount + '/' + (passCount + failCount)
    : 'Audit Flagged: ' + failCount + ' issue' + (failCount !== 1 ? 's' : '');

  // Left panel
  document.getElementById('guidedEvidence').innerHTML =
    '<div style="display:flex;flex-direction:column;align-items:center;justify-content:center;height:100%;color:rgba(255,255,255,0.8);text-align:center;padding:40px">' +
    '<div style="font-size:64px;margin-bottom:16px">' + (failCount === 0 ? '&#x2705;' : '&#x26A0;') + '</div>' +
    '<h2 style="color:#fff;margin:0 0 8px">Audit ' + (failCount === 0 ? 'Passed' : 'Complete') + '</h2>' +
    '<div style="background:' + badgeColor + ';color:#fff;padding:8px 20px;border-radius:20px;font-size:16px;font-weight:600;margin-top:8px">' + badgeText + '</div>' +
    '</div>';

  // Right panel
  var html = '<div style="padding:32px">' +
    '<h2 style="margin:0 0 8px">Post-Run Audit Complete</h2>';

  if (failCount > 0) {
    html += '<p style="color:#E53935;margin:0 0 16px">' + failCount + ' field(s) flagged for re-review. They have been routed back into the review queue.</p>';
    html += '<button class="btn btn-primary" onclick="openGuidedReview()" style="margin-bottom:12px">&#x21BB; Re-open Review Queue</button><br>';
  } else {
    html += '<p style="color:#4CAF50;margin:0 0 16px">All sampled values confirmed. Review is complete.</p>';
  }

  html += '<div style="display:flex;flex-wrap:wrap;gap:8px;margin-top:16px">' +
    '<button class="btn btn-success" onclick="window.open(\'/api/download/' + guidedJobId + '\',\'_blank\')">&#x2B73; Download Excel</button>' +
    '<button class="btn btn-secondary" onclick="window.open(\'/api/download-log/' + guidedJobId + '\',\'_blank\')">&#x2B73; Download JSON</button>' +
    '<button class="btn btn-ghost" onclick="openGridReview()">&#x2630; List View</button>' +
    '<button class="btn btn-ghost" onclick="showSection(\'history\')">&#x1F4CB; History</button>' +
    '</div></div>';

  document.getElementById('guidedDetail').innerHTML = html;
  document.getElementById('guidedDetail').style.opacity = '1';
}

// ─── Lite Findings Panel (3C-1.2) ───────────────────────────────────────────
function renderLiteFindings() {
  const panel = document.getElementById('liteFindingsPanel');
  if (!panel) return;

  // Data-level guard: only show when ardent_summary exists
  if (!reviewData || !reviewData.ardent_summary) {
    panel.style.display = 'none';
    return;
  }

  const s = reviewData.ardent_summary;
  panel.style.display = '';

  // Status badge
  const badge = document.getElementById('liteFindingsStatusBadge');
  if (s.blocked) {
    badge.textContent = 'BLOCKED';
    badge.style.background = 'var(--red)';
  } else if (s.needs_review) {
    badge.textContent = 'REVIEW REQUIRED';
    badge.style.background = 'var(--yellow)';
    badge.style.color = '#7C5800';
  } else {
    badge.textContent = 'OK';
    badge.style.background = 'var(--green)';
  }

  // Counts summary
  const counts = document.getElementById('liteFindingsCounts');
  counts.textContent = (s.findings || []).length + ' finding' + ((s.findings || []).length !== 1 ? 's' : '');

  // Severity pills
  const sevRow = document.getElementById('liteFindingsSeverityRow');
  const pills = [];
  if (s.critical_count) pills.push('<span style="background:var(--red-bg);color:var(--red);padding:2px 8px;border-radius:8px;font-size:11px;font-weight:600">Critical: ' + s.critical_count + '</span>');
  if (s.error_count) pills.push('<span style="background:var(--red-bg);color:var(--red);padding:2px 8px;border-radius:8px;font-size:11px;font-weight:600">Error: ' + s.error_count + '</span>');
  if (s.warning_count) pills.push('<span style="background:var(--yellow-bg);color:#B7791F;padding:2px 8px;border-radius:8px;font-size:11px;font-weight:600">Warning: ' + s.warning_count + '</span>');
  if (s.info_count) pills.push('<span style="background:#EBF5FB;color:var(--accent);padding:2px 8px;border-radius:8px;font-size:11px;font-weight:600">Info: ' + s.info_count + '</span>');
  if (s.verification_requests_count) pills.push('<span style="background:var(--purple-bg);color:var(--purple);padding:2px 8px;border-radius:8px;font-size:11px;font-weight:600">Verification Requests: ' + s.verification_requests_count + '</span>');
  sevRow.innerHTML = pills.join('');

  // Severity color map
  const sevColors = { critical: 'var(--red)', error: 'var(--red)', warning: 'var(--yellow)', info: 'var(--accent)' };
  const sevBgColors = { critical: 'var(--red-bg)', error: 'var(--red-bg)', warning: 'var(--yellow-bg)', info: '#EBF5FB' };

  // Findings list
  const list = document.getElementById('liteFindingsList');
  if (!s.findings || s.findings.length === 0) {
    list.innerHTML = '<div style="padding:12px;color:var(--text-light);font-size:12px;">No findings.</div>';
  } else {
    let html = '';
    s.findings.forEach(function(f) {
      const sevColor = sevColors[f.severity] || 'var(--text-secondary)';
      const sevBg = sevBgColors[f.severity] || '#F5F5F5';
      const passIcon = f.passed ? '<span style="color:var(--green)" title="Passed">&#x2713;</span>' : '<span style="color:' + sevColor + '" title="Failed">&#x2717;</span>';

      html += '<div style="padding:8px 0;border-bottom:1px solid var(--border-light);font-size:12px;">';
      html += '  <div style="display:flex;align-items:center;gap:8px;">';
      html += '    ' + passIcon;
      html += '    <span style="background:' + sevBg + ';color:' + sevColor + ';padding:1px 6px;border-radius:6px;font-size:10px;font-weight:700;text-transform:uppercase;">' + esc(f.severity) + '</span>';
      html += '    <span style="font-family:var(--mono);font-size:11px;color:var(--text-secondary);">' + esc(f.rule_id) + '</span>';
      html += '    <span style="color:var(--text);">' + esc(f.message || f.rule_name) + '</span>';
      html += '  </div>';

      // Evidence items
      if (f.evidence && f.evidence.length > 0) {
        html += '  <div style="margin:4px 0 0 24px;">';
        f.evidence.forEach(function(ev) {
          html += '<div style="font-size:11px;color:var(--text-secondary);padding:2px 0;">';
          if (ev.field) html += '<span style="font-family:var(--mono);">' + esc(ev.field) + '</span>: ';
          if (ev.extracted_value !== null && ev.extracted_value !== undefined) html += 'extracted=' + esc(String(ev.extracted_value));
          if (ev.expected_value !== null && ev.expected_value !== undefined) html += ', expected=' + esc(String(ev.expected_value));
          if (ev.expected_range) html += ' (' + esc(ev.expected_range) + ')';
          if (ev.detail) html += ' &mdash; ' + esc(ev.detail);
          html += '</div>';
        });
        html += '  </div>';
      }
      html += '</div>';
    });
    list.innerHTML = html;
  }

  // Provenance footer
  const prov = document.getElementById('liteFindingsProvenance');
  const parts = [];
  if (s.ruleset_version) parts.push('Ruleset v' + s.ruleset_version);
  if (s.evaluated_at) parts.push('Evaluated: ' + s.evaluated_at);
  if (s.deterministic_match_pct !== null && s.deterministic_match_pct !== undefined) parts.push('Match: ' + s.deterministic_match_pct.toFixed(1) + '%');
  if (s.schema_version) parts.push('Schema v' + s.schema_version);
  prov.textContent = parts.join(' \u2022 ');

  // ── Doctrine provenance + drift badge ──
  _renderDoctrineBadge();
}

function _renderDoctrineBadge() {
  // Find or create the doctrine provenance line in the Lite Findings header
  var header = document.querySelector('#liteFindingsPanel summary');
  if (!header) return;
  var existing = document.getElementById('doctrineBadge');
  if (existing) existing.remove();

  var dc = reviewData.doctrine_current;
  var drift = reviewData.doctrine_drift;
  if (!dc) return;

  var span = document.createElement('span');
  span.id = 'doctrineBadge';
  span.style.cssText = 'font-size:10px; font-family:var(--mono); color:var(--text-secondary); margin-left:8px; display:inline-flex; align-items:center; gap:4px;';
  span.textContent = 'Doctrine v' + dc.version + ' \u2022 ' + dc.hash_short;

  // Drift badge
  if (drift && drift.status !== 'ok' && drift.status !== 'legacy') {
    var badge = document.createElement('span');
    badge.style.cssText = 'background:var(--yellow);color:#7C5800;padding:1px 6px;border-radius:6px;font-size:9px;font-weight:700;cursor:help;';
    badge.textContent = 'DOCTRINE DRIFT';
    badge.title = drift.message || 'Governance rules have changed since this document was evaluated.';
    span.appendChild(badge);
  } else if (drift && drift.status === 'legacy') {
    var badge = document.createElement('span');
    badge.style.cssText = 'background:var(--border);color:var(--text-light);padding:1px 6px;border-radius:6px;font-size:9px;font-weight:600;cursor:help;';
    badge.textContent = 'PRE-DOCTRINE';
    badge.title = drift.message || 'This log was created before Doctrine governance was enabled.';
    span.appendChild(badge);
  }

  header.appendChild(span);
}

</script>

<!-- New Client Modal -->
<div class="modal-overlay" id="newClientOverlay">
  <div class="modal-content">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px">
      <h3 style="margin:0">New Client</h3>
      <button class="btn btn-ghost btn-sm" onclick="closeNewClientModal()">&times;</button>
    </div>
    <div class="form-group" style="margin-bottom:12px">
      <label class="form-label">Client Name <span style="color:var(--danger)">*</span></label>
      <input type="text" id="newClientName" class="form-input" placeholder="e.g. Watts, Stacy">
    </div>
    <div class="form-group" style="margin-bottom:12px">
      <label class="form-label">EIN / SSN (last 4)</label>
      <input type="text" id="newClientEin" class="form-input" placeholder="e.g. 1234" maxlength="4">
    </div>
    <div class="form-group" style="margin-bottom:12px">
      <label class="form-label">Contact</label>
      <input type="text" id="newClientContact" class="form-input" placeholder="e.g. email or phone">
    </div>
    <div class="form-group" style="margin-bottom:16px">
      <label class="form-label">Notes</label>
      <textarea id="newClientNotes" class="form-input" rows="2" placeholder="Optional notes about this client"></textarea>
    </div>
    <div style="display:flex;justify-content:flex-end;gap:8px">
      <button class="btn btn-ghost" onclick="closeNewClientModal()">Cancel</button>
      <button class="btn btn-primary" onclick="createNewClient()">Create Client</button>
    </div>
  </div>
</div>

<!-- Generate Report Modal -->
<div class="modal-overlay" id="reportOverlay">
  <div class="modal-content">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px">
      <h3 style="margin:0">Generate Report</h3>
      <button class="btn btn-ghost btn-sm" onclick="closeReportModal()">&times;</button>
    </div>
    <div class="form-group" style="margin-bottom:12px">
      <label class="form-label">Select Jobs</label>
      <div id="reportJobList" style="max-height:200px;overflow-y:auto;border:1px solid var(--border);border-radius:8px;padding:8px 12px"></div>
    </div>
    <div class="form-row" style="margin-bottom:16px">
      <div class="form-group">
        <label class="form-label">Output Format</label>
        <select id="reportFormat" class="form-input">
          <option value="tax_review">Tax Review</option>
          <option value="journal_entries">Journal Entries</option>
          <option value="account_balances">Account Balances</option>
          <option value="trial_balance">Trial Balance</option>
          <option value="transaction_register">Transaction Register</option>
        </select>
      </div>
      <div class="form-group">
        <label class="form-label">Year</label>
        <input type="number" id="reportYear" class="form-input" value="2025" min="2000" max="2030">
      </div>
    </div>
    <div style="display:flex;justify-content:flex-end;gap:8px">
      <button class="btn btn-ghost" onclick="closeReportModal()">Cancel</button>
      <button class="btn btn-primary" onclick="generateReport()">Generate</button>
    </div>
  </div>
</div>

<!-- Delete Client Modal -->
<div class="modal-overlay" id="deleteClientOverlay">
  <div class="modal-content">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px">
      <h3 style="margin:0;color:var(--red)">Delete Client</h3>
      <button class="btn btn-ghost btn-sm" onclick="closeDeleteClientModal()">&times;</button>
    </div>
    <p style="font-size:13px;margin-bottom:12px">This will permanently delete the client folder <strong id="deleteClientTarget"></strong> and all its contents (context documents, instructions, output files).</p>
    <p style="font-size:13px;margin-bottom:12px;color:var(--text-secondary)">Job history records will be preserved for audit purposes.</p>
    <div class="form-group" style="margin-bottom:16px">
      <label class="form-label">Type "delete" to confirm</label>
      <input type="text" id="deleteClientConfirm" class="form-input" placeholder="delete" autocomplete="off">
    </div>
    <div style="display:flex;justify-content:flex-end;gap:8px">
      <button class="btn btn-ghost" onclick="closeDeleteClientModal()">Cancel</button>
      <button class="btn btn-danger" id="deleteClientBtn" onclick="confirmDeleteClient()">Delete Client</button>
    </div>
  </div>
</div>

<!-- Merge Client Modal -->
<div class="modal-overlay" id="mergeClientOverlay">
  <div class="modal-content">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px">
      <h3 style="margin:0;color:var(--purple)">Merge Client</h3>
      <button class="btn btn-ghost btn-sm" onclick="closeMergeClientModal()">&times;</button>
    </div>
    <p style="font-size:13px;margin-bottom:12px">Merge all data from <strong id="mergeSourceName"></strong> into another client. This will move all documents, context, instructions, and update job history.</p>
    <div class="form-group" style="margin-bottom:12px">
      <label class="form-label">Merge Into (Target Client)</label>
      <select id="mergeTargetSelect" class="form-input form-select"></select>
    </div>
    <p style="font-size:12px;color:var(--red);margin-bottom:16px">&#x26A0; The source client folder will be deleted after merge. This cannot be undone.</p>
    <div style="display:flex;justify-content:flex-end;gap:8px">
      <button class="btn btn-ghost" onclick="closeMergeClientModal()">Cancel</button>
      <button class="btn btn-primary" id="mergeClientBtn" onclick="confirmMergeClient()" style="background:var(--purple)">Merge</button>
    </div>
  </div>
</div>

</body>
</html>"""


# ── Aftercare shutdown (T-UX-CONFIRM-FASTPATH) ──

def _drain_aftercare(timeout=5.0):
    """Drain aftercare queue on shutdown. Best-effort."""
    global _aftercare_running
    _aftercare_running = False
    _aftercare_event.set()
    _aftercare_thread.join(timeout=timeout)
    if _aftercare_queue:
        print(f"  Warning: {len(_aftercare_queue)} aftercare tasks abandoned at shutdown")

atexit.register(_drain_aftercare)


# ─── Main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))

    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("\n  WARNING:  ANTHROPIC_API_KEY not set!")
        print("  Option 1: Create .env file with:  ANTHROPIC_API_KEY=sk-ant-...")
        print("  Option 2: Set it from the Upload page in the browser")
        print()

    if not (BASE_DIR / "extract.py").exists():
        print("\n  WARNING:  extract.py not found in", BASE_DIR)
        print("  Place extract.py in the same folder as app.py\n")

    print("=" * 52)
    print(f"  Starting OathLedger on :{port}")
    print(f"  (Bearden Document Intake Platform v{_app_version})")
    print("  -------------------------------------")
    print(f"  Open in browser:  http://localhost:{port}")
    print(f"  Database:         {DB_PATH}")
    print(f"  Uploads:          {UPLOAD_DIR}")
    print(f"  Outputs:          {OUTPUT_DIR}")
    print(f"  Client folders:   {CLIENTS_DIR}")
    print("=" * 52)
    print()

    app.run(host="127.0.0.1", port=port, debug=False)
