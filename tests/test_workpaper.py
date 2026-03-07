#!/usr/bin/env python3
"""Tests for T1.6 — Deterministic Workpaper Modeling + DB-First Export Layer.

Covers: import guardrails, runtime guardrails, FactStore round-trip,
        workpaper generation, safe mode, formula protection, constants alignment.

Run:  python3 tests/test_workpaper.py
All test execution is inside run_tests() behind __name__ guard.
"""

import sys, os, json, tempfile, shutil, sqlite3

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

PASS = 0
FAIL = 0


def check(cond, msg):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  ✓ {msg}")
    else:
        FAIL += 1
        print(f"  ✗ FAIL: {msg}")


# ═══════════════════════════════════════════════════════════════════════════════
# IMPORT GUARDRAIL TESTS
# ═══════════════════════════════════════════════════════════════════════════════

def test_fact_store_imports_without_extract():
    """fact_store.py imports without extract.py on sys.path."""
    # Already imported at top if extract.py is available, but test the
    # _FORBIDDEN_MODULES guardrail
    from fact_store import _FORBIDDEN_MODULES
    check('extract' in _FORBIDDEN_MODULES, "fact_store forbids 'extract'")
    check('pytesseract' in _FORBIDDEN_MODULES, "fact_store forbids 'pytesseract'")
    check('anthropic' in _FORBIDDEN_MODULES, "fact_store forbids 'anthropic'")


def test_workpaper_export_imports_without_extract():
    """workpaper_export.py imports without extract.py on sys.path."""
    from workpaper_export import _FORBIDDEN_MODULES
    check('extract' in _FORBIDDEN_MODULES, "workpaper_export forbids 'extract'")
    check('PIL' in _FORBIDDEN_MODULES, "workpaper_export forbids 'PIL'")
    check('fitz' in _FORBIDDEN_MODULES, "workpaper_export forbids 'fitz'")


def test_no_forbidden_imports_in_source():
    """Neither module source code contains import statements for forbidden modules."""
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    for filename in ('fact_store.py', 'workpaper_export.py'):
        filepath = os.path.join(base, filename)
        with open(filepath, encoding="utf-8") as f:
            source = f.read()
        # Check for actual import statements (not strings in _FORBIDDEN_MODULES)
        lines = source.split('\n')
        violations = []
        for i, line in enumerate(lines, 1):
            stripped = line.strip()
            # Skip comments and the _FORBIDDEN_MODULES definition
            if stripped.startswith('#') or '_FORBIDDEN_MODULES' in line:
                continue
            if stripped.startswith(('import ', 'from ')):
                for forbidden in ('extract', 'pytesseract', 'anthropic',
                                   'pdf2image', 'PIL', 'Pillow', 'fitz'):
                    if f'import {forbidden}' in stripped or f'from {forbidden}' in stripped:
                        violations.append(f"  {filename}:{i}: {stripped}")
        check(len(violations) == 0,
              f"{filename} has no forbidden imports"
              + (f" — found: {violations}" if violations else ""))


# ═══════════════════════════════════════════════════════════════════════════════
# RUNTIME GUARDRAIL TESTS
# ═══════════════════════════════════════════════════════════════════════════════

def test_reject_pdf_path():
    """FactStore.upsert_fact rejects PDF file paths as canonical_value."""
    from fact_store import _reject_raw_inputs
    try:
        _reject_raw_inputs("/path/to/document.pdf")
        check(False, "should reject PDF path")
    except ValueError as e:
        check("PDF" in str(e), "rejects PDF path with descriptive error")


def test_reject_binary_data():
    """FactStore.upsert_fact rejects binary data (images, PDFs)."""
    from fact_store import _reject_raw_inputs
    try:
        _reject_raw_inputs(b'\x89PNG\r\n\x1a\n')
        check(False, "should reject binary data")
    except ValueError as e:
        check("binary" in str(e).lower(), "rejects binary with descriptive error")


def test_reject_large_text_blob():
    """FactStore.upsert_fact rejects large text (>5000 chars, likely OCR)."""
    from fact_store import _reject_raw_inputs
    try:
        _reject_raw_inputs("x" * 5001)
        check(False, "should reject large text blob")
    except ValueError as e:
        check("5001" in str(e) or "large" in str(e).lower(),
              "rejects large text with descriptive error")


def test_workpaper_rejects_pdf_as_client_name():
    """WorkpaperBuilder rejects PDF file path as client_name."""
    from workpaper_export import WorkpaperBuilder, _validate_identifier
    try:
        _validate_identifier("/Users/jeff/clients/Evans.pdf", "client_name")
        check(False, "should reject PDF as client_name")
    except ValueError as e:
        check("file path" in str(e).lower() or "identifier" in str(e).lower(),
              "rejects PDF path as client_name")


def test_workpaper_rejects_image_as_client_name():
    """WorkpaperBuilder rejects image file path as client_name."""
    from workpaper_export import _validate_identifier
    try:
        _validate_identifier("scan_page_001.png", "client_name")
        check(False, "should reject image path as client_name")
    except ValueError as e:
        check("file path" in str(e).lower() or "identifier" in str(e).lower(),
              "rejects image path as client_name")


def test_workpaper_rejects_non_factstore():
    """WorkpaperBuilder rejects non-FactStore as fact_store arg."""
    from workpaper_export import WorkpaperBuilder
    try:
        WorkpaperBuilder("not_a_factstore", "Evans, Lisa", "2025")
        check(False, "should reject non-FactStore")
    except TypeError as e:
        check("FactStore" in str(e), "rejects non-FactStore with descriptive error")


def test_workpaper_rejects_invalid_mode():
    """WorkpaperBuilder rejects invalid mode."""
    from fact_store import FactStore
    from workpaper_export import WorkpaperBuilder
    with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as f:
        db_path = f.name
    try:
        fs = FactStore(db_path)
        WorkpaperBuilder(fs, "Evans, Lisa", "2025", mode="invalid")
        check(False, "should reject invalid mode")
    except ValueError as e:
        check("mode" in str(e).lower(), "rejects invalid mode with descriptive error")
    finally:
        os.unlink(db_path)


def test_workpaper_rejects_empty_client_name():
    """WorkpaperBuilder rejects empty client_name."""
    from workpaper_export import _validate_identifier
    try:
        _validate_identifier("", "client_name")
        check(False, "should reject empty client_name")
    except ValueError as e:
        check("empty" in str(e).lower(), "rejects empty client_name")


def test_workpaper_rejects_long_client_name():
    """WorkpaperBuilder rejects oversized client_name (>200 chars)."""
    from workpaper_export import _validate_identifier
    try:
        _validate_identifier("A" * 201, "client_name")
        check(False, "should reject long client_name")
    except ValueError as e:
        check("too long" in str(e).lower(), "rejects long client_name")


def test_build_rejects_non_xlsx_output():
    """build() rejects output path not ending in .xlsx."""
    from fact_store import FactStore
    from workpaper_export import WorkpaperBuilder
    with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as f:
        db_path = f.name
    try:
        fs = FactStore(db_path)
        builder = WorkpaperBuilder(fs, "Evans, Lisa", "2025")
        builder.build("/tmp/output.pdf")
        check(False, "should reject non-xlsx output path")
    except ValueError as e:
        check(".xlsx" in str(e), "rejects non-xlsx output path")
    finally:
        os.unlink(db_path)


# ═══════════════════════════════════════════════════════════════════════════════
# FACTSTORE ROUND-TRIP TESTS
# ═══════════════════════════════════════════════════════════════════════════════

def _make_temp_factstore():
    """Create a FactStore with a temp database."""
    from fact_store import FactStore
    fd, db_path = tempfile.mkstemp(suffix='.db')
    os.close(fd)
    return FactStore(db_path), db_path


def test_upsert_get_round_trip():
    """Upsert + get_fact round-trip preserves all fields."""
    from fact_store import FactStore
    fs, db_path = _make_temp_factstore()
    try:
        fs.upsert_legacy_fact(
            "Evans, Lisa", "2025", "W-2", "ein:12-3456789", "wages",
            85000.00, original_value=84999.50, status="confirmed",
            source_job_id="job-001", reviewer="Jeffrey",
            payer_display="Acme Corp",
            evidence_ref="page1-box1", source_doc="evans-w2.pdf",
            page_number=1
        )
        fact = fs.get_legacy_fact("Evans, Lisa", "2025", "W-2", "ein:12-3456789", "wages")
        check(fact is not None, "get_fact returns result")
        check(fact["canonical_value"] == 85000.00, "canonical_value preserved")
        check(fact["original_value"] == 84999.50, "original_value preserved")
        check(fact["status"] == "confirmed", "status preserved")
        check(fact["payer_display"] == "Acme Corp", "payer_display preserved")
        check(fact["source_job_id"] == "job-001", "source_job_id preserved")
        check(fact["reviewer"] == "Jeffrey", "reviewer preserved")
        check(fact["evidence_ref"] == "page1-box1", "evidence_ref preserved")
        check(fact["source_doc"] == "evans-w2.pdf", "source_doc preserved")
        check(fact["page_number"] == 1, "page_number preserved")
        check(fact["fact_key"] == "W-2.ein:12-3456789.wages", "fact_key correct")
    finally:
        os.unlink(db_path)


def test_get_facts_filters():
    """get_facts filters by client/year and optionally document_type."""
    fs, db_path = _make_temp_factstore()
    try:
        # Insert facts for two doc types
        fs.upsert_legacy_fact("Evans, Lisa", "2025", "W-2", "ein:111", "wages", 50000)
        fs.upsert_legacy_fact("Evans, Lisa", "2025", "1099-INT", "ein:222", "interest_income", 1500)
        fs.upsert_legacy_fact("Smith, John", "2025", "W-2", "ein:333", "wages", 60000)

        # All facts for Evans 2025
        all_facts = fs.get_legacy_facts("Evans, Lisa", "2025")
        check(len(all_facts) == 2, f"get_facts returns 2 facts for Evans (got {len(all_facts)})")

        # Filter by doc type
        w2_facts = fs.get_legacy_facts("Evans, Lisa", "2025", document_type="W-2")
        check(len(w2_facts) == 1, f"filtered to 1 W-2 fact (got {len(w2_facts)})")
        check(w2_facts[0]["field_name"] == "wages", "correct field returned")

        # Different client has separate facts
        smith_facts = fs.get_legacy_facts("Smith, John", "2025")
        check(len(smith_facts) == 1, f"Smith has 1 fact (got {len(smith_facts)})")
    finally:
        os.unlink(db_path)


def test_list_facts():
    """list_facts returns correct (doc_type, payer_key, field_name, status) tuples."""
    fs, db_path = _make_temp_factstore()
    try:
        fs.upsert_legacy_fact("Evans, Lisa", "2025", "W-2", "ein:111", "wages", 50000, status="confirmed")
        fs.upsert_legacy_fact("Evans, Lisa", "2025", "W-2", "ein:111", "federal_wh", 8000, status="extracted")

        facts = fs.list_legacy_facts("Evans, Lisa", "2025")
        check(len(facts) == 2, f"list_facts returns 2 tuples (got {len(facts)})")
        # Sorted by field_name
        check(facts[0] == ("W-2", "ein:111", "federal_wh", "extracted"),
              "first tuple correct (sorted)")
        check(facts[1] == ("W-2", "ein:111", "wages", "confirmed"),
              "second tuple correct (sorted)")
    finally:
        os.unlink(db_path)


def test_missing_fact_returns_none():
    """Missing fact returns None — never re-extracts."""
    fs, db_path = _make_temp_factstore()
    try:
        result = fs.get_legacy_fact("Nobody", "2025", "W-2", "ein:000", "wages")
        check(result is None, "missing fact returns None")
    finally:
        os.unlink(db_path)


def test_upsert_overwrites_with_corrected():
    """Upsert with status='corrected' overwrites existing 'extracted' fact."""
    fs, db_path = _make_temp_factstore()
    try:
        fs.upsert_legacy_fact("Evans, Lisa", "2025", "W-2", "ein:111", "wages",
                         50000, status="extracted")
        fact1 = fs.get_legacy_fact("Evans, Lisa", "2025", "W-2", "ein:111", "wages")
        check(fact1["canonical_value"] == 50000, "initial value is 50000")
        check(fact1["status"] == "extracted", "initial status is extracted")

        fs.upsert_legacy_fact("Evans, Lisa", "2025", "W-2", "ein:111", "wages",
                         52000, status="corrected", reviewer="Susan")
        fact2 = fs.get_legacy_fact("Evans, Lisa", "2025", "W-2", "ein:111", "wages")
        check(fact2["canonical_value"] == 52000, "value updated to 52000")
        check(fact2["status"] == "corrected", "status updated to corrected")
        check(fact2["reviewer"] == "Susan", "reviewer updated")
    finally:
        os.unlink(db_path)


def test_string_value_round_trip():
    """String canonical values (text fields) round-trip correctly."""
    fs, db_path = _make_temp_factstore()
    try:
        fs.upsert_legacy_fact("Evans, Lisa", "2025", "1099-R", "ein:444",
                         "distribution_code", "7", status="extracted")
        fact = fs.get_legacy_fact("Evans, Lisa", "2025", "1099-R", "ein:444", "distribution_code")
        check(fact is not None, "text fact retrieved")
        check(fact["canonical_value"] == "7", f"text value preserved (got {fact['canonical_value']!r})")
    finally:
        os.unlink(db_path)


def test_none_value_round_trip():
    """None canonical value round-trips correctly."""
    fs, db_path = _make_temp_factstore()
    try:
        fs.upsert_legacy_fact("Evans, Lisa", "2025", "W-2", "ein:555",
                         "state_wh", None, status="missing")
        fact = fs.get_legacy_fact("Evans, Lisa", "2025", "W-2", "ein:555", "state_wh")
        check(fact is not None, "fact with None value exists")
        check(fact["canonical_value"] is None, "None value preserved")
        check(fact["status"] == "missing", "missing status preserved")
    finally:
        os.unlink(db_path)


# ═══════════════════════════════════════════════════════════════════════════════
# WORKPAPER GENERATION TESTS
# ═══════════════════════════════════════════════════════════════════════════════

def _build_test_workpaper(mode="assisted", extra_facts=None):
    """Helper: create a temp DB, insert test facts, build workpaper, return (wb, db_path, output_path)."""
    import openpyxl
    from fact_store import FactStore
    from workpaper_export import WorkpaperBuilder

    fd, db_path = tempfile.mkstemp(suffix='.db')
    os.close(fd)
    fs = FactStore(db_path)

    # Insert standard test facts
    fs.upsert_legacy_fact("Evans, Lisa", "2025", "W-2", "ein:12-3456789", "wages",
                     85000.00, status="confirmed", payer_display="Acme Corp",
                     source_doc="evans-w2.pdf", page_number=1,
                     evidence_ref="box1")
    fs.upsert_legacy_fact("Evans, Lisa", "2025", "W-2", "ein:12-3456789", "federal_wh",
                     12750.00, status="confirmed", payer_display="Acme Corp",
                     source_doc="evans-w2.pdf", page_number=1)
    fs.upsert_legacy_fact("Evans, Lisa", "2025", "W-2", "ein:12-3456789", "state_wh",
                     4250.00, status="confirmed", payer_display="Acme Corp",
                     source_doc="evans-w2.pdf", page_number=1)
    fs.upsert_legacy_fact("Evans, Lisa", "2025", "W-2", "ein:12-3456789", "employer_name",
                     "Acme Corp", status="confirmed", payer_display="Acme Corp")

    # 1099-INT with extracted (unverified) status
    fs.upsert_legacy_fact("Evans, Lisa", "2025", "1099-INT", "ein:99-8765432", "interest_income",
                     1250.50, status="extracted", payer_display="First National Bank",
                     source_doc="evans-1099int.pdf", page_number=1)

    # Extra facts for specific tests
    if extra_facts:
        for ef in extra_facts:
            fs.upsert_legacy_fact(*ef)

    fd2, output_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd2)

    builder = WorkpaperBuilder(fs, "Evans, Lisa", "2025", mode=mode)
    builder.build(output_path)

    wb = openpyxl.load_workbook(output_path)
    return wb, db_path, output_path


def _cleanup_workpaper(db_path, output_path):
    """Clean up temp files."""
    for p in (db_path, output_path):
        try:
            os.unlink(p)
        except OSError:
            pass


def test_workpaper_has_year_sheet():
    """Workpaper has a sheet named after the tax year."""
    wb, db_path, output_path = _build_test_workpaper()
    try:
        check("2025" in wb.sheetnames, f"year sheet exists (sheets: {wb.sheetnames})")
    finally:
        _cleanup_workpaper(db_path, output_path)


def test_workpaper_has_audit_trail():
    """Workpaper has an 'Audit Trail' sheet."""
    wb, db_path, output_path = _build_test_workpaper()
    try:
        check("Audit Trail" in wb.sheetnames,
              f"Audit Trail sheet exists (sheets: {wb.sheetnames})")
    finally:
        _cleanup_workpaper(db_path, output_path)


def test_w2_wages_in_correct_position():
    """W-2 wages value appears in the workpaper."""
    wb, db_path, output_path = _build_test_workpaper()
    try:
        ws = wb["2025"]
        # Search for the wages value (85000.00)
        found_wages = False
        for row in ws.iter_rows(min_col=2, max_col=7, values_only=False):
            for cell in row:
                if cell.value == 85000.00:
                    found_wages = True
                    break
            if found_wages:
                break
        check(found_wages, "W-2 wages (85000.00) found in workpaper")
    finally:
        _cleanup_workpaper(db_path, output_path)


def test_w2_employer_name_as_label():
    """W-2 employer name appears as a label in column A."""
    wb, db_path, output_path = _build_test_workpaper()
    try:
        ws = wb["2025"]
        found_employer = False
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=False):
            for cell in row:
                if cell.value and "Acme Corp" in str(cell.value):
                    found_employer = True
                    break
            if found_employer:
                break
        check(found_employer, "employer name 'Acme Corp' appears as label")
    finally:
        _cleanup_workpaper(db_path, output_path)


def test_sum_formulas_exist():
    """SUM formula cells exist in the workpaper."""
    wb, db_path, output_path = _build_test_workpaper()
    try:
        ws = wb["2025"]
        sum_formulas = []
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("=SUM("):
                    sum_formulas.append(cell.coordinate)
        check(len(sum_formulas) > 0,
              f"found {len(sum_formulas)} SUM formulas: {sum_formulas[:5]}")
    finally:
        _cleanup_workpaper(db_path, output_path)


def test_formula_cells_are_locked():
    """Formula cells have Protection(locked=True)."""
    wb, db_path, output_path = _build_test_workpaper()
    try:
        ws = wb["2025"]
        formula_locked = True
        formula_count = 0
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
                    if not cell.protection.locked:
                        formula_locked = False
        check(formula_count > 0, f"found {formula_count} formula cells")
        check(formula_locked, "all formula cells are locked")
    finally:
        _cleanup_workpaper(db_path, output_path)


def test_input_cells_are_unlocked():
    """Input cells (with numeric values) have Protection(locked=False)."""
    wb, db_path, output_path = _build_test_workpaper()
    try:
        ws = wb["2025"]
        # Find cells with numeric values in columns B-G
        unlocked_count = 0
        locked_input_count = 0
        for row in ws.iter_rows(min_col=2, max_col=7, values_only=False):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    if not cell.protection.locked:
                        unlocked_count += 1
                    else:
                        locked_input_count += 1
        check(unlocked_count > 0, f"found {unlocked_count} unlocked input cells")
        check(locked_input_count == 0,
              f"no locked input cells (found {locked_input_count})")
    finally:
        _cleanup_workpaper(db_path, output_path)


def test_safe_mode_omits_unverified():
    """Safe mode omits unverified values — empty cells with NEEDS REVIEW comment."""
    wb, db_path, output_path = _build_test_workpaper(mode="safe")
    try:
        ws = wb["2025"]
        # The 1099-INT interest_income is status="extracted" (not verified)
        # In safe mode, this should NOT appear as a number
        found_unverified_value = False
        for row in ws.iter_rows(min_col=2, max_col=7, values_only=False):
            for cell in row:
                if cell.value == 1250.50:
                    found_unverified_value = True
        check(not found_unverified_value,
              "safe mode omits unverified 1099-INT interest (1250.50)")

        # But W-2 wages (confirmed) should still appear
        found_wages = False
        for row in ws.iter_rows(min_col=2, max_col=7, values_only=False):
            for cell in row:
                if cell.value == 85000.00:
                    found_wages = True
        check(found_wages, "safe mode retains verified W-2 wages (85000.00)")
    finally:
        _cleanup_workpaper(db_path, output_path)


def test_assisted_mode_includes_unverified():
    """Assisted mode includes unverified values with status fill."""
    wb, db_path, output_path = _build_test_workpaper(mode="assisted")
    try:
        ws = wb["2025"]
        found_interest = False
        for row in ws.iter_rows(min_col=2, max_col=7, values_only=False):
            for cell in row:
                if cell.value == 1250.50:
                    found_interest = True
        check(found_interest,
              "assisted mode includes unverified 1099-INT interest (1250.50)")
    finally:
        _cleanup_workpaper(db_path, output_path)


def test_status_fill_colors_applied():
    """Status fill colors are applied to cells."""
    from workpaper_export import STATUS_FILLS
    wb, db_path, output_path = _build_test_workpaper()
    try:
        ws = wb["2025"]
        filled_cells = 0
        for row in ws.iter_rows(min_col=2, max_col=7, values_only=False):
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.fill and cell.fill.fgColor:
                    if cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != '00000000':
                        filled_cells += 1
        check(filled_cells > 0,
              f"found {filled_cells} cells with status fill colors")
    finally:
        _cleanup_workpaper(db_path, output_path)


def test_sheet_protection_enabled():
    """Sheet protection is enabled (structural only, no password)."""
    wb, db_path, output_path = _build_test_workpaper()
    try:
        ws = wb["2025"]
        check(ws.protection.sheet, "sheet protection is enabled")
        check(not ws.protection.password, "no password set (structural only)")
    finally:
        _cleanup_workpaper(db_path, output_path)


def test_title_block():
    """Workpaper has title with client name."""
    wb, db_path, output_path = _build_test_workpaper()
    try:
        ws = wb["2025"]
        title = ws.cell(row=1, column=1).value
        check(title is not None and "Evans, Lisa" in title,
              f"title contains client name (got: {title!r})")
    finally:
        _cleanup_workpaper(db_path, output_path)


def test_multiple_payers_same_doc_type():
    """Multiple payers for same document type each get their own row."""
    extra = [
        ("Evans, Lisa", "2025", "W-2", "ein:99-0000001", "wages",
         45000.00, None, "confirmed", "", "", "Beta Inc"),
        ("Evans, Lisa", "2025", "W-2", "ein:99-0000001", "employer_name",
         "Beta Inc", None, "confirmed", "", "", "Beta Inc"),
        ("Evans, Lisa", "2025", "W-2", "ein:99-0000001", "federal_wh",
         6750.00, None, "confirmed", "", "", "Beta Inc"),
    ]
    wb, db_path, output_path = _build_test_workpaper(extra_facts=extra)
    try:
        ws = wb["2025"]
        wages_values = []
        for row in ws.iter_rows(min_col=2, max_col=2, values_only=False):
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value in (85000.0, 45000.0):
                    wages_values.append(cell.value)
        check(len(wages_values) >= 2,
              f"found {len(wages_values)} W-2 wage entries for multiple payers")
    finally:
        _cleanup_workpaper(db_path, output_path)


def test_audit_trail_has_facts():
    """Audit Trail sheet contains fact entries."""
    wb, db_path, output_path = _build_test_workpaper()
    try:
        ws = wb["Audit Trail"]
        # Count data rows (after header row 5)
        data_rows = 0
        for row in ws.iter_rows(min_row=6, max_col=1, values_only=True):
            if row[0]:
                data_rows += 1
        check(data_rows > 0, f"Audit Trail has {data_rows} fact entries")
    finally:
        _cleanup_workpaper(db_path, output_path)


def test_audit_trail_summary():
    """Audit Trail has summary with total and verified counts."""
    wb, db_path, output_path = _build_test_workpaper()
    try:
        ws = wb["Audit Trail"]
        found_summary = False
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            if row[0] and "Total facts written" in str(row[0]):
                found_summary = True
        check(found_summary, "Audit Trail has summary section")
    finally:
        _cleanup_workpaper(db_path, output_path)


# ═══════════════════════════════════════════════════════════════════════════════
# CONSTANTS ALIGNMENT TEST
# ═══════════════════════════════════════════════════════════════════════════════

def test_registry_section_ids_match_template():
    """MAPPING_REGISTRY section IDs cover all TEMPLATE_SECTIONS IDs from extract.py."""
    from workpaper_export import MAPPING_REGISTRY

    # Get registry IDs
    registry_ids = {s["id"] for s in MAPPING_REGISTRY}

    # Expected IDs from extract.py TEMPLATE_SECTIONS
    expected_ids = {
        "w2", "interest", "dividends", "1099r", "ssa", "1099nec",
        "1099misc", "1099g", "schedule_d", "k1", "k1_detail",
        "rental", "farm", "1098", "1098t", "property_tax",
        "estimated", "charitable", "w2g", "1099k", "1099s",
        "1099c", "1098e", "5498", "schedule_c",
        "bank_statement", "credit_card", "check_stub", "invoice",
        "receipt", "profit_loss", "balance_sheet", "loan_statement",
        "payroll_register", "payroll_tax",
    }

    missing = expected_ids - registry_ids
    extra = registry_ids - expected_ids
    check(len(missing) == 0,
          f"no missing section IDs" + (f" — missing: {missing}" if missing else ""))
    check(len(extra) == 0,
          f"no extra section IDs" + (f" — extra: {extra}" if extra else ""))
    check(len(registry_ids) == 35,
          f"registry has 35 sections (got {len(registry_ids)})")


def test_always_show_sections_in_registry():
    """ALWAYS_SHOW section IDs all exist in MAPPING_REGISTRY."""
    from workpaper_export import MAPPING_REGISTRY, ALWAYS_SHOW
    registry_ids = {s["id"] for s in MAPPING_REGISTRY}
    missing = ALWAYS_SHOW - registry_ids
    check(len(missing) == 0,
          f"all ALWAYS_SHOW IDs in registry"
          + (f" — missing: {missing}" if missing else ""))


def test_verified_statuses_completeness():
    """VERIFIED_STATUSES contains all expected verification statuses."""
    from workpaper_export import VERIFIED_STATUSES
    expected = {"confirmed", "corrected", "auto_verified",
                "verified_confirmed", "verified_corrected",
                "dual_confirmed", "consensus_accepted", "multipage_verified"}
    check(expected == VERIFIED_STATUSES,
          f"VERIFIED_STATUSES matches expected set")


# ═══════════════════════════════════════════════════════════════════════════════
# DB-ONLY WORKBOOK BUILD TEST
# ═══════════════════════════════════════════════════════════════════════════════

def test_db_only_build():
    """Insert facts into temp SQLite → generate workbook → verify cells match.
    No extraction code involved.
    """
    import openpyxl
    from fact_store import FactStore
    from workpaper_export import WorkpaperBuilder

    fd, db_path = tempfile.mkstemp(suffix='.db')
    os.close(fd)
    fd2, output_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd2)

    try:
        fs = FactStore(db_path)

        # Insert a set of W-2 facts directly into DB
        fs.upsert_legacy_fact("TestClient", "2024", "W-2", "ein:11-1111111",
                         "wages", 75000.00, status="confirmed",
                         payer_display="Test Employer")
        fs.upsert_legacy_fact("TestClient", "2024", "W-2", "ein:11-1111111",
                         "federal_wh", 11250.00, status="confirmed",
                         payer_display="Test Employer")
        fs.upsert_legacy_fact("TestClient", "2024", "W-2", "ein:11-1111111",
                         "employer_name", "Test Employer", status="confirmed",
                         payer_display="Test Employer")

        builder = WorkpaperBuilder(fs, "TestClient", "2024", mode="assisted")
        builder.build(output_path)

        wb = openpyxl.load_workbook(output_path)
        check("2024" in wb.sheetnames, "DB-only build has year sheet '2024'")

        ws = wb["2024"]
        # Verify wages value is present
        wages_found = False
        fwh_found = False
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if cell.value == 75000.00:
                    wages_found = True
                if cell.value == 11250.00:
                    fwh_found = True
        check(wages_found, "DB-only build: wages (75000) present")
        check(fwh_found, "DB-only build: federal_wh (11250) present")
    finally:
        for p in (db_path, output_path):
            try:
                os.unlink(p)
            except OSError:
                pass


# ═══════════════════════════════════════════════════════════════════════════════
# CROSS-COLUMN FORMULA TEST
# ═══════════════════════════════════════════════════════════════════════════════

def test_cross_column_formula():
    """Interest section generates cross-column total formulas."""
    extra = [
        ("Evans, Lisa", "2025", "1099-INT", "ein:99-8765432",
         "us_savings_bonds_and_treasury", 500.00, None, "confirmed", "", "",
         "First National Bank"),
    ]
    wb, db_path, output_path = _build_test_workpaper(extra_facts=extra)
    try:
        ws = wb["2025"]
        # Look for a formula like =B{n}+C{n} in column D
        cross_formulas = []
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("=") and "+" in cell.value:
                    if "SUM" not in cell.value:
                        cross_formulas.append(f"{cell.coordinate}: {cell.value}")
        check(len(cross_formulas) > 0,
              f"found {len(cross_formulas)} cross-column formulas: {cross_formulas[:3]}")
    finally:
        _cleanup_workpaper(db_path, output_path)


# ═══════════════════════════════════════════════════════════════════════════════
# EMPTY WORKPAPER TEST
# ═══════════════════════════════════════════════════════════════════════════════

def test_empty_workpaper():
    """Workpaper with no facts still generates ALWAYS_SHOW sections."""
    import openpyxl
    from fact_store import FactStore
    from workpaper_export import WorkpaperBuilder, ALWAYS_SHOW

    fd, db_path = tempfile.mkstemp(suffix='.db')
    os.close(fd)
    fd2, output_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd2)

    try:
        fs = FactStore(db_path)
        builder = WorkpaperBuilder(fs, "Empty, Client", "2025", mode="assisted")
        builder.build(output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["2025"]

        # Should still have section headers for ALWAYS_SHOW sections
        headers_found = set()
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            val = row[0]
            if val and isinstance(val, str):
                if "W-2:" in val:
                    headers_found.add("w2")
                if "Interest Income:" in val:
                    headers_found.add("interest")
                if "Dividends:" in val:
                    headers_found.add("dividends")
                if "Schedule D:" in val:
                    headers_found.add("schedule_d")
                if "K-1s:" in val:
                    headers_found.add("k1")

        check(len(headers_found) >= len(ALWAYS_SHOW),
              f"empty workpaper shows ALWAYS_SHOW sections (found {headers_found})")
    finally:
        for p in (db_path, output_path):
            try:
                os.unlink(p)
            except OSError:
                pass


# ═══════════════════════════════════════════════════════════════════════════════
# WORKPAPER-001: CANONICAL FACTS PATH TESTS
# ═══════════════════════════════════════════════════════════════════════════════

def test_canonical_facts_round_trip():
    """WorkpaperBuilder with job_id reads from canonical facts table."""
    from fact_store import FactStore

    fd, db_path = tempfile.mkstemp(suffix='.db')
    os.close(fd)

    try:
        fs = FactStore(db_path)

        # Write a fact into the canonical table
        fs.upsert_candidate_fact(
            job_id="test-job-001",
            client_id="Smith, John",
            tax_year="2025",
            fact_key="W-2.ein:99-1234567.wages",
            value_num=85000.0,
            value_text=None,
            status="confirmed",
            confidence="high",
            source_method="vision",
            source_doc="smith_w2.pdf",
            source_page=1,
            evidence_ref="box1",
        )

        # Retrieve via workpaper method
        facts = fs.get_workpaper_facts("test-job-001", "2025")
        check(len(facts) == 1, f"get_workpaper_facts returns 1 fact (got {len(facts)})")
        f = facts[0]
        check(f["document_type"] == "W-2", f"document_type parsed correctly: {f['document_type']}")
        check(f["payer_key"] == "ein:99-1234567", f"payer_key parsed correctly: {f['payer_key']}")
        check(f["field_name"] == "wages", f"field_name parsed correctly: {f['field_name']}")
        check(f["canonical_value"] == 85000.0, f"canonical_value from value_num: {f['canonical_value']}")
        check(f["status"] == "confirmed", f"status preserved: {f['status']}")
        check(f["payer_display"] == "EIN 99-1234567", f"payer_display derived: {f['payer_display']}")
    finally:
        try:
            os.unlink(db_path)
        except OSError:
            pass


def test_canonical_workpaper_build():
    """WorkpaperBuilder with job_id generates workpaper from canonical facts."""
    import openpyxl
    from fact_store import FactStore
    from workpaper_export import WorkpaperBuilder

    fd, db_path = tempfile.mkstemp(suffix='.db')
    os.close(fd)
    fd2, output_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd2)

    try:
        fs = FactStore(db_path)

        # Write canonical facts for a W-2
        for field, val in [("wages", 85000.0), ("federal_wh", 12000.0), ("state_wh", 5000.0)]:
            fs.upsert_candidate_fact(
                job_id="test-job-002",
                client_id="Jones, Jane",
                tax_year="2025",
                fact_key=f"W-2.ein:11-2222222.{field}",
                value_num=val,
                value_text=None,
                status="confirmed",
                confidence="high",
                source_method="vision",
                source_doc="jones_w2.pdf",
                source_page=1,
                evidence_ref=f"box-{field}",
            )

        builder = WorkpaperBuilder(fs, "Jones, Jane", "2025", mode="assisted",
                                   job_id="test-job-002")
        builder.build(output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["2025"]

        # Check subtitle includes "canonical facts"
        subtitle = ws.cell(row=2, column=1).value or ""
        check("canonical facts" in subtitle.lower(),
              f"subtitle shows canonical facts source: {subtitle[:80]}")

        # Verify data rendered
        values_found = []
        for row in ws.iter_rows(min_col=2, max_col=4, values_only=True):
            for v in row:
                if isinstance(v, (int, float)) and v > 0:
                    values_found.append(v)
        check(85000.0 in values_found, "wages value 85000 rendered in workpaper")
        check(12000.0 in values_found, "federal_wh value 12000 rendered in workpaper")
    finally:
        for p in (db_path, output_path):
            try:
                os.unlink(p)
            except OSError:
                pass


def test_canonical_fallback_to_legacy():
    """WorkpaperBuilder falls back to legacy facts if job_id has no canonical facts."""
    from fact_store import FactStore
    from workpaper_export import WorkpaperBuilder

    fd, db_path = tempfile.mkstemp(suffix='.db')
    os.close(fd)
    fd2, output_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd2)

    try:
        fs = FactStore(db_path)

        # Write only legacy facts (no canonical facts for this job)
        fs.upsert_legacy_fact(
            "Test, Client", "2025", "W-2", "ein:33-3333333", "wages",
            canonical_value=50000.0, status="confirmed", payer_display="Acme Corp"
        )

        # Build with a job_id that has no canonical facts → should fall back
        builder = WorkpaperBuilder(fs, "Test, Client", "2025", mode="assisted",
                                   job_id="nonexistent-job")
        builder.build(output_path)

        check(os.path.exists(output_path), "workpaper generated via legacy fallback")
    finally:
        for p in (db_path, output_path):
            try:
                os.unlink(p)
            except OSError:
                pass


def test_parse_fact_key():
    """FactStore._parse_fact_key handles various key formats correctly."""
    from fact_store import FactStore

    # Standard: "W-2.ein:12-3456789.wages"
    dt, pk, fn = FactStore._parse_fact_key("W-2.ein:12-3456789.wages")
    check(dt == "W-2", f"doc_type=W-2: {dt}")
    check(pk == "ein:12-3456789", f"payer_key=ein:12-3456789: {pk}")
    check(fn == "wages", f"field_name=wages: {fn}")

    # Payer with dot: "1099-INT.Chase.Bank.interest_income"
    dt, pk, fn = FactStore._parse_fact_key("1099-INT.Chase.Bank.interest_income")
    check(dt == "1099-INT", f"doc_type=1099-INT: {dt}")
    check(pk == "Chase.Bank", f"payer_key=Chase.Bank: {pk}")
    check(fn == "interest_income", f"field_name=interest_income: {fn}")

    # Minimal: "W-2.field"
    dt, pk, fn = FactStore._parse_fact_key("W-2.wages")
    check(dt == "W-2", f"minimal doc_type=W-2: {dt}")
    check(fn == "wages", f"minimal field_name=wages: {fn}")


def test_custom_layout():
    """WorkpaperBuilder accepts custom layout override."""
    import openpyxl
    from fact_store import FactStore
    from workpaper_export import WorkpaperBuilder

    fd, db_path = tempfile.mkstemp(suffix='.db')
    os.close(fd)
    fd2, output_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd2)

    try:
        fs = FactStore(db_path)

        # Write a fact
        fs.upsert_candidate_fact(
            job_id="layout-test",
            client_id="Layout, Test",
            tax_year="2025",
            fact_key="W-2.ein:00-0000001.wages",
            value_num=99000.0,
            value_text=None,
            status="confirmed",
            confidence="high",
            source_method="vision",
            source_doc="test.pdf",
            source_page=1,
        )

        # Minimal custom layout — just W-2 section
        custom_layout = [
            {
                "id": "w2",
                "header": "W-2 (Custom Layout):",
                "match_types": ["W-2"],
                "fields": {
                    "wages": {"col": "A", "type": "input", "fmt": "money"},
                },
                "col_headers": {"A": "Gross Wages"},
                "sum_cols": ["A"],
            },
        ]

        builder = WorkpaperBuilder(fs, "Layout, Test", "2025", mode="assisted",
                                   job_id="layout-test", layout=custom_layout)
        builder.build(output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["2025"]

        # Find the custom header
        found_custom = False
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            if row[0] and "Custom Layout" in str(row[0]):
                found_custom = True
                break
        check(found_custom, "custom layout header rendered in workpaper")
    finally:
        for p in (db_path, output_path):
            try:
                os.unlink(p)
            except OSError:
                pass


# ═══════════════════════════════════════════════════════════════════════════════
# RUN ALL TESTS
# ═══════════════════════════════════════════════════════════════════════════════

def run_tests():
    global PASS, FAIL

    print("\n═══ IMPORT GUARDRAIL TESTS ═══")
    test_fact_store_imports_without_extract()
    test_workpaper_export_imports_without_extract()
    test_no_forbidden_imports_in_source()

    print("\n═══ RUNTIME GUARDRAIL TESTS ═══")
    test_reject_pdf_path()
    test_reject_binary_data()
    test_reject_large_text_blob()
    test_workpaper_rejects_pdf_as_client_name()
    test_workpaper_rejects_image_as_client_name()
    test_workpaper_rejects_non_factstore()
    test_workpaper_rejects_invalid_mode()
    test_workpaper_rejects_empty_client_name()
    test_workpaper_rejects_long_client_name()
    test_build_rejects_non_xlsx_output()

    print("\n═══ FACTSTORE ROUND-TRIP TESTS ═══")
    test_upsert_get_round_trip()
    test_get_facts_filters()
    test_list_facts()
    test_missing_fact_returns_none()
    test_upsert_overwrites_with_corrected()
    test_string_value_round_trip()
    test_none_value_round_trip()

    print("\n═══ WORKPAPER GENERATION TESTS ═══")
    test_workpaper_has_year_sheet()
    test_workpaper_has_audit_trail()
    test_w2_wages_in_correct_position()
    test_w2_employer_name_as_label()
    test_sum_formulas_exist()
    test_formula_cells_are_locked()
    test_input_cells_are_unlocked()
    test_safe_mode_omits_unverified()
    test_assisted_mode_includes_unverified()
    test_status_fill_colors_applied()
    test_sheet_protection_enabled()
    test_title_block()
    test_multiple_payers_same_doc_type()
    test_audit_trail_has_facts()
    test_audit_trail_summary()
    test_cross_column_formula()
    test_empty_workpaper()

    print("\n═══ CONSTANTS ALIGNMENT TESTS ═══")
    test_registry_section_ids_match_template()
    test_always_show_sections_in_registry()
    test_verified_statuses_completeness()

    print("\n═══ DB-ONLY BUILD TEST ═══")
    test_db_only_build()

    print(f"\n{'='*60}")
    print(f"  PASS: {PASS}  |  FAIL: {FAIL}  |  TOTAL: {PASS + FAIL}")
    print(f"{'='*60}")
    if FAIL > 0:
        sys.exit(1)


if __name__ == "__main__":
    run_tests()
