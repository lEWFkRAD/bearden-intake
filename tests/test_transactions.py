#!/usr/bin/env python3
"""
Tests for T-TXN-LEDGER-1: Transaction Ledger
=============================================
Tests TransactionStore, transaction_extract, and transaction_reports modules.

Run:
    python3 tests/test_transactions.py
"""
import os
import sys
import json
import tempfile

# Add parent dir to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

PASS = 0
FAIL = 0


def check(cond, msg):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  ✓ {msg}")
    else:
        FAIL += 1
        print(f"  ✗ FAIL: {msg}")


# ── Helpers ──────────────────────────────────────────────────────────────────

def _make_ts():
    """Create a TransactionStore with a temp database."""
    from transaction_store import TransactionStore
    fd, db_path = tempfile.mkstemp(suffix='.db')
    os.close(fd)
    return TransactionStore(db_path), db_path


def _cleanup(db_path):
    try:
        os.unlink(db_path)
    except OSError:
        pass


def _sample_bank_log():
    """Sample extraction log with bank statement transactions."""
    return {
        "extractions": [
            {
                "document_type": "bank_statement",
                "payer_or_entity": "First National Bank",
                "_page": 1,
                "_extraction_method": "ocr_text",
                "fields": {
                    "bank_name": {"value": "First National Bank", "confidence": "high"},
                    "beginning_balance": {"value": 5000.00, "confidence": "high"},
                    "ending_balance": {"value": 4250.50, "confidence": "high"},
                    "txn_1_date": {"value": "01/15/2025", "confidence": "high"},
                    "txn_1_desc": {"value": "GEORGIA POWER COMPANY #12345", "confidence": "high"},
                    "txn_1_amount": {"value": 150.00, "confidence": "high"},
                    "txn_1_type": {"value": "withdrawal", "confidence": "high"},
                    "txn_2_date": {"value": "01/20/2025", "confidence": "high"},
                    "txn_2_desc": {"value": "DIRECT DEPOSIT ACME INC", "confidence": "high"},
                    "txn_2_amount": {"value": 2500.00, "confidence": "high"},
                    "txn_2_type": {"value": "deposit", "confidence": "high"},
                    "txn_3_date": {"value": "01/25/2025", "confidence": "medium"},
                    "txn_3_desc": {"value": "WAL-MART SUPER CENTER 0423", "confidence": "medium"},
                    "txn_3_amount": {"value": 99.50, "confidence": "medium"},
                    "txn_3_type": {"value": "withdrawal", "confidence": "high"},
                },
            },
        ],
    }


def _sample_cc_log():
    """Sample extraction log with credit card transactions."""
    return {
        "extractions": [
            {
                "document_type": "credit_card_statement",
                "payer_or_entity": "Chase Visa",
                "_page": 1,
                "fields": {
                    "card_issuer": {"value": "Chase"},
                    "txn_1_date": {"value": "02/01/2025"},
                    "txn_1_desc": {"value": "AMAZON.COM"},
                    "txn_1_amount": {"value": "$45.99"},
                    "txn_1_type": {"value": "purchase"},
                    "txn_1_category": {"value": "Office Supplies"},
                    "txn_2_date": {"value": "02/05/2025"},
                    "txn_2_desc": {"value": "HILTON HOTELS"},
                    "txn_2_amount": {"value": "189.00"},
                    "txn_2_type": {"value": "purchase"},
                },
            },
        ],
    }


def _sample_check_log():
    """Sample extraction log with a check."""
    return {
        "extractions": [
            {
                "document_type": "check",
                "payer_or_entity": "Client Account",
                "_page": 1,
                "fields": {
                    "check_number": {"value": "1234"},
                    "check_date": {"value": "03/15/2025"},
                    "payee": {"value": "Office Depot"},
                    "check_amount": {"value": "250.00"},
                },
            },
        ],
    }


def _sample_tax_log():
    """Sample extraction log with tax documents (should be ignored)."""
    return {
        "extractions": [
            {
                "document_type": "W-2",
                "payer_or_entity": "Employer Inc",
                "fields": {
                    "wages": {"value": 50000.00},
                    "federal_wh": {"value": 8000.00},
                },
            },
        ],
    }


# ═════════════════════════════════════════════════════════════════════════════
# TESTS
# ═════════════════════════════════════════════════════════════════════════════

def test_import_guardrails():
    """Verify none of the transaction modules import forbidden modules."""
    print("\n=== Import Guardrail Tests ===")
    import transaction_store
    import transaction_extract

    for mod_name, mod in [("transaction_store", transaction_store),
                           ("transaction_extract", transaction_extract)]:
        source_file = mod.__file__
        with open(source_file) as f:
            source = f.read()
        for forbidden in transaction_store._FORBIDDEN_MODULES:
            has_import = (f"import {forbidden}" in source and
                          f"'{forbidden}'" not in source and
                          f'"{forbidden}"' not in source)
            check(not has_import,
                  f"{mod_name} does not import {forbidden}")


def test_schema():
    """Schema creation and idempotency."""
    print("\n=== Schema Tests ===")
    import sqlite3
    ts, db_path = _make_ts()
    try:
        conn = sqlite3.connect(db_path)
        tables = [r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
        ).fetchall()]
        conn.close()

        for tbl in ("txn_values", "txn_events", "txn_evidence",
                     "vendor_rules", "category_rules"):
            check(tbl in tables, f"Table {tbl} exists")

        # Idempotent — calling _ensure_schema again should not fail
        ts._ensure_schema()
        check(True, "Schema creation is idempotent")
    finally:
        _cleanup(db_path)


def test_txn_id():
    """Deterministic transaction ID generation."""
    print("\n=== txn_id Tests ===")
    from transaction_store import TransactionStore

    id1 = TransactionStore.make_txn_id("job1", "bank_statement", "BANK", 1)
    id2 = TransactionStore.make_txn_id("job1", "bank_statement", "BANK", 1)
    check(id1 == id2, "Same inputs produce same txn_id (deterministic)")
    check(len(id1) == 16, f"txn_id is 16 chars (got {len(id1)})")

    id3 = TransactionStore.make_txn_id("job1", "bank_statement", "BANK", 2)
    check(id1 != id3, "Different txn_index produces different txn_id")

    id4 = TransactionStore.make_txn_id("job2", "bank_statement", "BANK", 1)
    check(id1 != id4, "Different job_id produces different txn_id")


def test_ingest_bank_statement():
    """Ingest bank statement transactions."""
    print("\n=== Ingest Tests (Bank Statement) ===")
    ts, db_path = _make_ts()
    try:
        log_data = _sample_bank_log()
        result = ts.ingest_from_extraction("job-001", log_data, "Test Client", 2025)

        check(result["total_parsed"] == 3, f"Parsed 3 transactions (got {result['total_parsed']})")
        check(result["inserted"] == 3, f"Inserted 3 transactions (got {result['inserted']})")
        check(result["skipped_dup"] == 0, f"No duplicates (got {result['skipped_dup']})")

        # Check status = staged
        txn = ts.get_transactions("Test Client", 2025)
        check(txn["total"] == 3, f"3 transactions in DB (got {txn['total']})")
        for item in txn["items"]:
            check(item["status"] == "staged", f"Transaction {item['txn_id'][:8]} status is staged")
            break  # Check first one

        # Verify events were created
        first_txn = txn["items"][0]
        events = ts.get_events(first_txn["txn_id"])
        check(len(events) >= 1, f"At least 1 event for first transaction (got {len(events)})")
        check(events[0]["event_type"] == "staged", f"Event type is 'staged' (got {events[0]['event_type']})")
    finally:
        _cleanup(db_path)


def test_ingest_credit_card():
    """Ingest credit card transactions."""
    print("\n=== Ingest Tests (Credit Card) ===")
    ts, db_path = _make_ts()
    try:
        result = ts.ingest_from_extraction("job-cc1", _sample_cc_log(), "CC Client", 2025)
        check(result["inserted"] == 2, f"Inserted 2 CC transactions (got {result['inserted']})")

        txns = ts.get_transactions("CC Client", 2025)
        # CC with category pre-set from extraction
        found_category = False
        for item in txns["items"]:
            if item["category"] == "Office Supplies":
                found_category = True
        check(found_category, "Credit card transaction has pre-extracted category")
    finally:
        _cleanup(db_path)


def test_ingest_check():
    """Ingest check document as single transaction."""
    print("\n=== Ingest Tests (Check) ===")
    ts, db_path = _make_ts()
    try:
        result = ts.ingest_from_extraction("job-chk1", _sample_check_log(), "Check Client", 2025)
        check(result["inserted"] == 1, f"Inserted 1 check transaction (got {result['inserted']})")

        txns = ts.get_transactions("Check Client", 2025)
        check(txns["total"] == 1, "1 transaction in DB")
        item = txns["items"][0]
        check(item["txn_type"] == "check", f"Type is 'check' (got {item['txn_type']})")
        check("1234" in item["description"], f"Description contains check number (got {item['description']})")
        check(item["amount"] == 250.00, f"Amount is 250.00 (got {item['amount']})")
    finally:
        _cleanup(db_path)


def test_ingest_dedup():
    """Re-ingesting same job produces no duplicates."""
    print("\n=== Dedup Tests ===")
    ts, db_path = _make_ts()
    try:
        log_data = _sample_bank_log()
        r1 = ts.ingest_from_extraction("job-001", log_data, "Test Client", 2025)
        check(r1["inserted"] == 3, "First ingest: 3 inserted")

        r2 = ts.ingest_from_extraction("job-001", log_data, "Test Client", 2025)
        check(r2["inserted"] == 0, f"Second ingest: 0 inserted (got {r2['inserted']})")
        check(r2["skipped_dup"] == 3, f"Second ingest: 3 skipped (got {r2['skipped_dup']})")

        total = ts.get_transactions("Test Client", 2025)
        check(total["total"] == 3, f"Still only 3 in DB (got {total['total']})")
    finally:
        _cleanup(db_path)


def test_ingest_skips_tax_docs():
    """Tax documents should not produce transactions."""
    print("\n=== Skip Non-TXN Tests ===")
    ts, db_path = _make_ts()
    try:
        result = ts.ingest_from_extraction("job-tax", _sample_tax_log(), "Tax Client", 2025)
        check(result["total_parsed"] == 0, f"No transactions parsed from W-2 (got {result['total_parsed']})")
    finally:
        _cleanup(db_path)


def test_categorize():
    """Categorize a transaction."""
    print("\n=== Categorize Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)
        txns = ts.get_transactions("Test Client", 2025)
        txn_id = txns["items"][0]["txn_id"]

        # Categorize
        ok = ts.categorize(txn_id, "Utilities", reviewer="Jeffrey")
        check(ok, "categorize() returns True")

        txn = ts.get_transaction(txn_id)
        check(txn["category"] == "Utilities", f"Category set to 'Utilities' (got {txn['category']})")
        check(txn["category_group"] == "Operating Expenses > Facilities",
              f"Group set correctly (got {txn['category_group']})")
        check(txn["status"] == "suggested", f"Status advanced to 'suggested' (got {txn['status']})")

        # Event recorded
        events = ts.get_events(txn_id)
        cat_events = [e for e in events if e["event_type"] == "category_set"]
        check(len(cat_events) >= 1, "category_set event recorded")
        check(cat_events[-1]["new_value"] == "Utilities", "Event new_value is 'Utilities'")
        check(cat_events[-1]["reviewer"] == "Jeffrey", "Event reviewer is 'Jeffrey'")
    finally:
        _cleanup(db_path)


def test_categorize_refuses_invalid():
    """Categorize refuses invalid category."""
    print("\n=== Categorize Validation Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)
        txns = ts.get_transactions("Test Client", 2025)
        txn_id = txns["items"][0]["txn_id"]

        ok = ts.categorize(txn_id, "INVALID_CATEGORY")
        check(not ok, "Refuses invalid category")
    finally:
        _cleanup(db_path)


def test_categorize_refuses_locked():
    """Categorize refuses locked (corrected) transaction."""
    print("\n=== Categorize Lock Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)
        txns = ts.get_transactions("Test Client", 2025)
        txn_id = txns["items"][0]["txn_id"]

        # Correct the transaction (locks it)
        ts.correct(txn_id, {"amount": 999.99, "category": "Utilities"}, reviewer="Admin")

        ok = ts.categorize(txn_id, "Payroll")
        check(not ok, "Refuses to categorize locked (corrected) transaction")
    finally:
        _cleanup(db_path)


def test_verify():
    """Verify a categorized transaction."""
    print("\n=== Verify Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)
        txns = ts.get_transactions("Test Client", 2025)
        txn_id = txns["items"][0]["txn_id"]

        # Must categorize first
        ts.categorize(txn_id, "Utilities")

        ok = ts.verify(txn_id, reviewer="Susan")
        check(ok, "verify() returns True")

        txn = ts.get_transaction(txn_id)
        check(txn["status"] == "verified", f"Status is 'verified' (got {txn['status']})")

        events = ts.get_events(txn_id)
        ver_events = [e for e in events if e["event_type"] == "verified"]
        check(len(ver_events) >= 1, "verified event recorded")
    finally:
        _cleanup(db_path)


def test_verify_refuses_uncategorized():
    """Verify refuses transaction without category."""
    print("\n=== Verify Validation Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)
        txns = ts.get_transactions("Test Client", 2025)
        txn_id = txns["items"][0]["txn_id"]

        ok = ts.verify(txn_id)
        check(not ok, "Refuses to verify uncategorized transaction")
    finally:
        _cleanup(db_path)


def test_correct():
    """Correct a transaction."""
    print("\n=== Correct Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)
        txns = ts.get_transactions("Test Client", 2025)
        txn_id = txns["items"][0]["txn_id"]
        original_amount = txns["items"][0]["amount"]

        ok = ts.correct(txn_id, {"amount": 175.00, "category": "Utilities"}, reviewer="Charles")
        check(ok, "correct() returns True")

        txn = ts.get_transaction(txn_id)
        check(txn["amount"] == 175.00, f"Amount corrected to 175.00 (got {txn['amount']})")
        check(txn["category"] == "Utilities", f"Category set to 'Utilities' (got {txn['category']})")
        check(txn["status"] == "corrected", f"Status is 'corrected' (got {txn['status']})")

        events = ts.get_events(txn_id)
        corr_events = [e for e in events if e["event_type"] == "corrected"]
        check(len(corr_events) >= 1, "corrected event recorded")
        # Check old/new values in event
        evt = corr_events[-1]
        old_vals = json.loads(evt["old_value"])
        new_vals = json.loads(evt["new_value"])
        check(old_vals.get("amount") == original_amount,
              f"Event records old amount {old_vals.get('amount')}")
        check(new_vals.get("amount") == 175.00,
              f"Event records new amount {new_vals.get('amount')}")

        # Locked: cannot re-categorize
        ok2 = ts.categorize(txn_id, "Payroll")
        check(not ok2, "Corrected transaction is locked — cannot re-categorize")
    finally:
        _cleanup(db_path)


def test_bulk_categorize():
    """Bulk categorize multiple transactions."""
    print("\n=== Bulk Categorize Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)
        txns = ts.get_transactions("Test Client", 2025)
        txn_ids = [t["txn_id"] for t in txns["items"]]

        count = ts.bulk_categorize(txn_ids, "Office Supplies", reviewer="Jeffrey")
        check(count == 3, f"Bulk categorized 3 transactions (got {count})")

        # All should have the category
        for tid in txn_ids:
            txn = ts.get_transaction(tid)
            check(txn["category"] == "Office Supplies",
                  f"Txn {tid[:8]} category is 'Office Supplies'")
            break  # Just check the first
    finally:
        _cleanup(db_path)


def test_bulk_categorize_skips_locked():
    """Bulk categorize skips locked transactions."""
    print("\n=== Bulk Categorize Lock Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)
        txns = ts.get_transactions("Test Client", 2025)
        txn_ids = [t["txn_id"] for t in txns["items"]]

        # Lock one
        ts.correct(txn_ids[0], {"amount": 999.99, "category": "Utilities"})

        count = ts.bulk_categorize(txn_ids, "Payroll")
        check(count == 2, f"Bulk categorized 2 (skipped 1 locked) — got {count}")
    finally:
        _cleanup(db_path)


def test_vendor_rules():
    """Add, use, and delete vendor rules."""
    print("\n=== Vendor Rules Tests ===")
    ts, db_path = _make_ts()
    try:
        # Add exact rule
        rule_id = ts.add_vendor_rule("GEORGIA POWER", "exact", "Utilities", "Jeffrey")
        check(rule_id is not None, f"Created vendor rule (id={rule_id})")

        # Query rules
        rules = ts.get_vendor_rules()
        check(rules["total"] == 1, f"1 vendor rule exists (got {rules['total']})")
        check(rules["rules"][0]["vendor_pattern"] == "GEORGIA POWER", "Pattern matches")

        # Delete rule
        ok = ts.delete_vendor_rule(rule_id)
        check(ok, "Deleted vendor rule")

        rules = ts.get_vendor_rules()
        check(rules["total"] == 0, "0 vendor rules after delete")
    finally:
        _cleanup(db_path)


def test_category_rules():
    """Add and delete category rules."""
    print("\n=== Category Rules Tests ===")
    ts, db_path = _make_ts()
    try:
        rule_id = ts.add_category_rule("ELECTRIC", "Utilities", priority=50, created_by="Jeffrey")
        check(rule_id is not None, f"Created category rule (id={rule_id})")

        rules = ts.get_category_rules()
        check(len(rules) == 1, f"1 category rule exists (got {len(rules)})")
        check(rules[0]["keyword"] == "ELECTRIC", "Keyword matches")
        check(rules[0]["priority"] == 50, "Priority matches")

        ok = ts.delete_category_rule(rule_id)
        check(ok, "Deleted category rule")
    finally:
        _cleanup(db_path)


def test_apply_rules_exact_match():
    """Rules engine: exact vendor match."""
    print("\n=== Rules Engine (Exact) Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)
        # GEORGIA POWER COMPANY #12345 normalizes to GEORGIA POWER
        ts.add_vendor_rule("GEORGIA POWER", "exact", "Utilities")

        result = ts.apply_vendor_rules("Test Client", 2025)
        check(result["matched"] >= 1, f"At least 1 matched by exact rule (got {result['matched']})")

        # Verify the Georgia Power transaction got categorized
        txns = ts.get_transactions("Test Client", 2025, filters={"category": "Utilities"})
        check(txns["total"] >= 1, "Georgia Power transaction categorized as Utilities")

        # Status should be 'suggested', not 'verified'
        for item in txns["items"]:
            if item["category"] == "Utilities":
                check(item["status"] == "suggested",
                      f"Status is 'suggested' (not verified) — got {item['status']}")
                break
    finally:
        _cleanup(db_path)


def test_apply_rules_prefix_match():
    """Rules engine: prefix vendor match."""
    print("\n=== Rules Engine (Prefix) Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)
        ts.add_vendor_rule("WAL-MART", "prefix", "Office Supplies")

        result = ts.apply_vendor_rules("Test Client", 2025)
        txns = ts.get_transactions("Test Client", 2025, filters={"category": "Office Supplies"})
        check(txns["total"] >= 1, "WAL-MART transaction matched by prefix rule")
    finally:
        _cleanup(db_path)


def test_apply_rules_contains_match():
    """Rules engine: contains vendor match."""
    print("\n=== Rules Engine (Contains) Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)
        ts.add_vendor_rule("DEPOSIT", "contains", "Sales/Revenue")

        result = ts.apply_vendor_rules("Test Client", 2025)
        txns = ts.get_transactions("Test Client", 2025, filters={"category": "Sales/Revenue"})
        check(txns["total"] >= 1, "DIRECT DEPOSIT matched by contains rule")
    finally:
        _cleanup(db_path)


def test_apply_rules_keyword_match():
    """Rules engine: keyword category rule match."""
    print("\n=== Rules Engine (Keyword) Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)
        ts.add_category_rule("POWER", "Utilities", priority=50)

        result = ts.apply_vendor_rules("Test Client", 2025)
        txns = ts.get_transactions("Test Client", 2025, filters={"category": "Utilities"})
        check(txns["total"] >= 1, "GEORGIA POWER matched by keyword rule 'POWER'")
    finally:
        _cleanup(db_path)


def test_apply_rules_hierarchy():
    """Rules engine: exact rule takes priority over prefix and keyword."""
    print("\n=== Rules Engine (Hierarchy) Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)
        # Add conflicting rules — exact should win
        ts.add_vendor_rule("GEORGIA POWER", "exact", "Utilities")
        ts.add_category_rule("POWER", "Rent/Lease", priority=1)  # Lower prio keyword
        ts.add_vendor_rule("GEORGIA", "prefix", "Professional Fees")  # Prefix

        result = ts.apply_vendor_rules("Test Client", 2025)

        # Georgia Power should be Utilities (exact match wins)
        txns = ts.get_transactions("Test Client", 2025)
        georgia_txn = None
        for t in txns["items"]:
            if "GEORGIA" in (t["vendor_norm"] or "").upper():
                georgia_txn = t
                break
        check(georgia_txn is not None, "Found Georgia Power transaction")
        if georgia_txn:
            check(georgia_txn["category"] == "Utilities",
                  f"Exact rule wins: category is 'Utilities' (got {georgia_txn['category']})")
    finally:
        _cleanup(db_path)


def test_apply_rules_skips_locked():
    """Rules engine skips locked transactions."""
    print("\n=== Rules Engine (Skip Locked) Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)
        txns = ts.get_transactions("Test Client", 2025)
        first_id = txns["items"][0]["txn_id"]

        # Lock one
        ts.correct(first_id, {"category": "Payroll"})

        # Add rule
        ts.add_vendor_rule("GEORGIA POWER", "exact", "Utilities")
        result = ts.apply_vendor_rules("Test Client", 2025)

        # The locked one should still be Payroll
        txn = ts.get_transaction(first_id)
        check(txn["category"] == "Payroll", f"Locked txn category unchanged (got {txn['category']})")
    finally:
        _cleanup(db_path)


def test_learn_vendor_rule():
    """Learn vendor rule increments usage count."""
    print("\n=== Learn Vendor Rule Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.learn_vendor_rule("GEORGIA POWER", "Utilities", source="learned")
        ts.learn_vendor_rule("GEORGIA POWER", "Utilities", source="learned")
        ts.learn_vendor_rule("GEORGIA POWER", "Utilities", source="learned")

        rules = ts.get_vendor_rules()
        check(rules["total"] == 1, "Only 1 rule (upsert, not duplicate)")
        check(rules["rules"][0]["usage_count"] == 3,
              f"Usage count is 3 (got {rules['rules'][0]['usage_count']})")
    finally:
        _cleanup(db_path)


def test_get_transactions_pagination():
    """Pagination in get_transactions."""
    print("\n=== Pagination Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)

        result = ts.get_transactions("Test Client", 2025, page=1, per_page=2)
        check(result["total"] == 3, f"Total is 3 (got {result['total']})")
        check(len(result["items"]) == 2, f"Page 1 has 2 items (got {len(result['items'])})")
        check(result["pages"] == 2, f"Total pages is 2 (got {result['pages']})")

        result2 = ts.get_transactions("Test Client", 2025, page=2, per_page=2)
        check(len(result2["items"]) == 1, f"Page 2 has 1 item (got {len(result2['items'])})")
    finally:
        _cleanup(db_path)


def test_get_transactions_filter_status():
    """Filter transactions by status."""
    print("\n=== Filter by Status Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)
        txns = ts.get_transactions("Test Client", 2025)
        first_id = txns["items"][0]["txn_id"]

        ts.categorize(first_id, "Utilities")
        ts.verify(first_id)

        staged = ts.get_transactions("Test Client", 2025, filters={"status": "staged"})
        check(staged["total"] == 2, f"2 staged transactions (got {staged['total']})")

        verified = ts.get_transactions("Test Client", 2025, filters={"status": "verified"})
        check(verified["total"] == 1, f"1 verified transaction (got {verified['total']})")
    finally:
        _cleanup(db_path)


def test_get_transactions_filter_month():
    """Filter transactions by month."""
    print("\n=== Filter by Month Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)

        jan = ts.get_transactions("Test Client", 2025, filters={"month": 1})
        check(jan["total"] == 3, f"3 transactions in January (got {jan['total']})")

        feb = ts.get_transactions("Test Client", 2025, filters={"month": 2})
        check(feb["total"] == 0, f"0 transactions in February (got {feb['total']})")
    finally:
        _cleanup(db_path)


def test_get_transactions_filter_search():
    """Search transactions by description."""
    print("\n=== Search Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)

        results = ts.get_transactions("Test Client", 2025, filters={"search": "GEORGIA"})
        check(results["total"] == 1, f"1 transaction matching 'GEORGIA' (got {results['total']})")

        results2 = ts.get_transactions("Test Client", 2025, filters={"search": "MART"})
        check(results2["total"] == 1, f"1 transaction matching 'MART' (got {results2['total']})")
    finally:
        _cleanup(db_path)


def test_get_uncategorized():
    """Get uncategorized transactions."""
    print("\n=== Uncategorized Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)

        uncat = ts.get_uncategorized("Test Client", 2025)
        check(len(uncat) == 3, f"3 uncategorized initially (got {len(uncat)})")

        # Categorize one
        ts.categorize(uncat[0]["txn_id"], "Utilities")
        uncat2 = ts.get_uncategorized("Test Client", 2025)
        check(len(uncat2) == 2, f"2 uncategorized after categorizing 1 (got {len(uncat2)})")
    finally:
        _cleanup(db_path)


def test_monthly_summary():
    """Get monthly summary pivot data."""
    print("\n=== Monthly Summary Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)
        txns = ts.get_transactions("Test Client", 2025)

        # Categorize all
        for item in txns["items"]:
            ts.categorize(item["txn_id"], "Utilities")

        summary = ts.get_monthly_summary("Test Client", 2025)
        check("Utilities" in summary["categories"], "Utilities in category summary")
        check("1" in summary["monthly_totals"], "January in monthly totals")
        check(summary["grand_total"] > 0, f"Grand total > 0 (got {summary['grand_total']})")
    finally:
        _cleanup(db_path)


def test_count_by_status():
    """Count transactions by status."""
    print("\n=== Count by Status Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)

        counts = ts.count_by_status("Test Client", 2025)
        check(counts["staged"] == 3, f"3 staged (got {counts['staged']})")
        check(counts["total"] == 3, f"Total 3 (got {counts['total']})")
        check(counts["uncategorized"] == 3, f"3 uncategorized (got {counts['uncategorized']})")
    finally:
        _cleanup(db_path)


def test_clients_with_transactions():
    """Get distinct clients."""
    print("\n=== Clients with Transactions Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Client A", 2025)
        ts.ingest_from_extraction("job-002", _sample_cc_log(), "Client B", 2025)

        clients = ts.get_clients_with_transactions()
        check(len(clients) == 2, f"2 distinct clients (got {len(clients)})")
        check("Client A" in clients, "Client A in list")
        check("Client B" in clients, "Client B in list")
    finally:
        _cleanup(db_path)


# ── Transaction Extract Bridge Tests ─────────────────────────────────────────

def test_extract_bank_statement():
    """Parse bank statement transactions from log."""
    print("\n=== Extract Bridge (Bank) Tests ===")
    from transaction_extract import parse_transactions_from_log

    txns = list(parse_transactions_from_log(_sample_bank_log()))
    check(len(txns) == 3, f"Parsed 3 bank transactions (got {len(txns)})")

    first = txns[0]
    check(first["document_type"] == "bank_statement", "Document type is bank_statement")
    check(first["txn_index"] == 1, f"First txn_index is 1 (got {first['txn_index']})")
    check(first["txn_date"] == "01/15/2025", f"Date is 01/15/2025 (got {first['txn_date']})")
    check(first["amount"] == 150.00, f"Amount is 150.00 (got {first['amount']})")
    check(first["payer_entity"] == "First National Bank", "Payer entity is First National Bank")


def test_extract_credit_card():
    """Parse credit card transactions from log."""
    print("\n=== Extract Bridge (CC) Tests ===")
    from transaction_extract import parse_transactions_from_log

    txns = list(parse_transactions_from_log(_sample_cc_log()))
    check(len(txns) == 2, f"Parsed 2 CC transactions (got {len(txns)})")

    first = txns[0]
    check(first["amount"] == 45.99, f"Amount is 45.99 (got {first['amount']})")
    check(first["category"] == "Office Supplies", f"Category pre-set (got {first['category']})")


def test_extract_check():
    """Parse check document from log."""
    print("\n=== Extract Bridge (Check) Tests ===")
    from transaction_extract import parse_transactions_from_log

    txns = list(parse_transactions_from_log(_sample_check_log()))
    check(len(txns) == 1, f"Parsed 1 check transaction (got {len(txns)})")

    chk = txns[0]
    check(chk["txn_type"] == "check", f"Type is 'check' (got {chk['txn_type']})")
    check(chk["amount"] == 250.00, f"Amount is 250.00 (got {chk['amount']})")
    check("1234" in chk["description"], f"Description has check number (got {chk['description']})")


def test_extract_ignores_tax():
    """Tax documents produce no transactions."""
    print("\n=== Extract Bridge (Skip Tax) Tests ===")
    from transaction_extract import parse_transactions_from_log

    txns = list(parse_transactions_from_log(_sample_tax_log()))
    check(len(txns) == 0, "No transactions from W-2")


def test_extract_value_helpers():
    """Helper functions handle various field formats."""
    print("\n=== Extract Value Helpers Tests ===")
    from transaction_extract import _extract_value, _extract_numeric

    # Dict format
    check(_extract_value({"value": "hello"}) == "hello", "Dict value extracted")
    check(_extract_value({"value": 123}) == "123", "Dict numeric value extracted as string")
    check(_extract_value(None) == "", "None returns empty string")
    check(_extract_value("bare") == "bare", "Bare string returned as-is")

    # Numeric
    check(_extract_numeric({"value": "1,234.56"}) == 1234.56, "Strips commas")
    check(_extract_numeric({"value": "$99.00"}) == 99.00, "Strips dollar sign")
    check(_extract_numeric({"value": "(500.00)"}) == -500.00, "Parenthesized negative")
    check(_extract_numeric(None) is None, "None returns None")
    check(_extract_numeric({"value": ""}) is None, "Empty string returns None")


def test_vendor_normalize():
    """Vendor normalization function."""
    print("\n=== Vendor Normalize Tests ===")
    from transaction_store import normalize_vendor

    check(normalize_vendor("GEORGIA POWER COMPANY #12345") == "GEORGIA POWER",
          "Strips company suffix and store number")
    check(normalize_vendor("WAL-MART SUPER CENTER 0423") == "WAL-MART SUPER CENTER",
          "Strips trailing store number")
    check(normalize_vendor("ACME INC.") == "ACME",
          "Strips INC. suffix")
    check(normalize_vendor("") == "", "Empty string returns empty")
    check(normalize_vendor(None) == "", "None returns empty")


def test_category_taxonomy():
    """Category taxonomy is well-formed."""
    print("\n=== Category Taxonomy Tests ===")
    from transaction_store import (CATEGORY_TAXONOMY, ALL_TXN_CATEGORIES,
                                    CATEGORY_TO_GROUP)

    check(len(CATEGORY_TAXONOMY) == 10, f"10 top-level groups (got {len(CATEGORY_TAXONOMY)})")
    check(len(ALL_TXN_CATEGORIES) > 25, f"More than 25 categories (got {len(ALL_TXN_CATEGORIES)})")
    check("Utilities" in ALL_TXN_CATEGORIES, "Utilities in all categories")
    check(CATEGORY_TO_GROUP["Utilities"] == "Operating Expenses > Facilities",
          "Utilities maps to correct group")
    check("Uncategorized" in ALL_TXN_CATEGORIES, "Uncategorized exists")

    # No duplicates
    check(len(ALL_TXN_CATEGORIES) == len(set(ALL_TXN_CATEGORIES)),
          "No duplicate categories in taxonomy")


# ── Report Generation Tests ──────────────────────────────────────────────────

def test_report_import_guardrail():
    """transaction_reports.py does not import forbidden modules."""
    print("\n=== Report Import Guardrail Tests ===")
    import transaction_reports
    source_file = transaction_reports.__file__
    with open(source_file) as f:
        source = f.read()
    for forbidden in transaction_reports._FORBIDDEN_MODULES:
        has_import = (f"import {forbidden}" in source and
                      f"'{forbidden}'" not in source and
                      f'"{forbidden}"' not in source)
        check(not has_import,
              f"transaction_reports does not import {forbidden}")


def test_report_builds_xlsx():
    """Report generates a valid xlsx file."""
    print("\n=== Report Build Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)
        # Categorize all so summary has data
        txns = ts.get_transactions("Test Client", 2025)
        for item in txns["items"]:
            ts.categorize(item["txn_id"], "Utilities")

        from transaction_reports import TransactionReportBuilder
        import openpyxl

        fd2, xlsx_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd2)

        builder = TransactionReportBuilder(ts, "Test Client", 2025)
        result_path = builder.build(xlsx_path)
        check(os.path.exists(result_path), "xlsx file created")

        # Open and verify
        wb = openpyxl.load_workbook(xlsx_path)
        sheet_names = wb.sheetnames
        check("Monthly Summary" in sheet_names, "Monthly Summary sheet exists")
        check("Transaction Detail" in sheet_names, "Transaction Detail sheet exists")
        check("Vendor Summary" in sheet_names, "Vendor Summary sheet exists")
        check(len(sheet_names) == 3, f"Exactly 3 sheets (got {len(sheet_names)})")
        wb.close()

        os.unlink(xlsx_path)
    finally:
        _cleanup(db_path)


def test_report_summary_has_months():
    """Summary sheet has month columns."""
    print("\n=== Report Summary Month Columns Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)
        txns = ts.get_transactions("Test Client", 2025)
        for item in txns["items"]:
            ts.categorize(item["txn_id"], "Utilities")

        from transaction_reports import TransactionReportBuilder
        import openpyxl

        fd2, xlsx_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd2)

        builder = TransactionReportBuilder(ts, "Test Client", 2025)
        builder.build(xlsx_path)

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Monthly Summary"]

        # Find header row (row 4 based on title block)
        header_vals = [ws.cell(row=4, column=c).value for c in range(1, 15)]
        check("Category" in header_vals, f"Category header found")
        check("Jan" in header_vals, f"Jan header found")
        check("Dec" in header_vals, f"Dec header found")
        check("Total" in header_vals, f"Total header found")
        wb.close()
        os.unlink(xlsx_path)
    finally:
        _cleanup(db_path)


def test_report_detail_sorted_by_date():
    """Detail sheet transactions are sorted by date."""
    print("\n=== Report Detail Sort Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)

        from transaction_reports import TransactionReportBuilder
        import openpyxl

        fd2, xlsx_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd2)

        builder = TransactionReportBuilder(ts, "Test Client", 2025)
        builder.build(xlsx_path)

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Transaction Detail"]

        # Read dates from data rows (after header at row 4, data starts row 5)
        dates = []
        for row in range(5, ws.max_row + 1):
            d = ws.cell(row=row, column=1).value
            if d:
                dates.append(str(d))

        check(len(dates) == 3, f"3 transaction rows (got {len(dates)})")
        check(dates == sorted(dates), "Dates are sorted ascending")
        wb.close()
        os.unlink(xlsx_path)
    finally:
        _cleanup(db_path)


def test_report_vendor_sheet():
    """Vendor sheet has data."""
    print("\n=== Report Vendor Sheet Tests ===")
    ts, db_path = _make_ts()
    try:
        ts.ingest_from_extraction("job-001", _sample_bank_log(), "Test Client", 2025)

        from transaction_reports import TransactionReportBuilder
        import openpyxl

        fd2, xlsx_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd2)

        builder = TransactionReportBuilder(ts, "Test Client", 2025)
        builder.build(xlsx_path)

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Vendor Summary"]

        # Header row should have Vendor
        header_vals = [ws.cell(row=4, column=c).value for c in range(1, 5)]
        check("Vendor" in header_vals, "Vendor header found")
        check("Category" in header_vals, "Category header found")
        check("Total Amount" in header_vals, "Total Amount header found")

        # Should have data rows
        check(ws.max_row > 5, f"Has vendor data rows (max_row={ws.max_row})")
        wb.close()
        os.unlink(xlsx_path)
    finally:
        _cleanup(db_path)


# ═════════════════════════════════════════════════════════════════════════════
# RUNNER
# ═════════════════════════════════════════════════════════════════════════════

def run_tests():
    global PASS, FAIL

    # Import guardrails
    test_import_guardrails()

    # Schema
    test_schema()

    # txn_id
    test_txn_id()

    # Ingest
    test_ingest_bank_statement()
    test_ingest_credit_card()
    test_ingest_check()
    test_ingest_dedup()
    test_ingest_skips_tax_docs()

    # Categorize
    test_categorize()
    test_categorize_refuses_invalid()
    test_categorize_refuses_locked()

    # Verify
    test_verify()
    test_verify_refuses_uncategorized()

    # Correct
    test_correct()

    # Bulk
    test_bulk_categorize()
    test_bulk_categorize_skips_locked()

    # Rules engine
    test_vendor_rules()
    test_category_rules()
    test_apply_rules_exact_match()
    test_apply_rules_prefix_match()
    test_apply_rules_contains_match()
    test_apply_rules_keyword_match()
    test_apply_rules_hierarchy()
    test_apply_rules_skips_locked()
    test_learn_vendor_rule()

    # Queries
    test_get_transactions_pagination()
    test_get_transactions_filter_status()
    test_get_transactions_filter_month()
    test_get_transactions_filter_search()
    test_get_uncategorized()
    test_monthly_summary()
    test_count_by_status()
    test_clients_with_transactions()

    # Extract bridge
    test_extract_bank_statement()
    test_extract_credit_card()
    test_extract_check()
    test_extract_ignores_tax()
    test_extract_value_helpers()

    # Taxonomy + normalize
    test_vendor_normalize()
    test_category_taxonomy()

    # Report generation
    test_report_import_guardrail()
    test_report_builds_xlsx()
    test_report_summary_has_months()
    test_report_detail_sorted_by_date()
    test_report_vendor_sheet()

    print(f"\n{'='*60}")
    print(f"  Results: {PASS} passed, {FAIL} failed, {PASS + FAIL} total")
    print(f"{'='*60}")
    if FAIL > 0:
        sys.exit(1)


if __name__ == "__main__":
    run_tests()
