"""
Transaction Report Builder — T-TXN-LEDGER-1
=============================================
Generates monthly category/vendor summary Excel workbooks from the
TransactionStore. Reads ONLY from TransactionStore (SQLite).

ARCHITECTURAL RULE: This module must NEVER import extract.py, OCR,
vision, or PDF libraries. It reads only from TransactionStore.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from datetime import datetime

from transaction_store import (
    TransactionStore, CATEGORY_TAXONOMY, CATEGORY_TO_GROUP,
    ALL_TXN_CATEGORIES,
)

# ── Forbidden-module guardrail ───────────────────────────────────────────────
_FORBIDDEN_MODULES = frozenset({
    'extract', 'pytesseract', 'anthropic', 'pdf2image',
    'PIL', 'Pillow', 'fitz',
})

# ── Styling Constants (consistent with workpaper_export.py) ──────────────────

TITLE_FONT = Font(bold=True, size=14, color="1A252F")
SUBTITLE_FONT = Font(italic=True, color="888888", size=9)
SECTION_FONT = Font(bold=True, size=11, color="000000")
SECTION_FILL = PatternFill("solid", fgColor="D9D9D9")
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
DATA_FONT = Font(size=10)
TOTAL_FONT = Font(bold=True, size=10)
TOTAL_FILL = PatternFill("solid", fgColor="E8E8E8")
ALT_ROW_FILL = PatternFill("solid", fgColor="F7F7F7")
MONEY_FMT = '#,##0.00_);(#,##0.00)'
DATE_FMT = 'MM/DD/YYYY'
THIN_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)

MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# Status color fills
STATUS_FILLS = {
    "verified":  PatternFill("solid", fgColor="C8E6C9"),   # Green
    "corrected": PatternFill("solid", fgColor="BBDEFB"),   # Blue
    "suggested": PatternFill("solid", fgColor="FFF9C4"),   # Yellow
    "staged":    PatternFill("solid", fgColor="FFCDD2"),   # Red
}


class TransactionReportBuilder:
    """Generate monthly transaction summary Excel reports.

    Creates a workbook with three sheets:
      1. Monthly Summary — category × month pivot table
      2. Transaction Detail — all transactions sorted by date
      3. Vendor Summary — vendor × category × count × total

    Usage:
        from transaction_store import TransactionStore
        ts = TransactionStore("data/bearden.db")
        builder = TransactionReportBuilder(ts, "Client Name", 2025)
        builder.build("output/Client-txn-summary-2025.xlsx")
    """

    def __init__(self, store, client_name, year):
        if not isinstance(store, TransactionStore):
            raise TypeError("Requires TransactionStore instance")
        self.store = store
        self.client = client_name
        self.year = int(year)

    def build(self, output_path):
        """Generate the report. Creates three sheets."""
        if not str(output_path).lower().endswith('.xlsx'):
            raise ValueError("Output path must end with .xlsx")

        wb = openpyxl.Workbook()

        # Sheet 1: Monthly Summary
        ws_summary = wb.active
        ws_summary.title = "Monthly Summary"
        self._write_summary_sheet(ws_summary)

        # Sheet 2: Transaction Detail
        ws_detail = wb.create_sheet("Transaction Detail")
        self._write_detail_sheet(ws_detail)

        # Sheet 3: Vendor Summary
        ws_vendor = wb.create_sheet("Vendor Summary")
        self._write_vendor_sheet(ws_vendor)

        wb.save(output_path)
        return output_path

    # ── Sheet 1: Monthly Summary ─────────────────────────────────────────────

    def _write_summary_sheet(self, ws):
        """Write category × month pivot table."""
        summary = self.store.get_monthly_summary(self.client, self.year)
        categories = summary["categories"]
        monthly_totals = summary["monthly_totals"]
        category_totals = summary["category_totals"]
        grand_total = summary["grand_total"]

        # Title block
        row = 1
        ws.cell(row=row, column=1,
                value=f"Transaction Summary: {self.client}").font = TITLE_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
        row += 1
        ws.cell(row=row, column=1,
                value=f"Tax Year {self.year} — Generated {datetime.now().strftime('%m/%d/%Y %I:%M %p')}"
                ).font = SUBTITLE_FONT
        row += 2

        # Header row: Category | Jan | Feb | ... | Dec | Total
        headers = ["Category"] + MONTH_NAMES + ["Total"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        row += 1
        data_start_row = row

        # Group by taxonomy
        for group_name, group_cats in CATEGORY_TAXONOMY.items():
            # Check if any categories in this group have data
            group_has_data = any(cat in categories for cat in group_cats)
            if not group_has_data:
                continue

            # Group header row
            cell = ws.cell(row=row, column=1, value=group_name)
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
            for col in range(2, len(headers) + 1):
                ws.cell(row=row, column=col).fill = SECTION_FILL
            row += 1

            # Category rows
            for cat in group_cats:
                if cat not in categories:
                    continue

                ws.cell(row=row, column=1, value=f"  {cat}").font = DATA_FONT
                cat_data = categories[cat]
                cat_total = 0

                for m in range(1, 13):
                    month_key = str(m)
                    val = cat_data.get(month_key, 0)
                    col_idx = m + 1
                    if val:
                        cell = ws.cell(row=row, column=col_idx, value=val)
                        cell.number_format = MONEY_FMT
                    cat_total += val

                # Total column
                total_cell = ws.cell(row=row, column=14, value=cat_total)
                total_cell.number_format = MONEY_FMT
                total_cell.font = TOTAL_FONT

                # Alternating row fill
                if (row - data_start_row) % 2 == 1:
                    for c in range(1, len(headers) + 1):
                        cell = ws.cell(row=row, column=c)
                        if cell.fill == PatternFill():
                            cell.fill = ALT_ROW_FILL
                row += 1

        # Grand total row
        row += 1
        ws.cell(row=row, column=1, value="GRAND TOTAL").font = TOTAL_FONT
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).fill = TOTAL_FILL
        for m in range(1, 13):
            val = monthly_totals.get(str(m), 0)
            if val:
                cell = ws.cell(row=row, column=m + 1, value=val)
                cell.number_format = MONEY_FMT
                cell.font = TOTAL_FONT
        total_cell = ws.cell(row=row, column=14, value=grand_total)
        total_cell.number_format = MONEY_FMT
        total_cell.font = TOTAL_FONT

        # Column widths
        ws.column_dimensions["A"].width = 35
        for col in range(2, 15):
            ws.column_dimensions[get_column_letter(col)].width = 12

        # Print setup
        ws.freeze_panes = f"B{data_start_row}"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1

    # ── Sheet 2: Transaction Detail ──────────────────────────────────────────

    def _write_detail_sheet(self, ws):
        """Write all transactions sorted by date."""
        # Get ALL transactions (paginate through)
        all_txns = []
        page = 1
        while True:
            result = self.store.get_transactions(
                self.client, self.year, page=page, per_page=500
            )
            all_txns.extend(result["items"])
            if page >= result["pages"]:
                break
            page += 1

        # Sort by date
        all_txns.sort(key=lambda t: t.get("txn_date", ""))

        # Title
        row = 1
        ws.cell(row=row, column=1,
                value=f"Transaction Detail: {self.client}").font = TITLE_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1
        ws.cell(row=row, column=1,
                value=f"Tax Year {self.year} — {len(all_txns)} transactions"
                ).font = SUBTITLE_FONT
        row += 2

        # Headers
        headers = ["Date", "Description", "Amount", "Type", "Category",
                    "Vendor", "Status", "Source"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        row += 1
        data_start_row = row

        # Data rows
        for idx, txn in enumerate(all_txns):
            ws.cell(row=row, column=1, value=txn.get("txn_date", ""))
            ws.cell(row=row, column=2, value=txn.get("description", ""))

            amt_cell = ws.cell(row=row, column=3, value=txn.get("amount"))
            amt_cell.number_format = MONEY_FMT

            ws.cell(row=row, column=4, value=txn.get("txn_type", ""))
            ws.cell(row=row, column=5, value=txn.get("category", ""))
            ws.cell(row=row, column=6, value=txn.get("vendor_norm", ""))

            status = txn.get("status", "")
            status_cell = ws.cell(row=row, column=7, value=status)
            if status in STATUS_FILLS:
                status_cell.fill = STATUS_FILLS[status]

            ws.cell(row=row, column=8, value=txn.get("document_type", ""))

            # Alternating rows
            if idx % 2 == 1:
                for c in range(1, 9):
                    cell = ws.cell(row=row, column=c)
                    if cell.fill == PatternFill():
                        cell.fill = ALT_ROW_FILL

            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 22
        ws.column_dimensions["F"].width = 25
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 18

        ws.freeze_panes = f"A{data_start_row}"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1

    # ── Sheet 3: Vendor Summary ──────────────────────────────────────────────

    def _write_vendor_sheet(self, ws):
        """Write vendor summary: Vendor, Category, Count, Total Amount."""
        # Get all transactions and aggregate by vendor
        all_txns = []
        page = 1
        while True:
            result = self.store.get_transactions(
                self.client, self.year, page=page, per_page=500
            )
            all_txns.extend(result["items"])
            if page >= result["pages"]:
                break
            page += 1

        # Aggregate: vendor → {category, count, total}
        vendor_agg = {}
        for txn in all_txns:
            vendor = txn.get("vendor_norm", "") or "(no vendor)"
            cat = txn.get("category", "") or "Uncategorized"
            key = (vendor, cat)

            if key not in vendor_agg:
                vendor_agg[key] = {"vendor": vendor, "category": cat,
                                   "count": 0, "total": 0.0}
            vendor_agg[key]["count"] += 1
            vendor_agg[key]["total"] += abs(txn.get("amount", 0) or 0)

        # Sort by vendor, then category
        sorted_vendors = sorted(vendor_agg.values(),
                                key=lambda v: (v["vendor"], v["category"]))

        # Title
        row = 1
        ws.cell(row=row, column=1,
                value=f"Vendor Summary: {self.client}").font = TITLE_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
        ws.cell(row=row, column=1,
                value=f"Tax Year {self.year} — {len(sorted_vendors)} vendor/category combinations"
                ).font = SUBTITLE_FONT
        row += 2

        # Headers
        headers = ["Vendor", "Category", "Count", "Total Amount"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        row += 1

        # Data
        for idx, v in enumerate(sorted_vendors):
            ws.cell(row=row, column=1, value=v["vendor"])
            ws.cell(row=row, column=2, value=v["category"])
            ws.cell(row=row, column=3, value=v["count"])
            total_cell = ws.cell(row=row, column=4, value=v["total"])
            total_cell.number_format = MONEY_FMT

            if idx % 2 == 1:
                for c in range(1, 5):
                    cell = ws.cell(row=row, column=c)
                    if cell.fill == PatternFill():
                        cell.fill = ALT_ROW_FILL
            row += 1

        # Grand total
        row += 1
        ws.cell(row=row, column=1, value="TOTAL").font = TOTAL_FONT
        ws.cell(row=row, column=3, value=sum(v["count"] for v in sorted_vendors)).font = TOTAL_FONT
        total_amt = ws.cell(row=row, column=4,
                            value=sum(v["total"] for v in sorted_vendors))
        total_amt.number_format = MONEY_FMT
        total_amt.font = TOTAL_FONT
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = TOTAL_FILL

        # Column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 16
