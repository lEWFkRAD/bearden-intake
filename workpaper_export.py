"""Deterministic workpaper generator for the Bearden Document Intake Platform.

Generates professional accountant-style workpapers from the fact store.
Reads ONLY from the FactStore (SQLite). Never accesses PDFs, images,
OCR, or extraction logic.

ARCHITECTURAL RULE: This module must NEVER import extract.py, OCR,
vision, or PDF libraries. The MAPPING_REGISTRY is a standalone copy
of the field layout derived from extract.py's TEMPLATE_SECTIONS.

GUARDRAIL A: The builder accepts only identifiers (client_name, year).
It is impossible to pass PDF paths, image objects, or raw OCR text.

GUARDRAIL B: No imports from the extraction stack. Tested by
tests/test_workpaper.py.

Usage::

    from fact_store import FactStore
    from workpaper_export import WorkpaperBuilder

    fs = FactStore("data/bearden.db")
    builder = WorkpaperBuilder(fs, "Evans, Lisa", "2025", mode="assisted")
    builder.build("output/Evans-workpaper-2025.xlsx")
"""

import json
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, Protection,
)
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

from fact_store import FactStore

# ─── IMPORT GUARDRAIL ─────────────────────────────────────────────────────────
_FORBIDDEN_MODULES = frozenset({
    'extract', 'pytesseract', 'anthropic', 'pdf2image',
    'PIL', 'Pillow', 'fitz',
})

# ─── RUNTIME GUARDRAIL ────────────────────────────────────────────────────────

def _validate_identifier(value, param_name):
    """Reject non-string, file paths, binary, and oversized inputs.

    Workpaper identifiers must be short strings (client names, years).
    Anything that looks like a file path, binary blob, or OCR dump is rejected.
    """
    if not isinstance(value, str):
        raise TypeError(
            f"{param_name} must be a string, got {type(value).__name__}"
        )
    if not value.strip():
        raise ValueError(f"{param_name} must not be empty")
    lowered = value.strip().lower()
    if lowered.endswith(('.pdf', '.png', '.jpg', '.jpeg', '.tiff', '.tif')):
        raise ValueError(
            f"{param_name} looks like a file path ({value!r}) — "
            "WorkpaperBuilder accepts only identifiers"
        )
    if os.sep in value and len(value) > 60:
        raise ValueError(
            f"{param_name} looks like a file path ({value!r}) — "
            "WorkpaperBuilder accepts only identifiers"
        )
    if len(value) > 200:
        raise ValueError(
            f"{param_name} is too long ({len(value)} chars) — "
            "WorkpaperBuilder accepts only short identifiers"
        )


# ─── STATUS COLORS ────────────────────────────────────────────────────────────

STATUS_FILLS = {
    "auto_verified": PatternFill("solid", fgColor="C8E6C9"),   # Green
    "confirmed":     PatternFill("solid", fgColor="C8E6C9"),   # Green
    "dual_confirmed": PatternFill("solid", fgColor="A5D6A7"),  # Darker green
    "consensus_accepted": PatternFill("solid", fgColor="C8E6C9"),  # Green
    "verified_confirmed": PatternFill("solid", fgColor="C8E6C9"),  # Green
    "verified_corrected": PatternFill("solid", fgColor="BBDEFB"),  # Blue
    "extracted":     PatternFill("solid", fgColor="FFF9C4"),   # Yellow
    "needs_review":  PatternFill("solid", fgColor="FFCDD2"),   # Red/pink
    "corrected":     PatternFill("solid", fgColor="BBDEFB"),   # Blue
    "flagged":       PatternFill("solid", fgColor="FFE0B2"),   # Orange
    "missing":       PatternFill("solid", fgColor="F5F5F5"),   # Light gray
}

# ─── STYLING CONSTANTS ────────────────────────────────────────────────────────

TITLE_FONT = Font(bold=True, size=14, color="1A252F")
SUBTITLE_FONT = Font(italic=True, color="888888", size=9)
SECTION_FONT = Font(bold=True, size=11, color="000000")
SECTION_FILL = PatternFill("solid", fgColor="D9D9D9")
COL_HEADER_FONT = Font(size=10, color="000000", bold=True)
COL_HEADER_FILL = PatternFill("solid", fgColor="E8E8E8")
SUM_FONT = Font(bold=True, size=11, color="000000")
FLAG_FONT = Font(italic=True, color="CC0000", size=9)

MONEY_FMT = '#,##0.00_);(#,##0.00)'
PCT_FMT = '0.00%'
DATE_FMT = 'MM/DD/YYYY'

LOCKED = Protection(locked=True)
UNLOCKED = Protection(locked=False)

AUDIT_HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
AUDIT_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
AUDIT_ALT_FILL = PatternFill("solid", fgColor="F8F9FA")
THIN_BORDER = Border(bottom=Side(style="thin", color="E0E0E0"))

# Status labels considered "verified" for safe mode
VERIFIED_STATUSES = frozenset({
    "confirmed", "corrected", "auto_verified", "verified_confirmed",
    "verified_corrected", "dual_confirmed", "consensus_accepted",
    "multipage_verified",
})

# Sections always shown even when no matching facts exist
ALWAYS_SHOW = {"w2", "interest", "dividends", "schedule_d", "k1"}


# ═══════════════════════════════════════════════════════════════════════════════
# MAPPING REGISTRY
# ═══════════════════════════════════════════════════════════════════════════════
# Derived from extract.py TEMPLATE_SECTIONS (lines 1183-1496).
# Intentionally duplicated here so workpaper.py never imports extract.py.
# When TEMPLATE_SECTIONS changes, update this registry to match.

MAPPING_REGISTRY = [
    {
        "id": "w2",
        "header": "W-2:",
        "match_types": ["W-2"],
        "fields": {
            "employer_name": {"col": "A", "type": "label"},
            "wages":         {"col": "B", "type": "input", "fmt": "money"},
            "federal_wh":    {"col": "C", "type": "input", "fmt": "money"},
            "state_wh":      {"col": "D", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Gross", "C": "Federal WH", "D": "State WH"},
        "sum_cols": ["B", "C", "D"],
    },
    {
        "id": "interest",
        "header": "Interest Income:",
        "match_types": ["1099-INT", "_interest_rollup"],
        "fields": {
            "_source_name":                  {"col": "A", "type": "label"},
            "interest_income":               {"col": "B", "type": "input", "fmt": "money"},
            "us_savings_bonds_and_treasury": {"col": "C", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Interest", "C": "US Bonds"},
        "sum_cols": ["B", "C"],
        "total_formula_col": {"D": "B+C"},
    },
    {
        "id": "dividends",
        "header": "Dividends:",
        "match_types": ["1099-DIV", "_dividend_rollup"],
        "fields": {
            "_source_name":            {"col": "A", "type": "label"},
            "ordinary_dividends":      {"col": "B", "type": "input", "fmt": "money"},
            "qualified_dividends":     {"col": "C", "type": "input", "fmt": "money"},
            "capital_gain_distributions": {"col": "D", "type": "input", "fmt": "money"},
            "section_199a":            {"col": "E", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Total Ord", "C": "Qualified", "D": "Cap Gain Dist", "E": "Sec 199A"},
        "sum_cols": ["B", "C", "D", "E"],
    },
    {
        "id": "1099r",
        "header": "1099-R:",
        "match_types": ["1099-R"],
        "fields": {
            "payer_or_entity":   {"col": "A", "type": "label"},
            "gross_distribution": {"col": "B", "type": "input", "fmt": "money"},
            "taxable_amount":    {"col": "C", "type": "input", "fmt": "money"},
            "federal_wh":        {"col": "D", "type": "input", "fmt": "money"},
            "state_wh":          {"col": "E", "type": "input", "fmt": "money"},
            "distribution_code": {"col": "F", "type": "input", "fmt": "text"},
        },
        "col_headers": {"B": "Gross", "C": "Taxable", "D": "FWH", "E": "SWH", "F": "Code"},
        "sum_cols": ["B", "C", "D", "E"],
    },
    {
        "id": "ssa",
        "header": "SSA-1099:",
        "match_types": ["SSA-1099"],
        "fields": {
            "payer_or_entity": {"col": "A", "type": "label"},
            "net_benefits":    {"col": "B", "type": "input", "fmt": "money"},
            "federal_wh":      {"col": "C", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Net Benefits", "C": "Federal WH"},
        "sum_cols": ["B", "C"],
    },
    {
        "id": "1099nec",
        "header": "1099-NEC (Self-Employment):",
        "match_types": ["1099-NEC"],
        "fields": {
            "payer_or_entity":          {"col": "A", "type": "label"},
            "nonemployee_compensation": {"col": "B", "type": "input", "fmt": "money"},
            "federal_wh":               {"col": "C", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "NEC Income", "C": "FWH"},
        "sum_cols": ["B", "C"],
    },
    {
        "id": "1099misc",
        "header": "1099-MISC:",
        "match_types": ["1099-MISC"],
        "fields": {
            "payer_or_entity": {"col": "A", "type": "label"},
            "rents":           {"col": "B", "type": "input", "fmt": "money"},
            "royalties":       {"col": "C", "type": "input", "fmt": "money"},
            "other_income":    {"col": "D", "type": "input", "fmt": "money"},
            "federal_wh":      {"col": "E", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Rents", "C": "Royalties", "D": "Other Income", "E": "FWH"},
        "sum_cols": ["B", "C", "D", "E"],
    },
    {
        "id": "1099g",
        "header": "1099-G:",
        "match_types": ["1099-G"],
        "fields": {
            "payer_or_entity":  {"col": "A", "type": "label"},
            "unemployment":     {"col": "B", "type": "input", "fmt": "money"},
            "state_local_refund": {"col": "C", "type": "input", "fmt": "money"},
            "federal_wh":       {"col": "D", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Unemployment", "C": "State Refund", "D": "FWH"},
        "sum_cols": ["B", "C", "D"],
    },
    {
        "id": "schedule_d",
        "header": "Schedule D:",
        "match_types": ["1099-B", "_brokerage_gains"],
        "fields": {
            "_source_name":   {"col": "A", "type": "label"},
            "total_proceeds": {"col": "B", "type": "input", "fmt": "money"},
            "total_basis":    {"col": "C", "type": "input", "fmt": "money"},
            "wash_sale_loss": {"col": "D", "type": "input", "fmt": "money"},
            "total_gain_loss": {"col": "E", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Proceeds", "C": "Cost Basis", "D": "Wash Sale", "E": "Net Gain/Loss"},
        "sum_cols": ["B", "C", "D", "E"],
        "flags": ["Check for capital loss carryover from prior year"],
    },
    {
        "id": "k1",
        "header": "K-1s:",
        "match_types": ["K-1"],
        "fields": {
            "_display_name":       {"col": "A", "type": "label"},
            "_carryover_prior":    {"col": "B", "type": "input", "fmt": "money"},
            "box1_ordinary_income": {"col": "C", "type": "input", "fmt": "money"},
            "box2_rental_real_estate": {"col": "D", "type": "input", "fmt": "money"},
            "_allowed":            {"col": "E", "type": "input", "fmt": "money"},
            "_carryover_next":     {"col": "F", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "C/O from PY", "C": "Box 1", "D": "Box 2", "E": "Allowed", "F": "C/O to NY"},
        "sum_cols": ["B", "C", "D", "E", "F"],
        "flags": [
            "Column B (PY carryover): REQUIRES prior year data",
            "Column E (Allowed): REQUIRES basis/at-risk/passive analysis",
            "Column F (NY carryover): REQUIRES basis/at-risk/passive analysis",
        ],
    },
    {
        "id": "k1_detail",
        "header": "K-1 Additional Detail:",
        "match_types": [],  # Special section — populated by K-1 extra boxes
        "special": "k1_extras",
        "fields": {},
        "col_headers": {},
        "sum_cols": [],
    },
    {
        "id": "rental",
        "header": "Rental Income (Schedule E):",
        "match_types": ["rental_income_document"],
        "fields": {
            "property_address":   {"col": "A", "type": "label"},
            "gross_rents":        {"col": "B", "type": "input", "fmt": "money"},
            "total_expenses":     {"col": "C", "type": "input", "fmt": "money"},
            "net_rental_income":  {"col": "D", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Gross Rents", "C": "Expenses", "D": "Net Income"},
        "sum_cols": ["B", "C", "D"],
    },
    {
        "id": "farm",
        "header": "Farm Income (Schedule F):",
        "match_types": ["farm_income_document"],
        "fields": {
            "description":       {"col": "A", "type": "label"},
            "gross_farm_income":  {"col": "B", "type": "input", "fmt": "money"},
            "farm_expenses":      {"col": "C", "type": "input", "fmt": "money"},
            "net_farm_income":    {"col": "D", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Gross", "C": "Expenses", "D": "Net"},
        "sum_cols": ["B", "C", "D"],
    },
    {
        "id": "1098",
        "header": "Form 1098 (Mortgage Interest):",
        "match_types": ["1098"],
        "fields": {
            "payer_or_entity":            {"col": "A", "type": "label"},
            "mortgage_interest":          {"col": "B", "type": "input", "fmt": "money"},
            "property_tax":               {"col": "C", "type": "input", "fmt": "money"},
            "mortgage_insurance_premiums": {"col": "D", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Mortgage Int", "C": "Property Tax", "D": "PMI"},
        "sum_cols": ["B", "C", "D"],
    },
    {
        "id": "1098t",
        "header": "1098-T (Tuition):",
        "match_types": ["1098-T"],
        "fields": {
            "institution_name":      {"col": "A", "type": "label"},
            "payments_received":     {"col": "B", "type": "input", "fmt": "money"},
            "scholarships_grants":   {"col": "C", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Box 1", "C": "Box 5"},
        "sum_cols": ["B", "C"],
    },
    {
        "id": "property_tax",
        "header": "Property Tax Bills:",
        "match_types": ["property_tax_bill"],
        "fields": {
            "property_address": {"col": "A", "type": "label"},
            "tax_amount":       {"col": "B", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Tax Amount"},
        "sum_cols": ["B"],
    },
    {
        "id": "estimated",
        "header": "Estimated Tax Payments:",
        "match_types": ["estimated_tax_record"],
        "fields": {
            "payment_date":   {"col": "A", "type": "label"},
            "federal_amount": {"col": "B", "type": "input", "fmt": "money"},
            "state_amount":   {"col": "C", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Federal", "C": "State"},
        "sum_cols": ["B", "C"],
        "flags": ["Estimated payments often NOT on scanned docs -- verify with client"],
    },
    {
        "id": "charitable",
        "header": "Charitable Contributions:",
        "match_types": ["charitable_receipt"],
        "fields": {
            "organization_name": {"col": "A", "type": "label"},
            "donation_amount":   {"col": "B", "type": "input", "fmt": "money"},
            "donation_type":     {"col": "C", "type": "input", "fmt": "text"},
        },
        "col_headers": {"B": "Amount", "C": "Type"},
        "sum_cols": ["B"],
    },
    {
        "id": "w2g",
        "header": "W-2G (Gambling Winnings):",
        "match_types": ["W-2G"],
        "fields": {
            "payer_or_entity": {"col": "A", "type": "label"},
            "gross_winnings":  {"col": "B", "type": "input", "fmt": "money"},
            "federal_wh":      {"col": "C", "type": "input", "fmt": "money"},
            "type_of_wager":   {"col": "D", "type": "input", "fmt": "text"},
            "state_wh":        {"col": "E", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Winnings", "C": "FWH", "D": "Type", "E": "SWH"},
        "sum_cols": ["B", "C", "E"],
    },
    {
        "id": "1099k",
        "header": "1099-K (Payment Card / Third Party):",
        "match_types": ["1099-K"],
        "fields": {
            "payer_or_entity":       {"col": "A", "type": "label"},
            "gross_amount":          {"col": "B", "type": "input", "fmt": "money"},
            "number_of_transactions": {"col": "C", "type": "input", "fmt": "text"},
            "federal_wh":            {"col": "D", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Gross Amount", "C": "# Txns", "D": "FWH"},
        "sum_cols": ["B", "D"],
    },
    {
        "id": "1099s",
        "header": "1099-S (Real Estate Proceeds):",
        "match_types": ["1099-S"],
        "fields": {
            "address_of_property":           {"col": "A", "type": "label"},
            "gross_proceeds":                {"col": "B", "type": "input", "fmt": "money"},
            "date_of_closing":               {"col": "C", "type": "input", "fmt": "text"},
            "buyers_part_of_real_estate_tax": {"col": "D", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Gross Proceeds", "C": "Closing Date", "D": "RE Tax"},
        "sum_cols": ["B", "D"],
    },
    {
        "id": "1099c",
        "header": "1099-C (Cancellation of Debt):",
        "match_types": ["1099-C"],
        "fields": {
            "payer_or_entity":  {"col": "A", "type": "label"},
            "debt_cancelled":   {"col": "B", "type": "input", "fmt": "money"},
            "date_cancelled":   {"col": "C", "type": "input", "fmt": "text"},
            "fair_market_value": {"col": "D", "type": "input", "fmt": "money"},
            "debt_description": {"col": "E", "type": "input", "fmt": "text"},
        },
        "col_headers": {"B": "Debt Cancelled", "C": "Date", "D": "FMV", "E": "Description"},
        "sum_cols": ["B", "D"],
        "flags": ["Check insolvency exclusion -- may reduce taxable amount"],
    },
    {
        "id": "1098e",
        "header": "1098-E (Student Loan Interest):",
        "match_types": ["1098-E"],
        "fields": {
            "payer_or_entity":     {"col": "A", "type": "label"},
            "student_loan_interest": {"col": "B", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Interest Paid"},
        "sum_cols": ["B"],
    },
    {
        "id": "5498",
        "header": "5498 (IRA Contributions):",
        "match_types": ["5498"],
        "fields": {
            "payer_or_entity":     {"col": "A", "type": "label"},
            "ira_contributions":   {"col": "B", "type": "input", "fmt": "money"},
            "rollover_contributions": {"col": "C", "type": "input", "fmt": "money"},
            "roth_conversion":     {"col": "D", "type": "input", "fmt": "money"},
            "fair_market_value":   {"col": "E", "type": "input", "fmt": "money"},
            "rmd_amount":          {"col": "F", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Contributions", "C": "Rollovers", "D": "Roth Conv", "E": "FMV", "F": "RMD"},
        "sum_cols": ["B", "C", "D"],
    },
    {
        "id": "schedule_c",
        "header": "Schedule C (Business Income):",
        "match_types": ["schedule_c_summary"],
        "fields": {
            "business_name":  {"col": "A", "type": "label"},
            "gross_income":   {"col": "B", "type": "input", "fmt": "money"},
            "total_expenses": {"col": "C", "type": "input", "fmt": "money"},
            "net_profit":     {"col": "D", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Gross Income", "C": "Expenses", "D": "Net Profit"},
        "sum_cols": ["B", "C", "D"],
    },
    # ─── Bookkeeping Sections ───
    {
        "id": "bank_statement",
        "header": "Bank Statements:",
        "match_types": ["bank_statement"],
        "fields": {
            "bank_name":          {"col": "A", "type": "label"},
            "beginning_balance":  {"col": "B", "type": "input", "fmt": "money"},
            "total_deposits":     {"col": "C", "type": "input", "fmt": "money"},
            "total_withdrawals":  {"col": "D", "type": "input", "fmt": "money"},
            "fees_charged":       {"col": "E", "type": "input", "fmt": "money"},
            "interest_earned":    {"col": "F", "type": "input", "fmt": "money"},
            "ending_balance":     {"col": "G", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Begin Bal", "C": "Deposits", "D": "Withdrawals",
                         "E": "Fees", "F": "Interest", "G": "End Bal"},
        "sum_cols": ["C", "D", "E", "F"],
    },
    {
        "id": "credit_card",
        "header": "Credit Card Statements:",
        "match_types": ["credit_card_statement"],
        "fields": {
            "card_issuer":       {"col": "A", "type": "label"},
            "previous_balance":  {"col": "B", "type": "input", "fmt": "money"},
            "purchases":         {"col": "C", "type": "input", "fmt": "money"},
            "payments":          {"col": "D", "type": "input", "fmt": "money"},
            "interest_charged":  {"col": "E", "type": "input", "fmt": "money"},
            "fees_charged":      {"col": "F", "type": "input", "fmt": "money"},
            "new_balance":       {"col": "G", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Prev Bal", "C": "Purchases", "D": "Payments",
                         "E": "Interest", "F": "Fees", "G": "New Bal"},
        "sum_cols": ["C", "D", "E", "F"],
    },
    {
        "id": "check_stub",
        "header": "Pay Stubs / Check Stubs:",
        "match_types": ["check_stub"],
        "fields": {
            "employer_name":  {"col": "A", "type": "label"},
            "gross_pay":      {"col": "B", "type": "input", "fmt": "money"},
            "federal_wh":     {"col": "C", "type": "input", "fmt": "money"},
            "state_wh":       {"col": "D", "type": "input", "fmt": "money"},
            "social_security": {"col": "E", "type": "input", "fmt": "money"},
            "medicare":       {"col": "F", "type": "input", "fmt": "money"},
            "net_pay":        {"col": "G", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Gross Pay", "C": "FWH", "D": "SWH",
                         "E": "SS", "F": "Medicare", "G": "Net Pay"},
        "sum_cols": ["B", "C", "D", "E", "F", "G"],
    },
    {
        "id": "invoice",
        "header": "Invoices:",
        "match_types": ["invoice"],
        "fields": {
            "vendor_name":    {"col": "A", "type": "label"},
            "invoice_number": {"col": "B", "type": "input", "fmt": "text"},
            "invoice_date":   {"col": "C", "type": "input", "fmt": "text"},
            "subtotal":       {"col": "D", "type": "input", "fmt": "money"},
            "tax_amount":     {"col": "E", "type": "input", "fmt": "money"},
            "total_amount":   {"col": "F", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Invoice #", "C": "Date", "D": "Subtotal",
                         "E": "Tax", "F": "Total"},
        "sum_cols": ["D", "E", "F"],
    },
    {
        "id": "receipt",
        "header": "Receipts:",
        "match_types": ["receipt"],
        "fields": {
            "vendor_name":  {"col": "A", "type": "label"},
            "receipt_date": {"col": "B", "type": "input", "fmt": "text"},
            "category":     {"col": "C", "type": "input", "fmt": "text"},
            "subtotal":     {"col": "D", "type": "input", "fmt": "money"},
            "tax_amount":   {"col": "E", "type": "input", "fmt": "money"},
            "total_amount": {"col": "F", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Date", "C": "Category", "D": "Subtotal",
                         "E": "Tax", "F": "Total"},
        "sum_cols": ["D", "E", "F"],
    },
    # ─── Financial Statements ───
    {
        "id": "profit_loss",
        "header": "Profit & Loss Statements:",
        "match_types": ["profit_loss_statement"],
        "fields": {
            "payer_or_entity":          {"col": "A", "type": "label"},
            "total_revenue":            {"col": "B", "type": "input", "fmt": "money"},
            "total_cogs":               {"col": "C", "type": "input", "fmt": "money"},
            "gross_profit":             {"col": "D", "type": "input", "fmt": "money"},
            "total_operating_expenses": {"col": "E", "type": "input", "fmt": "money"},
            "net_income":               {"col": "F", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Revenue", "C": "COGS", "D": "Gross Profit",
                         "E": "Op Expenses", "F": "Net Income"},
        "sum_cols": ["B", "C", "D", "E", "F"],
    },
    {
        "id": "balance_sheet",
        "header": "Balance Sheets:",
        "match_types": ["balance_sheet"],
        "fields": {
            "payer_or_entity":  {"col": "A", "type": "label"},
            "total_assets":     {"col": "B", "type": "input", "fmt": "money"},
            "total_liabilities": {"col": "C", "type": "input", "fmt": "money"},
            "total_equity":     {"col": "D", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Total Assets", "C": "Total Liabilities", "D": "Total Equity"},
        "sum_cols": ["B", "C", "D"],
    },
    {
        "id": "loan_statement",
        "header": "Loan Statements:",
        "match_types": ["loan_statement", "mortgage_statement"],
        "fields": {
            "lender":          {"col": "A", "type": "label"},
            "current_balance": {"col": "B", "type": "input", "fmt": "money"},
            "interest_rate":   {"col": "C", "type": "input", "fmt": "text"},
            "payment_amount":  {"col": "D", "type": "input", "fmt": "money"},
            "principal_paid":  {"col": "E", "type": "input", "fmt": "money"},
            "interest_paid":   {"col": "F", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Balance", "C": "Rate", "D": "Payment",
                         "E": "Principal", "F": "Interest"},
        "sum_cols": ["D", "E", "F"],
    },
    # ─── Payroll Sections ───
    {
        "id": "payroll_register",
        "header": "Payroll Registers:",
        "match_types": ["payroll_register", "payroll_summary"],
        "fields": {
            "payer_or_entity":     {"col": "A", "type": "label"},
            "total_gross":         {"col": "B", "type": "input", "fmt": "money"},
            "total_federal_wh":    {"col": "C", "type": "input", "fmt": "money"},
            "total_state_wh":      {"col": "D", "type": "input", "fmt": "money"},
            "total_social_security": {"col": "E", "type": "input", "fmt": "money"},
            "total_medicare":      {"col": "F", "type": "input", "fmt": "money"},
            "total_net_pay":       {"col": "G", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Gross", "C": "FWH", "D": "SWH",
                         "E": "SS", "F": "Medicare", "G": "Net Pay"},
        "sum_cols": ["B", "C", "D", "E", "F", "G"],
    },
    {
        "id": "payroll_tax",
        "header": "Payroll Tax Forms (940/941/943/944/945):",
        "match_types": ["940", "941", "943", "944", "945"],
        "fields": {
            "payer_or_entity":       {"col": "A", "type": "label"},
            "total_wages":           {"col": "B", "type": "input", "fmt": "money"},
            "total_federal_tax":     {"col": "C", "type": "input", "fmt": "money"},
            "total_social_security_tax": {"col": "D", "type": "input", "fmt": "money"},
            "total_medicare_tax":    {"col": "E", "type": "input", "fmt": "money"},
            "balance_due":           {"col": "F", "type": "input", "fmt": "money"},
        },
        "col_headers": {"B": "Wages", "C": "Federal Tax", "D": "SS Tax",
                         "E": "Medicare Tax", "F": "Bal Due"},
        "sum_cols": ["B", "C", "D", "E", "F"],
    },
]

# Section ID lookup for fast access
_REGISTRY_BY_ID = {s["id"]: s for s in MAPPING_REGISTRY}


# ═══════════════════════════════════════════════════════════════════════════════
# WORKPAPER BUILDER
# ═══════════════════════════════════════════════════════════════════════════════

class WorkpaperBuilder:
    """Generate a deterministic accountant workpaper from the fact store.

    Accepts ONLY identifiers — never file paths, image data, or raw text.

    Args:
        fact_store: FactStore instance (DB-only gateway)
        client_name: Client name string (matches client_canonical_values.client_name)
        year: Tax year string (e.g. "2025")
        mode: "assisted" (all values, flagged) or "safe" (only verified values)

    Raises:
        TypeError: If fact_store is not a FactStore instance.
        ValueError: If client_name or year look like file paths or raw data.
    """

    def __init__(self, fact_store, client_name, year, mode="assisted"):
        # ── Guardrail A: Only accept identifiers ──────────────────────────
        if not isinstance(fact_store, FactStore):
            raise TypeError(
                f"WorkpaperBuilder requires a FactStore instance, "
                f"got {type(fact_store).__name__}"
            )
        _validate_identifier(client_name, "client_name")
        _validate_identifier(str(year), "year")
        if mode not in ("assisted", "safe"):
            raise ValueError(f"mode must be 'assisted' or 'safe', got {mode!r}")

        self.fs = fact_store
        self.client = client_name
        self.year = str(year)
        self.mode = mode

    def build(self, output_path):
        """Generate the workpaper Excel file.

        Args:
            output_path: Path for the output .xlsx file. Must end in .xlsx.

        Returns the output path on success.
        """
        if not isinstance(output_path, str) or not output_path.strip():
            raise ValueError("output_path must be a non-empty string")
        if not output_path.lower().endswith('.xlsx'):
            raise ValueError(
                f"output_path must end in .xlsx, got {output_path!r}"
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = str(self.year)

        # Load all facts for this client/year
        facts = self.fs.get_legacy_facts(self.client, self.year)
        fact_lookup = self._build_lookup(facts)

        # Title block
        row = 1
        ws.cell(row=row, column=1, value=f"Workpaper: {self.client}").font = TITLE_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1
        ws.cell(row=row, column=1,
                value=f"Tax Year {self.year} | Generated {datetime.now().strftime('%m/%d/%Y %I:%M %p')} | Mode: {self.mode.title()}"
                ).font = SUBTITLE_FONT
        row += 2  # blank row

        audit_rows = []

        # Render each section
        for section in MAPPING_REGISTRY:
            section_id = section["id"]
            matched_payers = self._match_payers(section, fact_lookup)

            # Skip empty sections (unless in ALWAYS_SHOW)
            if not matched_payers and section_id not in ALWAYS_SHOW:
                continue

            row = self._write_section(ws, section, matched_payers, fact_lookup,
                                       row, audit_rows)

        # Print setup
        ws.freeze_panes = "A4"
        ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.oddHeader.center.text = f"&BWorkpaper \u2014 {self.client} \u2014 {self.year}"
        ws.oddHeader.center.size = 10
        ws.oddFooter.left.text = "Bearden Accounting \u2014 Workpaper v1"
        ws.oddFooter.left.size = 8
        ws.oddFooter.right.text = "Page &P of &N"
        ws.oddFooter.right.size = 8
        ws.print_options.gridLines = True

        # Set column widths
        ws.column_dimensions["A"].width = 30
        for col_letter in "BCDEFG":
            ws.column_dimensions[col_letter].width = 16

        # Audit trail sheet
        self._write_audit_trail(wb, audit_rows)

        # Protect formulas
        self._protect_formulas(ws)

        # Remove default sheet if unused
        if "Sheet" in wb.sheetnames and wb["Sheet"].max_row <= 1:
            del wb["Sheet"]

        wb.save(output_path)
        return output_path

    # ── Internal methods ─────────────────────────────────────────────────────

    def _build_lookup(self, facts):
        """Build nested lookup: {doc_type: {payer_key: {field_name: fact_dict}}}."""
        lookup = {}
        for f in facts:
            dt = f["document_type"]
            pk = f["payer_key"]
            fn = f["field_name"]
            lookup.setdefault(dt, {}).setdefault(pk, {})[fn] = f
        return lookup

    def _match_payers(self, section, fact_lookup):
        """Find all payer_keys that match this section's match_types.

        Returns list of (doc_type, payer_key, payer_display) tuples.
        """
        matched = []
        seen = set()
        for mt in section.get("match_types", []):
            for dt, payers in fact_lookup.items():
                if dt == mt:
                    for pk, fields in payers.items():
                        key = (dt, pk)
                        if key not in seen:
                            seen.add(key)
                            # Get payer_display from any fact in this group
                            display = ""
                            for f in fields.values():
                                display = f.get("payer_display", "")
                                if display:
                                    break
                            matched.append((dt, pk, display))
        return matched

    def _write_section(self, ws, section, matched_payers, fact_lookup,
                        start_row, audit_rows):
        """Write one section block. Returns the next available row."""
        section_id = section["id"]
        row = start_row

        # Section header
        ws.cell(row=row, column=1, value=section["header"]).font = SECTION_FONT
        for col_idx in range(1, 8):
            ws.cell(row=row, column=col_idx).fill = SECTION_FILL
        row += 1

        # Column headers
        col_headers = section.get("col_headers", {})
        for col_letter, header_text in col_headers.items():
            col_idx = ord(col_letter) - ord("A") + 1
            cell = ws.cell(row=row, column=col_idx, value=header_text)
            cell.font = COL_HEADER_FONT
            cell.fill = COL_HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        row += 1

        data_start_row = row
        fields_spec = section.get("fields", {})

        # Data rows — one per payer
        for dt, pk, payer_display in matched_payers:
            payer_facts = fact_lookup.get(dt, {}).get(pk, {})

            for field_name, spec in fields_spec.items():
                col_letter = spec["col"]
                col_idx = ord(col_letter) - ord("A") + 1
                field_type = spec.get("type", "input")
                fmt = spec.get("fmt", "text")

                if field_type == "label":
                    # Label column — write payer display name
                    ws.cell(row=row, column=col_idx, value=payer_display or pk)
                    continue

                if field_name.startswith("_"):
                    # Special fields (_carryover_prior, _allowed, etc.)
                    # Leave empty — require manual entry
                    cell = ws.cell(row=row, column=col_idx)
                    cell.fill = PatternFill("solid", fgColor="FFF3E0")  # Light orange
                    cell.protection = UNLOCKED
                    if field_name in ("_carryover_prior", "_allowed", "_carryover_next"):
                        cell.comment = Comment("Manual entry required", "Workpaper")
                    continue

                # Look up fact
                fact = payer_facts.get(field_name)
                if fact:
                    value = fact["canonical_value"]
                    status = fact["status"]
                    evidence = fact.get("evidence_ref", "")
                    source = fact.get("source_doc", "")
                    page = fact.get("page_number")
                else:
                    value = None
                    status = "missing"
                    evidence = ""
                    source = ""
                    page = None

                # Safe mode: skip unverified values
                if self.mode == "safe" and status not in VERIFIED_STATUSES:
                    cell = ws.cell(row=row, column=col_idx)
                    cell.fill = STATUS_FILLS.get("needs_review", STATUS_FILLS["missing"])
                    cell.protection = UNLOCKED
                    cell.comment = Comment("NEEDS REVIEW -- not yet verified", "Workpaper")
                    # Audit trail entry
                    audit_rows.append({
                        "fact_key": FactStore.fact_key(dt, pk, field_name),
                        "value": "",
                        "status": "omitted (safe mode)",
                        "source_doc": source,
                        "page": page,
                        "evidence": evidence,
                    })
                    continue

                # Write the value
                cell = ws.cell(row=row, column=col_idx)
                if value is not None:
                    # Try to write as number for money fields
                    if fmt == "money":
                        try:
                            num_val = float(str(value).replace(",", "").replace("$", ""))
                            cell.value = num_val
                            cell.number_format = MONEY_FMT
                        except (ValueError, TypeError):
                            cell.value = value
                    else:
                        cell.value = value
                    cell.alignment = Alignment(horizontal="right" if fmt == "money" else "left")

                # Status fill
                fill = STATUS_FILLS.get(status, STATUS_FILLS.get("extracted"))
                if fill:
                    cell.fill = fill

                # Input cell — unlocked
                cell.protection = UNLOCKED

                # Evidence comment
                comment_parts = []
                if status and status != "confirmed":
                    comment_parts.append(f"Status: {status}")
                if evidence:
                    comment_parts.append(f"Evidence: {evidence}")
                if source:
                    page_str = f" p.{page}" if page else ""
                    comment_parts.append(f"Source: {source}{page_str}")
                if comment_parts:
                    cell.comment = Comment("\n".join(comment_parts), "Workpaper")

                # Audit trail
                audit_rows.append({
                    "fact_key": FactStore.fact_key(dt, pk, field_name),
                    "value": value if value is not None else "",
                    "status": status,
                    "source_doc": source,
                    "page": page,
                    "evidence": evidence,
                })

            row += 1

        # If no payers matched and section is in ALWAYS_SHOW, leave one empty row
        if not matched_payers:
            row += 1

        data_end_row = row - 1

        # SUM formulas in total row
        sum_cols = section.get("sum_cols", [])
        if sum_cols and matched_payers:
            ws.cell(row=row, column=1, value="Total:").font = SUM_FONT
            for col_letter in sum_cols:
                col_idx = ord(col_letter) - ord("A") + 1
                formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"
                cell = ws.cell(row=row, column=col_idx, value=formula)
                cell.font = SUM_FONT
                cell.number_format = MONEY_FMT
                cell.protection = LOCKED  # Formula cell — locked
            row += 1

        # Total formula columns (cross-column formulas per data row)
        total_formula = section.get("total_formula_col", {})
        if total_formula and matched_payers:
            for col_letter, formula_expr in total_formula.items():
                col_idx = ord(col_letter) - ord("A") + 1
                # Write column header if not already set
                hdr_cell = ws.cell(row=start_row + 1, column=col_idx)
                if not hdr_cell.value:
                    hdr_cell.value = "Total"
                    hdr_cell.font = COL_HEADER_FONT
                    hdr_cell.fill = COL_HEADER_FILL
                    hdr_cell.alignment = Alignment(horizontal="center")
                # Write formulas in each data row
                for r in range(data_start_row, data_end_row + 1):
                    parts = formula_expr.split("+")
                    formula = "=" + "+".join(f"{p}{r}" for p in parts)
                    fcell = ws.cell(row=r, column=col_idx, value=formula)
                    fcell.number_format = MONEY_FMT
                    fcell.protection = LOCKED  # Formula — locked

        # Flags / warnings
        for flag in section.get("flags", []):
            ws.cell(row=row, column=1, value=f"\u26A0 {flag}").font = FLAG_FONT
            row += 1

        # Blank separator row
        row += 1
        return row

    def _write_audit_trail(self, wb, audit_rows):
        """Create the Audit Trail worksheet."""
        ws = wb.create_sheet("Audit Trail")

        # Title
        ws["A1"] = "Workpaper Audit Trail"
        ws["A1"].font = Font(bold=True, size=14, color="1A252F")
        ws.merge_cells("A1:G1")
        ws["A2"] = f"Client: {self.client} | Year: {self.year} | Mode: {self.mode.title()}"
        ws["A2"].font = Font(italic=True, color="888888", size=9)
        ws["A3"] = f"Generated: {datetime.now().strftime('%m/%d/%Y %I:%M %p')}"
        ws["A3"].font = Font(italic=True, color="888888", size=9)

        # Headers
        row = 5
        headers = ["Fact Key", "Value", "Status", "Source Document", "Page", "Evidence"]
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=i + 1, value=h)
            cell.font = AUDIT_HEADER_FONT
            cell.fill = AUDIT_HEADER_FILL
            cell.alignment = Alignment(horizontal="center" if i in (2, 4) else "left")
        row += 1

        # Data rows
        for idx, ar in enumerate(audit_rows):
            ws.cell(row=row, column=1, value=ar.get("fact_key", ""))
            ws.cell(row=row, column=2, value=str(ar.get("value", "")))

            status_cell = ws.cell(row=row, column=3, value=ar.get("status", ""))
            status_cell.alignment = Alignment(horizontal="center")
            status = ar.get("status", "")
            if status in STATUS_FILLS:
                status_cell.fill = STATUS_FILLS[status]

            ws.cell(row=row, column=4, value=ar.get("source_doc", ""))
            ws.cell(row=row, column=5, value=ar.get("page") or "")
            ws.cell(row=row, column=6, value=ar.get("evidence", ""))

            # Alternating background
            if idx % 2 == 1:
                for c in range(1, 7):
                    cell = ws.cell(row=row, column=c)
                    if cell.fill == PatternFill():
                        cell.fill = AUDIT_ALT_FILL

            for c in range(1, 7):
                ws.cell(row=row, column=c).border = THIN_BORDER

            row += 1

        # Summary
        row += 1
        total = len(audit_rows)
        verified = sum(1 for a in audit_rows if a.get("status") in VERIFIED_STATUSES)
        ws.cell(row=row, column=1, value="Summary:").font = Font(bold=True, size=11)
        row += 1
        ws.cell(row=row, column=1, value=f"Total facts written: {total}")
        row += 1
        ws.cell(row=row, column=1, value=f"Verified: {verified}")
        row += 1
        ws.cell(row=row, column=1, value=f"Unverified: {total - verified}")

        # Column widths
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 8
        ws.column_dimensions["F"].width = 25

    def _protect_formulas(self, ws):
        """Lock formula cells and enable sheet protection.

        Formula cells (value starts with '=') get locked=True.
        Input cells were already set to locked=False during writing.
        Sheet protection is structural only (no password).
        """
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.protection = LOCKED
        ws.protection.sheet = True
        # No password — structural protection only (prevents accidental edits)
