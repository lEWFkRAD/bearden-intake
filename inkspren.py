# inkspren.py — OathLedger Output Engine
# Named for the Inkspren of the Stormlight Archive: spren of logic and
# creation, they bond with scholars to form Elsecallers. InkSpren transforms
# raw extracted data into structured, beautiful output.
#
# Supported formats: Tax Review, Journal Entries, Account Balances,
# Trial Balance, Transaction Register.

import os
import re
from datetime import datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment

# Shared helpers from extraction engine
from extract import get_val, get_str

# ─── SHARED STYLE SYSTEM (Pi for Excel–inspired) ─────────────────────────
# All formatting constants now live in workpaper_styles.py so inkspren,
# oathledger/renderer.py, and the Office.js add-in share identical styles.
from workpaper_styles import (
    # Named style helpers
    apply_styles, write_section_header, write_total_row,
    write_flag_rows, write_title_block, set_standard_widths, setup_print,
    # Number formats
    MONEY_FMT, PCT_FMT, DATE_FMT, INT_FMT, NUMBER_FORMATS,
    # Backward-compat constants (still used in _write_cell_value etc.)
    BOLD, SECTION_FONT, SECTION_FILL, COL_HEADER_FONT, COL_HEADER_FILL,
    SUM_FONT, SUM_FILL, SUM_BORDER, FLAG_FONT, FLAG_FILL,
    CORRECTED_FILL, REVIEW_FILL, CONFIRMED_FILL, DUAL_FILL,
    ALT_ROW_FILL, DARK_HEADER_FILL, DARK_HEADER_FONT,
    THIN_BORDER, SECTION_BORDER,
    # Color palette
    COLORS,
)


# ─── TEMPLATE SECTIONS ──────────────────────────────────────────────────────
# DEPRECATED (TMPL-007-B): These hardcoded section definitions are now
# superseded by templates/tax_review_1040.json (a structured TemplateMap).
# The new rendering path uses: TemplateMap + OutputPlan → renderer.
#
# This constant is preserved for backward compatibility with:
#   - oathledger/rules_engine.py (builds payload from these sections)
#   - extract.py (field display ordering)
#   - tests/test_workpaper.py (MAPPING_REGISTRY coverage check)
#
# Kill switch: INKSPREN_LEGACY_RENDER=1 forces the old _populate_tax_review path.
# Default: new TemplateMap-driven path (via template_bridge.py).

TEMPLATE_SECTIONS = [
    {
        "id": "w2",
        "header": "W-2:",
        "match_types": ["W-2"],
        "columns": {"A": "employer_name", "B": "wages", "C": "federal_wh", "D": "state_wh"},
        "col_headers": {"B": "Gross", "C": "Federal WH", "D": "State WH"},
        "sum_cols": ["B", "C", "D"],
    },
    {
        "id": "interest",
        "header": "Interest Income:",
        "match_types": ["1099-INT", "_interest_rollup"],
        "columns": {"A": "_source_name", "B": "interest_income", "C": "us_savings_bonds_and_treasury"},
        "col_headers": {"B": "Interest", "C": "US Bonds"},
        "sum_cols": ["B", "C"],
        "total_formula_col": {"D": "B+C"},
    },
    {
        "id": "dividends",
        "header": "Dividends:",
        "match_types": ["1099-DIV", "_dividend_rollup"],
        "columns": {"A": "_source_name", "B": "ordinary_dividends", "C": "qualified_dividends",
                     "D": "capital_gain_distributions", "E": "section_199a"},
        "col_headers": {"B": "Total Ord", "C": "Qualified", "D": "Cap Gain Dist", "E": "Sec 199A"},
        "sum_cols": ["B", "C", "D", "E"],
    },
    {
        "id": "1099r",
        "header": "1099-R:",
        "match_types": ["1099-R"],
        "columns": {"A": "payer_or_entity", "B": "gross_distribution", "C": "taxable_amount",
                     "D": "federal_wh", "E": "state_wh", "F": "distribution_code"},
        "col_headers": {"B": "Gross", "C": "Taxable", "D": "FWH", "E": "SWH", "F": "Code"},
        "sum_cols": ["B", "C", "D", "E"],
    },
    {
        "id": "ssa",
        "header": "SSA-1099:",
        "match_types": ["SSA-1099"],
        "columns": {"A": "payer_or_entity", "B": "net_benefits", "C": "federal_wh"},
        "col_headers": {"B": "Net Benefits", "C": "Federal WH"},
        "sum_cols": ["B", "C"],
    },
    {
        "id": "1099nec",
        "header": "1099-NEC (Self-Employment):",
        "match_types": ["1099-NEC"],
        "columns": {"A": "payer_or_entity", "B": "nonemployee_compensation", "C": "federal_wh"},
        "col_headers": {"B": "NEC Income", "C": "FWH"},
        "sum_cols": ["B", "C"],
    },
    {
        "id": "1099misc",
        "header": "1099-MISC:",
        "match_types": ["1099-MISC"],
        "columns": {"A": "payer_or_entity", "B": "rents", "C": "royalties", "D": "other_income", "E": "federal_wh"},
        "col_headers": {"B": "Rents", "C": "Royalties", "D": "Other Income", "E": "FWH"},
        "sum_cols": ["B", "C", "D", "E"],
    },
    {
        "id": "1099g",
        "header": "1099-G:",
        "match_types": ["1099-G"],
        "columns": {"A": "payer_or_entity", "B": "unemployment", "C": "state_local_refund", "D": "federal_wh"},
        "col_headers": {"B": "Unemployment", "C": "State Refund", "D": "FWH"},
        "sum_cols": ["B", "C", "D"],
    },
    {
        "id": "schedule_d",
        "header": "Schedule D:",
        "match_types": ["1099-B", "_brokerage_gains"],
        "columns": {"A": "_source_name", "B": "total_proceeds", "C": "total_basis",
                     "D": "wash_sale_loss", "E": "total_gain_loss"},
        "col_headers": {"B": "Proceeds", "C": "Cost Basis", "D": "Wash Sale", "E": "Net Gain/Loss"},
        "sum_cols": ["B", "C", "D", "E"],
        "flags": ["\u26a0 Check for capital loss carryover from prior year"],
    },
    {
        "id": "k1",
        "header": "K-1s:",
        "match_types": ["K-1"],
        "columns": {"A": "_display_name", "B": "_carryover_prior", "C": "box1_ordinary_income",
                     "D": "box2_rental_real_estate", "E": "_allowed", "F": "_carryover_next"},
        "col_headers": {"B": "C/O from PY", "C": "Box 1", "D": "Box 2", "E": "Allowed", "F": "C/O to NY"},
        "sum_cols": ["B", "C", "D", "E", "F"],
        "flags": [
            "\u26a0 Column B (PY carryover): REQUIRES prior year data \u2014 enter manually",
            "\u26a0 Column E (Allowed): REQUIRES basis/at-risk/passive analysis \u2014 enter manually",
            "\u26a0 Column F (NY carryover): REQUIRES basis/at-risk/passive analysis \u2014 enter manually",
        ],
    },
    {
        "id": "k1_detail",
        "header": "K-1 Additional Detail:",
        "special": "k1_extras",
    },
    {
        "id": "rental",
        "header": "Rental Income (Schedule E):",
        "match_types": ["rental_income_document"],
        "columns": {"A": "property_address", "B": "gross_rents", "C": "total_expenses", "D": "net_rental_income"},
        "col_headers": {"B": "Gross Rents", "C": "Expenses", "D": "Net Income"},
        "sum_cols": ["B", "C", "D"],
    },
    {
        "id": "farm",
        "header": "Farm Income (Schedule F):",
        "match_types": ["farm_income_document"],
        "columns": {"A": "description", "B": "gross_farm_income", "C": "farm_expenses", "D": "net_farm_income"},
        "col_headers": {"B": "Gross", "C": "Expenses", "D": "Net"},
        "sum_cols": ["B", "C", "D"],
    },
    {
        "id": "1098",
        "header": "Schedule A \u2014 Mortgage Interest (1098):",
        "match_types": ["1098"],
        "columns": {"A": "payer_or_entity", "B": "mortgage_interest", "C": "property_tax", "D": "mortgage_insurance_premiums"},
        "col_headers": {"B": "Mortgage Int", "C": "Property Tax", "D": "PMI"},
        "sum_cols": ["B", "C", "D"],
        "field_aliases": {"mortgage_interest": ["mortgage_interest_received"],
                          "property_tax": ["real_estate_tax", "property_taxes_paid"]},
    },
    {
        "id": "1098t",
        "header": "1098-T (Tuition):",
        "match_types": ["1098-T"],
        "columns": {"A": "institution_name", "B": "payments_received", "C": "scholarships_grants"},
        "col_headers": {"B": "Box 1", "C": "Box 5"},
        "sum_cols": ["B", "C"],
    },
    {
        "id": "property_tax",
        "header": "Schedule A \u2014 Property Taxes:",
        "match_types": ["property_tax_bill"],
        "columns": {"A": "property_address", "B": "tax_amount"},
        "field_aliases": {"tax_amount": ["total_due", "total_tax", "total_amount_due", "total_estimated_tax"]},
        "col_headers": {"B": "Tax Amount"},
        "sum_cols": ["B"],
    },
    {
        "id": "estimated",
        "header": "Estimated Tax Payments:",
        "match_types": ["estimated_tax_record"],
        "columns": {"A": "payment_date", "B": "federal_amount", "C": "state_amount"},
        "col_headers": {"B": "Federal", "C": "State"},
        "sum_cols": ["B", "C"],
        "flags": ["\u26a0 Estimated payments often NOT on scanned docs \u2014 verify with client"],
    },
    {
        "id": "charitable",
        "header": "Schedule A \u2014 Charitable Contributions:",
        "match_types": ["charitable_receipt"],
        "columns": {"A": "organization_name", "B": "donation_amount", "C": "donation_type"},
        "col_headers": {"B": "Amount", "C": "Type"},
        "sum_cols": ["B"],
    },
    # ─── Additional Tax Forms ───
    {
        "id": "w2g",
        "header": "W-2G (Gambling Winnings):",
        "match_types": ["W-2G"],
        "columns": {"A": "payer_or_entity", "B": "gross_winnings", "C": "federal_wh",
                     "D": "type_of_wager", "E": "state_wh"},
        "col_headers": {"B": "Winnings", "C": "FWH", "D": "Type", "E": "SWH"},
        "sum_cols": ["B", "C", "E"],
    },
    {
        "id": "1099k",
        "header": "1099-K (Payment Card / Third Party):",
        "match_types": ["1099-K"],
        "columns": {"A": "payer_or_entity", "B": "gross_amount", "C": "number_of_transactions",
                     "D": "federal_wh"},
        "col_headers": {"B": "Gross Amount", "C": "# Txns", "D": "FWH"},
        "sum_cols": ["B", "D"],
    },
    {
        "id": "1099s",
        "header": "1099-S (Real Estate Proceeds):",
        "match_types": ["1099-S"],
        "columns": {"A": "address_of_property", "B": "gross_proceeds", "C": "date_of_closing",
                     "D": "buyers_part_of_real_estate_tax"},
        "col_headers": {"B": "Gross Proceeds", "C": "Closing Date", "D": "RE Tax"},
        "sum_cols": ["B", "D"],
    },
    {
        "id": "1099c",
        "header": "1099-C (Cancellation of Debt):",
        "match_types": ["1099-C"],
        "columns": {"A": "payer_or_entity", "B": "debt_cancelled", "C": "date_cancelled",
                     "D": "fair_market_value", "E": "debt_description"},
        "col_headers": {"B": "Debt Cancelled", "C": "Date", "D": "FMV", "E": "Description"},
        "sum_cols": ["B", "D"],
        "flags": ["\u26a0 Check insolvency exclusion \u2014 may reduce taxable amount"],
    },
    {
        "id": "1098e",
        "header": "1098-E (Student Loan Interest):",
        "match_types": ["1098-E"],
        "columns": {"A": "payer_or_entity", "B": "student_loan_interest"},
        "col_headers": {"B": "Interest Paid"},
        "sum_cols": ["B"],
    },
    {
        "id": "5498",
        "header": "5498 (IRA Contributions):",
        "match_types": ["5498"],
        "columns": {"A": "payer_or_entity", "B": "ira_contributions", "C": "rollover_contributions",
                     "D": "roth_conversion", "E": "fair_market_value", "F": "rmd_amount"},
        "col_headers": {"B": "Contributions", "C": "Rollovers", "D": "Roth Conv", "E": "FMV", "F": "RMD"},
        "sum_cols": ["B", "C", "D"],
    },
    {
        "id": "schedule_c",
        "header": "Schedule C (Business Income):",
        "match_types": ["schedule_c_summary"],
        "columns": {"A": "business_name", "B": "gross_income", "C": "total_expenses", "D": "net_profit"},
        "col_headers": {"B": "Gross Income", "C": "Expenses", "D": "Net Profit"},
        "sum_cols": ["B", "C", "D"],
    },
    # ─── Bookkeeping Sections ───
    {
        "id": "bank_statement",
        "header": "Bank Statements:",
        "match_types": ["bank_statement"],
        "columns": {"A": "bank_name", "B": "beginning_balance", "C": "total_deposits",
                     "D": "total_withdrawals", "E": "fees_charged", "F": "interest_earned", "G": "ending_balance"},
        "col_headers": {"B": "Begin Bal", "C": "Deposits", "D": "Withdrawals", "E": "Fees", "F": "Interest", "G": "End Bal"},
        "sum_cols": ["C", "D", "E", "F"],
    },
    {
        "id": "credit_card",
        "header": "Credit Card Statements:",
        "match_types": ["credit_card_statement"],
        "columns": {"A": "card_issuer", "B": "previous_balance", "C": "purchases",
                     "D": "payments", "E": "interest_charged", "F": "fees_charged", "G": "new_balance"},
        "col_headers": {"B": "Prev Bal", "C": "Purchases", "D": "Payments", "E": "Interest", "F": "Fees", "G": "New Bal"},
        "sum_cols": ["C", "D", "E", "F"],
    },
    {
        "id": "check_stub",
        "header": "Pay Stubs / Check Stubs:",
        "match_types": ["check_stub"],
        "columns": {"A": "employer_name", "B": "gross_pay", "C": "federal_wh",
                     "D": "state_wh", "E": "social_security", "F": "medicare", "G": "net_pay"},
        "col_headers": {"B": "Gross Pay", "C": "FWH", "D": "SWH", "E": "SS", "F": "Medicare", "G": "Net Pay"},
        "sum_cols": ["B", "C", "D", "E", "F", "G"],
    },
    {
        "id": "invoice",
        "header": "Invoices:",
        "match_types": ["invoice"],
        "columns": {"A": "vendor_name", "B": "invoice_number", "C": "invoice_date",
                     "D": "subtotal", "E": "tax_amount", "F": "total_amount"},
        "col_headers": {"B": "Invoice #", "C": "Date", "D": "Subtotal", "E": "Tax", "F": "Total"},
        "sum_cols": ["D", "E", "F"],
    },
    {
        "id": "receipt",
        "header": "Receipts:",
        "match_types": ["receipt"],
        "columns": {"A": "vendor_name", "B": "receipt_date", "C": "category",
                     "D": "subtotal", "E": "tax_amount", "F": "total_amount"},
        "col_headers": {"B": "Date", "C": "Category", "D": "Subtotal", "E": "Tax", "F": "Total"},
        "sum_cols": ["D", "E", "F"],
    },
    # ─── Financial Statements ───
    {
        "id": "profit_loss",
        "header": "Profit & Loss Statements:",
        "match_types": ["profit_loss_statement"],
        "columns": {"A": "payer_or_entity", "B": "total_revenue", "C": "total_cogs",
                     "D": "gross_profit", "E": "total_operating_expenses", "F": "net_income"},
        "col_headers": {"B": "Revenue", "C": "COGS", "D": "Gross Profit", "E": "Op Expenses", "F": "Net Income"},
        "sum_cols": ["B", "C", "D", "E", "F"],
    },
    {
        "id": "balance_sheet",
        "header": "Balance Sheets:",
        "match_types": ["balance_sheet"],
        "columns": {"A": "payer_or_entity", "B": "total_assets", "C": "total_liabilities",
                     "D": "total_equity"},
        "col_headers": {"B": "Total Assets", "C": "Total Liabilities", "D": "Total Equity"},
        "sum_cols": ["B", "C", "D"],
    },
    {
        "id": "loan_statement",
        "header": "Loan Statements:",
        "match_types": ["loan_statement", "mortgage_statement"],
        "columns": {"A": "lender", "B": "current_balance", "C": "interest_rate",
                     "D": "payment_amount", "E": "principal_paid", "F": "interest_paid"},
        "col_headers": {"B": "Balance", "C": "Rate", "D": "Payment", "E": "Principal", "F": "Interest"},
        "sum_cols": ["D", "E", "F"],
    },
    # ─── Payroll Sections ───
    {
        "id": "payroll_register",
        "header": "Payroll Registers:",
        "match_types": ["payroll_register", "payroll_summary"],
        "columns": {"A": "payer_or_entity", "B": "total_gross", "C": "total_federal_wh",
                     "D": "total_state_wh", "E": "total_social_security", "F": "total_medicare", "G": "total_net_pay"},
        "col_headers": {"B": "Gross", "C": "FWH", "D": "SWH", "E": "SS", "F": "Medicare", "G": "Net Pay"},
        "sum_cols": ["B", "C", "D", "E", "F", "G"],
    },
    {
        "id": "payroll_tax",
        "header": "Payroll Tax Forms (940/941/943/944/945):",
        "match_types": ["940", "941", "943", "944", "945"],
        "columns": {"A": "payer_or_entity", "B": "total_wages", "C": "total_federal_tax",
                     "D": "total_social_security_tax", "E": "total_medicare_tax", "F": "balance_due"},
        "col_headers": {"B": "Wages", "C": "Federal Tax", "D": "SS Tax", "E": "Medicare Tax", "F": "Bal Due"},
        "sum_cols": ["B", "C", "D", "E", "F"],
    },
]

ALWAYS_SHOW = []  # Skip empty sections entirely — only show sections with actual documents


# ─── EXCEL STYLES ────────────────────────────────────────────────────────────
# All style constants are now imported from workpaper_styles.py (see imports above).
# This keeps the constants available at module scope for backward compatibility
# with code that does `import inkspren; inkspren.MONEY_FMT` etc.


# ─── ROUTER ──────────────────────────────────────────────────────────────────

def populate_template(extractions, template_path, output_path, year, output_format="tax_review"):
    """Router: create workbook, delegate to format-specific function, save.

    For the default 'tax_review' format, this now routes through the
    TemplateMap + OutputPlan pipeline (TMPL-007-B) unless the legacy
    kill switch INKSPREN_LEGACY_RENDER=1 is set.
    """
    fmt_labels = {
        "tax_review": "Tax Review", "tax_review_payload": "Tax Review (v2 Payload)",
        "tax_review_v3": "Tax Review (v3 TemplateMap)",
        "journal_entries": "Journal Entries",
        "account_balances": "Account Balances", "trial_balance": "Trial Balance",
        "transaction_register": "Transaction Register",
    }
    print(f"\n\u2500\u2500 [PASSION] Phase 6: Excel ({fmt_labels.get(output_format, output_format)}) \u2500\u2500")

    if template_path and os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
    else:
        wb = openpyxl.Workbook()

    sheet_name = str(year)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name, 0)

    if output_format == "journal_entries":
        _populate_journal_entries(ws, extractions, year)
    elif output_format == "account_balances":
        _populate_account_balances(ws, extractions, year)
    elif output_format == "trial_balance":
        _populate_trial_balance(ws, extractions, year)
    elif output_format == "transaction_register":
        _populate_transaction_register(ws, extractions, year)
    elif output_format == "tax_review_payload":
        from oathledger.rules_engine import build_tax_review_payload
        from oathledger.renderer import populate_tax_review_from_payload
        payload = build_tax_review_payload(extractions, year)
        populate_tax_review_from_payload(ws, payload, year)
        # Schema drift guard: log if TEMPLATE_SECTIONS changed since v2 was built
        expected_hash = payload.get("schema_hash", "")
        if expected_hash:
            import hashlib, json as _json
            current_hash = "sha256:" + hashlib.sha256(
                _json.dumps(TEMPLATE_SECTIONS, sort_keys=True, separators=(",", ":"), ensure_ascii=False).encode()
            ).hexdigest()
            if current_hash != expected_hash:
                print(f"  [OathLedger] WARNING: TEMPLATE_SECTIONS schema drift detected")
                print(f"    Expected: {expected_hash[:24]}...")
                print(f"    Current:  {current_hash[:24]}...")
    elif output_format == "tax_review_v3":
        # Explicit v3 TemplateMap path (TMPL-007-B)
        from template_bridge import render_tax_review_via_map
        render_tax_review_via_map(ws, extractions, year, draft_mode=True)
    else:
        # Default: tax_review
        # TMPL-007-B: Route through TemplateMap unless legacy kill switch is set
        if os.environ.get("INKSPREN_LEGACY_RENDER") == "1":
            _populate_tax_review(ws, extractions, year)
        else:
            try:
                from template_bridge import render_tax_review_via_map
                render_tax_review_via_map(ws, extractions, year, draft_mode=True)
            except Exception as e:
                print(f"  [TMPL-007-B] TemplateMap render failed, falling back to legacy: {e}")
                _populate_tax_review(ws, extractions, year)

    ws.freeze_panes = "A4"

    # Column widths (Pi-inspired character-unit math)
    set_standard_widths(ws, num_cols=6)

    # Print setup (shared across all renderers)
    setup_print(ws, fmt_labels.get(output_format, "Document Intake"), year)

    # Remove default sheet if it exists and is empty
    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row <= 1:
        del wb["Sheet"]

    wb.save(output_path)
    print(f"  \u2713 Saved: {output_path}")


# ─── SHARED HELPERS ──────────────────────────────────────────────────────────

def _match_exts(extractions, match_types):
    matched = []
    for ext in extractions:
        dtype = str(ext.get("document_type", ""))
        for mt in match_types:
            if mt in dtype:
                matched.append(ext)
                break
    return matched

def _dedup_by_ein(exts):
    """Deduplicate extractions by EIN/entity, keeping the BEST version.

    When two copies of the same document exist (common with scanned returns),
    keep the one with more high-confidence fields instead of first-seen.
    Missing fields from the discarded copy are merged into the winner.
    """
    seen = {}       # key → (ext, score)
    order = []      # preserve insertion order of keys

    def _confidence_score(ext):
        """Score an extraction by field confidence. Higher = better."""
        score = 0
        for fdata in ext.get("fields", {}).values():
            if not isinstance(fdata, dict):
                score += 1  # bare value = ok
                continue
            c = fdata.get("confidence", "")
            if c in ("dual_confirmed", "verified_confirmed"):
                score += 3
            elif c in ("verified_corrected", "found_in_verification"):
                score += 2
            elif c in ("high", "ocr_accepted"):
                score += 2
            elif c in ("medium", "from_continuation", "from_k1_box5"):
                score += 1
            elif c in ("low",):
                score += 0
            else:
                score += 1
        return score

    for ext in exts:
        ein = ext.get("payer_ein", "")
        fields = ext.get("fields", {})
        name = get_str(fields, "partnership_name") or get_str(fields, "employer_name") or ext.get("payer_or_entity", "")
        recip = ext.get("recipient", "")
        if ein:
            key = f"EIN:{ein}|{recip}".upper()
        else:
            norm = re.sub(r'\s*\(\d{4}\)\s*', '', name).upper().strip()
            key = f"NAME:{norm}|{recip}".upper()

        score = _confidence_score(ext)

        if key not in seen:
            seen[key] = (ext, score)
            order.append(key)
        else:
            existing_ext, existing_score = seen[key]
            if score > existing_score:
                # New copy is better — swap, merge missing fields from old
                for fname, fdata in existing_ext.get("fields", {}).items():
                    if fname not in ext.get("fields", {}):
                        ext.setdefault("fields", {})[fname] = fdata
                seen[key] = (ext, score)
                print(f"  Dedup: kept better copy of {name or ein} (score {score} > {existing_score})")
            else:
                # Existing is better — merge missing fields from new
                for fname, fdata in ext.get("fields", {}).items():
                    if fname not in existing_ext.get("fields", {}):
                        existing_ext.setdefault("fields", {})[fname] = fdata

    return [seen[k][0] for k in order]

def _write_title(ws, title, year, client_name=""):
    """Write title rows with optional client name, return next row number.

    Now delegates to workpaper_styles.write_title_block() for consistency
    across all renderers.
    """
    return write_title_block(ws, title, year, client_name=client_name)

def _write_cell_value(ws, col, row, fields, field_name, ext, matched):
    """Write a single field cell with formatting and confidence coloring."""
    if field_name == "_display_name":
        name = get_str(fields, "partnership_name") or ext.get("payer_or_entity", "")
        recip = ext.get("recipient", "")
        same = [e for e in matched if (e.get("payer_ein","") == ext.get("payer_ein","") and ext.get("payer_ein",""))
                or (get_str(e.get("fields",{}), "partnership_name") or e.get("payer_or_entity","")).upper() == name.upper()]
        if len(same) > 1 and recip:
            name = f"{name} - {recip}"
        ws[f"{col}{row}"] = name
    elif field_name == "_source_name":
        ws[f"{col}{row}"] = ext.get("_source_name", ext.get("payer_or_entity", ""))
    elif field_name == "_carryover_prior":
        cell = ws[f"{col}{row}"]
        cell.value = 0
        cell.number_format = MONEY_FMT
        cell.alignment = Alignment(horizontal="right")
        cell.fill = REVIEW_FILL
        cell.comment = Comment("REQUIRES PRIOR YEAR DATA", "System")
    elif field_name in ("_allowed", "_carryover_next"):
        ws[f"{col}{row}"] = None
        ws[f"{col}{row}"].fill = REVIEW_FILL
        ws[f"{col}{row}"].comment = Comment("REQUIRES PREPARER JUDGMENT", "System")
    elif field_name == "employer_name":
        name = get_str(fields, "employer_name") or ext.get("payer_or_entity", "")
        name = re.sub(r'\s*\(\d{4}\)\s*$', '', name)
        ws[f"{col}{row}"] = name
    elif field_name in ("payer_or_entity", "institution_name"):
        ws[f"{col}{row}"] = get_str(fields, field_name) or ext.get("payer_or_entity", "")
    else:
        val = get_val(fields, field_name)
        if val is None:
            val = get_str(fields, field_name)
        cell = ws[f"{col}{row}"]
        cell.value = val
        if isinstance(val, (int, float)):
            cell.number_format = MONEY_FMT
            cell.alignment = Alignment(horizontal="right")
        # Confidence comments (no cell coloring — keeps spreadsheet clean)
        fdata = fields.get(field_name)
        if isinstance(fdata, dict):
            conf = fdata.get("confidence", "")
            if conf == "verified_corrected":
                cell.comment = Comment(f"Corrected: was {fdata.get('original_value','?')}. {fdata.get('correction_note','')}", "System")
            elif conf == "low":
                cell.comment = Comment("Low confidence \u2014 check source", "System")
            elif conf == "operator_corrected":
                cell.comment = Comment(f"Operator corrected (was {fdata.get('_original_value','?')})", "Operator")


# ─── TAX REVIEW FORMAT ──────────────────────────────────────────────────────

def _populate_tax_review(ws, extractions, year):
    # Extract client name from first W-2 employee or first recipient
    client_name = ""
    for ext in extractions:
        fields = ext.get("fields", {})
        name = get_str(fields, "employee_name") or get_str(fields, "recipient_name") or get_str(fields, "recipient") or get_str(fields, "borrower_name")
        if name and len(name) > 2:
            client_name = name
            break
    if not client_name:
        for ext in extractions:
            name = ext.get("recipient", "")
            if name and len(name) > 2:
                client_name = name
                break

    row = _write_title(ws, "Document Intake", year, client_name=client_name)
    k1_extras = []

    for section in TEMPLATE_SECTIONS:
        sid = section["id"]

        if section.get("special") == "k1_extras":
            if k1_extras:
                ws[f"A{row}"] = section["header"]
                ws[f"A{row}"].font = SECTION_FONT
                ws[f"A{row}"].fill = SECTION_FILL
                for hcol, hlabel in [("B", "Line Ref"), ("C", "Description"), ("D", "Amount")]:
                    cell = ws[f"{hcol}{row}"]
                    cell.value = hlabel
                    cell.font = COL_HEADER_FONT
                    cell.fill = COL_HEADER_FILL
                    cell.alignment = Alignment(horizontal="center")
                for c in ["E", "F"]:
                    ws[f"{c}{row}"].fill = SECTION_FILL
                row += 1
                for item in k1_extras:
                    ws[f"A{row}"] = item.get("entity", "")
                    ws[f"B{row}"] = item.get("line_reference", "")
                    ws[f"C{row}"] = item.get("description", "")
                    amt_cell = ws[f"D{row}"]
                    amt_cell.value = item.get("amount")
                    if isinstance(amt_cell.value, (int, float)):
                        amt_cell.number_format = MONEY_FMT
                        amt_cell.alignment = Alignment(horizontal="right")
                    for c in ["A", "B", "C", "D"]:
                        ws[f"{c}{row}"].border = THIN_BORDER
                    row += 1
                row += 1
            continue

        match_types = section.get("match_types", [])
        matched = _match_exts(extractions, match_types) if match_types else []
        col_headers = section.get("col_headers", {})
        columns = section.get("columns", {})
        sum_cols = section.get("sum_cols", [])
        flags = section.get("flags", [])

        matched = _dedup_by_ein(matched) if matched else []

        # Filter out zero-value entries for interest/dividend sections
        if sid in ("interest", "dividends") and matched:
            matched = [e for e in matched if any(
                (get_val(e.get("fields", {}), fn) or 0) != 0
                for fn in columns.values() if not fn.startswith("_")
            )]

        # Skip empty sections entirely
        if not matched:
            continue

        ws[f"A{row}"] = section["header"]
        ws[f"A{row}"].font = SECTION_FONT
        ws[f"A{row}"].fill = SECTION_FILL
        for col, label in col_headers.items():
            cell = ws[f"{col}{row}"]
            cell.value = label
            cell.font = COL_HEADER_FONT
            cell.fill = COL_HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        # Fill remaining columns in header row with gray
        for c in ["A", "B", "C", "D", "E", "F"]:
            if c not in col_headers:
                ws[f"{c}{row}"].fill = SECTION_FILL
        row += 1

        data_start = row
        all_cols = list(columns.keys())

        field_aliases = section.get("field_aliases", {})

        for ext_idx, ext in enumerate(matched):
            fields = ext.get("fields", {})
            # Resolve field aliases: if primary field not found, try alternates
            if field_aliases:
                for primary, alternates in field_aliases.items():
                    if primary not in fields:
                        for alt in alternates:
                            if alt in fields:
                                fields[primary] = fields[alt]
                                break
            for col, field_name in columns.items():
                _write_cell_value(ws, col, row, fields, field_name, ext, matched)

            # K-1 extras
            if sid == "k1":
                entity_name = get_str(fields, "partnership_name") or ext.get("payer_or_entity", "")
                extras_map = {
                    "box5_interest": "Box 5 (Interest)",
                    "box6a_ordinary_dividends": "Box 6a (Ordinary Dividends)",
                    "box6b_qualified_dividends": "Box 6b (Qualified Dividends)",
                    "box7_royalties": "Box 7 (Royalties)",
                    "box8_short_term_capital_gain": "Box 8 (ST Cap Gain)",
                    "box9a_long_term_capital_gain": "Box 9a (LT Cap Gain)",
                    "box9c_unrecaptured_1250": "Box 9c (Unrec 1250)",
                    "box10_net_1231_gain": "Box 10 (1231 Gain)",
                    "box11_other_income": "Box 11 (Other Income)",
                    "box12_section_179": "Box 12 (Sec 179)",
                    "box13_other_deductions": "Box 13 (Other Ded)",
                    "box14_self_employment": "Box 14 (SE Earnings)",
                    "box15_credits": "Box 15 (Credits)",
                    "box17_alt_min_tax": "Box 17 (AMT)",
                    "box18_tax_exempt_income": "Box 18 (Tax Exempt)",
                    "box19_distributions": "Box 19 (Distributions)",
                    "box20_other_info": "Box 20 (Other Info)",
                }
                for fkey, label in extras_map.items():
                    val = get_val(fields, fkey)
                    if val and val != 0:
                        k1_extras.append({"entity": entity_name, "line_reference": label, "description": "", "amount": val})
                for ci in ext.get("continuation_items", []):
                    if ci.get("amount") and ci["amount"] != 0:
                        k1_extras.append({"entity": entity_name, "line_reference": ci.get("line_reference", ""), "description": ci.get("description", ""), "amount": ci["amount"]})

            for bcol in list(columns.keys()):
                ws[f"{bcol}{row}"].border = THIN_BORDER
            row += 1

        data_end = row - 1
        if data_end >= data_start and sum_cols:
            for col in sum_cols:
                cell = ws[f"{col}{row}"]
                cell.value = f"=SUM({col}{data_start}:{col}{data_end})"
                cell.font = SUM_FONT
                cell.fill = SUM_FILL
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right")
                cell.border = SUM_BORDER
            ws[f"A{row}"].font = SUM_FONT
            ws[f"A{row}"].fill = SUM_FILL
            ws[f"A{row}"].value = "TOTAL"
            ws[f"A{row}"].border = SUM_BORDER
            if section.get("total_formula_col"):
                for col, formula in section["total_formula_col"].items():
                    parts = formula.split("+")
                    cell = ws[f"{col}{row}"]
                    cell.value = "=" + "+".join([f"{p}{row}" for p in parts])
                    cell.font = SUM_FONT
                    cell.number_format = MONEY_FMT
                    cell.alignment = Alignment(horizontal="right")
            row += 1

        for flag in flags:
            ws[f"A{row}"] = flag
            ws[f"A{row}"].font = FLAG_FONT
            row += 1
        row += 1

    # Schedule A — cross-references extracted data
    ws[f"A{row}"] = "Schedule A:"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"A{row}"].fill = SECTION_FILL
    # Column headers for Schedule A
    ws[f"C{row}"] = "Total"
    ws[f"C{row}"].font = COL_HEADER_FONT
    ws[f"C{row}"].fill = COL_HEADER_FILL
    ws[f"C{row}"].alignment = Alignment(horizontal="center")
    ws[f"E{row}"] = "Allowed"
    ws[f"E{row}"].font = COL_HEADER_FONT
    ws[f"E{row}"].fill = COL_HEADER_FILL
    ws[f"E{row}"].alignment = Alignment(horizontal="center")
    for c in ["B", "D", "F"]:
        ws[f"{c}{row}"].fill = SECTION_FILL
    row += 1

    # Medical
    ws[f"A{row}"] = "Medical:"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="333333")
    row += 1
    total_medical = 0
    medical_exts = [e for e in extractions if e.get("document_type") in ("receipt",)
                    and get_str(e.get("fields", {}), "category") == "medical"]
    for mext in medical_exts:
        mfields = mext.get("fields", {})
        mamt = get_val(mfields, "total_amount") or 0
        ws[f"A{row}"] = get_str(mfields, "vendor_name") or mext.get("payer_or_entity", "")
        mcell = ws[f"C{row}"]
        mcell.value = mamt
        mcell.number_format = MONEY_FMT
        mcell.alignment = Alignment(horizontal="right")
        total_medical += mamt
        row += 1
    if not medical_exts:
        ws[f"A{row}"] = "(none found)"
        ws[f"A{row}"].font = Font(italic=True, color="BBBBBB")
        row += 1
    ws[f"A{row}"] = "Total Medical"
    ws[f"A{row}"].font = SUM_FONT
    ws[f"A{row}"].fill = SUM_FILL
    ws[f"A{row}"].border = SUM_BORDER
    mcell = ws[f"D{row}"]
    mcell.value = total_medical
    mcell.number_format = MONEY_FMT
    mcell.alignment = Alignment(horizontal="right")
    mcell.font = SUM_FONT
    mcell.fill = SUM_FILL
    mcell.border = SUM_BORDER
    row += 1

    # Taxes
    ws[f"A{row}"] = "Taxes:"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="333333")
    row += 1

    # Income Taxes — State WH from W-2s
    total_state_wh = sum(get_val(e.get("fields", {}), "state_wh") or 0
                         for e in extractions if e.get("document_type") == "W-2")
    ws[f"A{row}"] = "Income Taxes:"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="333333")
    row += 1
    ws[f"A{row}"] = "State Withholding"
    tcell = ws[f"C{row}"]
    tcell.value = total_state_wh
    tcell.number_format = MONEY_FMT
    tcell.alignment = Alignment(horizontal="right")
    row += 1
    ws[f"A{row}"] = "Total State Tax"
    ws[f"A{row}"].font = SUM_FONT
    ws[f"A{row}"].fill = SUM_FILL
    ws[f"A{row}"].border = SUM_BORDER
    tcell = ws[f"D{row}"]
    tcell.value = total_state_wh
    tcell.number_format = MONEY_FMT
    tcell.alignment = Alignment(horizontal="right")
    tcell.font = SUM_FONT
    tcell.fill = SUM_FILL
    tcell.border = SUM_BORDER
    row += 1

    # Real Estate Taxes — from 1098 property_tax + property_tax_bill
    ws[f"A{row}"] = "Real Estate Taxes:"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="333333")
    row += 1
    total_re_tax = 0
    for rext in extractions:
        if rext.get("document_type") == "1098":
            prop_tax = (get_val(rext.get("fields", {}), "property_tax")
                        or get_val(rext.get("fields", {}), "real_estate_tax") or 0)
            if prop_tax:
                rentity = rext.get("payer_or_entity", "1098")
                ws[f"A{row}"] = f"{rentity} (1098)"
                rcell = ws[f"C{row}"]
                rcell.value = prop_tax
                rcell.number_format = MONEY_FMT
                rcell.alignment = Alignment(horizontal="right")
                total_re_tax += prop_tax
                row += 1
    for rext in extractions:
        if rext.get("document_type") == "property_tax_bill":
            tax_amt = (get_val(rext.get("fields", {}), "tax_amount")
                       or get_val(rext.get("fields", {}), "total_due")
                       or get_val(rext.get("fields", {}), "total_tax")
                       or get_val(rext.get("fields", {}), "total_estimated_tax") or 0)
            if tax_amt:
                addr = get_str(rext.get("fields", {}), "property_address") or "Property"
                ws[f"A{row}"] = addr
                rcell = ws[f"C{row}"]
                rcell.value = tax_amt
                rcell.number_format = MONEY_FMT
                rcell.alignment = Alignment(horizontal="right")
                total_re_tax += tax_amt
                row += 1
    if total_re_tax == 0:
        ws[f"A{row}"] = "(none found)"
        ws[f"A{row}"].font = Font(italic=True, color="BBBBBB")
        row += 1
    ws[f"A{row}"] = "Total Real Estate Tax"
    ws[f"A{row}"].font = SUM_FONT
    ws[f"A{row}"].fill = SUM_FILL
    ws[f"A{row}"].border = SUM_BORDER
    rcell = ws[f"D{row}"]
    rcell.value = total_re_tax
    rcell.number_format = MONEY_FMT
    rcell.alignment = Alignment(horizontal="right")
    rcell.font = SUM_FONT
    rcell.fill = SUM_FILL
    rcell.border = SUM_BORDER
    row += 1

    # Total Taxes
    total_taxes = total_state_wh + total_re_tax
    ws[f"A{row}"] = "Total Taxes:"
    ws[f"A{row}"].font = SUM_FONT
    ws[f"A{row}"].border = SUM_BORDER
    tcell = ws[f"D{row}"]
    tcell.value = total_taxes
    tcell.number_format = MONEY_FMT
    tcell.alignment = Alignment(horizontal="right")
    tcell.font = SUM_FONT
    tcell.border = SUM_BORDER
    ecell = ws[f"E{row}"]
    ecell.value = total_taxes
    ecell.number_format = MONEY_FMT
    ecell.alignment = Alignment(horizontal="right")
    ecell.font = SUM_FONT
    ecell.border = SUM_BORDER
    row += 1

    # Mortgage Interest — from 1098
    ws[f"A{row}"] = "Mortgage Interest:"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="333333")
    row += 1
    total_mortgage = 0
    for mext in extractions:
        if mext.get("document_type") == "1098":
            mort_int = (get_val(mext.get("fields", {}), "mortgage_interest")
                        or get_val(mext.get("fields", {}), "mortgage_interest_received") or 0)
            if mort_int:
                mentity = mext.get("payer_or_entity", "1098")
                ws[f"A{row}"] = mentity
                mcell = ws[f"C{row}"]
                mcell.value = mort_int
                mcell.number_format = MONEY_FMT
                mcell.alignment = Alignment(horizontal="right")
                total_mortgage += mort_int
                row += 1
    if total_mortgage == 0:
        ws[f"A{row}"] = "(none found)"
        ws[f"A{row}"].font = Font(italic=True, color="BBBBBB")
        row += 1
    ws[f"A{row}"] = "Total Mortgage Interest"
    ws[f"A{row}"].font = SUM_FONT
    ws[f"A{row}"].border = SUM_BORDER
    mcell = ws[f"D{row}"]
    mcell.value = total_mortgage
    mcell.number_format = MONEY_FMT
    mcell.alignment = Alignment(horizontal="right")
    mcell.font = SUM_FONT
    mcell.border = SUM_BORDER
    ecell = ws[f"E{row}"]
    ecell.value = total_mortgage
    ecell.number_format = MONEY_FMT
    ecell.alignment = Alignment(horizontal="right")
    ecell.font = SUM_FONT
    ecell.border = SUM_BORDER
    row += 1

    # Charitable Contributions
    ws[f"A{row}"] = "Donations:"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="333333")
    row += 1
    total_donations = 0
    for dext in extractions:
        if dext.get("document_type") == "charitable_receipt":
            damt = get_val(dext.get("fields", {}), "donation_amount") or 0
            if damt:
                dorg = get_str(dext.get("fields", {}), "organization_name") or dext.get("payer_or_entity", "")
                ws[f"A{row}"] = dorg
                dcell = ws[f"C{row}"]
                dcell.value = damt
                dcell.number_format = MONEY_FMT
                dcell.alignment = Alignment(horizontal="right")
                total_donations += damt
                row += 1
    if total_donations == 0:
        ws[f"A{row}"] = "(none found)"
        ws[f"A{row}"].font = Font(italic=True, color="BBBBBB")
        row += 1
    ws[f"A{row}"] = "Total Donations"
    ws[f"A{row}"].font = SUM_FONT
    ws[f"A{row}"].border = SUM_BORDER
    dcell = ws[f"D{row}"]
    dcell.value = total_donations
    dcell.number_format = MONEY_FMT
    dcell.alignment = Alignment(horizontal="right")
    dcell.font = SUM_FONT
    dcell.border = SUM_BORDER
    ecell = ws[f"E{row}"]
    ecell.value = total_donations
    ecell.number_format = MONEY_FMT
    ecell.alignment = Alignment(horizontal="right")
    ecell.font = SUM_FONT
    ecell.border = SUM_BORDER
    row += 1

    # Schedule A grand total
    sched_a_total = total_taxes + total_mortgage + total_donations
    row += 1
    ws[f"A{row}"] = "Total Schedule A"
    ws[f"A{row}"].font = SUM_FONT
    ws[f"A{row}"].border = SUM_BORDER
    gcell = ws[f"E{row}"]
    gcell.value = sched_a_total
    gcell.number_format = MONEY_FMT
    gcell.alignment = Alignment(horizontal="right")
    gcell.font = SUM_FONT
    gcell.border = SUM_BORDER
    row += 1

    # Column widths
    ws.column_dimensions["A"].width = 33
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 15


# ─── JOURNAL ENTRIES FORMAT ──────────────────────────────────────────────────

# Fields on tax documents that are informational, not dollar amounts to post.
# These must never become journal entry lines.
_NON_POSTING_FIELDS = {
    # K-1 informational
    "partnership_name", "partnership_ein", "partner_name", "entity_type",
    "partner_type", "profit_share_begin", "profit_share_end",
    "loss_share_begin", "loss_share_end", "capital_share_begin",
    "capital_share_end", "beginning_capital_account", "ending_capital_account",
    "current_year_net_income", "withdrawals_distributions", "capital_contributed",
    # Identifiers on any form
    "employer_ein", "payer_ein", "recipient_ssn_last4", "tax_year",
    "state_id", "account_number", "account_number_last4",
    # Statement period / dates
    "statement_period_start", "statement_period_end",
    "pay_period_start", "pay_period_end",
    # Counts / metadata
    "num_deposits", "num_withdrawals", "number_of_transactions",
    "hours_regular", "hours_overtime", "rate_regular", "rate_overtime",
    "distribution_code", "identifiable_event_code", "type_of_wager",
}


def _build_journal_entries(extractions, year):
    """Transform extractions into balanced double-entry journal entries.

    ACCOUNTING RULES:
    1. Every entry MUST balance: total debits == total credits.
       Entries that fail are flagged "UNBALANCED" and still included
       (so the operator sees them) but marked for review.
    2. Tax information documents (W-2, 1099-*, K-1, SSA-1099, 1098-*,
       5498-*, 1095-*) are NOT source transactions. They report income
       for tax return preparation. They do NOT generate journal entries.
    3. A bank deposit is not revenue until classified. Posts to
       "Unclassified Deposits" pending operator review.
    4. A withdrawal is not a known expense until classified. Posts to
       "Unclassified Expenses" pending operator review.
    5. Credit cards are liabilities (credit-normal). Purchases increase
       the liability (credit CC, debit expense). Payments decrease it.
    6. Payroll: debit Wages Expense for gross, credit each tax payable,
       credit Cash for net. If net is missing, compute it.

    Returns list of:
      {"date", "description", "lines": [{"account", "debit", "credit"}],
       "balanced": bool, "source_type": str}
    """
    journal = []

    def _post(date, description, lines, source_type=""):
        """Validate balance and append entry. Flag if unbalanced."""
        # Round everything to 2 decimals
        for line in lines:
            if line.get("debit"):
                line["debit"] = round(line["debit"], 2)
            else:
                line["debit"] = None
            if line.get("credit"):
                line["credit"] = round(line["credit"], 2)
            else:
                line["credit"] = None

        total_dr = round(sum(l["debit"] or 0 for l in lines), 2)
        total_cr = round(sum(l["credit"] or 0 for l in lines), 2)
        diff = round(total_dr - total_cr, 2)
        balanced = abs(diff) < 0.02

        # Auto-fix small rounding differences (< $1.00)
        if not balanced and 0.02 <= abs(diff) < 1.00 and total_dr > 0 and total_cr > 0:
            if diff > 0:
                lines.append({"account": "Rounding", "debit": None, "credit": abs(diff)})
            else:
                lines.append({"account": "Rounding", "debit": abs(diff), "credit": None})
            balanced = True

        entry = {
            "date": date,
            "description": description,
            "lines": lines,
            "balanced": balanced,
            "source_type": source_type,
        }
        if not balanced:
            entry["description"] = (
                f"\u26a0 UNBALANCED \u2014 {description} "
                f"(DR {total_dr:,.2f} \u2260 CR {total_cr:,.2f})"
            )
        journal.append(entry)

    def _get_cat(fields, field_name, default=""):
        """Read operator-assigned category from a field, if present."""
        fdata = fields.get(field_name)
        if isinstance(fdata, dict):
            return fdata.get("_operator_category", default)
        return default

    for ext in extractions:
        dtype = str(ext.get("document_type", ""))
        fields = ext.get("fields", {})
        entity = ext.get("payer_or_entity", "")

        # ═══════════════════════════════════════════════════════════════
        # TAX INFORMATION DOCUMENTS — DO NOT POST
        # ═══════════════════════════════════════════════════════════════
        # W-2, 1099-*, K-1, SSA-1099, 1098-*, 5498-*, 1095-*, W-2G
        # are information returns. They tell the IRS (and the taxpayer)
        # what happened, but they are NOT the transactions themselves.
        # The actual transactions are on bank statements, checks, etc.
        # These belong in tax_review format, not journal entries.
        if any(tag in dtype for tag in (
            "W-2", "1099", "K-1", "SSA", "1098", "5498", "1095", "W-2G",
        )):
            continue

        # ═══════════════════════════════════════════════════════════════
        # BANK STATEMENTS
        # ═══════════════════════════════════════════════════════════════
        # Cash is an ASSET (debit-normal).
        #   Deposit:    DR Cash / CR Unclassified Deposits
        #   Withdrawal: DR Unclassified Expenses / CR Cash
        #   Fee:        DR Bank Service Charges / CR Cash
        #   Interest:   DR Cash / CR Interest Income
        if "bank_statement" in dtype:
            stmt_date = get_str(fields, "statement_period_end") or f"12/31/{year}"
            bank = get_str(fields, "bank_name") or entity
            acct = get_str(fields, "account_number_last4") or ""
            cash_acct = f"Cash \u2014 {bank}" + (f" (...{acct[-4:]})" if acct else "")

            txn_nums = sorted(set(
                int(m.group(1)) for k in fields
                for m in [re.match(r"txn_(\d+)_", k)] if m
            ))

            if txn_nums:
                for n in txn_nums:
                    tdate = get_str(fields, f"txn_{n}_date") or stmt_date
                    tdesc = get_str(fields, f"txn_{n}_desc") or f"Transaction #{n}"
                    tamt = get_val(fields, f"txn_{n}_amount")
                    ttype = (get_str(fields, f"txn_{n}_type") or "").lower()
                    if not tamt:
                        continue
                    amt = round(abs(tamt), 2)
                    # Operator may have categorized this transaction
                    cat = _get_cat(fields, f"txn_{n}_amount")

                    if ttype in ("deposit", "transfer in", "credit"):
                        cr_acct = cat or "Unclassified Deposits"
                        _post(tdate, tdesc, [
                            {"account": cash_acct, "debit": amt, "credit": None},
                            {"account": cr_acct, "debit": None, "credit": amt},
                        ], "bank_txn")
                    elif ttype == "fee":
                        _post(tdate, f"{bank}: {tdesc}", [
                            {"account": cat or "Bank Service Charges", "debit": amt, "credit": None},
                            {"account": cash_acct, "debit": None, "credit": amt},
                        ], "bank_txn")
                    elif ttype == "interest":
                        _post(tdate, f"{bank}: {tdesc}", [
                            {"account": cash_acct, "debit": amt, "credit": None},
                            {"account": cat or "Interest Income", "debit": None, "credit": amt},
                        ], "bank_txn")
                    else:
                        dr_acct = cat or "Unclassified Expenses"
                        _post(tdate, tdesc, [
                            {"account": dr_acct, "debit": amt, "credit": None},
                            {"account": cash_acct, "debit": None, "credit": amt},
                        ], "bank_txn")
            else:
                # Summary-level entries
                deposits = get_val(fields, "total_deposits")
                if deposits:
                    _post(stmt_date, f"{cash_acct} \u2014 Total deposits", [
                        {"account": cash_acct, "debit": abs(deposits), "credit": None},
                        {"account": "Unclassified Deposits", "debit": None, "credit": abs(deposits)},
                    ], "bank_summary")
                withdrawals = get_val(fields, "total_withdrawals")
                if withdrawals:
                    _post(stmt_date, f"{cash_acct} \u2014 Total withdrawals", [
                        {"account": "Unclassified Expenses", "debit": abs(withdrawals), "credit": None},
                        {"account": cash_acct, "debit": None, "credit": abs(withdrawals)},
                    ], "bank_summary")

            # Fees / interest as separate entries (known categories)
            fees = get_val(fields, "fees_charged")
            if fees and not txn_nums:
                _post(stmt_date, f"{bank} \u2014 Service charges", [
                    {"account": "Bank Service Charges", "debit": abs(fees), "credit": None},
                    {"account": cash_acct, "debit": None, "credit": abs(fees)},
                ], "bank_summary")
            interest = get_val(fields, "interest_earned")
            if interest and not txn_nums:
                _post(stmt_date, f"{bank} \u2014 Interest earned", [
                    {"account": cash_acct, "debit": abs(interest), "credit": None},
                    {"account": "Interest Income", "debit": None, "credit": abs(interest)},
                ], "bank_summary")

        # ═══════════════════════════════════════════════════════════════
        # CREDIT CARD STATEMENTS
        # ═══════════════════════════════════════════════════════════════
        # Credit card is a LIABILITY (credit-normal).
        #   Purchase: DR Unclassified Expenses / CR Credit Card Payable
        #   Payment:  DR Credit Card Payable / CR Cash
        #   Interest: DR Interest Expense / CR Credit Card Payable
        elif "credit_card" in dtype:
            stmt_date = get_str(fields, "statement_period_end") or f"12/31/{year}"
            issuer = get_str(fields, "card_issuer") or entity
            acct_num = get_str(fields, "account_number_last4") or ""
            cc_acct = f"Credit Card Payable \u2014 {issuer}" + (
                f" (...{acct_num[-4:]})" if acct_num else "")

            txn_nums = sorted(set(
                int(m.group(1)) for k in fields
                for m in [re.match(r"txn_(\d+)_", k)] if m
            ))

            if txn_nums:
                for n in txn_nums:
                    tdate = get_str(fields, f"txn_{n}_date") or stmt_date
                    tdesc = get_str(fields, f"txn_{n}_desc") or f"CC Transaction #{n}"
                    tamt = get_val(fields, f"txn_{n}_amount")
                    if not tamt:
                        continue
                    amt = round(abs(tamt), 2)
                    cat = _get_cat(fields, f"txn_{n}_amount")
                    if tamt < 0:
                        # Payment / credit / return
                        cr_acct = cat or "Cash"
                        _post(tdate, f"{issuer}: {tdesc}", [
                            {"account": cc_acct, "debit": amt, "credit": None},
                            {"account": cr_acct, "debit": None, "credit": amt},
                        ], "cc_txn")
                    else:
                        # Purchase / charge
                        dr_acct = cat or "Unclassified Expenses"
                        _post(tdate, tdesc, [
                            {"account": dr_acct, "debit": amt, "credit": None},
                            {"account": cc_acct, "debit": None, "credit": amt},
                        ], "cc_txn")
            else:
                purchases = get_val(fields, "purchases")
                if purchases:
                    _post(stmt_date, f"{cc_acct} \u2014 Purchases", [
                        {"account": "Unclassified Expenses", "debit": abs(purchases), "credit": None},
                        {"account": cc_acct, "debit": None, "credit": abs(purchases)},
                    ], "cc_summary")
                payments = get_val(fields, "payments")
                if payments:
                    _post(stmt_date, f"{cc_acct} \u2014 Payment", [
                        {"account": cc_acct, "debit": abs(payments), "credit": None},
                        {"account": "Cash", "debit": None, "credit": abs(payments)},
                    ], "cc_summary")

            interest_chg = get_val(fields, "interest_charged")
            if interest_chg and not txn_nums:
                _post(stmt_date, f"{issuer} \u2014 Finance charges", [
                    {"account": "Interest Expense", "debit": abs(interest_chg), "credit": None},
                    {"account": cc_acct, "debit": None, "credit": abs(interest_chg)},
                ], "cc_summary")

        # ═══════════════════════════════════════════════════════════════
        # CHECKS
        # ═══════════════════════════════════════════════════════════════
        elif dtype == "check":
            check_date = get_str(fields, "check_date") or f"12/31/{year}"
            check_num = get_str(fields, "check_number") or ""
            payee = get_str(fields, "payee") or get_str(fields, "pay_to") or ""
            check_amt = get_val(fields, "check_amount")
            memo = get_str(fields, "memo_line") or ""
            cat = _get_cat(fields, "check_amount")
            desc = f"Check #{check_num}" if check_num else "Check"
            if payee:
                desc += f" to {payee}"
            if memo:
                desc += f" ({memo})"
            if check_amt:
                dr_acct = cat or "Unclassified Expenses"
                _post(check_date, desc, [
                    {"account": dr_acct, "debit": abs(check_amt), "credit": None},
                    {"account": "Cash", "debit": None, "credit": abs(check_amt)},
                ], "check")

        # ═══════════════════════════════════════════════════════════════
        # INVOICES
        # ═══════════════════════════════════════════════════════════════
        # DR Expense / CR Accounts Payable
        elif "invoice" in dtype:
            inv_date = get_str(fields, "invoice_date") or f"12/31/{year}"
            vendor = get_str(fields, "vendor_name") or entity
            inv_num = get_str(fields, "invoice_number") or ""
            total = get_val(fields, "total_amount")
            cat = _get_cat(fields, "total_amount")
            desc_text = get_str(fields, "description") or ""
            desc = f"{vendor}" + (f" Inv #{inv_num}" if inv_num else "")
            if desc_text:
                desc += f" \u2014 {desc_text}"
            if total:
                dr_acct = cat or "Unclassified Expenses"
                _post(inv_date, desc, [
                    {"account": dr_acct, "debit": abs(total), "credit": None},
                    {"account": "Accounts Payable", "debit": None, "credit": abs(total)},
                ], "invoice")

        # ═══════════════════════════════════════════════════════════════
        # RECEIPTS
        # ═══════════════════════════════════════════════════════════════
        # DR Expense (category if known) / CR Cash or Credit Card
        elif "receipt" in dtype:
            r_date = get_str(fields, "receipt_date") or f"12/31/{year}"
            vendor = get_str(fields, "vendor_name") or entity
            total = get_val(fields, "total_amount")
            cat = _get_cat(fields, "total_amount")
            category = get_str(fields, "category") or ""
            pay_method = (get_str(fields, "payment_method") or "").lower()

            # Priority: operator category > AI-extracted category > unclassified
            if cat:
                expense_acct = cat
            else:
                cat_map = {
                    "meals": "Meals & Entertainment",
                    "supplies": "Office Supplies",
                    "travel": "Auto & Travel",
                    "equipment": "Equipment",
                    "utilities": "Utilities",
                    "rent": "Rent",
                    "insurance": "Insurance",
                    "professional_services": "Legal & Professional",
                }
                expense_acct = cat_map.get(category.lower(), "Unclassified Expenses")

            if "credit" in pay_method or "card" in pay_method:
                contra = "Credit Card Payable"
            elif "check" in pay_method:
                contra = "Cash"
            else:
                contra = "Cash"

            if total:
                _post(r_date, vendor, [
                    {"account": expense_acct, "debit": abs(total), "credit": None},
                    {"account": contra, "debit": None, "credit": abs(total)},
                ], "receipt")

        # ═══════════════════════════════════════════════════════════════
        # PAYROLL (Check Stubs)
        # ═══════════════════════════════════════════════════════════════
        # DR Wages Expense (gross)
        # CR Federal WH Payable, State WH Payable, SS Payable, Medicare Payable
        # CR Cash (net pay)
        # If net is missing: compute as gross - sum(known deductions)
        elif "check_stub" in dtype:
            pay_date = get_str(fields, "pay_date") or f"12/31/{year}"
            employer = get_str(fields, "employer_name") or entity
            employee = get_str(fields, "employee_name") or ""
            label = f"Payroll \u2014 {employer}: {employee}" if employee else f"Payroll \u2014 {employer}"

            gross = get_val(fields, "gross_pay")
            if not gross:
                continue

            lines = [{"account": "Wages Expense", "debit": round(abs(gross), 2), "credit": None}]
            known_deductions = 0.0

            for acct_name, field_name in [
                ("Federal WH Payable", "federal_wh"),
                ("State WH Payable", "state_wh"),
                ("Social Security Payable", "social_security"),
                ("Medicare Payable", "medicare"),
            ]:
                val = get_val(fields, field_name)
                if val:
                    amt = round(abs(val), 2)
                    lines.append({"account": acct_name, "debit": None, "credit": amt})
                    known_deductions += amt

            net = get_val(fields, "net_pay")
            if net:
                lines.append({"account": "Cash", "debit": None, "credit": round(abs(net), 2)})
            else:
                # Compute net to force balance
                computed_net = round(abs(gross) - known_deductions, 2)
                if computed_net > 0:
                    lines.append({"account": "Cash (net \u2014 computed)", "debit": None, "credit": computed_net})

            _post(pay_date, label, lines, "payroll")

        # ═══════════════════════════════════════════════════════════════
        # LOAN / MORTGAGE PAYMENTS
        # ═══════════════════════════════════════════════════════════════
        # DR Loan Payable (principal) + DR Interest Expense + DR Escrow
        # CR Cash (total payment)
        elif "loan_statement" in dtype or "mortgage_statement" in dtype:
            pay_date = (get_str(fields, "payment_date")
                        or get_str(fields, "next_due_date")
                        or f"12/31/{year}")
            lender = get_str(fields, "lender") or entity
            principal = get_val(fields, "principal_paid")
            interest = get_val(fields, "interest_paid")
            escrow = get_val(fields, "escrow_paid")

            if principal or interest:
                lines = []
                if principal:
                    lines.append({"account": f"Loan Payable \u2014 {lender}",
                                  "debit": round(abs(principal), 2), "credit": None})
                if interest:
                    lines.append({"account": "Interest Expense",
                                  "debit": round(abs(interest), 2), "credit": None})
                if escrow:
                    lines.append({"account": "Escrow Deposit",
                                  "debit": round(abs(escrow), 2), "credit": None})
                total = round(sum(l["debit"] or 0 for l in lines), 2)
                lines.append({"account": "Cash", "debit": None, "credit": total})
                _post(pay_date, f"Loan payment \u2014 {lender}", lines, "loan")

        # ═══════════════════════════════════════════════════════════════
        # PROFIT & LOSS STATEMENTS (summary-level)
        # ═══════════════════════════════════════════════════════════════
        elif "profit_loss" in dtype:
            period_end = get_str(fields, "period_end") or f"12/31/{year}"
            revenue = get_val(fields, "total_revenue")
            expenses = (get_val(fields, "total_operating_expenses")
                        or get_val(fields, "total_expenses"))
            if revenue:
                _post(period_end, f"Revenue \u2014 {entity}", [
                    {"account": "Accounts Receivable", "debit": abs(revenue), "credit": None},
                    {"account": "Revenue", "debit": None, "credit": abs(revenue)},
                ], "pnl")
            if expenses:
                _post(period_end, f"Operating expenses \u2014 {entity}", [
                    {"account": "Operating Expenses", "debit": abs(expenses), "credit": None},
                    {"account": "Accounts Payable", "debit": None, "credit": abs(expenses)},
                ], "pnl")

        # All other document types (payroll_register, balance_sheet, etc.)
        # are reporting documents, not source transactions. Skip.

    journal.sort(key=lambda e: e.get("date", ""))
    return journal


def _populate_journal_entries(ws, extractions, year):
    """Write proper double-entry journal entries.

    Layout per entry:
      Row 1: Date | Account (debit) | Description | Debit amount |
      Row 2:      |   Account (credit, indented) |             |  | Credit amount
      ...repeat for multi-line entries
      [blank row separator between entries]

    Columns: A=Date, B=Account, C=Description, D=Debit, E=Credit
    Credits are indented with leading spaces in column B.
    """
    row = _write_title(ws, "Journal Entries", year)

    # Column headers
    for col, label in [("A", "Date"), ("B", "Account"), ("C", "Description"), ("D", "Debit"), ("E", "Credit")]:
        cell = ws[f"{col}{row}"]
        cell.value = label
        cell.font = COL_HEADER_FONT
        cell.fill = COL_HEADER_FILL
        cell.border = SECTION_BORDER
        if col in ("D", "E"):
            cell.alignment = Alignment(horizontal="right")
    row += 1

    journal = _build_journal_entries(extractions, year)

    if not journal:
        ws[f"A{row}"] = "(no journal entries generated)"
        ws[f"A{row}"].font = Font(italic=True, color="BBBBBB")
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 40
        for col in ["D", "E"]:
            ws.column_dimensions[col].width = 16
        return

    # Styles for journal entries
    credit_font = Font(size=10, color="444444")
    entry_sep_border = openpyxl.styles.Border(
        bottom=openpyxl.styles.Side(style="thin", color="CCCCCC"))
    entry_num_font = Font(size=8, color="999999")

    data_start = row
    entry_count = 0
    unbalanced_count = 0

    for entry in journal:
        entry_count += 1
        date_str = entry.get("date", "")
        desc = entry.get("description", "")
        lines = entry.get("lines", [])
        balanced = entry.get("balanced", True)
        entry_start_row = row

        if not balanced:
            unbalanced_count += 1

        # Separate debit lines (listed first) and credit lines
        debit_lines = [l for l in lines if l.get("debit")]
        credit_lines = [l for l in lines if l.get("credit")]

        # Write debit lines first
        first_line = True
        for dl in debit_lines:
            if first_line:
                ws[f"A{row}"] = date_str
                ws[f"C{row}"] = desc
                first_line = False
            ws[f"B{row}"] = dl["account"]
            ws[f"B{row}"].font = Font(bold=True, size=10)
            cell_d = ws[f"D{row}"]
            cell_d.value = dl["debit"]
            cell_d.number_format = MONEY_FMT
            cell_d.alignment = Alignment(horizontal="right")
            for c in ["A", "B", "C", "D", "E"]:
                ws[f"{c}{row}"].border = THIN_BORDER
            row += 1

        # Write credit lines (indented)
        for cl in credit_lines:
            if first_line:
                ws[f"A{row}"] = date_str
                ws[f"C{row}"] = desc
                first_line = False
            ws[f"B{row}"] = f"    {cl['account']}"
            ws[f"B{row}"].font = credit_font
            cell_e = ws[f"E{row}"]
            cell_e.value = cl["credit"]
            cell_e.number_format = MONEY_FMT
            cell_e.alignment = Alignment(horizontal="right")
            for c in ["A", "B", "C", "D", "E"]:
                ws[f"{c}{row}"].border = THIN_BORDER
            row += 1

        # Highlight unbalanced entries in orange
        if not balanced:
            for r in range(entry_start_row, row):
                for c in ["A", "B", "C", "D", "E"]:
                    ws[f"{c}{r}"].fill = REVIEW_FILL

        # Blank separator row between entries
        for c in ["A", "B", "C", "D", "E"]:
            ws[f"{c}{row}"].border = entry_sep_border
        row += 1

    # Grand totals
    data_end = row - 2  # skip last separator
    ws[f"B{row}"] = "TOTALS"
    ws[f"B{row}"].font = SUM_FONT
    ws[f"B{row}"].fill = SUM_FILL
    for col in ["D", "E"]:
        cell = ws[f"{col}{row}"]
        cell.value = f"=SUM({col}{data_start}:{col}{data_end})"
        cell.font = SUM_FONT
        cell.fill = SUM_FILL
        cell.number_format = MONEY_FMT
        cell.alignment = Alignment(horizontal="right")
    for c in ["A", "B", "C", "D", "E"]:
        ws[f"{c}{row}"].border = openpyxl.styles.Border(
            top=openpyxl.styles.Side(style="double", color="333333"))
    ws[f"A{row}"].fill = SUM_FILL
    ws[f"C{row}"].fill = SUM_FILL
    row += 1

    # Balance check row
    ws[f"B{row}"] = "BALANCE CHECK (should be zero):"
    ws[f"B{row}"].font = Font(bold=True, size=9, color="666666")
    bal_cell = ws[f"D{row}"]
    bal_cell.value = f"=D{row-1}-E{row-1}"
    bal_cell.number_format = MONEY_FMT
    bal_cell.font = Font(bold=True, size=10, color="CC0000")
    bal_cell.alignment = Alignment(horizontal="right")
    row += 1

    # Entry count
    count_text = f"{entry_count} journal entries"
    if unbalanced_count:
        count_text += f" ({unbalanced_count} UNBALANCED \u2014 highlighted orange)"
    ws[f"B{row}"] = count_text
    ws[f"B{row}"].font = Font(italic=True, size=9, color="CC0000" if unbalanced_count else "999999")

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 42
    for col in ["D", "E"]:
        ws.column_dimensions[col].width = 16


# ─── ACCOUNT BALANCES FORMAT ────────────────────────────────────────────────

ACCT_BAL_TEMPLATE_SECTIONS = [
    {
        "id": "acct_balances",
        "header": "Bank Account Balances:",
        "match_types": ["bank_statement", "bank_statement_deposit_slip"],
        "columns": {
            "A": "bank_name", "B": "account_number_last4",
            "C": "statement_period_start", "D": "statement_period_end",
            "E": "beginning_balance", "F": "total_deposits",
            "G": "total_withdrawals", "H": "fees_charged",
            "I": "interest_earned", "J": "ending_balance",
        },
        "col_headers": {
            "B": "Acct #", "C": "Period Start", "D": "Period End",
            "E": "Begin Bal", "F": "Deposits", "G": "Withdrawals",
            "H": "Fees", "I": "Interest", "J": "End Bal",
        },
        "sum_cols": ["E", "F", "G", "H", "I", "J"],
        "flags": [],
    },
    {
        "id": "credit_card_balances",
        "header": "Credit Card Balances:",
        "match_types": ["credit_card_statement"],
        "columns": {
            "A": "card_issuer", "B": "account_number_last4",
            "C": "statement_period_start", "D": "statement_period_end",
            "E": "previous_balance", "F": "purchases",
            "G": "payments", "H": "interest_charged", "I": "new_balance",
        },
        "col_headers": {
            "B": "Acct #", "C": "Period Start", "D": "Period End",
            "E": "Prev Bal", "F": "Purchases", "G": "Payments",
            "H": "Interest", "I": "New Bal",
        },
        "sum_cols": ["E", "F", "G", "H", "I"],
        "flags": [],
    },
]

ACCT_BAL_ALWAYS_SHOW = ["acct_balances", "credit_card_balances"]

def _populate_account_balances(ws, extractions, year):
    """Write account balance format: one row per bank statement."""
    row = _write_title(ws, "Account Balances", year)

    for section in ACCT_BAL_TEMPLATE_SECTIONS:
        sid = section["id"]
        match_types = section.get("match_types", [])
        matched = _match_exts(extractions, match_types) if match_types else []
        col_headers = section.get("col_headers", {})
        columns = section.get("columns", {})
        sum_cols = section.get("sum_cols", [])

        if not matched and sid not in ACCT_BAL_ALWAYS_SHOW:
            continue

        ws[f"A{row}"] = section["header"]
        ws[f"A{row}"].font = SECTION_FONT
        ws[f"A{row}"].fill = SECTION_FILL
        for col, label in col_headers.items():
            cell = ws[f"{col}{row}"]
            cell.value = label
            cell.font = COL_HEADER_FONT
            cell.fill = COL_HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        # Fill remaining columns in header row with gray
        for c in ["A", "B", "C", "D", "E", "F"]:
            if c not in col_headers:
                ws[f"{c}{row}"].fill = SECTION_FILL
        row += 1

        if not matched:
            ws[f"A{row}"] = "(no bank statements found)"
            ws[f"A{row}"].font = Font(italic=True, color="BBBBBB")
            row += 2
            continue

        matched = _dedup_by_ein(matched)
        data_start = row
        all_cols = list(columns.keys())

        for ext_idx, ext in enumerate(matched):
            fields = ext.get("fields", {})
            for col, field_name in columns.items():
                if field_name in ("bank_name",):
                    ws[f"{col}{row}"] = get_str(fields, field_name) or ext.get("payer_or_entity", "")
                else:
                    val = get_val(fields, field_name)
                    if val is None:
                        val = get_str(fields, field_name)
                    cell = ws[f"{col}{row}"]
                    cell.value = val
                    if isinstance(val, (int, float)):
                        cell.number_format = MONEY_FMT
                        cell.alignment = Alignment(horizontal="right")
            # Alternating row fill
            if ext_idx % 2 == 1:
                for bcol in all_cols:
                    cell = ws[f"{bcol}{row}"]
                    if cell.fill == PatternFill() or cell.fill is None:
                        cell.fill = ALT_ROW_FILL
            for bcol in all_cols:
                ws[f"{bcol}{row}"].border = THIN_BORDER
            row += 1

        data_end = row - 1
        if data_end >= data_start and sum_cols:
            ws[f"A{row}"] = "TOTAL"
            ws[f"A{row}"].font = SUM_FONT
            ws[f"A{row}"].fill = SUM_FILL
            for col in sum_cols:
                cell = ws[f"{col}{row}"]
                cell.value = f"=SUM({col}{data_start}:{col}{data_end})"
                cell.font = SUM_FONT
                cell.fill = SUM_FILL
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right")
                cell.border = SUM_BORDER
            ws[f"A{row}"].border = SUM_BORDER
            row += 1

    # ─── Transaction Detail Section ───
    row += 1
    all_txns = []
    for ext in extractions:
        if ext.get("document_type") not in ("bank_statement", "bank_statement_deposit_slip", "check", "check_stub", "credit_card_statement"):
            continue
        fields = ext.get("fields", {})
        entity = ext.get("payer_or_entity", "")
        # Gather numbered transaction fields: txn_N_date, txn_N_desc, txn_N_amount, txn_N_type
        txn_nums = set()
        for k in fields:
            m = __import__("re").match(r"txn_(\d+)_", k)
            if m:
                txn_nums.add(int(m.group(1)))
        for n in sorted(txn_nums):
            txn_date = get_str(fields, f"txn_{n}_date") or ""
            txn_desc = get_str(fields, f"txn_{n}_desc") or ""
            txn_amt = get_val(fields, f"txn_{n}_amount")
            txn_type = get_str(fields, f"txn_{n}_type") or ""
            all_txns.append((entity, txn_date, txn_desc, txn_amt, txn_type))

    if all_txns:
        ws[f"A{row}"] = "Transaction Detail:"
        ws[f"A{row}"].font = SECTION_FONT
        ws[f"A{row}"].border = SECTION_BORDER
        for col, label in [("B", "Date"), ("C", "Description"), ("D", "Amount"), ("E", "Type")]:
            cell = ws[f"{col}{row}"]
            cell.value = label
            cell.font = COL_HEADER_FONT
            cell.fill = COL_HEADER_FILL
            cell.alignment = Alignment(horizontal="right") if col == "D" else Alignment(horizontal="left")
            cell.border = SECTION_BORDER
        row += 1
        txn_data_start = row
        for txn_idx, (entity, tdate, tdesc, tamt, ttype) in enumerate(all_txns):
            ws[f"A{row}"] = entity
            ws[f"B{row}"] = tdate
            ws[f"C{row}"] = tdesc
            cell = ws[f"D{row}"]
            cell.value = tamt
            if isinstance(tamt, (int, float)):
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right")
            ws[f"E{row}"] = ttype
            if txn_idx % 2 == 1:
                for bcol in ["A", "B", "C", "D", "E"]:
                    cell = ws[f"{bcol}{row}"]
                    if cell.fill == PatternFill() or cell.fill is None:
                        cell.fill = ALT_ROW_FILL
            for bcol in ["A", "B", "C", "D", "E"]:
                ws[f"{bcol}{row}"].border = THIN_BORDER
            row += 1
        txn_data_end = row - 1
        # Total row for transaction amounts
        ws[f"A{row}"] = "TOTAL"
        ws[f"A{row}"].font = SUM_FONT
        ws[f"A{row}"].fill = SUM_FILL
        cell = ws[f"D{row}"]
        cell.value = f"=SUM(D{txn_data_start}:D{txn_data_end})"
        cell.font = SUM_FONT
        cell.fill = SUM_FILL
        cell.number_format = MONEY_FMT
        cell.alignment = Alignment(horizontal="right")
        cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style="thin", color="999999"))
        row += 1

    ws.column_dimensions["A"].width = 30
    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        ws.column_dimensions[col].width = 16


# ─── TRIAL BALANCE FORMAT ───────────────────────────────────────────────────

def _populate_trial_balance(ws, extractions, year):
    """Write a trial balance: accounts with debit and credit totals.

    Builds journal entries first, then aggregates by account into
    a standard trial balance (Account | Debit | Credit) with totals.
    """
    row = _write_title(ws, "Trial Balance", year)

    # Column headers
    for col, label, align in [("A", "Account", "left"), ("B", "Debit", "right"),
                               ("C", "Credit", "right"), ("D", "Net Balance", "right")]:
        cell = ws[f"{col}{row}"]
        cell.value = label
        cell.font = DARK_HEADER_FONT
        cell.fill = DARK_HEADER_FILL
        cell.border = SECTION_BORDER
        cell.alignment = Alignment(horizontal=align)
    row += 1

    # Build journal entries and aggregate by account
    journal = _build_journal_entries(extractions, year)
    account_totals = {}  # account_name → {"debit": float, "credit": float}

    for entry in journal:
        for line in entry.get("lines", []):
            acct = line.get("account", "Uncategorized")
            dr = line.get("debit") or 0
            cr = line.get("credit") or 0
            if acct not in account_totals:
                account_totals[acct] = {"debit": 0.0, "credit": 0.0}
            account_totals[acct]["debit"] += dr
            account_totals[acct]["credit"] += cr

    if not account_totals:
        ws[f"A{row}"] = "(no account data \u2014 upload bank statements, invoices, or receipts)"
        ws[f"A{row}"].font = Font(italic=True, color="BBBBBB")
        _tb_col_widths(ws)
        return

    # Sort accounts: Assets/Expenses (debit-normal) first, then Liabilities/Revenue (credit-normal)
    # Simple heuristic: sort by net balance direction, then alphabetically
    def sort_key(item):
        acct, tots = item
        net = tots["debit"] - tots["credit"]
        # Debit-normal accounts first (positive net = debit heavy)
        return (0 if net >= 0 else 1, acct.lower())

    data_start = row
    for idx, (acct, tots) in enumerate(sorted(account_totals.items(), key=sort_key)):
        ws[f"A{row}"] = acct
        dr_cell = ws[f"B{row}"]
        cr_cell = ws[f"C{row}"]
        net_cell = ws[f"D{row}"]

        if tots["debit"] > 0:
            dr_cell.value = round(tots["debit"], 2)
            dr_cell.number_format = MONEY_FMT
            dr_cell.alignment = Alignment(horizontal="right")
        if tots["credit"] > 0:
            cr_cell.value = round(tots["credit"], 2)
            cr_cell.number_format = MONEY_FMT
            cr_cell.alignment = Alignment(horizontal="right")

        net = round(tots["debit"] - tots["credit"], 2)
        net_cell.value = net
        net_cell.number_format = MONEY_FMT
        net_cell.alignment = Alignment(horizontal="right")
        if net < 0:
            net_cell.font = Font(color="CC0000")

        # Alternating rows
        if idx % 2 == 1:
            for c in ["A", "B", "C", "D"]:
                ws[f"{c}{row}"].fill = ALT_ROW_FILL
        for c in ["A", "B", "C", "D"]:
            ws[f"{c}{row}"].border = THIN_BORDER
        row += 1

    data_end = row - 1

    # Totals row
    for c in ["A", "B", "C", "D"]:
        ws[f"{c}{row}"].fill = SUM_FILL
        ws[f"{c}{row}"].font = SUM_FONT
        ws[f"{c}{row}"].border = openpyxl.styles.Border(
            top=openpyxl.styles.Side(style="double", color="333333"))
    ws[f"A{row}"] = "TOTALS"
    for col in ["B", "C", "D"]:
        cell = ws[f"{col}{row}"]
        cell.value = f"=SUM({col}{data_start}:{col}{data_end})"
        cell.number_format = MONEY_FMT
        cell.alignment = Alignment(horizontal="right")
    row += 1

    # Balance check
    ws[f"A{row}"] = "BALANCE CHECK (Debits \u2212 Credits, should be zero):"
    ws[f"A{row}"].font = Font(bold=True, size=9, color="666666")
    bal_cell = ws[f"B{row}"]
    bal_cell.value = f"=B{row-1}-C{row-1}"
    bal_cell.number_format = MONEY_FMT
    bal_cell.font = Font(bold=True, size=10, color="CC0000")
    bal_cell.alignment = Alignment(horizontal="right")
    row += 2

    # Account count
    ws[f"A{row}"] = f"{len(account_totals)} accounts from {len(journal)} journal entries"
    ws[f"A{row}"].font = Font(italic=True, size=9, color="999999")

    _tb_col_widths(ws)


def _tb_col_widths(ws):
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18


# ─── TRANSACTION REGISTER FORMAT ────────────────────────────────────────────

def _populate_transaction_register(ws, extractions, year):
    """Write a chronological transaction register with running balance.

    Like a checkbook register: Date | Description | Source | Debit | Credit | Balance
    Every transaction from all documents in one flat, time-ordered list.
    """
    row = _write_title(ws, "Transaction Register", year)

    # Column headers
    headers = [("A", "Date", "left"), ("B", "Description", "left"), ("C", "Source", "left"),
               ("D", "Type", "left"), ("E", "Debit (In)", "right"), ("F", "Credit (Out)", "right"),
               ("G", "Balance", "right")]
    for col, label, align in headers:
        cell = ws[f"{col}{row}"]
        cell.value = label
        cell.font = DARK_HEADER_FONT
        cell.fill = DARK_HEADER_FILL
        cell.border = SECTION_BORDER
        cell.alignment = Alignment(horizontal=align)
    row += 1

    # Gather all transactions from all extractions
    txns = []  # list of (date_str, description, source, type, amount_signed)

    for ext in extractions:
        dtype = str(ext.get("document_type", ""))
        fields = ext.get("fields", {})
        entity = ext.get("payer_or_entity", "")

        # Bank statement / credit card individual transactions
        if "bank_statement" in dtype or "credit_card" in dtype:
            bank = get_str(fields, "bank_name") or get_str(fields, "card_issuer") or entity
            acct = get_str(fields, "account_number_last4") or ""
            source = f"{bank}" + (f" (...{acct[-4:]})" if acct else "")

            txn_nums = sorted(set(int(m.group(1)) for k in fields
                                  for m in [re.match(r"txn_(\d+)_", k)] if m))
            for n in txn_nums:
                tdate = get_str(fields, f"txn_{n}_date") or ""
                tdesc = get_str(fields, f"txn_{n}_desc") or ""
                tamt = get_val(fields, f"txn_{n}_amount")
                ttype = get_str(fields, f"txn_{n}_type") or "transaction"
                if tamt is not None:
                    txns.append((tdate, tdesc, source, ttype, tamt))

            # If no individual txns, add summary lines
            if not txn_nums:
                stmt_date = get_str(fields, "statement_period_end") or f"12/31/{year}"
                deposits = get_val(fields, "total_deposits")
                if deposits:
                    txns.append((stmt_date, f"{source} \u2014 Total deposits", source, "deposit", abs(deposits)))
                withdrawals = get_val(fields, "total_withdrawals")
                if withdrawals:
                    txns.append((stmt_date, f"{source} \u2014 Total withdrawals", source, "withdrawal", -abs(withdrawals)))
                fees = get_val(fields, "fees_charged")
                if fees:
                    txns.append((stmt_date, f"{source} \u2014 Fees", source, "fee", -abs(fees)))
                interest = get_val(fields, "interest_earned")
                if interest:
                    txns.append((stmt_date, f"{source} \u2014 Interest", source, "interest", abs(interest)))

        elif dtype == "check":
            check_date = get_str(fields, "check_date") or f"12/31/{year}"
            check_num = get_str(fields, "check_number") or ""
            payee = get_str(fields, "payee") or get_str(fields, "pay_to") or ""
            check_amt = get_val(fields, "check_amount")
            desc = f"Check #{check_num}" if check_num else "Check"
            if payee:
                desc += f" to {payee}"
            if check_amt:
                txns.append((check_date, desc, "Check", "check", -abs(check_amt)))

        elif "invoice" in dtype:
            inv_date = get_str(fields, "invoice_date") or f"12/31/{year}"
            vendor = get_str(fields, "vendor_name") or entity
            inv_num = get_str(fields, "invoice_number") or ""
            total = get_val(fields, "total_amount")
            desc = f"{vendor}" + (f" Inv #{inv_num}" if inv_num else "")
            if total:
                txns.append((inv_date, desc, "Invoice", "expense", -abs(total)))

        elif "receipt" in dtype:
            r_date = get_str(fields, "receipt_date") or f"12/31/{year}"
            vendor = get_str(fields, "vendor_name") or entity
            total = get_val(fields, "total_amount")
            category = get_str(fields, "category") or "expense"
            if total:
                txns.append((r_date, vendor, "Receipt", category, -abs(total)))

        elif "check_stub" in dtype:
            pay_date = get_str(fields, "pay_date") or f"12/31/{year}"
            employer = get_str(fields, "employer_name") or entity
            net = get_val(fields, "net_pay")
            if net:
                txns.append((pay_date, f"Payroll \u2014 {employer}", "Payroll", "payroll", abs(net)))

    if not txns:
        ws[f"A{row}"] = "(no transactions found \u2014 upload bank statements, invoices, or receipts)"
        ws[f"A{row}"].font = Font(italic=True, color="BBBBBB")
        _tr_col_widths(ws)
        return

    # Sort by date
    def date_sort_key(t):
        d = t[0]
        # Try to parse common date formats for proper sorting
        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y"):
            try:
                return datetime.strptime(d, fmt)
            except (ValueError, TypeError):
                continue
        return datetime(int(year), 12, 31)  # fallback: end of year

    txns.sort(key=date_sort_key)

    # Start running balance from earliest beginning balance (if available)
    # This is fundamental — a register without a starting point is meaningless
    opening_balance = 0.0
    for ext in extractions:
        if "bank_statement" in str(ext.get("document_type", "")):
            begin = get_val(ext.get("fields", {}), "beginning_balance")
            if begin is not None:
                opening_balance = begin
                break  # Use the first one found (earliest in document order)

    data_start = row
    running_balance = opening_balance

    # Write opening balance row if we have one
    if opening_balance != 0.0:
        ws[f"A{row}"] = ""
        ws[f"B{row}"] = "Opening Balance"
        ws[f"B{row}"].font = Font(bold=True, italic=True, size=10)
        ws[f"G{row}"] = round(opening_balance, 2)
        ws[f"G{row}"].number_format = MONEY_FMT
        ws[f"G{row}"].alignment = Alignment(horizontal="right")
        ws[f"G{row}"].font = Font(bold=True)
        for c in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{c}{row}"].fill = SUM_FILL
            ws[f"{c}{row}"].border = THIN_BORDER
        row += 1

    debit_font = Font(color="006600")
    credit_font = Font(color="CC0000")

    for idx, (tdate, tdesc, source, ttype, amount) in enumerate(txns):
        ws[f"A{row}"] = tdate
        ws[f"B{row}"] = tdesc
        ws[f"C{row}"] = source
        ws[f"D{row}"] = ttype.title()

        if amount >= 0:
            # Debit (money in)
            dr_cell = ws[f"E{row}"]
            dr_cell.value = round(amount, 2)
            dr_cell.number_format = MONEY_FMT
            dr_cell.alignment = Alignment(horizontal="right")
            dr_cell.font = debit_font
        else:
            # Credit (money out)
            cr_cell = ws[f"F{row}"]
            cr_cell.value = round(abs(amount), 2)
            cr_cell.number_format = MONEY_FMT
            cr_cell.alignment = Alignment(horizontal="right")
            cr_cell.font = credit_font

        running_balance += amount
        bal_cell = ws[f"G{row}"]
        bal_cell.value = round(running_balance, 2)
        bal_cell.number_format = MONEY_FMT
        bal_cell.alignment = Alignment(horizontal="right")
        if running_balance < 0:
            bal_cell.font = Font(bold=True, color="CC0000")

        # Alternating rows
        if idx % 2 == 1:
            for c in ["A", "B", "C", "D", "E", "F", "G"]:
                cell = ws[f"{c}{row}"]
                if cell.fill == PatternFill():
                    cell.fill = ALT_ROW_FILL
        for c in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{c}{row}"].border = THIN_BORDER
        row += 1

    data_end = row - 1

    # Totals row
    for c in ["A", "B", "C", "D", "E", "F", "G"]:
        ws[f"{c}{row}"].fill = SUM_FILL
        ws[f"{c}{row}"].font = SUM_FONT
        ws[f"{c}{row}"].border = openpyxl.styles.Border(
            top=openpyxl.styles.Side(style="double", color="333333"))
    ws[f"A{row}"] = "TOTALS"
    for col in ["E", "F"]:
        cell = ws[f"{col}{row}"]
        cell.value = f"=SUM({col}{data_start}:{col}{data_end})"
        cell.number_format = MONEY_FMT
        cell.alignment = Alignment(horizontal="right")
    ws[f"G{row}"] = round(running_balance, 2)
    ws[f"G{row}"].number_format = MONEY_FMT
    ws[f"G{row}"].alignment = Alignment(horizontal="right")
    if running_balance < 0:
        ws[f"G{row}"].font = Font(bold=True, color="CC0000")
    row += 2

    # Summary
    ws[f"A{row}"] = f"{len(txns)} transactions"
    ws[f"A{row}"].font = Font(italic=True, size=9, color="999999")

    _tr_col_widths(ws)


def _tr_col_widths(ws):
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
