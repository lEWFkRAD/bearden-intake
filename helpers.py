"""
Bearden Document Intake — Shared Helpers
=========================================
Pure helper functions and shared state extracted from app.py
for use across Flask blueprints.

NO Flask imports here — these are pure Python helpers.
"""

import os
import sys
import json
import threading
import uuid
import re
import subprocess
from pathlib import Path
from datetime import datetime
from io import BytesIO

import db as appdb

# ─── Configuration / Directory Paths ─────────────────────────────────────────

BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
UPLOAD_DIR = DATA_DIR / "uploads"
OUTPUT_DIR = DATA_DIR / "outputs"
CLIENTS_DIR = BASE_DIR / "clients"
PAGES_DIR = DATA_DIR / "page_images"
VERIFY_DIR = BASE_DIR / "verifications"
JOBS_FILE = DATA_DIR / "jobs_history.json"

for _d in [DATA_DIR, UPLOAD_DIR, OUTPUT_DIR, CLIENTS_DIR, PAGES_DIR, VERIFY_DIR]:
    _d.mkdir(exist_ok=True)

VENDOR_CATEGORIES_FILE = DATA_DIR / "vendor_categories.json"
DB_PATH = DATA_DIR / "bearden.db"

# ─── Shared State ────────────────────────────────────────────────────────────

_start_time = datetime.now()
_app_version = "5.2"

VALID_DOC_TYPES = {"tax_returns", "bank_statements", "trust_documents", "bookkeeping", "payroll", "other"}

# In-memory job tracking (persisted to SQLite via appdb)
jobs = {}
_jobs_lock = threading.Lock()
_active_procs = {}  # job_id -> subprocess.Popen for cancellation

# ─── File Security ───────────────────────────────────────────────────────────

def _secure_file(path):
    """Set restrictive permissions on sensitive files (owner-only read/write)."""
    try:
        os.chmod(str(path), 0o600)
    except OSError:
        pass  # Non-fatal: may fail on some filesystems


def _client_dir(client_name, doc_type, year):
    """Build a per-client output directory:  clients/<Client Name>/<doc_type>/<year>/"""
    safe_client = re.sub(r'[^\w\s\-\.,()]', '', client_name).strip() or "Unknown Client"
    safe_client = safe_client.title()
    type_labels = {
        "tax_returns": "Tax Returns",
        "bank_statements": "Bank Statements",
        "trust_documents": "Trust Documents",
        "bookkeeping": "Bookkeeping",
        "payroll": "Payroll",
        "other": "Other Documents",
    }
    type_folder = type_labels.get(doc_type, "Other")
    client_dir = CLIENTS_DIR / safe_client / type_folder / str(year)
    client_dir.mkdir(parents=True, exist_ok=True)
    return client_dir

# ─── Job Persistence ────────────────────────────────────────────────────────

def load_jobs():
    """Load all jobs from SQLite into the in-memory dict.
    Uses .clear() + in-place updates to preserve reference identity
    so all modules sharing this dict stay in sync."""
    try:
        all_jobs = appdb.list_jobs()
        jobs.clear()
        for j in all_jobs:
            jid = j.get("id", "")
            if jid:
                j["log"] = []
                jobs[jid] = j
    except Exception as e:
        print(f"  Warning: Could not load jobs from database: {e}")
        jobs.clear()


def save_jobs():
    """Persist all jobs to SQLite (upsert)."""
    with _jobs_lock:
        for jid, j in jobs.items():
            try:
                appdb.save_job(j)
            except Exception as e:
                print(f"  Warning: Could not save job {jid}: {e}")


def _sync_job_to_db(job_id):
    """Sync a single job to the database."""
    job = jobs.get(job_id)
    if job:
        try:
            appdb.save_job(job)
        except Exception:
            pass

# ─── Chart of Accounts + Vendor Memory ───────────────────────────────────────

CHART_OF_ACCOUNTS = {
    "Expense": [
        "Advertising & Marketing",
        "Auto & Travel",
        "Bank Service Charges",
        "Computer & Internet",
        "Depreciation",
        "Dues & Subscriptions",
        "Equipment",
        "Insurance",
        "Interest Expense",
        "Legal & Professional",
        "Meals & Entertainment",
        "Office Supplies",
        "Payroll Expenses",
        "Rent",
        "Repairs & Maintenance",
        "Taxes & Licenses",
        "Telephone",
        "Utilities",
        "Miscellaneous Expense",
    ],
    "Income": [
        "Service Revenue",
        "Product Sales",
        "Interest Income",
        "Rental Income",
        "Refund / Rebate",
        "Other Income",
    ],
    "Other": [
        "Owner Draw / Distribution",
        "Owner Contribution / Investment",
        "Loan Proceeds",
        "Loan Payment",
        "Transfer Between Accounts",
    ],
}

# Flat list for validation
ALL_ACCOUNTS = []
for _grp in CHART_OF_ACCOUNTS.values():
    ALL_ACCOUNTS.extend(_grp)


def _normalize_vendor(desc):
    """Normalize a vendor/payee name for matching.
    'GEORGIA POWER COMPANY #12345' -> 'GEORGIA POWER'
    'WAL-MART SUPER CENTER 0423' -> 'WAL-MART SUPER CENTER'
    """
    if not desc:
        return ""
    s = str(desc).upper().strip()
    # Strip trailing reference/store numbers
    s = re.sub(r'[\s#*]+\d{2,}$', '', s)
    # Strip common suffixes
    s = re.sub(r'\s+(LLC|INC|CORP|CO|COMPANY|LTD|LP|NA|N\.A\.)\.?\s*$', '', s, flags=re.IGNORECASE)
    # Strip trailing punctuation
    s = s.rstrip(' .,;:*#-')
    return s.strip()


def _load_vendor_categories():
    try:
        return appdb.get_vendor_categories()
    except Exception:
        return {}


def _save_vendor_categories(data):
    for vendor, info in data.items():
        cat = info.get("category", "") if isinstance(info, dict) else str(info)
        try:
            appdb.set_vendor_category(vendor, cat, info)
        except Exception as e:
            print(f"  Warning: Could not save vendor category {vendor}: {e}")


def _learn_vendor_category(vendor_desc, category):
    """Record that vendor_desc was categorized as category."""
    if not vendor_desc or not category:
        return
    norm = _normalize_vendor(vendor_desc)
    if not norm or len(norm) < 2:
        return
    data = _load_vendor_categories()
    existing = data.get(norm, {})
    data[norm] = {
        "category": category,
        "count": existing.get("count", 0) + 1,
        "last_used": datetime.now().isoformat(),
        "original": vendor_desc,  # keep one raw example
    }
    _save_vendor_categories(data)


def _suggest_category(vendor_desc):
    """Look up a vendor in the learned map. Returns category or ''."""
    if not vendor_desc:
        return ""
    norm = _normalize_vendor(vendor_desc)
    if not norm:
        return ""
    data = _load_vendor_categories()
    entry = data.get(norm)
    if entry:
        return entry.get("category", "")
    # Try prefix match (e.g., "WALMART SUPERCENTER" matches "WALMART")
    for known, info in data.items():
        if norm.startswith(known) or known.startswith(norm):
            return info.get("category", "")
    return ""

# ─── Job Sanitization ───────────────────────────────────────────────────────

def _sanitize_job(j):
    """Make a JSON-safe copy of a job dict (no None keys, no log/pdf_path)."""
    safe = {}
    for k, v in j.items():
        if k in ("log", "pdf_path"):
            continue
        if k == "stats" and isinstance(v, dict):
            v = dict(v)
            for sk in ("methods", "confidences"):
                if sk in v and isinstance(v[sk], dict):
                    v[sk] = {str(dk) if dk is None else dk: dv for dk, dv in v[sk].items()}
        safe[k] = v
    return safe

# ─── Verification Helpers ────────────────────────────────────────────────────

def _verify_path(job_id):
    return VERIFY_DIR / f"{job_id}.json"


def _load_verifications(job_id):
    """Load verifications from JSON file (backward compat for _regen_excel, batch-categories)."""
    p = _verify_path(job_id)
    if p.exists():
        try:
            with open(p) as f:
                return json.load(f)
        except (ValueError, OSError):
            pass
    return {"fields": {}, "updated": None, "reviewer": ""}


def _save_verifications(job_id, data):
    """Save verifications to JSON file AND update job summary."""
    data["updated"] = datetime.now().isoformat()
    with open(_verify_path(job_id), "w") as f:
        json.dump(data, f, indent=2, default=str)
    _update_verify_summary(job_id, data)


def _update_verify_summary(job_id, vdata):
    job = jobs.get(job_id)
    if not job:
        return
    fields = vdata.get("fields", {})
    total = len(fields)
    confirmed = sum(1 for f in fields.values() if f.get("status") == "confirmed")
    corrected = sum(1 for f in fields.values() if f.get("status") == "corrected")
    flagged = sum(1 for f in fields.values() if f.get("status") == "flagged")
    job["verification"] = {
        "reviewed": confirmed + corrected + flagged,
        "confirmed": confirmed,
        "corrected": corrected,
        "flagged": flagged,
        "reviewer": vdata.get("reviewer", ""),
        "updated": vdata.get("updated"),
    }
    save_jobs()

# ─── Batch Categorization Helpers ────────────────────────────────────────────

def _gather_uncategorized(job_ids=None, client_name=None):
    """Gather all transactions needing categorization across jobs.

    Returns list of:
      {job_id, page, ext_idx, txn_num, date, desc, amount, type, source,
       current_category, suggested_category, vendor_norm}
    """
    target_jobs = {}
    for jid, j in jobs.items():
        if j.get("status") != "complete":
            continue
        if job_ids and jid not in job_ids:
            continue
        if client_name and j.get("client_name", "").lower() != client_name.lower():
            continue
        target_jobs[jid] = j

    items = []
    for jid, j in target_jobs.items():
        log_path = j.get("output_log")
        if not log_path or not os.path.exists(log_path):
            continue
        try:
            with open(log_path) as f:
                log_data = json.load(f)
        except (json.JSONDecodeError, IOError):
            continue

        vdata = _load_verifications(jid)
        vfields = vdata.get("fields", {})

        for ext in log_data.get("extractions", []):
            dtype = str(ext.get("document_type", ""))
            fields = ext.get("fields", {})
            entity = ext.get("payer_or_entity", "")
            page = ext.get("_page", 0)

            # Find ext index for this page
            page_exts = [e for e in log_data.get("extractions", []) if e.get("_page") == page]
            ext_idx = 0
            for ei, pe in enumerate(page_exts):
                if pe is ext:
                    ext_idx = ei
                    break

            # Bank/CC transactions
            if "bank_statement" in dtype or "credit_card" in dtype:
                bank = ""
                for k in ["bank_name", "card_issuer"]:
                    v = fields.get(k)
                    bank = (v.get("value", "") if isinstance(v, dict) else str(v or "")) if v else ""
                    if bank:
                        break
                source = bank or entity

                txn_nums = sorted(set(
                    int(m.group(1)) for k in fields
                    for m in [re.match(r"txn_(\d+)_", k)] if m
                ))
                for n in txn_nums:
                    amt_key = f"txn_{n}_amount"
                    vk = f"{page}:{ext_idx}:{amt_key}"
                    vstate = vfields.get(vk, {})
                    current_cat = vstate.get("category", "")

                    date_f = fields.get(f"txn_{n}_date")
                    desc_f = fields.get(f"txn_{n}_desc")
                    amt_f = fields.get(amt_key)
                    type_f = fields.get(f"txn_{n}_type")

                    date_v = (date_f.get("value", "") if isinstance(date_f, dict) else str(date_f or "")) if date_f else ""
                    desc_v = (desc_f.get("value", "") if isinstance(desc_f, dict) else str(desc_f or "")) if desc_f else ""
                    amt_v = (amt_f.get("value") if isinstance(amt_f, dict) else amt_f) if amt_f else None
                    type_v = (type_f.get("value", "") if isinstance(type_f, dict) else str(type_f or "")) if type_f else ""

                    if amt_v is None:
                        continue

                    norm = _normalize_vendor(desc_v)
                    suggested = _suggest_category(desc_v) if not current_cat else ""

                    items.append({
                        "job_id": jid, "page": page, "ext_idx": ext_idx,
                        "field_key": vk, "date": date_v, "desc": desc_v,
                        "amount": amt_v, "type": type_v, "source": source,
                        "doc_type": dtype,
                        "current_category": current_cat,
                        "suggested_category": suggested,
                        "vendor_norm": norm,
                        "client_name": j.get("client_name", ""),
                    })

            # Checks
            elif dtype == "check":
                check_amt_f = fields.get("check_amount")
                amt_v = (check_amt_f.get("value") if isinstance(check_amt_f, dict) else check_amt_f) if check_amt_f else None
                if amt_v is None:
                    continue
                vk = f"{page}:{ext_idx}:check_amount"
                vstate = vfields.get(vk, {})
                current_cat = vstate.get("category", "")

                payee_f = fields.get("payee") or fields.get("pay_to")
                payee = (payee_f.get("value", "") if isinstance(payee_f, dict) else str(payee_f or "")) if payee_f else ""
                date_f = fields.get("check_date")
                date_v = (date_f.get("value", "") if isinstance(date_f, dict) else str(date_f or "")) if date_f else ""
                num_f = fields.get("check_number")
                num_v = (num_f.get("value", "") if isinstance(num_f, dict) else str(num_f or "")) if num_f else ""

                norm = _normalize_vendor(payee)
                suggested = _suggest_category(payee) if not current_cat else ""

                items.append({
                    "job_id": jid, "page": page, "ext_idx": ext_idx,
                    "field_key": vk, "date": date_v,
                    "desc": f"Check #{num_v} to {payee}" if num_v else payee,
                    "amount": amt_v, "type": "check", "source": "Check",
                    "doc_type": dtype,
                    "current_category": current_cat,
                    "suggested_category": suggested,
                    "vendor_norm": norm,
                    "client_name": j.get("client_name", ""),
                })

    return items

# ─── Excel Regeneration ─────────────────────────────────────────────────────

def _regen_excel(job_id):
    """Regenerate the Excel file with operator verification corrections applied."""
    job = jobs.get(job_id)
    if not job or job.get("status") != "complete":
        return False

    log_path = job.get("output_log")
    xlsx_path = job.get("output_xlsx")
    if not log_path or not xlsx_path or not os.path.exists(log_path):
        return False

    vdata = _load_verifications(job_id)
    corrections = vdata.get("fields", {})
    if not corrections:
        return True  # Nothing to apply

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.comments import Comment

        # Load the extraction log
        with open(log_path) as f:
            log_data = json.load(f)

        extractions = log_data.get("extractions", [])
        if not extractions:
            return False

        # Apply corrections to the extraction data
        corrected_count = 0
        for key, decision in corrections.items():
            if decision.get("status") != "corrected":
                continue
            parts = key.split(":")
            if len(parts) != 3:
                continue
            page_str, ext_idx_str, field_name = parts

            corrected_value = decision.get("corrected_value")
            if corrected_value is None:
                continue

            # Find the matching extraction by page
            page_num = int(page_str) if page_str.isdigit() else None
            ext_idx = int(ext_idx_str) if ext_idx_str.isdigit() else None
            if page_num is None or ext_idx is None:
                continue

            # Match: find extractions for this page
            page_exts = [e for e in extractions if e.get("_page") == page_num]
            if ext_idx < len(page_exts):
                ext = page_exts[ext_idx]
                fields = ext.get("fields", {})
                if field_name in fields:
                    fdata = fields[field_name]
                    old_val = fdata.get("value") if isinstance(fdata, dict) else fdata
                    # Try to convert to number if it looks numeric
                    try:
                        new_val = float(str(corrected_value).replace(",", ""))
                    except (ValueError, TypeError):
                        new_val = corrected_value
                    if isinstance(fdata, dict):
                        fdata["_original_value"] = old_val
                        fdata["value"] = new_val
                        fdata["confidence"] = "operator_corrected"
                    else:
                        fields[field_name] = {
                            "value": new_val,
                            "_original_value": old_val,
                            "confidence": "operator_corrected",
                        }
                    corrected_count += 1

        # Inject operator-assigned categories into extraction fields
        # These flow through to _build_journal_entries to replace "Unclassified"
        for key, decision in corrections.items():
            cat = decision.get("category", "")
            if not cat:
                continue
            parts = key.split(":")
            if len(parts) != 3:
                continue
            page_str, ext_idx_str, field_name = parts
            page_num = int(page_str) if page_str.isdigit() else None
            ext_idx = int(ext_idx_str) if ext_idx_str.isdigit() else None
            if page_num is None or ext_idx is None:
                continue
            page_exts = [e for e in extractions if e.get("_page") == page_num]
            if ext_idx < len(page_exts):
                ext = page_exts[ext_idx]
                fields = ext.get("fields", {})
                # Store category on the field itself so _build_journal_entries can read it
                if field_name in fields:
                    fdata = fields[field_name]
                    if isinstance(fdata, dict):
                        fdata["_operator_category"] = cat
                    else:
                        fields[field_name] = {"value": fdata, "_operator_category": cat}
                # Also store vendor description for the extraction log
                vendor = decision.get("vendor_desc", "")
                if vendor:
                    if isinstance(fields.get(field_name), dict):
                        fields[field_name]["_vendor_desc"] = vendor

        # Now re-run populate_template via extract.py as a subprocess
        # This is the safest way — extract.py has all the Excel formatting logic
        import subprocess
        year = job.get("year", "2024")
        cmd = [
            sys.executable, str(BASE_DIR / "extract.py"),
            "--regen-excel",
            "--log-input", log_path,
            "--output", xlsx_path,
            "--year", str(year),
        ]

        # Write the corrected log to a temp file for extract.py to read
        corrected_log_path = log_path.replace("_log.json", "_corrected_log.json")
        with open(corrected_log_path, "w") as f:
            json.dump(log_data, f, indent=2, default=str)

        cmd[4] = corrected_log_path  # --log-input uses the corrected version

        proc = subprocess.run(cmd, capture_output=True, text=True, cwd=str(BASE_DIR), timeout=30)

        # Clean up temp file
        try:
            os.remove(corrected_log_path)
        except OSError:
            pass

        if proc.returncode != 0:
            print(f"  Regen subprocess failed (rc={proc.returncode}): {proc.stderr[:500] if proc.stderr else 'no stderr'}")
            # Fallback: apply corrections directly to existing Excel
            _apply_corrections_to_excel(xlsx_path, corrections, vdata.get("reviewer", ""))
        elif corrected_count > 0:
            # Verify corrections actually made it into the new Excel
            # (The subprocess rewrote the file using corrected extraction data,
            # so operator_corrected confidence values should appear)
            print(f"  Regen complete: {corrected_count} corrections applied via extract.py")

        # Always add the audit trail worksheet (primary path doesn't create one)
        _add_audit_trail_worksheet(xlsx_path, corrections, vdata.get("reviewer", ""))

        # Copy updated Excel to client folder
        if job.get("client_folder"):
            import shutil
            client_dir = Path(job["client_folder"])
            if client_dir.exists():
                try:
                    dst = client_dir / Path(xlsx_path).name
                    shutil.copy2(xlsx_path, str(dst))
                except Exception:
                    pass

        return True

    except Exception as e:
        print(f"  Excel regen error: {e}")
        # Fallback: try direct Excel patching
        try:
            _apply_corrections_to_excel(xlsx_path, corrections, vdata.get("reviewer", ""))
        except Exception:
            pass
        return False


def _add_audit_trail_worksheet(xlsx_path, corrections, reviewer=""):
    """Add or replace the Audit Trail worksheet with all verification decisions."""
    if not os.path.exists(xlsx_path) or not corrections:
        return
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.load_workbook(xlsx_path)
        audit_name = "Audit Trail"
        if audit_name in wb.sheetnames:
            del wb[audit_name]
        ws = wb.create_sheet(audit_name)

        # Title
        ws["A1"] = "Operator Verification Audit Trail"
        ws["A1"].font = Font(bold=True, size=14, color="1A252F")
        ws.merge_cells("A1:G1")
        ws["A2"] = f"Generated {datetime.now().strftime('%m/%d/%Y %I:%M %p')}"
        ws["A2"].font = Font(italic=True, color="888888", size=9)
        ws["A3"] = f"Reviewer: {reviewer}" if reviewer else "Reviewer: (not specified)"
        ws["A3"].font = Font(bold=True, size=11)

        # Headers
        row = 5
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        headers = ["Page:Field", "Status", "Original Value", "Corrected Value", "Reviewer", "Timestamp", "Note"]
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=i+1, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center" if i == 1 else "left")
        row += 1

        # Data rows with alternating colors
        alt_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        status_fills = {
            "corrected": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
            "confirmed": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
            "flagged": PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
        }
        thin_border = Border(bottom=Side(style="thin", color="E0E0E0"))

        for idx, key in enumerate(sorted(corrections.keys())):
            decision = corrections[key]
            parts = key.split(":")
            if len(parts) == 3:
                page_label = f"Pg {parts[0]}"
                field_label = parts[2].replace("_", " ").title()
                field_display = f"{page_label}: {field_label}"
            else:
                field_display = key

            status = decision.get("status", "")
            original = decision.get("original_value", "")
            corrected = decision.get("corrected_value", "")
            rev = decision.get("reviewer", reviewer)
            ts = decision.get("timestamp", "")
            note = decision.get("note", "")

            ws.cell(row=row, column=1, value=field_display)
            status_cell = ws.cell(row=row, column=2, value=status.upper())
            status_cell.alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=3, value=str(original) if original else "")
            ws.cell(row=row, column=4, value=str(corrected) if corrected else "")
            ws.cell(row=row, column=5, value=rev)
            ws.cell(row=row, column=6, value=ts)
            ws.cell(row=row, column=7, value=note)

            # Status color
            if status in status_fills:
                status_cell.fill = status_fills[status]
            # Alternating row background
            if idx % 2 == 1:
                for c in range(1, 8):
                    cell = ws.cell(row=row, column=c)
                    if cell.fill == PatternFill():  # only if not already colored
                        cell.fill = alt_fill
            # Border
            for c in range(1, 8):
                ws.cell(row=row, column=c).border = thin_border
            row += 1

        # Summary
        row += 1
        confirmed = sum(1 for d in corrections.values() if d.get("status") == "confirmed")
        corrected = sum(1 for d in corrections.values() if d.get("status") == "corrected")
        flagged = sum(1 for d in corrections.values() if d.get("status") == "flagged")
        ws.cell(row=row, column=1, value="Summary:").font = Font(bold=True, size=11)
        row += 1
        summary_items = [
            ("Confirmed", confirmed, "C8E6C9"),
            ("Corrected", corrected, "FFF9C4"),
            ("Flagged", flagged, "FFE0B2"),
            ("Total Reviewed", len(corrections), "E0E0E0"),
        ]
        for label, count, color in summary_items:
            ws.cell(row=row, column=1, value=label)
            ct_cell = ws.cell(row=row, column=2, value=count)
            ct_cell.font = Font(bold=True, size=11)
            ct_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ct_cell.alignment = Alignment(horizontal="center")
            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 24
        ws.column_dimensions["G"].width = 32

        # Print setup
        ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.orientation = "landscape"

        wb.save(xlsx_path)
    except Exception as e:
        print(f"  Audit trail worksheet error: {e}")


def _apply_corrections_to_excel(xlsx_path, corrections, reviewer=""):
    """Direct fallback: patch existing Excel cells with corrected values + add audit sheet."""
    if not os.path.exists(xlsx_path):
        return

    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.comments import Comment

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Color for operator-corrected cells
    op_fill = PatternFill(start_color="B3E5FC", end_color="B3E5FC", fill_type="solid")

    # Build a list of corrections to apply: (original_value, corrected_value, field_name)
    pending = []
    for key, decision in corrections.items():
        if decision.get("status") != "corrected" or decision.get("corrected_value") is None:
            continue
        parts = key.split(":")
        field_name = parts[2] if len(parts) == 3 else key
        orig = decision.get("original_value")
        corr = decision.get("corrected_value")
        # Normalize to number if possible
        try:
            corr_num = float(str(corr).replace(",", ""))
        except (ValueError, TypeError):
            corr_num = None
        try:
            orig_num = float(str(orig).replace(",", "")) if orig is not None else None
        except (ValueError, TypeError):
            orig_num = None
        pending.append({
            "field": field_name,
            "orig": orig, "orig_num": orig_num,
            "corr": corr, "corr_num": corr_num,
            "reviewer": decision.get("reviewer", reviewer),
            "applied": False,
        })

    # Scan all data cells and match by original value
    # Strategy: for each cell with a value, check if it matches any pending correction's
    # original value. Apply the first match and mark it done. This is best-effort but
    # far better than the previous no-op.
    patched = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is None:
                continue
            for p in pending:
                if p["applied"]:
                    continue
                matched = False
                # Numeric match (within tolerance for floating point)
                if p["orig_num"] is not None and isinstance(cell.value, (int, float)):
                    if abs(cell.value - p["orig_num"]) < 0.005:
                        matched = True
                # String match
                elif p["orig"] is not None and isinstance(cell.value, str):
                    if str(cell.value).strip() == str(p["orig"]).strip():
                        matched = True
                # String-to-number: cell has number but orig was stored as string
                elif p["orig_num"] is not None and isinstance(cell.value, (int, float)):
                    pass  # already handled above
                elif p["orig"] is not None and isinstance(cell.value, (int, float)):
                    try:
                        if abs(cell.value - float(str(p["orig"]).replace(",", ""))) < 0.005:
                            matched = True
                    except (ValueError, TypeError):
                        pass

                if matched:
                    # Apply correction
                    if p["corr_num"] is not None and isinstance(cell.value, (int, float)):
                        cell.value = p["corr_num"]
                    else:
                        cell.value = p["corr"]
                    cell.fill = op_fill
                    old_display = str(p["orig"]) if p["orig"] is not None else "?"
                    cell.comment = Comment(
                        f"Operator corrected (was {old_display}) — {p['reviewer']}",
                        "Operator"
                    )
                    p["applied"] = True
                    patched += 1
                    break  # move to next cell

    wb.save(xlsx_path)

    # Add audit trail worksheet
    _add_audit_trail_worksheet(xlsx_path, corrections, reviewer)
