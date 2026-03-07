"""
workpaper_styles — Centralized style system for Excel workpaper generation.

Inspired by Pi for Excel's composable named styles ("conventions"):
  - Named styles can be combined: apply_styles(cell, ["currency", "total-row"])
  - Number format presets with configurable parameters
  - Column width math (character units → openpyxl width units)
  - Consistent formatting across all renderers (inkspren, oathledger, add-in)

All renderers (inkspren.py, oathledger/renderer.py, and the Office.js add-in)
share these constants so workpapers look identical regardless of render path.
"""

from __future__ import annotations

from typing import Any, Dict, List, Optional, Union

import openpyxl.styles
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)


# ─── NUMBER FORMAT PRESETS ─────────────────────────────────────────────────
# Mirrors Pi for Excel's format-builder: named presets with consistent strings.

NUMBER_FORMATS = {
    "currency":  '#,##0.00_);(#,##0.00)',   # Accounting: parentheses for negatives
    "integer":   '#,##0_);(#,##0)',          # Whole number with thousands sep
    "number":    '#,##0.00',                 # Standard 2dp
    "percent":   '0.00%',                    # Percentage
    "ratio":     '0.00x',                    # Multiple/ratio (e.g. 1.25x)
    "date":      'MM/DD/YYYY',               # Standard US date
    "text":      '@',                        # Force text
}

# Backward compat aliases
MONEY_FMT = NUMBER_FORMATS["currency"]
PCT_FMT   = NUMBER_FORMATS["percent"]
DATE_FMT  = NUMBER_FORMATS["date"]
INT_FMT   = NUMBER_FORMATS["integer"]


# ─── COLOR PALETTE ─────────────────────────────────────────────────────────
# Centralized colors so all renderers use identical values.

COLORS = {
    # Structural
    "section_gray":       "D9D9D9",   # Section header background
    "header_dark":        "2C3E50",   # Dark header (used in some formats)
    "header_blue":        "4472C4",   # Pi-style blue header
    "blank_section":      "F2F2F2",   # Blank separator area

    # Confidence / status fills
    "flag_yellow":        "FFFDE7",   # Soft yellow — low confidence / preparer judgment
    "input_yellow":       "FFFD78",   # Bright yellow — user input cell
    "corrected_green":    "C8E6C9",   # Green — corrected value
    "confirmed_green":    "E8F5E9",   # Light green — confirmed
    "dual_green":         "A5D6A7",   # Darker green — OCR + image agree
    "review_orange":      "FFE0B2",   # Orange — needs human review

    # Text
    "black":              "000000",
    "white":              "FFFFFF",
    "dark_gray":          "333333",
    "medium_gray":        "999999",
    "light_gray":         "BBBBBB",
    "flag_red":           "CC0000",   # Flag / warning text
}


# ─── BORDER DEFINITIONS ──────────────────────────────────────────────────
# Pi-style border system: named borders with priority ordering.

BORDERS = {
    "none":        Border(),
    "thin":        Border(
        top=Side(style="thin", color=COLORS["black"]),
        bottom=Side(style="thin", color=COLORS["black"]),
        left=Side(style="thin", color=COLORS["black"]),
        right=Side(style="thin", color=COLORS["black"]),
    ),
    "thin_bottom": Border(
        bottom=Side(style="thin", color=COLORS["black"]),
    ),
    "total":       Border(   # Classic total row: thin top + double bottom
        top=Side(style="thin", color=COLORS["black"]),
        bottom=Side(style="double", color=COLORS["black"]),
    ),
    "subtotal":    Border(
        top=Side(style="thin", color=COLORS["black"]),
    ),
}


# ─── NAMED STYLES ──────────────────────────────────────────────────────────
# Composable named styles, inspired by Pi for Excel's conventions system.
# Each style is a dict of openpyxl-compatible properties.
# Apply one or more by name: apply_styles(cell, ["currency", "total-row"])

NAMED_STYLES: Dict[str, Dict[str, Any]] = {
    # ── Format styles ──
    "currency": {
        "number_format": MONEY_FMT,
        "alignment": Alignment(horizontal="right"),
    },
    "integer": {
        "number_format": INT_FMT,
        "alignment": Alignment(horizontal="right"),
    },
    "percent": {
        "number_format": PCT_FMT,
        "alignment": Alignment(horizontal="right"),
    },
    "date": {
        "number_format": DATE_FMT,
    },
    "text": {
        "number_format": NUMBER_FORMATS["text"],
    },

    # ── Structural styles (Pi-inspired) ──
    "section-header": {
        "font": Font(bold=True, size=11, color=COLORS["black"]),
        "fill": PatternFill("solid", fgColor=COLORS["section_gray"]),
    },
    "col-header": {
        "font": Font(size=11, color=COLORS["black"]),
        "fill": PatternFill("solid", fgColor=COLORS["section_gray"]),
        "alignment": Alignment(horizontal="center"),
    },
    "header-blue": {
        "font": Font(bold=True, color=COLORS["white"], size=10),
        "fill": PatternFill("solid", fgColor=COLORS["header_blue"]),
    },
    "header-dark": {
        "font": Font(bold=True, color=COLORS["white"], size=10),
        "fill": PatternFill("solid", fgColor=COLORS["header_dark"]),
    },
    "total-row": {
        "font": Font(bold=True, size=11, color=COLORS["black"]),
        "fill": PatternFill(),  # No fill on total rows
        "border": BORDERS["total"],
    },
    "subtotal": {
        "font": Font(bold=True),
        "border": BORDERS["subtotal"],
    },
    "blank-section": {
        "fill": PatternFill("solid", fgColor=COLORS["blank_section"]),
    },

    # ── Data row styles ──
    "data": {
        "font": Font(size=11, color=COLORS["black"]),
        "border": BORDERS["none"],
    },
    "data-label": {
        "font": Font(size=11, color=COLORS["black"]),
    },

    # ── Status/confidence styles ──
    "flag": {
        "font": Font(italic=True, color=COLORS["flag_red"]),
    },
    "flag-row": {
        "font": Font(italic=True, color=COLORS["flag_red"], size=9),
    },
    "input": {
        "fill": PatternFill("solid", fgColor=COLORS["input_yellow"]),
    },
    "review": {
        "fill": PatternFill("solid", fgColor=COLORS["review_orange"]),
    },
    "confirmed": {
        "fill": PatternFill("solid", fgColor=COLORS["confirmed_green"]),
    },
    "corrected": {
        "fill": PatternFill("solid", fgColor=COLORS["corrected_green"]),
    },
    "low-confidence": {
        "fill": PatternFill("solid", fgColor=COLORS["flag_yellow"]),
    },

    # ── Title styles ──
    "title": {
        "font": Font(bold=True, size=14, color=COLORS["black"]),
        "alignment": Alignment(horizontal="center"),
    },
    "client-name": {
        "font": Font(bold=True, size=16, color=COLORS["black"]),
        "alignment": Alignment(horizontal="center"),
    },
    "subtitle": {
        "font": Font(italic=True, color=COLORS["medium_gray"], size=9),
        "alignment": Alignment(horizontal="center"),
    },
    "sub-header": {
        "font": Font(bold=True, size=10, color=COLORS["dark_gray"]),
    },
}


# ─── STYLE APPLICATION ─────────────────────────────────────────────────────

def apply_styles(cell, style_names: Union[str, List[str]], **overrides) -> None:
    """Apply one or more named styles to an openpyxl cell.

    Styles are applied in order — later styles override earlier ones.
    Additional keyword overrides are applied last.

    Usage:
        apply_styles(cell, "currency")
        apply_styles(cell, ["currency", "total-row"])
        apply_styles(cell, "currency", number_format='#,##0')
    """
    if isinstance(style_names, str):
        style_names = [style_names]

    for name in style_names:
        style = NAMED_STYLES.get(name)
        if not style:
            continue
        _apply_props(cell, style)

    if overrides:
        _apply_props(cell, overrides)


def _apply_props(cell, props: Dict[str, Any]) -> None:
    """Apply a dict of properties to an openpyxl cell."""
    for key, value in props.items():
        if key == "font":
            cell.font = value
        elif key == "fill":
            cell.fill = value
        elif key == "border":
            cell.border = value
        elif key == "alignment":
            cell.alignment = value
        elif key == "number_format":
            cell.number_format = value


def apply_styles_to_range(ws, range_str: str, style_names: Union[str, List[str]], **overrides) -> None:
    """Apply named styles to all cells in a range string like 'A1:F1'."""
    for row in ws[range_str]:
        if hasattr(row, '__iter__'):
            for cell in row:
                apply_styles(cell, style_names, **overrides)
        else:
            # Single cell
            apply_styles(row, style_names, **overrides)


# ─── COLUMN WIDTH MATH ────────────────────────────────────────────────────
# Pi for Excel uses POINTS_PER_CHAR_ARIAL_10 = 7.2
# openpyxl column_dimensions.width is in "character units" (approx width of '0' in default font).
# Excel default font is Calibri 11 where 1 char ≈ 7.0 pixels.
#
# For our purposes:
#   - A (entity names): ~35 chars
#   - B-F (numbers): ~15 chars

COLUMN_WIDTHS = {
    "entity":   35,    # Column A — entity/employer names
    "number":   15,    # B-F — numeric columns
    "narrow":   10,    # Distribution codes, flags
    "wide":     40,    # Description columns
}

def set_column_widths(ws, widths: Dict[str, int]) -> None:
    """Set column widths from a dict of {column_letter: char_width}.

    Example:
        set_column_widths(ws, {"A": 35, "B": 15, "C": 15})
    """
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def set_standard_widths(ws, num_cols: int = 6) -> None:
    """Set standard workpaper column widths: A=35 (entity), B-onwards=15 (numbers)."""
    ws.column_dimensions["A"].width = COLUMN_WIDTHS["entity"]
    for i in range(1, num_cols):
        col_letter = chr(65 + i)  # B, C, D, E, F
        ws.column_dimensions[col_letter].width = COLUMN_WIDTHS["number"]


# ─── PRINT SETUP ───────────────────────────────────────────────────────────

def setup_print(ws, title: str, year: int, footer_text: str = "Bearden Accounting — Document Intake v5") -> None:
    """Configure print settings for professional output.

    - Landscape, fit to 1 page wide, unlimited pages tall
    - Header with title + year, footer with source + page numbers
    - Gridlines on
    """
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.oddHeader.center.text = f"&B{title} — {year}"
    ws.oddHeader.center.size = 10
    ws.oddFooter.left.text = footer_text
    ws.oddFooter.left.size = 8
    ws.oddFooter.right.text = "Page &P of &N"
    ws.oddFooter.right.size = 8
    ws.print_options.gridLines = True


# ─── CONVENIENCE HELPERS ───────────────────────────────────────────────────

def write_section_header(ws, row: int, header_text: str, col_headers: Dict[str, str],
                         max_col: str = "F") -> int:
    """Write a section header row with column headers, return next row.

    Applies section-header style to header cell, col-header style to column headers,
    and fills remaining cells with section gray.
    """
    cell = ws[f"A{row}"]
    cell.value = header_text
    apply_styles(cell, "section-header")

    for col, label in col_headers.items():
        c = ws[f"{col}{row}"]
        c.value = label
        apply_styles(c, "col-header")

    # Fill remaining columns with section gray
    max_col_idx = ord(max_col) - 64
    for i in range(max_col_idx):
        col_letter = chr(65 + i)
        if col_letter != "A" and col_letter not in col_headers:
            ws[f"{col_letter}{row}"].fill = PatternFill("solid", fgColor=COLORS["section_gray"])

    return row + 1


def write_total_row(ws, row: int, data_start: int, data_end: int,
                    sum_cols: List[str], total_formula_col: Optional[Dict[str, str]] = None,
                    label: str = "TOTAL") -> int:
    """Write a TOTAL row with SUM formulas and total-row styling.

    Returns the next row after the total.
    """
    if data_end < data_start:
        return row

    # Label cell
    cell_a = ws[f"A{row}"]
    cell_a.value = label
    apply_styles(cell_a, "total-row")

    # SUM formula cells
    for col in sum_cols:
        cell = ws[f"{col}{row}"]
        cell.value = f"=SUM({col}{data_start}:{col}{data_end})"
        apply_styles(cell, ["currency", "total-row"])

    # Cross-column formulas (e.g., D = B + C)
    if total_formula_col:
        for col, formula_spec in total_formula_col.items():
            parts = formula_spec.split("+")
            cell = ws[f"{col}{row}"]
            cell.value = "=" + "+".join([f"{p}{row}" for p in parts])
            apply_styles(cell, ["currency", "total-row"])

    return row + 1


def write_flag_rows(ws, row: int, flags: List[str]) -> int:
    """Write flag/note rows with flag-row styling. Returns next row."""
    for flag in flags:
        cell = ws[f"A{row}"]
        cell.value = flag
        apply_styles(cell, "flag-row")
        row += 1
    return row


def write_title_block(ws, title: str, year: int, client_name: str = "",
                      merge_cols: str = "A{}:F{}") -> int:
    """Write title rows (client name, title, timestamp). Returns next data row.

    Mirrors inkspren._write_title but uses named styles.
    """
    from datetime import datetime

    row = 1
    if client_name:
        cell = ws[f"A{row}"]
        cell.value = client_name
        apply_styles(cell, "client-name")
        ws.merge_cells(merge_cols.format(row, row))
        row += 1

    cell = ws[f"A{row}"]
    cell.value = f"{title} — {year}"
    apply_styles(cell, "title")
    ws.merge_cells(merge_cols.format(row, row))
    row += 1

    cell = ws[f"A{row}"]
    cell.value = f"Extracted {datetime.now().strftime('%m/%d/%Y %I:%M %p')}"
    apply_styles(cell, "subtitle")
    ws.merge_cells(merge_cols.format(row, row))

    return row + 2  # Skip a blank row after title block


# ─── EXPORT BACKWARD COMPAT ──────────────────────────────────────────────
# These aliases let existing code import from here without changes.

BOLD              = Font(bold=True)
SECTION_FONT      = NAMED_STYLES["section-header"]["font"]
SECTION_FILL      = NAMED_STYLES["section-header"]["fill"]
COL_HEADER_FONT   = NAMED_STYLES["col-header"]["font"]
COL_HEADER_FILL   = NAMED_STYLES["col-header"]["fill"]
SUM_FONT          = NAMED_STYLES["total-row"]["font"]
SUM_FILL          = NAMED_STYLES["total-row"]["fill"]
SUM_BORDER        = NAMED_STYLES["total-row"]["border"]
FLAG_FONT         = NAMED_STYLES["flag"]["font"]
FLAG_FILL         = PatternFill("solid", fgColor=COLORS["flag_yellow"])
CORRECTED_FILL    = NAMED_STYLES["corrected"]["fill"]
REVIEW_FILL       = NAMED_STYLES["review"]["fill"]
CONFIRMED_FILL    = NAMED_STYLES["confirmed"]["fill"]
DUAL_FILL         = PatternFill("solid", fgColor=COLORS["dual_green"])
ALT_ROW_FILL      = PatternFill()
DARK_HEADER_FILL  = NAMED_STYLES["header-dark"]["fill"]
DARK_HEADER_FONT  = NAMED_STYLES["header-dark"]["font"]
THIN_BORDER       = BORDERS["none"]
SECTION_BORDER    = BORDERS["none"]
